from __future__ import annotations

import csv
import json
import math
import os
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from .errors import EngineError
from .loaders import ROOT, import_path, load_file

Progress = Callable[[str, int, int, str, str, list[str] | None, Any | None], None]


def _required(params: dict, key: str) -> Any:
    value = params.get(key)
    if value is None or value == "" or value == []:
        raise EngineError("INVALID_ARGUMENT", f"缺少必填参数：{key}")
    return value


def _assert_exists(path: str, *, directory: bool = False) -> Path:
    p = Path(path).expanduser().resolve()
    ok = p.is_dir() if directory else p.is_file()
    if not ok:
        raise EngineError("PATH_NOT_FOUND", f"找不到{'文件夹' if directory else '文件'}：{p}")
    return p


def _json_value(value: Any) -> Any:
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(k): _json_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_value(v) for v in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def inspect_tabular(path: Path, sheet: str | None = None, header_row: int = 1, limit: int = 20) -> dict:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt", ".tsv"}:
        raw = path.read_bytes()[:65536]
        encoding = "utf-8-sig"
        for candidate in ("utf-8-sig", "gb18030", "utf-16", "latin-1"):
            try:
                text = raw.decode(candidate)
                encoding = candidate
                break
            except UnicodeDecodeError:
                continue
        sample = text[:8192]
        delimiter = "\t" if suffix == ".tsv" else ","
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            pass
        with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = list(row for _, row in zip(range(header_row - 1 + limit + 1), reader))
        headers = rows[header_row - 1] if len(rows) >= header_row else []
        preview = rows[header_row:header_row + limit]
        return {"path": str(path), "kind": "text", "encoding": encoding, "delimiter": delimiter, "sheets": [], "headers": headers, "preview": preview}
    if suffix == ".parquet":
        raise EngineError("PARQUET_RUST_ROUTE_REQUIRED", "Parquet 文件请通过 Rust Polars 工具读取。")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=False)
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        selected = sheet if sheet in wb.sheetnames else (visible[0] if visible else wb.sheetnames[0])
        ws = wb[selected]
        rows = list(ws.iter_rows(min_row=max(1, header_row), max_row=max(1, header_row) + limit, values_only=True))
        headers = [_json_value(v) for v in (rows[0] if rows else [])]
        preview = [[_json_value(v) for v in row] for row in rows[1:]]
        dimensions = {"rows": ws.max_row, "columns": ws.max_column}
        wb.close()
        return {"path": str(path), "kind": "excel", "sheets": visible, "selectedSheet": selected, "headers": headers, "preview": preview, "dimensions": dimensions}
    except EngineError:
        raise
    except Exception as exc:
        raise EngineError("WORKBOOK_READ_FAILED", "无法读取工作簿，请确认文件未损坏且未被占用。", detail=str(exc)) from exc


def file_list_scan(params: dict, **_: Any) -> dict:
    source = _assert_exists(str(_required(params, "sourceDir")), directory=True)
    count = 0; max_depth = 0; sample: list[dict] = []
    for root, _dirs, files in os.walk(source):
        rel = Path(root).relative_to(source)
        max_depth = max(max_depth, len(rel.parts))
        for name in files:
            count += 1
            if len(sample) < 20:
                sample.append({"name": name, "relativePath": str(rel / name), "fullPath": str(Path(root) / name)})
    return {"sourceDir": str(source), "fileCount": count, "maxDepth": max_depth, "preview": sample}


def file_list_export(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    source = _assert_exists(str(_required(params, "sourceDir")), directory=True)
    output = Path(str(_required(params, "outputPath"))).expanduser().resolve()
    if output.suffix.lower() != ".xlsx": output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    files = [(Path(root), name) for root, _dirs, names in os.walk(source) for name in names]
    max_depth = max((len(root.relative_to(source).parts) for root, _ in files), default=0)
    import xlsxwriter
    workbook = xlsxwriter.Workbook(output)
    try:
        sheet = workbook.add_worksheet("文件清单")
        header = workbook.add_format({"bold": True, "bg_color": "#DCECE8", "border": 1})
        link = workbook.add_format({"font_color": "blue", "underline": 1})
        headers = [f"{i + 1}级文件夹" for i in range(max_depth + 1)] + ["文件名称", "超链接", "文件路径"]
        for col, value in enumerate(headers): sheet.write(0, col, value, header)
        total = len(files)
        for index, (root, name) in enumerate(files, start=1):
            if cancel.is_set(): raise EngineError("JOB_CANCELLED", "任务已取消。")
            parts = [source.name, *root.relative_to(source).parts]
            row = index
            for col in range(max_depth + 1): sheet.write(row, col, parts[col] if col < len(parts) else "")
            path = (root / name).resolve(); base = max_depth + 1
            sheet.write(row, base, name)
            try: sheet.write_url(row, base + 1, path.as_uri(), link, name)
            except Exception: sheet.write(row, base + 1, name)
            sheet.write(row, base + 2, str(path))
            if index == total or index % 100 == 0: progress("export", index, total, f"正在写入 {index}/{total}", "info", None, None)
        sheet.freeze_panes(1, 0); sheet.autofilter(0, 0, max(total, 1), len(headers) - 1)
        sheet.set_column(0, max_depth, 18); sheet.set_column(max_depth + 1, max_depth + 2, 30); sheet.set_column(max_depth + 3, max_depth + 3, 70)
    finally:
        workbook.close()
    return {"fileCount": len(files), "outputPaths": [str(output)]}


def wp_validate(params: dict, **_: Any) -> dict:
    folder = _assert_exists(str(_required(params, "folder")), directory=True)
    required = ["FY27 WP服务单.xlsx", "FY27 section list.xlsx"]
    missing = [name for name in required if not (folder / name).is_file()]
    return {"folder": str(folder), "valid": not missing, "missing": missing, "outputPath": str(folder / "FY27+WP服务单汇总.xlsx")}


def wp_generate(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    check = wp_validate(params)
    if not check["valid"]: raise EngineError("WP_INPUT_MISSING", "缺少 WP 服务单输入文件：" + "、".join(check["missing"]))
    if cancel.is_set(): raise EngineError("JOB_CANCELLED", "任务已取消。")
    progress("validate", 1, 3, "输入文件检查通过", "info", None, None)
    folder = Path(check["folder"])
    module = load_file("audit_wp_adapter", "modules/wp-service-generator/FY27_WP服务单生成工具.py")
    progress("generate", 2, 3, "正在生成服务方案和汇总文件", "info", None, None)
    result = module.run_generation(folder)
    output = folder / "FY27+WP服务单汇总.xlsx"
    progress("verify", 3, 3, "生成完成", "success", [str(output)], None)
    return {**_json_value(result), "outputPaths": [str(output)]}


def confirmation_inspect(params: dict, **_: Any) -> dict:
    path = _assert_exists(str(_required(params, "inputPath")))
    from .confirmation_adapter import inspect_confirmation

    return _json_value(inspect_confirmation(path, params.get("mode") or "both"))


def confirmation_process(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    path = _assert_exists(str(_required(params, "inputPath")))
    from .confirmation_adapter import process_confirmation

    return _json_value(process_confirmation(path, params.get("mode") or "both", progress, cancel))


def generic_inspect(params: dict, **_: Any) -> dict:
    path_value = params.get("inputPath") or params.get("beginPath")
    if not path_value: raise EngineError("INVALID_ARGUMENT", "请选择输入文件。")
    return inspect_tabular(_assert_exists(str(path_value)), params.get("sheet") or params.get("beginSheet"), int(params.get("headerRow") or 1))


def _fa_header_value(value: Any) -> int | None:
    if value in (None, "", "auto", "自动"):
        return None
    try:
        return max(1, int(value))
    except (TypeError, ValueError):
        raise EngineError("FA_HEADER_INVALID", "标题行必须是大于等于 1 的整数，或留空自动识别。")


def _fa_inspect_one(path: Path, sheet: str | None, header_value: Any, limit: int = 12) -> dict:
    """Inspect through the legacy FileHandler so preview and execution use identical columns."""
    import pandas as pd

    with import_path(ROOT / "tools" / "fa_list"):
        from file_handler import FileHandler

        helper = FileHandler()
        requested_header = _fa_header_value(header_value)
        suffix = path.suffix.lower()
        sheets: list[str] = []
        if suffix in {".xlsx", ".xls", ".xlsm"}:
            ok, message, available = helper.get_excel_sheets(str(path))
            if not ok:
                raise EngineError("FA_LOAD_FAILED", str(message))
            sheets = [str(value) for value in available]

        def detect(candidate: str | None) -> tuple[int, list[str], dict, int]:
            if requested_header is not None:
                header_index = requested_header - 1
            else:
                try:
                    if suffix in {".xlsx", ".xls", ".xlsm"}:
                        engine = "xlrd" if suffix == ".xls" else "openpyxl"
                        raw = pd.read_excel(
                            path,
                            sheet_name=candidate if candidate else 0,
                            header=None,
                            nrows=20,
                            engine=engine,
                        )
                    else:
                        raw = pd.read_csv(path, header=None, nrows=20, low_memory=False)
                    header_index = helper._detect_header_row(raw, max_rows=20)
                except Exception:
                    header_index = 0
            try:
                if suffix in {".xlsx", ".xls", ".xlsm"}:
                    engine = "xlrd" if suffix == ".xls" else "openpyxl"
                    sample = pd.read_excel(
                        path,
                        sheet_name=candidate if candidate else 0,
                        header=header_index,
                        nrows=3,
                        engine=engine,
                    )
                else:
                    sample = pd.read_csv(path, header=header_index, nrows=3, low_memory=False)
                headers = [str(value) for value in sample.columns]
                row_count = len(sample)
            except Exception:
                headers, row_count = [], 0
            mapping = _fa_suggest_columns(headers)
            return header_index, headers, mapping, row_count

        selected = sheet if sheet in sheets else None
        detected_header = 0
        if selected is None and sheets:
            candidates = []
            for position, candidate in enumerate(sheets):
                header_index, headers, mapping, _sample_rows = detect(candidate)
                mapped = sum(1 for value in mapping.values() if value)
                core = sum(
                    1
                    for key in ("matchKey", "category", "name", "originalValue", "depreciation")
                    if mapping.get(key)
                )
                lowered = candidate.lower()
                summary_penalty = 5 if any(word in lowered for word in ("合计", "汇总", "summary", "pivot")) else 0
                score = mapped * 2 + core * 4 + (6 if mapping.get("matchKey") else 0) - summary_penalty - position * 0.01
                candidates.append((score, candidate, header_index))
            _score, selected, detected_header = max(candidates, key=lambda item: item[0])
        else:
            detected_header, _headers, _mapping, _rows = detect(selected)

        success, message, frame = helper.load_file(str(path), selected, detected_header)
        if not success or frame is None:
            raise EngineError("FA_LOAD_FAILED", str(message))
        frame = frame.copy()
        frame.columns = [str(value) for value in frame.columns]
        suggested = _fa_suggest_columns(list(frame.columns), frame)
        id_terms = ("编号", "编码", "卡片", "coding", "code", "assetid", "assetnumber")
        id_candidates = []
        for column in frame.columns:
            normalized = "".join(ch.lower() for ch in str(column) if ch.isalnum())
            if not any(term in normalized for term in id_terms):
                continue
            values = frame[column].dropna()
            non_empty = values.astype(str).str.strip()
            non_empty = non_empty[non_empty.ne("")]
            if non_empty.empty:
                continue
            unique_ratio = non_empty.nunique(dropna=True) / len(non_empty)
            coverage = len(non_empty) / max(len(frame), 1)
            id_candidates.append((unique_ratio * 5 + coverage, str(column)))
        if id_candidates:
            suggested["matchKey"] = max(id_candidates, key=lambda item: item[0])[1]
        preview_frame = frame.head(limit).where(pd.notna(frame.head(limit)), None)
        preview = [_json_value(list(row)) for row in preview_frame.itertuples(index=False, name=None)]
        return {
            "path": str(path),
            "kind": "excel" if sheets else "text",
            "sheets": sheets,
            "selectedSheet": selected,
            "displayName": f"{path.name} & {selected}" if selected else path.name,
            "detectedHeaderRow": detected_header + 1,
            "headerMode": "manual" if requested_header is not None else "auto",
            "headers": list(frame.columns),
            "suggestedMapping": suggested,
            "preview": preview,
            "dimensions": {"rows": len(frame), "columns": len(frame.columns)},
        }


def fa_inspect(params: dict, **_: Any) -> dict:
    begin = _fa_inspect_one(
        _assert_exists(str(_required(params, "beginPath"))),
        params.get("beginSheet"),
        params.get("beginHeaderRow"),
    )
    end = _fa_inspect_one(
        _assert_exists(str(_required(params, "endPath"))),
        params.get("endSheet"),
        params.get("endHeaderRow"),
    )
    begin_mapping = begin.pop("suggestedMapping")
    end_mapping = end.pop("suggestedMapping")
    begin_keys = [begin_mapping.get("matchKey")] if begin_mapping.get("matchKey") else []
    end_keys = [end_mapping.get("matchKey")] if end_mapping.get("matchKey") else []
    # The effective legacy workflow automatically extends its primary asset ID
    # with the mapped asset-name column. Preserve that composite-key default.
    begin_name = begin_mapping.get("name")
    end_name = end_mapping.get("name")
    if (
        begin_keys
        and end_keys
        and begin_name
        and end_name
        and begin_name not in begin_keys
        and end_name not in end_keys
    ):
        begin_keys.append(begin_name)
        end_keys.append(end_name)
    begin_mapping["matchKeys"] = begin_keys
    end_mapping["matchKeys"] = end_keys
    return {
        "begin": begin,
        "end": end,
        "suggestedMapping": {
            "begin": begin_mapping,
            "end": end_mapping,
        },
    }


def fa_supplement_inspect(params: dict, **_: Any) -> dict:
    path = _assert_exists(str(_required(params, "path")))
    result = _fa_inspect_one(path, params.get("sheet"), params.get("headerRow"), 12)
    mapping = result.get("suggestedMapping") or {}
    headers = [str(value) for value in result.get("headers") or []]
    references = [str(value) for value in params.get("referenceKeys") or []]
    with import_path(ROOT / "tools" / "fa_list"):
        from mapping_rules import normalize_header, score_match_id
    selected: list[str] = []
    for reference in references:
        reference_normalized = normalize_header(reference)
        picked = next(
            (
                column for column in headers
                if column not in selected and normalize_header(column) == reference_normalized
            ),
            None,
        )
        if not picked and (
            score_match_id(reference) is not None
            or any(token in reference_normalized for token in ("coding", "assetcode", "code", "assetid", "id", "编码", "编号", "卡片号"))
        ):
            picked = mapping.get("matchKey")
        if not picked and any(
            token in reference_normalized
            for token in ("资产名称", "固定资产名称", "名称", "资产描述", "固定资产描述", "描述", "assetname", "name", "description", "desc")
        ):
            picked = mapping.get("name")
        if not picked:
            picked = next(
                (
                    column for column in headers
                    if column not in selected
                    and reference_normalized
                    and (
                        reference_normalized in normalize_header(column)
                        or normalize_header(column) in reference_normalized
                    )
                ),
                None,
            )
        if picked and picked in headers and picked not in selected:
            selected.append(picked)
    if not selected:
        selected = [
            value for value in (mapping.get("matchKey"), mapping.get("name"))
            if value and value not in selected
        ]
    mapping["matchKeys"] = selected
    return result


def _fa_llm_samples(frame: Any) -> dict[str, list[str]]:
    samples: dict[str, list[str]] = {}
    for column in list(frame.columns)[:80]:
        values: list[str] = []
        for value in frame[column].dropna().astype(str).head(3).tolist():
            text = value.strip()
            if text:
                values.append(text[:60])
        samples[str(column)] = values
    return samples


def _fa_llm_profiles(frame: Any) -> dict[str, dict[str, Any]]:
    import re

    profiles: dict[str, dict[str, Any]] = {}
    for column in list(frame.columns)[:80]:
        series = frame[column].dropna().astype(str).map(lambda value: value.strip())
        series = series[series.ne("")]
        sample = series.head(200).tolist()
        lengths = [len(value) for value in sample]
        denominator = len(sample) or 1
        profiles[str(column)] = {
            "non_empty_count": int(series.size),
            "unique_count": int(series.nunique(dropna=True)),
            "avg_text_len": round(sum(lengths) / len(lengths), 1) if lengths else 0,
            "max_text_len": max(lengths) if lengths else 0,
            "looks_like_code_ratio": round(
                sum(1 for value in sample if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.\-/]{0,11}", value))
                / denominator,
                2,
            ),
            "cjk_short_name_ratio": round(
                sum(1 for value in sample if re.search(r"[\u4e00-\u9fff]", value) and len(value) <= 15)
                / denominator,
                2,
            ),
            "long_text_ratio": round(sum(1 for value in sample if len(value) > 15) / denominator, 2),
        }
    return profiles


def _fa_key_profile(frame: Any, columns: list[str]) -> dict[str, Any]:
    import pandas as pd

    valid_columns = [column for column in columns if column in frame.columns]
    profile = {
        "columns": valid_columns,
        "row_count": int(len(frame)),
        "valid_count": 0,
        "blank_count": int(len(frame)),
        "blank_rate": 1.0 if len(frame) else 0.0,
        "unique_count": 0,
        "duplicate_key_count": 0,
        "duplicate_row_count": 0,
        "unique_rate": 0.0,
        "is_unique": False,
        "duplicate_examples": [],
        "missing_columns": [column for column in columns if column not in frame.columns],
    }
    if not valid_columns:
        return profile

    def normalize(value: Any) -> str:
        if value is None or pd.isna(value):
            return ""
        text = " ".join(str(value).split())
        return text[:-2] if text.endswith(".0") and text[:-2].replace("-", "", 1).isdigit() else text

    keys = frame[valid_columns].apply(
        lambda row: " | ".join(normalize(value) for value in row),
        axis=1,
    )
    blank = keys.str.replace(" | ", "", regex=False).eq("")
    non_blank = keys[~blank]
    counts = non_blank.value_counts(dropna=False)
    duplicates = counts[counts.gt(1)]
    profile.update({
        "valid_count": int(len(non_blank)),
        "blank_count": int(blank.sum()),
        "blank_rate": round(float(blank.mean()), 4) if len(keys) else 0.0,
        "unique_count": int(counts.shape[0]),
        "duplicate_key_count": int(duplicates.shape[0]),
        "duplicate_row_count": int(duplicates.sum()) if not duplicates.empty else 0,
        "unique_rate": round(float(counts.shape[0] / len(non_blank)), 4) if len(non_blank) else 0.0,
        "is_unique": bool(len(keys) and not blank.any() and duplicates.empty),
        "duplicate_examples": [
            {"key": str(key)[:120], "count": int(count)}
            for key, count in duplicates.head(5).items()
        ],
    })
    return profile


def fa_review(params: dict, **_: Any) -> dict:
    """Run the same effective FA mapping/review LLM used by the legacy wizard."""
    from dataclasses import asdict

    with import_path(ROOT / "tools" / "fa_list"):
        from file_handler import FileHandler
        from launcher.llm_client import AUTO_APPLY_CONFIDENCE, generate_combined_fa_list_assistance
        from launcher.llm_settings import is_llm_enabled, load_llm_settings

        if not is_llm_enabled():
            return {
                "enabled": False,
                "passed": True,
                "message": "Hub 尚未启用 LLM，保留当前脚本映射。",
                "autoApplied": [],
                "fieldReviews": [],
                "matchReview": None,
            }
        handler = FileHandler()
        begin_header = max(0, int(params.get("beginHeaderRow") or 1) - 1)
        end_header = max(0, int(params.get("endHeaderRow") or 1) - 1)
        ok1, message1 = handler.set_file1(
            str(_assert_exists(str(_required(params, "beginPath")))),
            params.get("beginSheet"),
            begin_header,
        )
        ok2, message2 = handler.set_file2(
            str(_assert_exists(str(_required(params, "endPath")))),
            params.get("endSheet"),
            end_header,
        )
        if not ok1 or not ok2:
            raise EngineError("FA_REVIEW_LOAD_FAILED", message1 if not ok1 else message2)
        frame1, frame2 = handler.file1_df, handler.file2_df
        columns1 = [str(column) for column in frame1.columns]
        columns2 = [str(column) for column in frame2.columns]
        begin_mapping = dict(params.get("beginMapping") or {})
        end_mapping = dict(params.get("endMapping") or {})
        role_map = {
            "original_value": "originalValue",
            "depreciation": "depreciation",
            "category": "category",
            "name": "name",
            "date": "startDate",
            "life": "life",
            "residual": "residualRate",
            "current_year_dep": "currentYearDep",
            "addition_method": "additionMethod",
            "addition_date": "additionDate",
        }
        current_mapping: dict[str, Any] = {
            "match": {
                "file1": [str(value) for value in params.get("beginKeys") or []],
                "file2": [str(value) for value in params.get("endKeys") or []],
            }
        }
        for role, frontend_name in role_map.items():
            current_mapping[role] = {
                "file1": "" if role in {"current_year_dep", "addition_method", "addition_date"} else str(begin_mapping.get(frontend_name) or ""),
                "file2": str(end_mapping.get(frontend_name) or ""),
            }
        role_labels = [
            ("match", "匹配列/固定资产编号/资产编码/卡片号"),
            ("original_value", "原值/资产原值/成本"),
            ("depreciation", "累计折旧"),
            ("category", "资产类别"),
            ("name", "固定资产名称"),
            ("date", "入账开始日期/取得日期/资本化日期"),
            ("life", "使用寿命(月)/使用年限"),
            ("residual", "残值率"),
            ("current_year_dep", "本年折旧"),
            ("addition_method", "新增方式"),
            ("addition_date", "新增时间"),
        ]
        role_definitions = [
            {"role": role, "label": label, "description": label}
            for role, label in role_labels
        ]
        files = [
            {
                "file_side": "file1",
                "headers": columns1,
                "samples": _fa_llm_samples(frame1),
                "column_profiles": _fa_llm_profiles(frame1),
            },
            {
                "file_side": "file2",
                "headers": columns2,
                "samples": _fa_llm_samples(frame2),
                "column_profiles": _fa_llm_profiles(frame2),
            },
        ]
        current_match = current_mapping["match"]
        local_profile = {
            "file1": _fa_key_profile(frame1, current_match["file1"]),
            "file2": _fa_key_profile(frame2, current_match["file2"]),
        }
        forbidden = {"file1": [], "file2": []}
        for role, sides in current_mapping.items():
            if role in {"match", "name"} or not isinstance(sides, dict):
                continue
            for side in ("file1", "file2"):
                value = str(sides.get(side) or "")
                if value:
                    forbidden[side].append(value)
        combined = generate_combined_fa_list_assistance(
            load_llm_settings(),
            tool_name="FA List",
            role_definitions=role_definitions,
            files=files,
            current_mapping=current_mapping,
            current_match=current_match,
            local_profile=local_profile,
            candidate_profiles=[],
            include_match_review=True,
            mapping_extra_instructions=(
                "file1为期初，file2为期末。普通模式下新增方式和新增时间只服务file2。"
                "仅填补未映射字段；已映射字段通过field_review复核。"
            ),
            review_extra_instructions=(
                "样例值和列画像优先于表头；累计折旧不得映射本年或本月折旧；"
                "资产类别必须是描述性类别而非编码。"
            ),
            match_extra_instructions="两侧匹配列数量必须一致，优先稳定的资产或卡片编号。",
            forbidden_columns=forbidden,
        )
        suggestions = [asdict(value) for value in combined.suggestions]
        auto_applied = [
            value for value in suggestions
            if value.get("action") == "fill"
            and float(value.get("confidence") or 0) >= AUTO_APPLY_CONFIDENCE
            and value.get("suggested_column") in (
                columns1 if value.get("file_side") == "file1" else columns2
            )
        ]
        auto_count = len(auto_applied)
        review_count = len(combined.fa_review)
        if auto_count == 0 and review_count == 0:
            review_message = "LLM 复核完成：现有脚本映射无需补充，匹配键已复核。"
        else:
            review_message = (
                f"LLM 复核完成：自动补充 {auto_count} 项，"
                f"人工复核 {review_count} 项。"
            )
        return {
            "enabled": True,
            "passed": True,
            "message": review_message,
            "autoApplied": auto_applied,
            "fieldReviews": [asdict(value) for value in combined.fa_review],
            "matchReview": asdict(combined.match_review) if combined.match_review else None,
            "localProfile": _json_value(local_profile),
        }


def fa_supplement_review(params: dict, **_: Any) -> dict:
    """Review optional addition/disposal lists against the first-step ID shape."""
    from dataclasses import asdict

    with import_path(ROOT / "tools" / "fa_list"):
        from file_handler import FileHandler
        from launcher.llm_client import (
            AUTO_APPLY_CONFIDENCE,
            generate_combined_fa_list_assistance,
            review_supplement_match_key_columns,
        )
        from launcher.llm_settings import is_llm_enabled, load_llm_settings

        if not is_llm_enabled():
            return {
                "enabled": False,
                "passed": True,
                "message": "Hub 尚未启用 LLM，保留当前补充清单映射。",
                "autoApplied": [],
                "fieldReviews": [],
                "matchReview": None,
            }

        addition = dict(params.get("addition") or {})
        disposal = dict(params.get("disposal") or {})
        files: list[dict[str, Any]] = []
        frames: dict[str, Any] = {}
        handler = FileHandler()

        def load(config: dict, side: str):
            path_value = str(config.get("path") or "").strip()
            if not path_value:
                return
            path = _assert_exists(path_value)
            ok, message, frame = handler.load_file(
                str(path),
                config.get("sheet"),
                max(0, int(config.get("headerRow") or 1) - 1),
            )
            if not ok or frame is None:
                raise EngineError("FA_SUPPLEMENT_REVIEW_LOAD_FAILED", str(message))
            frame.columns = [str(column) for column in frame.columns]
            frames[side] = frame
            files.append({
                "file_side": side,
                "headers": list(frame.columns),
                "samples": _fa_llm_samples(frame),
                "column_profiles": _fa_llm_profiles(frame),
            })

        load(addition, "file1")
        load(disposal, "file2")
        if not files:
            raise EngineError("FA_SUPPLEMENT_REQUIRED", "请先选择新增清单或处置清单。")

        current_match = {
            "file1": [str(value) for value in addition.get("keys") or []],
            "file2": [str(value) for value in disposal.get("keys") or []],
        }
        reference_match = {
            "file1": [
                str(value)
                for value in (params.get("endKeys") or params.get("beginKeys") or [])
            ],
            "file2": [
                str(value)
                for value in (params.get("beginKeys") or params.get("endKeys") or [])
            ],
        }
        current_mapping = {
            "match": current_match,
            "addition_method": {"file1": str(addition.get("method") or "")},
            "addition_date": {"file1": str(addition.get("date") or "")},
            "disposal_method": {"file2": str(disposal.get("method") or "")},
            "disposal_date": {"file2": str(disposal.get("date") or "")},
            "disposal_orig": {"file2": str(disposal.get("originalValue") or "")},
            "disposal_dep": {"file2": str(disposal.get("depreciation") or "")},
        }
        roles = [
            ("match", "匹配列/唯一识别码"),
            ("addition_method", "新增方式"),
            ("addition_date", "新增时间"),
            ("disposal_method", "处置方式"),
            ("disposal_date", "处置时间"),
            ("disposal_orig", "处置原值/原值减少"),
            ("disposal_dep", "处置折旧/累计折旧减少"),
        ]
        settings = load_llm_settings()
        combined = generate_combined_fa_list_assistance(
            settings,
            tool_name="FA List",
            role_definitions=[
                {"role": role, "label": label, "description": label}
                for role, label in roles
            ],
            files=files,
            current_mapping=current_mapping,
            current_match=current_match,
            local_profile={
                side: _fa_key_profile(frame, current_match[side])
                for side, frame in frames.items()
            },
            candidate_profiles=[],
            include_match_review=False,
            mapping_extra_instructions=(
                "file1为新增清单，只映射新增方式和新增时间；file2为处置清单，"
                "只映射处置方式、处置时间、处置原值和处置折旧。"
            ),
            review_extra_instructions="补充清单字段以样例值和列画像为准。",
            match_extra_instructions="",
            forbidden_columns={"file1": [], "file2": []},
        )
        match_review = review_supplement_match_key_columns(
            settings,
            tool_name="FA List",
            files=files,
            current_match=current_match,
            reference_match=reference_match,
            extra_instructions="只判断补充清单匹配ID是否完整对齐第一步ID口径。",
        )
        suggestions = [asdict(value) for value in combined.suggestions]
        headers = {
            row["file_side"]: set(row["headers"])
            for row in files
        }
        auto_applied = [
            value for value in suggestions
            if value.get("action") == "fill"
            and float(value.get("confidence") or 0) >= AUTO_APPLY_CONFIDENCE
            and value.get("suggested_column") in headers.get(value.get("file_side"), set())
        ]
        return {
            "enabled": True,
            "passed": True,
            "message": (
                f"补充清单 LLM 复核完成：自动补充 {len(auto_applied)} 项，"
                f"人工复核 {len(combined.fa_review)} 项。"
            ),
            "autoApplied": auto_applied,
            "fieldReviews": [asdict(value) for value in combined.fa_review],
            "matchReview": asdict(match_review),
        }


def _fa_suggest_columns(columns: list[Any], frame: Any = None) -> dict:
    names = [str(value) for value in columns]
    with import_path(ROOT / "tools" / "fa_list"):
        from mapping_rules import pick_category, pick_life, pick_name
    category = pick_category(names, frame)

    def find(*terms: str) -> str | None:
        normalized = [(name, "".join(ch.lower() for ch in name if ch.isalnum())) for name in names]
        wanted = ["".join(ch.lower() for ch in term if ch.isalnum()) for term in terms]
        for name, value in normalized:
            if any(value == term for term in wanted):
                return name
        for name, value in normalized:
            if any(term and term in value for term in wanted):
                return name
        return None

    return {
        "matchKey": find(
            "卡片编号", "资产编号", "固定资产编号", "资产卡片号", "资产编码",
            "coding", "code", "asset id", "asset number",
        ),
        "category": category or find("category"),
        "name": pick_name(names, frame, [category]) or find(
            "资产名称", "固定资产名称", "资产描述", "设备名称", "asset name", "description",
        ),
        "originalValue": find("原值", "资产原值", "期末原值", "original cost", "cost"),
        "depreciation": find("累计折旧", "期末累计折旧", "accumulated depreciation"),
        "startDate": find(
            "入账日期", "开始日期", "开始使用日期", "启用日期", "购置日期",
            "取得日期", "资本化日期", "in service date",
        ),
        "life": pick_life(names) or find("useful life"),
        # Preserve the legacy wizard's fallback: a monetary residual-value
        # column is a valid source and Exporter later normalizes it as
        # residual value / original value.
        "residualRate": find(
            "残值率", "预计残值率", "净残值率", "residual rate",
            "残值", "预计残值", "净残值", "residual value", "salvage value",
        ),
        "currentYearDep": find("本年折旧", "本期折旧", "当年折旧", "current year depreciation"),
        "additionMethod": find(
            "新增方式", "增加方式", "取得方式", "资产来源", "新增来源",
            "变动方式", "变动类型", "增减方式", "增减类型", "增减类别",
            "addition method",
        ),
        "additionDate": find(
            "新增时间", "新增日期", "增加时间", "增加日期", "取得日期", "购置日期",
            "资本化日期", "入账日期", "开始使用日期", "开始使用时间", "变动时间",
            "变动日期", "addition date",
        ),
        "disposalMethod": find("处置方式", "减少方式", "disposal method"),
        "disposalDate": find("处置日期", "减少日期", "disposal date"),
        "disposalOriginal": find("处置原值", "减少原值", "原值减少", "处置成本"),
        "disposalDepreciation": find(
            "处置折旧", "减少折旧", "累计折旧处置", "累计折旧减少",
        ),
    }


def _fa_merge(params: dict, progress: Progress, cancel: threading.Event):
    begin_path = _assert_exists(str(_required(params, "beginPath")))
    end_path = _assert_exists(str(_required(params, "endPath")))
    begin_keys = [str(value) for value in _required(params, "beginKeys")]
    end_keys = [str(value) for value in _required(params, "endKeys")]
    if len(begin_keys) != len(end_keys):
        raise EngineError("FA_KEY_COUNT_MISMATCH", "期初和期末匹配列数量必须一致。")
    if not begin_keys:
        raise EngineError("FA_KEY_REQUIRED", "请至少选择一个匹配列。")
    with import_path(ROOT / "tools" / "fa_list"):
        from file_handler import FileHandler
        from merge_engine import MergeEngine

        handler = FileHandler()
        progress("load", 0, 4, "正在读取期初固定资产清单", "info", None, None)
        header1 = int(params.get("beginHeaderRow") or 1) - 1
        header2 = int(params.get("endHeaderRow") or 1) - 1
        ok1, message1 = handler.set_file1(str(begin_path), params.get("beginSheet"), max(0, header1))
        if cancel.is_set():
            raise EngineError("JOB_CANCELLED", "任务已取消。")
        progress("load", 1, 4, "正在读取期末固定资产清单", "info", None, None)
        ok2, message2 = handler.set_file2(str(end_path), params.get("endSheet"), max(0, header2))
        if not ok1 or not ok2:
            raise EngineError("FA_LOAD_FAILED", message1 if not ok1 else message2)
        if cancel.is_set():
            raise EngineError("JOB_CANCELLED", "任务已取消。")
        progress("match", 2, 4, "正在执行多键全外连接", "info", None, None)
        engine = MergeEngine()
        success, message, merged = engine.perform_full_outer_join(
            handler.file1_df,
            handler.file2_df,
            begin_keys,
            end_keys,
            str(params.get("beginDataType") or "auto"),
            str(params.get("endDataType") or "auto"),
            bool(params.get("removeSpaces", False)),
            bool(params.get("caseSensitive", True)),
            str(params.get("handleDuplicates") or "pivot"),
            params.get("beginOriginalValue"),
            params.get("endOriginalValue"),
            params.get("beginDepreciation"),
            params.get("endDepreciation"),
            params.get("endResidualRate"),
        )
        if not success or merged is None:
            raise EngineError("FA_MATCH_FAILED", str(message))
        begin_name = str(params.get("beginDisplayName") or "期初")
        end_name = str(params.get("endDisplayName") or "期末")
        if "数据来源" in merged.columns:
            merged["数据来源"] = merged["数据来源"].replace(
                {"仅文件1": f"仅{begin_name}", "仅文件2": f"仅{end_name}", "两文件都有": "两文件都有"}
            )
        return handler, engine, merged, begin_keys, end_keys, begin_name, end_name, str(message)


def _fa_apply_supplements(
    merged: Any,
    begin_keys: list[str],
    end_keys: list[str],
    params: dict,
) -> tuple[Any, dict[str, str], Any, Any]:
    """Apply the legacy supplemental-list rules without importing its Tk UI."""
    import re
    import pandas as pd

    def actual_columns(base_columns: list[str], side: int) -> list[str]:
        found: list[str] = []
        for base in base_columns:
            wanted = f"{base}_文件{side}"
            actual = next(
                (str(column) for column in merged.columns if str(column) == wanted or str(column).startswith(wanted + "_")),
                None,
            )
            if actual and actual not in found:
                found.append(actual)
        return found

    def normalize(series: Any):
        def one(value: Any) -> str:
            if value is None or pd.isna(value):
                return ""
            text = re.sub(r"\s+", "", str(value).replace("\u3000", " ").strip()).upper()
            if re.fullmatch(r"[+-]?\d+\.0+", text):
                text = text.split(".", 1)[0]
            elif re.fullmatch(r"[+-]?\d+(?:\.\d+)?E[+-]?\d+", text):
                try:
                    number = float(text)
                    if number.is_integer():
                        text = str(int(number))
                except ValueError:
                    pass
            return text
        return series.apply(one)

    def composite(frame: Any, columns: list[str]):
        valid = [column for column in columns if column in frame.columns]
        if not valid:
            return pd.Series([""] * len(frame), index=frame.index)
        parts = [normalize(frame[column]) for column in valid]
        empty = pd.Series(True, index=frame.index)
        for part in parts:
            empty &= part.eq("")
        result = parts[0].astype(str)
        for part in parts[1:]:
            result = result + "||" + part.astype(str)
        return result.where(~empty, "")

    def merge_text(values: Any) -> str:
        seen: list[str] = []
        for value in values:
            if value is None or pd.isna(value):
                continue
            text = str(value).strip()
            if text and text not in seen:
                seen.append(text)
        return "；".join(seen)

    key1 = composite(merged, actual_columns(begin_keys, 1))
    key2 = composite(merged, actual_columns(end_keys, 2))
    main_key = key1.where(key1.ne(""), key2)
    valid_main_keys = set(main_key[main_key.ne("")].astype(str))
    field_updates: dict[str, str] = {}
    unmatched_add = unmatched_disposal = None

    def load(config: dict):
        path = _assert_exists(str(_required(config, "path")))
        with import_path(ROOT / "tools" / "fa_list"):
            from file_handler import FileHandler
            supplement_handler = FileHandler()
            ok, message = supplement_handler.set_file1(
                str(path), config.get("sheet"), max(0, int(config.get("headerRow") or 1) - 1)
            )
        if not ok:
            raise EngineError("FA_SUPPLEMENT_LOAD_FAILED", str(message))
        return supplement_handler.file1_df

    addition = dict(params.get("additionSupplement") or {})
    if addition.get("path"):
        frame = load(addition)
        keys = [str(value) for value in addition.get("keys") or []]
        work = frame.copy()
        work["__k__"] = composite(work, keys)
        work = work[work["__k__"].ne("")]
        unmatched_add = work[~work["__k__"].isin(valid_main_keys)].drop(columns=["__k__"], errors="ignore")
        rules = {}
        for key in ("method", "date"):
            column = addition.get(key)
            if column in work.columns:
                rules[column] = merge_text
        if rules:
            aggregate = work.groupby("__k__", sort=False).agg(rules)
            if addition.get("method") in aggregate.columns:
                merged["新增方式_辅助_文件2"] = main_key.map(aggregate[addition["method"]])
                field_updates["addition_method_col2"] = "新增方式_辅助_文件2"
            if addition.get("date") in aggregate.columns:
                merged["新增时间_辅助_文件2"] = main_key.map(aggregate[addition["date"]])
                field_updates["addition_date_col2"] = "新增时间_辅助_文件2"

    disposal = dict(params.get("disposalSupplement") or {})
    if disposal.get("path"):
        frame = load(disposal)
        keys = [str(value) for value in disposal.get("keys") or []]
        work = frame.copy()
        work["__k__"] = composite(work, keys)
        work = work[work["__k__"].ne("")]
        unmatched_disposal = work[~work["__k__"].isin(valid_main_keys)].drop(columns=["__k__"], errors="ignore")
        rules = {}
        for key in ("method", "date"):
            column = disposal.get(key)
            if column in work.columns:
                rules[column] = merge_text
        for key in ("originalValue", "depreciation"):
            column = disposal.get(key)
            if column in work.columns:
                rules[column] = lambda values: float(pd.to_numeric(values, errors="coerce").fillna(0).abs().sum())
        if rules:
            aggregate = work.groupby("__k__", sort=False).agg(rules)
            outputs = {
                "method": ("处置方式_辅助_文件1", "disposal_method_col1"),
                "date": ("处置时间_辅助_文件1", "disposal_date_col1"),
                "originalValue": ("处置原值_辅助_文件1", "disposal_orig_col1"),
                "depreciation": ("处置折旧_辅助_文件1", "disposal_dep_col1"),
            }
            for key, (column_name, mapping_name) in outputs.items():
                source = disposal.get(key)
                if source in aggregate.columns:
                    merged[column_name] = main_key.map(aggregate[source])
                    field_updates[mapping_name] = column_name
    return merged, field_updates, unmatched_add, unmatched_disposal


def fa_match(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    _handler, engine, merged, keys1, keys2, name1, name2, message = _fa_merge(params, progress, cancel)
    merged, _field_updates, unmatched_add, unmatched_disposal = _fa_apply_supplements(
        merged, keys1, keys2, params
    )
    if "_merge" in merged.columns:
        counts = merged["_merge"].astype(str).value_counts().to_dict()
        both = int(counts.get("both", 0))
        begin_only = int(counts.get("left_only", 0))
        end_only = int(counts.get("right_only", 0))
    elif "数据来源" in merged.columns:
        counts = merged["数据来源"].astype(str).value_counts().to_dict()
        both = int(counts.get("两文件都有", 0))
        begin_only = int(counts.get(f"仅{name1}", 0))
        end_only = int(counts.get(f"仅{name2}", 0))
    else:
        both = begin_only = end_only = 0
    duplicate_info = engine.get_duplicate_info() or {}
    compact_duplicates = {
        "hasDuplicates": bool(duplicate_info.get("has_duplicates")),
        "duplicateValueCount": int(duplicate_info.get("total_duplicate_values") or 0),
        "duplicateRowCount": int(duplicate_info.get("total_duplicate_rows") or 0),
    }
    stats = {
        "rows": len(merged),
        "both": both,
        "beginOnly": begin_only,
        "endOnly": end_only,
        "duplicates": compact_duplicates,
        "unmatchedAddition": 0 if unmatched_add is None else len(unmatched_add),
        "unmatchedDisposal": 0 if unmatched_disposal is None else len(unmatched_disposal),
    }
    progress("preview", 4, 4, "匹配预览完成", "success", None, None)
    return {
        "engine": "python-fa-kernel",
        "message": str(message),
        "stats": _json_value(stats),
        "columns": [str(column) for column in merged.columns],
        "preview": _json_value(merged.head(50).to_dict(orient="records")),
    }


def fa_export(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    handler, _engine, merged, begin_keys, end_keys, begin_name, end_name, message = _fa_merge(params, progress, cancel)
    merged, supplement_fields, unmatched_add, unmatched_disposal = _fa_apply_supplements(
        merged, begin_keys, end_keys, params
    )
    if cancel.is_set():
        raise EngineError("JOB_CANCELLED", "任务已取消。")
    output_value = params.get("outputPath")
    if output_value:
        output = Path(str(output_value)).expanduser().resolve()
    else:
        source = Path(str(_required(params, "endPath"))).expanduser().resolve()
        output = source.parent / f"FA_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_format = "csv" if output.suffix.lower() == ".csv" else "xlsx"
    if export_format == "xlsx" and output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")

    begin_mapping = dict(params.get("beginMapping") or {})
    end_mapping = dict(params.get("endMapping") or {})
    begin_original = params.get("beginOriginalValue") or begin_mapping.get("originalValue")
    end_original = params.get("endOriginalValue") or end_mapping.get("originalValue")
    begin_dep = params.get("beginDepreciation") or begin_mapping.get("depreciation")
    end_dep = params.get("endDepreciation") or end_mapping.get("depreciation")

    raw_merged_columns = [str(column) for column in merged.columns]

    def raw_side_column(raw: Any, side: int) -> str | None:
        if not raw:
            return None
        wanted = f"{raw}_文件{side}"
        if wanted in raw_merged_columns:
            return wanted
        return next(
            (column for column in raw_merged_columns if column.startswith(wanted + "_")),
            None,
        )

    def format_column(value: Any) -> str | None:
        if not value:
            return None
        return str(value).replace("_文件1", f"_{begin_name}").replace("_文件2", f"_{end_name}")

    role_names = {
        "category": "category_col",
        "name": "name_col",
        "startDate": "date_col",
        "life": "life_col",
        "residualRate": "residual_col",
        "currentYearDep": "current_year_dep_col",
        "additionMethod": "addition_method_col",
        "additionDate": "addition_date_col",
        "disposalMethod": "disposal_method_col",
        "disposalDate": "disposal_date_col",
        "disposalOriginal": "disposal_orig_col",
        "disposalDepreciation": "disposal_dep_col",
    }
    raw_field_mapping: dict[str, Any] = {}
    formatted_field_mapping: dict[str, Any] = {}
    for frontend_name, kernel_name in role_names.items():
        for side, source in ((1, begin_mapping), (2, end_mapping)):
            raw = source.get(frontend_name)
            raw_field_mapping[f"{kernel_name}{side}"] = raw
            formatted_field_mapping[f"{kernel_name}{side}"] = format_column(raw_side_column(raw, side))
    for key, value in supplement_fields.items():
        raw_field_mapping[key] = (
            str(value).replace("_文件1", "").replace("_文件2", "")
            if value
            else None
        )
    formatted_field_mapping.update({
        key: format_column(value) for key, value in supplement_fields.items()
    })

    # The legacy wizard replaces the technical “_文件1/_文件2” suffixes only
    # at export time. Keep matching on raw names, then format the export copy.
    merged = merged.copy()
    merged.columns = [format_column(column) or str(column) for column in merged.columns]

    def fallback_category(side: int) -> str | None:
        suffix = f"_文件{side}"
        candidates = [column for column in raw_merged_columns if column.endswith(suffix)]

        def base(column: str) -> str:
            return column[:-len(suffix)] if column.endswith(suffix) else column

        def numeric(column: str) -> bool:
            return any(
                term in column
                for term in ("原值", "累计折旧", "成本", "净值", "残值", "减值", "折旧", "金额", "价值")
            )

        for predicate in (
            lambda value: base(value) in ("资产大类", "资产类别"),
            lambda value: "资产大类" in value or "资产类别" in value,
            lambda value: any(term in value for term in ("类别", "种类", "大类")),
        ):
            found = next(
                (column for column in candidates if not numeric(column) and predicate(column)),
                None,
            )
            if found:
                return found
        return None

    category1 = formatted_field_mapping.get("category_col1")
    category2 = formatted_field_mapping.get("category_col2")
    pivot_category1 = category1 or format_column(fallback_category(1))
    pivot_category2 = category2 or format_column(fallback_category(2))
    pivot_df = None
    pivot_config = None
    with import_path(ROOT / "tools" / "fa_list"):
        from pivot_engine import PivotEngine
        from exporter import Exporter

        index_fields = [value for value in (pivot_category1, pivot_category2) if value]
        value_fields = [
            value for value in (
                format_column(raw_side_column(begin_original, 1)),
                format_column(raw_side_column(end_original, 2)),
                format_column(raw_side_column(begin_dep, 1)),
                format_column(raw_side_column(end_dep, 2)),
            ) if value
        ]
        if index_fields:
            pivot_engine = PivotEngine()
            ok, _pivot_message, generated = pivot_engine.create_pivot_table(
                merged, index=index_fields, columns=None, values=value_fields or None, aggfunc="sum"
            )
            if ok:
                pivot_df = generated
                pivot_config = pivot_engine.get_pivot_config()

        summary_config = {
            "category_col": category1 or category2,
            "category_col1": category1,
            "category_col2": category2,
            "match_col": format_column(raw_side_column(begin_keys[0], 1)),
            "match_col2": format_column(raw_side_column(end_keys[0], 2)),
            "match_cols": [format_column(raw_side_column(value, 1)) for value in begin_keys],
            "match_cols2": [format_column(raw_side_column(value, 2)) for value in end_keys],
            "original_value_col1": format_column(raw_side_column(begin_original, 1)),
            "original_value_col2": format_column(raw_side_column(end_original, 2)),
            "depreciation_col1": format_column(raw_side_column(begin_dep, 1)),
            "depreciation_col2": format_column(raw_side_column(end_dep, 2)),
            "file1_display_name": begin_name,
            "file2_display_name": end_name,
            "balance_sheet_date": params.get("balanceSheetDate") or datetime.now().strftime("%Y/12/31"),
            "field_mapping": formatted_field_mapping,
            "extended_summary_mode": True,
            "use_supplement_lists": bool(params.get("additionSupplement") or params.get("disposalSupplement")),
            "unmatched_add_df": unmatched_add,
            "unmatched_disp_df": unmatched_disposal,
            "has_unmatched_supplement": bool(
                (unmatched_add is not None and not unmatched_add.empty)
                or (unmatched_disposal is not None and not unmatched_disposal.empty)
            ),
            "source_file1_df": handler.file1_df,
            "source_file2_df": handler.file2_df,
            "source_match_cols1_raw": begin_keys,
            "source_match_cols2_raw": end_keys,
            "source_field_mapping_raw": raw_field_mapping,
            "source_original_value_col1_raw": begin_original,
            "source_original_value_col2_raw": end_original,
            "source_depreciation_col1_raw": begin_dep,
            "source_depreciation_col2_raw": end_dep,
            "pivot_export_config": pivot_config,
        }
        progress("export", 3, 4, "正在生成 FA List、变动清单、汇总与透视表", "info", None, None)
        exporter = Exporter()
        exporter.set_progress_callback(
            lambda current, text="": progress("export", min(99, max(3, int(current))), 100, str(text or "正在导出"), "info", None, None)
        )
        selected = [str(value) for value in params.get("selectedColumns") or []] or None
        success, export_message = exporter.export_dataframe(
            merged,
            str(output),
            selected,
            export_format,
            pivot_df=pivot_df,
            full_df=merged,
            summary_config=summary_config,
        )
    if not success:
        raise EngineError("FA_EXPORT_FAILED", str(export_message))
    if cancel.is_set():
        raise EngineError("JOB_CANCELLED", "任务已取消，但导出文件可能已经生成。")
    progress("completed", 4, 4, "FA List 导出完成", "success", [str(output)], None)
    return {
        "engine": "python-fa-kernel",
        "message": str(message),
        "exportMessage": str(export_message),
        "rows": len(merged),
        "columns": len(merged.columns),
        "outputPaths": [str(output)],
    }


def _roll_forward_config() -> dict:
    return json.loads((ROOT / "modules/audit-roll-forward/subjects_config.json").read_text(encoding="utf-8"))


def roll_forward_catalog(_params: dict, **_: Any) -> dict:
    subjects = _roll_forward_config().get("subjects", {})
    return {
        "version": _roll_forward_config().get("version"),
        "subjects": [
            {
                "code": code,
                "name": item.get("name", code),
                "templateFile": item.get("template_file", ""),
                "priorPatterns": item.get("prior_file_patterns") or [item.get("prior_file_pattern", "")],
                "hasCra": True,
            }
            for code, item in subjects.items()
        ],
    }


def _roll_forward_normalize_match_text(value: Any) -> str:
    return re.sub(r"[\s_\-./\\：:；;（）()\[\]【】]+", "", str(value or "").upper())


def roll_forward_detect_subjects(params: dict, **_: Any) -> dict:
    """Mirror the legacy GUI's subject auto-selection for a prior path."""
    source = Path(str(_required(params, "priorPath"))).expanduser().resolve()
    if not source.exists():
        raise EngineError("PATH_NOT_FOUND", f"找不到上年底稿路径：{source}")
    if source.is_file():
        if source.suffix.lower() != ".xlsx" or source.name.startswith("~$"):
            raise EngineError("ROLL_FORWARD_PRIOR_INVALID", "单文件模式请选择有效的 XLSX 上年底稿。")
        files = [source]
    elif source.is_dir():
        files = [item for item in source.rglob("*.xlsx") if item.is_file() and not item.name.startswith("~$")][:500]
    else:
        raise EngineError("ROLL_FORWARD_PRIOR_INVALID", "上年底稿路径不是文件或目录。")

    catalog = roll_forward_catalog({})["subjects"]
    detected: set[str] = set()
    matched_files: dict[str, list[str]] = {}
    for item in catalog:
        code = str(item["code"])
        name = str(item.get("name") or "")
        variants = {code, code.replace("_", ""), name}
        if code.lower() == "uexp":
            variants.update({"U_EXP", "UEXP"})
        elif code.lower() == "uexpvcvd":
            variants.update({"U_EXPVCVD", "UEXPVCVD", "VCVD", "VC&VD"})
        tokens = [_roll_forward_normalize_match_text(value) for value in variants if value]
        code_token = _roll_forward_normalize_match_text(code)
        for file_path in files:
            filename = file_path.name
            upper_name = filename.upper()
            normalized_name = _roll_forward_normalize_match_text(filename)
            has_code = bool(re.search(rf"(^|[^A-Z0-9]){re.escape(code.upper())}([^A-Z0-9]|$)", upper_name))
            if code.lower() == "uexp":
                is_vcvd = "VC&VD" in upper_name or "VCVD" in normalized_name
                has_code = (not is_vcvd) and (has_code or "U_EXP" in upper_name or "UEXP" in normalized_name)
            elif code.lower() == "uexpvcvd":
                has_code = has_code or "VC&VD" in upper_name or "VCVD" in normalized_name
            has_name = any(token and token != code_token and token in normalized_name for token in tokens)
            if has_code or has_name:
                detected.add(code)
                matched_files.setdefault(code, []).append(str(file_path))

    ordered = [str(item["code"]) for item in catalog if str(item["code"]) in detected]
    return {
        "subjects": ordered,
        "matchedFiles": matched_files,
        "scannedWorkbookCount": len(files),
        "message": (
            f"已根据上年底稿默认识别科目：{', '.join(ordered)}。请复核后再执行。"
            if ordered else "未能从上年底稿路径自动识别科目，请手动选择。"
        ),
    }


def roll_forward_project_export(params: dict, **_: Any) -> dict:
    project = params.get("project")
    if not isinstance(project, dict):
        raise EngineError("INVALID_ARGUMENT", "项目数据格式不正确。")
    output = Path(str(_required(params, "outputPath"))).expanduser().resolve()
    if output.suffix.lower() not in {".auditproj", ".json"}:
        output = output.with_suffix(".auditproj")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.name}.tmp")
    try:
        temporary.write_text(json.dumps(project, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temporary, output)
    finally:
        temporary.unlink(missing_ok=True)
    return {"message": "项目已导出。", "outputPaths": [str(output)]}


def roll_forward_cra_parse(params: dict, **_: Any) -> dict:
    text = str(_required(params, "text"))
    subjects = params.get("subjectCodes") or []
    module = load_file("audit_roll_forward_cra_support", "modules/audit-roll-forward/cra_support.py")
    records = module.parse_cra_paste_text(
        text,
        selected_subjects=subjects,
        cra_header_preference=str(params.get("headerPreference") or ""),
    )
    return {
        "records": _json_value(records),
        "headerOptions": _json_value(module.detect_cra_header_options(text)),
        "writeCount": sum(1 for row in records if row.get("match_status") == "将写入"),
        "issueCount": sum(1 for row in records if row.get("match_status") != "将写入"),
    }


def roll_forward_validate(params: dict, **_: Any) -> dict:
    template = _assert_exists(str(_required(params, "templateDir")), directory=True)
    prior = Path(str(_required(params, "priorDir"))).expanduser().resolve()
    if not prior.exists():
        raise EngineError("PATH_NOT_FOUND", f"找不到上年底稿路径：{prior}")
    if not prior.is_dir() and prior.suffix.lower() != ".xlsx":
        raise EngineError("ROLL_FORWARD_PRIOR_INVALID", "上年底稿必须是目录或 XLSX 文件。")
    output = Path(str(_required(params, "outputDir"))).expanduser().resolve()
    subjects = params.get("subjectCodes") or []
    if isinstance(subjects, str): subjects = [s.strip() for s in subjects.split(",") if s.strip()]
    cfg = _roll_forward_config()
    known = cfg.get("subjects", {})
    missing_subjects = [s for s in subjects if s not in known]
    missing_templates = [known[s].get("template_file") for s in subjects if s in known and not (template / known[s].get("template_file", "")).is_file()]
    date_text = str(params.get("bsDate") or "").strip()
    parsed_date = None
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            parsed_date = datetime.strptime(date_text, date_format)
            break
        except ValueError:
            continue
    date_valid = parsed_date is not None
    prior_year = str(parsed_date.year - 1) if parsed_date else ""
    core = load_file("audit_roll_forward_validate_core", "modules/audit-roll-forward/roll_forward_core.py")
    details = []
    for code in subjects:
        config = known.get(code) or {}
        match = None
        if config and prior_year:
            if prior.is_file():
                match = str(prior) if len(subjects) == 1 else None
            else:
                match = core.find_prior_file(str(prior), code, prior_year, config)
        details.append({
            "code": code,
            "name": config.get("name", code),
            "templatePath": str(template / config.get("template_file", "")) if config else "",
            "templateReady": bool(config and (template / config.get("template_file", "")).is_file()),
            "priorPath": str(match or ""),
            "priorReady": bool(match),
        })
    prior_files = [prior.name] if prior.is_file() else [p.name for p in prior.glob("*.xlsx") if not p.name.startswith("~$")]
    output_parent = output if output.exists() else next((p for p in output.parents if p.exists()), output.parent)
    output_writable = output_parent.is_dir() and os.access(output_parent, os.W_OK)
    company_valid = bool(str(params.get("companyName") or "").strip())
    pmte_text = str(params.get("pmtePath") or "").strip()
    pmte_ready = not pmte_text or Path(pmte_text).expanduser().is_file()
    llm_requested = bool(params.get("llmEnhanced") or params.get("llmWordingRevision"))
    llm_options = params.get("__llmOptions") if isinstance(params.get("__llmOptions"), dict) else {}
    llm_api_type = str(llm_options.get("api_type") or "openai")
    llm_ready = bool(
        not llm_requested
        or (
            llm_options.get("enabled")
            and llm_options.get("api_key")
            and llm_options.get("base_url")
            and llm_options.get("model")
            and llm_api_type == "openai"
        )
    )
    valid = bool(subjects) and company_valid and date_valid and pmte_ready and llm_ready and not missing_subjects and not missing_templates and all(row["priorReady"] for row in details) and output_writable
    return {
        "valid": valid,
        "subjects": subjects,
        "unknownSubjects": missing_subjects,
        "missingTemplates": missing_templates,
        "priorWorkbookCount": len(prior_files),
        "dateValid": date_valid,
        "companyValid": company_valid,
        "pmteReady": pmte_ready,
        "llmRequested": llm_requested,
        "llmReady": llm_ready,
        "llmMessage": (
            "全局 LLM 配置已就绪。"
            if llm_requested and llm_ready
            else "Roll Forward 的 LLM 增强仅支持已配置的 OpenAI 兼容接口，请先在工具箱设置中完成配置。"
            if llm_requested
            else "未启用 LLM 增强。"
        ),
        "outputWritable": output_writable,
        "details": details,
        "outputDir": str(output),
    }


def roll_forward_process(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    check = roll_forward_validate(params)
    if not check["valid"]: raise EngineError("ROLL_FORWARD_INVALID", "模板、科目或上年底稿检查未通过。")
    module = load_file("audit_roll_forward_core", "modules/audit-roll-forward/roll_forward_core.py")
    def callback(current: int, total: int, message: str):
        progress("process", current, total, message, "info", None, None)
    pause_text = str(params.get("__pausePath") or "").strip()
    pause_path = Path(pause_text) if pause_text else None
    def control(event: str, subject_code: str, current: int, total: int):
        if cancel.is_set():
            return "terminate"
        if event == "before_subject" and pause_path is not None and pause_path.exists():
            progress("paused", current - 1, total, f"[{subject_code}] 已暂停，将在继续后处理该科目", "warning", None, None)
        while event == "before_subject" and pause_path is not None and pause_path.exists() and not cancel.is_set():
            time.sleep(0.2)
        return "terminate" if cancel.is_set() else "continue"
    prior_path = Path(str(_required(params, "priorDir"))).expanduser().resolve()
    common = {
        "functional_currency": params.get("functionalCurrency"),
        "accounting_standard": params.get("accountingStandard"),
        "pm_value": params.get("pmValue"),
        "te_value": params.get("teValue"),
        "sad_value": params.get("sadValue"),
        "cra_records": params.get("craRecords") or [],
        "roll_forward_wording": bool(params.get("rollForwardWording", False)),
        "generate_summary": bool(params.get("generateSummary", True)),
        "llm_enhanced": bool(params.get("llmEnhanced", False)),
        "llm_wording_revision": bool(params.get("llmWordingRevision", False)),
        "llm_options": params.get("__llmOptions") or {},
    }
    if prior_path.is_file():
        code = check["subjects"][0]
        config = _roll_forward_config()["subjects"][code]
        result = module.process_single_subject(
            code,
            str(Path(str(_required(params, "templateDir"))) / config["template_file"]),
            str(prior_path),
            str(params.get("pmtePath") or ""),
            str(_required(params, "companyName")),
            str(_required(params, "bsDate")),
            str(_required(params, "outputDir")),
            config,
            progress_callback=lambda message: callback(0, 1, f"[{code}] {message}"),
            **common,
        )
        results = [(code, *result)]
    else:
        results = module.process_multiple_subjects(
            check["subjects"], str(_required(params,"templateDir")), str(prior_path), str(params.get("pmtePath") or ""),
            str(_required(params,"companyName")), str(_required(params,"bsDate")), str(_required(params,"outputDir")),
            progress_callback=callback, control_callback=control, **common,
        )
    outputs = [str(row[3]) for row in results if row[1] and row[3]]
    if cancel.is_set():
        raise EngineError("JOB_CANCELLED", "任务已取消；已完成科目的结果已保留。")
    rows = []
    for code, success, message, output_path, warnings in results:
        rows.append({
            "subjectCode": code,
            "success": bool(success),
            "message": str(message),
            "outputPath": str(output_path or ""),
            "warnings": _json_value(list(warnings or [])),
            "metadata": _json_value(getattr(warnings, "metadata", {})),
        })
    return {"results": rows, "outputPaths": outputs}


def roll_forward_process_companies(params: dict, progress: Progress, cancel: threading.Event) -> dict:
    companies = params.get("companies") or []
    if not isinstance(companies, list) or not companies:
        raise EngineError("INVALID_ARGUMENT", "请至少选择一家公司。")
    all_results = []
    all_outputs = []
    total = len(companies)
    for index, company in enumerate(companies, start=1):
        if cancel.is_set():
            raise EngineError("JOB_CANCELLED", "批量处理已取消；已完成公司的结果已保留。")
        company_params = dict(params)
        company_params.pop("companies", None)
        company_params.update(company if isinstance(company, dict) else {})
        company_name = str(company_params.get("companyName") or f"公司{index}")
        progress("company", index - 1, total, f"[{company_name}] 开始处理", "info", None, None)
        result = roll_forward_process(
            company_params,
            lambda phase, current, subject_total, message, severity, paths, value:
                progress(phase, index - 1, total, f"[{company_name}] {message}", severity, paths, value),
            cancel,
        )
        all_results.append({"companyName": company_name, **result})
        all_outputs.extend(result.get("outputPaths") or [])
        progress("company", index, total, f"[{company_name}] 处理完成", "success", result.get("outputPaths"), None)
    return {"companies": all_results, "outputPaths": all_outputs}


def audipick_projects(params: dict, **_: Any) -> dict:
    return {"projects": [], "projectName": params.get("projectName"), "storage": "tauri-sqlite", "migrationRequired": True}


def unsupported_job(params: dict, **_: Any) -> dict:
    raise EngineError("MIGRATION_PENDING", "该处理动作仍在双轨迁移中；当前 Tauri 版已可检查文件结构，但尚未开放最终导出。", retryable=False)


HANDLERS: dict[str, Callable[..., dict]] = {
    "file_list.scan": file_list_scan, "file_list.export": file_list_export,
    "wp.validate": wp_validate, "wp.generate": wp_generate,
    "confirmation.inspect": confirmation_inspect, "confirmation.process": confirmation_process,
    "fa.inspect": fa_inspect, "fa.supplement_inspect": fa_supplement_inspect,
    "fa.review": fa_review, "fa.supplement_review": fa_supplement_review,
    "fa.match": fa_match, "fa.preview": fa_match, "fa.export": fa_export,
    "roll_forward.catalog": roll_forward_catalog,
    "roll_forward.detect_subjects": roll_forward_detect_subjects,
    "roll_forward.project_export": roll_forward_project_export,
    "roll_forward.cra.parse": roll_forward_cra_parse,
    "roll_forward.validate": roll_forward_validate,
    "roll_forward.process": roll_forward_process,
    "roll_forward.process_companies": roll_forward_process_companies,
    "audipick.projects": audipick_projects, "audipick.import": unsupported_job, "audipick.ocr": unsupported_job,
    "audipick.extract": unsupported_job, "audipick.workpaper": unsupported_job, "audipick.pdf": unsupported_job,
}


def dispatch(method: str, params: dict, progress: Progress | None = None, cancel: threading.Event | None = None) -> dict:
    handler = HANDLERS.get(method)
    if handler is None: raise EngineError("METHOD_NOT_ALLOWED", f"未授权的引擎方法：{method}")
    noop: Progress = lambda *_args: None
    return handler(params, progress=progress or noop, cancel=cancel or threading.Event())
