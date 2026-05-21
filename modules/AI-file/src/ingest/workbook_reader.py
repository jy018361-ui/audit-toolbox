from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from ingest.field_mapping import check_required_fields, map_headers
from ingest.header_detection import scan_rows_for_headers
from ingest.models import SheetClassification, SheetKind, WorkbookDiagnostic
from ingest.sheet_classifier import classify_sheet, score_by_name


def _read_sheet_rows(ws, max_rows: int = 100) -> list[tuple[Any, ...]]:
    max_row = min(ws.max_row or 0, max_rows)
    max_col = min(ws.max_column or 0, 100)
    if max_row == 0:
        return []
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        rows.append(row)
    return rows


def diagnose_workbook(path: str | Path, max_rows: int = 100) -> WorkbookDiagnostic:
    path = Path(path)
    diag = WorkbookDiagnostic(path=str(path))

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        diag.errors.append(f"{type(e).__name__}: {e}")
        return diag

    for ws in wb.worksheets:
        title = ws.title
        rows = _read_sheet_rows(ws, max_rows=max_rows)

        kind, confidence, name_score, content_score, name_hint, header_row = classify_sheet(
            title, rows
        )

        notes: list[str] = []
        if kind != SheetKind.SKIP:
            name_kind, ns, _ = score_by_name(title)
            if name_kind != kind and ns >= 0.7 and kind != SheetKind.UNCLASSIFIED:
                notes.append(
                    f"name_content_mismatch: name->{name_kind.value}, selected->{kind.value}"
                )

        mapped_fields = []
        unmapped: list[str] = []
        missing_req: list[str] = []
        missing_rec: list[str] = []

        if kind not in (SheetKind.SKIP, SheetKind.UNCLASSIFIED, SheetKind.SUMMARY, SheetKind.LEAD):
            hr, cells, unmapped = scan_rows_for_headers(rows, sheet_kind=kind)
            if hr:
                header_row = hr
            if cells:
                mapped_fields, unmapped = map_headers(cells, sheet_kind=kind)
                missing_req, missing_rec = check_required_fields(mapped_fields, kind)

        diag.sheets.append(
            SheetClassification(
                sheet_name=title,
                kind=kind,
                confidence=round(confidence, 3),
                name_score=round(name_score, 3),
                content_score=round(content_score, 3),
                name_hint=name_hint,
                header_row=header_row,
                mapped_fields=mapped_fields,
                missing_required=missing_req,
                missing_recommended=missing_rec,
                unmapped_headers=unmapped[:24],
                notes=notes,
            )
        )

    wb.close()
    return diag
