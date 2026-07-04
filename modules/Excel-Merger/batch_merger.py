import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import threading
import warnings
import csv
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from launcher.ui_theme import (
    add_standard_button,
    apply_app_theme,
    create_button_group,
    create_section,
    create_standard_layout,
    fit_window_to_screen,
)

# 忽略警告
warnings.filterwarnings("ignore")

SUPPORTED_TABLE_EXTENSIONS = ('.xlsx', '.xls', '.csv', '.txt')
EXCEL_MAX_ROWS = 1048576
STREAM_XLSX_EXTENSIONS = ('.xlsx', '.xlsm')
HYPERLINK_WARNING_ROW_THRESHOLD = 50000
HYPERLINK_WARNING_SIZE_THRESHOLD = 40 * 1024 * 1024
EXACT_ESTIMATE_MAX_FILE_SIZE = 20 * 1024 * 1024
ONE_WORKBOOK_FALLBACK_REWRITE_LIMIT = 100 * 1024 * 1024


class MergeCancelled(Exception):
    pass

try:
    import windnd
    HAS_WINDND = True
except ImportError:
    windnd = None
    HAS_WINDND = False

# 尝试导入 xlsxwriter
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

try:
    import polars as pl
    import fastexcel
    HAS_POLARS = True
except ImportError:
    pl = None
    fastexcel = None
    HAS_POLARS = False

try:
    from python_calamine import load_workbook as load_calamine_workbook
    HAS_CALAMINE = True
except ImportError:
    load_calamine_workbook = None
    HAS_CALAMINE = False


def get_xlsx_sheet_names_from_zip_file(file_path):
    with zipfile.ZipFile(file_path) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        sheet_names = []
        for sheet in workbook_root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
            name = sheet.attrib.get("name")
            state = sheet.attrib.get("state", "visible")
            if name and state == "visible":
                sheet_names.append(name)
        return sheet_names


def get_sheet_names_lightweight_file(file_path):
    fp_lower = file_path.lower()
    if fp_lower.endswith(STREAM_XLSX_EXTENSIONS):
        return get_xlsx_sheet_names_from_zip_file(file_path)
    xls = pd.ExcelFile(file_path)
    return xls.sheet_names

# ==========================================
# V2.3: 页签选择弹窗 (保持 V2.2 逻辑不变)
# ==========================================
class SheetSelectDialog(tk.Toplevel):
    def __init__(self, parent, file_list, default_file_index):
        super().__init__(parent)
        apply_app_theme(self)
        self.title("检测到多Sheet - 请定义合并范围")
        fit_window_to_screen(self, 520, 620, 460, 420)
        self.transient(parent)
        self.grab_set()
        
        self.result_action = "cancel"
        self.selected_sheets = []
        self.file_list = file_list
        
        # 底部按钮
        f_bot = tk.Frame(self, pady=10, bg="#f0f0f0")
        f_bot.pack(side="bottom", fill="x")
        ttk.Button(f_bot, text="确定合并", command=self.on_confirm).pack(side="right", padx=20)
        ttk.Button(f_bot, text="取消", command=self.destroy).pack(side="right")
        
        # 1. 基准文件
        f_ref = tk.LabelFrame(self, text="1. 选择基准文件 (用于查看页签)", padx=10, pady=5)
        f_ref.pack(side="top", fill="x", padx=10, pady=5)
        
        self.excel_files = [f for f in file_list if f.lower().endswith(('.xlsx', '.xls'))]
        self.excel_basenames = [os.path.basename(f) for f in self.excel_files]
        
        self.cb_files = ttk.Combobox(f_ref, values=self.excel_basenames, state="readonly")
        self.cb_files.pack(fill="x", pady=5)
        self.cb_files.bind("<<ComboboxSelected>>", self.on_file_change)
        
        default_path = file_list[default_file_index]
        if default_path in self.excel_files:
            self.cb_files.current(self.excel_files.index(default_path))
        elif self.excel_files:
            self.cb_files.current(0)
            
        # 2. 合并模式
        f_mode = tk.LabelFrame(self, text="2. 合并逻辑", padx=10, pady=5)
        f_mode.pack(side="top", fill="x", padx=10, pady=5)
        
        self.var_mode = tk.StringVar(value="match")
        r1 = ttk.Radiobutton(f_mode, text="A. 按名称匹配 (勾选下方Sheet)", variable=self.var_mode, value="match", command=self.toggle_list)
        r1.pack(anchor="w")
        tk.Label(f_mode, text="   (仅提取所有文件中与勾选名称一致的Sheet)", fg="gray", font=("size", 8)).pack(anchor="w", pady=(0,5))
        
        r2 = ttk.Radiobutton(f_mode, text="B. 合并所有Sheet (无差别堆叠)", variable=self.var_mode, value="all", command=self.toggle_list)
        r2.pack(anchor="w")
        
        # 3. 列表区
        self.f_list = tk.LabelFrame(self, text="3. 请勾选目标Sheet (可多选)")
        self.f_list.pack(side="top", fill="both", expand=True, padx=10, pady=10)
        
        f_tool = tk.Frame(self.f_list)
        f_tool.pack(fill="x")
        ttk.Button(f_tool, text="全选", command=lambda: self.set_all(True), width=6).pack(side="left")
        ttk.Button(f_tool, text="全不选", command=lambda: self.set_all(False), width=6).pack(side="left")

        list_body = tk.Frame(self.f_list)
        list_body.pack(fill="both", expand=True)
        list_body.grid_rowconfigure(0, weight=1)
        list_body.grid_columnconfigure(0, weight=1)

        self.cvs = tk.Canvas(list_body, highlightthickness=0)
        sb = ttk.Scrollbar(list_body, orient="vertical", command=self.cvs.yview)
        self.frm_inner = tk.Frame(self.cvs)
        self.inner_window = self.cvs.create_window((0,0), window=self.frm_inner, anchor="nw")
        self.cvs.configure(yscrollcommand=sb.set)
        self.cvs.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self.frm_inner.bind("<Configure>", lambda e: self.cvs.configure(scrollregion=self.cvs.bbox("all")))
        self.cvs.bind("<Configure>", self._sync_sheet_list_width)
        
        self.vars = {}
        self.refresh_sheet_list(self.excel_files[self.cb_files.current()])

    def on_file_change(self, event):
        idx = self.cb_files.current()
        if idx >= 0: self.refresh_sheet_list(self.excel_files[idx])

    def refresh_sheet_list(self, filepath):
        for w in self.frm_inner.winfo_children(): w.destroy()
        self.vars = {}
        try:
            for sn in get_sheet_names_lightweight_file(filepath):
                v = tk.BooleanVar(value=True)
                self.vars[sn] = v
                tk.Checkbutton(self.frm_inner, text=sn, variable=v, anchor="w").pack(fill="x", padx=5)
            self.toggle_list()
        except Exception as e:
            tk.Label(self.frm_inner, text=f"读取失败: {e}", fg="red").pack()

    def _sync_sheet_list_width(self, event):
        self.cvs.itemconfigure(self.inner_window, width=event.width)

    def toggle_list(self):
        state = "normal" if self.var_mode.get() == "match" else "disabled"
        for child in self.frm_inner.winfo_children(): child.configure(state=state)

    def set_all(self, val):
        for v in self.vars.values(): v.set(val)

    def on_confirm(self):
        if self.var_mode.get() == "all":
            self.result_action = "merge_all"
        else:
            self.result_action = "match_selected"
            self.selected_sheets = [k for k,v in self.vars.items() if v.get()]
            if not self.selected_sheets:
                return messagebox.showwarning("提示", "请至少勾选一个Sheet")
        self.destroy()

# ==========================================
# 主程序 V2.3 (增加目录索引功能)
# ==========================================
class BatchMergeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel/CSV 批量合并工具 By CSDC      !!!大文件合并建议选CSV格式!!!")
        apply_app_theme(self.root)
        fit_window_to_screen(self.root, 900, 680, 760, 560)

        self.file_list = []
        self.read_warnings = []
        self.var_mode = tk.StringVar(value="one_sheet")
        self.var_direction = tk.StringVar(value="vertical")
        self.var_add_hyperlinks = tk.BooleanVar(value=True)
        self.is_processing = False
        self.cancel_requested = False
        self.btn_start = None
        
        self.setup_ui()
        self.update_ui_state()

    def setup_ui(self):
        _header, body, footer = create_standard_layout(
            self.root,
            "Excel/CSV 批量合并",
            "添加文件、配置合并规则，然后执行合并",
        )
        body.rowconfigure(0, weight=1)
        body.rowconfigure(1, weight=0)
        body.columnconfigure(0, weight=1)

        frame_top = create_section(body, "1. 文件源", row=0, pady=(0, 10))
        frame_top.columnconfigure(0, weight=1)
        frame_top.columnconfigure(1, weight=0)
        frame_top.rowconfigure(0, weight=1)
        frame_top.rowconfigure(1, weight=0)

        f_list = ttk.Frame(frame_top)
        f_list.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        self.lb_files = tk.Listbox(f_list, selectmode="extended", height=10, bg="#f9f9f9")
        sb = ttk.Scrollbar(f_list, orient="vertical", command=self.lb_files.yview)
        self.lb_files.configure(yscrollcommand=sb.set)
        self.lb_files.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        f_btns = ttk.Frame(frame_top)
        f_btns.grid(row=0, column=1, sticky="n")
        ttk.Button(f_btns, text="添加文件", command=self.add_files, width=14).pack(fill="x", pady=3)
        ttk.Button(f_btns, text="扫描文件夹", command=self.add_folder, width=14).pack(fill="x", pady=3)
        ttk.Separator(f_btns, orient="horizontal").pack(fill="x", pady=8)
        ttk.Button(f_btns, text="移除选中", command=self.remove_files, width=14).pack(fill="x", pady=3)
        ttk.Button(f_btns, text="清空列表", command=self.clear_files, width=14).pack(fill="x", pady=3)

        self.lbl_status = ttk.Label(frame_top, text="待处理 0 个文件", style="Muted.TLabel")
        self.lbl_status.grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))
        self.setup_drag_drop(self.root, frame_top, f_list, self.lb_files)
        if HAS_WINDND:
            self.lbl_status.config(text="待处理 0 个文件（可拖拽文件/文件夹到文件源区域）")

        frame_mid = create_section(body, "2. 合并规则", row=1)

        f_mode = ttk.Frame(frame_mid)
        f_mode.pack(fill="x", pady=2)
        ttk.Label(f_mode, text="输出目标:", width=10).pack(side="left")
        ttk.Radiobutton(f_mode, text="合并成一张大表 (One Sheet)", variable=self.var_mode, value="one_sheet", command=self.update_ui_state).pack(side="left", padx=(0, 18))
        ttk.Radiobutton(f_mode, text="合并成一个工作簿 (多 Sheet)", variable=self.var_mode, value="one_workbook", command=self.update_ui_state).pack(side="left")

        ttk.Separator(frame_mid, orient="horizontal").pack(fill="x", pady=10)

        self.f_opts = ttk.Frame(frame_mid)
        self.f_opts.pack(fill="x")
        ttk.Label(self.f_opts, text="拼接方向:", width=10).pack(side="left")
        self.rb_v = ttk.Radiobutton(self.f_opts, text="纵向堆叠 (上下拼)", variable=self.var_direction, value="vertical", command=self.update_ui_state)
        self.rb_v.pack(side="left", padx=(0, 18))
        self.rb_h = ttk.Radiobutton(self.f_opts, text="横向拼接 (左右拼)", variable=self.var_direction, value="horizontal", command=self.update_ui_state)
        self.rb_h.pack(side="left")

        f_link = ttk.Frame(frame_mid)
        f_link.pack(fill="x", pady=(8, 0))
        ttk.Label(f_link, text="来源追溯:", width=10).pack(side="left")
        self.cb_hyperlinks = ttk.Checkbutton(
            f_link,
            text="加入源文件超链接",
            variable=self.var_add_hyperlinks,
        )
        self.cb_hyperlinks.pack(side="left", padx=(0, 12))
        ttk.Label(
            f_link,
            text="勾选后可点击来源文件跳转原文件，但大文件会明显降低导出速度。",
            style="Muted.TLabel",
        ).pack(side="left")

        self.pb = ttk.Progressbar(footer, mode="indeterminate")
        self.pb.pack(side="left", fill="x", expand=True, padx=(0, 12), pady=2)
        btn_group = create_button_group(footer)
        self.btn_start = add_standard_button(btn_group, "开始合并", self.on_start_stop_clicked)

    def on_start_stop_clicked(self):
        if self.is_processing:
            self.request_cancel()
        else:
            self.prepare_and_start()

    def update_ui_state(self):
        mode = self.var_mode.get()
        children = self.f_opts.winfo_children()
        if mode == "one_workbook":
            for c in children: 
                try: c.configure(state="disabled")
                except: pass
        else:
            for c in children: 
                try: c.configure(state="normal")
                except: pass

    # --- 文件操作 ---
    def setup_drag_drop(self, *widgets):
        if not HAS_WINDND:
            return
        for widget in widgets:
            try:
                windnd.hook_dropfiles(widget, func=self.on_drop_files, force_unicode=True)
            except Exception:
                pass

    def on_drop_files(self, files):
        paths = []
        for item in files:
            if isinstance(item, bytes):
                try:
                    paths.append(item.decode("gbk"))
                except UnicodeDecodeError:
                    paths.append(item.decode("utf-8", errors="ignore"))
            else:
                paths.append(str(item))
        self.root.after(0, lambda: self.add_paths(paths))

    def add_paths(self, paths):
        added = 0
        for path in paths:
            if not path:
                continue
            path = os.path.normpath(path)
            if os.path.isdir(path):
                for root, _, files in os.walk(path):
                    for name in files:
                        full_path = os.path.join(root, name)
                        if self.add_file_path(full_path):
                            added += 1
            elif self.add_file_path(path):
                added += 1
        self.lbl_status.config(text=f"待处理: {len(self.file_list)} 个文件")
        if paths and added == 0:
            messagebox.showinfo("提示", "未发现支持的表格文件（支持 xlsx/xls/csv/txt）")

    def add_file_path(self, file_path):
        if not file_path.lower().endswith(SUPPORTED_TABLE_EXTENSIONS):
            return False
        if file_path in self.file_list:
            return False
        self.file_list.append(file_path)
        self.lb_files.insert(tk.END, os.path.basename(file_path))
        return True

    def add_files(self):
        files = filedialog.askopenfilenames(filetypes=[("表格文件", "*.xlsx *.xls *.csv *.txt")])
        self.add_paths(files)

    def add_folder(self):
        folder = filedialog.askdirectory()
        if not folder: return
        self.add_paths([folder])

    def remove_files(self):
        indices = list(self.lb_files.curselection())
        indices.reverse()
        for i in indices:
            self.lb_files.delete(i)
            del self.file_list[i]
        self.lbl_status.config(text=f"待处理: {len(self.file_list)} 个文件")

    def clear_files(self):
        self.file_list = []
        self.lb_files.delete(0, tk.END)
        self.lbl_status.config(text=f"待处理: 0 个文件")
    
    def move_item(self, direction):
        sel = self.lb_files.curselection()
        if not sel: return
        idx = sel[0]
        new_idx = idx + direction
        if 0 <= new_idx < len(self.file_list):
            val = self.file_list.pop(idx)
            self.file_list.insert(new_idx, val)
            self.lb_files.delete(0, tk.END)
            for f in self.file_list: self.lb_files.insert(tk.END, os.path.basename(f))
            self.lb_files.selection_set(new_idx)

    # --- 核心逻辑 ---
    def prepare_and_start(self):
        if self.is_processing:
            return
        if not self.file_list:
            return messagebox.showwarning("提示", "请先添加需要合并的文件！")
        self.cancel_requested = False
        
        trigger_index = -1
        for i, f in enumerate(self.file_list):
            if f.lower().endswith(('.xlsx', '.xls')):
                try:
                    if self.get_sheet_names_lightweight(f):
                        trigger_index = i
                        break
                except: continue
        
        sheet_config = {"action": "default", "targets": []}
        if trigger_index != -1:
            dlg = SheetSelectDialog(self.root, self.file_list, trigger_index)
            self.root.wait_window(dlg)
            if dlg.result_action == "cancel": return
            sheet_config["action"] = dlg.result_action
            sheet_config["targets"] = dlg.selected_sheets
        
        default_name = f"Excel合并结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")],
            initialfile=default_name
        )
        if not save_path: return

        if not self._confirm_hyperlink_choice(save_path, sheet_config):
            return

        self._set_processing_state(True, f"正在合并 {len(self.file_list)} 个文件，请稍候...")
        self.pb.start(10)
        threading.Thread(target=self.run_process, args=(save_path, sheet_config), daemon=True).start()

    def _format_size(self, size_bytes):
        if size_bytes >= 1024 * 1024 * 1024:
            return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"
        if size_bytes >= 1024 * 1024:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
        if size_bytes >= 1024:
            return f"{size_bytes / 1024:.1f} KB"
        return f"{size_bytes} B"

    def _format_duration_range(self, low_seconds, high_seconds):
        low_seconds = max(10, int(low_seconds))
        high_seconds = max(low_seconds + 10, int(high_seconds))

        def fmt(seconds):
            if seconds < 60:
                return f"{seconds} 秒"
            minutes = seconds / 60
            if minutes < 60:
                return f"{minutes:.0f} 分钟"
            hours = minutes / 60
            return f"{hours:.1f} 小时"

        return f"约 {fmt(low_seconds)} - {fmt(high_seconds)}"

    def _estimate_duration_text(self, save_path, scale, include_hyperlinks=True):
        input_mb = max(scale.get("input_size", 0) / (1024 * 1024), 0.1)
        output_mb = max(scale.get("output_size", 0) / (1024 * 1024), 0.1)
        rows = scale.get("rows") or 0
        mode = self.var_mode.get()
        direction = self.var_direction.get()
        is_xlsx = save_path.lower().endswith(".xlsx")

        if mode == "one_workbook":
            # Excel COM 原样复制主要受 Excel 打开/复制大工作表影响。
            low = input_mb * 0.8
            high = input_mb * 2.5
        elif save_path.lower().endswith(".csv"):
            # CSV 省掉 xlsx 打包写入，通常明显更快。
            low = input_mb * 0.5 + output_mb * 0.05
            high = input_mb * 1.5 + output_mb * 0.15
        elif is_xlsx and direction == "vertical":
            low = input_mb * 0.9 + output_mb * 0.5
            high = input_mb * 2.2 + output_mb * 1.2
        else:
            low = input_mb * 0.7 + output_mb * 0.4
            high = input_mb * 1.8 + output_mb * 1.0

        if include_hyperlinks and is_xlsx and self.var_add_hyperlinks.get():
            if rows:
                low += rows / 4500
                high += rows / 1200
            else:
                low += input_mb * 0.4
                high += input_mb * 1.0

        return self._format_duration_range(low, high)

    def _get_total_input_size(self):
        total = 0
        for fp in self.file_list:
            try:
                total += os.path.getsize(fp)
            except OSError:
                pass
        return total

    def _post_status(self, status_text):
        self.root.after(0, lambda text=status_text: self.lbl_status.config(text=text))

    def _check_cancelled(self):
        if self.cancel_requested:
            raise MergeCancelled("用户已停止本次合并。")

    def request_cancel(self):
        if not self.is_processing or self.cancel_requested:
            return
        should_stop = messagebox.askyesno(
            "确认停止",
            "当前合并正在执行。\n\n确认停止后，正在写入的结果文件可能不完整，需要重新导出。\n\n是否确认停止？",
            parent=self.root,
        )
        if not should_stop:
            return
        self.cancel_requested = True
        self._set_processing_state(True, "正在停止，请稍候... 当前文件处理到安全点后会退出。")

    def _ask_hyperlink_choice(self, title, message):
        try:
            self.root.lift()
            self.root.focus_force()
            self.root.attributes("-topmost", True)
            self.root.after(200, lambda: self.root.attributes("-topmost", False))
        except Exception:
            pass

        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        dialog.result = None

        body = ttk.Frame(dialog, padding=(22, 20, 22, 12))
        body.pack(fill="both", expand=True)
        ttk.Label(body, text="?", width=3, anchor="center", font=("Arial", 28, "bold"), foreground="#1686d9").pack(side="left", anchor="n", padx=(0, 18))
        ttk.Label(body, text=message, justify="left", wraplength=520).pack(side="left", fill="both", expand=True)

        footer = ttk.Frame(dialog, padding=(12, 12, 12, 14))
        footer.pack(fill="x")

        def choose(value):
            dialog.result = value
            dialog.destroy()

        ttk.Button(footer, text="关闭超链接并继续", command=lambda: choose(True), width=18).pack(side="right", padx=(8, 0))
        ttk.Button(footer, text="保留超链接继续", command=lambda: choose(False), width=18).pack(side="right")
        dialog.protocol("WM_DELETE_WINDOW", lambda: choose(False))

        dialog.update_idletasks()
        x = self.root.winfo_rootx() + max((self.root.winfo_width() - dialog.winfo_width()) // 2, 0)
        y = self.root.winfo_rooty() + max((self.root.winfo_height() - dialog.winfo_height()) // 2, 0)
        dialog.geometry(f"+{x}+{y}")
        dialog.wait_window()
        return bool(dialog.result)

    def _estimate_text_rows(self, file_path):
        try:
            if os.path.getsize(file_path) > EXACT_ESTIMATE_MAX_FILE_SIZE:
                return 0
        except OSError:
            return 0
        try:
            with open(file_path, "rb") as f:
                return sum(chunk.count(b"\n") for chunk in iter(lambda: f.read(1024 * 1024), b""))
        except Exception:
            return 0

    def _estimate_xlsx_rows(self, file_path, sheet_config):
        try:
            if os.path.getsize(file_path) > EXACT_ESTIMATE_MAX_FILE_SIZE:
                return 0
        except OSError:
            return 0
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
            try:
                target_sheets = self.get_target_sheet_names_for_openpyxl(wb, sheet_config)
                return sum(max((wb[name].max_row or 0) - 1, 0) for name in target_sheets)
            finally:
                wb.close()
        except Exception:
            return 0

    def get_xlsx_sheet_names_from_zip(self, file_path):
        return get_xlsx_sheet_names_from_zip_file(file_path)

    def get_sheet_names_lightweight(self, file_path):
        return get_sheet_names_lightweight_file(file_path)

    def _estimate_selected_sheet_count(self, file_path, sheet_config):
        if file_path.lower().endswith(('.csv', '.txt')):
            return 1
        try:
            all_sheets = self.get_sheet_names_lightweight(file_path)
            action = sheet_config.get("action", "default")
            if action == "match_selected":
                targets = set(sheet_config.get("targets", []))
                return len([name for name in all_sheets if name in targets])
            if action == "merge_all":
                return len(all_sheets)
            return 1 if all_sheets else 0
        except Exception:
            return 0

    def _estimate_export_scale(self, sheet_config):
        estimated_rows = 0
        input_size = 0
        selected_sheet_count = 0
        for fp in self.file_list:
            try:
                input_size += os.path.getsize(fp)
            except OSError:
                pass
            selected_sheet_count += self._estimate_selected_sheet_count(fp, sheet_config)
            fp_lower = fp.lower()
            if fp_lower.endswith(('.csv', '.txt')):
                estimated_rows += self._estimate_text_rows(fp)
            elif fp_lower.endswith(STREAM_XLSX_EXTENSIONS):
                estimated_rows += self._estimate_xlsx_rows(fp, sheet_config)

        if self.var_mode.get() == "one_sheet" and self.var_direction.get() == "horizontal":
            estimated_hyperlinks = len(self.file_list)
        elif self.var_mode.get() == "one_sheet":
            estimated_hyperlinks = estimated_rows if estimated_rows else None
        else:
            estimated_hyperlinks = selected_sheet_count

        estimated_output_size = int(max(input_size * 1.15, estimated_rows * 120))
        return {
            "rows": estimated_rows,
            "input_size": input_size,
            "output_size": estimated_output_size,
            "hyperlinks": estimated_hyperlinks,
            "sheets": selected_sheet_count,
        }

    def _confirm_hyperlink_choice(self, save_path, sheet_config):
        if not self.var_add_hyperlinks.get():
            return True
        if not save_path.lower().endswith(".xlsx"):
            return True

        scale = self._estimate_export_scale(sheet_config)
        hyperlinks = scale["hyperlinks"]
        hyperlink_count_is_large = (
            hyperlinks is not None
            and hyperlinks >= HYPERLINK_WARNING_ROW_THRESHOLD
        )
        if (
            not hyperlink_count_is_large
            and scale["input_size"] < HYPERLINK_WARNING_SIZE_THRESHOLD
            and scale["output_size"] < HYPERLINK_WARNING_SIZE_THRESHOLD
        ):
            return True

        rows_text = f"{scale['rows']:,}" if scale["rows"] else "无法准确预估"
        links_text = f"{hyperlinks:,}" if hyperlinks is not None else "无法准确预估"
        duration_text = self._estimate_duration_text(save_path, scale, include_hyperlinks=True)
        no_link_duration_text = self._estimate_duration_text(save_path, scale, include_hyperlinks=False)
        message = (
            "预计本次导出规模较大。\n\n"
            f"预计数据行数：{rows_text}\n"
            f"输入文件总大小：约 {self._format_size(scale['input_size'])}\n"
            f"预计导出文件大小：约 {self._format_size(scale['output_size'])}\n"
            f"预计写入超链接数量：约 {links_text}\n\n"
            f"预计耗时：{duration_text}\n"
            f"关闭超链接后预计：{no_link_duration_text}\n\n"
            "保留源文件超链接会方便追溯，但会明显拖慢 .xlsx 导出。\n\n"
            "请选择本次导出的超链接策略。"
        )
        disable_links = self._ask_hyperlink_choice("大文件导出提示", message)
        if disable_links:
            self.var_add_hyperlinks.set(False)
        return True

    def _set_processing_state(self, processing, status_text=None):
        self.is_processing = processing
        if self.btn_start is not None:
            self.btn_start.configure(
                state="normal",
                text=("停止执行" if processing else "开始合并"),
            )
        if status_text:
            self.lbl_status.config(text=status_text)

    def _reset_read_warnings(self):
        self.read_warnings = []

    def _record_read_warning(self, file_path, sheet_name=None, reason="读取失败"):
        file_name = os.path.basename(str(file_path))
        sheet_part = f" / {sheet_name}" if sheet_name else ""
        message = f"{file_name}{sheet_part}: {reason}"
        if message not in self.read_warnings:
            self.read_warnings.append(message)

    def _build_read_warning_message(self, extra_warning=None):
        parts = []
        if extra_warning:
            parts.append(str(extra_warning))
        if self.read_warnings:
            shown = self.read_warnings[:12]
            warning = "以下文件或 Sheet 未成功合并，请核对源文件：\n" + "\n".join(f"- {item}" for item in shown)
            if len(self.read_warnings) > len(shown):
                warning += f"\n- 另有 {len(self.read_warnings) - len(shown)} 项未显示"
            parts.append(warning)
        return "\n\n".join(parts)

    def _notify_success(self, save_path, extra_warning=None):
        warning = self._build_read_warning_message(extra_warning)
        if warning:
            self.on_success_with_warning(save_path, warning)
        else:
            self.on_success(save_path)

    def get_target_sheet_names_for_com(self, workbook, sheet_config):
        all_sheets = [workbook.Worksheets(i).Name for i in range(1, workbook.Worksheets.Count + 1)]
        action = sheet_config.get("action", "default")
        if action == "match_selected":
            targets = set(sheet_config.get("targets", []))
            return [name for name in all_sheets if name in targets]
        if action == "merge_all":
            return all_sheets
        return all_sheets[:1]

    def make_unique_com_sheet_name(self, workbook, preferred):
        existing = {workbook.Worksheets(i).Name for i in range(1, workbook.Worksheets.Count + 1)}
        base = self.clean_sheet_name(preferred) or "Sheet"
        if base not in existing:
            return base
        for n in range(1, 10000):
            suffix = f"_{n}"
            candidate = self.clean_sheet_name(base[:31 - len(suffix)] + suffix)
            if candidate not in existing:
                return candidate
        raise Exception(f"无法生成不重复的Sheet名称: {preferred}")

    def add_reference_sheet_com(self, workbook, toc_data):
        if not toc_data:
            return
        ref_name = self.make_unique_com_sheet_name(workbook, "Reference")
        ws = workbook.Worksheets.Add(Before=workbook.Worksheets(1))
        ws.Name = ref_name
        ws.Cells(1, 1).Value = "Source File Name"
        ws.Cells(1, 2).Value = "Target Sheet Link"
        ws.Range("A1:B1").Font.Bold = True
        ws.Columns(1).ColumnWidth = 40
        ws.Columns(2).ColumnWidth = 40
        for row, item in enumerate(toc_data, start=2):
            target_sheet = item["Target Sheet"]
            ws.Cells(row, 1).Value = item["Source File"]
            if self.var_add_hyperlinks.get():
                ws.Hyperlinks.Add(
                    Anchor=ws.Cells(row, 2),
                    Address="",
                    SubAddress=f"'{target_sheet}'!A1",
                    TextToDisplay=target_sheet,
                )
            else:
                ws.Cells(row, 2).Value = target_sheet
        ws.Activate()

    def copy_workbook_sheets_with_com(self, save_path, sheet_config):
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            raise Exception(
                "缺少 Excel 直接复制所需的 pywin32 依赖。\n"
                "请先安装 pywin32，或使用 One Sheet/CSV 普通合并模式。\n"
                "安装命令：pip install pywin32"
            ) from exc

        pythoncom.CoInitialize()
        excel = None
        dest_wb = None
        src_wbs = []
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.EnableEvents = False

            dest_wb = excel.Workbooks.Add(1)
            placeholder = dest_wb.Worksheets(1)
            placeholder.Name = "__placeholder__"
            placeholder_name = placeholder.Name
            toc_data = []
            copied_count = 0

            for fp in self.file_list:
                self._check_cancelled()
                self._post_status(f"正在直接复制 Sheet：{os.path.basename(fp)}")
                src_wb = excel.Workbooks.Open(str(Path(fp).resolve()), ReadOnly=True, UpdateLinks=0)
                src_wbs.append(src_wb)
                sheet_names = self.get_target_sheet_names_for_com(src_wb, sheet_config)
                for sheet_name in sheet_names:
                    self._check_cancelled()
                    self._post_status(f"正在复制：{os.path.basename(fp)} / {sheet_name}")
                    src_ws = src_wb.Worksheets(sheet_name)
                    src_ws.Copy(None, dest_wb.Worksheets(dest_wb.Worksheets.Count))
                    new_ws = excel.ActiveSheet
                    preferred = (
                        f"{Path(fp).stem}_{sheet_name}"
                        if len(sheet_names) > 1 or sheet_config.get("action") != "default"
                        else Path(fp).stem
                    )
                    new_name = self.make_unique_com_sheet_name(dest_wb, preferred)
                    new_ws.Name = new_name
                    toc_data.append({
                        "Source File": os.path.basename(fp),
                        "Target Sheet": new_name,
                    })
                    copied_count += 1

            if copied_count == 0:
                raise Exception("没有读取到有效Sheet")

            try:
                dest_wb.Worksheets(placeholder_name).Delete()
            except Exception:
                pass
            self.add_reference_sheet_com(dest_wb, toc_data)
            dest_wb.SaveAs(str(Path(save_path).resolve()), FileFormat=51)
        finally:
            for wb in reversed(src_wbs):
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if dest_wb is not None:
                try:
                    dest_wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def write_vertical_csv_polars_stream(self, save_path, sheet_config, is_physical):
        wrote_any = False
        wrote_header = False
        for fp in self.file_list:
            self._check_cancelled()
            self._post_status(f"正在读取并写入 CSV：{os.path.basename(fp)}")
            f_base = os.path.basename(fp)
            df_dict = self.load_all_sheets_polars(fp, sheet_config, auto_header=(not is_physical))
            if not df_dict:
                continue
            for sn, df in df_dict.items():
                self._check_cancelled()
                if df.is_empty():
                    continue
                if is_physical:
                    df.columns = [str(i) for i in range(df.width)]
                source_cols = [pl.lit(f_base).alias("【来源文件】")]
                if len(df_dict) > 1 or sheet_config['action'] != 'default':
                    source_cols.append(pl.lit(sn).alias("【来源Sheet】"))
                df = df.with_columns(source_cols)
                leading = [expr.meta.output_name() for expr in source_cols]
                df = df.select(leading + [c for c in df.columns if c not in leading])
                self._post_status(f"正在写入 CSV：{os.path.basename(fp)} / {sn}")
                with open(save_path, "ab" if wrote_any else "wb") as f:
                    if not wrote_any:
                        f.write(b"\xef\xbb\xbf")
                    df.write_csv(f, include_header=((not is_physical) and not wrote_header))
                wrote_any = True
                wrote_header = True
        if not wrote_any:
            raise Exception("没有读取到有效数据")

    def write_vertical_csv_pandas_stream(self, save_path, sheet_config, is_physical):
        wrote_any = False
        wrote_header = False
        for fp in self.file_list:
            self._check_cancelled()
            self._post_status(f"正在读取并写入 CSV：{os.path.basename(fp)}")
            f_base = os.path.basename(fp)
            df_dict = self.load_all_sheets(fp, sheet_config, auto_header=(not is_physical))
            if not df_dict:
                continue
            for sn, df in df_dict.items():
                self._check_cancelled()
                if df.empty:
                    continue
                if is_physical:
                    df.columns = range(len(df.columns))
                df.insert(0, "【来源文件】", f_base)
                if len(df_dict) > 1 or sheet_config['action'] != 'default':
                    df.insert(1, "【来源Sheet】", sn)
                self._post_status(f"正在写入 CSV：{os.path.basename(fp)} / {sn}")
                df.to_csv(
                    save_path,
                    mode=("w" if not wrote_any else "a"),
                    index=False,
                    header=((not is_physical) and not wrote_header),
                    encoding=("utf-8-sig" if not wrote_any else "utf-8"),
                )
                wrote_any = True
                wrote_header = True
        if not wrote_any:
            raise Exception("没有读取到有效数据")

    def write_vertical_csv_calamine_stream(self, save_path, sheet_config, is_physical):
        wrote_any = False
        wrote_header = False
        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for fp, sheet_name, include_sheet_col, row_iter in self.iter_calamine_sheet_sources(sheet_config, is_physical):
                self._check_cancelled()
                self._post_status(f"正在流式写入 CSV：{os.path.basename(fp)} / {sheet_name}")
                f_base = os.path.basename(fp)
                sheet_rows = 0
                for values in row_iter:
                    if sheet_rows % 1000 == 0:
                        self._check_cancelled()
                    if not is_physical and not wrote_header:
                        header_row = ["Source File"]
                        if include_sheet_col:
                            header_row.append("Source Sheet")
                        header_row.extend(values)
                        writer.writerow(header_row)
                        wrote_header = True
                        continue

                    out_row = [f_base]
                    if include_sheet_col:
                        out_row.append(sheet_name)
                    out_row.extend(values)
                    writer.writerow(out_row)
                    sheet_rows += 1
                    wrote_any = True
                if sheet_rows == 0:
                    self._record_read_warning(fp, sheet_name, "未读取到有效数据")
        if not wrote_any:
            raise Exception("没有读取到有效数据")

    def can_stream_vertical_xlsx(self):
        return all(f.lower().endswith(STREAM_XLSX_EXTENSIONS) for f in self.file_list)

    def get_target_sheet_names_for_openpyxl(self, workbook, sheet_config):
        all_sheets = workbook.sheetnames
        action = sheet_config.get("action", "default")
        if action == "match_selected":
            targets = set(sheet_config.get("targets", []))
            return [name for name in all_sheets if name in targets]
        if action == "merge_all":
            return all_sheets
        return all_sheets[:1]

    def get_target_sheet_names_for_calamine(self, workbook, sheet_config):
        all_sheets = workbook.sheet_names
        action = sheet_config.get("action", "default")
        if action == "match_selected":
            targets = set(sheet_config.get("targets", []))
            return [name for name in all_sheets if name in targets]
        if action == "merge_all":
            return all_sheets
        return all_sheets[:1]

    def iter_openpyxl_rows_for_merge(self, worksheet, auto_header):
        rows_iter = worksheet.iter_rows(values_only=True)
        if not auto_header:
            for row in rows_iter:
                if row and any(value is not None and str(value) != "" for value in row):
                    yield list(row)
            return

        preview = []
        first_row_index = None
        first_col_index = 0
        for idx, row in enumerate(rows_iter):
            row_values = list(row or [])
            preview.append(row_values)
            if first_row_index is None:
                for col_idx, value in enumerate(row_values):
                    if value is not None and str(value) != "":
                        first_row_index = idx
                        first_col_index = col_idx
                        break
            if len(preview) >= 20 and first_row_index is not None:
                break

        if first_row_index is None:
            return

        first_data_row = preview[first_row_index][first_col_index:]
        has_num = False
        for value in first_data_row:
            if value is None or str(value) == "":
                continue
            try:
                float(str(value).replace(',', ''))
                has_num = True
                break
            except Exception:
                continue

        start_index = first_row_index if has_num else first_row_index + 1
        header = None if has_num else [str(v) if v is not None else "" for v in first_data_row]

        if header is not None:
            yield header

        for row in preview[start_index:]:
            row_values = row[first_col_index:]
            if row_values and any(value is not None and str(value) != "" for value in row_values):
                yield row_values

        for row in rows_iter:
            row_values = list(row or [])[first_col_index:]
            if row_values and any(value is not None and str(value) != "" for value in row_values):
                yield row_values

    def iter_calamine_rows_for_merge(self, worksheet, auto_header):
        rows_iter = worksheet.iter_rows()
        if not auto_header:
            for row in rows_iter:
                if row and any(value is not None and str(value) != "" for value in row):
                    yield list(row)
            return

        preview = []
        first_row_index = None
        first_col_index = 0
        for idx, row in enumerate(rows_iter):
            row_values = list(row or [])
            preview.append(row_values)
            if first_row_index is None:
                for col_idx, value in enumerate(row_values):
                    if value is not None and str(value) != "":
                        first_row_index = idx
                        first_col_index = col_idx
                        break
            if len(preview) >= 20 and first_row_index is not None:
                break

        if first_row_index is None:
            return

        first_data_row = preview[first_row_index][first_col_index:]
        has_num = False
        for value in first_data_row:
            if value is None or str(value) == "":
                continue
            try:
                float(str(value).replace(',', ''))
                has_num = True
                break
            except Exception:
                continue

        start_index = first_row_index if has_num else first_row_index + 1
        header = None if has_num else [str(v) if v is not None else "" for v in first_data_row]

        if header is not None:
            yield header

        for row in preview[start_index:]:
            row_values = row[first_col_index:]
            if row_values and any(value is not None and str(value) != "" for value in row_values):
                yield row_values

        for row in rows_iter:
            row_values = list(row or [])[first_col_index:]
            if row_values and any(value is not None and str(value) != "" for value in row_values):
                yield row_values

    def write_vertical_xlsx_rows_stream(self, save_path, sheet_config, is_physical, sheet_sources):
        from openpyxl import Workbook

        out_wb = Workbook(write_only=True)
        current_ws = None
        current_row = 0
        sheet_no = 0
        wrote_any = False
        wrote_header = False
        header_row = None

        def create_output_sheet():
            nonlocal current_ws, current_row, sheet_no
            sheet_no += 1
            title = "Merged" if sheet_no == 1 else f"Merged_{sheet_no}"
            current_ws = out_wb.create_sheet(title=title)
            current_row = 0
            if header_row is not None and not is_physical:
                current_ws.append(header_row)
                current_row = 1

        create_output_sheet()

        for fp, sheet_name, include_sheet_col, row_iter in sheet_sources:
            self._check_cancelled()
            self._post_status(f"正在流式写入 Excel：{os.path.basename(fp)} / {sheet_name}")
            f_base = os.path.basename(fp)
            sheet_rows = 0
            for values in row_iter:
                if sheet_rows % 1000 == 0:
                    self._check_cancelled()
                if not is_physical and not wrote_header:
                    header_row = ["Source File"]
                    if include_sheet_col:
                        header_row.append("Source Sheet")
                    header_row.extend(values)
                    current_ws.append(header_row)
                    current_row += 1
                    wrote_header = True
                    continue

                out_row = [f_base]
                if include_sheet_col:
                    out_row.append(sheet_name)
                out_row.extend(values)

                if current_row >= EXCEL_MAX_ROWS:
                    create_output_sheet()
                current_ws.append(out_row)
                current_row += 1
                sheet_rows += 1
                wrote_any = True
            if sheet_rows == 0:
                self._record_read_warning(fp, sheet_name, "未读取到有效数据")

        if not wrote_any:
            raise Exception("没有读取到有效数据")
        out_wb.save(save_path)

    def write_vertical_xlsx_rows_xlsxwriter_stream(self, save_path, sheet_config, is_physical, sheet_sources):
        workbook = xlsxwriter.Workbook(save_path, {'constant_memory': True})
        workbook.use_zip64()
        link_fmt = workbook.add_format({'font_color': 'blue', 'underline': 1}) if self.var_add_hyperlinks.get() else None
        current_ws = None
        current_row = 0
        sheet_no = 0
        wrote_any = False
        wrote_header = False
        header_row = None

        def create_output_sheet():
            nonlocal current_ws, current_row, sheet_no
            sheet_no += 1
            title = "Merged" if sheet_no == 1 else f"Merged_{sheet_no}"
            current_ws = workbook.add_worksheet(title)
            current_row = 0
            if header_row is not None and not is_physical:
                current_ws.write_row(current_row, 0, header_row)
                current_row = 1

        try:
            create_output_sheet()

            for fp, sheet_name, include_sheet_col, row_iter in sheet_sources:
                self._check_cancelled()
                self._post_status(f"正在流式写入 Excel：{os.path.basename(fp)} / {sheet_name}")
                f_base = os.path.basename(fp)
                sheet_rows = 0
                for values in row_iter:
                    if sheet_rows % 1000 == 0:
                        self._check_cancelled()
                    if not is_physical and not wrote_header:
                        header_row = ["Source File"]
                        if include_sheet_col:
                            header_row.append("Source Sheet")
                        header_row.extend(values)
                        current_ws.write_row(current_row, 0, header_row)
                        current_row += 1
                        wrote_header = True
                        continue

                    out_row = [f_base]
                    if include_sheet_col:
                        out_row.append(sheet_name)
                    out_row.extend(values)

                    if current_row >= EXCEL_MAX_ROWS:
                        create_output_sheet()
                    if link_fmt is not None:
                        try:
                            current_ws.write_url(current_row, 0, f"external:{fp}", link_fmt, string=f_base)
                            if len(out_row) > 1:
                                current_ws.write_row(current_row, 1, out_row[1:])
                        except Exception:
                            current_ws.write_row(current_row, 0, out_row)
                    else:
                        current_ws.write_row(current_row, 0, out_row)
                    current_row += 1
                    sheet_rows += 1
                    wrote_any = True
                if sheet_rows == 0:
                    self._record_read_warning(fp, sheet_name, "未读取到有效数据")

            if not wrote_any:
                raise Exception("没有读取到有效数据")
        finally:
            workbook.close()

    def iter_openpyxl_sheet_sources(self, sheet_config, is_physical):
        from openpyxl import load_workbook

        for fp in self.file_list:
            src_wb = None
            try:
                src_wb = load_workbook(fp, read_only=True, data_only=True, keep_links=False)
                target_sheets = self.get_target_sheet_names_for_openpyxl(src_wb, sheet_config)
                if not target_sheets:
                    self._record_read_warning(fp, reason="未找到符合条件的 Sheet")
                include_sheet_col = len(target_sheets) > 1 or sheet_config.get('action') != 'default'
                for sheet_name in target_sheets:
                    ws = src_wb[sheet_name]
                    row_iter = self.iter_openpyxl_rows_for_merge(ws, auto_header=(not is_physical))
                    yield fp, sheet_name, include_sheet_col, row_iter
            finally:
                if src_wb is not None:
                    src_wb.close()

    def iter_calamine_sheet_sources(self, sheet_config, is_physical):
        for fp in self.file_list:
            src_wb = load_calamine_workbook(fp)
            target_sheets = self.get_target_sheet_names_for_calamine(src_wb, sheet_config)
            if not target_sheets:
                self._record_read_warning(fp, reason="未找到符合条件的 Sheet")
            include_sheet_col = len(target_sheets) > 1 or sheet_config.get('action') != 'default'
            for sheet_name in target_sheets:
                ws = src_wb.get_sheet_by_name(sheet_name)
                row_iter = self.iter_calamine_rows_for_merge(ws, auto_header=(not is_physical))
                yield fp, sheet_name, include_sheet_col, row_iter

    def write_vertical_xlsx_calamine_stream(self, save_path, sheet_config, is_physical):
        sheet_sources = self.iter_calamine_sheet_sources(sheet_config, is_physical)
        if HAS_XLSXWRITER:
            self.write_vertical_xlsx_rows_xlsxwriter_stream(save_path, sheet_config, is_physical, sheet_sources)
        else:
            self.write_vertical_xlsx_rows_stream(save_path, sheet_config, is_physical, sheet_sources)

    def write_vertical_xlsx_openpyxl_stream(self, save_path, sheet_config, is_physical):
        sheet_sources = self.iter_openpyxl_sheet_sources(sheet_config, is_physical)
        if HAS_XLSXWRITER:
            self.write_vertical_xlsx_rows_xlsxwriter_stream(save_path, sheet_config, is_physical, sheet_sources)
        else:
            self.write_vertical_xlsx_rows_stream(save_path, sheet_config, is_physical, sheet_sources)

    def run_process(self, save_path, sheet_config):
        try:
            self._reset_read_warnings()
            mode = self.var_mode.get()
            
            # === 1. 多 Sheet 模式 (带 Reference) ===
            if mode == "one_workbook":
                if save_path.lower().endswith(".csv"):
                    raise Exception("多Sheet模式必须保存为 .xlsx 格式")

                try:
                    self.copy_workbook_sheets_with_com(save_path, sheet_config)
                    self.root.after(0, lambda: self._notify_success(save_path))
                    return
                except Exception as com_err:
                    total_input_size = self._get_total_input_size()
                    if total_input_size >= ONE_WORKBOOK_FALLBACK_REWRITE_LIMIT:
                        raise Exception(
                            "多 Sheet 模式的 Excel 直接复制失败，且源文件体积较大，已停止自动回退重写。\n\n"
                            f"源文件总大小：约 {self._format_size(total_input_size)}\n"
                            "如果继续回退为普通导出，工具需要重新读取并写入所有 Sheet，可能等待很久且丢失部分原始格式。\n\n"
                            "建议：关闭已打开的 Excel 文件后重试；如果只是要合成一张明细表，请改选 One Sheet 并优先导出 CSV。\n\n"
                            f"直接复制错误信息：{com_err}"
                        )
                    fallback_warning = (
                        "Excel COM 原样复制失败，已自动改用普通导出。\n"
                        "导出文件已生成，但源 Sheet 的完整格式、对象、部分公式/超链接可能无法保留。\n\n"
                        f"COM 错误信息：{com_err}"
                    )
                
                with pd.ExcelWriter(save_path, engine="xlsxwriter" if HAS_XLSXWRITER else "openpyxl") as writer:
                    # 【核心更新】提前创建目录页，并记录目录数据
                    toc_data = [] # [(filename, sheet_name), ...]
                    
                    # 1. 如果使用 xlsxwriter，可以先预留一个Sheet位置
                    # 但为了简单，我们可以先遍历写数据，最后再把 Reference 页移动到最前？
                    # 不，xlsxwriter 写入顺序决定了显示顺序。
                    # 更好的方法：先创建一个空的 'Reference' 页
                    
                    if HAS_XLSXWRITER:
                        ws_ref = writer.book.add_worksheet('Reference')
                    
                    # 2. 写入数据页
                    for idx, fp in enumerate(self.file_list):
                        self._check_cancelled()
                        self._post_status(f"正在普通导出：{os.path.basename(fp)}")
                        df_dict = self.load_all_sheets(fp, sheet_config)
                        if not df_dict: continue
                        f_base = self.clean_sheet_name(os.path.basename(fp))
                        
                        for sn, df in df_dict.items():
                            self._check_cancelled()
                            self._post_status(f"正在写入：{os.path.basename(fp)} / {sn}")
                            target = f"{f_base}_{sn}" if len(df_dict)>1 or sheet_config['action']!='default' else f_base
                            target = self.clean_sheet_name(target)
                            if target in writer.sheets:
                                target = self.make_unique_sheet_name(target, writer.sheets, idx)
                            
                            # 写入数据
                            df.to_excel(writer, sheet_name=target, index=False)
                            
                            # 记录到目录
                            # 注意：Source File 显示原名，Sheet Name 显示最终名
                            toc_data.append({
                                "Source File": os.path.basename(fp),
                                "Target Sheet": target
                            })

                    # 3. 填充 Reference 页内容 (如果支持 xlsxwriter)
                    if HAS_XLSXWRITER and toc_data:
                        ws_ref = writer.sheets['Reference']
                        # 样式
                        fmt_head = writer.book.add_format({'bold':True, 'border':1, 'bg_color':'#D9D9D9', 'align':'center'})
                        fmt_link = writer.book.add_format({'font_color':'blue', 'underline':1, 'border':1})
                        fmt_norm = writer.book.add_format({'border':1})
                        
                        # 写表头
                        ws_ref.write(0, 0, "Source File Name", fmt_head)
                        ws_ref.write(0, 1, "Target Sheet Link", fmt_head)
                        ws_ref.set_column(0, 0, 40)
                        ws_ref.set_column(1, 1, 40)
                        
                        # 写内容
                        for i, item in enumerate(toc_data):
                            row = i + 1
                            ws_ref.write(row, 0, item["Source File"], fmt_norm)
                            # 写入超链接: internal:'SheetName'!A1
                            # 注意 Excel Sheet 名如果包含空格或特殊字符，需用单引号包裹
                            s_name = item["Target Sheet"]
                            if self.var_add_hyperlinks.get():
                                link = f"internal:'{s_name}'!A1"
                                ws_ref.write_url(row, 1, link, fmt_link, string=s_name)
                            else:
                                ws_ref.write(row, 1, s_name, fmt_norm)
                        
                        # 激活 Reference 页为默认打开页
                        ws_ref.activate()
                if 'fallback_warning' in locals():
                    self.root.after(0, lambda msg=fallback_warning: self._notify_success(save_path, msg))
                    return

            # === 2. 单 Sheet 合并模式 ===
            else:
                direction = self.var_direction.get()
                is_physical = (direction == "vertical")
                if save_path.lower().endswith(".csv") and direction == "vertical":
                    if HAS_CALAMINE and self.can_stream_vertical_xlsx():
                        self.write_vertical_csv_calamine_stream(save_path, sheet_config, is_physical)
                        self.root.after(0, lambda: self._notify_success(save_path))
                        return
                    if HAS_POLARS:
                        warning_checkpoint = len(self.read_warnings)
                        try:
                            self.write_vertical_csv_polars_stream(save_path, sheet_config, is_physical)
                            self.root.after(0, lambda: self._notify_success(save_path))
                            return
                        except Exception as fast_err:
                            self.read_warnings = self.read_warnings[:warning_checkpoint]
                            print(f"polars csv stream failed, fallback to pandas: {fast_err}")
                    self.write_vertical_csv_pandas_stream(save_path, sheet_config, is_physical)
                    self.root.after(0, lambda: self._notify_success(save_path))
                    return

                if save_path.lower().endswith(".xlsx") and direction == "vertical" and self.can_stream_vertical_xlsx():
                    if HAS_CALAMINE:
                        self.write_vertical_xlsx_calamine_stream(save_path, sheet_config, is_physical)
                    else:
                        self.write_vertical_xlsx_openpyxl_stream(save_path, sheet_config, is_physical)
                    self.root.after(0, lambda: self._notify_success(save_path))
                    return

                if HAS_POLARS and HAS_XLSXWRITER:
                    warning_checkpoint = len(self.read_warnings)
                    try:
                        self.run_process_single_polars(save_path, sheet_config)
                        self.root.after(0, lambda: self._notify_success(save_path))
                        return
                    except Exception as fast_err:
                        self.read_warnings = self.read_warnings[:warning_checkpoint]
                        print(f"polars fast path failed, fallback to pandas: {fast_err}")

                dfs = []
                dfs_metadata = [] 

                for fp in self.file_list:
                    self._check_cancelled()
                    self._post_status(f"正在读取：{os.path.basename(fp)}")
                    f_base = os.path.basename(fp)
                    df_dict = self.load_all_sheets(fp, sheet_config, auto_header=(not is_physical))
                    if not df_dict: continue

                    for sn, df in df_dict.items():
                        self._check_cancelled()
                        self._post_status(f"正在整理：{os.path.basename(fp)} / {sn}")
                        if df.empty: continue
                        meta = {"path": fp, "fname": f_base, "sheet": sn, "rows": len(df), "cols": len(df.columns)}
                        
                        if direction == "horizontal":
                            df.reset_index(drop=True, inplace=True)
                            dfs.append(df)
                            dfs_metadata.append(meta)
                        elif direction == "vertical":
                            if is_physical: df.columns = range(len(df.columns))
                            df.insert(0, "【来源文件】", f_base)
                            if len(df_dict)>1 or sheet_config['action']!='default':
                                df.insert(1, "【来源Sheet】", sn)
                            dfs.append(df)
                            dfs_metadata.append(meta)

                if not dfs: raise Exception("没有读取到有效数据")

                if direction == "horizontal":
                    self._check_cancelled()
                    self._post_status("正在横向拼接所有表格...")
                    keys = [f"{m['fname']} - {m['sheet']}" for m in dfs_metadata]
                    final_df = pd.concat(dfs, axis=1, keys=keys)
                else:
                    self._check_cancelled()
                    self._post_status("正在纵向堆叠所有表格...")
                    final_df = pd.concat(dfs, axis=0, ignore_index=True, sort=False)

                if save_path.lower().endswith(".csv"):
                    self._check_cancelled()
                    self._post_status("正在写出 CSV 文件...")
                    final_df.to_csv(save_path, index=False, header=(not is_physical), encoding="utf-8-sig")
                else:
                    self._check_cancelled()
                    self._post_status("正在写出 Excel 文件...")
                    with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
                        if direction == "horizontal":
                            flat_df = final_df.copy()
                            flat_df.columns = flat_df.columns.get_level_values(1)
                            flat_df.to_excel(writer, sheet_name="合并结果", startrow=2, header=False, index=False)
                            ws = writer.sheets['合并结果']
                            fmt_h = writer.book.add_format({'bold':True, 'border':1, 'align':'center', 'bg_color':'#D7E4BC'})
                            fmt_l = writer.book.add_format({'font_color':'blue', 'underline':1, 'bold':True, 'border':1, 'align':'center', 'bg_color':'#D7E4BC'})
                            
                            start_col = 0
                            for meta in dfs_metadata:
                                w = meta['cols']
                                if w == 0: continue
                                display_name = meta['fname']
                                if sheet_config['action'] != 'default': display_name += f" ({meta['sheet']})"
                                for offset in range(w):
                                    if self.var_add_hyperlinks.get():
                                        try: ws.write_url(0, start_col+offset, f"external:{meta['path']}", fmt_l, string=display_name)
                                        except: ws.write(0, start_col+offset, display_name, fmt_h)
                                    else:
                                        ws.write(0, start_col+offset, display_name, fmt_h)
                                sub_df = dfs[dfs_metadata.index(meta)]
                                for c_i, c_name in enumerate(sub_df.columns):
                                    ws.write(1, start_col+c_i, str(c_name), fmt_h)
                                start_col += w
                        else:
                            final_df.to_excel(writer, sheet_name="合并结果", index=False, header=(not is_physical))
                            if HAS_XLSXWRITER:
                                ws = writer.sheets['合并结果']
                                link_fmt = writer.book.add_format({'font_color':'blue', 'underline':1})
                                current_row = 0 if is_physical else 1
                                for i, meta in enumerate(dfs_metadata):
                                    row_count = len(dfs[i]) 
                                    f_path = meta['path']; f_name = meta['fname']
                                    for _ in range(row_count):
                                        if current_row % 1000 == 0:
                                            self._check_cancelled()
                                        if self.var_add_hyperlinks.get():
                                            try: ws.write_url(current_row, 0, f"external:{f_path}", link_fmt, string=f_name)
                                            except: pass
                                        current_row += 1

            self.root.after(0, lambda: self._notify_success(save_path))

        except MergeCancelled as e:
            self.root.after(0, lambda msg=str(e): self.on_cancelled(msg))
        except Exception as e:
            err_msg = str(e); print(err_msg)
            self.root.after(0, lambda: self.on_error(err_msg))

    def run_process_single_polars(self, save_path, sheet_config):
        direction = self.var_direction.get()
        is_physical = (direction == "vertical")
        dfs = []
        dfs_metadata = []

        for fp in self.file_list:
            self._check_cancelled()
            self._post_status(f"正在快速读取：{os.path.basename(fp)}")
            f_base = os.path.basename(fp)
            df_dict = self.load_all_sheets_polars(fp, sheet_config, auto_header=(not is_physical))
            if not df_dict:
                continue

            for sn, df in df_dict.items():
                self._check_cancelled()
                self._post_status(f"正在快速整理：{os.path.basename(fp)} / {sn}")
                if df.is_empty():
                    continue

                meta = {
                    "path": fp,
                    "fname": f_base,
                    "sheet": sn,
                    "rows": df.height,
                    "cols": df.width,
                    "columns": list(df.columns),
                }

                if direction == "horizontal":
                    df = df.rename({col: f"__{len(dfs)}_{i}" for i, col in enumerate(df.columns)})
                    dfs.append(df)
                    dfs_metadata.append(meta)
                elif direction == "vertical":
                    if is_physical:
                        df.columns = [str(i) for i in range(df.width)]
                    source_cols = [pl.lit(f_base).alias("【来源文件】")]
                    if len(df_dict) > 1 or sheet_config['action'] != 'default':
                        source_cols.append(pl.lit(sn).alias("【来源Sheet】"))
                    df = df.with_columns(source_cols)
                    leading = [expr.meta.output_name() for expr in source_cols]
                    df = df.select(leading + [c for c in df.columns if c not in leading])
                    dfs.append(df)
                    dfs_metadata.append(meta)

        if not dfs:
            raise Exception("没有读取到有效数据")

        if direction == "horizontal":
            self._check_cancelled()
            self._post_status("正在快速横向拼接所有表格...")
            final_df = pl.concat(dfs, how="horizontal")
        else:
            self._check_cancelled()
            self._post_status("正在快速纵向堆叠所有表格...")
            final_df = pl.concat(dfs, how="diagonal_relaxed")

        if save_path.lower().endswith(".csv"):
            self._check_cancelled()
            self._post_status("正在写出 CSV 文件...")
            final_df.write_csv(save_path, include_header=(not is_physical), include_bom=True)
            return

        self._check_cancelled()
        self._post_status("正在写出 Excel 文件...")
        self.write_polars_excel(save_path, final_df, dfs_metadata, direction, is_physical, sheet_config)

    def write_polars_excel(self, save_path, final_df, dfs_metadata, direction, is_physical, sheet_config):
        workbook = xlsxwriter.Workbook(save_path)
        try:
            ws = workbook.add_worksheet("合并结果")
            if direction == "horizontal":
                fmt_h = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'bg_color': '#D7E4BC'})
                fmt_l = workbook.add_format({'font_color': 'blue', 'underline': 1, 'bold': True, 'border': 1, 'align': 'center', 'bg_color': '#D7E4BC'})

                start_col = 0
                flat_columns = []
                for meta in dfs_metadata:
                    display_name = meta['fname']
                    if sheet_config['action'] != 'default':
                        display_name += f" ({meta['sheet']})"
                    for col_name in meta["columns"]:
                        if self.var_add_hyperlinks.get():
                            try:
                                ws.write_url(0, start_col, f"external:{meta['path']}", fmt_l, string=display_name)
                            except Exception:
                                ws.write(0, start_col, display_name, fmt_h)
                        else:
                            ws.write(0, start_col, display_name, fmt_h)
                        ws.write(1, start_col, str(col_name), fmt_h)
                        flat_columns.append(str(col_name))
                        start_col += 1

                self.write_polars_table(ws, final_df, startrow=2, header=False)
            else:
                self.write_polars_table(ws, final_df, startrow=0, header=(not is_physical))
                link_fmt = workbook.add_format({'font_color': 'blue', 'underline': 1})
                current_row = 0 if is_physical else 1
                for i, meta in enumerate(dfs_metadata):
                    row_count = dfs_metadata[i]["rows"]
                    for _ in range(row_count):
                        if current_row % 1000 == 0:
                            self._check_cancelled()
                        if self.var_add_hyperlinks.get():
                            try:
                                ws.write_url(current_row, 0, f"external:{meta['path']}", link_fmt, string=meta['fname'])
                            except Exception:
                                pass
                        current_row += 1
        finally:
            workbook.close()

    def write_polars_table(self, worksheet, df, startrow=0, header=True):
        row_offset = startrow
        if header:
            for c_i, col in enumerate(df.columns):
                worksheet.write(row_offset, c_i, str(col))
            row_offset += 1

        for r_i, row in enumerate(df.iter_rows()):
            for c_i, value in enumerate(row):
                if value is None:
                    worksheet.write_blank(row_offset + r_i, c_i, None)
                else:
                    worksheet.write(row_offset + r_i, c_i, value)

    def load_all_sheets_polars(self, fp, sheet_config, auto_header=True):
        result = {}
        if fp.lower().endswith(('.csv', '.txt')):
            try:
                df = self.load_single_sheet_data_polars(fp, None, auto_header)
                if df is not None and not df.is_empty():
                    result["CSV"] = df
                else:
                    self._record_read_warning(fp, reason="未读取到有效数据")
            except Exception as exc:
                self._record_read_warning(fp, reason=f"读取失败：{exc}")
            return result

        try:
            all_sheets = fastexcel.read_excel(fp).sheet_names
        except Exception as exc:
            self._record_read_warning(fp, reason=f"读取工作簿失败：{exc}")
            return result
        action = sheet_config.get("action", "default")
        if action == "match_selected":
            req = sheet_config.get("targets", [])
            targets = [s for s in all_sheets if s in req]
        elif action == "merge_all":
            targets = all_sheets
        else:
            targets = [all_sheets[0]]

        if not targets:
            self._record_read_warning(fp, reason="未找到符合条件的 Sheet")
            return result

        for s in targets:
            try:
                df = self.load_single_sheet_data_polars(fp, s, auto_header)
                if df is not None and not df.is_empty():
                    result[s] = df
                else:
                    self._record_read_warning(fp, s, "未读取到有效数据")
            except Exception as exc:
                self._record_read_warning(fp, s, f"读取失败：{exc}")
        return result

    def load_single_sheet_data_polars(self, fp, sheet_name, auto_header):
        is_text = fp.lower().endswith(('.csv', '.txt'))

        if not auto_header:
            return self.read_polars_data(fp, sheet_name, has_header=False, is_text=is_text)

        df_p = self.read_polars_data(fp, sheet_name, has_header=False, is_text=is_text, n_rows=20)
        if df_p.is_empty():
            return None

        s_r = 0
        s_c = 0
        found = False
        for r, row in enumerate(df_p.iter_rows()):
            for c, value in enumerate(row):
                if value is not None and str(value) != "":
                    s_r = r
                    s_c = c
                    found = True
                    break
            if found:
                break
        if not found:
            return None

        vals = df_p.row(s_r)[s_c:]
        has_num = False
        for v in vals:
            if v is None:
                continue
            try:
                float(str(v).replace(',', ''))
                has_num = True
                break
            except Exception:
                continue

        df = self.read_polars_data(
            fp,
            sheet_name,
            has_header=(not has_num),
            is_text=is_text,
            skip_rows=s_r,
            header_row=(s_r if not has_num else None),
        )
        if s_c > 0:
            df = df[:, s_c:]
        return df

    def read_polars_data(self, fp, sheet_name, has_header, is_text, n_rows=None, skip_rows=0, header_row=None):
        if is_text:
            return pl.read_csv(
                fp,
                has_header=has_header,
                infer_schema=False,
                n_rows=n_rows,
                skip_rows=skip_rows,
                encoding="utf8",
            )

        read_options = {}
        if n_rows is not None:
            read_options["n_rows"] = n_rows
        if header_row is not None:
            read_options["header_row"] = header_row
        elif skip_rows:
            read_options["header_row"] = None
            read_options["skip_rows"] = skip_rows

        return pl.read_excel(
            fp,
            sheet_name=sheet_name,
            has_header=has_header,
            infer_schema_length=0,
            read_options=(read_options or None),
        )

    def load_all_sheets(self, fp, sheet_config, auto_header=True):
        result = {}
        try:
            if fp.lower().endswith(('.csv', '.txt')):
                df = self.load_single_sheet_data(fp, None, auto_header)
                if df is not None and not df.empty:
                    result["CSV"] = df
                else:
                    self._record_read_warning(fp, reason="未读取到有效数据")
                return result

            xls = pd.ExcelFile(fp)
            all_sheets = xls.sheet_names
            targets = []
            action = sheet_config.get("action", "default")
            
            if action == "match_selected":
                req = sheet_config.get("targets", [])
                targets = [s for s in all_sheets if s in req]
            elif action == "merge_all":
                targets = all_sheets
            else:
                targets = [all_sheets[0]]

            if not targets:
                self._record_read_warning(fp, reason="未找到符合条件的 Sheet")
                return result
            
            for s in targets:
                df = self.load_single_sheet_data(fp, s, auto_header)
                if df is not None and not df.empty:
                    result[s] = df
                else:
                    self._record_read_warning(fp, s, "未读取到有效数据")
            return result
        except Exception as exc:
            self._record_read_warning(fp, reason=f"读取工作簿失败：{exc}")
            return {}

    def load_single_sheet_data(self, fp, sheet_name, auto_header):
        try:
            read_func = pd.read_csv if fp.lower().endswith(('.csv','.txt')) else pd.read_excel
            args = {'dtype': str}
            if sheet_name and not fp.lower().endswith(('.csv','.txt')): args['sheet_name'] = sheet_name
            
            if not auto_header:
                args['header'] = None
                if fp.lower().endswith(('.csv','.txt')):
                    try: return read_func(fp, encoding='utf-8-sig', **args)
                    except: return read_func(fp, encoding='gbk', **args)
                else: return read_func(fp, **args)

            p_args = args.copy(); p_args.update({'nrows':20, 'header':None})
            if fp.lower().endswith(('.csv','.txt')):
                try: df_p = read_func(fp, encoding='utf-8-sig', **p_args)
                except: df_p = read_func(fp, encoding='gbk', **p_args)
            else: df_p = read_func(fp, **p_args)
            
            if df_p.empty: return None
            s_r=0; s_c=0; found=False
            for r, row in df_p.iterrows():
                if row.isna().all(): continue
                idx = row.first_valid_index()
                if idx is not None: s_r=r; s_c=idx; found=True; break
            if not found: return None
            
            vals = df_p.iloc[s_r, s_c:]
            has_num = False
            for v in vals:
                if pd.isna(v): continue
                try: float(str(v).replace(',','')); has_num=True; break
                except: continue
            
            f_args = args.copy(); f_args.update({'skiprows':s_r, 'header':(None if has_num else 0)})
            if fp.lower().endswith(('.csv','.txt')):
                try: df = read_func(fp, encoding='utf-8-sig', **f_args)
                except: df = read_func(fp, encoding='gbk', **f_args)
            else: df = read_func(fp, **f_args)
            
            if s_c > 0: df = df.iloc[:, s_c:]
            return df
        except Exception:
            return None

    def clean_sheet_name(self, filename):
        name = os.path.splitext(filename)[0]
        for char in '[]:*?/\\': name = name.replace(char, "_")
        return name[:31]

    def make_unique_sheet_name(self, preferred, existing_sheets, index_hint=1):
        existing = set(existing_sheets)
        base = self.clean_sheet_name(preferred) or "Sheet"
        if base not in existing:
            return base
        for n in [index_hint] + list(range(1, 10000)):
            suffix = f"_{n}"
            candidate = self.clean_sheet_name(base[:31 - len(suffix)] + suffix)
            if candidate not in existing:
                return candidate
        raise Exception(f"无法生成不重复的Sheet名称: {preferred}")

    def on_success(self, path):
        self.pb.stop()
        self.cancel_requested = False
        self._set_processing_state(False, f"合并完成：{len(self.file_list)} 个文件")
        messagebox.showinfo("完成", f"合并成功！\n文件已保存至：\n{path}")

    def on_success_with_warning(self, path, warning):
        self.pb.stop()
        self.cancel_requested = False
        self._set_processing_state(False, f"合并完成：{len(self.file_list)} 个文件（有提示）")
        messagebox.showwarning("完成（已降级导出）", f"合并已完成！\n文件已保存至：\n{path}\n\n{warning}")

    def on_cancelled(self, msg):
        self.pb.stop()
        self.cancel_requested = False
        self._set_processing_state(False, "合并已停止")
        messagebox.showinfo("已停止", msg or "本次合并已停止。")

    def on_error(self, msg):
        self.pb.stop()
        self.cancel_requested = False
        self._set_processing_state(False, "合并失败，请检查文件或规则后重试")
        messagebox.showerror("发生错误", msg)

if __name__ == "__main__":
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: pass
    app = BatchMergeApp(root)
    root.mainloop()
