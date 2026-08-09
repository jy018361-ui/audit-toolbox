#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CRA paste parsing and workbook application helpers.

The first CRA implementation is intentionally conservative: it parses pasted
table-like text into records, lets the user confirm them in the UI, then writes
only the small CRA/threshold table in each generated workpaper.
"""

from __future__ import annotations

import re
import csv
import io
import os
from copy import copy
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill


CRA_PARSER_VERSION = "2026-07-crafix1"


ASSERTION_ALIASES = {
    "C": ("C", "COMPLETENESS", "\u5b8c\u6574"),
    "E/O": ("E/O", "EO", "E", "O", "EXISTENCE", "OCCURRENCE", "\u5b58\u5728", "\u53d1\u751f"),
    "V/M": ("V/M", "M/V", "VM", "MV", "V", "M", "VALUATION", "MEASUREMENT", "\u8ba1\u4ef7", "\u8ba1\u91cf"),
    "R&O": ("R&O", "RO", "RIGHT", "OBLIGATION", "\u6743\u5229", "\u4e49\u52a1"),
    "P&D": ("P&D", "PD", "PRESENTATION", "DISCLOSURE", "\u5217\u62a5", "\u62ab\u9732"),
}

RISK_ALIASES = {
    "Minimal": ("MINIMAL", "\u6700\u4f4e", "\u6781\u4f4e"),
    "Low": ("LOW", "\u4f4e"),
    "Moderate": ("MODERATE", "MEDIUM", "\u4e2d"),
    "High": ("HIGH", "\u9ad8"),
    "N/A": ("N/A", "NA", "\u4e0d\u9002\u7528"),
}

SUBJECT_ALIASES = {
    "C": ("C", "\u8d27\u5e01\u8d44\u91d1", "\u73b0\u91d1", "\u94f6\u884c\u5b58\u6b3e", "CASH", "BANK"),
    "F": ("F", "\u5b58\u8d27", "INVENTORY"),
    "F1": ("F1", "\u5b58\u8d27\u8dcc\u4ef7\u51c6\u5907", "\u8dcc\u4ef7\u51c6\u5907", "INVENTORY WRITE-DOWN"),
    "E1": ("E1", "\u5e94\u6536\u8d26\u6b3e", "\u5176\u4ed6\u5e94\u6536\u6b3e", "ACCOUNTS RECEIVABLE", "RECEIVABLE"),
    "ECL": ("ECL", "\u574f\u8d26\u51c6\u5907", "\u9884\u671f\u4fe1\u7528\u635f\u5931", "ECL", "ALLOWANCE"),
    "G3": ("G3", "\u9884\u4ed8\u8d26\u6b3e", "PREPAYMENT", "PREPAID"),
    "J1": ("J1", "\u5728\u5efa\u5de5\u7a0b", "CIP", "CONSTRUCTION IN PROGRESS"),
    "K1": ("K1", "\u56fa\u5b9a\u8d44\u4ea7", "FIXED ASSET", "PPE"),
    "L1": ("L1", "\u65e0\u5f62\u8d44\u4ea7", "INTANGIBLE"),
    "L2": ("L2", "\u957f\u671f\u5f85\u644a", "LONG-TERM DEFERRED", "DEFERRED EXPENSE"),
    "M": ("M", "\u5e94\u4ed8\u7968\u636e", "NOTES PAYABLE"),
    "N": ("N", "\u5e94\u4ed8\u8d26\u6b3e", "ACCOUNTS PAYABLE", "PAYABLE"),
    "Q1": ("Q1", "\u94f6\u884c\u501f\u6b3e", "\u501f\u6b3e", "BORROWING", "LOAN"),
    "Uexp": ("UEXP", "U_EXP", "\u8d22\u52a1\u8d39\u7528", "FINANCE EXPENSE", "FINANCIAL EXPENSE"),
    "UexpVCVD": (
        "UEXPVCVD",
        "VC&VD",
        "\u9500\u552e\u8d39\u7528",
        "\u7ba1\u7406\u8d39\u7528",
        "SELLING EXPENSE",
        "ADMINISTRATIVE EXPENSE",
        "GENERAL AND ADMIN",
    ),
}

EXPLICIT_SUBJECT_CODES = {
    "C": "C",
    "F": "F",
    "F1": "F1",
    "E1": "E1",
    "ECL": "ECL",
    "G3": "G3",
    "J1": "J1",
    "K1": "K1",
    "L1": "L1",
    "L2": "L2",
    "M": "M",
    "N": "N",
    "Q1": "Q1",
    "UEXP": "Uexp",
    "U_EXP": "Uexp",
    "UEXPVCVD": "UexpVCVD",
    "VC&VD": "UexpVCVD",
    "VC": "UexpVCVD",
    "VD": "UexpVCVD",
}

LIABILITY_EXPENSE_SUBJECTS = {"M", "N", "Q1", "Uexp", "UexpVCVD"}
ASSET_REVENUE_SUBJECTS = {"C", "F", "F1", "E1", "ECL", "G3", "J1", "K1", "L1", "L2"}
ASSERTION_ONLY_KEYS = {
    "C",
    "E",
    "O",
    "V",
    "M",
    "R",
    "P",
    "D",
    "EO",
    "E/O",
    "VM",
    "MV",
    "V/M",
    "M/V",
    "RO",
    "R&O",
    "R & O",
    "PD",
    "P&D",
    "P & D",
}
LIABILITY_EXPENSE_ACCOUNT_TOKENS = (
    "应付",
    "应交",
    "借款",
    "负债",
    "费用",
    "税费",
    "税金",
    "PAYABLE",
    "BORROWING",
    "LOAN",
    "LIABILITY",
    "EXPENSE",
    "TAX",
)
ASSET_REVENUE_ACCOUNT_TOKENS = (
    "货币资金",
    "现金",
    "银行存款",
    "存货",
    "应收",
    "预付",
    "在建工程",
    "固定资产",
    "无形资产",
    "长期待摊",
    "资产",
    "收入",
    "CASH",
    "BANK",
    "INVENTORY",
    "RECEIVABLE",
    "ASSET",
    "REVENUE",
    "INCOME",
)

RATIO_RANGES = {
    "asset_revenue": {
        "Minimal": (0.75, 1.00),
        "Low": (0.50, 0.75),
        "Moderate": (0.25, 0.50),
        "High": (0.10, 0.25),
    },
    "liability_expense": {
        "Minimal": (0.25, 0.50),
        "Low": (0.15, 0.25),
        "Moderate": (0.10, 0.15),
        "High": (0.05, 0.10),
    },
}

RISK_TOKEN_RE = re.compile(
    r"(HIGH\+SC|LOW\+SC|MINIMAL|MODERATE|MEDIUM|HIGH|LOW|N/A|NA|较高|较低|中等程度|中等|最低|极低|高|中|低)",
    re.IGNORECASE,
)

CRA_TOKEN_RE = re.compile(
    r"(HIGH\+SC|LOW\+SC|MINIMAL|MODERATE|MEDIUM|HIGH|LOW|N/A|NA|中等程度|中等|最低|极低|高|中|低)",
    re.IGNORECASE,
)


@dataclass
class CRARecord:
    subject_code: str
    account_name: str
    assertion: str
    cra_level: str
    ratio: float | None = None
    ratio_text: str = ""
    applicable: bool = True
    ratio_status: str = ""
    range_status: str = ""
    source: str = "paste"
    confidence: float = 0.0
    apply: bool = True
    applicable: bool = True
    match_status: str = ""
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\uff08", "(").replace("\uff09", ")")
    text = text.replace("\u3001", "/").replace("\uff0c", ",")
    return text


def normalize_assertion(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return ""
    compact = re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", key)
    if key in {"C"} or compact in {"C"} or "COMPLETENESS" in key or "\u5b8c\u6574" in key:
        return "C"
    if (
        "E/O" in key
        or compact in {"EO", "E", "O"}
        or "EXISTENCE" in key
        or "OCCURRENCE" in key
        or "\u5b58\u5728" in key
        or "\u53d1\u751f" in key
    ):
        return "E/O"
    if (
        "V/M" in key
        or "M/V" in key
        or compact in {"VM", "MV", "V", "M"}
        or "VALUATION" in key
        or "MEASUREMENT" in key
        or "\u8ba1\u4ef7" in key
        or "\u8ba1\u91cf" in key
    ):
        return "V/M"
    if (
        "R&O" in key
        or "R & O" in key
        or compact in {"RO", "R"}
        or "RIGHT" in key
        or "OBLIGATION" in key
        or "\u6743\u5229" in key
        or "\u4e49\u52a1" in key
    ):
        return "R&O"
    if (
        "P&D" in key
        or "P & D" in key
        or compact in {"PD", "P", "D"}
        or "PRESENTATION" in key
        or "DISCLOSURE" in key
        or "\u5217\u62a5" in key
        or "\u62ab\u9732" in key
        or "\u8868\u8fbe" in key
    ):
        return "P&D"
    return ""


def normalize_risk_level(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return ""
    if "LOW+SC" in key:
        return "Low"
    for standard, aliases in RISK_ALIASES.items():
        if any(alias in key for alias in aliases):
            return standard
    return ""


def normalize_cra_level(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return ""
    if "HIGH+SC" in key:
        return "High"
    if "LOW+SC" in key:
        return "Low"
    exact_aliases = {
        "Minimal": {"MINIMAL"},
        "Low": {"LOW"},
        "Moderate": {"MODERATE", "MEDIUM"},
        "High": {"HIGH"},
        "N/A": {"N/A", "NA"},
    }
    compact = re.sub(r"[^A-Z/+]+", "", key)
    for standard, aliases in exact_aliases.items():
        if compact in aliases:
            return standard
    for standard, aliases in RISK_ALIASES.items():
        if any(alias in key for alias in aliases):
            return standard
    return ""


def is_not_applicable_value(value: Any) -> bool:
    key = normalize_key(value)
    if not key:
        return False
    compact = re.sub(r"[\s/_-]+", "", key)
    return compact in {
        "N",
        "NO",
        "NA",
        "N/A",
        "\u5426",
        "\u4e0d\u9002\u7528",
    } or "\u4e0d\u9002\u7528" in key


def match_subject(account_name: Any) -> str:
    key = normalize_key(account_name)
    if not key:
        return ""
    compact = re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", key)
    if is_assertion_only_text(key):
        return ""

    code_match = re.match(
        r"^([A-Z][A-Z0-9_]*(?:&[A-Z0-9_]+)?)\s*(?=[.．、:：\-\s]|$)",
        key,
    )
    if code_match:
        explicit_code = code_match.group(1)
        if explicit_code in EXPLICIT_SUBJECT_CODES:
            return EXPLICIT_SUBJECT_CODES[explicit_code]

    exact_code = re.sub(r"[^A-Z0-9_&]+", "", key)
    if exact_code in EXPLICIT_SUBJECT_CODES:
        return EXPLICIT_SUBJECT_CODES[exact_code]

    best_subject = ""
    best_length = 0
    for subject_code, aliases in SUBJECT_ALIASES.items():
        for alias in aliases:
            alias_key = normalize_key(alias)
            alias_compact = re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", alias_key)
            if alias_compact == compact and alias_compact in ASSERTION_ONLY_KEYS:
                continue
            if not alias_key:
                continue
            code_like = bool(re.fullmatch(r"[A-Z0-9_&]+", alias_key))
            if code_like and len(alias_compact) <= 3:
                matched = bool(
                    re.search(
                        rf"(?<![A-Z0-9_]){re.escape(alias_key)}(?![A-Z0-9_])",
                        key,
                    )
                )
            else:
                matched = alias_key in key or alias_compact in compact
            if matched:
                score = max(len(alias_key), len(alias_compact))
                if score > best_length:
                    best_subject = subject_code
                    best_length = score
    return best_subject


def parse_ratio(value: Any) -> tuple[float | None, str, str]:
    if value is None:
        return None, "", "\u672a\u63d0\u4f9b\u6bd4\u4f8b\uff0c\u8bf7\u786e\u8ba4\u590d\u5236\u533a\u57df\u5305\u542b Preliminary Scope/\u6d4b\u8bd5\u754c\u9650\u5217\uff1b\u91c7\u7528\u6a21\u677f\u9ed8\u8ba4\u516c\u5f0f"
    text = str(value).strip()
    if not text:
        return None, "", "\u672a\u63d0\u4f9b\u6bd4\u4f8b\uff0c\u8bf7\u786e\u8ba4\u590d\u5236\u533a\u57df\u5305\u542b Preliminary Scope/\u6d4b\u8bd5\u754c\u9650\u5217\uff1b\u91c7\u7528\u6a21\u677f\u9ed8\u8ba4\u516c\u5f0f"
    key = normalize_key(text)
    if key in {"N/A", "NA"} or key.startswith("N/A "):
        return None, "", "\u672a\u63d0\u4f9b\u5355\u4e00\u6bd4\u4f8b\uff08N/A\uff09\uff0c\u5df2\u4fdd\u7559\u6a21\u677f\u9ed8\u8ba4\u516c\u5f0f"
    if re.search(r"\d+(?:\.\d+)?\s*[%\uff05]?\s*[-~]\s*\d+(?:\.\d+)?\s*[%\uff05]", text):
        return None, text, "\u63d0\u4f9b\u7684\u662f\u533a\u95f4\uff0c\u8bf7\u5728\u9884\u89c8\u8868\u786e\u8ba4\u5355\u4e00\u6bd4\u4f8b"
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*[%\uff05]", text)
    if match:
        ratio = float(match.group(1)) / 100
        return ratio, f"{match.group(1)}%", "\u5df2\u8bc6\u522b\u6bd4\u4f8b"
    if re.search(r"\d+(?:\.\d+)?\s*[%％﹪]?\s*[-~－—]\s*\d+(?:\.\d+)?\s*[%％﹪]", text):
        return None, text, "\u63d0\u4f9b\u7684\u662f\u533a\u95f4\uff0c\u8bf7\u5728\u9884\u89c8\u8868\u786e\u8ba4\u5355\u4e00\u6bd4\u4f8b"
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*[%％﹪]", text)
    if match:
        ratio = float(match.group(1)) / 100
        return ratio, f"{match.group(1)}%", "\u5df2\u8bc6\u522b\u6bd4\u4f8b"
    match = re.search(r"(-?\d+(?:\.\d+)?)", text.replace(",", ""))
    if not match:
        return None, text, "\u672a\u8bc6\u522b\u6bd4\u4f8b\uff0c\u91c7\u7528\u6a21\u677f\u9ed8\u8ba4\u516c\u5f0f"
    number = float(match.group(1))
    ratio = number / 100 if number > 1 else number
    return ratio, f"{ratio:.0%}", "\u5df2\u8bc6\u522b\u6bd4\u4f8b"


def is_assertion_only_text(value: Any) -> bool:
    key = normalize_key(value)
    if not key:
        return False
    if re.search(r"[\u4e00-\u9fff]", key):
        return False
    compact = re.sub(r"[^A-Z/&]+", "", key)
    return key in ASSERTION_ONLY_KEYS or compact in ASSERTION_ONLY_KEYS


def account_threshold_family(subject_code: str = "", account_name: Any = "") -> str:
    if subject_code in LIABILITY_EXPENSE_SUBJECTS:
        return "liability_expense"
    if subject_code in ASSET_REVENUE_SUBJECTS:
        return "asset_revenue"
    key = normalize_key(account_name)
    if any(token in key for token in LIABILITY_EXPENSE_ACCOUNT_TOKENS):
        return "liability_expense"
    if any(token in key for token in ASSET_REVENUE_ACCOUNT_TOKENS):
        return "asset_revenue"
    return ""


def threshold_family(subject_code: str, assertion: str = "", account_name: Any = "") -> str:
    return account_threshold_family(subject_code, account_name)


def check_ratio_range(subject_code: str, assertion: str, cra_level: str, ratio: float | None, account_name: Any = "") -> str:
    if ratio is None:
        return "\u672a\u68c0\u67e5\uff08\u65e0\u5355\u4e00\u6bd4\u4f8b\uff09"
    family = threshold_family(subject_code, assertion, account_name)
    if not family:
        return "\u672a\u68c0\u67e5\uff08\u65e0\u6cd5\u5224\u65ad\u8d26\u6237\u6027\u8d28\uff09"
    ranges = RATIO_RANGES.get(family, {})
    if cra_level not in ranges:
        return "\u672a\u68c0\u67e5\uff08CRA\u7b49\u7ea7\u4e0d\u9002\u7528\uff09"
    low, high = ranges[cra_level]
    if low <= ratio <= high:
        return "\u901a\u8fc7"
    return f"\u8d85\u51fa\u5efa\u8bae\u533a\u95f4 {low:.0%}-{high:.0%}"


def _basic_split_pasted_rows(text: str) -> list[list[str]]:
    if "\t" in text:
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        return [[str(cell or "").strip() for cell in row] for row in reader if any(str(cell or "").strip() for cell in row)]

    rows = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "\t" in raw_line:
            cells = [cell.strip() for cell in raw_line.split("\t")]
        elif "," in raw_line and raw_line.count(",") >= 3:
            cells = [cell.strip() for cell in raw_line.split(",")]
        else:
            cells = [cell.strip() for cell in re.split(r"\s{2,}", raw_line.strip()) if cell.strip()]
        if cells:
            rows.append(cells)
    return rows


def _rebuild_verticalized_excel_rows(text: str) -> list[list[str]]:
    cells = [line.strip() for line in text.splitlines() if line.strip()]
    if len(cells) < 12:
        return []

    header_start = None
    header_end = None
    for start in range(min(len(cells), 60)):
        roles = [(offset, header_role(cell)) for offset, cell in enumerate(cells[start : start + 14])]
        role_names = {role for _, role in roles if role}
        if "cra" not in role_names or "account" not in role_names:
            continue
        ratio_offsets = [offset for offset, role in roles if role == "ratio"]
        if ratio_offsets:
            end = ratio_offsets[0]
        else:
            cra_offsets = [offset for offset, role in roles if role == "cra"]
            end = cra_offsets[0] if cra_offsets else 0
        if 2 <= end <= 12:
            header_start = start
            header_end = start + end
            break
    if header_start is None or header_end is None:
        return []

    header = cells[header_start : header_end + 1]
    rows: list[list[str]] = [header]
    data = cells[header_end + 1 :]
    max_width = max(len(header) + 2, 6)
    last_account = ""

    def starts_account_assertion(cell: str) -> bool:
        return is_account_assertion_cell(cell)

    def starts_account_then_assertion(index: int) -> bool:
        if index + 1 >= len(data):
            return False
        cell = data[index]
        next_cell = data[index + 1]
        if not cell or normalize_risk_level(cell) or is_percent_like(cell) or is_numeric_only(cell):
            return False
        return bool(match_subject(cell) and normalize_assertion(next_cell))

    def starts_assertion_only(cell: str) -> bool:
        return bool(normalize_assertion(cell) and not normalize_risk_level(cell) and not is_percent_like(cell) and not is_numeric_only(cell))

    index = 0
    while index < len(data):
        cell = data[index]
        if starts_account_assertion(cell):
            row = [cell]
            account, _ = split_account_assertion(cell)
            last_account = account or last_account
            index += 1
        elif starts_account_then_assertion(index):
            row = [cell, data[index + 1]]
            last_account = clean_account_name(cell) or last_account
            index += 2
        elif starts_assertion_only(cell) and last_account:
            row = ["", cell]
            index += 1
        else:
            index += 1
            continue

        while index < len(data) and len(row) < max_width:
            next_cell = data[index]
            if starts_account_assertion(next_cell) or starts_account_then_assertion(index) or (
                starts_assertion_only(next_cell) and last_account
            ):
                break
            row.append(next_cell)
            index += 1
        if len(row) > 1:
            rows.append(row)

    return rows if len(rows) > 1 else []


def split_pasted_rows(text: str) -> list[list[str]]:
    rows = _basic_split_pasted_rows(text)
    if "\t" not in text and rows and sum(1 for row in rows if len(row) == 1) / max(len(rows), 1) > 0.85:
        rebuilt = _rebuild_verticalized_excel_rows(text)
        if rebuilt:
            return rebuilt
    return rows


def header_role(value: str) -> str:
    key = normalize_key(value)
    if any(token in key for token in ("APPLICABLE", "是否适用", "适用性", "适用")):
        return "applicable"
    if any(token in key for token in ("ACCOUNT", "ACCOUNTS", "SIGNIFICANT ACCOUNT", "\u79d1\u76ee", "\u8d26\u6237", "\u5e10\u6237", "\u62ab\u9732")):
        return "account"
    if any(token in key for token in ("ASSERTION", "ASSERTIONS", "\u8ba4\u5b9a")):
        return "assertion"
    if "CRA" in key or "\u7efc\u5408\u98ce\u9669" in key or "\u98ce\u9669\u7b49\u7ea7" in key:
        return "cra"
    if any(token in key for token in ("RATIO", "% OF TE", "THRESHOLD", "SCOPE", "\u6bd4\u4f8b", "\u6d4b\u8bd5\u754c\u9650", "\u6d4b\u8bd5\u9608\u503c")):
        return "ratio"
    if any(token in key for token in ("APPLICABLE", "APPLY", "\u662f\u5426\u9002\u7528", "\u9002\u7528\u6027")):
        return "applicable"
    return ""


def is_not_applicable_value(value: Any) -> bool:
    key = normalize_key(value)
    if not key:
        return False
    compact = re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", key)
    return key in {"N", "NO", "N/A", "NA"} or compact in {
        "N",
        "NO",
        "NA",
        "\u5426",
        "\u4e0d\u9002\u7528",
        "\u4e0d\u9069\u7528",
    }


def find_header(rows: list[list[str]], cra_header_preference: str = "") -> tuple[int | None, dict[str, int]]:
    best_index = None
    best_map = {}
    best_score = 0
    preference = normalize_key(cra_header_preference)
    for index, row in enumerate(rows[:20]):
        role_map = {}
        for col, value in enumerate(row):
            role = header_role(value)
            if role == "cra" and preference and preference in normalize_key(value):
                role_map[role] = col
                continue
            if role and role not in role_map:
                role_map[role] = col
        score = sum(role in role_map for role in ("account", "assertion", "cra"))
        if score > best_score:
            best_index = index
            best_map = role_map
            best_score = score
    if best_score >= 2:
        return best_index, best_map
    return None, {}


def detect_cra_header_options(text: str) -> list[str]:
    rows = split_pasted_rows(text)
    options: list[str] = []
    for row in rows[:20]:
        row_options = [str(value).strip() for value in row if header_role(value) == "cra" and str(value).strip()]
        if not row_options:
            continue
        role_count = sum(1 for value in row if header_role(value) in {"account", "assertion", "cra", "ratio", "applicable"})
        if role_count >= 1:
            for option in row_options:
                if option not in options:
                    options.append(option)
            break
    return options


def clean_account_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.split(
        r"\s+(?:HIGH\+SC|LOW\+SC|MINIMAL|MODERATE|MEDIUM|HIGH|LOW|N/A|NA|较高|较低|中等|最低|极低|高|中|低)",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0].strip()
    text = re.sub(r"^[A-Z]{1,4}\d*\s*[.．、]\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^\d+\s*[.．、]\s*", "", text)
    return text.strip()


def split_account_assertion(value: Any) -> tuple[str, str]:
    account = clean_account_name(value)
    match = re.match(r"(.+)\s*[-－–—]\s*([^-－–—]+)$", account)
    if not match:
        return account, ""
    base = match.group(1).strip()
    suffix = match.group(2).strip()
    assertion = normalize_assertion(suffix)
    if not assertion:
        return account, ""
    return base, assertion


def is_percent_like(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(re.search(r"\d+(?:\.\d+)?\s*[%％﹪]?(?:\s*[-~－—]\s*\d+(?:\.\d+)?\s*)?[%％﹪]", text))


def is_numeric_only(value: Any) -> bool:
    text = str(value or "").strip().replace(",", "")
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text))


def account_candidate_score(value: Any) -> int:
    text = str(value or "").strip()
    if not text or is_percent_like(text) or is_numeric_only(text):
        return -1
    if is_assertion_only_text(text):
        return -1
    if normalize_risk_level(text) and not (normalize_assertion(text) or match_subject(text)):
        return -1
    score = 0
    if match_subject(text):
        score += 8
    if normalize_assertion(text):
        score += 4
    if re.search(r"[\u4e00-\u9fff]", text):
        score += 3
    if "-" in text or "－" in text or "/" in text:
        score += 1
    return score


def find_ratio_cell(row: list[str]) -> tuple[int | None, str]:
    for index in range(len(row) - 1, -1, -1):
        cell = row[index]
        if is_percent_like(cell):
            return index, cell
    for index in range(len(row) - 1, -1, -1):
        cell = row[index]
        if not is_numeric_only(cell):
            continue
        try:
            number = float(str(cell).strip().replace(",", ""))
        except ValueError:
            continue
        if 0 <= number <= 1:
            return index, cell
    return None, ""


def infer_applicable_cell(row: list[str], ratio_index: int | None = None) -> str:
    """Find a Y/N-style applicability marker when the exported header is blank."""
    search_end = ratio_index if ratio_index is not None else len(row)
    for index in range(search_end - 1, -1, -1):
        value = str(row[index] or "").strip()
        key = normalize_key(value)
        compact = re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", key)
        if key in {"Y", "YES", "N", "NO", "是", "否", "适用", "不适用", "不適用"}:
            return value
        if key in {"N/A", "NA"} and index + 1 < len(row):
            next_value = row[index + 1]
            if is_percent_like(next_value) or normalize_key(next_value) in {"N/A", "NA"}:
                return value
        if compact in {"YES", "NO", "是", "否", "适用", "不适用", "不適用"}:
            return value
    return ""


def detect_section_account(row: list[str]) -> str:
    """Recognize a standalone CRA account heading before its assertion rows."""
    populated = [str(cell or "").strip() for cell in row if str(cell or "").strip()]
    if not populated:
        return ""
    if any(normalize_risk_level(cell) or is_percent_like(cell) for cell in populated):
        return ""
    if any(is_account_assertion_cell(cell) for cell in populated):
        return ""

    for index, cell in enumerate(populated):
        subject_code = match_subject(cell)
        if not subject_code:
            continue
        if cell.upper() in EXPLICIT_SUBJECT_CODES and index + 1 < len(populated):
            next_cell = populated[index + 1]
            if match_subject(next_cell) == subject_code:
                return clean_account_name(next_cell)
        return clean_account_name(cell)
    return ""


def find_cra_cell(row: list[str], ratio_index: int | None) -> tuple[int | None, str, str]:
    search_end = ratio_index if ratio_index is not None else len(row)
    for index in range(search_end - 1, -1, -1):
        raw = row[index]
        level = normalize_cra_level(raw)
        if level:
            return index, raw, level
    return None, "", ""


def find_embedded_risk(text: str, before_index: int | None = None) -> tuple[str, str, int | None]:
    search_text = text[:before_index] if before_index is not None else text
    best_raw = ""
    best_level = ""
    best_start = None
    for match in CRA_TOKEN_RE.finditer(search_text):
        raw = match.group(0)
        level = normalize_cra_level(raw)
        if not level:
            continue
        best_raw = raw
        best_level = level
        best_start = match.start()
    return best_raw, best_level, best_start


def guess_embedded_record_from_text(text: str) -> tuple[str, str, str, str, str]:
    joined = str(text or "").strip()
    if not joined:
        return "", "", "", "", ""

    ratio_match = re.search(r"\d+(?:\.\d+)?\s*[%％﹪]?(?:\s*[-~－—]\s*\d+(?:\.\d+)?\s*)?[%％﹪]", joined)
    ratio_raw = ratio_match.group(0) if ratio_match else ""
    ratio_start = ratio_match.start() if ratio_match else None
    cra_raw, cra_level, risk_start = find_embedded_risk(joined, ratio_start)
    note = ""
    if cra_raw and "+SC" in normalize_key(cra_raw):
        note = f"CRA原值 {cra_raw} 已标准化为 {cra_level}"

    account_part = joined[:risk_start].strip() if risk_start is not None else joined
    account = clean_account_name(account_part)
    assertion = normalize_assertion(account) or normalize_assertion(joined)
    return account, assertion, cra_level, ratio_raw, note


def is_account_assertion_cell(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if is_percent_like(text) or is_numeric_only(text):
        return False
    if normalize_cra_level(text):
        return False
    if normalize_key(text) in {"+SC", "SC"}:
        return False
    if is_assertion_only_text(text):
        return False
    account, assertion = split_account_assertion(text)
    if not account or not assertion:
        return False
    if is_assertion_only_text(account):
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z]", account))


def extract_cra_table_record(row: list[str]) -> tuple[str, str, str, str, str]:
    if len(row) <= 1:
        return "", "", "", "", ""

    ratio_index, ratio = find_ratio_cell(row)
    cra_index, cra_raw, cra_level = find_cra_cell(row, ratio_index)
    if not cra_level:
        return "", "", "", "", ""

    account_search_end = cra_index if cra_index is not None else len(row)
    candidates = row[:account_search_end] if account_search_end > 0 else row
    account_cell = ""
    for cell in candidates:
        if is_account_assertion_cell(cell):
            account_cell = cell
            break
    if not account_cell:
        for cell in row:
            if is_account_assertion_cell(cell):
                account_cell = cell
                break
    if not account_cell:
        return "", "", "", "", ""

    account, assertion = split_account_assertion(account_cell)
    note = ""
    if "+SC" in normalize_key(cra_raw):
        note = f"CRA原值 {cra_raw} 已标准化为 {cra_level}"
    return account, assertion, cra_level, ratio, note


def guess_record_from_row(row: list[str]) -> tuple[str, str, str, str, str]:
    account = ""
    assertion = ""
    cra_level = ""
    ratio = ""
    note = ""
    joined = " ".join(str(cell or "").strip() for cell in row if str(cell or "").strip())

    if len(row) == 1:
        return guess_embedded_record_from_text(joined)

    ratio_index, ratio = find_ratio_cell(row)
    cra_index, cra_raw, cra_level = find_cra_cell(row, ratio_index)
    if cra_raw and normalize_key(cra_raw) != normalize_key(cra_level):
        if "+SC" in normalize_key(cra_raw):
            note = f"CRA原值 {cra_raw} 已标准化为 {cra_level}"

    account_search_end = cra_index if cra_index is not None else (ratio_index if ratio_index is not None else len(row))
    candidates = row[:account_search_end] if account_search_end > 0 else row
    best = max(candidates, key=account_candidate_score, default="")
    if account_candidate_score(best) >= 0:
        account = clean_account_name(best)

    if account:
        assertion = normalize_assertion(account)
    if not assertion:
        for cell in candidates:
            if normalize_risk_level(cell) or is_percent_like(cell):
                continue
            assertion = normalize_assertion(cell)
            if assertion:
                break
    if not (account and assertion and cra_level):
        embedded_account, embedded_assertion, embedded_cra, embedded_ratio, embedded_note = guess_embedded_record_from_text(joined)
        account = account or embedded_account
        assertion = assertion or embedded_assertion
        cra_level = cra_level or embedded_cra
        ratio = ratio or embedded_ratio
        note = note or embedded_note
    return account, assertion, cra_level, ratio, note


def parse_cra_paste_text(
    text: str,
    selected_subjects: list[str] | None = None,
    cra_header_preference: str = "",
    _debug_traces: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = split_pasted_rows(text)
    if not rows:
        return []

    selected_subjects = selected_subjects or []
    header_index, col_map = find_header(rows, cra_header_preference)
    data_rows = rows[header_index + 1 :] if header_index is not None else rows
    records: list[CRARecord] = []
    current_account = ""

    for row_index, row in enumerate(data_rows):
        section_account = detect_section_account(row)
        if section_account:
            current_account = section_account
            continue

        account_raw = ""
        assertion_raw = ""
        cra_raw = ""
        ratio_raw = ""
        applicable_raw = ""
        if header_index is not None:
            has_account_col = "account" in col_map and col_map.get("account", 999) < len(row)
            account_raw = row[col_map["account"]] if has_account_col else ""
            assertion_raw = row[col_map["assertion"]] if col_map.get("assertion", 999) < len(row) else ""
            cra_raw = row[col_map["cra"]] if col_map.get("cra", 999) < len(row) else ""
            ratio_raw = row[col_map["ratio"]] if col_map.get("ratio", 999) < len(row) else ""
            applicable_raw = row[col_map["applicable"]] if col_map.get("applicable", 999) < len(row) else ""
            note = ""
            guessed_account, guessed_assertion, guessed_cra, guessed_ratio, guessed_note = guess_record_from_row(row)
            cleaned_account = clean_account_name(account_raw)
            account_raw_assertion = normalize_assertion(account_raw)
            if cleaned_account and not is_assertion_only_text(cleaned_account):
                account = cleaned_account
            elif guessed_account and not is_assertion_only_text(guessed_account):
                account = clean_account_name(guessed_account)
            elif current_account:
                account = current_account
            else:
                account = guessed_account
            assertion = normalize_assertion(assertion_raw) or account_raw_assertion or normalize_assertion(account) or guessed_assertion
            cra_level = normalize_cra_level(cra_raw)
            if not cra_level:
                cra_level = guessed_cra
            if guessed_ratio and (not ratio_raw or not is_percent_like(ratio_raw)):
                ratio_raw = guessed_ratio
            elif not ratio_raw:
                ratio_raw = guessed_ratio
            if not applicable_raw:
                guessed_ratio_index, _ = find_ratio_cell(row)
                applicable_raw = infer_applicable_cell(row, guessed_ratio_index)
            if "+SC" in normalize_key(cra_raw):
                note = f"CRA原值 {cra_raw} 已标准化为 {cra_level}"
            if not note:
                note = guessed_note
        else:
            table_account, table_assertion, table_cra, table_ratio, table_note = extract_cra_table_record(row)
            if table_account and table_assertion and table_cra:
                account = table_account
                assertion = table_assertion
                cra_level = table_cra
                ratio_raw = table_ratio
                note = table_note
            else:
                account, assertion, cra_level, ratio_raw, note = guess_record_from_row(row)
                if (not account or is_assertion_only_text(account)) and current_account:
                    account = current_account

        if account and assertion and (not cra_level or not ratio_raw):
            for extra_row in data_rows[row_index + 1 : row_index + 4]:
                extra_account, extra_assertion, extra_cra, extra_ratio, extra_note = guess_record_from_row(extra_row)
                if extra_account and extra_assertion:
                    break
                if not cra_level and extra_cra:
                    cra_level = extra_cra
                    note = note or extra_note
                if not ratio_raw and extra_ratio:
                    ratio_raw = extra_ratio
                if cra_level and ratio_raw:
                    break

        if account and not is_assertion_only_text(account) and not normalize_cra_level(account) and not is_percent_like(account) and not is_numeric_only(account):
            current_account = account

        if not assertion or not cra_level:
            continue

        account, suffix_assertion = split_account_assertion(account)
        if suffix_assertion:
            assertion = suffix_assertion

        if (not account or is_assertion_only_text(account)) and current_account:
            account = current_account

        if not account or normalize_key(account) in {"+SC", "SC"}:
            continue
        if normalize_cra_level(account) or is_percent_like(account) or is_numeric_only(account):
            continue
        if is_assertion_only_text(account):
            continue
        if normalize_assertion(account) and not match_subject(account):
            continue

        subject_code = match_subject(account)
        if subject_code:
            current_account = account
        if subject_code and selected_subjects and subject_code not in selected_subjects:
            confidence = 0.55
        else:
            confidence = 0.90 if subject_code else 0.45

        applicable = not is_not_applicable_value(applicable_raw)
        if not applicable:
            cra_level = "N/A"
            ratio_raw = "N/A"

        ratio, ratio_text, ratio_status = parse_ratio(ratio_raw)
        if not applicable:
            ratio = None
            ratio_text = "N/A"
            ratio_status = "\u4e0d\u9002\u7528"
        range_status = check_ratio_range(subject_code, assertion, cra_level, ratio, account)
        note_parts = []
        if note:
            note_parts.append(note)
        if not applicable:
            note_parts.append("\u662f\u5426\u9002\u7528=N\uff0c\u5df2\u5199\u5165 N/A")
        if not subject_code:
            note_parts.append("\u672a\u5339\u914d\u5e95\u7a3f\u79d1\u76ee")
        elif not selected_subjects or subject_code not in selected_subjects:
            note_parts.append(f"\u672c\u6b21\u672a\u9009\u62e9 {subject_code}")
        if ratio is None:
            note_parts.append(ratio_status)
        if range_status.startswith("\u8d85\u51fa"):
            note_parts.append(range_status)

        if not subject_code:
            match_status = "\u4e0d\u5199\u5165-\u672a\u5339\u914d\u5230\u5e95\u7a3f\u79d1\u76ee"
        elif ratio_status.startswith("\u63d0\u4f9b\u7684\u662f\u533a\u95f4"):
            match_status = "\u9700\u786e\u8ba4-\u6bd4\u4f8b\u4e3a\u533a\u95f4"
        elif not selected_subjects or subject_code not in selected_subjects:
            match_status = "\u4e0d\u5199\u5165-\u672a\u9009\u62e9\u8be5\u79d1\u76ee"
        else:
            match_status = "\u5c06\u5199\u5165"

        record = CRARecord(
            subject_code=subject_code,
            account_name=str(account or ""),
            assertion=assertion,
            cra_level=cra_level,
            ratio=ratio,
            ratio_text=ratio_text,
            ratio_status=ratio_status,
            range_status=range_status,
            confidence=confidence,
            apply=(match_status == "\u5c06\u5199\u5165"),
            applicable=applicable,
            match_status=match_status,
            note="; ".join(note_parts),
        )
        records.append(record)
        if _debug_traces is not None:
            _debug_traces.append(
                {
                    "row": list(row),
                    "account_raw": account_raw,
                    "assertion_raw": assertion_raw,
                    "cra_raw": cra_raw,
                    "ratio_raw": ratio_raw,
                    "applicable_raw": applicable_raw,
                    "record": record.to_dict(),
                }
            )

    return [record.to_dict() for record in records]


def build_cra_parse_debug_report(
    text: str,
    selected_subjects: list[str] | None = None,
    cra_header_preference: str = "",
) -> str:
    rows = split_pasted_rows(text)
    header_index, col_map = find_header(rows, cra_header_preference)
    traces: list[dict[str, Any]] = []
    parse_cra_paste_text(text, selected_subjects, cra_header_preference, _debug_traces=traces)
    lines = [
        "CRA parse debug",
        f"generated_at={datetime.now().isoformat(timespec='seconds')}",
        f"selected_subjects={repr(selected_subjects or [])}",
        f"cra_header_preference={repr(cra_header_preference)}",
        "",
        "raw_text_repr_first_30_lines:",
    ]
    raw_lines = text.splitlines()
    for index, line in enumerate(raw_lines[:30], 1):
        lines.append(f"{index:02d}: {repr(line)}")
    lines.extend([
        "",
        "split_pasted_rows_first_30:",
    ])
    for index, row in enumerate(rows[:30], 1):
        rendered = ", ".join(f"{col}:{repr(cell)}" for col, cell in enumerate(row))
        lines.append(f"{index:02d}: cell_count={len(row)} [{rendered}]")
    lines.extend([
        "",
        f"header_index={repr(header_index)}",
        f"col_map={repr(col_map)}",
    ])
    if header_index is not None and header_index < len(rows):
        lines.append("header_cells:")
        for col, cell in enumerate(rows[header_index]):
            lines.append(f"  {col}: {repr(cell)} role={repr(header_role(cell))}")
    lines.extend([
        "",
        "first_10_record_trace:",
    ])
    for index, trace in enumerate(traces[:10], 1):
        final = trace.get("record", {})
        lines.append(f"{index:02d}: row={repr(trace.get('row', []))}")
        lines.append(
            "    raw="
            + repr(
                {
                    "account_raw": trace.get("account_raw", ""),
                    "assertion_raw": trace.get("assertion_raw", ""),
                    "cra_raw": trace.get("cra_raw", ""),
                    "ratio_raw": trace.get("ratio_raw", ""),
                }
            )
        )
        lines.append(
            "    final="
            + repr(
                {
                    "account": final.get("account_name", ""),
                    "assertion": final.get("assertion", ""),
                    "cra_level": final.get("cra_level", ""),
                    "ratio_text": final.get("ratio_text", ""),
                }
            )
        )
    return "\n".join(lines) + "\n"


def write_cra_parse_debug_log(
    text: str,
    selected_subjects: list[str] | None = None,
    cra_header_preference: str = "",
    path: str | Path | None = None,
) -> Path:
    log_path = Path(path) if path else Path(os.getenv("APPDATA", str(Path.home()))) / "AuditRollForward" / "logs" / "cra_parse_debug.txt"
    report = build_cra_parse_debug_report(text, selected_subjects, cra_header_preference)
    try:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text(report, encoding="utf-8")
        return log_path
    except OSError:
        if path is not None:
            raise
        fallback_path = Path.cwd() / "cra_parse_debug.txt"
        fallback_path.write_text(report, encoding="utf-8")
        return fallback_path


def normalize_records(records: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    normalized = []
    for record in records or []:
        if not record or not record.get("apply", True):
            continue
        subject_code = str(record.get("subject_code") or "").strip()
        assertion = normalize_assertion(record.get("assertion"))
        cra_level = normalize_risk_level(record.get("cra_level"))
        applicable = bool(record.get("applicable", True))
        if not applicable:
            cra_level = "N/A"
        if not subject_code or not assertion or not cra_level:
            continue
        ratio = record.get("ratio")
        if not applicable:
            ratio = None
        elif ratio in ("", None):
            ratio = None
        else:
            try:
                ratio = float(ratio)
                if ratio > 1:
                    ratio = ratio / 100
            except (TypeError, ValueError):
                ratio = parse_ratio(record.get("ratio_text"))[0]
        normalized.append({
            "subject_code": subject_code,
            "account_name": str(record.get("account_name") or ""),
            "assertion": assertion,
            "cra_level": cra_level,
            "ratio": ratio,
            "ratio_text": "N/A" if not applicable else record.get("ratio_text") or (f"{ratio:.0%}" if ratio is not None else ""),
            "range_status": check_ratio_range(subject_code, assertion, cra_level, ratio, record.get("account_name")),
            "applicable": applicable,
            "note": str(record.get("note") or ""),
        })
    return normalized


def find_cra_tables(ws) -> list[dict[str, int]]:
    tables = []
    for row in range(1, min(ws.max_row, 80) + 1):
        roles = {}
        for col in range(1, min(ws.max_column, 12) + 1):
            role = header_role(ws.cell(row=row, column=col).value or "")
            if role in {"assertion", "cra", "ratio"} and role not in roles:
                roles[role] = col
        if "assertion" in roles and "cra" in roles:
            threshold_col = roles.get("ratio")
            if not threshold_col:
                for col in range(roles["cra"] + 1, min(ws.max_column, roles["cra"] + 4) + 1):
                    value = normalize_key(ws.cell(row=row, column=col).value)
                    if value and any(token in value for token in ("THRESHOLD", "\u5404\u9879\u8ba4\u5b9a", "\u6240\u6709\u76f8\u5173\u8ba4\u5b9a")):
                        threshold_col = col
                        break
            tables.append({
                "header_row": row,
                "assertion_col": roles["assertion"],
                "cra_col": roles["cra"],
                "threshold_col": threshold_col or 0,
            })
    return tables


def iter_cra_table_assertion_rows(ws, table: dict[str, int]) -> list[tuple[int, str]]:
    assertion_col = table["assertion_col"]
    rows: list[tuple[int, str]] = []
    blank_after_start = 0
    max_scan_row = min(ws.max_row, table["header_row"] + 20)
    for row in range(table["header_row"] + 1, max_scan_row + 1):
        assertion = normalize_assertion(ws.cell(row=row, column=assertion_col).value)
        if assertion:
            rows.append((row, assertion))
            blank_after_start = 0
            continue
        if rows:
            blank_after_start += 1
            if blank_after_start >= 1:
                break
    return rows


def extract_te_reference(formula: Any) -> str:
    if not isinstance(formula, str):
        return ""

    def is_row_five(ref: str) -> bool:
        match = re.search(r"(\d+)$", ref)
        return bool(match and int(match.group(1)) == 5)

    refs_after_multiply = re.findall(r"\*\s*((?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+)", formula)
    for ref in reversed(refs_after_multiply):
        if is_row_five(ref):
            return ref
    refs = re.findall(r"((?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+)", formula)
    for ref in refs:
        if is_row_five(ref):
            return ref
    return ""


def excel_percent_literal(ratio: float) -> str:
    percent = ratio * 100
    if abs(percent - round(percent)) < 0.000001:
        return f"{int(round(percent))}%"
    return f"{percent:.4f}%"


def build_record_lookup(records: list[dict[str, Any]], subject_code: str) -> dict[str, dict[str, Any]]:
    lookup = {}
    for record in normalize_records(records):
        if record["subject_code"] != subject_code:
            continue
        lookup[record["assertion"]] = record
    return lookup


def apply_cra_records_to_workbook(wb, subject_code: str, records: list[dict[str, Any]] | None) -> dict[str, Any]:
    lookup = build_record_lookup(records or [], subject_code)
    result = {
        "applied": 0,
        "missing_ratio": 0,
        "out_of_range": 0,
        "default_na": 0,
        "matched_assertions": [],
        "messages": [],
    }
    if not lookup:
        return result

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF99")
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for table in find_cra_tables(ws):
            cra_col = table["cra_col"]
            threshold_col = table["threshold_col"]
            for row, assertion in iter_cra_table_assertion_rows(ws, table):
                cra_cell = ws.cell(row=row, column=cra_col)
                if isinstance(cra_cell, MergedCell):
                    continue
                if assertion in lookup:
                    record = lookup[assertion]
                    cra_cell.value = record["cra_level"]
                    cra_cell.fill = copy(yellow_fill)
                    result["applied"] += 1
                    result["matched_assertions"].append(assertion)

                    if record.get("applicable") is False or record.get("cra_level") == "N/A":
                        if threshold_col:
                            threshold_cell = ws.cell(row=row, column=threshold_col)
                            if not isinstance(threshold_cell, MergedCell):
                                threshold_cell.value = "N/A"
                                threshold_cell.fill = copy(yellow_fill)
                        continue

                    ratio = record.get("ratio")
                    if ratio is None:
                        result["missing_ratio"] += 1
                        continue
                    if threshold_col:
                        threshold_cell = ws.cell(row=row, column=threshold_col)
                        if not isinstance(threshold_cell, MergedCell):
                            te_ref = extract_te_reference(threshold_cell.value)
                            if te_ref:
                                threshold_cell.value = f"={te_ref}*{excel_percent_literal(float(ratio))}"
                            else:
                                threshold_cell.value = excel_percent_literal(float(ratio))
                            threshold_cell.fill = copy(yellow_fill)

                    if str(record.get("range_status", "")).startswith("\u8d85\u51fa"):
                        result["out_of_range"] += 1
                else:
                    cra_cell.value = "N/A"
                    cra_cell.fill = copy(yellow_fill)
                    result["default_na"] += 1

    if result["applied"]:
        result["messages"].append(
            f"CRA\u5df2\u5199\u5165 {result['applied']} \u4e2a\u8ba4\u5b9a\uff0c\u8bf7\u590d\u6838\u9ec4\u8272\u533a\u57df"
        )
    if result["missing_ratio"]:
        result["messages"].append(
            f"{result['missing_ratio']} \u4e2aCRA\u8bb0\u5f55\u672a\u63d0\u4f9b\u6bd4\u4f8b\uff0c\u5df2\u4fdd\u7559\u6a21\u677f\u9ed8\u8ba4\u516c\u5f0f"
        )
    if result["default_na"]:
        result["messages"].append(
            f"{result['default_na']} \u4e2a\u540c\u79d1\u76ee\u672a\u89e3\u6790\u8ba4\u5b9a\u5df2\u6309\u89c4\u5219\u5199\u5165 N/A"
        )
    if result["out_of_range"]:
        result["messages"].append(
            f"{result['out_of_range']} \u4e2aCRA\u6bd4\u4f8b\u8d85\u51fa\u5efa\u8bae\u533a\u95f4\uff0c\u8bf7\u9879\u76ee\u7ec4\u786e\u8ba4"
        )
    if not result["applied"] and lookup:
        result["messages"].append("CRA\u8bb0\u5f55\u5df2\u63d0\u4f9b\uff0c\u4f46\u672a\u5339\u914d\u5230\u5f53\u524d\u5e95\u7a3fCRA\u8868")
    return result
