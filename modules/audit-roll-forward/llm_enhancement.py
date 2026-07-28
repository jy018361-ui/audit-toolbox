#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Optional LLM enhancement layer for Audit Roll Forward.

The core roll-forward logic remains deterministic. This module only:
1. prechecks workbook structure and asks for mapping suggestions,
2. revises copied wording cells that are already marked for review,
3. writes review/report sheets.
"""

import datetime
import json
import os
import re
import urllib.error
import urllib.request
from copy import copy
from dataclasses import dataclass

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


LLM_REVIEW_SHEET_NAME = "LLM Review"
LLM_WORDING_SHEET_NAME = "LLM Wording Changes"
LLM_MAX_STRUCTURE_ROWS = 80
LLM_MAX_STRUCTURE_COLS = 20
LLM_MAX_WORDING_CELLS = 120


@dataclass
class LLMConfig:
    enabled: bool
    api_key: str = ""
    model: str = "gpt-4o-mini"
    base_url: str = "https://api.openai.com/v1"
    timeout: int = 45

    @property
    def available(self):
        return self.enabled and bool(self.api_key)


def get_llm_config(enabled, options=None):
    """Load LLM settings from environment variables."""
    options = options or {}
    return LLMConfig(
        enabled=bool(enabled),
        api_key=(options.get("api_key") or os.getenv("OPENAI_API_KEY", "")).strip(),
        model=(
            options.get("model")
            or os.getenv("AUDIT_RF_LLM_MODEL")
            or os.getenv("OPENAI_MODEL")
            or "gpt-4o-mini"
        ).strip(),
        base_url=(options.get("base_url") or os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")).rstrip("/"),
        timeout=int(os.getenv("AUDIT_RF_LLM_TIMEOUT", "45")),
    )


def parse_date(value):
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def normalize_text(value):
    text = str(value or "")
    return re.sub(r"\s+", "", text)


def cell_fill_key(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return str(color.rgb or color.indexed or color.theme or "")


def is_yellow_fill(cell):
    key = cell_fill_key(cell)
    return bool(key and str(key).upper().endswith("FFFF99"))


def safe_cell_text(value, limit=220):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def find_header_row(ws, search_text, search_range=(1, 80)):
    start, end = search_range
    target = normalize_text(search_text)
    for row in range(start, min(end, ws.max_row) + 1):
        for col in range(1, min(LLM_MAX_STRUCTURE_COLS, ws.max_column) + 1):
            if target and target in normalize_text(ws.cell(row=row, column=col).value):
                return row
    return None


def find_header_col(ws, row, keywords):
    if not row:
        return None
    for col in range(1, min(LLM_MAX_STRUCTURE_COLS, ws.max_column) + 1):
        text = normalize_text(ws.cell(row=row, column=col).value)
        if any(normalize_text(keyword) in text for keyword in keywords):
            return col
    return None


def call_openai_json(config, system_prompt, payload, schema_name, schema):
    """Call an OpenAI-compatible API and parse a JSON object response."""
    if not config.enabled:
        return {"ok": False, "error": "LLM未启用"}
    if not config.api_key:
        return {"ok": False, "error": "未设置OPENAI_API_KEY，已跳过LLM调用"}

    response_result = call_responses_json(config, system_prompt, payload, schema_name, schema)
    if response_result.get("ok"):
        return response_result

    if response_result.get("status_code") == 404:
        chat_result = call_chat_completions_json(config, system_prompt, payload, schema_name, schema)
        if chat_result.get("ok"):
            return chat_result
        return {
            "ok": False,
            "error": (
                "Responses接口不存在，已尝试Chat Completions兼容接口但仍失败；"
                f"Responses错误: {response_result.get('error')}; "
                f"Chat错误: {chat_result.get('error')}"
            ),
        }

    return response_result


def call_responses_json(config, system_prompt, payload, schema_name, schema):
    """Call OpenAI Responses API."""
    body = {
        "model": config.model,
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": schema_name,
                "schema": schema,
                "strict": True,
            }
        },
    }
    request = urllib.request.Request(
        f"{config.base_url}/responses",
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config.api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=config.timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        return {
            "ok": False,
            "status_code": exc.code,
            "error": format_openai_error(exc.code, detail),
        }
    except Exception as exc:
        return {"ok": False, "error": f"LLM请求失败: {exc}"}

    text = data.get("output_text")
    if not text:
        parts = []
        for item in data.get("output", []):
            for content in item.get("content", []):
                if content.get("type") in ("output_text", "text"):
                    parts.append(content.get("text", ""))
        text = "\n".join(part for part in parts if part)

    try:
        return {"ok": True, "data": json.loads(text)}
    except Exception as exc:
        return {"ok": False, "error": f"LLM返回JSON解析失败: {exc}; raw={str(text)[:400]}"}


def call_chat_completions_json(config, system_prompt, payload, schema_name, schema):
    """Fallback for OpenAI-compatible gateways that only expose Chat Completions."""
    body = {
        "model": config.model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": (
                    "请严格返回一个JSON对象，不要输出Markdown或解释文字。"
                    f"\nJSON schema name: {schema_name}"
                    f"\nPayload:\n{json.dumps(payload, ensure_ascii=False)}"
                ),
            },
        ],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    request = urllib.request.Request(
        f"{config.base_url}/chat/completions",
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config.api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=config.timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        if exc.code in (400, 404):
            legacy_result = call_chat_completions_json_legacy(config, system_prompt, payload)
            if legacy_result.get("ok"):
                return legacy_result
        return {
            "ok": False,
            "status_code": exc.code,
            "error": format_openai_error(exc.code, detail),
        }
    except Exception as exc:
        return {"ok": False, "error": f"Chat Completions请求失败: {exc}"}

    try:
        text = data["choices"][0]["message"]["content"]
        return {"ok": True, "data": parse_json_text(text)}
    except Exception as exc:
        return {"ok": False, "error": f"Chat Completions返回解析失败: {exc}"}


def call_chat_completions_json_legacy(config, system_prompt, payload):
    """Last-resort fallback for gateways that reject response_format."""
    body = {
        "model": config.model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": (
                    "请严格返回一个JSON对象，不要输出Markdown或解释文字。"
                    f"\nPayload:\n{json.dumps(payload, ensure_ascii=False)}"
                ),
            },
        ],
        "temperature": 0,
    }
    request = urllib.request.Request(
        f"{config.base_url}/chat/completions",
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config.api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=config.timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
        text = data["choices"][0]["message"]["content"]
        return {"ok": True, "data": parse_json_text(text)}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        return {
            "ok": False,
            "status_code": exc.code,
            "error": format_openai_error(exc.code, detail),
        }
    except Exception as exc:
        return {"ok": False, "error": f"Chat Completions兼容请求失败: {exc}"}


def parse_json_text(text):
    """Parse a JSON object, tolerating fenced code blocks."""
    raw = str(text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        if match:
            return json.loads(match.group(0))
        raise


def format_openai_error(status_code, detail):
    """Return a concise user-facing message for OpenAI-compatible API errors."""
    message = str(detail or "").strip()
    code = ""
    try:
        payload = json.loads(detail)
        error = payload.get("error", {})
        message = error.get("message") or message
        code = error.get("code") or error.get("type") or ""
    except Exception:
        pass

    if status_code == 401:
        prefix = "认证失败：API Key无效或未被服务端接受"
    elif status_code == 403:
        prefix = "权限不足：当前API Key无权访问该模型或服务"
    elif status_code == 404:
        prefix = "接口或模型不存在：请检查Base URL是否应填写到/v1层级，以及模型名称是否正确"
    elif status_code == 429:
        prefix = "请求受限：额度不足、限流或配额耗尽"
    else:
        prefix = f"LLM请求失败：HTTP {status_code}"

    suffix = f" ({code})" if code else ""
    if message:
        return f"{prefix}{suffix}；{message[:300]}"
    return f"{prefix}{suffix}"


def test_llm_connection(llm_options=None):
    """Send a tiny structured request to validate API key, base URL, and model."""
    config = get_llm_config(True, llm_options)
    if not config.api_key:
        return {
            "ok": False,
            "error": "未填写API Key，且未设置OPENAI_API_KEY环境变量",
            "model": config.model,
            "base_url": config.base_url,
        }

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "status": {"type": "string", "enum": ["ok"]},
            "message": {"type": "string"},
        },
        "required": ["status", "message"],
    }
    response = call_openai_json(
        config,
        "你是连接测试助手。只返回JSON，不要输出额外文本。",
        {"task": "connection_test", "instruction": "Return ok."},
        "audit_llm_connection_test",
        schema,
    )
    if response.get("ok"):
        return {
            "ok": True,
            "message": "LLM连接测试成功",
            "model": config.model,
            "base_url": config.base_url,
        }
    return {
        "ok": False,
        "error": response.get("error", "LLM连接测试失败"),
        "model": config.model,
        "base_url": config.base_url,
    }


def summarize_sheet(ws):
    rows = []
    for row in range(1, min(LLM_MAX_STRUCTURE_ROWS, ws.max_row) + 1):
        cells = []
        for col in range(1, min(LLM_MAX_STRUCTURE_COLS, ws.max_column) + 1):
            value = ws.cell(row=row, column=col).value
            if value in (None, ""):
                continue
            if isinstance(value, (int, float)):
                continue
            cells.append({
                "cell": f"{get_column_letter(col)}{row}",
                "text": safe_cell_text(value, 120),
            })
        if cells:
            rows.append({"row": row, "cells": cells[:12]})
    return {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "sample_rows": rows[:40],
    }


def build_structure_summary(path, subject_config):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        expected_sheets = set()
        lead_name = subject_config.get("lead_sheet", {}).get("sheet_name")
        k01_name = subject_config.get("k01", {}).get("sheet_name")
        if lead_name:
            expected_sheets.add(lead_name)
        if k01_name:
            expected_sheets.add(k01_name)

        selected = []
        for sheet_name in wb.sheetnames:
            norm = normalize_text(sheet_name)
            if (
                sheet_name in expected_sheets
                or "Lead" in sheet_name
                or "lead" in sheet_name
                or "Agree" in sheet_name
                or "BKD" in sheet_name
                or len(selected) < 4
            ):
                selected.append(summarize_sheet(wb[sheet_name]))
            if len(selected) >= 8:
                break
        return {
            "file": os.path.basename(path),
            "sheet_names": wb.sheetnames,
            "selected_sheets": selected,
        }
    finally:
        wb.close()


def deterministic_precheck(template_path, prior_path, subject_code, subject_config):
    """Run non-LLM structure checks before any model call."""
    issues = []
    suggestions = []
    lead_config = subject_config.get("lead_sheet", {})
    k01_config = subject_config.get("k01", {})
    header_text = lead_config.get("header_search_text", "期末审定数")
    lead_name = lead_config.get("sheet_name")

    for label, path in (("模板", template_path), ("上年底稿", prior_path)):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        try:
            if lead_name not in wb.sheetnames:
                lead_candidates = [s for s in wb.sheetnames if "Lead" in s or "lead" in s]
                issues.append(f"{label}缺少配置中的Lead Sheet: {lead_name}")
                if lead_candidates:
                    suggestions.append(f"{label}疑似Lead Sheet: {', '.join(lead_candidates[:5])}")
            else:
                ws = wb[lead_name]
                header_row = find_header_row(ws, header_text)
                if not header_row:
                    issues.append(f"{label}的{lead_name}未找到表头关键词: {header_text}")
                else:
                    closing_col = find_header_col(ws, header_row, ["期末审定数", "本期期末审定数"])
                    opening_col = find_header_col(ws, header_row, ["期初审定数", "上期末审定数", "上年审定数"])
                    if not closing_col:
                        issues.append(f"{label}的{lead_name}第{header_row}行未定位到期末列")
                    if not opening_col and label == "模板":
                        issues.append(f"{label}的{lead_name}第{header_row}行未定位到期初列")

            if k01_config.get("has_k01"):
                k01_name = k01_config.get("sheet_name")
                if k01_name and k01_name not in wb.sheetnames:
                    issues.append(f"{label}缺少配置中的K.01 Sheet: {k01_name}")
        finally:
            wb.close()

    return {
        "subject_code": subject_code,
        "status": "needs_review" if issues else "ok",
        "issues": issues,
        "suggestions": suggestions,
    }


def run_llm_precheck(template_path, prior_path, subject_code, subject_config, enabled, llm_options=None):
    config = get_llm_config(enabled, llm_options)
    result = deterministic_precheck(template_path, prior_path, subject_code, subject_config)
    result["llm_used"] = False
    result["llm_error"] = ""
    result["llm_suggestions"] = []

    if not config.available:
        if config.enabled:
            result["llm_error"] = "未设置OPENAI_API_KEY，LLM预检仅执行本地规则检查"
        return result

    payload = {
        "task": "audit_roll_forward_precheck",
        "subject_code": subject_code,
        "subject_config": {
            "name": subject_config.get("name", ""),
            "lead_sheet": subject_config.get("lead_sheet", {}),
            "k01": subject_config.get("k01", {}),
            "sub_sheets": subject_config.get("sub_sheets", []),
        },
        "local_precheck": result,
        "template": build_structure_summary(template_path, subject_config),
        "prior": build_structure_summary(prior_path, subject_config),
    }
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "status": {"type": "string", "enum": ["ok", "needs_review", "blocked"]},
            "confidence": {"type": "number"},
            "mapping_suggestions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "sheet": {"type": "string"},
                        "role": {"type": "string"},
                        "header_row": {"type": ["integer", "null"]},
                        "closing_col": {"type": ["integer", "null"]},
                        "opening_col": {"type": ["integer", "null"]},
                        "reason": {"type": "string"},
                    },
                    "required": ["sheet", "role", "header_row", "closing_col", "opening_col", "reason"],
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
            "manual_checks": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["status", "confidence", "mapping_suggestions", "warnings", "manual_checks"],
    }
    response = call_openai_json(
        config,
        "你是审计底稿Roll Forward预检助手。只根据提供的工作簿结构摘要判断，不能编造单元格位置。输出结构化JSON。",
        payload,
        "audit_roll_forward_precheck",
        schema,
    )
    if response.get("ok"):
        data = response["data"]
        result["llm_used"] = True
        result["status"] = data.get("status", result["status"])
        result["confidence"] = data.get("confidence")
        result["llm_suggestions"] = data.get("mapping_suggestions", [])
        result["llm_warnings"] = data.get("warnings", [])
        result["manual_checks"] = data.get("manual_checks", [])
    else:
        result["llm_error"] = response.get("error", "LLM预检失败")
    return result


def collect_wording_candidates(wb):
    candidates = []
    skip_names = {LLM_REVIEW_SHEET_NAME, LLM_WORDING_SHEET_NAME, "Roll Forward Summary", "Post-roll Validation"}
    for ws in wb.worksheets:
        if ws.title in skip_names:
            continue
        for (row, col), cell in sorted(ws._cells.items()):
            if len(candidates) >= LLM_MAX_WORDING_CELLS:
                return candidates
            if isinstance(cell, MergedCell) or not is_yellow_fill(cell):
                continue
            value = cell.value
            if not isinstance(value, str) or not value.strip() or value.startswith("="):
                continue
            if len(value.strip()) < 4:
                continue
            candidates.append({
                "sheet": ws.title,
                "cell": f"{get_column_letter(col)}{row}",
                "row": row,
                "col": col,
                "value": value,
            })
    return candidates


def run_llm_wording_revision(wb, subject_code, company_name, bs_date, enabled, llm_options=None):
    config = get_llm_config(enabled, llm_options)
    candidates = collect_wording_candidates(wb)
    result = {
        "llm_used": False,
        "llm_error": "",
        "candidate_count": len(candidates),
        "changes": [],
        "warnings": [],
    }
    if not candidates:
        result["warnings"].append("未找到可供LLM修订的黄色wording单元格")
        return result
    if not config.available:
        result["llm_error"] = "未设置OPENAI_API_KEY，已跳过LLM wording修订" if config.enabled else "LLM未启用"
        return result

    date_obj = parse_date(bs_date)
    current_year = date_obj.year if date_obj else None
    prior_year = current_year - 1 if current_year else None
    payload = {
        "task": "revise_copied_audit_wording",
        "subject_code": subject_code,
        "company_name": company_name,
        "current_bs_date": bs_date,
        "current_year": current_year,
        "prior_year": prior_year,
        "rules": [
            "只修订文字说明，不修改金额、公式、表格结构或审计结论。",
            "只有当旧年份、旧日期、旧公司名或明显错期表述需要更新时才返回edit。",
            "涉及金额、比例、重大波动原因、管理层解释、合同/银行/供应商名称时只给warning，不自动改。",
            "new_text必须保留原语言风格，不能编造事实。",
        ],
        "cells": [
            {"sheet": item["sheet"], "cell": item["cell"], "text": item["value"]}
            for item in candidates
        ],
    }
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "edits": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "sheet": {"type": "string"},
                        "cell": {"type": "string"},
                        "old_text": {"type": "string"},
                        "new_text": {"type": "string"},
                        "reason": {"type": "string"},
                        "confidence": {"type": "number"},
                    },
                    "required": ["sheet", "cell", "old_text", "new_text", "reason", "confidence"],
                },
            },
            "warnings": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["edits", "warnings"],
    }
    response = call_openai_json(
        config,
        "你是审计底稿wording修订助手。必须保守，只能返回允许的单元格级文字修改建议。",
        payload,
        "audit_wording_revision",
        schema,
    )
    if not response.get("ok"):
        result["llm_error"] = response.get("error", "LLM wording修订失败")
        return result

    result["llm_used"] = True
    data = response["data"]
    result["warnings"].extend(data.get("warnings", []))
    candidate_map = {(item["sheet"], item["cell"]): item for item in candidates}
    cell_ref_re = re.compile(r"^([A-Z]{1,3})([1-9][0-9]*)$")

    for edit in data.get("edits", []):
        key = (edit.get("sheet"), edit.get("cell"))
        candidate = candidate_map.get(key)
        if not candidate:
            result["warnings"].append(f"拒绝非白名单wording单元格修改: {key}")
            continue
        if float(edit.get("confidence") or 0) < 0.75:
            result["warnings"].append(f"拒绝低置信度wording修改: {key}")
            continue
        if edit.get("old_text") != candidate["value"]:
            result["warnings"].append(f"拒绝old_text不匹配的wording修改: {key}")
            continue
        new_text = str(edit.get("new_text") or "")
        if not new_text.strip() or new_text.strip().startswith("="):
            result["warnings"].append(f"拒绝空文本或公式wording修改: {key}")
            continue
        if len(new_text) > max(800, len(candidate["value"]) * 4):
            result["warnings"].append(f"拒绝异常变长wording修改: {key}")
            continue
        if re.search(r"\d[\d,]*(\.\d+)?", candidate["value"]) and re.search(r"\d[\d,]*(\.\d+)?", new_text):
            old_numbers = re.findall(r"\d[\d,]*(?:\.\d+)?", candidate["value"])
            new_numbers = re.findall(r"\d[\d,]*(?:\.\d+)?", new_text)
            allowed_numbers = {str(current_year), str(prior_year)}
            if [n for n in old_numbers if n not in allowed_numbers] != [n for n in new_numbers if n not in allowed_numbers]:
                result["warnings"].append(f"拒绝疑似金额/比例变化的wording修改: {key}")
                continue

        match = cell_ref_re.match(candidate["cell"])
        if not match:
            continue
        ws = wb[candidate["sheet"]]
        cell = ws[candidate["cell"]]
        if isinstance(cell, MergedCell) or (isinstance(cell.value, str) and cell.value.startswith("=")):
            result["warnings"].append(f"拒绝公式或合并占位单元格修改: {key}")
            continue
        before = cell.value
        cell.value = new_text
        cell.fill = copy(PatternFill(fill_type="solid", fgColor="FFFF99"))
        result["changes"].append({
            "sheet": candidate["sheet"],
            "cell": candidate["cell"],
            "before": before,
            "after": new_text,
            "reason": edit.get("reason", ""),
            "confidence": edit.get("confidence"),
        })

    return result


def write_table(ws, start_row, headers, rows):
    row = start_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
    for item in rows:
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    return row + 1


def add_llm_wording_changes_sheet(wb, wording_result):
    if LLM_WORDING_SHEET_NAME in wb.sheetnames:
        del wb[LLM_WORDING_SHEET_NAME]
    ws = wb.create_sheet(LLM_WORDING_SHEET_NAME)
    ws.cell(row=1, column=1, value="LLM Wording Changes")
    ws.cell(row=2, column=1, value="LLM used")
    ws.cell(row=2, column=2, value="Yes" if wording_result.get("llm_used") else "No")
    ws.cell(row=3, column=1, value="Candidate cells")
    ws.cell(row=3, column=2, value=wording_result.get("candidate_count", 0))
    ws.cell(row=4, column=1, value="Applied changes")
    ws.cell(row=4, column=2, value=len(wording_result.get("changes", [])))
    ws.cell(row=5, column=1, value="Error")
    ws.cell(row=5, column=2, value=wording_result.get("llm_error", ""))

    row = write_table(
        ws,
        7,
        ["Sheet", "Cell", "Before", "After", "Reason", "Confidence"],
        [
            (
                item.get("sheet"),
                item.get("cell"),
                item.get("before"),
                item.get("after"),
                item.get("reason"),
                item.get("confidence"),
            )
            for item in wording_result.get("changes", [])
        ],
    )
    write_table(ws, row, ["Warnings"], [(item,) for item in wording_result.get("warnings", [])])
    for col, width in {"A": 26, "B": 14, "C": 44, "D": 44, "E": 40, "F": 14}.items():
        ws.column_dimensions[col].width = width
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A7"


def generate_review_text(subject_code, company_name, bs_date, precheck_result, wording_result, warnings_list, validation_items=None):
    lines = [
        f"科目: {subject_code}",
        f"公司: {company_name}",
        f"资产负债表日: {bs_date}",
        "",
        "LLM增强结论:",
    ]
    if precheck_result:
        lines.append(f"- 预检状态: {precheck_result.get('status', '未执行')}")
        for item in precheck_result.get("issues", [])[:8]:
            lines.append(f"- 本地预检问题: {item}")
        for item in precheck_result.get("llm_warnings", [])[:8]:
            lines.append(f"- LLM预检提示: {item}")
        for item in precheck_result.get("manual_checks", [])[:8]:
            lines.append(f"- 建议人工检查: {item}")
        if precheck_result.get("llm_error"):
            lines.append(f"- LLM预检未完成: {precheck_result.get('llm_error')}")
    else:
        lines.append("- 预检未启用")

    validation_items = validation_items or (precheck_result or {}).get("post_roll_validation", [])
    if validation_items:
        counts = {"PASS": 0, "WARN": 0, "FAIL": 0}
        for item in validation_items:
            counts[item.get("status", "")] = counts.get(item.get("status", ""), 0) + 1
        lines.append("")
        lines.append("程序规则复核:")
        lines.append(f"- PASS: {counts.get('PASS', 0)}; WARN: {counts.get('WARN', 0)}; FAIL: {counts.get('FAIL', 0)}")
        for item in [i for i in validation_items if i.get("status") in {"FAIL", "WARN"}][:10]:
            lines.append(f"- {item.get('status')}: {item.get('check')} {item.get('sheet', '')} {item.get('cell', '')} - {item.get('detail', '')}")

    if wording_result:
        lines.append(f"- Wording候选单元格: {wording_result.get('candidate_count', 0)}")
        lines.append(f"- LLM已修改wording: {len(wording_result.get('changes', []))}")
        if wording_result.get("llm_error"):
            lines.append(f"- LLM wording修订未完成: {wording_result.get('llm_error')}")
        for item in wording_result.get("warnings", [])[:8]:
            lines.append(f"- Wording提示: {item}")
    else:
        lines.append("- Wording修订未启用")

    if warnings_list:
        lines.append("")
        lines.append("程序warnings:")
        for item in list(dict.fromkeys(str(w) for w in warnings_list))[:12]:
            lines.append(f"- {item}")

    lines.append("")
    lines.append("注意: LLM仅用于预检、wording白名单修订和复核提示；金额、公式和核心roll forward规则仍由确定性代码处理。")
    return "\n".join(lines)


def add_llm_review_sheet(wb, subject_code, company_name, bs_date, precheck_result, wording_result, warnings_list, validation_items=None):
    if LLM_REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[LLM_REVIEW_SHEET_NAME]
    ws = wb.create_sheet(LLM_REVIEW_SHEET_NAME)
    validation_items = validation_items or (precheck_result or {}).get("post_roll_validation", [])
    review_text = generate_review_text(subject_code, company_name, bs_date, precheck_result, wording_result, warnings_list, validation_items)
    ws.cell(row=1, column=1, value="LLM Review")
    ws.cell(row=2, column=1, value=review_text)
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row = 4

    if precheck_result:
        row = write_table(
            ws,
            row,
            ["LLM mapping suggestion", "Role", "Header row", "Closing col", "Opening col", "Reason"],
            [
                (
                    item.get("sheet"),
                    item.get("role"),
                    item.get("header_row"),
                    item.get("closing_col"),
                    item.get("opening_col"),
                    item.get("reason"),
                )
                for item in precheck_result.get("llm_suggestions", [])
            ],
        )

    if validation_items:
        row = write_table(
            ws,
            row,
            ["Rule validation", "Status", "Sheet", "Cell/Row", "Expected", "Actual", "Detail"],
            [
                (
                    item.get("check", ""),
                    item.get("status", ""),
                    item.get("sheet", ""),
                    item.get("cell", ""),
                    item.get("expected", ""),
                    item.get("actual", ""),
                    item.get("detail", ""),
                )
                for item in validation_items
            ],
        )

    if wording_result:
        row = write_table(
            ws,
            row,
            ["Wording changed sheet", "Cell", "Reason", "Confidence"],
            [
                (
                    item.get("sheet"),
                    item.get("cell"),
                    item.get("reason"),
                    item.get("confidence"),
                )
                for item in wording_result.get("changes", [])
            ],
        )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 50
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
