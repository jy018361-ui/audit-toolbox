#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
审计底稿 Roll Forward 核心模块 (v4.0)
功能：支持多科目差异的批量Roll Forward处理
作者：AI Assistant
"""

import os
import re
import shutil
import json
import datetime
import warnings
import sys
import fnmatch
import posixpath
import zipfile
from collections import OrderedDict
from copy import copy, deepcopy
from pathlib import Path
from xml.etree import ElementTree

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.formula.translate import Translator
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.formula import ArrayFormula

# 忽略openpyxl的警告
warnings.filterwarnings('ignore', category=UserWarning)


def resource_path(relative_path):
    """Return an absolute path for bundled resources in source or PyInstaller."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_dir = Path(sys._MEIPASS)
    else:
        base_dir = Path(__file__).resolve().parent
    return str(base_dir / relative_path)

class RollForwardError(Exception):
    """Roll Forward 自定义异常"""
    pass


class SubjectConfig:
    """科目配置管理器"""

    def __init__(self, config_path=None):
        if config_path is None:
            config_path = resource_path("subjects_config.json")
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

    def get_subject(self, subject_code):
        """获取单个科目配置"""
        return self.config.get("subjects", {}).get(subject_code)

    def get_all_subjects(self):
        """获取所有科目配置"""
        return self.config.get("subjects", {})

    def get_subject_list(self):
        """获取科目列表 [(code, name), ...]"""
        subjects = []
        for code, info in self.config.get("subjects", {}).items():
            subjects.append((code, info.get("name", "")))
        return subjects


def find_prior_file(prior_dir, subject_code, prior_year, config):
    """Recursively find the best prior-year workbook for one subject."""
    if not prior_dir or not os.path.isdir(prior_dir):
        return None

    patterns = config.get("prior_file_patterns") or [
        config.get("prior_file_pattern", f"{subject_code}*")
    ]
    concrete_patterns = [
        pattern.replace("{prior_year}", str(prior_year))
        for pattern in patterns
        if pattern
    ]

    candidates = []
    for root_dir, _, filenames in os.walk(prior_dir):
        for filename in filenames:
            lower_name = filename.lower()
            if not lower_name.endswith(".xlsx") or filename.startswith(("~$", "~")):
                continue
            if any(marker in filename for marker in ("XYZ公司", "202YMMDD", "标准模板")):
                continue
            candidates.append(os.path.join(root_dir, filename))

    subject_name = str(config.get("name") or "").strip()

    def candidate_score(path):
        filename = os.path.basename(path)
        normalized = filename.upper().replace(" ", "")

        # Finance expense and VC/VD use the same U_exp prefix and must be
        # mutually exclusive before any generic alias scoring is applied.
        has_vcvd = "VC&VD" in normalized or ("销售费用" in filename and "管理费用" in filename)
        if subject_code == "UexpVCVD" and not has_vcvd:
            return None
        if subject_code == "Uexp" and has_vcvd:
            return None

        score = 0
        for pattern in concrete_patterns:
            if fnmatch.fnmatch(filename, pattern) or fnmatch.fnmatch(filename, f"*{pattern}"):
                score = max(score, 200)

        if str(prior_year) in filename:
            score += 40
        if subject_name and subject_name in filename:
            score += 100

        if subject_code == "UexpVCVD":
            score += 160
        elif subject_code == "Uexp":
            if "财务费用" in filename:
                score += 160
            elif "OTHER" in normalized:
                score += 100
        elif subject_code == "L1" and any(alias in normalized for alias in ("LAR", "LRA")):
            score += 80

        code_key = subject_code.upper().replace(" ", "")
        if len(code_key) > 1 and code_key in normalized:
            score += 60

        for pattern in patterns:
            keyword_text = pattern.replace("{prior_year}", "").replace(".xlsx", "")
            keywords = [part for part in keyword_text.split("*") if len(part.strip()) > 1]
            score += 15 * sum(1 for keyword in keywords if keyword.strip().upper() in normalized)

        return score if score >= 40 else None

    ranked = []
    for path in candidates:
        score = candidate_score(path)
        if score is not None:
            ranked.append((score, os.path.getmtime(path), path))
    if not ranked:
        return None
    ranked.sort(reverse=True)
    return ranked[0][2]


def find_header_row(ws, search_text, search_range=(1, 80)):
    """在工作表中查找包含特定文本的行"""
    start, end = search_range
    for row in range(start, min(end + 1, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and search_text in str(cell_value):
                return row
    return None


def parse_date_value(value):
    """Parse common user-entered date formats."""
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime(value.year, value.month, value.day)
    if value is None:
        return None

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def normalize_text(value):
    """Normalize labels for cross-template matching."""
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    return text


def is_visible_worksheet(workbook, sheet_name):
    """Return True only for a worksheet that users can see in Excel."""
    return (
        sheet_name in workbook.sheetnames
        and workbook[sheet_name].sheet_state == "visible"
    )


def find_visible_sheet_name(workbook, predicate):
    """Find the first visible worksheet whose name satisfies predicate."""
    return next(
        (
            name
            for name in workbook.sheetnames
            if workbook[name].sheet_state == "visible" and predicate(name)
        ),
        None,
    )


def fill_adjacent_header_value(ws, keywords, value, search_range=(1, 20)):
    """Fill the cell to the right of a matching header label."""
    if value is None or str(value).strip() == "":
        return False

    start, end = search_range
    for row in range(start, min(end + 1, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            text = str(cell_value)
            if any(keyword in text for keyword in keywords):
                return set_cell_value(ws, row, col + 1, value)

    return False


def normalize_lead_date_formats(ws):
    """Keep report-date cells and date formulas in a consistent display format."""
    for row in range(1, min(80, ws.max_row) + 1):
        for col in range(1, min(20, ws.max_column) + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if isinstance(value, (datetime.datetime, datetime.date)):
                cell.number_format = "yyyy/mm/dd"
            elif isinstance(value, str) and (
                "$C$3" in value
                or "DATE(YEAR(" in value
                or "汇总!D5" in value
                or "汇总!$D$5" in value
            ):
                cell.number_format = "yyyy/mm/dd"


def set_cell_value(ws, row, column, value):
    """Set a cell value unless the target is a merged placeholder."""
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def find_header_col(ws, row, keywords, max_col=None):
    """Find a column in one header row by keyword."""
    if not row:
        return None
    max_col = max_col or ws.max_column
    for col in range(1, min(max_col, ws.max_column) + 1):
        value = ws.cell(row=row, column=col).value
        if not value:
            continue
        text = normalize_text(value)
        if any(normalize_text(keyword) in text for keyword in keywords):
            return col
    return None


def find_header_col_near(ws, header_row, keywords, row_offsets=(-1, 0, 1), max_col=None):
    """Find a header column in rows around the main header row."""
    for offset in row_offsets:
        row = header_row + offset
        if row < 1:
            continue
        col = find_header_col(ws, row, keywords, max_col)
        if col:
            return col
    return None


def row_contains_any(ws, row, keywords, columns=None):
    """Return True when any target column in a row contains a keyword."""
    if columns is None:
        columns = range(1, min(20, ws.max_column) + 1)
    row_text = " ".join(str(ws.cell(row=row, column=col).value or "") for col in columns)
    normalized = normalize_text(row_text)
    return any(normalize_text(keyword) in normalized for keyword in keywords)


def is_table_end_marker_row(ws, row, max_col=3):
    """Return True for worksheet marker rows such as /T1."""
    for col in range(1, min(max_col, ws.max_column) + 1):
        value = ws.cell(row=row, column=col).value
        if isinstance(value, str) and normalize_text(value).upper().startswith("/T"):
            return True
    return False


def copy_cell_shape(source_cell, target_cell, translate_formula=False):
    """Copy style and optionally formula/value from one cell to another."""
    if isinstance(target_cell, MergedCell):
        return

    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)

    value = source_cell.value
    if translate_formula and isinstance(value, ArrayFormula):
        formula_text = value.text
        try:
            formula_text = Translator(
                formula_text,
                origin=source_cell.coordinate,
            ).translate_formula(target_cell.coordinate)
        except Exception:
            pass
        value = ArrayFormula(ref=target_cell.coordinate, text=formula_text)
    elif isinstance(value, ArrayFormula):
        value = ArrayFormula(ref=target_cell.coordinate, text=value.text)
    elif translate_formula and isinstance(value, str) and value.startswith("="):
        try:
            value = Translator(value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
        except Exception:
            pass
    target_cell.value = value


def copy_row_shape(ws, source_row, target_row, translate_formula=True):
    """Copy a row's visible structure into another row."""
    for col in range(1, ws.max_column + 1):
        copy_cell_shape(
            ws.cell(row=source_row, column=col),
            ws.cell(row=target_row, column=col),
            translate_formula=translate_formula,
        )
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def transform_cell_range_for_insert(cell_range, idx, amount, axis="row", expand_adjacent=False):
    """Return a range adjusted for inserted rows or columns."""
    new_range = CellRange(str(cell_range))

    if axis == "row":
        if new_range.min_row >= idx:
            new_range.shift(row_shift=amount)
        elif new_range.min_row < idx <= new_range.max_row:
            new_range.expand(down=amount)
        elif expand_adjacent and new_range.max_row == idx - 1:
            new_range.expand(down=amount)
    else:
        if new_range.min_col >= idx:
            new_range.shift(col_shift=amount)
        elif new_range.min_col < idx <= new_range.max_col:
            new_range.expand(right=amount)
        elif expand_adjacent and new_range.max_col == idx - 1:
            new_range.expand(right=amount)

    return new_range


def transform_multi_cell_range_for_insert(sqref, idx, amount, axis="row", expand_adjacent=False):
    """Adjust all ranges in a MultiCellRange for an insertion."""
    if sqref is None:
        return sqref

    ranges = getattr(sqref, "ranges", None)
    if ranges is None:
        ranges = [CellRange(str(sqref))]

    return MultiCellRange([
        transform_cell_range_for_insert(cell_range, idx, amount, axis, expand_adjacent)
        for cell_range in ranges
    ])


def update_conditional_formatting_for_insert(ws, idx, amount, axis="row"):
    """Move and extend conditional formatting ranges around inserted rows/columns."""
    cf_rules = getattr(ws.conditional_formatting, "_cf_rules", None)
    if not cf_rules:
        return

    updated_rules = OrderedDict()
    for conditional_formatting, rules in cf_rules.items():
        new_sqref = transform_multi_cell_range_for_insert(
            conditional_formatting.sqref,
            idx,
            amount,
            axis=axis,
            expand_adjacent=True,
        )
        new_cf = ConditionalFormatting(
            sqref=new_sqref,
            pivot=conditional_formatting.pivot,
            cfRule=conditional_formatting.cfRule,
        )
        updated_rules[new_cf] = rules

    ws.conditional_formatting._cf_rules = updated_rules


def update_data_validations_for_insert(ws, idx, amount, axis="row"):
    """Move and extend data validation ranges around inserted rows/columns."""
    validations = getattr(ws.data_validations, "dataValidation", [])
    for validation in validations:
        validation.sqref = transform_multi_cell_range_for_insert(
            validation.sqref,
            idx,
            amount,
            axis=axis,
            expand_adjacent=True,
        )


def insert_rows_preserving_sheet_metadata(ws, idx, amount):
    """Insert rows without leaving sheet-level metadata at stale coordinates."""
    if amount <= 0:
        return

    affected_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= idx:
            affected_ranges.append((merged_range, "shift"))
        elif merged_range.min_row < idx <= merged_range.max_row:
            affected_ranges.append((merged_range, "expand"))

    ws.insert_rows(idx, amount)

    for merged_range, action in affected_ranges:
        if action == "shift":
            merged_range.shift(row_shift=amount)
        elif action == "expand":
            merged_range.expand(down=amount)

    update_conditional_formatting_for_insert(ws, idx, amount, axis="row")
    update_data_validations_for_insert(ws, idx, amount, axis="row")


def insert_cols_preserving_sheet_metadata(ws, idx, amount):
    """Insert columns without leaving sheet-level metadata at stale coordinates."""
    if amount <= 0:
        return

    affected_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col >= idx:
            affected_ranges.append((merged_range, "shift"))
        elif merged_range.min_col < idx <= merged_range.max_col:
            affected_ranges.append((merged_range, "expand"))

    ws.insert_cols(idx, amount)

    for merged_range, action in affected_ranges:
        if action == "shift":
            merged_range.shift(col_shift=amount)
        elif action == "expand":
            merged_range.expand(right=amount)

    update_conditional_formatting_for_insert(ws, idx, amount, axis="col")
    update_data_validations_for_insert(ws, idx, amount, axis="col")


def clone_worksheet_contents(ws_source, ws_target):
    """Replace a worksheet's contents with another worksheet's contents."""
    for merged_range in list(ws_target.merged_cells.ranges):
        ws_target.unmerge_cells(str(merged_range))

    ws_target._cells = {}
    ws_target.merged_cells = deepcopy(ws_source.merged_cells)
    ws_target.sheet_format = copy(ws_source.sheet_format)
    ws_target.sheet_properties = copy(ws_source.sheet_properties)
    ws_target.page_margins = copy(ws_source.page_margins)
    ws_target.page_setup = copy(ws_source.page_setup)
    ws_target.print_options = copy(ws_source.print_options)
    ws_target.freeze_panes = ws_source.freeze_panes
    ws_target.auto_filter.ref = ws_source.auto_filter.ref

    ws_target.row_dimensions = deepcopy(ws_source.row_dimensions)
    ws_target.column_dimensions = deepcopy(ws_source.column_dimensions)
    ws_target.conditional_formatting = deepcopy(ws_source.conditional_formatting)
    ws_target.data_validations = deepcopy(ws_source.data_validations)
    ws_target._tables = deepcopy(ws_source._tables)

    for row in ws_source.iter_rows():
        for source_cell in row:
            target_cell = ws_target.cell(row=source_cell.row, column=source_cell.column)
            if source_cell.value is not None:
                target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def find_row_containing(ws, text, search_range=(1, 120)):
    """Find the first row containing a text fragment."""
    start, end = search_range
    for row in range(start, min(end, ws.max_row) + 1):
        for col in range(1, min(30, ws.max_column) + 1):
            value = ws.cell(row=row, column=col).value
            if value and text in str(value):
                return row
    return None


def find_total_row_after(ws, start_row):
    """Find the next total row after a starting row."""
    if not start_row:
        return None
    for row in range(start_row + 1, ws.max_row + 1):
        row_text = " ".join(
            str(ws.cell(row=row, column=col).value or "")
            for col in range(1, min(20, ws.max_column) + 1)
        )
        if "合计" in row_text:
            return row
    return None


def find_group_child_cols(ws, group_header_row, child_header_row, group_keywords, stop_keywords=None):
    """Find child columns under a grouped header area."""
    stop_keywords = stop_keywords or []
    start_col = None
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=group_header_row, column=col).value
        if value and any(keyword in str(value) for keyword in group_keywords):
            start_col = col
            break
    if not start_col:
        return []

    cols = []
    for col in range(start_col, ws.max_column + 1):
        header = ws.cell(row=child_header_row, column=col).value
        if col > start_col and header and any(keyword in str(header) for keyword in stop_keywords):
            break
        if header not in (None, ""):
            cols.append(col)
    return cols


def clear_constant_cells(ws, row_start, row_end, columns):
    """Clear constants in target columns while keeping formulas."""
    for row in range(row_start, row_end + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None and not (isinstance(value, str) and value.startswith("=")):
                set_cell_value(ws, row, col, None)


def clear_borders(ws, row_start, row_end, col_start, col_end):
    """Remove dense copied borders from a body area."""
    no_border = Border()
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.border = no_border


def highlight_rows(ws, rows, col_start=1, col_end=None):
    """Highlight rows that need manual refresh."""
    fill = PatternFill(fill_type="solid", fgColor="FFFF99")
    col_end = col_end or ws.max_column
    for row in rows:
        if row < 1 or row > ws.max_row:
            continue
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.fill = copy(fill)


WORDING_START_KEYWORDS = (
    "预期",
    "波动说明",
    "波动分析",
    "Notes",
    "Notes:",
    "Notes：",
    "调整汇总",
    "调整分录",
    "调整事项",
    "ARP",
    "变动不在范围",
    "在下文中描述",
    "对于单项变动金额",
)

WORDING_END_ROW_KEYWORDS = (
    "账套名称",
    "账套编码",
    "总账科目编码",
    "科目名称",
    "期末账面数",
    "期末审定数",
    "本期期末审定数",
    "公司名称",
    "银行/存款机构",
    "账号",
    "项目编码",
    "债务描述",
    "资产类别",
)

WORDING_HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFFF99")
SUMMARY_SHEET_NAME = "Roll Forward Summary"
SUMMARY_DETAIL_LIMIT = 1000


class RollForwardWarnings(list):
    """Warnings list with optional run metadata for the GUI."""

    def __init__(self):
        super().__init__()
        self.metadata = {}


def cell_fill_key(cell):
    """Return a compact fill key used for lightweight diff reporting."""
    fill = cell.fill
    if not fill or not fill.fill_type:
        return None

    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return str(color.rgb or color.indexed or color.theme or "")


def is_yellow_fill(fill_key):
    """Return True for the yellow review marker used by wording roll-forward."""
    if not fill_key:
        return False
    return str(fill_key).upper().endswith("FFFF99")


def workbook_snapshot(wb):
    """Take a value/fill snapshot without changing workbook contents."""
    snapshot = {}
    for ws in wb.worksheets:
        if ws.title == SUMMARY_SHEET_NAME:
            continue
        sheet_cells = {}
        for key, cell in ws._cells.items():
            sheet_cells[key] = (cell.value, cell_fill_key(cell))
        snapshot[ws.title] = sheet_cells
    return snapshot


def short_cell_value(value, limit=160):
    """Format a cell value for the summary sheet."""
    if value is None:
        return ""
    text = str(value)
    if len(text) > limit:
        return text[:limit - 3] + "..."
    return text


def build_workbook_diff(before_snapshot, wb):
    """Compare the initial template snapshot with the processed workbook."""
    after_snapshot = workbook_snapshot(wb)
    updated_cells = []
    yellow_cells = []
    updated_sheets = set()

    sheet_names = sorted(set(before_snapshot) | set(after_snapshot))
    for sheet_name in sheet_names:
        before_cells = before_snapshot.get(sheet_name, {})
        after_cells = after_snapshot.get(sheet_name, {})
        cell_keys = set(before_cells) | set(after_cells)

        for row, col in sorted(cell_keys):
            before_value, before_fill = before_cells.get((row, col), (None, None))
            after_value, after_fill = after_cells.get((row, col), (None, None))
            address = f"{get_column_letter(col)}{row}"

            if before_value != after_value:
                updated_sheets.add(sheet_name)
                updated_cells.append({
                    "sheet": sheet_name,
                    "cell": address,
                    "before": before_value,
                    "after": after_value,
                })

            if is_yellow_fill(after_fill) and before_fill != after_fill:
                yellow_cells.append({
                    "sheet": sheet_name,
                    "cell": address,
                    "value": after_value,
                })

    return {
        "updated_sheets": sorted(updated_sheets),
        "updated_cells": updated_cells,
        "yellow_cells": yellow_cells,
    }


def write_rows(ws, start_row, headers, rows, limit=SUMMARY_DETAIL_LIMIT):
    """Write a bounded table and return the next available row."""
    row = start_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

    shown_rows = rows[:limit]
    for item in shown_rows:
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    if len(rows) > limit:
        ws.cell(row=row, column=1, value=f"... truncated, {len(rows) - limit} more rows")
        row += 1

    return row + 1


def add_roll_forward_summary_sheet(wb, subject_code, subject_name, company_name, bs_date,
                                   prior_path, output_path, warnings_list, options,
                                   wording_count, wording_sheets, before_snapshot):
    """Append a lightweight run summary sheet to the generated workbook."""
    diff = build_workbook_diff(before_snapshot, wb)

    if SUMMARY_SHEET_NAME in wb.sheetnames:
        del wb[SUMMARY_SHEET_NAME]
    ws = wb.create_sheet(SUMMARY_SHEET_NAME)

    warning_items = list(dict.fromkeys(warnings_list))
    unmatched_items = list(warning_items)
    if options.get("roll_wording") and wording_count == 0:
        unmatched_items.append("未匹配到可复制的 wording 区域，或上年底稿无可复制 wording 内容")

    metadata = {
        "updated_sheet_count": len(diff["updated_sheets"]),
        "updated_cell_count": len(diff["updated_cells"]),
        "yellow_cell_count": len(diff["yellow_cells"]),
        "wording_copied_count": wording_count,
        "wording_sheets": wording_sheets,
        "warnings_count": len(warning_items),
        "unmatched_count": len(unmatched_items),
        "summary_sheet": SUMMARY_SHEET_NAME,
    }
    if hasattr(warnings_list, "metadata"):
        warnings_list.metadata.update(metadata)

    ws.cell(row=1, column=1, value="Roll Forward Summary")
    ws.cell(row=2, column=1, value="Generated at")
    ws.cell(row=2, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=3, column=1, value="Subject")
    ws.cell(row=3, column=2, value=f"{subject_code} {subject_name}".strip())
    ws.cell(row=4, column=1, value="Company")
    ws.cell(row=4, column=2, value=company_name)
    ws.cell(row=5, column=1, value="Balance sheet date")
    ws.cell(row=5, column=2, value=bs_date)
    ws.cell(row=6, column=1, value="Prior file")
    ws.cell(row=6, column=2, value=prior_path)
    ws.cell(row=7, column=1, value="Output file")
    ws.cell(row=7, column=2, value=output_path)

    option_text = []
    option_text.append(f"roll wording: {'Yes' if options.get('roll_wording') else 'No'}")
    option_text.append(f"generate summary: {'Yes' if options.get('generate_summary') else 'No'}")
    ws.cell(row=8, column=1, value="Options")
    ws.cell(row=8, column=2, value="; ".join(option_text))

    rows = [
        ("Updated sheets", metadata["updated_sheet_count"]),
        ("Updated cells", metadata["updated_cell_count"]),
        ("Yellow cells", metadata["yellow_cell_count"]),
        ("Wording copied cells", wording_count),
        ("Warnings / unmatched", metadata["unmatched_count"]),
    ]
    write_rows(ws, 10, ["Metric", "Count"], rows)

    row = 18
    row = write_rows(
        ws,
        row,
        ["Updated sheet"],
        [(sheet_name,) for sheet_name in diff["updated_sheets"]],
    )
    row = write_rows(
        ws,
        row,
        ["Sheet", "Cell", "Before", "After"],
        [
            (
                item["sheet"],
                item["cell"],
                short_cell_value(item["before"]),
                short_cell_value(item["after"]),
            )
            for item in diff["updated_cells"]
        ],
    )
    row = write_rows(
        ws,
        row,
        ["Sheet", "Yellow cell", "Value"],
        [
            (item["sheet"], item["cell"], short_cell_value(item["value"]))
            for item in diff["yellow_cells"]
        ],
    )
    write_rows(
        ws,
        row,
        ["Warnings / unmatched areas"],
        [(item,) for item in unmatched_items],
    )

    for col, width in {"A": 28, "B": 32, "C": 44, "D": 44}.items():
        ws.column_dimensions[col].width = width
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A10"

    return metadata


def get_row_text(ws, row, col_start=1, col_end=None):
    """Return a compact text representation of a worksheet row."""
    col_end = col_end or min(ws.max_column, 40)
    values = []
    for col in range(col_start, min(col_end, ws.max_column) + 1):
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            values.append(str(value))
    return " ".join(values)


def text_has_any_keyword(text, keywords):
    """Case-insensitive keyword match with whitespace normalized."""
    text_key = normalize_text(text).lower()
    return any(normalize_text(keyword).lower() in text_key for keyword in keywords)


def row_has_any_keyword(ws, row, keywords, col_start=1, col_end=None):
    """Return True when a row contains any keyword."""
    return text_has_any_keyword(get_row_text(ws, row, col_start, col_end), keywords)


def find_wording_start_rows(ws):
    """Find likely wording/commentary section anchors."""
    starts = []
    last_start = 0
    for row in range(1, ws.max_row + 1):
        if row <= last_start + 1:
            continue
        if row_has_any_keyword(ws, row, WORDING_END_ROW_KEYWORDS):
            continue
        if row_has_any_keyword(ws, row, WORDING_START_KEYWORDS):
            starts.append(row)
            last_start = row
    return starts


def find_wording_section_end(ws, start_row, next_start_row=None, max_rows=80):
    """Find the end of a wording section using the next anchor and blank runs."""
    hard_end = min(ws.max_row, start_row + max_rows - 1)
    if next_start_row:
        hard_end = min(hard_end, next_start_row - 1)

    last_content_row = start_row
    blank_run = 0
    for row in range(start_row, hard_end + 1):
        if row > start_row and is_table_end_marker_row(ws, row):
            return max(start_row, row - 1)
        if row > start_row and row_has_any_keyword(ws, row, WORDING_END_ROW_KEYWORDS):
            return max(start_row, row - 1)

        has_content = row_has_content(ws, row, 1, min(ws.max_column, 40))
        if has_content:
            last_content_row = row
            blank_run = 0
        else:
            blank_run += 1
            if row > start_row and blank_run >= 3:
                return last_content_row

    return last_content_row


def get_matching_wording_keywords(ws, row):
    """Return the wording keywords present in an anchor row."""
    text = get_row_text(ws, row)
    text_key = normalize_text(text).lower()
    return [
        keyword
        for keyword in WORDING_START_KEYWORDS
        if normalize_text(keyword).lower() in text_key
    ]


def find_target_wording_start(ws_target, source_anchor_row, source_keywords):
    """Find the corresponding wording anchor in the target worksheet."""
    if not source_keywords:
        source_keywords = WORDING_START_KEYWORDS

    target_starts = find_wording_start_rows(ws_target)
    windows = [
        (max(1, source_anchor_row - 25), min(ws_target.max_row, source_anchor_row + 25)),
        (1, ws_target.max_row),
    ]
    for start, end in windows:
        for row in target_starts:
            if row < start or row > end:
                continue
            if row_has_any_keyword(ws_target, row, source_keywords):
                return row

    if source_anchor_row <= ws_target.max_row:
        return source_anchor_row
    return ws_target.max_row


def is_adjustment_wording_section(ws, start_row):
    """Return True for adjustment summary style sections where numbers matter too."""
    return row_has_any_keyword(
        ws,
        start_row,
        ("调整汇总", "调整分录", "调整事项"),
    )


def source_wording_value(formula_cell, value_cell, table_like=False):
    """Return the value to roll for a wording cell, or None when it should be skipped."""
    formula_value = formula_cell.value
    value = value_cell.value

    if formula_value in (None, "") and value in (None, ""):
        return None

    if isinstance(formula_value, str) and formula_value.startswith("="):
        return value if table_like and value not in (None, "") else None

    if isinstance(formula_value, str):
        return formula_value

    if table_like and value not in (None, ""):
        return value

    return None


def highlight_wording_cell(ws, row, col):
    """Highlight one copied wording cell."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.fill = copy(WORDING_HIGHLIGHT_FILL)
        try:
            cell.font = cell.font.copy(color="000000")
        except Exception:
            pass


def ensure_target_wording_capacity(ws_target, target_start, target_len, required_len):
    """Insert rows when the target wording area is shorter than the source section."""
    extra_rows = required_len - target_len
    if extra_rows <= 0:
        return

    insert_at = target_start + max(target_len, 1)
    source_shape_row = max(target_start, insert_at - 1)
    insert_rows_preserving_sheet_metadata(ws_target, insert_at, extra_rows)
    for offset in range(extra_rows):
        copy_row_shape(ws_target, source_shape_row, insert_at + offset, translate_formula=True)


def copy_wording_section(ws_prior_formula, ws_prior_values, ws_target, source_start, source_end, target_start):
    """Copy one detected wording section into the target sheet and mark copied cells."""
    table_like = is_adjustment_wording_section(ws_prior_formula, source_start)
    source_len = source_end - source_start + 1
    target_starts = find_wording_start_rows(ws_target)
    next_target_start = next((row for row in target_starts if row > target_start), None)
    target_end = find_wording_section_end(ws_target, target_start, next_target_start)
    target_len = max(1, target_end - target_start + 1)
    if table_like:
        ensure_target_wording_capacity(ws_target, target_start, target_len, source_len)

    copied = 0
    max_col = min(max(ws_prior_formula.max_column, ws_target.max_column), 40)
    for offset, source_row in enumerate(range(source_start, source_end + 1)):
        target_row = target_start + offset
        for col in range(1, max_col + 1):
            value = source_wording_value(
                ws_prior_formula.cell(row=source_row, column=col),
                ws_prior_values.cell(row=source_row, column=col),
                table_like=table_like,
            )
            if value in (None, ""):
                continue
            if set_cell_value(ws_target, target_row, col, value):
                highlight_wording_cell(ws_target, target_row, col)
                copied += 1

    return copied


def process_wording_sections(wb_prior_formula, wb_prior_values, wb_new, subject_code, subject_config, warnings_list=None):
    """Roll prior-year wording sections into the new workbook as an optional post-step."""
    copied = 0
    touched_sheets = set()

    for sheet_name in wb_prior_formula.sheetnames:
        if wb_prior_formula[sheet_name].sheet_state != "visible":
            continue
        if normalize_text(sheet_name).startswith("汇总"):
            continue
        if subject_code == "C" and normalize_text(sheet_name).upper().startswith("C.03CUTOFF"):
            continue
        if subject_code == "J1" and (
            normalize_text(sheet_name).startswith("J.00")
            or normalize_text(sheet_name).startswith("J.01")
            or normalize_text(sheet_name).startswith("J.03")
        ):
            continue
        if subject_code == "K1" and normalize_text(sheet_name).lower() == "falist":
            continue
        if subject_code == "L2" and (
            normalize_text(sheet_name).startswith("L2.00")
            or normalize_text(sheet_name).startswith("L2.01.1")
            or normalize_text(sheet_name).startswith("L2.02")
        ):
            continue
        if subject_code == "N" and (
            normalize_text(sheet_name).startswith("N.02")
            or normalize_text(sheet_name).startswith("N.03")
        ):
            continue
        if subject_code == "Q1" and (
            normalize_text(sheet_name).upper().startswith("Q1.02B")
            or normalize_text(sheet_name).upper().startswith("Q1.02C")
            or normalize_text(sheet_name).startswith("Q1.05")
        ):
            continue
        if subject_code == "UexpVCVD" and (
            normalize_text(sheet_name).startswith("VC.00")
            or normalize_text(sheet_name).startswith("VD.00")
        ):
            continue
        if subject_code == "Uexp":
            continue
        if sheet_name not in wb_new.sheetnames or sheet_name not in wb_prior_values.sheetnames:
            continue

        ws_prior_formula = wb_prior_formula[sheet_name]
        ws_prior_values = wb_prior_values[sheet_name]
        ws_target = wb_new[sheet_name]
        source_starts = find_wording_start_rows(ws_prior_formula)
        if subject_code == "L1" and normalize_text(sheet_name).startswith("L1.00"):
            source_starts = [
                row
                for row in source_starts
                if not row_has_any_keyword(ws_prior_formula, row, ("预期",))
            ]
        if not source_starts:
            continue

        last_target_start = 0
        for idx, source_start in enumerate(source_starts):
            next_source_start = source_starts[idx + 1] if idx + 1 < len(source_starts) else None
            source_end = find_wording_section_end(ws_prior_formula, source_start, next_source_start)
            if source_end < source_start:
                continue

            source_keywords = get_matching_wording_keywords(ws_prior_formula, source_start)
            target_start = find_target_wording_start(ws_target, source_start, source_keywords)
            if target_start <= last_target_start:
                continue

            section_copied = copy_wording_section(
                ws_prior_formula,
                ws_prior_values,
                ws_target,
                source_start,
                source_end,
                target_start,
            )
            if section_copied:
                copied += section_copied
                touched_sheets.add(sheet_name)
                last_target_start = target_start

    if copied and warnings_list is not None:
        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

    return copied, sorted(touched_sheets)


def copy_wording_value_to_cell(ws_prior_formula, ws_prior_values, ws_target, source_row, source_col, target_row=None, target_col=None, table_like=False):
    """Copy one wording value and highlight only the copied target cell."""
    target_row = target_row or source_row
    target_col = target_col or source_col
    value = source_wording_value(
        ws_prior_formula.cell(row=source_row, column=source_col),
        ws_prior_values.cell(row=source_row, column=source_col),
        table_like=table_like,
    )
    if value in (None, ""):
        return 0
    if set_cell_value(ws_target, target_row, target_col, value):
        highlight_wording_cell(ws_target, target_row, target_col)
        return 1
    return 0


def clear_target_cells(ws_target, row_start, row_end, col_start, col_end):
    """Clear a small target area before placing a wording table."""
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            set_cell_value(ws_target, row, col, None)


def find_label_cell(ws, keywords, search_range=None, max_col=30):
    """Find a label cell by its visible text instead of a fixed coordinate."""
    start_row, end_row = search_range or (1, ws.max_row)
    for row in range(max(1, start_row), min(ws.max_row, end_row) + 1):
        for col in range(1, min(ws.max_column, max_col) + 1):
            value = ws.cell(row=row, column=col).value
            if value in (None, ""):
                continue
            if text_has_any_keyword(value, keywords):
                return row, col
    return None, None


def rollable_cell_value(formula_cell, value_cell):
    """Return a prior-year constant or cached formula result for a labeled field."""
    formula_value = formula_cell.value
    cached_value = value_cell.value
    if isinstance(formula_value, str) and formula_value.startswith("="):
        return cached_value
    if formula_value not in (None, ""):
        return formula_value
    return cached_value


def copy_adjacent_labeled_value(ws_prior_formula, ws_prior_values, ws_new, label_keywords):
    """Copy the nearest value beside a matching label and mark it for review."""
    source_row, source_label_col = find_label_cell(ws_prior_formula, label_keywords)
    target_row, target_label_col = find_label_cell(ws_new, label_keywords)
    if not source_row or not target_row:
        return 0

    source_col = None
    value = None
    for distance in range(1, min(ws_prior_formula.max_column, 12) + 1):
        for candidate_col in (source_label_col - distance, source_label_col + distance):
            if candidate_col < 1 or candidate_col > ws_prior_formula.max_column:
                continue
            candidate = rollable_cell_value(
                ws_prior_formula.cell(source_row, candidate_col),
                ws_prior_values.cell(source_row, candidate_col),
            )
            if candidate in (None, ""):
                continue
            source_col = candidate_col
            value = candidate
            break
        if source_col:
            break
    if source_col is None:
        return 0

    relative_col = source_col - source_label_col
    target_col = target_label_col + relative_col
    if target_col < 1 or target_col > ws_new.max_column or isinstance(ws_new.cell(target_row, target_col), MergedCell):
        target_col = None
        for distance in range(1, min(ws_new.max_column, 12) + 1):
            for candidate_col in (target_label_col - distance, target_label_col + distance):
                if candidate_col < 1 or candidate_col > ws_new.max_column:
                    continue
                if not isinstance(ws_new.cell(target_row, candidate_col), MergedCell):
                    target_col = candidate_col
                    break
            if target_col:
                break
    if not target_col or not set_cell_value(ws_new, target_row, target_col, value):
        return 0
    highlight_wording_cell(ws_new, target_row, target_col)
    return 1


def copy_following_labeled_text(ws_prior_formula, ws_prior_values, ws_new, label_keywords, max_rows=8):
    """Copy the first populated response below a matching prompt."""
    source_anchor, _ = find_label_cell(ws_prior_formula, label_keywords)
    target_anchor, _ = find_label_cell(ws_new, label_keywords)
    if not source_anchor or not target_anchor:
        return 0

    source_row = source_col = None
    value = None
    for row in range(source_anchor + 1, min(ws_prior_formula.max_row, source_anchor + max_rows) + 1):
        for col in range(1, min(ws_prior_formula.max_column, 30) + 1):
            candidate = rollable_cell_value(
                ws_prior_formula.cell(row, col),
                ws_prior_values.cell(row, col),
            )
            if candidate in (None, ""):
                continue
            if isinstance(candidate, str) and candidate.startswith("="):
                continue
            source_row, source_col, value = row, col, candidate
            break
        if source_row:
            break
    if not source_row:
        return 0

    target_row = target_anchor + 1
    target_col = source_col
    if target_row > ws_new.max_row:
        return 0
    if target_col > ws_new.max_column or isinstance(ws_new.cell(target_row, target_col), MergedCell):
        target_col = next(
            (
                col for col in range(1, min(ws_new.max_column, 30) + 1)
                if not isinstance(ws_new.cell(target_row, col), MergedCell)
            ),
            None,
        )
    if not target_col or not set_cell_value(ws_new, target_row, target_col, value):
        return 0
    highlight_wording_cell(ws_new, target_row, target_col)
    return 1


def find_c_bkd_structure(ws):
    """Locate the C.00 BKD detail header, mapped base columns, and detail boundary."""
    field_keywords = OrderedDict([
        ("company", ("公司名称",)),
        ("subject", ("科目名称",)),
        ("bank", ("银行/存款机构名称", "存款机构名称")),
        ("account", ("账号",)),
        ("currency", ("币种",)),
        ("purpose", ("货币资金账户的性质和用途", "账户的性质和用途")),
    ])
    for row in range(1, min(ws.max_row, 80) + 1):
        columns = {}
        for field, keywords in field_keywords.items():
            col = find_header_col(ws, row, keywords, max_col=min(ws.max_column, 30))
            if col:
                columns[field] = col
        if len(columns) != len(field_keywords):
            continue
        marker_row = next(
            (
                candidate for candidate in range(row + 1, ws.max_row + 1)
                if row_has_any_keyword(
                    ws,
                    candidate,
                    ("本表格包括所有的货币资金信息", "本表格包括所有货币资金信息"),
                    1,
                    min(ws.max_column, 30),
                )
            ),
            None,
        )
        if marker_row:
            return row, columns, marker_row
    return None, {}, None


def process_c_bkd_basic_info(ws_prior_values, ws_new):
    """Always roll the six C.00 BKD account identity fields from the prior workbook."""
    source_header, source_cols, source_marker = find_c_bkd_structure(ws_prior_values)
    target_header, target_cols, target_marker = find_c_bkd_structure(ws_new)
    if not source_header or not target_header or not source_marker or not target_marker:
        return 0

    records = []
    for row in range(source_header + 1, source_marker):
        record = {field: ws_prior_values.cell(row, col).value for field, col in source_cols.items()}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    if not records:
        return 0

    data_start = target_header + 1
    available_rows = target_marker - data_start
    extra_rows = max(0, len(records) - available_rows)
    old_detail_end = target_marker - 1
    if extra_rows:
        shape_row = data_start
        insert_rows_preserving_sheet_metadata(ws_new, target_marker, extra_rows)
        for offset in range(extra_rows):
            copy_row_shape(ws_new, shape_row, target_marker + offset, translate_formula=True)
        target_marker += extra_rows

    copied = 0
    for index, record in enumerate(records):
        target_row = data_start + index
        if target_row > old_detail_end:
            copy_row_shape(ws_new, data_start, target_row, translate_formula=True)
        for field, value in record.items():
            if set_cell_value(ws_new, target_row, target_cols[field], value):
                copied += 1

    for row in range(data_start + len(records), target_marker):
        for col in target_cols.values():
            set_cell_value(ws_new, row, col, None)

    last_record_row = data_start + len(records) - 1
    total_row = next(
        (
            row for row in range(target_marker + 1, min(ws_new.max_row, target_marker + 8) + 1)
            if any(
                isinstance(ws_new.cell(row, col).value, str)
                and ws_new.cell(row, col).value.upper().startswith("=SUM(")
                for col in range(1, min(ws_new.max_column, 30) + 1)
            )
        ),
        None,
    )
    if total_row:
        for col in range(1, min(ws_new.max_column, 30) + 1):
            formula = ws_new.cell(total_row, col).value
            if isinstance(formula, str) and formula.upper().startswith("=SUM("):
                letter = get_column_letter(col)
                set_cell_value(ws_new, total_row, col, f"=SUM({letter}{data_start}:{letter}{last_record_row})")
        for row in range(total_row + 1, min(ws_new.max_row, total_row + 5) + 1):
            if not row_has_any_keyword(ws_new, row, ("C_Lead",), 1, min(ws_new.max_column, 30)):
                continue
            for col in range(1, min(ws_new.max_column, 30) + 1):
                formula = ws_new.cell(row, col).value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                letter = get_column_letter(col)
                updated = re.sub(rf"(?<![A-Za-z0-9_!']){letter}\d+", f"{letter}{total_row}", formula, count=1)
                set_cell_value(ws_new, row, col, updated)

    return copied


def process_c_cutoff_wording(ws_prior_formula, ws_prior_values, ws_new):
    """Roll the labeled C.03 cutoff period and its rationale when wording is enabled."""
    copied = copy_adjacent_labeled_value(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        ("使用的截止期间",),
    )
    copied += copy_following_labeled_text(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        ("描述所使用截止期间的理由", "所使用截止期间的理由"),
    )
    return copied


def process_j1_lead_wording(ws_prior_formula, ws_prior_values, ws_new):
    """Roll J.00 Lead wording by labels so shifted prior layouts stay aligned."""
    copied = 0

    source_header = next(
        (
            row
            for row in range(1, ws_prior_formula.max_row + 1)
            if normalize_text(ws_prior_formula.cell(row=row, column=2).value)
            in {"账户变更", "账户变动"}
        ),
        None,
    )
    target_header = next(
        (
            row
            for row in range(1, ws_new.max_row + 1)
            if normalize_text(ws_new.cell(row=row, column=2).value)
            in {"账户变更", "账户变动"}
        ),
        None,
    )
    if source_header and target_header:
        source_stop = find_row_containing(
            ws_prior_formula, "波动范围", (source_header + 1, ws_prior_formula.max_row)
        ) or min(ws_prior_formula.max_row + 1, source_header + 10)
        target_stop = find_row_containing(
            ws_new, "波动范围", (target_header + 1, ws_new.max_row)
        ) or min(ws_new.max_row + 1, target_header + 10)
        target_rows = {
            normalize_text(ws_new.cell(row=row, column=2).value): row
            for row in range(target_header + 1, target_stop)
            if normalize_text(ws_new.cell(row=row, column=2).value)
        }
        for source_row in range(source_header + 1, source_stop):
            label = normalize_text(ws_prior_formula.cell(row=source_row, column=2).value)
            target_row = target_rows.get(label)
            if not label or not target_row:
                continue
            for source_col in range(3, min(ws_prior_formula.max_column, 9) + 1):
                value = rollable_cell_value(
                    ws_prior_formula.cell(source_row, source_col),
                    ws_prior_values.cell(source_row, source_col),
                )
                if value in (None, ""):
                    continue
                if set_cell_value(ws_new, target_row, 3, value):
                    highlight_wording_cell(ws_new, target_row, 3)
                    ws_new.cell(target_row, 3).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
                    copied += 1
                break

    source_note = find_row_containing(
        ws_prior_formula, "波动说明", (1, ws_prior_formula.max_row)
    )
    target_note = find_row_containing(ws_new, "波动说明", (1, ws_new.max_row))
    if source_note and target_note:
        note_value = None
        for row in range(source_note + 1, min(ws_prior_formula.max_row, source_note + 5) + 1):
            for col in range(2, min(ws_prior_formula.max_column, 7) + 1):
                note_value = rollable_cell_value(
                    ws_prior_formula.cell(row, col),
                    ws_prior_values.cell(row, col),
                )
                if note_value not in (None, ""):
                    break
            if note_value not in (None, ""):
                break
        if note_value not in (None, "") and set_cell_value(ws_new, target_note + 1, 2, note_value):
            highlight_wording_cell(ws_new, target_note + 1, 2)
            ws_new.cell(target_note + 1, 2).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            copied += 1

    source_adjustment = find_row_containing(
        ws_prior_formula, "调整汇总表", (1, ws_prior_formula.max_row)
    )
    target_adjustment = find_row_containing(ws_new, "调整汇总表", (1, ws_new.max_row))
    if source_adjustment and target_adjustment:
        source_end = find_wording_section_end(
            ws_prior_formula, source_adjustment, max_rows=80
        )
        if source_end > source_adjustment:
            copied += copy_section_with_shape(
                ws_prior_formula,
                ws_new,
                source_adjustment,
                source_end,
                target_adjustment,
                col_start=2,
                col_end=min(18, ws_prior_formula.max_column, ws_new.max_column),
                highlight=True,
            )

    return copied


def process_j1_agree_notes(ws_prior_formula, ws_prior_values, ws_new):
    """Roll the J.01 Notes body into the existing Notes box."""
    copied = 0
    notes_row = None
    for row in range(1, ws_new.max_row + 1):
        if row_has_any_keyword(ws_new, row, ("Notes", "Notes：", "Notes:")):
            notes_row = row + 1
            break
    if not notes_row:
        return 0

    prior_notes_row = None
    for row in range(1, ws_prior_formula.max_row + 1):
        if row_has_any_keyword(ws_prior_formula, row, ("Notes", "Notes：", "Notes:")):
            prior_notes_row = row + 1
            break
    if not prior_notes_row:
        return 0

    copied += copy_wording_value_to_cell(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        prior_notes_row,
        2,
        notes_row,
        2,
    )
    return copied


def process_j1_cip_long_aging(ws_prior_formula, ws_prior_values, ws_new):
    """Roll populated J.03 CIP long-aging test rows when prior year has content."""
    copied = 0
    copied += copy_following_labeled_text(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        ("选择待测试项目的理由", "选取待测试项目的理由"),
    )
    copied += copy_following_labeled_text(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        ("标记图例", "记录已识别的异常情况"),
    )

    header_row = find_header_row(ws_prior_formula, "样本序号", (1, 40))
    target_header_row = find_header_row(ws_new, "样本序号", (1, 40))
    if not header_row or not target_header_row:
        return copied

    prior_records = []
    for row in range(header_row + 1, ws_prior_formula.max_row + 1):
        if row_has_any_keyword(ws_prior_formula, row, ("标记图例",)):
            break
        if row_has_content(ws_prior_formula, row, 3, min(ws_prior_formula.max_column, 9)):
            prior_records.append(row)

    if not prior_records:
        return copied

    marker_row = find_row_containing(ws_new, "标记图例", (target_header_row + 1, ws_new.max_row))
    if not marker_row:
        marker_row = ws_new.max_row + 1
    available_rows = max(0, marker_row - target_header_row - 1)
    extra_rows = max(0, len(prior_records) - available_rows)
    if extra_rows:
        insert_rows_preserving_sheet_metadata(ws_new, marker_row, extra_rows)
        for offset in range(extra_rows):
            copy_row_shape(ws_new, marker_row - 1, marker_row + offset, translate_formula=True)

    for idx, source_row in enumerate(prior_records):
        target_row = target_header_row + 1 + idx
        for col in range(2, min(ws_prior_formula.max_column, 9) + 1):
            copied += copy_wording_value_to_cell(
                ws_prior_formula,
                ws_prior_values,
                ws_new,
                source_row,
                col,
                target_row,
                col,
                table_like=True,
            )
    return copied


def process_j1_wording_sections(wb_prior_formula, wb_prior_values, wb_new, warnings_list=None):
    """Roll J1 wording sections that need fixed template placement."""
    copied = 0
    if is_visible_worksheet(wb_prior_formula, "J.00  Lead Sheet") and "J.00  Lead Sheet" in wb_new.sheetnames:
        copied += process_j1_lead_wording(
            wb_prior_formula["J.00  Lead Sheet"],
            wb_prior_values["J.00  Lead Sheet"],
            wb_new["J.00  Lead Sheet"],
        )
    if is_visible_worksheet(wb_prior_formula, "J.01 Agree SL to GL") and "J.01 Agree SL to GL" in wb_new.sheetnames:
        copied += process_j1_agree_notes(
            wb_prior_formula["J.01 Agree SL to GL"],
            wb_prior_values["J.01 Agree SL to GL"],
            wb_new["J.01 Agree SL to GL"],
        )

    prior_j03_sheet = find_visible_sheet_name(
        wb_prior_formula, lambda name: normalize_text(name).startswith("J.03")
    )
    new_j03_sheet = next((s for s in wb_new.sheetnames if normalize_text(s).startswith("J.03")), None)
    if prior_j03_sheet and new_j03_sheet:
        copied += process_j1_cip_long_aging(
            wb_prior_formula[prior_j03_sheet],
            wb_prior_values[prior_j03_sheet],
            wb_new[new_j03_sheet],
        )

    if copied and warnings_list is not None:
        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")
    return copied


def unmerge_ranges_intersecting(ws, row_start, row_end, col_start, col_end):
    """Unmerge ranges touching a bounded area."""
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.max_row < row_start
            or merged_range.min_row > row_end
            or merged_range.max_col < col_start
            or merged_range.min_col > col_end
        ):
            continue
        try:
            ws.unmerge_cells(str(merged_range))
        except KeyError:
            continue


def process_l1_lead_expectation_table(ws_prior_formula, ws_new):
    """Roll the complete L1 account-change expectation table, including added rows."""
    source_start = next(
        (
            row
            for row in range(1, ws_prior_formula.max_row + 1)
            if normalize_text(ws_prior_formula.cell(row=row, column=2).value) == "账户变动"
        ),
        None,
    )
    if not source_start:
        return 0

    source_stop = next(
        (
            row
            for row in range(source_start + 1, ws_prior_formula.max_row + 1)
            if normalize_text(ws_prior_formula.cell(row=row, column=2).value) == "波动范围"
        ),
        None,
    )
    target_stop = next(
        (
            row
            for row in range(1, ws_new.max_row + 1)
            if normalize_text(ws_new.cell(row=row, column=2).value) == "波动范围"
        ),
        None,
    )
    if not source_stop or not target_stop:
        return 0

    target_start = next(
        (
            row
            for row in range(1, target_stop)
            if normalize_text(ws_new.cell(row=row, column=2).value) == "账户变动"
        ),
        None,
    )
    if not target_start:
        prompt_row = find_row_containing(
            ws_new, "以下描述我们对无形资产的预期", (1, target_stop)
        )
        target_start = (prompt_row + 1) if prompt_row else None
    if not target_start:
        return 0

    source_end = source_stop - 1
    while source_end > source_start and not row_has_content(
        ws_prior_formula, source_end, 2, min(ws_prior_formula.max_column, 9)
    ):
        source_end -= 1

    source_len = source_end - source_start + 1
    target_capacity = max(0, target_stop - target_start)
    if source_len > target_capacity:
        extra_rows = source_len - target_capacity
        insert_rows_preserving_sheet_metadata(ws_new, target_stop, extra_rows)
        shift_local_formula_refs_after_insert(ws_new, target_stop, extra_rows)

    return copy_section_with_shape(
        ws_prior_formula,
        ws_new,
        source_start,
        source_end,
        target_start,
        col_start=2,
        col_end=min(9, ws_prior_formula.max_column, ws_new.max_column),
        highlight=True,
    )


def rebuild_l1_lead_total_formulas(ws):
    """Rebuild L1 Lead totals from every actual asset, amortization and impairment row."""
    header_row = find_header_row(ws, "期末审定数", (1, 100))
    if not header_row:
        return False
    total_row = find_total_row_after(ws, header_row)
    if not total_row:
        return False

    detail_rows = []
    for row in range(header_row + 1, total_row):
        name = normalize_text(ws.cell(row=row, column=3).value)
        if not name or name in {"合计", "净值"}:
            continue
        sign = -1 if ("累计摊销" in name or "减值准备" in name) else 1
        detail_rows.append((row, sign))
    if not detail_rows:
        return False

    for col in range(5, 12):
        col_letter = get_column_letter(col)
        terms = []
        for index, (row, sign) in enumerate(detail_rows):
            prefix = "-" if sign < 0 else ("+" if index else "")
            terms.append(f"{prefix}{col_letter}{row}")
        formula = "=" + "".join(terms)
        set_cell_value(ws, total_row, col, formula)

    set_cell_value(ws, total_row, 12, f"=IF(J{total_row}<>0,K{total_row}/J{total_row},1)")
    return True


def repair_l1_lead_detail_formulas(ws_prior_formula, ws_new):
    """Populate formulas for L1 detail rows that had to be added to the template."""
    prior_header = find_header_row(ws_prior_formula, "期末审定数", (1, 100))
    new_header = find_header_row(ws_new, "期末审定数", (1, 100))
    prior_total = find_total_row_after(ws_prior_formula, prior_header)
    new_total = find_total_row_after(ws_new, new_header)
    if not prior_header or not new_header or not prior_total or not new_total:
        return 0

    source_rows = {
        normalize_text(ws_prior_formula.cell(row=row, column=3).value): row
        for row in range(prior_header + 1, prior_total)
        if normalize_text(ws_prior_formula.cell(row=row, column=3).value)
    }
    fluctuation_row = next(
        (
            row
            for row in range(1, ws_new.max_row + 1)
            if normalize_text(ws_new.cell(row=row, column=2).value) == "波动范围"
        ),
        31,
    )
    amount_row = fluctuation_row + 1
    percent_row = fluctuation_row + 2
    repaired = 0
    for row in range(new_header + 1, new_total):
        name = normalize_text(ws_new.cell(row=row, column=3).value)
        source_row = source_rows.get(name)
        if not source_row:
            continue
        source_current = ws_prior_formula.cell(source_row, 5).value
        if ws_new.cell(row, 5).value in (None, "") and isinstance(source_current, str) and source_current.startswith("="):
            set_cell_value(ws_new, row, 5, source_current)
            repaired += 1
        formula_map = {
            7: f"=E{row}+F{row}",
            9: f"=G{row}+H{row}",
            11: f"=I{row}-J{row}",
            12: f"=IF(J{row}<>0,K{row}/J{row},1)",
            14: f'=IF(AND(ABS(K{row})>=$C${amount_row},ABS(L{row})>=$C${percent_row}),"是","否")',
        }
        for col, formula in formula_map.items():
            if ws_new.cell(row, col).value in (None, ""):
                set_cell_value(ws_new, row, col, formula)
                repaired += 1
    return repaired


def process_l1_wording_sections(wb_prior_formula, wb_new, warnings_list=None):
    """Roll only the structured L1 Lead expectation table handled outside generic wording."""
    prior_name = find_visible_sheet_name(
        wb_prior_formula, lambda name: normalize_text(name).lower().startswith("l1.00lead")
    )
    new_name = next(
        (name for name in wb_new.sheetnames if normalize_text(name).lower().startswith("l1.00lead")),
        None,
    )
    if not prior_name or not new_name:
        return 0
    copied = process_l1_lead_expectation_table(
        wb_prior_formula[prior_name], wb_new[new_name]
    )
    if copied and warnings_list is not None:
        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")
    return copied


def process_l2_lead_expectation_table(ws_prior_formula, ws_prior_values, ws_new):
    """Roll L2 prior-year expectation table into the current placeholder area."""
    prior_header_row = None
    for row in range(1, min(ws_prior_formula.max_row, 40) + 1):
        if normalize_text(ws_prior_formula.cell(row=row, column=3).value) == "账户变动":
            prior_header_row = row
            break
    if not prior_header_row:
        return 0

    target_start_row = 14
    data_rows = []
    for row in range(prior_header_row + 1, prior_header_row + 6):
        label = ws_prior_values.cell(row=row, column=3).value
        detail = ws_prior_values.cell(row=row, column=4).value
        if label not in (None, "") or detail not in (None, ""):
            data_rows.append((label, detail))

    if not data_rows:
        return 0

    unmerge_ranges_intersecting(ws_new, target_start_row, target_start_row + 2, 3, 7)
    clear_target_cells(ws_new, target_start_row, target_start_row + 2, 3, 7)

    set_cell_value(ws_new, target_start_row, 3, "账户变动")
    set_cell_value(ws_new, target_start_row, 4, "预期的依据和理由")
    set_cell_value(ws_new, target_start_row + 1, 3, "\n".join(str(label or "") for label, _ in data_rows))
    set_cell_value(ws_new, target_start_row + 1, 4, "\n".join(str(detail or "") for _, detail in data_rows))

    try:
        ws_new.merge_cells(start_row=target_start_row, start_column=4, end_row=target_start_row, end_column=7)
        ws_new.merge_cells(start_row=target_start_row + 1, start_column=3, end_row=target_start_row + 2, end_column=3)
        ws_new.merge_cells(start_row=target_start_row + 1, start_column=4, end_row=target_start_row + 2, end_column=7)
    except ValueError:
        pass

    for row in range(target_start_row, target_start_row + 3):
        ws_new.row_dimensions[row].height = 24 if row == target_start_row else 68

    copied = 0
    for row, col in (
        (target_start_row, 3),
        (target_start_row, 4),
        (target_start_row + 1, 3),
        (target_start_row + 1, 4),
    ):
        cell = ws_new.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        highlight_wording_cell(ws_new, row, col)
        copied += 1

    return copied


def process_l2_lead_notes(ws_prior_formula, ws_prior_values, ws_new):
    """Roll L2 Lead Notes body into the current Notes box."""
    prior_notes_row = next(
        (
            row for row in range(1, ws_prior_formula.max_row + 1)
            if normalize_text(ws_prior_formula.cell(row=row, column=2).value) == "Notes"
        ),
        None,
    )
    target_notes_row = next(
        (
            row for row in range(1, ws_new.max_row + 1)
            if normalize_text(ws_new.cell(row=row, column=2).value) == "Notes"
        ),
        None,
    )
    if not prior_notes_row or not target_notes_row:
        return 0

    return copy_wording_value_to_cell(
        ws_prior_formula,
        ws_prior_values,
        ws_new,
        prior_notes_row + 1,
        3,
        target_notes_row + 1,
        3,
    )


def process_l2_lead_adjustment_summary(ws_prior_formula, ws_prior_values, ws_new):
    """Roll a populated L2 Lead adjustment summary into the current template section."""
    source_start = find_row_containing(ws_prior_formula, "调整汇总表", (1, ws_prior_formula.max_row))
    target_start = find_row_containing(ws_new, "调整汇总表", (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0
    source_end = find_wording_section_end(ws_prior_formula, source_start, max_rows=20)
    if source_end <= source_start:
        return 0

    target_end = find_wording_section_end(ws_new, target_start, max_rows=20)
    source_data_len = source_end - source_start
    target_data_len = max(1, target_end - target_start)
    ensure_target_wording_capacity(
        ws_new,
        target_start + 1,
        target_data_len,
        source_data_len,
    )

    copied = 0
    for offset, source_row in enumerate(range(source_start + 1, source_end + 1)):
        target_row = target_start + 1 + offset
        for col in range(1, min(ws_prior_formula.max_column, ws_new.max_column, 40) + 1):
            value = source_wording_value(
                ws_prior_formula.cell(source_row, col),
                ws_prior_values.cell(source_row, col),
                table_like=True,
            )
            if value in (None, ""):
                continue
            if set_cell_value(ws_new, target_row, col, value):
                highlight_wording_cell(ws_new, target_row, col)
                copied += 1
    return copied


def find_l2_note_anchors(ws):
    """Return each actual L2 Notes prompt in worksheet order."""
    anchors = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 12) + 1):
            value = normalize_text(ws.cell(row=row, column=col).value)
            if value.lower().startswith("notes"):
                anchors.append((row, col))
                break
    return anchors


def l2_note_body_bounds(ws, anchor_row, anchor_col):
    """Locate the formatted response box immediately below an L2 Notes prompt."""
    candidates = [
        merged_range
        for merged_range in ws.merged_cells.ranges
        if anchor_row < merged_range.min_row <= anchor_row + 4
        and merged_range.max_col > merged_range.min_col
    ]
    if candidates:
        merged_range = min(candidates, key=lambda item: (item.min_row, item.min_col))
        return merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col
    body_row = anchor_row + 1
    body_col = min(ws.max_column, anchor_col + 1)
    return body_row, min(ws.max_row, body_row + 2), body_col, min(ws.max_column, body_col + 4)


def duplicate_bounded_row_block(ws, source_start, source_end, insert_at, col_end=40):
    """Clone a formatted row block without copying trailing worksheet plug-in columns."""
    block_length = source_end - source_start + 1
    source_merges = [
        deepcopy(merged_range)
        for merged_range in ws.merged_cells.ranges
        if merged_range.min_row >= source_start
        and merged_range.max_row <= source_end
        and merged_range.max_col <= col_end
    ]
    insert_rows_preserving_sheet_metadata(ws, insert_at, block_length)
    for offset in range(block_length):
        source_row = source_start + offset
        target_row = insert_at + offset
        for col in range(1, min(ws.max_column, col_end) + 1):
            copy_cell_shape(
                ws.cell(source_row, col),
                ws.cell(target_row, col),
                translate_formula=True,
            )
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    row_shift = insert_at - source_start
    for merged_range in source_merges:
        ws.merge_cells(
            start_row=merged_range.min_row + row_shift,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + row_shift,
            end_column=merged_range.max_col,
        )
    return insert_at, insert_at + block_length - 1


def process_l2_bkd_notes(ws_prior_formula, ws_prior_values, ws_new):
    """Roll every L2 BKD Notes response by occurrence, cloning a template box if needed."""
    source_anchors = find_l2_note_anchors(ws_prior_formula)
    target_anchors = find_l2_note_anchors(ws_new)
    if not source_anchors or not target_anchors:
        return 0

    while len(target_anchors) < len(source_anchors):
        template_row, template_col = target_anchors[-1]
        _, body_end, _, _ = l2_note_body_bounds(ws_new, template_row, template_col)
        new_start, _ = duplicate_bounded_row_block(
            ws_new,
            template_row,
            body_end,
            body_end + 1,
        )
        target_anchors.append((new_start, template_col))

    copied = 0
    for (source_row, source_col), (target_row, target_col) in zip(source_anchors, target_anchors):
        source_body_start, source_body_end, source_body_col, source_body_col_end = l2_note_body_bounds(
            ws_prior_formula,
            source_row,
            source_col,
        )
        value = None
        for row in range(source_body_start, source_body_end + 1):
            for col in range(source_body_col, source_body_col_end + 1):
                candidate = rollable_cell_value(
                    ws_prior_formula.cell(row, col),
                    ws_prior_values.cell(row, col),
                )
                if candidate not in (None, ""):
                    value = candidate
                    break
            if value not in (None, ""):
                break
        if value in (None, ""):
            continue

        target_body_row, _, target_body_col, _ = l2_note_body_bounds(ws_new, target_row, target_col)
        if set_cell_value(ws_new, target_body_row, target_body_col, value):
            highlight_wording_cell(ws_new, target_body_row, target_body_col)
            copied += 1
    return copied


def find_l2_bkd_header_row(ws):
    """Find either supported L2 detail header: 项目编码 or 序号 plus 项目名称."""
    for row in range(1, min(ws.max_row, 100) + 1):
        row_text = normalize_text(
            " ".join(
                str(ws.cell(row=row, column=col).value or "")
                for col in range(1, min(ws.max_column, 20) + 1)
            )
        )
        if "项目名称" in row_text and ("项目编码" in row_text or "序号" in row_text):
            return row
    return None


def find_visible_l2_bkd_sheet_name(workbook):
    """Choose the populated visible L2.01.1 sheet and ignore hidden duplicates."""
    candidates = []
    for name in workbook.sheetnames:
        if workbook[name].sheet_state != "visible" or "L2.01.1" not in name:
            continue
        ws = workbook[name]
        header_row = find_l2_bkd_header_row(ws)
        populated_rows = 0
        if header_row:
            total_row = find_total_row_after(ws, header_row) or min(ws.max_row + 1, header_row + 5000)
            for row in range(header_row + 1, total_row):
                if any(ws.cell(row=row, column=col).value not in (None, "") for col in (3, 4)):
                    populated_rows += 1
        candidates.append(((1 if header_row else 0, populated_rows), name))
    if not candidates:
        return None
    return max(candidates, key=lambda item: item[0])[1]


def process_l2_wording_sections(wb_prior_formula, wb_prior_values, wb_new, warnings_list=None):
    """Roll L2 wording sections that need placement across changed templates."""
    sheet_name = "L2.00 Lead"
    copied = 0
    if (
        is_visible_worksheet(wb_prior_formula, sheet_name)
        and is_visible_worksheet(wb_prior_values, sheet_name)
        and sheet_name in wb_new.sheetnames
    ):
        copied += process_l2_lead_expectation_table(
            wb_prior_formula[sheet_name],
            wb_prior_values[sheet_name],
            wb_new[sheet_name],
        )
        copied += process_l2_lead_notes(
            wb_prior_formula[sheet_name],
            wb_prior_values[sheet_name],
            wb_new[sheet_name],
        )
        copied += process_l2_lead_adjustment_summary(
            wb_prior_formula[sheet_name],
            wb_prior_values[sheet_name],
            wb_new[sheet_name],
        )

    prior_bkd_name = find_visible_l2_bkd_sheet_name(wb_prior_formula)
    new_bkd_name = find_visible_l2_bkd_sheet_name(wb_new)
    if prior_bkd_name and new_bkd_name and prior_bkd_name in wb_prior_values.sheetnames:
        copied += process_l2_bkd_notes(
            wb_prior_formula[prior_bkd_name],
            wb_prior_values[prior_bkd_name],
            wb_new[new_bkd_name],
        )

    if copied and warnings_list is not None:
        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")
    return copied


def clear_blank_borders(ws, row_start=1, row_end=None, col_start=1, col_end=None):
    """Remove borders from blank cells only."""
    no_border = Border()
    row_end = row_end or ws.max_row
    col_end = col_end or ws.max_column
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if cell.value in (None, ""):
                cell.border = no_border


def clear_cell_format(cell, clear_fill=True, clear_border=True):
    """Clear visual-only formatting on a single worksheet cell."""
    if isinstance(cell, MergedCell):
        return
    if clear_border:
        cell.border = Border()
    if clear_fill:
        cell.fill = PatternFill()


def clear_area_format(ws, row_start, row_end, col_start, col_end, clear_fill=True, clear_border=True):
    """Clear fill/border formatting in a bounded area."""
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            clear_cell_format(ws.cell(row=row, column=col), clear_fill=clear_fill, clear_border=clear_border)


def row_has_content(ws, row, col_start, col_end):
    """Return True when a row has actual content in a column band."""
    return any(
        ws.cell(row=row, column=col).value not in (None, "")
        for col in range(col_start, col_end + 1)
    )


def apply_thin_borders(ws, row_start, row_end, col_start, col_end):
    """Apply a simple table grid to a bounded area."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.border = copy(border)


def apply_thin_outer_border(ws, row_start, row_end, col_start, col_end):
    """Apply one continuous thin outline without internal fragmented borders."""
    thin = Side(style="thin", color="000000")
    empty = Side()
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=thin if col == col_start else empty,
                right=thin if col == col_end else empty,
                top=thin if row == row_start else empty,
                bottom=thin if row == row_end else empty,
            )
    anchor = ws.cell(row=row_start, column=col_start)
    if not isinstance(anchor, MergedCell):
        anchor.alignment = Alignment(wrap_text=True, vertical="top")


def format_uexp_wording_boxes(ws, row_start, row_end, col_start=3, col_end=7):
    """Normalize merged Uexp wording boxes and their thin outer borders."""
    covered_rows = set()
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row < row_start
            or merged_range.max_row > row_end
            or merged_range.min_col != col_start
            or merged_range.max_col != col_end
        ):
            continue
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        if anchor.value in (None, ""):
            continue
        apply_thin_outer_border(
            ws,
            merged_range.min_row,
            merged_range.max_row,
            merged_range.min_col,
            merged_range.max_col,
        )
        covered_rows.update(range(merged_range.min_row, merged_range.max_row + 1))

    for row in range(row_start, row_end + 1):
        if row in covered_rows:
            continue
        values = [
            ws.cell(row=row, column=col).value
            for col in range(col_start, col_end + 1)
            if ws.cell(row=row, column=col).value not in (None, "")
        ]
        if len(values) != 1 or ws.cell(row=row, column=col_start).value in (None, ""):
            continue
        unmerge_ranges_intersecting(ws, row, row, col_start, col_end)
        for col in range(col_start + 1, col_end + 1):
            set_cell_value(ws, row, col, None)
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        apply_thin_outer_border(ws, row, row, col_start, col_end)


def apply_borders_to_content_rows(ws, row_start, row_end, col_start, col_end):
    """Apply borders only to rows with content in a bounded area."""
    for row in range(row_start, row_end + 1):
        if row_has_content(ws, row, col_start, col_end):
            apply_thin_borders(ws, row, row, col_start, col_end)


def reset_dimension_styles(ws, last_row, last_col):
    """Drop whole-row/whole-column styles that create borders in empty sheet areas."""
    from openpyxl.styles.cell_style import StyleArray

    for row_idx in list(ws.row_dimensions):
        row_dim = ws.row_dimensions[row_idx]
        row_dim._style = StyleArray()
        if row_idx > last_row:
            del ws.row_dimensions[row_idx]

    for col_key in list(ws.column_dimensions):
        col_dim = ws.column_dimensions[col_key]
        min_col = col_dim.min or column_index_from_string(col_key)
        max_col = col_dim.max or min_col
        if min_col > last_col:
            del ws.column_dimensions[col_key]
            continue
        if max_col > last_col:
            col_dim.max = last_col
        col_dim._style = StyleArray()


def find_last_useful_row(ws):
    """Find the last row containing values or merged text."""
    last_row = 1
    for row in range(1, ws.max_row + 1):
        if row_has_content(ws, row, 1, ws.max_column):
            last_row = row
    for merged_range in ws.merged_cells.ranges:
        if ws.cell(merged_range.min_row, merged_range.min_col).value not in (None, ""):
            last_row = max(last_row, merged_range.max_row)
    return last_row


def prune_empty_cells_outside_area(ws, last_row, last_col):
    """Remove blank style-only cells outside the area we want Excel to display."""
    for key, cell in list(ws._cells.items()):
        row, col = key
        if row <= last_row and col <= last_col:
            continue
        if cell.value in (None, ""):
            del ws._cells[key]


def tidy_n_detail_sheet_borders(ws, header_row, total_row):
    """Remove copied empty-grid borders from N.01.01 while keeping useful tables framed."""
    clear_borders(ws, 1, ws.max_row, 1, ws.max_column)

    useful_col_end = min(ws.max_column, 24)
    useful_row_end = find_last_useful_row(ws)
    reset_dimension_styles(ws, useful_row_end, useful_col_end)
    prune_empty_cells_outside_area(ws, useful_row_end, useful_col_end)

    table_start_row = max(1, header_row - 1)
    table_end_row = total_row or ws.max_row
    if table_end_row + 1 <= ws.max_row and row_has_content(ws, table_end_row + 1, 2, ws.max_column):
        table_end_row += 1
    apply_thin_borders(ws, table_start_row, table_end_row, 2, useful_col_end)

    aging_summary_row = find_row_containing(ws, "账龄汇总表", (table_end_row + 1, ws.max_row))
    if aging_summary_row:
        apply_borders_to_content_rows(
            ws,
            aging_summary_row,
            min(aging_summary_row + 12, ws.max_row),
            2,
            min(useful_col_end, 6),
        )

    wording_start_row = find_row_containing(ws, "对于单项变动金额", (table_end_row + 1, ws.max_row))
    if wording_start_row:
        apply_borders_to_content_rows(
            ws,
            wording_start_row,
            ws.max_row,
            2,
            min(useful_col_end, 8),
        )

        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row >= wording_start_row
                and merged_range.min_col >= 2
                and merged_range.max_col <= min(useful_col_end, 8)
                and ws.cell(merged_range.min_row, merged_range.min_col).value not in (None, "")
            ):
                apply_thin_borders(
                    ws,
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )


def update_total_row_formulas(ws, total_row, old_total_row, data_start_row, data_end_row):
    """After inserting detail rows, keep total-row formulas pointed at the new range."""
    if not total_row or total_row == old_total_row:
        return

    def expand_sum(match):
        start_col, start_row, end_col, end_row = match.groups()
        start_row_num = int(start_row)
        end_row_num = int(end_row)
        if start_row_num == data_start_row and end_row_num == old_total_row - 1:
            return f"SUM({start_col}{start_row}:{end_col}{data_end_row})"
        return match.group(0)

    total_ref = re.compile(r"(\$?[A-Z]{1,3}\$?)(%d)(?!\d)" % old_total_row)
    sum_ref = re.compile(r"SUM\((\$?[A-Z]{1,3}\$?)(\d+):(\$?[A-Z]{1,3}\$?)(\d+)\)")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=col)
        formula = cell.value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        formula = sum_ref.sub(expand_sum, formula)
        formula = total_ref.sub(lambda m: f"{m.group(1)}{total_row}", formula)
        cell.value = formula


def shift_local_formula_refs_after_insert(ws, start_row, offset):
    """Shift simple same-sheet row references in formulas moved by row insertion."""
    if offset <= 0:
        return

    cell_ref = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3}\$?)(\d+)(?!\d)")

    for row in range(start_row + offset + 1, min(ws.max_row, start_row + offset + 30) + 1):
        for col in range(1, min(ws.max_column, 30) + 1):
            cell = ws.cell(row=row, column=col)
            formula = cell.value
            if not isinstance(formula, str) or not formula.startswith("=") or "!" in formula:
                continue

            def shift_ref(match):
                row_num = int(match.group(2))
                if row_num >= start_row:
                    return f"{match.group(1)}{row_num + offset}"
                return match.group(0)

            cell.value = cell_ref.sub(shift_ref, formula)


def assertion_key(value):
    """Normalize assertion labels used by CRA tables."""
    text = normalize_text(value)
    text = re.split(r"[（(]", text)[0]
    mapping = {
        "完整性": "完整性",
        "存在性": "存在性",
        "计价": "计价",
        "权利和义务": "权利义务",
        "权利义务": "权利义务",
        "列报和披露": "列报",
        "列报": "列报",
    }
    for key, mapped in mapping.items():
        if key in text:
            return mapped
    return text


def numeric_value(value):
    """Return numeric cell values for aggregation; ignore text/formula errors."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return None


def process_summary_sheet(ws_new, company_info, bs_date, company_name, warnings_list):
    """Fill common fields on the summary sheet."""
    date_obj = parse_date_value(bs_date)

    fill_adjacent_header_value(ws_new, ["客户名称"], company_name, search_range=(1, 15))

    for row in range(1, min(15, ws_new.max_row) + 1):
        for col in range(1, min(12, ws_new.max_column) + 1):
            value = ws_new.cell(row=row, column=col).value
            if value and "期末" in str(value):
                target = ws_new.cell(row=row, column=col + 1)
                if not isinstance(target, MergedCell):
                    target.value = date_obj if date_obj else bs_date
                    if date_obj:
                        target.number_format = "yyyy/mm/dd"
                break

    fill_adjacent_header_value(ws_new, ["记账本位币", "本位币"], company_info.get("functional_currency"), search_range=(1, 15))
    fill_adjacent_header_value(ws_new, ["适用会计准则", "会计准则"], company_info.get("accounting_standard"), search_range=(1, 15))

    pm_fields = [
        ("PM", ["重要性水平", "PM"]),
        ("TE", ["可容忍误差", "TE"]),
        ("SAD", ["名义金额", "SAD"]),
    ]
    for field, keywords in pm_fields:
        if company_info.get(field):
            fill_adjacent_header_value(ws_new, keywords, company_info[field], search_range=(1, 15))
        else:
            warnings_list.append(f"PMTE信息表中未找到{field}数据，请手动填写")

    cra_data = company_info.get("cra_data", {})
    if cra_data:
        for row in range(1, min(30, ws_new.max_row) + 1):
            assertion = assertion_key(ws_new.cell(row=row, column=3).value)
            if assertion and assertion in cra_data:
                level = cra_data[assertion].get("level")
                if level:
                    set_cell_value(ws_new, row, 4, level)

    for row in range(1, min(30, ws_new.max_row) + 1):
        for col in range(1, min(15, ws_new.max_column) + 1):
            value = ws_new.cell(row=row, column=col).value
            if isinstance(value, (datetime.datetime, datetime.date)):
                ws_new.cell(row=row, column=col).number_format = "yyyy/mm/dd"


def fill_basic_lead_header(ws_new, company_info, bs_date, company_name):
    """Fill common Lead Sheet header fields when no prior Lead Sheet exists."""
    date_obj = parse_date_value(bs_date)

    fill_adjacent_header_value(ws_new, ["客户"], company_name)

    for row in [2, 3, 4]:
        for col in [2, 3]:
            cell_val = ws_new.cell(row=row, column=col).value
            if cell_val and any(x in str(cell_val) for x in ["期末", "资产负债表日"]):
                target = ws_new.cell(row=row, column=col + 1)
                target.value = date_obj if date_obj else bs_date
                if date_obj:
                    target.number_format = "yyyy/mm/dd"
                break
        else:
            continue
        break

    for row in [2, 3, 4, 5]:
        for col in [2, 3]:
            cell_val = ws_new.cell(row=row, column=col).value
            if cell_val and "分析日期" in str(cell_val):
                target = ws_new.cell(row=row, column=col + 1)
                target.value = datetime.datetime.now()
                target.number_format = "yyyy/mm/dd"
                break
        else:
            continue
        break

    fill_adjacent_header_value(ws_new, ["记账本位币", "本位币"], company_info.get("functional_currency"))
    fill_adjacent_header_value(ws_new, ["适用会计准则", "会计准则"], company_info.get("accounting_standard"))
    normalize_lead_date_formats(ws_new)


def extract_company_info_from_pmte(pmte_path, company_name):
    if not pmte_path:
        return {
            "PM": None,
            "TE": None,
            "SAD": None,
            "RP": None,
            "Level": None,
        }
    """从PMTE信息表中提取公司信息"""
    if not os.path.exists(pmte_path):
        return {}

    wb = openpyxl.load_workbook(pmte_path, data_only=True)
    info = {
        "PM": None,
        "TE": None,
        "SAD": None,
        "RP": None,
        "Level": None,
    }

    # 读取PMTE Sheet
    if "PMTE" in wb.sheetnames:
        ws = wb["PMTE"]
        for row in range(2, ws.max_row + 1):
            cell_company = ws.cell(row=row, column=1).value
            company_key = normalize_text(company_name)
            cell_company_key = normalize_text(cell_company)
            if cell_company and (
                company_key in cell_company_key
                or cell_company_key in company_key
            ):
                info["Level"] = ws.cell(row=row, column=2).value
                info["RP"] = ws.cell(row=row, column=3).value
                info["PM"] = ws.cell(row=row, column=4).value
                info["TE"] = ws.cell(row=row, column=5).value
                info["SAD"] = ws.cell(row=row, column=6).value
                break

    wb.close()
    return info


def load_cra_data(cra_path, subject_code):
    """
    从CRA等级表中读取指定科目的认定等级和比例

    Args:
        cra_path: CRA等级表路径（与PMTE信息表同一个文件）
        subject_code: 科目代码（如"C"）

    Returns:
        dict: {
            认定名称: {
                "ratio": 比例值（如1.0表示100%）,
                "level": 等级文字（如"Low"）,
                "applicable": 是否适用（True/False）
            },
            ...
        }
    """
    if not os.path.exists(cra_path):
        return {}

    try:
        wb = openpyxl.load_workbook(cra_path, data_only=True)
        if "CRA" not in wb.sheetnames:
            wb.close()
            return {}

        ws = wb["CRA"]
        cra_data = {}

        # 中文等级到英文等级的映射
        level_map = {
            "极低": "Minimal",
            "低": "Low",
            "中等": "Moderate",
            "中": "Moderate",
            "高": "High",
            "很高": "High",
            "极高": "High",
            "不适用": "N/A",
            "N/A": "N/A",
        }

        # 遍历所有行，查找包含 "subject_code." 的认定行
        for row in range(1, ws.max_row + 1):
            c2_val = ws.cell(row=row, column=2).value
            if not c2_val or not isinstance(c2_val, str):
                continue

            # 检查是否以 "subject_code." 开头
            prefix = f"{subject_code}."
            if c2_val.startswith(prefix):
                # 提取认定名称
                # 格式如 "C. 货币资金-存在性"
                parts = c2_val.split("-")
                if len(parts) >= 2:
                    assertion = parts[-1].strip()

                    # 标准化认定名称（去除括号内容）
                    assertion_clean = assertion.split("(")[0].strip()

                    # 统一认定名称映射
                    assertion_map = {
                        "存在性": "存在性",
                        "完整性": "完整性",
                        "计价": "计价",
                        "计价/分摊": "计价",
                        "权利义务": "权利义务",
                        "列报": "列报",
                    }
                    for key, mapped in assertion_map.items():
                        if key in assertion_clean:
                            assertion_clean = mapped
                            break

                    # 读取C3列（是否适用）
                    c3_val = ws.cell(row=row, column=3).value
                    applicable = str(c3_val).strip().upper() != "X" if c3_val is not None else True

                    # 读取C6列（风险等级）
                    c6_val = ws.cell(row=row, column=6).value
                    level_cn = str(c6_val).strip() if c6_val else None
                    level_en = None
                    if level_cn:
                        for cn, en in level_map.items():
                            if cn in level_cn:
                                level_en = en
                                break

                    # 读取C9列（比例值）
                    c9_val = ws.cell(row=row, column=9).value
                    ratio = None
                    if c9_val is not None:
                        try:
                            ratio = float(c9_val)
                        except (ValueError, TypeError):
                            pass

                    cra_data[assertion_clean] = {
                        "ratio": ratio,
                        "level": level_en,
                        "applicable": applicable
                    }

        wb.close()
        return cra_data
    except Exception:
        return {}


def process_lead_sheet(ws_prior, ws_new, ws_prior_values, company_info, bs_date, company_name, lead_config, warnings_list):
    """
    处理Lead Sheet：填写表头 + Roll Forward期初数

    Args:
        ws_prior: 上年底稿的Lead Sheet（含公式）
        ws_new: 新底稿的Lead Sheet
        ws_prior_values: 上年底稿的Lead Sheet（含计算值，data_only=True）
        company_info: 公司信息字典
        bs_date: 资产负债表日期
        company_name: 公司名称
        lead_config: Lead Sheet配置
        warnings_list: 警告列表

    Returns:
        复制的期初数数量
    """
    date_obj = parse_date_value(bs_date)

    # 1. 填写表头（尝试多个位置）
    # 客户名称通常在第2行第2列或附近
    for row in [2, 3, 4]:
        for col in [2, 3, 4]:
            cell_val = ws_new.cell(row=row, column=col).value
            if cell_val and "客户" in str(cell_val):
                ws_new.cell(row=row, column=col + 1, value=company_name)
                break
        else:
            continue
        break

    # 资产负债表日期 - 统一格式
    for row in [2, 3, 4]:
        for col in [2, 3]:
            cell_val = ws_new.cell(row=row, column=col).value
            if cell_val and any(x in str(cell_val) for x in ["期末", "资产负债表日"]):
                # 找到日期单元格（通常在旁边一列）
                target = ws_new.cell(row=row, column=col + 1)
                if target.value is None or "202" in str(target.value):
                    target.value = date_obj if date_obj else bs_date
                    if date_obj:
                        target.number_format = "yyyy/mm/dd"
                break
        else:
            continue
        break

    # 分析日期
    for row in [2, 3, 4, 5]:
        for col in [2, 3]:
            cell_val = ws_new.cell(row=row, column=col).value
            if cell_val and "分析日期" in str(cell_val):
                target = ws_new.cell(row=row, column=col + 1)
                target.value = datetime.datetime.now()
                target.number_format = "yyyy/mm/dd"
                break
        else:
            continue
        break

    # 用户手工输入的基础信息
    fill_adjacent_header_value(
        ws_new,
        ["记账本位币", "本位币"],
        company_info.get("functional_currency")
    )
    fill_adjacent_header_value(
        ws_new,
        ["适用会计准则", "会计准则"],
        company_info.get("accounting_standard")
    )

    normalize_lead_date_formats(ws_new)

    # TE和SAD - 检查PMTE中是否找到
    te_found = False
    sad_found = False

    if company_info.get("TE"):
        for row in [2, 3, 4, 5, 6, 7]:
            for col in [2, 3]:
                cell_val = ws_new.cell(row=row, column=col).value
                if cell_val and ("可容忍误差" in str(cell_val) or "TE" in str(cell_val)):
                    ws_new.cell(row=row, column=col + 1, value=company_info["TE"])
                    te_found = True
                    break
            else:
                continue
            break
    else:
        warnings_list.append("PMTE信息表中未找到TE数据，请手动填写")

    if company_info.get("SAD"):
        for row in [2, 3, 4, 5, 6, 7]:
            for col in [2, 3]:
                cell_val = ws_new.cell(row=row, column=col).value
                if cell_val and ("名义金额" in str(cell_val) or "SAD" in str(cell_val)):
                    ws_new.cell(row=row, column=col + 1, value=company_info["SAD"])
                    sad_found = True
                    break
            else:
                continue
            break
    else:
        warnings_list.append("PMTE信息表中未找到SAD数据，请手动填写")

    # CRA信息 - 填写风险等级(Level)和风险系数(RP)
    if company_info.get("Level"):
        # 在CRA区域查找"实质性"等行，填入Level
        for row in range(13, 25):
            c2_val = ws_new.cell(row=row, column=2).value
            if c2_val and "实质性" in str(c2_val):
                # C列填入Level（风险等级，如"Low", "Moderate", "High"）
                ws_new.cell(row=row, column=3).value = company_info["Level"]
                break

    if company_info.get("RP"):
        # 在CRA区域查找"实质性"等行，填入RP
        for row in range(13, 25):
            c2_val = ws_new.cell(row=row, column=2).value
            if c2_val and "实质性" in str(c2_val):
                # D列填入RP（风险系数百分比）
                ws_new.cell(row=row, column=4).value = company_info["RP"]
                break

    # 新增：根据CRA等级表动态填写每个认定的等级和比例
    cra_data = company_info.get("cra_data", {})
    if cra_data:
        # 获取C5的值（Threshold，用于计算比例）
        c5_value = ws_new.cell(row=5, column=3).value
        # C5可能是公式，尝试获取数值
        try:
            if c5_value is not None:
                c5_num = float(c5_value)
            else:
                c5_num = None
        except (ValueError, TypeError):
            c5_num = None

        # 遍历CRA区域的认定行（R15-R19）
        for row in range(15, 20):
            c2_val = ws_new.cell(row=row, column=2).value
            if not c2_val:
                continue

            # 提取认定名称（去掉括号内容，同时处理全角和半角括号）
            assertion_name = re.split(r'[（(]', str(c2_val))[0].strip()
            # 标准化认定名称
            assertion_map = {
                "存在性": "存在性",
                "完整性": "完整性",
                "计价": "计价",
                "计价/分摊": "计价",
                "权利义务": "权利义务",
                "列报": "列报",
            }
            for key, mapped in assertion_map.items():
                if key in assertion_name:
                    assertion_name = mapped
                    break

            # 在CRA数据中查找对应认定
            if assertion_name in cra_data:
                data = cra_data[assertion_name]

                # 填入等级（C列）
                if data.get("level"):
                    ws_new.cell(row=row, column=3).value = data["level"]

                # 如果比例与公式默认不同，覆盖D列
                if data.get("ratio") is not None:
                    # 计算当前等级对应的默认比例
                    current_level = data.get("level", "")
                    default_ratio = None
                    if current_level == "Minimal":
                        default_ratio = 1.0
                    elif current_level == "Low":
                        default_ratio = 0.75
                    elif current_level == "Moderate":
                        default_ratio = 0.50
                    elif current_level == "High":
                        default_ratio = 0.25

                    # 如果用户比例与默认比例不同，覆盖D列
                    if default_ratio is None or abs(data["ratio"] - default_ratio) > 0.001:
                        # 将D列的公式替换为计算值
                        ratio_value = data["ratio"]
                        # 如果C5有值，计算实际金额；否则直接写入比例值
                        if c5_num is not None:
                            d_value = c5_num * ratio_value
                        else:
                            d_value = ratio_value
                        ws_new.cell(row=row, column=4).value = d_value

    # 2. Roll Forward期初数（从计算值工作表读取）
    overwrite_opening_formulas = lead_config.get("overwrite_opening_formulas", False)

    header_search_text = lead_config.get("header_search_text", "期末审定数")
    prior_header_row = find_header_row(ws_prior_values, header_search_text, (1, 80))
    new_header_row = find_header_row(ws_new, header_search_text, (1, 80))

    if not prior_header_row or not new_header_row:
        return 0

    prior_closing_col = (
        find_header_col(ws_prior_values, prior_header_row, ["期末审定数"])
        or lead_config.get("closing_col", 9)
    )
    opening_col = (
        find_header_col(ws_new, new_header_row, ["上期末审定数", "期初审定数", "上年审定数"])
        or lead_config.get("opening_col", 10)
    )

    descriptor_rules = [
        (["账套名称", "账套编码", "账套"], ["账套名称", "账套编码", "账套"]),
        (["总账科目编码", "科目编码"], ["总账科目编码", "科目编码"]),
        (["科目名称", "报表科目", "项目名称"], ["科目名称", "报表科目", "项目名称"]),
        (["索引号", "索引"], ["索引号", "索引"]),
    ]
    descriptor_cols = []
    for prior_keywords, new_keywords in descriptor_rules:
        source_col = find_header_col_near(ws_prior_values, prior_header_row, prior_keywords)
        target_col = find_header_col_near(ws_new, new_header_row, new_keywords)
        if source_col and target_col:
            descriptor_cols.append((source_col, target_col))
    if not descriptor_cols:
        descriptor_cols = [(2, 2), (3, 3), (4, 4)]

    prior_data_rows = []
    for row in range(prior_header_row + 1, ws_prior_values.max_row + 1):
        if row_contains_any(ws_prior_values, row, ["合计"], columns=range(2, min(8, ws_prior_values.max_column) + 1)):
            break
        if row_contains_any(ws_prior_values, row, ["check with", "Diff", "波动说明", "Notes：", "Notes:"], columns=range(2, min(8, ws_prior_values.max_column) + 1)):
            break
        if is_table_end_marker_row(ws_prior_values, row):
            continue
        if row_contains_any(ws_prior_values, row, ["Rx"], columns=range(2, min(8, ws_prior_values.max_column) + 1)):
            continue

        closing_val = ws_prior_values.cell(row=row, column=prior_closing_col).value
        descriptor_values = {
            target_col: ws_prior_values.cell(row=row, column=source_col).value
            for source_col, target_col in descriptor_cols
        }
        if closing_val is not None or any(value not in (None, "") for value in descriptor_values.values()):
            prior_data_rows.append((row, descriptor_values, closing_val))

    new_data_start_row = new_header_row + 1
    new_total_row = None
    total_row_keywords = ["合计", *lead_config.get("total_row_keywords", [])]
    for row in range(new_data_start_row, ws_new.max_row + 1):
        if row_contains_any(
            ws_new,
            row,
            total_row_keywords,
            columns=range(2, min(8, ws_new.max_column) + 1),
        ):
            new_total_row = row
            break

    if lead_config.get("match_existing_rows_only", False):
        copied = 0
        if new_total_row:
            for _, descriptor_values, closing_val in prior_data_rows:
                if closing_val is None:
                    continue
                descriptor_keys = [
                    normalize_text(value)
                    for value in descriptor_values.values()
                    if value not in (None, "")
                ]
                if not descriptor_keys:
                    continue
                for row in range(new_data_start_row, new_total_row):
                    if row_contains_any(ws_new, row, ["Rx", "A3", "Diff", "波动说明"], columns=range(2, min(8, ws_new.max_column) + 1)):
                        continue
                    row_keys = [
                        normalize_text(ws_new.cell(row=row, column=col).value)
                        for col in sorted({col for _, col in descriptor_cols})
                    ]
                    match_descriptor_col = lead_config.get("match_descriptor_col")
                    if match_descriptor_col:
                        source_key = normalize_text(descriptor_values.get(match_descriptor_col))
                        target_key = normalize_text(
                            ws_new.cell(row=row, column=match_descriptor_col).value
                        )
                        if not source_key or source_key != target_key:
                            continue
                    elif not any(key and key in row_keys for key in descriptor_keys):
                        continue
                    if lead_config.get("copy_matched_descriptors", False):
                        for target_col, value in descriptor_values.items():
                            if value not in (None, ""):
                                set_cell_value(ws_new, row, target_col, value)
                    existing_value = ws_new.cell(row=row, column=opening_col).value
                    keeps_formula = isinstance(existing_value, str) and existing_value.startswith("=")
                    if overwrite_opening_formulas or not keeps_formula:
                        if set_cell_value(ws_new, row, opening_col, closing_val):
                            copied += 1
                    break
        return copied

    template_data_count = 0
    if new_total_row:
        available_template_rows = 0
        for row in range(new_data_start_row, new_total_row):
            if row_contains_any(ws_new, row, ["Rx", "A3", "Diff", "波动说明"], columns=range(2, min(8, ws_new.max_column) + 1)):
                continue
            available_template_rows += 1
            if any(
                ws_new.cell(row=row, column=col).value not in (None, "")
                for col in sorted({col for _, col in descriptor_cols} | {opening_col})
            ):
                template_data_count += 1
        template_data_count = max(template_data_count, available_template_rows)
    else:
        template_data_count = len(prior_data_rows)

    new_total_row_adjusted = new_total_row
    rows_to_insert = 0
    formula_source_row = None
    if new_total_row:
        for row in range(new_total_row - 1, new_data_start_row - 1, -1):
            if any(
                isinstance(ws_new.cell(row=row, column=col).value, str)
                and ws_new.cell(row=row, column=col).value.startswith("=")
                for col in range(1, min(ws_new.max_column, 20) + 1)
            ):
                formula_source_row = row
                break
    formula_source_row = formula_source_row or new_data_start_row

    if new_total_row and len(prior_data_rows) > template_data_count:
        rows_to_insert = len(prior_data_rows) - template_data_count
        insert_rows_preserving_sheet_metadata(ws_new, new_total_row, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, formula_source_row, new_total_row + offset, translate_formula=True)

        new_total_row_adjusted = new_total_row + rows_to_insert
        update_total_row_formulas(
            ws_new,
            new_total_row_adjusted,
            new_total_row,
            new_data_start_row,
            new_total_row_adjusted - 1,
        )
        shift_local_formula_refs_after_insert(ws_new, new_total_row, rows_to_insert)

    copied = 0
    for idx, (prior_row, descriptor_values, closing_val) in enumerate(prior_data_rows):
        target_row = new_data_start_row + idx if new_total_row else prior_row

        if lead_config.get("copy_formula_row_for_all_details", False) and target_row != formula_source_row:
            copy_row_shape(ws_new, formula_source_row, target_row, translate_formula=True)

        for target_col, value in descriptor_values.items():
            set_cell_value(ws_new, target_row, target_col, value)

        existing_value = ws_new.cell(row=target_row, column=opening_col).value
        keeps_formula = isinstance(existing_value, str) and existing_value.startswith("=")
        if closing_val is not None and (overwrite_opening_formulas or not keeps_formula):
            if set_cell_value(ws_new, target_row, opening_col, closing_val):
                copied += 1

    if lead_config.get("strict_detail_sum_formulas", False) and new_total_row_adjusted and prior_data_rows:
        data_end_row = new_data_start_row + len(prior_data_rows) - 1
        sum_formula = re.compile(r"^=SUM\((\$?[A-Z]{1,3}\$?)\d+:(\$?[A-Z]{1,3}\$?)\d+\)$", re.IGNORECASE)
        for col in range(1, ws_new.max_column + 1):
            cell = ws_new.cell(row=new_total_row_adjusted, column=col)
            match = sum_formula.match(str(cell.value or ""))
            if match:
                cell.value = f"=SUM({match.group(1)}{new_data_start_row}:{match.group(2)}{data_end_row})"

    if lead_config.get("clear_extra_template_rows", False) and new_total_row:
        first_extra_row = new_data_start_row + len(prior_data_rows)
        for row in range(first_extra_row, new_total_row):
            for col in range(1, min(ws_new.max_column, 20) + 1):
                set_cell_value(ws_new, row, col, None)

    clear_current_cols = lead_config.get("clear_current_period_cols", [])
    if clear_current_cols:
        for idx, (prior_row, descriptor_values, closing_val) in enumerate(prior_data_rows):
            target_row = new_data_start_row + idx if new_total_row else prior_row
            if not any(value not in (None, "") for value in descriptor_values.values()):
                continue
            for col in clear_current_cols:
                set_cell_value(ws_new, target_row, col, None)

    # 3. 处理For Disclosure表（如果存在）
    # 查找"For Disclosure"相关行
    disclosure_start_row = None
    for row in range(1, ws_prior_values.max_row + 1):
        for col in range(1, min(10, ws_prior_values.max_column + 1)):
            val = ws_prior_values.cell(row=row, column=col).value
            if val and "Disclosure" in str(val):
                disclosure_start_row = row
                break
        if disclosure_start_row:
            break

    if disclosure_start_row:
        # 在For Disclosure区域内查找包含"期末审定数"的表头行（旧文件）
        fd_prior_header_row = None
        fd_prior_closing_col = None
        fd_prior_opening_col = None

        for row in range(disclosure_start_row, min(disclosure_start_row + 10, ws_prior_values.max_row + 1)):
            for col in range(1, min(20, ws_prior_values.max_column + 1)):
                val = ws_prior_values.cell(row=row, column=col).value
                if val and "期末审定数" in str(val):
                    fd_prior_header_row = row
                    fd_prior_closing_col = col
                    # 查找"期初审定数"列（通常在下一列）
                    for next_col in range(col + 1, min(col + 5, ws_prior_values.max_column + 1)):
                        next_val = ws_prior_values.cell(row=row, column=next_col).value
                        if next_val and "期初" in str(next_val):
                            fd_prior_opening_col = next_col
                            break
                    break
            if fd_prior_header_row:
                break

        # 在新文件中查找For Disclosure区域的表头行（通过"Disclosure"和"报表科目"定位）
        fd_new_header_row = None
        fd_new_opening_col = None

        for row in range(1, ws_new.max_row + 1):
            val = ws_new.cell(row=row, column=2).value  # C2通常是"报表科目"
            if val and ("报表科目" in str(val) or "项目名称" in str(val)):
                # 检查上方几行是否有"Disclosure"或"For"
                found_disclosure = False
                for check_row in range(max(1, row - 5), row):
                    for check_col in range(1, 5):
                        check_val = ws_new.cell(row=check_row, column=check_col).value
                        if check_val and "Disclosure" in str(check_val):
                            found_disclosure = True
                            break
                    if found_disclosure:
                        break
                if found_disclosure:
                    fd_new_header_row = row
                    # 查找"期初"列（在表头行中搜索"期初"）
                    # 注意：新文件可能包含公式而非文本，所以尝试多种方式
                    for col in range(1, min(20, ws_new.max_column + 1)):
                        col_val = ws_new.cell(row=row, column=col).value
                        if col_val and "期初" in str(col_val):
                            fd_new_opening_col = col
                            break
                    # 如果没找到"期初"，尝试找"期末审定数"然后+1
                    if not fd_new_opening_col:
                        for col in range(1, min(20, ws_new.max_column + 1)):
                            col_val = ws_new.cell(row=row, column=col).value
                            if col_val and "期末审定数" in str(col_val):
                                fd_new_opening_col = col + 1
                                break
                    # 如果还是没找到，默认C4（通常For Disclosure表结构固定）
                    if not fd_new_opening_col:
                        fd_new_opening_col = 4
                    break

        # 按项目名称匹配复制
        if fd_prior_header_row and fd_new_header_row and fd_prior_closing_col and fd_new_opening_col:
            # 构建旧文件的报表科目→期末审定数字典
            prior_items = {}
            for row in range(fd_prior_header_row + 1, ws_prior_values.max_row + 1):
                item_name = ws_prior_values.cell(row=row, column=2).value  # C2通常是报表科目
                if item_name and str(item_name).strip():
                    val = ws_prior_values.cell(row=row, column=fd_prior_closing_col).value
                    if val is not None:
                        try:
                            prior_items[str(item_name).strip()] = float(val)
                        except (ValueError, TypeError):
                            prior_items[str(item_name).strip()] = val

            # 在新文件中查找匹配的项目并填充
            for row in range(fd_new_header_row + 1, ws_new.max_row + 1):
                item_name = ws_new.cell(row=row, column=2).value
                if item_name and str(item_name).strip() in prior_items:
                    ws_new.cell(row=row, column=fd_new_opening_col).value = prior_items[str(item_name).strip()]
                    copied += 1

    return copied


def process_k01(ws_prior, ws_new, k01_config):
    """
    处理K.01 Agree SL to GL：复制表头 + 年初余额

    Args:
        ws_prior: 上年底稿的K.01工作表（含计算值）
        ws_new: 新底稿的K.01工作表
        k01_config: K.01配置

    Returns:
        复制的数据数量
    """
    if not k01_config.get("has_k01", False):
        return 0

    header_row = k01_config.get("header_row", 10)

    roll_forward_groups = k01_config.get("roll_forward_groups", [])
    if roll_forward_groups:
        copied = 0
        for group in roll_forward_groups:
            group_name = group.get("group")
            source_detail = group.get("source_detail", "期末余额")
            target_detail = group.get("target_detail", "期初余额")
            value_cols = group.get("value_cols", [5, 7, 9])

            source_row = find_group_detail_row(ws_prior, group_name, source_detail)
            target_row = find_group_detail_row(ws_new, group_name, target_detail)
            if not source_row or not target_row:
                continue

            for col in value_cols:
                value = ws_prior.cell(row=source_row, column=col).value
                if value is not None:
                    ws_new.cell(row=target_row, column=col).value = value
                    copied += 1
        return copied

    if k01_config.get("match_categories", False):
        return process_k01_by_category(ws_prior, ws_new, header_row, k01_config)

    dynamic_copied = process_k01_by_category(ws_prior, ws_new, header_row, {})
    if dynamic_copied:
        return dynamic_copied

    # 2. 复制年初余额数据
    opening_balance_rows = k01_config.get("opening_balance_rows", [])
    copied = 0

    for balance_row in opening_balance_rows:
        for col in range(1, ws_prior.max_column + 1):
            prior_cell = ws_prior.cell(row=balance_row, column=col)
            new_cell = ws_new.cell(row=balance_row, column=col)
            if prior_cell.value is not None:
                new_cell.value = prior_cell.value
                copied += 1

    return copied


def find_k01_category_groups(ws, header_row):
    """Find K.01 category groups by the category label row."""
    groups = []
    for col in range(1, ws.max_column + 1):
        label = ws.cell(row=header_row, column=col).value
        if not label:
            continue
        label_key = normalize_text(label)
        book_header = normalize_text(ws.cell(row=header_row + 1, column=col - 1).value) if col > 1 else ""
        adjust_header = normalize_text(ws.cell(row=header_row + 1, column=col).value)
        audit_header = normalize_text(ws.cell(row=header_row + 1, column=col + 1).value)
        if "账面数" not in book_header or "调整" not in adjust_header or "审定数" not in audit_header:
            continue
        groups.append({
            "name": str(label).strip(),
            "name_key": label_key,
            "book_col": col - 1,
            "adjust_col": col,
            "audit_col": col + 1,
        })
    return groups


def find_k01_section_rows(ws):
    """Return K.01 roll-forward source/target rows by section label."""
    ordered_keys = ("cost", "depreciation", "impairment")
    sections = OrderedDict((key, {}) for key in ordered_keys)
    opening_rows = []
    closing_rows = []

    for row in range(1, ws.max_row + 1):
        detail_value = normalize_text(ws.cell(row=row, column=3).value)
        if "年初余额" in detail_value:
            opening_rows.append(row)
        if "年末余额" in detail_value or "期末余额" in detail_value:
            closing_rows.append(row)

    for idx, key in enumerate(ordered_keys):
        if idx < len(opening_rows):
            sections[key]["opening"] = opening_rows[idx]
        if idx < len(closing_rows):
            sections[key]["closing"] = closing_rows[idx]
    return sections


def category_matches(source_key, target_key):
    """Return True when a prior category should roll to a target category."""
    if not source_key or not target_key:
        return False
    if "合计" in source_key or "合计" in target_key:
        return False
    if "…" in target_key or "【" in target_key:
        return False
    return source_key in target_key or target_key in source_key


def is_placeholder_category(category_key):
    if not category_key:
        return True
    return any(marker in category_key for marker in ("…", "...", "【", "】", "[…]"))


def is_total_category(category_key):
    return "合计" in category_key or "总计" in category_key


def process_k01_by_category(ws_prior, ws_new, header_row, k01_config):
    """Roll prior year-end K.01 balances to current opening rows by category."""
    prior_groups = [
        group for group in find_k01_category_groups(ws_prior, header_row)
        if not is_total_category(group["name_key"]) and not is_placeholder_category(group["name_key"])
    ]
    configured_categories = k01_config.get("categories", [])
    dynamic_target_groups = [
        group for group in find_k01_category_groups(ws_new, header_row)
        if not is_total_category(group["name_key"])
    ]
    if dynamic_target_groups:
        target_groups = dynamic_target_groups
    elif configured_categories:
        target_groups = []
        for category in configured_categories:
            audit_col = category.get("audit_col")
            book_col = category.get("book_col")
            if not audit_col or not book_col:
                continue
            target_groups.append({
                "name": category.get("name", ""),
                "name_key": normalize_text(category.get("name", "")),
                "book_col": book_col,
                "adjust_col": audit_col - 1,
                "audit_col": audit_col,
            })
    else:
        target_groups = []

    sections_prior = find_k01_section_rows(ws_prior)
    sections_new = find_k01_section_rows(ws_new)

    copied = 0
    used_source_indexes = set()
    for target in target_groups:
        target_key = target["name_key"]
        if is_total_category(target_key):
            continue

        source = None
        if not is_placeholder_category(target_key):
            for source_index, group in enumerate(prior_groups):
                if source_index in used_source_indexes:
                    continue
                if category_matches(group["name_key"], target_key):
                    source = group
                    used_source_indexes.add(source_index)
                    break

        if source is None:
            for source_index, group in enumerate(prior_groups):
                if source_index not in used_source_indexes:
                    source = group
                    used_source_indexes.add(source_index)
                    break

        if not source:
            continue

        set_cell_value(ws_new, header_row, target["adjust_col"], source["name"])
        for section_key in ("cost", "depreciation", "impairment"):
            source_row = sections_prior.get(section_key, {}).get("closing")
            target_row = sections_new.get(section_key, {}).get("opening")
            if not source_row or not target_row:
                continue
            for source_col, target_col in (
                (source["book_col"], target["book_col"]),
                (source["adjust_col"], target["adjust_col"]),
            ):
                value = ws_prior.cell(row=source_row, column=source_col).value
                if value is not None:
                    set_cell_value(ws_new, target_row, target_col, value)
                    copied += 1

    return copied


def find_group_detail_row(ws, group_name, detail_name):
    """Find a row by group label in column B and detail label in column C."""
    group_key = normalize_text(group_name)
    detail_key = normalize_text(detail_name)
    if not group_key or not detail_key:
        return None

    in_group = False
    for row in range(1, ws.max_row + 1):
        group_value = normalize_text(ws.cell(row=row, column=2).value)
        detail_value = normalize_text(ws.cell(row=row, column=3).value)

        if group_value:
            in_group = group_key in group_value or group_value in group_key

        if in_group and detail_key in detail_value:
            return row

    return None


def find_schedule_section_row(ws, section_name, detail_name):
    """Find a row in a roll-forward schedule by section in column B and detail in column C."""
    section_key = normalize_text(section_name)
    detail_key = normalize_text(detail_name)
    in_section = False

    for row in range(1, ws.max_row + 1):
        section_value = normalize_text(ws.cell(row=row, column=2).value)
        detail_value = normalize_text(ws.cell(row=row, column=3).value)

        if section_value:
            in_section = section_key in section_value or section_value in section_key

        if in_section and detail_key in detail_value:
            return row

    return None


def find_schedule_group_columns(ws):
    """Return roll-forward schedule category columns keyed by category label."""
    groups = []
    for col in range(1, ws.max_column + 1):
        label = ws.cell(row=2, column=col).value
        if not label:
            continue
        label_text = str(label).strip()
        if not label_text:
            continue
        book_header = normalize_text(ws.cell(row=3, column=col - 1).value) if col > 1 else ""
        audit_header = normalize_text(ws.cell(row=3, column=col + 1).value)
        if "账面数" not in book_header or "审定数" not in audit_header:
            continue

        groups.append({
            "name": label_text,
            "name_key": normalize_text(label_text),
            "book_col": col - 1,
            "adjust_col": col,
            "audit_col": col + 1,
        })

    return groups


def get_schedule_total_col(ws):
    for group in find_schedule_group_columns(ws):
        if "合计" in group["name_key"]:
            return group["audit_col"]
    return None


def l1_category_key(name):
    key = normalize_text(name)
    if "土地" in key:
        return "土地"
    if "非专利" in key:
        return "非专利"
    if "专利" in key:
        return "专利"
    if "软件" in key or "计算机" in key:
        return "软件"
    if "其他" in key:
        return "其他"
    return key


def process_l1_from_rollforward_schedule(ws_schedule, ws_new_lead, ws_new_k01, subject_config):
    """Populate L1 from an LAR/LRA roll-forward schedule."""
    copied = 0

    total_audit_col = get_schedule_total_col(ws_schedule)
    sections = [
        (["原值"], "无形资产", "1701"),
        (["累计摊销", "累计折旧"], "累计摊销", "1702"),
        (["减值准备"], "减值准备", "1703"),
    ]

    # L1.00 Lead sheet: PY values and GL account codes.
    lead_rows = {}
    for row in range(1, ws_new_lead.max_row + 1):
        label = normalize_text(ws_new_lead.cell(row=row, column=3).value)
        for section_names, lead_label, account_code in sections:
            if normalize_text(lead_label) in label:
                lead_rows[lead_label] = row

    if total_audit_col:
        for section_names, lead_label, account_code in sections:
            source_row = None
            for section_name in section_names:
                source_row = find_schedule_section_row(ws_schedule, section_name, "年末余额")
                if source_row:
                    break
            target_row = lead_rows.get(section_name)
            target_row = lead_rows.get(lead_label)
            if source_row and target_row:
                set_cell_value(ws_new_lead, target_row, 2, account_code)
                value = ws_schedule.cell(row=source_row, column=total_audit_col).value
                if value is not None:
                    set_cell_value(ws_new_lead, target_row, 10, value)
                    copied += 1

    # L1.01.1 Agree SL to GL: prior year-end rolls to current opening rows.
    schedule_groups = [g for g in find_schedule_group_columns(ws_schedule) if "合计" not in g["name_key"]]
    available_targets = [
        group for group in find_k01_category_groups(ws_new_k01, 10)
        if not is_total_category(group["name_key"])
    ]
    assigned_targets = []
    used_target_indexes = set()
    for schedule_group in schedule_groups:
        target = None
        source_key = l1_category_key(schedule_group["name"])
        for target_index, candidate in enumerate(available_targets):
            if target_index in used_target_indexes or is_placeholder_category(candidate["name_key"]):
                continue
            if category_matches(source_key, l1_category_key(candidate["name"])):
                target = candidate
                used_target_indexes.add(target_index)
                break
        if target is None:
            for target_index, candidate in enumerate(available_targets):
                if target_index not in used_target_indexes:
                    target = candidate
                    used_target_indexes.add(target_index)
                    break
        if target is not None:
            assigned_targets.append((schedule_group, target))

    section_rows_new = find_k01_section_rows(ws_new_k01)
    section_targets = [
        (["原值"], section_rows_new.get("cost", {}).get("opening")),
        (["累计摊销", "累计折旧"], section_rows_new.get("depreciation", {}).get("opening")),
        (["减值准备"], section_rows_new.get("impairment", {}).get("opening")),
    ]

    target_values = {}

    for schedule_group, target in assigned_targets:
        target_key = f"col_{target['adjust_col']}"
        if target_key not in target_values:
            target_values[target_key] = {
                "target": target,
                "source_names": [],
                "values": {},
            }
        target_values[target_key]["source_names"].append(schedule_group["name"])

        for section_names, target_row in section_targets:
            if not target_row:
                continue
            source_row = None
            for section_name in section_names:
                source_row = find_schedule_section_row(ws_schedule, section_name, "年末余额")
                if source_row:
                    break
            if not source_row:
                continue

            values = target_values[target_key]["values"]
            values.setdefault(target_row, {"book": 0, "adjust": 0, "book_seen": False, "adjust_seen": False})
            book_value = ws_schedule.cell(row=source_row, column=schedule_group["book_col"]).value
            adjust_value = ws_schedule.cell(row=source_row, column=schedule_group["adjust_col"]).value
            book_value = numeric_value(book_value)
            adjust_value = numeric_value(adjust_value)
            if book_value is not None:
                values[target_row]["book"] += book_value
                values[target_row]["book_seen"] = True
                copied += 1
            if adjust_value is not None:
                values[target_row]["adjust"] += adjust_value
                values[target_row]["adjust_seen"] = True
                copied += 1

    for item in target_values.values():
        target = item["target"]
        source_names = list(dict.fromkeys(item["source_names"]))
        ws_new_k01.cell(row=10, column=target["adjust_col"]).value = (
            source_names[0] if len(source_names) == 1 else "+".join(source_names)
        )
        for target_row, values in item["values"].items():
            if values["book_seen"]:
                ws_new_k01.cell(row=target_row, column=target["book_col"]).value = values["book"]
            if values["adjust_seen"]:
                ws_new_k01.cell(row=target_row, column=target["adjust_col"]).value = values["adjust"]

    return copied


def find_l2_business_end_col(ws, header_row):
    """Find the last real L2 BKD business column without touching trailing plug-in cells."""
    last_col = 0
    for col in range(1, min(ws.max_column, 60) + 1):
        if any(
            ws.cell(row=row, column=col).value not in (None, "")
            for row in range(max(1, header_row - 2), header_row + 1)
        ):
            last_col = col
    return last_col


def extend_detail_data_validations(ws, data_start_row, old_detail_end, new_detail_end, business_end_col):
    """Extend table data validations over every newly created business row."""
    for validation in getattr(ws.data_validations, "dataValidation", []):
        updated_ranges = []
        for cell_range in validation.sqref.ranges:
            updated = CellRange(str(cell_range))
            if (
                updated.min_col <= business_end_col
                and updated.max_col >= 1
                and updated.min_row <= old_detail_end
                and updated.max_row >= data_start_row
            ):
                updated.min_row = min(updated.min_row, data_start_row)
                updated.max_row = max(updated.max_row, new_detail_end)
            updated_ranges.append(updated)
        validation.sqref = MultiCellRange(updated_ranges)


def shift_unqualified_formula_rows(formula, insert_row, amount):
    """Shift same-sheet A1 row references while leaving sheet-qualified links unchanged."""
    if not isinstance(formula, str) or not formula.startswith("=") or amount <= 0:
        return formula

    pattern = re.compile(r"(?<![A-Za-z0-9_!'\"])(?P<col>\$?[A-Z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)")

    def replace(match):
        row = int(match.group("row"))
        if row < insert_row:
            return match.group(0)
        return f"{match.group('col')}{match.group('row_abs')}{row + amount}"

    return pattern.sub(replace, formula)


def process_l2_bkd(ws_prior_values, ws_new):
    """Roll L2.01.1 openings and extend every visible business section cleanly."""
    header_row = find_l2_bkd_header_row(ws_prior_values)
    new_header_row = find_l2_bkd_header_row(ws_new)
    if not header_row or not new_header_row:
        return 0

    prior_total_row = find_total_row_after(ws_prior_values, header_row)
    new_total_row = find_total_row_after(ws_new, new_header_row)
    if not prior_total_row or not new_total_row:
        return 0

    business_end_col = find_l2_business_end_col(ws_new, new_header_row)
    if business_end_col < 18:
        return 0

    original_total_row = new_total_row
    old_detail_end = original_total_row - 1
    total_formulas = {
        col: ws_new.cell(original_total_row, col).value
        for col in range(1, business_end_col + 1)
        if isinstance(ws_new.cell(original_total_row, col).value, str)
        and ws_new.cell(original_total_row, col).value.startswith("=")
    }

    def copy_business_row(source_row, target_row):
        for col in range(1, business_end_col + 1):
            copy_cell_shape(
                ws_new.cell(row=source_row, column=col),
                ws_new.cell(row=target_row, column=col),
                translate_formula=True,
            )
        if source_row in ws_new.row_dimensions:
            ws_new.row_dimensions[target_row].height = ws_new.row_dimensions[source_row].height

    records = []
    for row in range(header_row + 1, prior_total_row):
        if is_table_end_marker_row(ws_prior_values, row):
            continue
        project_code = ws_prior_values.cell(row=row, column=3).value
        project_name = ws_prior_values.cell(row=row, column=4).value
        if project_code in (None, "") and project_name in (None, ""):
            continue

        original_closing = ws_prior_values.cell(row=row, column=14).value
        if original_closing is None:
            original_closing = ws_prior_values.cell(row=row, column=12).value
        amort_closing = ws_prior_values.cell(row=row, column=18).value

        records.append({
            "code": project_code,
            "name": project_name,
            "initial_date": ws_prior_values.cell(row=row, column=5).value,
            "life": ws_prior_values.cell(row=row, column=6).value,
            "original_opening": original_closing,
            "amort_opening": amort_closing,
        })

    if not records:
        return 0

    data_start_row = new_header_row + 1
    available_rows = max(0, new_total_row - data_start_row)
    extra_rows = max(0, len(records) - available_rows)

    formula_source_row = None
    for row in range(new_total_row - 1, data_start_row - 1, -1):
        if any(
            isinstance(ws_new.cell(row=row, column=col).value, str)
            and ws_new.cell(row=row, column=col).value.startswith("=")
            for col in (12, 14, 18)
        ):
            formula_source_row = row
            break
    formula_source_row = formula_source_row or data_start_row

    if extra_rows:
        insert_rows_preserving_sheet_metadata(ws_new, new_total_row, extra_rows)
        for offset in range(extra_rows):
            copy_business_row(formula_source_row, new_total_row + offset)
        new_total_row += extra_rows
        for row in range(new_total_row + 1, ws_new.max_row + 1):
            for col in range(1, business_end_col + 1):
                formula = ws_new.cell(row, col).value
                updated_formula = shift_unqualified_formula_rows(
                    formula,
                    original_total_row,
                    extra_rows,
                )
                if updated_formula != formula:
                    set_cell_value(ws_new, row, col, updated_formula)

    copied = 0
    for idx, record in enumerate(records):
        target_row = data_start_row + idx
        if target_row != formula_source_row:
            copy_business_row(formula_source_row, target_row)

        values = {
            3: record["code"],
            4: record["name"],
            5: record["initial_date"],
            6: record["life"],
            7: record["original_opening"],
            8: None,
            9: None,
            10: None,
            11: None,
            13: None,
            15: record["amort_opening"],
            16: None,
            17: None,
        }
        for col, value in values.items():
            if set_cell_value(ws_new, target_row, col, value):
                copied += 1

        set_cell_value(ws_new, target_row, 12, f"=G{target_row}+H{target_row}-I{target_row}")
        set_cell_value(ws_new, target_row, 14, f"=L{target_row}+M{target_row}")
        set_cell_value(ws_new, target_row, 18, f"=O{target_row}+P{target_row}-Q{target_row}")
        ws_new.cell(row=target_row, column=5).number_format = "yyyy/mm/dd"

    for row in range(data_start_row + len(records), new_total_row):
        for col in (3, 4, 5, 6, 7, 8, 9, 10, 11, 13, 15, 16, 17):
            set_cell_value(ws_new, row, col, None)
        if row < new_total_row:
            set_cell_value(ws_new, row, 12, f"=G{row}+H{row}-I{row}")
            set_cell_value(ws_new, row, 14, f"=L{row}+M{row}")
            set_cell_value(ws_new, row, 18, f"=O{row}+P{row}-Q{row}")

    last_sum_row = data_start_row + len(records) - 1
    for col, formula in total_formulas.items():
        col_letter = get_column_letter(col)
        if formula.upper().startswith("=SUM("):
            updated_formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_sum_row})"
        else:
            try:
                updated_formula = Translator(
                    formula,
                    origin=f"{col_letter}{original_total_row}",
                ).translate_formula(f"{col_letter}{new_total_row}")
            except Exception:
                updated_formula = formula
        set_cell_value(ws_new, new_total_row, col, updated_formula)

    extend_detail_data_validations(
        ws_new,
        data_start_row,
        old_detail_end,
        last_sum_row,
        business_end_col,
    )

    return copied


def find_first_table_end_marker(ws, start_row, end_row):
    """Find the first /T marker row in a bounded table area."""
    for row in range(start_row, end_row):
        if is_table_end_marker_row(ws, row):
            return row
    return None


def find_expense_bkd_header_row(ws):
    """Find the Uexp BKD detail header row from stable header labels."""
    for row in range(1, min(ws.max_row, 100) + 1):
        row_text = " ".join(
            str(ws.cell(row=row, column=col).value or "")
            for col in range(1, min(ws.max_column, 30) + 1)
        )
        key = normalize_text(row_text)
        if "科目编码" in key and "科目名称" in key and ("上期末审定数" in key or "上期审定数" in key):
            return row
    return None


def find_expense_bkd_total_row(ws, header_row):
    """Find a Uexp BKD total row by label or by SUM formulas."""
    total_row = find_total_row_after(ws, header_row)
    if total_row:
        return total_row

    for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
        sum_formula_count = 0
        for col in range(1, min(ws.max_column, 25) + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.startswith("=SUM("):
                sum_formula_count += 1
        if sum_formula_count >= 3:
            return row

    for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
        if row_contains_any(ws, row, ["Rx", "A3", "Diff", "波动说明"], columns=range(1, min(18, ws.max_column) + 1)):
            candidate = row - 1
            while candidate > header_row and not row_has_content(ws, candidate, 1, min(18, ws.max_column)):
                candidate -= 1
            return candidate if candidate > header_row else None
    return None


def set_expense_bkd_row_formulas(ws, row, header_cols, total_row=None, data_last_row=None):
    """Refresh row formulas in an expense BKD table using discovered columns."""
    book_col = header_cols.get("book")
    book_adjust_col = header_cols.get("book_adjust")
    unaudited_col = header_cols.get("unaudited")
    structure_col = header_cols.get("structure")
    audit_adjust_col = header_cols.get("audit_adjust")
    audited_col = header_cols.get("audited")
    py_col = header_cols.get("py")
    variance_col = header_cols.get("variance")
    rate_col = header_cols.get("rate")

    if total_row and data_last_row:
        for key in ("book", "book_adjust", "unaudited", "structure", "audit_adjust", "audited", "py"):
            col = header_cols.get(key)
            if col:
                col_letter = get_column_letter(col)
                set_cell_value(ws, total_row, col, f"=SUM({col_letter}{header_cols['data_start_row']}:{col_letter}{data_last_row})")
        if variance_col and audited_col and py_col:
            set_cell_value(ws, total_row, variance_col, f"={get_column_letter(audited_col)}{total_row}-{get_column_letter(py_col)}{total_row}")
        if rate_col and py_col and variance_col:
            set_cell_value(
                ws,
                total_row,
                rate_col,
                f"=IF({get_column_letter(py_col)}{total_row}<>0,{get_column_letter(variance_col)}{total_row}/{get_column_letter(py_col)}{total_row},1)"
            )
        return

    if unaudited_col and book_col and book_adjust_col:
        set_cell_value(ws, row, unaudited_col, f"={get_column_letter(book_col)}{row}+{get_column_letter(book_adjust_col)}{row}")
    if structure_col and unaudited_col and header_cols.get("total_row"):
        total_row = header_cols["total_row"]
        unaudited_letter = get_column_letter(unaudited_col)
        set_cell_value(ws, row, structure_col, f'=IF(${unaudited_letter}${total_row}<>0,{unaudited_letter}{row}/${unaudited_letter}${total_row},"")')
    if audited_col and unaudited_col and audit_adjust_col:
        set_cell_value(ws, row, audited_col, f"={get_column_letter(unaudited_col)}{row}+{get_column_letter(audit_adjust_col)}{row}")
    if variance_col and audited_col and py_col:
        set_cell_value(ws, row, variance_col, f"={get_column_letter(audited_col)}{row}-{get_column_letter(py_col)}{row}")
    if rate_col and py_col and variance_col:
        set_cell_value(
            ws,
            row,
            rate_col,
            f"=IF({get_column_letter(py_col)}{row}<>0,{get_column_letter(variance_col)}{row}/{get_column_letter(py_col)}{row},1)"
        )


def process_expense_bkd_prior_current_to_py(ws_prior_values, ws_new):
    """Roll prior Uexp BKD current audited values into the current PY column."""
    prior_header_row = find_expense_bkd_header_row(ws_prior_values)
    new_header_row = find_expense_bkd_header_row(ws_new)
    if not prior_header_row or not new_header_row:
        return 0, None

    source_current_col = (
        find_header_col(ws_prior_values, prior_header_row, ["本期账面审定数", "本期审定数", "本期数"])
        or find_header_col_near(ws_prior_values, prior_header_row, ["本期账面审定数", "本期审定数"], row_offsets=(-1, 0))
    )
    target_py_col = (
        find_header_col(ws_new, new_header_row, ["上期末审定数", "上期审定数", "上年数", "PY"])
        or find_header_col_near(ws_new, new_header_row, ["上期末审定数", "上期审定数", "上年数", "PY"], row_offsets=(-1, 0))
    )
    if not source_current_col or not target_py_col:
        return 0, None

    descriptor_rules = [
        (["账套名称/账套编码", "账套名称", "账套编码"], ["账套名称/账套编码", "账套名称", "账套编码"]),
        (["科目编码"], ["科目编码"]),
        (["科目名称"], ["科目名称"]),
    ]
    descriptor_cols = []
    for source_keywords, target_keywords in descriptor_rules:
        source_col = find_header_col(ws_prior_values, prior_header_row, source_keywords)
        target_col = find_header_col(ws_new, new_header_row, target_keywords)
        if source_col and target_col:
            descriptor_cols.append((source_col, target_col))

    if not descriptor_cols:
        return 0, None

    prior_total_row = find_expense_bkd_total_row(ws_prior_values, prior_header_row)
    new_total_row = find_expense_bkd_total_row(ws_new, new_header_row)
    if not prior_total_row or not new_total_row:
        return 0, None

    records = []
    for row in range(prior_header_row + 1, prior_total_row):
        if is_table_end_marker_row(ws_prior_values, row):
            continue
        descriptors = {
            target_col: ws_prior_values.cell(row=row, column=source_col).value
            for source_col, target_col in descriptor_cols
        }
        if not any(value not in (None, "") for value in descriptors.values()):
            continue
        records.append((descriptors, ws_prior_values.cell(row=row, column=source_current_col).value))

    if not records:
        return 0, new_total_row

    data_start_row = new_header_row + 1
    marker_row = find_first_table_end_marker(ws_new, data_start_row, new_total_row) or new_total_row
    available_rows = max(0, marker_row - data_start_row)
    extra_rows = max(0, len(records) - available_rows)
    formula_source_row = max(data_start_row, marker_row - 1)
    old_total_row = new_total_row

    if extra_rows:
        insert_rows_preserving_sheet_metadata(ws_new, marker_row, extra_rows)
        for offset in range(extra_rows):
            copy_row_shape(ws_new, formula_source_row, marker_row + offset, translate_formula=True)
        marker_row += extra_rows
        new_total_row += extra_rows

    header_cols = {
        "header_row": new_header_row,
        "data_start_row": data_start_row,
        "total_row": new_total_row,
        "book": find_header_col(ws_new, new_header_row, ["本期账面数"]),
        "book_adjust": find_header_col(ws_new, new_header_row, ["本期账表调整数", "账表调整数"]),
        "unaudited": find_header_col(ws_new, new_header_row, ["本期账面未审数", "本期未审数"]),
        "structure": find_header_col(ws_new, new_header_row, ["结构比"]),
        "audit_adjust": find_header_col(ws_new, new_header_row, ["本期审计调整金数", "本期审计调整金额", "审计调整数"]),
        "audited": find_header_col(ws_new, new_header_row, ["本期账面审定数", "本期审定数"]),
        "py": target_py_col,
        "variance": find_header_col(ws_new, new_header_row, ["变动额", "变动金额"]),
        "rate": find_header_col(ws_new, new_header_row, ["变动率", "变动%"]),
    }

    copied = 0
    target_descriptor_cols = [target_col for _, target_col in descriptor_cols]
    for idx, (descriptors, current_value) in enumerate(records):
        target_row = data_start_row + idx
        copy_row_shape(ws_new, formula_source_row, target_row, translate_formula=True)
        for target_col, value in descriptors.items():
            if set_cell_value(ws_new, target_row, target_col, value):
                copied += 1
        if set_cell_value(ws_new, target_row, target_py_col, current_value):
            copied += 1
        for key in ("book", "book_adjust", "audit_adjust"):
            col = header_cols.get(key)
            if col:
                set_cell_value(ws_new, target_row, col, None)
        set_expense_bkd_row_formulas(ws_new, target_row, header_cols)

    for row in range(data_start_row + len(records), marker_row):
        for col in set(target_descriptor_cols + [target_py_col]):
            set_cell_value(ws_new, row, col, None)
        for key in ("book", "book_adjust", "audit_adjust"):
            col = header_cols.get(key)
            if col:
                set_cell_value(ws_new, row, col, None)
        set_expense_bkd_row_formulas(ws_new, row, header_cols)

    set_expense_bkd_row_formulas(
        ws_new,
        new_total_row,
        header_cols,
        total_row=new_total_row,
        data_last_row=data_start_row + len(records) - 1,
    )

    if extra_rows:
        shift_local_formula_refs_after_insert(ws_new, old_total_row, extra_rows)

    return copied, new_total_row


def update_lead_bkd_total_references(ws_lead, bkd_sheet_name, total_row):
    """Point Uexp lead formulas at the actual BKD total row after dynamic insertion."""
    if not total_row:
        return 0

    updated = 0
    sheet_ref_pattern = re.compile(r"('?" + re.escape(bkd_sheet_name) + r"'?!\$?[A-Z]{1,3})\$?\d+")
    for row in range(1, min(ws_lead.max_row, 80) + 1):
        row_refs_sheet = any(
            ws_lead.cell(row=row, column=col).value == bkd_sheet_name
            for col in (3, 4)
        )
        if not row_refs_sheet:
            continue
        for col in range(1, min(ws_lead.max_column, 20) + 1):
            value = ws_lead.cell(row=row, column=col).value
            if not isinstance(value, str) or bkd_sheet_name not in value:
                continue
            new_value = sheet_ref_pattern.sub(lambda match: f"{match.group(1)}{total_row}", value)
            if new_value != value:
                ws_lead.cell(row=row, column=col).value = new_value
                updated += 1
        break
    return updated


def find_q1_bkd_header_row(ws):
    """Find Q1.01 movement table header row."""
    for row in range(1, min(ws.max_row, 80) + 1):
        row_text = " ".join(
            str(ws.cell(row=row, column=col).value or "")
            for col in range(1, min(ws.max_column, 25) + 1)
        )
        key = normalize_text(row_text)
        if "债务描述" in key and "期初余额" in key and "期末余额" in key:
            return row
    return None


def is_q1_interest_row(ws, row, desc_col):
    """Return True when a Q1.01 row represents accrued interest."""
    return "利息" in normalize_text(ws.cell(row=row, column=desc_col).value)


def set_q1_bkd_row_formulas(ws, row, cols, interest=False):
    """Refresh Q1.01 movement formulas for a detail row."""
    opening = cols.get("opening")
    add = cols.get("add")
    repayment = cols.get("repayment")
    paid_interest = cols.get("paid_interest")
    accrued_interest = cols.get("accrued_interest")
    reclass = cols.get("reclass")
    fx = cols.get("fx")
    ending = cols.get("ending")
    tb = cols.get("tb")
    diff = cols.get("diff")
    current = cols.get("current")
    noncurrent = cols.get("noncurrent")

    if ending:
        if interest and opening and accrued_interest and paid_interest and reclass and fx:
            set_cell_value(ws, row, ending, f"={get_column_letter(opening)}{row}+{get_column_letter(accrued_interest)}{row}-{get_column_letter(paid_interest)}{row}+{get_column_letter(reclass)}{row}+{get_column_letter(fx)}{row}")
        elif opening and add and repayment and reclass and fx:
            set_cell_value(ws, row, ending, f"={get_column_letter(opening)}{row}+{get_column_letter(add)}{row}-{get_column_letter(repayment)}{row}+{get_column_letter(reclass)}{row}+{get_column_letter(fx)}{row}")
    if diff and ending and tb:
        set_cell_value(ws, row, diff, f"={get_column_letter(ending)}{row}-{get_column_letter(tb)}{row}")
    if noncurrent and ending and current:
        set_cell_value(ws, row, noncurrent, f"={get_column_letter(ending)}{row}-{get_column_letter(current)}{row}")


def process_q1_bkd(ws_prior_values, ws_new):
    """Roll prior Q1.01 ending balances to current opening balances."""
    prior_header_row = find_q1_bkd_header_row(ws_prior_values)
    new_header_row = find_q1_bkd_header_row(ws_new)
    if not prior_header_row or not new_header_row:
        return 0, None

    prior_total_row = find_total_row_after(ws_prior_values, prior_header_row)
    new_total_row = find_total_row_after(ws_new, new_header_row)
    if not prior_total_row or not new_total_row:
        return 0, None

    prior_desc_col = find_header_col(ws_prior_values, prior_header_row, ["债务描述/债务工具", "债务描述"])
    prior_account_col = find_header_col(ws_prior_values, prior_header_row, ["总账账户"])
    prior_ending_col = find_header_col(ws_prior_values, prior_header_row, ["期末余额"])
    prior_current_col = find_header_col(ws_prior_values, prior_header_row, ["债务工具的流动部分"])
    prior_covenant_col = find_header_col(ws_prior_values, prior_header_row, ["契约条件"])

    new_desc_col = find_header_col(ws_new, new_header_row, ["债务描述/债务工具", "债务描述"])
    new_account_col = find_header_col(ws_new, new_header_row, ["总账账户"])
    new_opening_col = find_header_col(ws_new, new_header_row, ["期初余额"])
    new_current_col = find_header_col(ws_new, new_header_row, ["债务工具的流动部分"])
    new_covenant_col = find_header_col(ws_new, new_header_row, ["契约条件"])

    new_current_col = None
    new_covenant_col = None

    if not all([prior_desc_col, prior_ending_col, new_desc_col, new_opening_col]):
        return 0, None

    records = []
    for row in range(prior_header_row + 1, prior_total_row):
        desc = ws_prior_values.cell(row=row, column=prior_desc_col).value
        account = ws_prior_values.cell(row=row, column=prior_account_col).value if prior_account_col else None
        ending = ws_prior_values.cell(row=row, column=prior_ending_col).value
        if desc in (None, "") and account in (None, ""):
            continue
        records.append({
            "desc": desc,
            "account": account,
            "opening": ending,
            "current": ws_prior_values.cell(row=row, column=prior_current_col).value if prior_current_col else None,
            "covenant": ws_prior_values.cell(row=row, column=prior_covenant_col).value if prior_covenant_col else None,
            "interest": "利息" in normalize_text(desc),
        })

    if not records:
        return 0, new_total_row

    data_start_row = new_header_row + 1
    available_rows = max(0, new_total_row - data_start_row)
    extra_rows = max(0, len(records) - available_rows)
    principal_shape_row = data_start_row
    interest_shape_row = next(
        (row for row in range(data_start_row, new_total_row) if is_q1_interest_row(ws_new, row, new_desc_col)),
        data_start_row,
    )
    old_total_row = new_total_row

    if extra_rows:
        insert_rows_preserving_sheet_metadata(ws_new, new_total_row, extra_rows)
        for offset in range(extra_rows):
            record_idx = available_rows + offset
            source_row = interest_shape_row if records[record_idx]["interest"] else principal_shape_row
            copy_row_shape(ws_new, source_row, old_total_row + offset, translate_formula=True)
        new_total_row += extra_rows
        shift_local_formula_refs_after_insert(ws_new, old_total_row, extra_rows)

    cols = {
        "opening": new_opening_col,
        "add": find_header_col(ws_new, new_header_row, ["本金增加/支取"]),
        "repayment": find_header_col(ws_new, new_header_row, ["本金减免/偿还"]),
        "paid_interest": find_header_col(ws_new, new_header_row, ["偿还的利息费用"]),
        "accrued_interest": find_header_col(ws_new, new_header_row, ["计提的利息费用"]),
        "reclass": find_header_col(ws_new, new_header_row, ["重分类"]),
        "fx": find_header_col(ws_new, new_header_row, ["外币重新计量"]),
        "ending": find_header_col(ws_new, new_header_row, ["期末余额"]),
        "tb": find_header_col(ws_new, new_header_row, ["试算表余额"]),
        "diff": find_header_col(ws_new, new_header_row, ["差额"]),
        "current": None,
        "noncurrent": find_header_col(ws_new, new_header_row, ["非流动部分"]),
    }

    copied = 0
    for idx, record in enumerate(records):
        target_row = data_start_row + idx
        source_row = interest_shape_row if record["interest"] else principal_shape_row
        if target_row != source_row:
            copy_row_shape(ws_new, source_row, target_row, translate_formula=True)
        for col, value in (
            (new_desc_col, record["desc"]),
            (new_account_col, record["account"] if new_account_col else None),
            (new_opening_col, record["opening"]),
            (new_current_col, record["current"] if new_current_col else None),
            (new_covenant_col, record["covenant"] if new_covenant_col else None),
        ):
            if col and set_cell_value(ws_new, target_row, col, value):
                copied += 1
        for key in ("add", "repayment", "paid_interest", "accrued_interest", "reclass", "fx", "tb"):
            col = cols.get(key)
            if col:
                set_cell_value(ws_new, target_row, col, None)
        set_q1_bkd_row_formulas(ws_new, target_row, cols, interest=record["interest"])

    for row in range(data_start_row + len(records), new_total_row):
        for col in range(2, min(ws_new.max_column, 18) + 1):
            set_cell_value(ws_new, row, col, None)

    data_last_row = data_start_row + len(records) - 1
    for key in ("opening", "add", "repayment", "paid_interest", "accrued_interest", "reclass", "fx", "ending", "tb", "diff", "current"):
        col = cols.get(key)
        if col:
            col_letter = get_column_letter(col)
            set_cell_value(ws_new, new_total_row, col, f"=SUM({col_letter}{data_start_row}:{col_letter}{data_last_row})")
    if cols.get("noncurrent") and cols.get("ending") and cols.get("current"):
        set_cell_value(ws_new, new_total_row, cols["noncurrent"], f"={get_column_letter(cols['ending'])}{new_total_row}-{get_column_letter(cols['current'])}{new_total_row}")
    if new_covenant_col:
        col_letter = get_column_letter(new_covenant_col)
        set_cell_value(ws_new, new_total_row, new_covenant_col, f'=COUNTIF({col_letter}{data_start_row}:{col_letter}{data_last_row},"是")')

    interest_row = next(
        (data_start_row + idx for idx, record in enumerate(records) if record["interest"]),
        None,
    )
    if interest_row and cols.get("paid_interest") and cols.get("opening") and cols.get("ending") and cols.get("accrued_interest"):
        check_row = find_row_containing(ws_new, "利息计提合理性检查", (new_total_row + 1, min(ws_new.max_row, new_total_row + 10)))
        if check_row:
            set_cell_value(
                ws_new,
                check_row,
                3,
                f"={get_column_letter(cols['paid_interest'])}{new_total_row}-{get_column_letter(cols['opening'])}{interest_row}+{get_column_letter(cols['ending'])}{interest_row}-{get_column_letter(cols['accrued_interest'])}{new_total_row}",
            )

    return copied, new_total_row


def copy_changed_constant_cell(ws_prior, ws_new, source_row, source_col, target_row=None, target_col=None):
    """Copy a non-formula constant when it adds information to the target."""
    target_row = target_row or source_row
    target_col = target_col or source_col
    value = ws_prior.cell(row=source_row, column=source_col).value
    if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
        return 0

    target_cell = ws_new.cell(row=target_row, column=target_col)
    if target_cell.value == value:
        return 0
    if set_cell_value(ws_new, target_row, target_col, value):
        highlight_wording_cell(ws_new, target_row, target_col)
        return 1
    return 0


def copy_q1_covenant_rows_between_markers(ws_prior, ws_new, start_keyword, end_keyword=None, columns=range(2, 10)):
    """Copy changed constants in a covenant section located by row markers."""
    source_start = find_row_containing(ws_prior, start_keyword, (1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, start_keyword, (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    if end_keyword:
        source_end_marker = find_row_containing(ws_prior, end_keyword, (source_start + 1, ws_prior.max_row))
        target_end_marker = find_row_containing(ws_new, end_keyword, (target_start + 1, ws_new.max_row))
        source_end = (source_end_marker - 1) if source_end_marker else min(source_start + 20, ws_prior.max_row)
        target_end = (target_end_marker - 1) if target_end_marker else min(target_start + 20, ws_new.max_row)
    else:
        source_end = ws_prior.max_row
        target_end = ws_new.max_row

    source_len = max(0, source_end - source_start)
    target_len = max(0, target_end - target_start)
    rows_to_insert = max(0, source_len - target_len)
    if rows_to_insert and end_keyword:
        insert_at = target_end + 1
        source_shape_row = max(target_start + 1, target_end)
        insert_rows_preserving_sheet_metadata(ws_new, insert_at, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, source_shape_row, insert_at + offset, translate_formula=True)

    copied = 0
    for offset in range(1, source_len + 1):
        source_row = source_start + offset
        target_row = target_start + offset
        for col in columns:
            copied += copy_changed_constant_cell(ws_prior, ws_new, source_row, col, target_row, col)
    return copied


def copy_q1_covenant_labeled_values(ws_prior, ws_new):
    """Copy numeric/input values beside stable covenant labels."""
    copied = 0
    labels = (
        "具有限制性契约的借款数量",
        "减：本金已分类为流动负债的借款数量",
    )
    for label in labels:
        source_row = find_row_containing(ws_prior, label, (1, ws_prior.max_row))
        target_row = find_row_containing(ws_new, label, (1, ws_new.max_row))
        if not source_row or not target_row:
            continue
        for col in range(3, min(10, ws_prior.max_column, ws_new.max_column) + 1):
            copied += copy_changed_constant_cell(ws_prior, ws_new, source_row, col, target_row, col)
    return copied


def find_q1_covenant_note_starts(ws):
    """Find project covenant evidence blocks headed by an actual numbered note."""
    starts = []
    for row in range(1, ws.max_row + 1):
        value = normalize_text(ws.cell(row=row, column=2).value)
        if value.startswith("附注") and "针对编号" in value:
            starts.append(row)
    return starts


def duplicate_q1_template_note_block(ws, source_start, source_end, insert_at):
    """Duplicate one current-template covenant note block with its merges and dimensions."""
    block_length = source_end - source_start + 1
    source_merges = [
        deepcopy(merged_range)
        for merged_range in ws.merged_cells.ranges
        if merged_range.min_row >= source_start and merged_range.max_row <= source_end
    ]
    insert_rows_preserving_sheet_metadata(ws, insert_at, block_length)
    for offset in range(block_length):
        source_row = source_start + offset
        target_row = insert_at + offset
        copy_row_shape(ws, source_row, target_row, translate_formula=True)
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    row_shift = insert_at - source_start
    for merged_range in source_merges:
        ws.merge_cells(
            start_row=merged_range.min_row + row_shift,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + row_shift,
            end_column=merged_range.max_col,
        )
    return insert_at, insert_at + block_length - 1


def copy_q1_covenant_note_values(ws_prior, ws_new, source_start, source_end, target_start, target_end):
    """Map prior covenant note constants into the matching current-template labels."""
    labels = ("合同编号", "借款金额", "借款银行", "Step1", "条件1", "测试过程", "测试结论", "结论")

    def label_rows(ws, start, end):
        result = {}
        for row in range(start, end + 1):
            key = normalize_text(ws.cell(row=row, column=2).value)
            for label in labels:
                if key == normalize_text(label):
                    result[label] = row
                    break
        return result

    source_rows = label_rows(ws_prior, source_start, source_end)
    target_rows = label_rows(ws_new, target_start, target_end)
    copied = 0

    heading = ws_prior.cell(row=source_start, column=2).value
    if heading not in (None, "") and set_cell_value(ws_new, target_start, 2, heading):
        highlight_wording_cell(ws_new, target_start, 2)
        copied += 1

    ordered_source_rows = sorted(source_rows.values())
    for label in labels:
        source_row = source_rows.get(label)
        target_row = target_rows.get(label)
        if not source_row or not target_row:
            continue
        next_source_row = next((row for row in ordered_source_rows if row > source_row), source_end + 1)
        source_segment_end = next_source_row - 1

        for col in range(3, min(9, ws_prior.max_column, ws_new.max_column) + 1):
            values = []
            for row in range(source_row, source_segment_end + 1):
                value = ws_prior.cell(row=row, column=col).value
                if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
                    continue
                values.append(value)
            if not values:
                continue
            value = values[0] if len(values) == 1 else "\n".join(str(item) for item in values)
            if normalize_text(ws_new.cell(row=target_row, column=col).value) == normalize_text(value):
                continue
            if set_cell_value(ws_new, target_row, col, value):
                highlight_wording_cell(ws_new, target_row, col)
                cell = ws_new.cell(row=target_row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                copied += 1
    return copied


def process_q1_covenant_sheet(ws_prior, ws_new):
    """Roll Q1.05 values while keeping the current template's layout authoritative."""
    copied = 0
    copied += copy_q1_covenant_labeled_values(ws_prior, ws_new)

    source_header = find_row_containing(ws_prior, "编号", (1, ws_prior.max_row))
    target_header = find_row_containing(ws_new, "编号", (1, ws_new.max_row))
    source_legend = find_row_containing(ws_prior, "标记图例", ((source_header or 1) + 1, ws_prior.max_row))
    target_legend = find_row_containing(ws_new, "标记图例", ((target_header or 1) + 1, ws_new.max_row))
    if source_header and target_header and source_legend and target_legend:
        source_records = []
        for row in range(source_header + 1, source_legend):
            values = [ws_prior.cell(row=row, column=col).value for col in range(2, 10)]
            if any(value not in (None, "") for value in values):
                source_records.append(values)
        capacity = max(0, target_legend - target_header - 1)
        if len(source_records) > capacity:
            extra_rows = len(source_records) - capacity
            shape_row = max(target_header + 1, target_legend - 1)
            insert_rows_preserving_sheet_metadata(ws_new, target_legend, extra_rows)
            for offset in range(extra_rows):
                copy_row_shape(
                    ws_new,
                    shape_row,
                    target_legend + offset,
                    translate_formula=True,
                )
                ws_new.row_dimensions[target_legend + offset].height = (
                    ws_new.row_dimensions[shape_row].height
                )
            target_legend += extra_rows
            capacity += extra_rows
        for offset, values in enumerate(source_records[:capacity]):
            target_row = target_header + 1 + offset
            for col, value in enumerate(values, start=2):
                if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
                    continue
                if set_cell_value(ws_new, target_row, col, value):
                    highlight_wording_cell(ws_new, target_row, col)
                    copied += 1

    source_note_starts = find_q1_covenant_note_starts(ws_prior)
    target_note_starts = find_q1_covenant_note_starts(ws_new)
    final_marker = find_row_containing(ws_new, "针对已经触发限制性条款", (1, ws_new.max_row))
    if source_note_starts and target_note_starts and final_marker:
        template_block_start = target_note_starts[0]
        template_block_end = final_marker - 1
        while template_block_end > template_block_start and not row_has_content(ws_new, template_block_end, 2, 9):
            template_block_end -= 1
        template_block_end += 1

        target_blocks = [(template_block_start, template_block_end)]
        insert_at = final_marker
        for _ in range(1, len(source_note_starts)):
            block = duplicate_q1_template_note_block(
                ws_new,
                template_block_start,
                template_block_end,
                insert_at,
            )
            target_blocks.append(block)
            insert_at = block[1] + 1

        source_final_marker = find_row_containing(ws_prior, "针对已经触发限制性条款", (1, ws_prior.max_row))
        for index, source_start in enumerate(source_note_starts):
            source_end = (
                source_note_starts[index + 1] - 1
                if index + 1 < len(source_note_starts)
                else (source_final_marker - 1 if source_final_marker else ws_prior.max_row)
            )
            target_start, target_end = target_blocks[index]
            copied += copy_q1_covenant_note_values(
                ws_prior,
                ws_new,
                source_start,
                source_end,
                target_start,
                target_end,
            )

    return copied


def process_vcvd_cutoff_table2(ws_prior, ws_new):
    """Roll VC&VD.01.4 cutoff test table 2 strategy text and highlight copied fields."""
    copied = 0
    source_start = find_row_containing(ws_prior, "表2", (1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, "表2", (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    source_end_marker = find_row_containing(ws_prior, "表3", (source_start + 1, ws_prior.max_row))
    target_end_marker = find_row_containing(ws_new, "表3", (target_start + 1, ws_new.max_row))
    source_end = (source_end_marker - 1) if source_end_marker else min(source_start + 20, ws_prior.max_row)
    target_end = (target_end_marker - 1) if target_end_marker else min(target_start + 20, ws_new.max_row)

    source_len = max(0, source_end - source_start)
    target_len = max(0, target_end - target_start)
    rows_to_insert = max(0, source_len - target_len)
    if rows_to_insert and target_end_marker:
        insert_at = target_end_marker
        source_shape_row = max(target_start + 1, target_end)
        insert_rows_preserving_sheet_metadata(ws_new, insert_at, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, source_shape_row, insert_at + offset, translate_formula=True)

    max_col = min(ws_prior.max_column, ws_new.max_column, 12)
    for offset in range(1, source_len + 1):
        source_row = source_start + offset
        target_row = target_start + offset
        for col in range(1, max_col + 1):
            copied += copy_changed_constant_cell(ws_prior, ws_new, source_row, col, target_row, col)

    return copied


def find_row_containing_any(ws, keywords, search_range=None):
    """Find the first row containing any of the provided keywords."""
    start, end = search_range or (1, ws.max_row)
    for row in range(start, min(end, ws.max_row) + 1):
        if row_has_any_keyword(ws, row, keywords, 1, min(ws.max_column, 30)):
            return row
    return None


def find_next_vcvd_bkd_section_row(ws, start_row, markers):
    """Find the next BKD wording section marker after a source/target anchor."""
    for row in range(start_row + 1, min(ws.max_row, start_row + 120) + 1):
        if row_has_any_keyword(ws, row, markers, 1, min(ws.max_column, 30)):
            return row
    return None


def copy_vcvd_bkd_section(ws_prior, ws_new, anchor_keywords, next_markers, insert_before_keywords=None, columns=range(2, 10)):
    """Copy one VC/VD BKD wording section by anchors without touching table 1 judgment columns."""
    source_start = find_row_containing_any(ws_prior, anchor_keywords)
    if not source_start:
        return 0

    source_next = find_next_vcvd_bkd_section_row(ws_prior, source_start, next_markers)
    if source_next:
        source_end = source_next - 1
    else:
        source_end = min(ws_prior.max_row, source_start + 20)
        while source_end > source_start and not row_has_content(ws_prior, source_end, 1, min(ws_prior.max_column, 30)):
            source_end -= 1

    if source_end < source_start:
        return 0

    target_start = find_row_containing_any(ws_new, anchor_keywords)
    if target_start:
        target_next = find_next_vcvd_bkd_section_row(ws_new, target_start, next_markers)
        target_end = (target_next - 1) if target_next else min(ws_new.max_row, target_start + 20)
    else:
        insert_before = find_row_containing_any(ws_new, insert_before_keywords or next_markers)
        if not insert_before:
            insert_before = ws_new.max_row + 1
        source_len = source_end - source_start + 1
        insert_rows_preserving_sheet_metadata(ws_new, insert_before, source_len)
        source_shape_row = max(1, insert_before - 1)
        for offset in range(source_len):
            copy_row_shape(ws_new, source_shape_row, insert_before + offset, translate_formula=True)
        target_start = insert_before
        target_end = insert_before + source_len - 1

    source_len = source_end - source_start + 1
    target_len = max(1, target_end - target_start + 1)
    rows_to_insert = max(0, source_len - target_len)
    if rows_to_insert:
        insert_at = target_end + 1
        source_shape_row = max(target_start, target_end)
        insert_rows_preserving_sheet_metadata(ws_new, insert_at, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, source_shape_row, insert_at + offset, translate_formula=True)

    copied = 0
    for offset in range(source_len):
        source_row = source_start + offset
        target_row = target_start + offset
        for col in columns:
            copied += copy_changed_constant_cell(ws_prior, ws_new, source_row, col, target_row, col)
    return copied


def process_vcvd_bkd_wording(ws_prior, ws_new):
    """Roll VC/VD BKD expected fluctuation, ARP wording, and adjustment summary."""
    copied = 0
    section_markers = (
        "波动范围",
        "表1",
        "波动说明",
        "预期波动分析",
        "ARP波动说明",
        "ARP",
        "调整汇总表",
    )

    copied += copy_vcvd_bkd_section(
        ws_prior,
        ws_new,
        ("在下文中描述我们当期对账户波动的预期", "预期波动分析"),
        ("波动范围", "表1"),
        columns=range(2, 10),
    )
    copied += copy_vcvd_bkd_section(
        ws_prior,
        ws_new,
        ("波动说明",),
        ("ARP波动说明", "ARP", "调整汇总表"),
        insert_before_keywords=("调整汇总表",),
        columns=range(2, 10),
    )
    copied += copy_vcvd_bkd_section(
        ws_prior,
        ws_new,
        ("ARP波动说明", "ARP"),
        ("调整汇总表",),
        insert_before_keywords=("调整汇总表",),
        columns=range(2, 10),
    )
    copied += copy_vcvd_bkd_section(
        ws_prior,
        ws_new,
        ("调整汇总表",),
        section_markers,
        insert_before_keywords=None,
        columns=range(2, 18),
    )

    return copied


def copy_section_with_shape(ws_prior, ws_new, source_start, source_end, target_start, col_start=1, col_end=None, highlight=True):
    """Copy a small worksheet section with styles so bordered wording boxes stay intact."""
    col_end = col_end or min(ws_prior.max_column, ws_new.max_column)
    source_len = source_end - source_start + 1
    target_end = target_start + source_len - 1
    if target_end > ws_new.max_row:
        insert_rows_preserving_sheet_metadata(ws_new, ws_new.max_row + 1, target_end - ws_new.max_row)

    unmerge_ranges_intersecting(ws_new, target_start, target_end, col_start, col_end)

    copied = 0
    for offset in range(source_len):
        source_row = source_start + offset
        target_row = target_start + offset
        try:
            ws_new.row_dimensions[target_row].height = ws_prior.row_dimensions[source_row].height
        except Exception:
            pass
        for col in range(col_start, min(col_end, ws_prior.max_column, ws_new.max_column) + 1):
            source_cell = ws_prior.cell(row=source_row, column=col)
            target_cell = ws_new.cell(row=target_row, column=col)
            if isinstance(target_cell, MergedCell):
                continue
            copy_cell_shape(source_cell, target_cell, translate_formula=False)
            if source_cell.value not in (None, ""):
                copied += 1
                if highlight:
                    highlight_wording_cell(ws_new, target_row, col)

    for merged_range in list(ws_prior.merged_cells.ranges):
        if (
            merged_range.min_row < source_start
            or merged_range.max_row > source_end
            or merged_range.min_col < col_start
            or merged_range.max_col > col_end
        ):
            continue
        row_shift = target_start - source_start
        try:
            ws_new.merge_cells(
                start_row=merged_range.min_row + row_shift,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row + row_shift,
                end_column=merged_range.max_col,
            )
        except ValueError:
            pass
    return copied


def copy_exact_section_with_shape(ws_prior, ws_new, source_start, source_end, target_start,
                                  target_existing_end=None, col_start=1, col_end=None,
                                  highlight=True):
    """Replace a target section with a source section, preserving row shape and merged cells."""
    source_len = source_end - source_start + 1
    if source_len <= 0:
        return 0

    if target_existing_end is None or target_existing_end < target_start:
        target_existing_len = 1
    else:
        target_existing_len = target_existing_end - target_start + 1

    if source_len > target_existing_len:
        insert_rows_preserving_sheet_metadata(ws_new, target_start + target_existing_len, source_len - target_existing_len)

    return copy_section_with_shape(
        ws_prior,
        ws_new,
        source_start,
        source_end,
        target_start,
        col_start=col_start,
        col_end=col_end,
        highlight=highlight,
    )


def find_section_end_by_blank_or_markers(ws, start_row, markers=(), max_rows=30):
    """Find a compact section end by next marker or trailing blank run."""
    last_content = start_row
    blank_run = 0
    for row in range(start_row, min(ws.max_row, start_row + max_rows) + 1):
        if row > start_row and markers and row_has_any_keyword(ws, row, markers, 1, min(ws.max_column, 30)):
            return max(start_row, row - 1)
        if row_has_content(ws, row, 1, min(ws.max_column, 30)):
            last_content = row
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row <= row <= merged_range.max_row:
                    anchor = ws.cell(merged_range.min_row, merged_range.min_col)
                    if anchor.value not in (None, ""):
                        last_content = max(last_content, merged_range.max_row)
            blank_run = 0
        elif row <= last_content:
            blank_run = 0
        else:
            blank_run += 1
            if row > start_row and blank_run >= 3:
                return last_content
    return last_content


def process_uexp_lead_adjustment_summary(ws_prior, ws_new):
    """Roll Uexp Lead adjustment summary with borders/styles intact."""
    source_start = find_row_containing(ws_prior, "调整汇总表", (1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, "调整汇总表", (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    source_end = find_section_end_by_blank_or_markers(
        ws_prior,
        source_start,
        markers=("Notes", "表1", "目标"),
        max_rows=20,
    )
    copied = copy_exact_section_with_shape(
        ws_prior,
        ws_new,
        source_start,
        source_end,
        target_start,
        target_existing_end=target_start + (source_end - source_start),
        col_start=2,
        col_end=min(ws_prior.max_column, ws_new.max_column, 12),
        highlight=True,
    )
    apply_thin_borders(ws_new, target_start, target_start + (source_end - source_start), 3, 7)
    return copied


def process_uexp_lead_expected_wording(ws_prior, ws_new):
    """Roll Uexp Lead expected fluctuation wording."""
    source_start = find_row_containing(ws_prior, "在下文中描述我们当期对账户波动的预期", (1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, "在下文中描述我们当期对账户波动的预期", (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    source_end = find_section_end_by_blank_or_markers(
        ws_prior,
        source_start,
        markers=("波动幅度", "波动范围", "表1"),
        max_rows=20,
    )
    target_next = find_row_containing_any(ws_new, ("波动幅度", "波动范围", "表1"), (target_start + 1, ws_new.max_row))
    target_existing_end = (target_next - 1) if target_next else target_start + (source_end - source_start)

    copied = copy_exact_section_with_shape(
        ws_prior,
        ws_new,
        source_start,
        source_end,
        target_start,
        target_existing_end=target_existing_end,
        col_start=2,
        col_end=min(ws_prior.max_column, ws_new.max_column, 12),
        highlight=True,
    )
    format_uexp_wording_boxes(
        ws_new,
        min(target_start + 1, target_start + (source_end - source_start)),
        target_start + (source_end - source_start),
    )
    return copied


def get_q1_covenant_image_row_shifts(ws_prior, ws_new):
    """Map prior Q1.05 note blocks to the rebuilt current-template blocks."""
    shifts = []
    prior_starts = find_q1_covenant_note_starts(ws_prior)
    output_starts = find_q1_covenant_note_starts(ws_new)
    prior_final = find_row_containing(
        ws_prior, "针对已经触发限制性条款", (1, ws_prior.max_row)
    )
    for index, (prior_start, output_start) in enumerate(zip(prior_starts, output_starts)):
        prior_end = (
            prior_starts[index + 1] - 1
            if index + 1 < len(prior_starts)
            else ((prior_final - 1) if prior_final else ws_prior.max_row)
        )
        shifts.append((prior_start, prior_end, output_start - prior_start))
    return shifts


def restore_template_drawing_parts(template_path, output_path):
    """Restore template drawings that openpyxl cannot round-trip safely."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_type_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    drawing_rel_type = f"{office_rel_ns}/drawing"

    def resolve_part(source_part, target):
        if target.startswith("/"):
            return target.lstrip("/")
        return posixpath.normpath(
            posixpath.join(posixpath.dirname(source_part), target)
        )

    def rels_part(source_part):
        return posixpath.join(
            posixpath.dirname(source_part),
            "_rels",
            posixpath.basename(source_part) + ".rels",
        )

    def sheet_parts(parts):
        workbook = ElementTree.fromstring(parts["xl/workbook.xml"])
        workbook_rels = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
        targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in workbook_rels}
        result = {}
        for sheet in workbook.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"):
            rel_id = sheet.attrib.get(f"{{{office_rel_ns}}}id")
            if rel_id in targets:
                result[sheet.attrib["name"]] = resolve_part(
                    "xl/workbook.xml", targets[rel_id]
                )
        return result

    with zipfile.ZipFile(template_path, "r") as template_zip:
        source_parts = {
            name: template_zip.read(name) for name in template_zip.namelist()
        }
        template_parts = {
            name: data
            for name, data in source_parts.items()
            if name.startswith(("xl/drawings/", "xl/media/"))
        }
        template_content_types = source_parts["[Content_Types].xml"]

    if not template_parts:
        return 0

    with zipfile.ZipFile(output_path, "r") as output_zip:
        output_parts = {
            name: output_zip.read(name) for name in output_zip.namelist()
        }

    output_parts.update(template_parts)

    source_sheets = sheet_parts(source_parts)
    output_sheets = sheet_parts(output_parts)
    ElementTree.register_namespace("", main_ns)
    ElementTree.register_namespace("r", office_rel_ns)
    ElementTree.register_namespace("", package_rel_ns)

    for sheet_name, source_sheet_part in source_sheets.items():
        output_sheet_part = output_sheets.get(sheet_name)
        source_sheet_rels_part = rels_part(source_sheet_part)
        if not output_sheet_part or source_sheet_rels_part not in source_parts:
            continue

        source_sheet_rels = ElementTree.fromstring(
            source_parts[source_sheet_rels_part]
        )
        source_drawing_rel = next(
            (
                rel
                for rel in source_sheet_rels
                if rel.attrib.get("Type") == drawing_rel_type
            ),
            None,
        )
        if source_drawing_rel is None:
            continue

        drawing_part = resolve_part(
            source_sheet_part, source_drawing_rel.attrib["Target"]
        )
        output_sheet_rels_part = rels_part(output_sheet_part)
        output_sheet_rels = (
            ElementTree.fromstring(output_parts[output_sheet_rels_part])
            if output_sheet_rels_part in output_parts
            else ElementTree.Element(f"{{{package_rel_ns}}}Relationships")
        )
        output_drawing_rel = next(
            (
                rel
                for rel in output_sheet_rels
                if rel.attrib.get("Type") == drawing_rel_type
            ),
            None,
        )
        if output_drawing_rel is None:
            existing_ids = {
                rel.attrib.get("Id", "") for rel in output_sheet_rels
            }
            rel_number = 1
            while f"rId{rel_number}" in existing_ids:
                rel_number += 1
            output_drawing_rel = ElementTree.SubElement(
                output_sheet_rels,
                f"{{{package_rel_ns}}}Relationship",
                {
                    "Id": f"rId{rel_number}",
                    "Type": drawing_rel_type,
                    "Target": posixpath.relpath(
                        drawing_part, posixpath.dirname(output_sheet_part)
                    ),
                },
            )
        else:
            output_drawing_rel.attrib["Target"] = posixpath.relpath(
                drawing_part, posixpath.dirname(output_sheet_part)
            )

        output_sheet = ElementTree.fromstring(output_parts[output_sheet_part])
        drawing_nodes = output_sheet.findall(f"{{{main_ns}}}drawing")
        if drawing_nodes:
            drawing_node = drawing_nodes[0]
            for duplicate in drawing_nodes[1:]:
                output_sheet.remove(duplicate)
        else:
            drawing_node = ElementTree.Element(f"{{{main_ns}}}drawing")
            trailing_tags = {
                "legacyDrawing", "legacyDrawingHF", "picture", "oleObjects",
                "controls", "webPublishItems", "tableParts", "extLst",
            }
            insert_at = len(output_sheet)
            for index, child in enumerate(list(output_sheet)):
                if child.tag.rsplit("}", 1)[-1] in trailing_tags:
                    insert_at = index
                    break
            output_sheet.insert(insert_at, drawing_node)
        drawing_node.attrib[f"{{{office_rel_ns}}}id"] = output_drawing_rel.attrib["Id"]

        output_parts[output_sheet_part] = ElementTree.tostring(
            output_sheet, encoding="utf-8", xml_declaration=True
        )
        output_parts[output_sheet_rels_part] = ElementTree.tostring(
            output_sheet_rels, encoding="utf-8", xml_declaration=True
        )

    content_types = ElementTree.fromstring(output_parts["[Content_Types].xml"])
    source_content_types = ElementTree.fromstring(template_content_types)
    existing_defaults = {
        node.attrib.get("Extension", "").lower()
        for node in content_types
        if node.tag == f"{{{content_type_ns}}}Default"
    }
    existing_overrides = {
        node.attrib.get("PartName", "")
        for node in content_types
        if node.tag == f"{{{content_type_ns}}}Override"
    }
    copied_extensions = {
        posixpath.splitext(name)[1].lstrip(".").lower()
        for name in template_parts
        if posixpath.splitext(name)[1]
    }
    copied_part_names = {"/" + name for name in template_parts}

    for node in source_content_types:
        if node.tag == f"{{{content_type_ns}}}Default":
            extension = node.attrib.get("Extension", "").lower()
            if extension in copied_extensions and extension not in existing_defaults:
                content_types.append(deepcopy(node))
                existing_defaults.add(extension)
        elif node.tag == f"{{{content_type_ns}}}Override":
            part_name = node.attrib.get("PartName", "")
            if part_name in copied_part_names and part_name not in existing_overrides:
                content_types.append(deepcopy(node))
                existing_overrides.add(part_name)

    ElementTree.register_namespace("", content_type_ns)
    output_parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )

    temp_path = str(output_path) + ".drawing-restore.tmp"
    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, data in output_parts.items():
                archive.writestr(name, data)
        os.replace(temp_path, output_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return len(template_parts)


def copy_q1_review_images(prior_path, output_path, q105_row_shifts=None):
    """Copy only Q1 Note-area and covenant evidence images at package level."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    content_type_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    drawing_rel_type = f"{office_rel_ns}/drawing"
    image_rel_type = f"{office_rel_ns}/image"
    drawing_content_type = "application/vnd.openxmlformats-officedocument.drawing+xml"

    ElementTree.register_namespace("", main_ns)
    ElementTree.register_namespace("r", office_rel_ns)
    ElementTree.register_namespace("xdr", drawing_ns)
    ElementTree.register_namespace("", package_rel_ns)

    q105_row_shifts = q105_row_shifts or []

    def resolve_part(source_part, target):
        if target.startswith("/"):
            return target.lstrip("/")
        return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))

    def rels_part(source_part):
        return posixpath.join(
            posixpath.dirname(source_part),
            "_rels",
            posixpath.basename(source_part) + ".rels",
        )

    def sheet_parts(parts):
        workbook = ElementTree.fromstring(parts["xl/workbook.xml"])
        workbook_rels = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
        targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in workbook_rels}
        result = {}
        for sheet in workbook.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"):
            rel_id = sheet.attrib.get(f"{{{office_rel_ns}}}id")
            if rel_id in targets:
                result[sheet.attrib["name"]] = resolve_part("xl/workbook.xml", targets[rel_id])
        return result

    with zipfile.ZipFile(prior_path, "r") as source_zip:
        prior_parts = {name: source_zip.read(name) for name in source_zip.namelist()}
    with zipfile.ZipFile(output_path, "r") as output_zip:
        output_parts = {name: output_zip.read(name) for name in output_zip.namelist()}

    prior_sheets = sheet_parts(prior_parts)
    output_sheets = sheet_parts(output_parts)
    selectors = {
        "Q1.01": 30,
        "Q1.05": 40,
    }
    copied_count = 0

    drawing_numbers = [
        int(match.group(1))
        for name in output_parts
        if (match := re.fullmatch(r"xl/drawings/drawing(\d+)\.xml", name))
    ]
    next_drawing_number = max(drawing_numbers, default=0) + 1

    for sheet_prefix, minimum_anchor_row in selectors.items():
        prior_sheet_name = next((name for name in prior_sheets if name.startswith(sheet_prefix)), None)
        output_sheet_name = next((name for name in output_sheets if name.startswith(sheet_prefix)), None)
        if not prior_sheet_name or not output_sheet_name:
            continue

        prior_sheet_part = prior_sheets[prior_sheet_name]
        prior_sheet_rels_part = rels_part(prior_sheet_part)
        if prior_sheet_rels_part not in prior_parts:
            continue
        prior_sheet_rels = ElementTree.fromstring(prior_parts[prior_sheet_rels_part])
        prior_drawing_rel = next(
            (rel for rel in prior_sheet_rels if rel.attrib.get("Type") == drawing_rel_type),
            None,
        )
        if prior_drawing_rel is None:
            continue

        prior_drawing_part = resolve_part(prior_sheet_part, prior_drawing_rel.attrib["Target"])
        prior_drawing_rels_part = rels_part(prior_drawing_part)
        if prior_drawing_part not in prior_parts or prior_drawing_rels_part not in prior_parts:
            continue
        prior_drawing = ElementTree.fromstring(prior_parts[prior_drawing_part])
        prior_drawing_rels = ElementTree.fromstring(prior_parts[prior_drawing_rels_part])
        prior_rel_lookup = {rel.attrib["Id"]: rel for rel in prior_drawing_rels}

        selected_anchors = []
        used_rel_ids = set()
        for anchor in list(prior_drawing):
            from_node = anchor.find(f"{{{drawing_ns}}}from")
            row_node = from_node.find(f"{{{drawing_ns}}}row") if from_node is not None else None
            if row_node is None or int(row_node.text or 0) < minimum_anchor_row:
                continue
            if anchor.find(f"{{{drawing_ns}}}pic") is None:
                continue
            anchor_copy = deepcopy(anchor)
            if sheet_prefix == "Q1.05" and q105_row_shifts:
                source_row = int(row_node.text or 0) + 1
                row_shift = next(
                    (shift for start, end, shift in q105_row_shifts if start <= source_row <= end),
                    0,
                )
                if row_shift:
                    for marker_name in ("from", "to"):
                        marker = anchor_copy.find(f"{{{drawing_ns}}}{marker_name}")
                        marker_row = marker.find(f"{{{drawing_ns}}}row") if marker is not None else None
                        if marker_row is not None:
                            marker_row.text = str(max(0, int(marker_row.text or 0) + row_shift))
            for node in anchor_copy.iter():
                for attr_name, attr_value in node.attrib.items():
                    if attr_name.startswith(f"{{{office_rel_ns}}}"):
                        used_rel_ids.add(attr_value)
            selected_anchors.append(anchor_copy)

        if not selected_anchors:
            continue

        output_sheet_part = output_sheets[output_sheet_name]
        output_sheet_rels_part = rels_part(output_sheet_part)
        if output_sheet_rels_part in output_parts:
            output_sheet_rels = ElementTree.fromstring(output_parts[output_sheet_rels_part])
        else:
            output_sheet_rels = ElementTree.Element(f"{{{package_rel_ns}}}Relationships")
        existing_drawing_rel = next(
            (rel for rel in output_sheet_rels if rel.attrib.get("Type") == drawing_rel_type),
            None,
        )
        is_new_drawing = existing_drawing_rel is None
        if existing_drawing_rel is not None:
            drawing_part = resolve_part(output_sheet_part, existing_drawing_rel.attrib["Target"])
            drawing_rels_part = rels_part(drawing_part)
            if drawing_part not in output_parts:
                continue
            drawing_root = ElementTree.fromstring(output_parts[drawing_part])
            drawing_rels_root = (
                ElementTree.fromstring(output_parts[drawing_rels_part])
                if drawing_rels_part in output_parts
                else ElementTree.Element(f"{{{package_rel_ns}}}Relationships")
            )
        else:
            drawing_part = f"xl/drawings/drawing{next_drawing_number}.xml"
            drawing_rels_part = rels_part(drawing_part)
            next_drawing_number += 1
            drawing_root = ElementTree.Element(prior_drawing.tag, prior_drawing.attrib)
            drawing_rels_root = ElementTree.Element(f"{{{package_rel_ns}}}Relationships")

        existing_drawing_rel_ids = {
            rel.attrib.get("Id", "") for rel in drawing_rels_root
        }
        rel_id_map = {}
        for rel_id in sorted(used_rel_ids):
            prior_rel = prior_rel_lookup.get(rel_id)
            if prior_rel is None:
                continue
            rel_copy = deepcopy(prior_rel)
            rel_number = 1
            while f"rId{rel_number}" in existing_drawing_rel_ids:
                rel_number += 1
            output_rel_id = f"rId{rel_number}"
            rel_copy.attrib["Id"] = output_rel_id
            if rel_copy.attrib.get("Type") == image_rel_type:
                prior_media_part = resolve_part(prior_drawing_part, rel_copy.attrib["Target"])
                if prior_media_part not in prior_parts:
                    continue
                media_part = prior_media_part
                if media_part in output_parts and output_parts[media_part] != prior_parts[prior_media_part]:
                    stem, extension = posixpath.splitext(posixpath.basename(media_part))
                    index = 1
                    while f"xl/media/{stem}_q1_{index}{extension}" in output_parts:
                        index += 1
                    media_part = f"xl/media/{stem}_q1_{index}{extension}"
                output_parts[media_part] = prior_parts[prior_media_part]
                rel_copy.attrib["Target"] = posixpath.relpath(media_part, posixpath.dirname(drawing_part))
            drawing_rels_root.append(rel_copy)
            existing_drawing_rel_ids.add(output_rel_id)
            rel_id_map[rel_id] = output_rel_id

        if not rel_id_map:
            continue

        existing_shape_ids = []
        for node in drawing_root.iter():
            if node.tag.rsplit("}", 1)[-1] == "cNvPr":
                try:
                    existing_shape_ids.append(int(node.attrib.get("id", "0")))
                except ValueError:
                    pass
        next_shape_id = max(existing_shape_ids, default=0) + 1
        added_anchors = []
        for anchor in selected_anchors:
            for node in anchor.iter():
                for attr_name, attr_value in list(node.attrib.items()):
                    if attr_name.startswith(f"{{{office_rel_ns}}}") and attr_value in rel_id_map:
                        node.attrib[attr_name] = rel_id_map[attr_value]
                if node.tag.rsplit("}", 1)[-1] == "cNvPr":
                    node.attrib["id"] = str(next_shape_id)
                    node.attrib["name"] = f"Q1 Roll Image {next_shape_id}"
                    next_shape_id += 1
            referenced = {
                attr_value
                for node in anchor.iter()
                for attr_name, attr_value in node.attrib.items()
                if attr_name.startswith(f"{{{office_rel_ns}}}")
            }
            if referenced and not referenced.issubset(existing_drawing_rel_ids):
                continue
            drawing_root.append(anchor)
            added_anchors.append(anchor)
        if not added_anchors:
            continue

        output_parts[drawing_part] = ElementTree.tostring(
            drawing_root, encoding="utf-8", xml_declaration=True
        )
        output_parts[drawing_rels_part] = ElementTree.tostring(
            drawing_rels_root, encoding="utf-8", xml_declaration=True
        )

        if is_new_drawing:
            existing_ids = {
                rel.attrib.get("Id", "") for rel in output_sheet_rels
            }
            rel_number = 1
            while f"rId{rel_number}" in existing_ids:
                rel_number += 1
            sheet_drawing_rel_id = f"rId{rel_number}"
            ElementTree.SubElement(
                output_sheet_rels,
                f"{{{package_rel_ns}}}Relationship",
                {
                    "Id": sheet_drawing_rel_id,
                    "Type": drawing_rel_type,
                    "Target": posixpath.relpath(drawing_part, posixpath.dirname(output_sheet_part)),
                },
            )
            output_sheet = ElementTree.fromstring(output_parts[output_sheet_part])
            drawing_node = ElementTree.Element(
                f"{{{main_ns}}}drawing",
                {f"{{{office_rel_ns}}}id": sheet_drawing_rel_id},
            )
            trailing_tags = {
                "legacyDrawing", "legacyDrawingHF", "picture", "oleObjects", "controls",
                "webPublishItems", "tableParts", "extLst",
            }
            insert_at = len(output_sheet)
            for index, child in enumerate(list(output_sheet)):
                if child.tag.rsplit("}", 1)[-1] in trailing_tags:
                    insert_at = index
                    break
            output_sheet.insert(insert_at, drawing_node)
            output_parts[output_sheet_part] = ElementTree.tostring(
                output_sheet, encoding="utf-8", xml_declaration=True
            )
        output_parts[output_sheet_rels_part] = ElementTree.tostring(
            output_sheet_rels, encoding="utf-8", xml_declaration=True
        )
        copied_count += len(added_anchors)

    if not copied_count:
        return 0

    content_types = ElementTree.fromstring(output_parts["[Content_Types].xml"])
    existing_defaults = {
        node.attrib.get("Extension", "").lower() for node in content_types
        if node.tag == f"{{{content_type_ns}}}Default"
    }
    media_types = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg"}
    for media_part in [name for name in output_parts if name.startswith("xl/media/")]:
        extension = posixpath.splitext(media_part)[1].lstrip(".").lower()
        if extension in media_types and extension not in existing_defaults:
            ElementTree.SubElement(
                content_types,
                f"{{{content_type_ns}}}Default",
                {"Extension": extension, "ContentType": media_types[extension]},
            )
            existing_defaults.add(extension)
    existing_overrides = {
        node.attrib.get("PartName") for node in content_types
        if node.tag == f"{{{content_type_ns}}}Override"
    }
    for drawing_part in [name for name in output_parts if re.fullmatch(r"xl/drawings/drawing\d+\.xml", name)]:
        part_name = "/" + drawing_part
        if part_name not in existing_overrides:
            ElementTree.SubElement(
                content_types,
                f"{{{content_type_ns}}}Override",
                {"PartName": part_name, "ContentType": drawing_content_type},
            )
            existing_overrides.add(part_name)
    ElementTree.register_namespace("", content_type_ns)
    output_parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )

    temp_path = str(output_path) + ".q1-images.tmp"
    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, data in output_parts.items():
                archive.writestr(name, data)
        os.replace(temp_path, output_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
    return copied_count


def process_uexp_lead_notes(ws_prior, ws_new):
    """Roll Uexp Lead Notes text into the current notes box."""
    source_row = find_row_containing(ws_prior, "Notes", (1, ws_prior.max_row))
    target_row = find_row_containing(ws_new, "Notes", (1, ws_new.max_row))
    if not source_row or not target_row:
        return 0

    copied = 0
    for offset in range(0, 6):
        source = source_row + offset
        target = target_row + offset
        for col in range(2, min(ws_prior.max_column, ws_new.max_column, 12) + 1):
            copied += copy_changed_constant_cell(ws_prior, ws_new, source, col, target, col)
    format_uexp_wording_boxes(ws_new, target_row, min(ws_new.max_row, target_row + 5))
    return copied


def process_uexp_lead_wording(ws_prior, ws_new):
    """Roll Uexp Lead notes and adjustment summary using dedicated sections."""
    copied = 0
    copied += process_uexp_lead_expected_wording(ws_prior, ws_new)
    copied += process_uexp_lead_notes(ws_prior, ws_new)
    copied += process_uexp_lead_adjustment_summary(ws_prior, ws_new)
    return copied


def process_uexp_bkd_expected_wording(ws_prior, ws_new):
    """Roll Uexp finance BKD expected fluctuation wording, excluding the procedure table."""
    source_start = find_row_containing(ws_prior, "在下文中描述我们当期对账户波动的预期", (1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, "在下文中描述我们当期对账户波动的预期", (1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    source_end = find_section_end_by_blank_or_markers(
        ws_prior,
        source_start,
        markers=("波动范围", "表1"),
        max_rows=20,
    )
    target_next = find_row_containing_any(ws_new, ("波动范围", "表1"), (target_start + 1, ws_new.max_row))
    target_existing_end = (target_next - 1) if target_next else target_start + (source_end - source_start)

    copied = copy_exact_section_with_shape(
        ws_prior,
        ws_new,
        source_start,
        source_end,
        target_start,
        target_existing_end=target_existing_end,
        col_start=2,
        col_end=min(ws_prior.max_column, ws_new.max_column, 12),
        highlight=True,
    )
    format_uexp_wording_boxes(
        ws_new,
        min(target_start + 1, target_start + (source_end - source_start)),
        target_start + (source_end - source_start),
    )
    return copied


def process_uexp_bkd_notes(ws_prior, ws_new):
    """Roll Uexp finance BKD bottom Notes/NB explanations."""
    source_header = find_expense_bkd_header_row(ws_prior)
    target_header = find_expense_bkd_header_row(ws_new)
    source_total = find_expense_bkd_total_row(ws_prior, source_header) if source_header else None
    target_total = find_expense_bkd_total_row(ws_new, target_header) if target_header else None

    source_start = find_row_containing(ws_prior, "Notes", ((source_total or 1) + 1, ws_prior.max_row))
    target_start = find_row_containing(ws_new, "Notes", ((target_total or 1) + 1, ws_new.max_row))
    if not source_start or not target_start:
        return 0

    source_end = source_start
    for row in range(source_start, min(ws_prior.max_row, source_start + 80) + 1):
        if row > source_start and row_has_any_keyword(ws_prior, row, ("调整汇总表", "表2", "截止"), 1, min(ws_prior.max_column, 30)):
            break
        if row_has_content(ws_prior, row, 1, min(ws_prior.max_column, 30)):
            source_end = row
    target_end_marker = find_row_containing(ws_new, "表2", (target_start + 1, ws_new.max_row))
    target_existing_end = (target_end_marker - 1) if target_end_marker else target_start + (source_end - source_start)

    copied = copy_exact_section_with_shape(
        ws_prior,
        ws_new,
        source_start,
        source_end,
        target_start,
        target_existing_end=target_existing_end,
        col_start=2,
        col_end=min(ws_prior.max_column, ws_new.max_column, 18),
        highlight=True,
    )
    format_uexp_wording_boxes(
        ws_new,
        target_start,
        target_start + (source_end - source_start),
    )
    return copied


def process_uexp_finance_bkd_wording(ws_prior, ws_new):
    """Roll the required Uexp finance BKD wording areas only."""
    copied = 0
    copied += process_uexp_bkd_expected_wording(ws_prior, ws_new)
    copied += process_uexp_bkd_notes(ws_prior, ws_new)
    return copied


def process_l103_policy_table(ws_prior, ws_new):
    """Roll L1.03 table 2 prior-year policy information into the new workbook."""
    copied = 0
    prior_rows = {}

    for row in range(1, ws_prior.max_row + 1):
        category = ws_prior.cell(row=row, column=2).value
        if not category:
            continue
        category_key = normalize_text(category)
        if not category_key or "资产类别" in category_key or "表2" in category_key:
            continue
        current_life = ws_prior.cell(row=row, column=3).value
        reason = ws_prior.cell(row=row, column=7).value
        if current_life is not None or reason is not None:
            prior_rows[category_key] = {
                "life": current_life,
                "reason": reason,
            }

    for row in range(1, ws_new.max_row + 1):
        category = ws_new.cell(row=row, column=2).value
        if not category:
            continue
        category_key = normalize_text(category)
        if category_key not in prior_rows:
            continue

        data = prior_rows[category_key]
        if data["life"] is not None:
            set_cell_value(ws_new, row, 4, data["life"])
            copied += 1
        if data["reason"] is not None:
            set_cell_value(ws_new, row, 7, data["reason"])
            copied += 1

    return copied


def process_k033_depreciation_policy(ws_prior, ws_new):
    """Roll prior K.03.3 current depreciation policy into the new review table."""
    copied = 0
    policy_rows = []

    for row in range(1, ws_prior.max_row + 1):
        category = ws_prior.cell(row=row, column=2).value
        useful_life = ws_prior.cell(row=row, column=3).value
        residual_rate = ws_prior.cell(row=row, column=4).value
        if category in (None, "") or useful_life in (None, ""):
            continue
        category_key = normalize_text(category)
        if (
            not category_key
            or "折旧" in category_key
            or "资产类别" in category_key
            or "Notes" in category_key
            or "表1" in category_key
            or "公司折旧政策" in category_key
        ):
            continue
        policy_rows.append({
            "category": category,
            "life": useful_life,
            "residual_rate": residual_rate,
        })

    if not policy_rows:
        return 0

    start_row = None
    for row in range(1, ws_new.max_row + 1):
        row_values = [normalize_text(ws_new.cell(row=row, column=col).value) for col in range(2, 8)]
        if any("折旧政策" in value for value in row_values) and any("使用寿命" in value for value in row_values):
            start_row = row + 1
            break

    for row in range(1, ws_new.max_row + 1):
        if start_row is not None:
            break
        if normalize_text(ws_new.cell(row=row, column=2).value) == "":
            formula_cell = ws_new.cell(row=row, column=5).value
            if isinstance(formula_cell, str) and formula_cell.startswith("="):
                start_row = row
                break
    if start_row is None:
        start_row = 13

    existing_capacity = 0
    for row in range(start_row, ws_new.max_row + 1):
        if normalize_text(ws_new.cell(row=row, column=2).value) == "Notes":
            break
        if ws_new.cell(row=row, column=5).value is not None:
            existing_capacity += 1

    rows_to_insert = max(0, len(policy_rows) - existing_capacity)
    if rows_to_insert:
        insert_at = start_row + existing_capacity
        source_row = max(start_row, insert_at - 1)
        insert_rows_preserving_sheet_metadata(ws_new, insert_at, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, source_row, insert_at + offset, translate_formula=True)

    for offset, policy in enumerate(policy_rows):
        row = start_row + offset
        for col, value in (
            (2, policy["category"]),
            (6, policy["life"]),
            (7, policy["residual_rate"]),
        ):
            set_cell_value(ws_new, row, col, value)
            copied += 1
        set_cell_value(ws_new, row, 3, None)
        set_cell_value(ws_new, row, 4, None)

    notes_value = None
    for row in range(1, ws_prior.max_row + 1):
        if normalize_text(ws_prior.cell(row=row, column=2).value) == "Notes":
            for notes_row in range(row + 1, ws_prior.max_row + 1):
                value = ws_prior.cell(row=notes_row, column=2).value
                if value not in (None, ""):
                    notes_value = value
                    break
            break

    if notes_value is not None:
        notes_row = None
        for row in range(1, ws_new.max_row + 1):
            if normalize_text(ws_new.cell(row=row, column=2).value) == "Notes":
                notes_row = row + 1
                break
        if notes_row:
            set_cell_value(ws_new, notes_row, 2, notes_value)
            copied += 1

    return copied


def process_n_lead_turnover_analysis(ws_prior_values, ws_new):
    """Roll N.00 table 2 prior-year turnover analysis values into PY column."""
    section_text = "表2 应付账款周转率分析"
    prior_section_row = find_row_containing(ws_prior_values, section_text, (1, ws_prior_values.max_row))
    new_section_row = find_row_containing(ws_new, section_text, (1, ws_new.max_row))
    if not prior_section_row or not new_section_row:
        return 0

    prior_source_col = 4
    new_target_col = None
    for row in range(new_section_row, min(new_section_row + 5, ws_new.max_row) + 1):
        for col in range(1, min(12, ws_new.max_column) + 1):
            value = ws_new.cell(row=row, column=col).value
            if value and "PY" in str(value):
                new_target_col = col
                break
        if new_target_col:
            break
    if not new_target_col:
        new_target_col = 5

    prior_values = {}
    for row in range(prior_section_row + 1, min(prior_section_row + 20, ws_prior_values.max_row) + 1):
        label = normalize_text(ws_prior_values.cell(row=row, column=2).value)
        if not label:
            continue
        value = ws_prior_values.cell(row=row, column=prior_source_col).value
        if value is not None:
            prior_values[label] = value

    copied = 0
    for row in range(new_section_row + 1, min(new_section_row + 20, ws_new.max_row) + 1):
        label = normalize_text(ws_new.cell(row=row, column=2).value)
        if label in prior_values:
            set_cell_value(ws_new, row, new_target_col, prior_values[label])
            copied += 1

    return copied


def find_n_turnover_analysis_marker_row(ws, section_row):
    """Find the actual N turnover-analysis wording marker, not the table title."""
    for row in range(section_row + 1, min(section_row + 80, ws.max_row) + 1):
        for col in range(1, min(6, ws.max_column) + 1):
            text = normalize_text(ws.cell(row=row, column=col).value)
            if text.startswith("分析"):
                return row
    return None


def find_n_turnover_analysis_text_rows(ws, section_row):
    """Find text rows below the N turnover analysis marker."""
    marker_row = find_n_turnover_analysis_marker_row(ws, section_row)
    if not marker_row:
        return None, []

    text_rows = []
    blank_run = 0
    for row in range(marker_row + 1, min(marker_row + 80, ws.max_row) + 1):
        row_text = get_row_text(ws, row, 1, min(ws.max_column, 12))
        normalized = normalize_text(row_text)

        if row_has_any_keyword(ws, row, ("调整汇总", "调整分录", "异常项复核", "截止测试")):
            break
        if re.match(r"^表[0-9一二三四五六七八九十]+", normalized):
            break

        if not row_text:
            blank_run += 1
            if text_rows and blank_run >= 3:
                break
            continue

        blank_run = 0
        text_rows.append(row)

    return marker_row, text_rows


def find_n_turnover_analysis_target_end(ws, marker_row):
    """Find the row before the next N turnover-analysis section."""
    last_content_row = marker_row
    blank_run = 0
    for row in range(marker_row + 1, min(marker_row + 80, ws.max_row) + 1):
        if row_has_any_keyword(ws, row, ("调整汇总", "调整分录", "异常项复核", "截止测试")):
            return row - 1

        row_text = get_row_text(ws, row, 1, min(ws.max_column, 12))
        normalized = normalize_text(row_text)
        if re.match(r"^表[0-9一二三四五六七八九十]+", normalized):
            return row - 1

        if row_text:
            last_content_row = row
            blank_run = 0
        else:
            blank_run += 1
            if row > marker_row + 1 and blank_run >= 3:
                return max(last_content_row, marker_row + 1)

    return max(last_content_row, marker_row + 1)


def process_n_lead_turnover_wording(ws_prior_values, ws_new):
    """Roll N.00 turnover-analysis wording using dynamic section anchors."""
    section_text = "表2 应付账款周转率分析"
    prior_section_row = find_row_containing(ws_prior_values, section_text, (1, ws_prior_values.max_row))
    new_section_row = find_row_containing(ws_new, section_text, (1, ws_new.max_row))
    if not prior_section_row or not new_section_row:
        return 0

    _, source_rows = find_n_turnover_analysis_text_rows(ws_prior_values, prior_section_row)
    target_marker_row, _ = find_n_turnover_analysis_text_rows(ws_new, new_section_row)
    if not source_rows or not target_marker_row:
        return 0

    target_end_row = find_n_turnover_analysis_target_end(ws_new, target_marker_row)
    available_rows = max(1, target_end_row - target_marker_row)
    rows_to_insert = max(0, len(source_rows) - available_rows)
    if rows_to_insert:
        insert_at = target_end_row + 1
        source_shape_row = max(target_marker_row + 1, target_end_row)
        insert_rows_preserving_sheet_metadata(ws_new, insert_at, rows_to_insert)
        for offset in range(rows_to_insert):
            copy_row_shape(ws_new, source_shape_row, insert_at + offset, translate_formula=True)

    copied = 0
    max_col = min(ws_prior_values.max_column, ws_new.max_column, 12)
    for offset, source_row in enumerate(source_rows):
        target_row = target_marker_row + 1 + offset
        for col in range(1, max_col + 1):
            value = ws_prior_values.cell(row=source_row, column=col).value
            if value in (None, ""):
                set_cell_value(ws_new, target_row, col, None)
                continue
            if set_cell_value(ws_new, target_row, col, value):
                highlight_wording_cell(ws_new, target_row, col)
                copied += 1

    return copied


def process_n_detail_sheet(ws_prior_formula, ws_prior_values, ws_new, bs_date, ws_prior_lead=None, ws_new_lead=None, roll_forward_wording=False):
    """Copy N.01.01 detail sheet and roll prior closing values into the PY column."""
    clone_worksheet_contents(ws_prior_formula, ws_new)

    date_obj = parse_date_value(bs_date)
    if date_obj:
        set_cell_value(ws_new, 5, 8, date_obj)
        ws_new.cell(row=5, column=8).number_format = "yyyy/mm/dd"
        prior_date = datetime.datetime(date_obj.year - 1, 12, 31)
        set_cell_value(ws_new, 5, 14, prior_date)
        ws_new.cell(row=5, column=14).number_format = "yyyy/mm/dd"
        set_cell_value(ws_new, 225, 3, date_obj)
        ws_new.cell(row=225, column=3).number_format = "yyyy/mm/dd"
        set_cell_value(ws_new, 225, 5, prior_date)
        ws_new.cell(row=225, column=5).number_format = "yyyy/mm/dd"
        set_cell_value(ws_new, 257, 4, date_obj)
        ws_new.cell(row=257, column=4).number_format = "yyyy/mm/dd"

    header_row = find_header_row(ws_new, "期末审定数", (1, 30))
    if not header_row:
        return 0

    total_row = find_total_row_after(ws_new, header_row)
    data_end_row = (total_row - 1) if total_row else ws_new.max_row
    closing_col = find_header_col(ws_new, header_row, ["期末审定数"]) or 13
    py_col = find_header_col(ws_new, header_row, ["上期末审定数", "上年数"]) or 14
    current_cols = [
        col for col in [
            find_header_col(ws_new, header_row, ["原币金额"]),
            find_header_col(ws_new, header_row, ["期末账面数"]),
            find_header_col(ws_new, header_row, ["本期审计调整编号"]),
            find_header_col(ws_new, header_row, ["审计调整"]),
            find_header_col(ws_new, header_row, ["重分类调整"]),
            closing_col,
        ]
        if col
    ]
    aging_cols = find_group_child_cols(
        ws_new,
        header_row - 1,
        header_row,
        ["账龄"],
        stop_keywords=["check", "检查"],
    )
    current_cols = sorted(set(current_cols + aging_cols))

    copied = 0
    for row in range(header_row + 1, data_end_row + 1):
        source_value = ws_prior_values.cell(row=row, column=closing_col).value
        if source_value is not None:
            set_cell_value(ws_new, row, py_col, source_value)
            copied += 1

    clear_constant_cells(ws_new, header_row + 1, data_end_row, current_cols)

    if ws_prior_lead is not None and ws_new_lead is not None:
        prior_header_row = find_header_row(ws_prior_lead, "期末审定数", (1, 100))
        new_header_row = find_header_row(ws_new_lead, "期末审定数", (1, 120))
        prior_total_row = find_total_row_after(ws_prior_lead, prior_header_row)
        new_total_row = find_total_row_after(ws_new_lead, new_header_row)
        if prior_total_row and new_total_row and prior_total_row != new_total_row:
            for row in range(1, ws_new.max_row + 1):
                for col in range(1, ws_new.max_column + 1):
                    cell = ws_new.cell(row=row, column=col)
                    formula = cell.value
                    if isinstance(formula, str) and formula.startswith("=") and "'N.00 Lead sheet'!" in formula:
                        cell.value = re.sub(
                            rf"('N\.00 Lead sheet'!\$?[A-Z]{{1,3}})\$?{prior_total_row}(?!\d)",
                            rf"\g<1>{new_total_row}",
                            formula,
                        )

    wording_start_row = None
    for row in range(1, ws_new.max_row + 1):
        row_text = " ".join(
            str(ws_new.cell(row=row, column=col).value or "")
            for col in range(1, min(ws_new.max_column, 10) + 1)
        )
        if "对于单项变动金额" in row_text:
            wording_start_row = row
            break
    if wording_start_row:
        wording_rows = []
        for row in range(wording_start_row, ws_new.max_row + 1):
            if any(
                ws_new.cell(row=row, column=col).value not in (None, "")
                for col in range(2, min(ws_new.max_column, 8) + 1)
            ):
                wording_rows.append(row)

        clear_area_format(
            ws_new,
            wording_start_row,
            ws_new.max_row,
            9,
            min(ws_new.max_column, 24),
            clear_fill=True,
            clear_border=True,
        )
        for row in range(wording_start_row, ws_new.max_row + 1):
            if not row_has_content(ws_new, row, 2, min(ws_new.max_column, 8)):
                clear_area_format(
                    ws_new,
                    row,
                    row,
                    1,
                    min(ws_new.max_column, 24),
                    clear_fill=True,
                    clear_border=True,
                )

        if roll_forward_wording:
            for row in wording_rows:
                for col in range(2, min(ws_new.max_column, 8) + 1):
                    if ws_new.cell(row=row, column=col).value not in (None, ""):
                        highlight_wording_cell(ws_new, row, col)
        else:
            for row in wording_rows:
                for col in range(2, min(ws_new.max_column, 8) + 1):
                    set_cell_value(ws_new, row, col, None)

    tidy_n_detail_sheet_borders(ws_new, header_row, total_row)
    ws_new.sheet_view.showGridLines = False

    return copied


def generate_output_filename(template_name, bs_date, company_name):
    """生成输出文件名"""
    output_name = template_name
    # 替换各种日期占位符
    output_name = re.sub(r'202[A-Za-z]{4,6}', bs_date.replace('-', ''), output_name)
    output_name = re.sub(r'\d{4}-\d{2}-\d{2}', bs_date, output_name)
    output_name = re.sub(r'\d{8}', bs_date.replace('-', ''), output_name)

    # 替换公司名称占位符
    output_name = re.sub(r'XYZ公司', company_name, output_name)
    output_name = re.sub(r'[<>:"/\\|?*]', '_', output_name)

    return output_name


def process_single_subject(subject_code, template_path, prior_path, pmte_path,
                           company_name, bs_date, output_dir, subject_config,
                           cra_path=None, functional_currency=None, accounting_standard=None,
                           pm_value=None, te_value=None, sad_value=None, cra_records=None,
                           roll_forward_wording=False, generate_summary=True,
                           progress_callback=None, llm_enhanced=False,
                           llm_wording_revision=False, llm_options=None):
    """
    处理单个科目的Roll Forward

    Args:
        subject_code: 科目代码（如"K1"）
        template_path: 标准模板路径
        prior_path: 上年底稿路径
        pmte_path: PMTE信息表路径
        company_name: 公司名称
        bs_date: 资产负债表日期（格式: YYYY-MM-DD）
        output_dir: 输出目录
        subject_config: 科目配置
        cra_path: CRA等级表路径（可选，默认与pmte_path相同）

    Returns:
        (success: bool, message: str, output_path: str, warnings: list)
    """
    warnings_list = RollForwardWarnings()
    work_output_path = None

    try:
        # 1. 提取公司信息
        company_info = extract_company_info_from_pmte(pmte_path, company_name)
        company_info["functional_currency"] = functional_currency
        company_info["accounting_standard"] = accounting_standard
        if pm_value not in (None, ""):
            company_info["PM"] = pm_value
        if te_value not in (None, ""):
            company_info["TE"] = te_value
        if sad_value not in (None, ""):
            company_info["SAD"] = sad_value

        # 1.1 加载CRA等级表数据
        # 如果未指定CRA路径，尝试从PMTE同目录查找CRA文件
        cra_data = None
        if cra_path and os.path.exists(cra_path):
            cra_data = load_cra_data(cra_path, subject_code)
        elif pmte_path and os.path.exists(pmte_path):
            cra_data = load_cra_data(pmte_path, subject_code)

        if cra_data:
            company_info["cra_data"] = cra_data

        # 2. 生成输出文件名
        template_name = os.path.basename(template_path)
        output_name = generate_output_filename(template_name, bs_date, company_name)
        final_output_path = os.path.join(output_dir, output_name)
        work_output_path = final_output_path + f".{os.getpid()}.partial.xlsx"
        output_path = work_output_path

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 3. 复制标准模板
        shutil.copy2(template_path, output_path)
        # 移除只读属性（Windows）
        os.chmod(output_path, 0o666)

        # 4. 打开新底稿和上年底稿
        # 大文件用 openpyxl 加载很慢。只有 wording/L1/N 等需要公式或完整复制结构时，
        # 才同时打开公式副本；普通 roll forward 只打开 data_only 工作簿。
        needs_formula_workbook = roll_forward_wording or subject_code in {"L1", "N"}
        if progress_callback:
            size_mb = os.path.getsize(prior_path) / (1024 * 1024)
            progress_callback(f"正在加载上年底稿: {os.path.basename(prior_path)} ({size_mb:.1f} MB)")
        wb_new = openpyxl.load_workbook(
            output_path,
            keep_links=(subject_code != "UexpVCVD"),
        )
        wb_prior_values = openpyxl.load_workbook(prior_path, data_only=True)
        wb_prior_formula = (
            openpyxl.load_workbook(prior_path, data_only=False)
            if needs_formula_workbook
            else wb_prior_values
        )
        before_snapshot = workbook_snapshot(wb_new)

        try:
            lead_config = subject_config.get("lead_sheet", {})
            k01_config = subject_config.get("k01", {})
            wording_copied_count = 0
            wording_touched_sheets = set()

            if "汇总" in wb_new.sheetnames:
                process_summary_sheet(wb_new["汇总"], company_info, bs_date, company_name, warnings_list)

            # 5. 处理Lead Sheet
            lead_sheet_name = lead_config.get("sheet_name", "")
            if lead_sheet_name and lead_sheet_name in wb_new.sheetnames and is_visible_worksheet(wb_prior_formula, lead_sheet_name):
                ws_new_lead = wb_new[lead_sheet_name]
                ws_prior_formula_lead = wb_prior_formula[lead_sheet_name]
                ws_prior_values_lead = wb_prior_values[lead_sheet_name]
                process_lead_sheet(ws_prior_formula_lead, ws_new_lead, ws_prior_values_lead,
                                   company_info, bs_date, company_name, lead_config, warnings_list)
                if subject_code == "L1":
                    repair_l1_lead_detail_formulas(
                        ws_prior_formula_lead, ws_new_lead
                    )
                    rebuild_l1_lead_total_formulas(ws_new_lead)
            elif lead_sheet_name and lead_sheet_name in wb_new.sheetnames:
                ws_new_lead = wb_new[lead_sheet_name]
                fill_basic_lead_header(ws_new_lead, company_info, bs_date, company_name)

            # 6. 处理K.01
            if k01_config.get("has_k01", False):
                k01_sheet_name = k01_config.get("sheet_name", "")
                if k01_sheet_name and k01_sheet_name in wb_new.sheetnames and is_visible_worksheet(wb_prior_values, k01_sheet_name):
                    ws_new_k01 = wb_new[k01_sheet_name]
                    ws_prior_k01 = wb_prior_values[k01_sheet_name]
                    process_k01(ws_prior_k01, ws_new_k01, k01_config)

            # 6.1 处理非标准L1 LAR/LRA后推明细表
            if subject_code == "L1" and is_visible_worksheet(wb_prior_values, "后推明细表"):
                lead_sheet_name = lead_config.get("sheet_name", "")
                k01_sheet_name = k01_config.get("sheet_name", "")
                if lead_sheet_name in wb_new.sheetnames and k01_sheet_name in wb_new.sheetnames:
                    process_l1_from_rollforward_schedule(
                        wb_prior_values["后推明细表"],
                        wb_new[lead_sheet_name],
                        wb_new[k01_sheet_name],
                        subject_config
                    )

            if subject_code == "L1":
                prior_l103_sheet = find_visible_sheet_name(wb_prior_formula, lambda name: "L1.03" in name)
                new_l103_sheet = next((s for s in wb_new.sheetnames if "L1.03" in s), None)
                if prior_l103_sheet and new_l103_sheet:
                    process_l103_policy_table(
                        wb_prior_formula[prior_l103_sheet],
                        wb_new[new_l103_sheet]
                    )

            if subject_code == "K1":
                prior_k033_sheet = find_visible_sheet_name(wb_prior_values, lambda name: "K.03.3" in name)
                new_k033_sheet = next((s for s in wb_new.sheetnames if "K.03.3" in s), None)
                if prior_k033_sheet and new_k033_sheet:
                    process_k033_depreciation_policy(
                        wb_prior_values[prior_k033_sheet],
                        wb_new[new_k033_sheet]
                    )

            if subject_code == "C":
                prior_c_bkd_sheet = find_visible_sheet_name(
                    wb_prior_values, lambda name: normalize_text(name) == "C.00BKD"
                )
                new_c_bkd_sheet = next((s for s in wb_new.sheetnames if normalize_text(s) == "C.00BKD"), None)
                if prior_c_bkd_sheet and new_c_bkd_sheet:
                    process_c_bkd_basic_info(
                        wb_prior_values[prior_c_bkd_sheet],
                        wb_new[new_c_bkd_sheet],
                    )

            if subject_code == "L2":
                prior_l2_bkd_sheet = find_visible_l2_bkd_sheet_name(wb_prior_values)
                new_l2_bkd_sheet = next((s for s in wb_new.sheetnames if "L2.01.1" in s), None)
                if prior_l2_bkd_sheet and new_l2_bkd_sheet:
                    process_l2_bkd(
                        wb_prior_values[prior_l2_bkd_sheet],
                        wb_new[new_l2_bkd_sheet]
                    )

            if subject_code == "Q1":
                prior_q1_bkd_sheet = find_visible_sheet_name(wb_prior_values, lambda name: "Q1.01" in name)
                new_q1_bkd_sheet = next((s for s in wb_new.sheetnames if "Q1.01" in s), None)
                if prior_q1_bkd_sheet and new_q1_bkd_sheet:
                    process_q1_bkd(
                        wb_prior_values[prior_q1_bkd_sheet],
                        wb_new[new_q1_bkd_sheet]
                    )
                if roll_forward_wording:
                    prior_q1_covenant_sheet = find_visible_sheet_name(
                        wb_prior_formula, lambda name: "Q1.05" in name
                    )
                    new_q1_covenant_sheet = next((s for s in wb_new.sheetnames if "Q1.05" in s), None)
                    if prior_q1_covenant_sheet and new_q1_covenant_sheet:
                        copied = process_q1_covenant_sheet(
                            wb_prior_formula[prior_q1_covenant_sheet],
                            wb_new[new_q1_covenant_sheet],
                        )
                        wording_copied_count += copied
                        if copied:
                            wording_touched_sheets.add(new_q1_covenant_sheet)
                            warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

            if subject_code == "N":
                lead_sheet_name = lead_config.get("sheet_name", "")
                if lead_sheet_name in wb_new.sheetnames and is_visible_worksheet(wb_prior_values, lead_sheet_name):
                    process_n_lead_turnover_analysis(
                        wb_prior_values[lead_sheet_name],
                        wb_new[lead_sheet_name]
                    )
                    if roll_forward_wording:
                        copied = process_n_lead_turnover_wording(
                            wb_prior_values[lead_sheet_name],
                            wb_new[lead_sheet_name],
                        )
                        wording_copied_count += copied
                        if copied:
                            wording_touched_sheets.add(lead_sheet_name)

                detail_sheet_name = "N.01.01明细账"
                if (
                    detail_sheet_name in wb_new.sheetnames
                    and is_visible_worksheet(wb_prior_formula, detail_sheet_name)
                    and is_visible_worksheet(wb_prior_values, detail_sheet_name)
                ):
                    ws_prior_lead = wb_prior_formula[lead_sheet_name] if is_visible_worksheet(wb_prior_formula, lead_sheet_name) else None
                    ws_new_lead = wb_new[lead_sheet_name] if lead_sheet_name in wb_new.sheetnames else None
                    process_n_detail_sheet(
                        wb_prior_formula[detail_sheet_name],
                        wb_prior_values[detail_sheet_name],
                        wb_new[detail_sheet_name],
                        bs_date,
                        ws_prior_lead,
                        ws_new_lead,
                        roll_forward_wording=roll_forward_wording
                    )

            # 7. 处理子表（如U_exp的财务费用子表）
            sub_sheets = subject_config.get("sub_sheets", [])
            for sub_sheet in sub_sheets:
                sub_sheet_name = sub_sheet.get("sheet_name", "")
                if sub_sheet_name and sub_sheet_name in wb_new.sheetnames and is_visible_worksheet(wb_prior_values, sub_sheet_name):
                    ws_new_sub = wb_new[sub_sheet_name]
                    ws_prior_sub = wb_prior_values[sub_sheet_name]

                    if sub_sheet.get("dynamic_prior_current_to_py", False):
                        _, total_row = process_expense_bkd_prior_current_to_py(ws_prior_sub, ws_new_sub)
                        lead_sheet_name = lead_config.get("sheet_name", "")
                        if lead_sheet_name in wb_new.sheetnames:
                            update_lead_bkd_total_references(wb_new[lead_sheet_name], sub_sheet_name, total_row)
                        continue

                    # 查找表头行
                    header_search_text = sub_sheet.get("header_search_text", "期末审定数")
                    header_row = find_header_row(ws_prior_sub, header_search_text, (1, 80))

                    if header_row:
                        closing_col = sub_sheet.get("closing_col", 9)
                        opening_col = sub_sheet.get("opening_col", 10)

                        for row in range(header_row + 1, ws_prior_sub.max_row + 1):
                            prior_cell = ws_prior_sub.cell(row=row, column=closing_col)
                            new_cell = ws_new_sub.cell(row=row, column=opening_col)

                            if prior_cell.value is not None:
                                new_cell.value = prior_cell.value

            if subject_code == "UexpVCVD" and roll_forward_wording:
                for bkd_sheet_name in ("VC.00 销售费用BKD", "VD.00 管理费用BKD"):
                    if is_visible_worksheet(wb_prior_formula, bkd_sheet_name) and bkd_sheet_name in wb_new.sheetnames:
                        copied = process_vcvd_bkd_wording(
                            wb_prior_formula[bkd_sheet_name],
                            wb_new[bkd_sheet_name],
                        )
                        wording_copied_count += copied
                        if copied:
                            wording_touched_sheets.add(bkd_sheet_name)
                            warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

                prior_cutoff_sheet = find_visible_sheet_name(
                    wb_prior_formula, lambda name: "01.4" in name or "截止" in name
                )
                new_cutoff_sheet = next((s for s in wb_new.sheetnames if "01.4" in s or "截止" in s), None)
                if prior_cutoff_sheet and new_cutoff_sheet:
                    copied = process_vcvd_cutoff_table2(
                        wb_prior_formula[prior_cutoff_sheet],
                        wb_new[new_cutoff_sheet],
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.add(new_cutoff_sheet)
                        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

            if subject_code == "Uexp" and roll_forward_wording:
                lead_sheet_name = lead_config.get("sheet_name", "")
                if is_visible_worksheet(wb_prior_formula, lead_sheet_name) and lead_sheet_name in wb_new.sheetnames:
                    copied = process_uexp_lead_wording(
                        wb_prior_formula[lead_sheet_name],
                        wb_new[lead_sheet_name],
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.add(lead_sheet_name)
                        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

                finance_bkd_sheet = next((s for s in wb_new.sheetnames if "BKD" in s and "财务" in s), None)
                prior_finance_bkd_sheet = find_visible_sheet_name(
                    wb_prior_formula, lambda name: "BKD" in name and "财务" in name
                )
                if finance_bkd_sheet and prior_finance_bkd_sheet:
                    copied = process_uexp_finance_bkd_wording(
                        wb_prior_formula[prior_finance_bkd_sheet],
                        wb_new[finance_bkd_sheet],
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.add(finance_bkd_sheet)
                        warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")

            if roll_forward_wording:
                copied, touched_sheets = process_wording_sections(
                    wb_prior_formula,
                    wb_prior_values,
                    wb_new,
                    subject_code,
                    subject_config,
                    warnings_list,
                )
                wording_copied_count += copied
                wording_touched_sheets.update(touched_sheets)
                if subject_code == "L1":
                    copied = process_l1_wording_sections(
                        wb_prior_formula,
                        wb_new,
                        warnings_list,
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.add("L1.00 Lead sheet")
                    l1_lead_name = next(
                        (
                            name
                            for name in wb_new.sheetnames
                            if normalize_text(name).lower().startswith("l1.00lead")
                        ),
                        None,
                    )
                    if l1_lead_name:
                        prior_l1_lead_name = find_visible_sheet_name(
                            wb_prior_formula,
                            lambda name: normalize_text(name).lower().startswith("l1.00lead"),
                        )
                        if prior_l1_lead_name:
                            repair_l1_lead_detail_formulas(
                                wb_prior_formula[prior_l1_lead_name],
                                wb_new[l1_lead_name],
                            )
                        rebuild_l1_lead_total_formulas(wb_new[l1_lead_name])
                if subject_code == "C":
                    prior_cutoff_sheet = find_visible_sheet_name(
                        wb_prior_formula,
                        lambda name: normalize_text(name).upper().startswith("C.03CUTOFF"),
                    )
                    new_cutoff_sheet = next((s for s in wb_new.sheetnames if normalize_text(s).upper().startswith("C.03CUTOFF")), None)
                    if prior_cutoff_sheet and new_cutoff_sheet and prior_cutoff_sheet in wb_prior_values.sheetnames:
                        copied = process_c_cutoff_wording(
                            wb_prior_formula[prior_cutoff_sheet],
                            wb_prior_values[prior_cutoff_sheet],
                            wb_new[new_cutoff_sheet],
                        )
                        wording_copied_count += copied
                        if copied:
                            wording_touched_sheets.add(new_cutoff_sheet)
                            warnings_list.append("已 roll forward wording，请项目组更新黄色标注区域")
                if subject_code == "J1":
                    copied = process_j1_wording_sections(
                        wb_prior_formula,
                        wb_prior_values,
                        wb_new,
                        warnings_list,
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.update(["J.00  Lead Sheet", "J.01 Agree SL to GL", "J.03"])
                if subject_code == "L2":
                    copied = process_l2_wording_sections(
                        wb_prior_formula,
                        wb_prior_values,
                        wb_new,
                        warnings_list,
                    )
                    wording_copied_count += copied
                    if copied:
                        wording_touched_sheets.update(["L2.00 Lead", "L2.01.1 BKD"])

            # 8. 保存
            if cra_records:
                try:
                    from cra_support import apply_cra_records_to_workbook
                    cra_result = apply_cra_records_to_workbook(wb_new, subject_code, cra_records)
                    for message in cra_result.get("messages", []):
                        warnings_list.append(message)
                except Exception as exc:
                    warnings_list.append(f"CRA写入失败，请手工复核: {exc}")

            wb_new.calculation.fullCalcOnLoad = True
            wb_new.calculation.forceFullCalc = True
            wb_new.calculation.calcMode = "auto"

            # 生成警告消息
            warning_msg = ""
            if warnings_list:
                warnings_list[:] = list(dict.fromkeys(warnings_list))
                warning_msg = "; ".join(warnings_list)

            if generate_summary:
                add_roll_forward_summary_sheet(
                    wb_new,
                    subject_code,
                    subject_config.get("name", ""),
                    company_name,
                    bs_date,
                    prior_path,
                    output_path,
                    warnings_list,
                    {
                        "roll_wording": roll_forward_wording,
                        "generate_summary": generate_summary,
                    },
                    wording_copied_count,
                    sorted(wording_touched_sheets),
                    before_snapshot,
                )

            if progress_callback:
                progress_callback(f"正在保存输出文件: {os.path.basename(output_path)}")
            wb_new.save(output_path)
            restore_template_drawing_parts(template_path, output_path)

            if subject_code == "Q1":
                try:
                    q105_row_shifts = []
                    prior_q105_name = find_visible_sheet_name(
                        wb_prior_formula, lambda name: name.startswith("Q1.05")
                    )
                    output_q105_name = next(
                        (name for name in wb_new.sheetnames if name.startswith("Q1.05")),
                        None,
                    )
                    if prior_q105_name and output_q105_name:
                        q105_row_shifts = get_q1_covenant_image_row_shifts(
                            wb_prior_formula[prior_q105_name],
                            wb_new[output_q105_name],
                        )
                    copy_q1_review_images(prior_path, output_path, q105_row_shifts)
                except Exception as exc:
                    warnings_list.append(f"Q1 Note/限制性契约图片复制失败，请手工复核: {exc}")
                    warning_msg = "; ".join(dict.fromkeys(warnings_list))

            wb_new.close()
            if wb_prior_formula is not wb_prior_values:
                wb_prior_formula.close()
            wb_prior_values.close()
            os.replace(work_output_path, final_output_path)
            work_output_path = None
            return True, f"处理成功{(' - ' + warning_msg if warning_msg else '')}", final_output_path, warnings_list

        finally:
            wb_new.close()
            if wb_prior_formula is not wb_prior_values:
                wb_prior_formula.close()
            wb_prior_values.close()

    except Exception as e:
        if work_output_path:
            try:
                os.remove(work_output_path)
            except OSError:
                pass
        return False, f"处理失败: {str(e)}", None, warnings_list


def process_multiple_subjects(subject_codes, template_dir, prior_dir, pmte_path,
                              company_name, bs_date, output_dir, config_path=None,
                              functional_currency=None, accounting_standard=None,
                              pm_value=None, te_value=None, sad_value=None, cra_records=None,
                              roll_forward_wording=False, generate_summary=True,
                              llm_enhanced=False, llm_wording_revision=False,
                              llm_options=None, progress_callback=None,
                              control_callback=None):
    """
    批量处理多个科目
    """
    config_manager = SubjectConfig(config_path)
    results = []

    # 计算上年日期
    prior_year = str(int(bs_date[:4]) - 1)

    total = len(subject_codes)
    for index, subject_code in enumerate(subject_codes, start=1):
        if control_callback:
            decision = control_callback("before_subject", subject_code, index, total)
            if decision == "terminate":
                break
        def emit(message):
            if progress_callback:
                progress_callback(index - 1, total, f"[{subject_code}] {message}")

        emit("正在检查配置和模板")
        subject_config = config_manager.get_subject(subject_code)
        if not subject_config:
            results.append((subject_code, False, "找不到科目配置", None, []))
            continue

        # 查找标准模板
        template_file = subject_config.get("template_file", "")
        template_path = os.path.join(template_dir, template_file)
        if not os.path.exists(template_path):
            results.append((subject_code, False, f"找不到标准模板: {template_file}", None, []))
            continue

        # 查找上年底稿
        emit("正在查找上年底稿")
        prior_path = find_prior_file(prior_dir, subject_code, prior_year, subject_config)
        if not prior_path:
            results.append((subject_code, False, f"找不到上年底稿: {subject_code}", None, []))
            continue
        prior_size = os.path.getsize(prior_path) / (1024 * 1024)
        emit(f"已匹配上年底稿: {os.path.basename(prior_path)} ({prior_size:.1f} MB)")

        # 处理单个科目
        emit("开始处理")
        success, message, output_path, warnings_list = process_single_subject(
            subject_code, template_path, prior_path, pmte_path,
            company_name, bs_date, output_dir, subject_config,
            functional_currency=functional_currency,
            accounting_standard=accounting_standard,
            pm_value=pm_value,
            te_value=te_value,
            sad_value=sad_value,
            cra_records=cra_records,
            roll_forward_wording=roll_forward_wording,
            generate_summary=generate_summary,
            llm_enhanced=llm_enhanced,
            llm_wording_revision=llm_wording_revision,
            llm_options=llm_options,
            progress_callback=lambda msg, code=subject_code: progress_callback(index - 1, total, f"[{code}] {msg}") if progress_callback else None,
        )

        results.append((subject_code, success, message, output_path, warnings_list))
        if progress_callback:
            status = "成功" if success else "失败"
            progress_callback(index, total, f"[{status}] {subject_code}: {message}")
        if control_callback:
            decision = control_callback("after_subject", subject_code, index, total)
            if decision == "terminate":
                break

    return results


if __name__ == "__main__":
    # 测试代码
    print("Roll Forward Core Module v4.0")
    print("请通过GUI或命令行调用此模块")
