import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import os
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import numpy as np
import traceback
import re

# 全局配置
DATE_FORMAT = '%Y-%m-%d'
DATE_TIME_FORMAT = f'{DATE_FORMAT} %H:%M:%S'


def resource_path(relative_path):
    """解决打包后资源路径问题"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def select_input_file(root=None):
    """使用Tkinter对话框选择输入文件"""
    need_cleanup = root is None
    if need_cleanup:
        root = tk.Tk()
        root.withdraw()
    try:
        print("=== 调试：开始文件选择 ===")
        root.attributes('-topmost', True)
        file_path = filedialog.askopenfilename(
            title="选择函证列表Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("Excel文件", "*.xls")],
            initialdir=os.getcwd()
        )
        return file_path
    except Exception as e:
        print(f"文件选择错误: {e}")
        return None
    finally:
        if need_cleanup:
            root.destroy()


def auto_generate_output_path(input_path, file_type):
    """自动生成输出文件路径"""
    try:
        input_path = Path(input_path)
        output_dir = input_path.parent / "函证统计结果"
        output_dir.mkdir(exist_ok=True)

        file_stem = input_path.stem
        current_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        output_filename = f"{file_stem}_{file_type}_进度报告_{current_date}.xlsx"
        output_path = output_dir / output_filename
        print(f"=== 调试：生成输出路径 = {output_path} ===")
        return output_path
    except Exception as e:
        print(f"生成输出路径错误: {e}")
        return None


def add_total_row(summary_df, is_bank=True, group_by="发函单位名称"):
    """添加合计行 - 使用concat替代已弃用的append"""
    try:
        if summary_df is None or len(summary_df) == 0:
            return summary_df

        # 确定分组列名
        group_column = "项目名称" if group_by == "项目名称" else "发函单位名称"

        total_row = {group_column: '合计'}

        if is_bank:
            numeric_columns = ['函证总数', '银行函证', '纸质', '电子', '格式一', '格式二', '其他模版',
                               '未发出', '已发出', '已签收（纸质）', '已回函', '回函相符', '回函不符']
        else:
            numeric_columns = ['往来总数', '标准往来', '其他函证', '未发出', '已发出', '已签收（纸质）',
                               '已回函', '回函相符', '回函不符']

        for col in numeric_columns:
            if col in summary_df.columns:
                total_row[col] = summary_df[col].sum()
            else:
                total_row[col] = 0

        # 计算百分比
        if is_bank:
            denominator = total_row.get('银行函证', 0)
        else:
            denominator = total_row.get('往来总数', 0)

        if denominator > 0:
            total_row['发函率'] = round(total_row.get('已发出', 0) / denominator * 100, 2)
            total_row['回函率'] = round(total_row.get('已回函', 0) / denominator * 100, 2)
        else:
            total_row['发函率'] = 0.0
            total_row['回函率'] = 0.0

        if total_row.get('已回函', 0) > 0:
            total_row['相符率'] = round(total_row.get('回函相符', 0) / total_row.get('已回函', 0) * 100, 2)
        else:
            total_row['相符率'] = 0.0

        # 使用concat替代append
        total_df = pd.DataFrame([total_row])
        # 确保列顺序一致
        for col in summary_df.columns:
            if col not in total_df.columns:
                total_df[col] = 0
        total_df = total_df[summary_df.columns]

        result_df = pd.concat([summary_df, total_df], ignore_index=True)
        print("=== 调试：合计行添加完成 ===")
        return result_df
    except Exception as e:
        print(f"添加合计行错误: {e}")
        return summary_df


def calculate_optimal_column_width(worksheet, df, column_index=1):
    """计算最优列宽 - 根据A列内容自适应调整"""
    try:
        if df is None or len(df) == 0:
            return 50

        # 获取A列的所有文本内容
        texts = []

        # 添加表头文本
        if df.columns[0] in ['发函单位名称', '项目名称']:
            texts.append(df.columns[0])

        # 添加数据行文本
        if column_index == 1:  # A列
            first_col_values = df.iloc[:, 0].dropna().astype(str).tolist()
            texts.extend(first_col_values)

        # 计算最大宽度
        max_width = 0
        for text in texts:
            if pd.notna(text):
                # 计算文本宽度：中文字符算2，英文字符算1
                width = sum(2 if ord(char) > 127 else 1 for char in str(text))
                max_width = max(max_width, width)

        # 添加边距并设置合理的宽度范围
        optimal_width = max(30, min(100, max_width + 6))
        print(f"=== 调试：计算出的A列最优宽度 = {optimal_width}，基于最长文本长度 {max_width} ===")
        return optimal_width
    except Exception as e:
        print(f"计算列宽错误: {e}")
        return 50


def safe_cell_write(worksheet, row, col, value):
    """安全写入单元格 - 避免合并单元格冲突"""
    try:
        cell = worksheet.cell(row=row, column=col)

        # 检查是否是合并单元格的非左上角部分
        for merged_range in worksheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                # 如果是合并单元格且不是左上角，跳过写入
                if not (row == merged_range.min_row and col == merged_range.min_col):
                    return False

        # 安全写入
        cell.value = value
        return True

    except Exception as e:
        print(f"写入单元格({row}, {col})错误: {e}")
        return False


def safe_merge_cells(worksheet, start_row, start_col, end_row, end_col, value=None):
    """安全合并单元格 - 先写入值再合并"""
    try:
        # 如果有值，先写入左上角单元格
        if value is not None:
            worksheet.cell(row=start_row, column=start_col, value=value)

        # 然后合并单元格
        worksheet.merge_cells(start_row=start_row, start_column=start_col,
                              end_row=end_row, end_column=end_col)
        return True
    except Exception as e:
        print(f"合并单元格错误: {e}")
        return False


def create_beautified_worksheet(worksheet, df, title, is_bank=True, group_by="发函单位名称"):
    """创建美化的工作表 - 确保A列显示为文本"""
    try:
        print("=== 调试：开始创建工作表美化 ===")

        if df is None or len(df) == 0:
            # 添加空数据提示
            worksheet.append(["无有效数据"])
            return

        # 清空现有内容
        worksheet.delete_rows(1, worksheet.max_row)

        nrows, ncols = len(df), len(df.columns)
        headers = list(df.columns)
        max_data_row = nrows + 3  # 标题行+表头行+数据行

        # ==================== 1. 计算并设置自适应列宽 ====================
        optimal_a_width = calculate_optimal_column_width(worksheet, df, 1)

        # 设置列宽
        column_widths = {
            'A': optimal_a_width,
            'B': 12, 'C': 12, 'D': 12, 'E': 12,
            'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
            'K': 15, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12
        } if is_bank else {
            'A': optimal_a_width,
            'B': 12, 'C': 12, 'D': 12, 'E': 12,
            'F': 12, 'G': 15, 'H': 12, 'I': 12, 'J': 12,
            'K': 12, 'L': 12, 'M': 12
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # ==================== 2. 设置行高 ====================
        for row in range(1, max_data_row + 1):
            worksheet.row_dimensions[row].height = 25
            if row == 1:
                worksheet.row_dimensions[row].height = 35
            elif row in [2, 3]:
                worksheet.row_dimensions[row].height = 30

        # ==================== 3. 定义样式常量 ====================
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 字体定义
        title_font = Font(name='微软雅黑', bold=True, size=16, color='FFFFFF')
        group_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        data_font = Font(name='微软雅黑', size=10)
        total_font = Font(name='微软雅黑', bold=True, size=11)

        # 填充色定义
        title_fill = PatternFill(start_color='2E5B8F', end_color='2E5B8F', fill_type='solid')
        group_fill = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        even_row_fill = PatternFill(start_color='F0F7FF', end_color='F0F7FF', fill_type='solid')
        odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        total_fill = PatternFill(start_color='FFE599', end_color='FFE599', fill_type='solid')

        # 对齐方式
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # ==================== 4. 创建表头结构 ====================

        # 第一行：主标题
        if is_bank:
            safe_merge_cells(worksheet, 1, 1, 1, 17, title)  # A1:Q1
        else:
            safe_merge_cells(worksheet, 1, 1, 1, 13, title)  # A1:M1

        title_cell = worksheet.cell(1, 1)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        title_cell.border = thin_border

        # 第二行：逻辑分组标题
        if is_bank:
            # 银行函证逻辑分组
            groups = [
                ("函证类别", 2, 3),  # B2:C2
                ("函证类型", 4, 5),  # D2:E2
                ("发函模版", 6, 8),  # F2:H2
                ("函证进度", 9, 12),  # I2:L2
                ("回函结果", 13, 14),  # M2:N2
                ("进度百分比", 15, 17)  # O2:Q2
            ]
        else:
            # 往来函证逻辑分组
            groups = [
                ("函证类型", 2, 4),  # B2:D2
                ("函证进度", 5, 8),  # E2:H2
                ("回函结果", 9, 10),  # I2:J2
                ("进度百分比", 11, 13)  # K2:M2
            ]

        # A2:A3 合并（发函单位名称或项目名称）
        group_label = "项目名称" if group_by == "项目名称" else "发函单位名称"
        safe_merge_cells(worksheet, 2, 1, 3, 1, group_label)
        cell_a2 = worksheet.cell(2, 1)
        cell_a2.font = header_font
        cell_a2.fill = header_fill
        cell_a2.alignment = center_align
        cell_a2.border = thin_border

        # 设置逻辑分组标题
        for group_name, start_col, end_col in groups:
            safe_merge_cells(worksheet, 2, start_col, 2, end_col, group_name)
            group_cell = worksheet.cell(2, start_col)
            group_cell.font = group_font
            group_cell.fill = group_fill
            group_cell.alignment = center_align
            group_cell.border = thin_border

        # 第三行：列标题
        for col_idx, header in enumerate(headers, 1):
            safe_cell_write(worksheet, 3, col_idx, header)
            cell = worksheet.cell(3, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # ==================== 5. 写入数据 ====================
        for row_idx, row_data in df.iterrows():
            for col_idx, col_name in enumerate(headers, 1):
                value = row_data[col_name]

                # 关键修复：确保A列（第一列）显示为文本格式
                if col_idx == 1:
                    # 强制转换为字符串，保留原始值
                    if pd.isna(value):
                        value = ""
                    else:
                        value = str(value)

                    # 设置单元格格式为文本
                    cell = worksheet.cell(row=row_idx + 4, column=col_idx)
                    cell.number_format = '@'  # 设置为文本格式

                if safe_cell_write(worksheet, row_idx + 4, col_idx, value):
                    cell = worksheet.cell(row_idx + 4, col_idx)
                    cell.border = thin_border
                    cell.font = data_font

                    # 设置交替行颜色
                    if row_idx % 2 == 0:
                        cell.fill = even_row_fill
                    else:
                        cell.fill = odd_row_fill

                    # 设置对齐方式
                    if col_idx == 1:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

        # ==================== 6. 设置合计行样式 ====================
        if nrows > 0:
            total_row = nrows + 3
            for col_idx in range(1, len(headers) + 1):
                cell = worksheet.cell(total_row, col_idx)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = center_align
                cell.border = thin_border

                # 确保合计行的第一列也是文本格式
                if col_idx == 1:
                    cell.number_format = '@'

        # ==================== 7. 修复边框覆盖问题 ====================
        for row in range(1, max_data_row + 1):
            for col in range(1, len(headers) + 1):
                try:
                    cell = worksheet.cell(row, col)
                    cell.border = thin_border
                except Exception as e:
                    print(f"设置边框错误({row}, {col}): {e}")

        # ==================== 8. 添加进度条可视化 ====================
        try:
            if is_bank:
                # 银行函证百分比列
                percent_cols = [
                    (15, '发函率', '63C384'),  # O列
                    (16, '回函率', '4A90E2'),  # P列
                    (17, '相符率', 'FF6B6B')  # Q列
                ]
            else:
                # 往来函证百分比列
                percent_cols = [
                    (11, '发函率', '63C384'),  # K列
                    (12, '回函率', '4A90E2'),  # L列
                    (13, '相符率', 'FF6B6B')  # M列
                ]

            for col_idx, col_name, color in percent_cols:
                if col_idx <= len(headers):
                    data_range = f'{get_column_letter(col_idx)}4:{get_column_letter(col_idx)}{nrows + 3}'
                    worksheet.conditional_formatting.add(
                        data_range,
                        DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=100,
                            color=color, showValue=True
                        )
                    )
        except Exception as e:
            print(f"进度条设置警告: {e}")

        print("=== 调试：工作表美化完成 ===")

    except Exception as e:
        print(f"工作表美化错误: {e}")
        traceback.print_exc()


def validate_date_time(date_str):
    """验证日期时间格式"""
    if pd.isna(date_str):
        return False
    try:
        stripped = date_str.strip()
        if re.fullmatch(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', stripped):
            pd.to_datetime(stripped, format=DATE_TIME_FORMAT)
            return True
        return False
    except:
        return False


def aggregate_bank_data(df, group_column='发函单位名称'):
    """银行函证数据聚合函数 - 确保分组列保持原始类型"""
    # 关键修复：保持分组列的原始值不变
    result = df.groupby(group_column, as_index=False).agg(
        函证总数=('函证编号', 'count'),
        银行函证=('函证类型', lambda x: x.isin(['银行', '银行-电子函证']).sum()),
        纸质=('函证类型', lambda x: (x == '银行').sum()),
        电子=('函证类型', lambda x: (x == '银行-电子函证').sum()),
        格式一=('发函模版', lambda x: x.str.contains('格式一', na=False).sum()),
        格式二=('发函模版', lambda x: x.str.contains('格式二', na=False).sum()),
        其他模版=('发函模版', lambda x: x.notna().sum() -
                                        (x.str.contains('格式一', na=False).sum() +
                                         x.str.contains('格式二', na=False).sum())),
        未发出=('函证状态', lambda x: (x == '未发出').sum()),
        已发出=('函证状态', lambda x: x.isin(['已发出', '已收回']).sum()),
        已签收_纸质=('发函签收时间', lambda x:
        x.apply(lambda y: 1 if validate_date_time(y) else 0).sum() if '发函签收时间' in df.columns else 0),
        已回函=('函证状态', lambda x: (x == '已收回').sum()),
        回函相符=('询证项回函结果', lambda x: (x == '相符').sum()),
        回函不符=('询证项回函结果', lambda x: x.str.contains('不符', na=False).sum())
    )

    return result


def aggregate_trade_data(df, group_column='发函单位名称'):
    """往来函证数据聚合函数 - 确保分组列保持原始类型"""
    # 关键修复：保持分组列的原始值不变
    result = df.groupby(group_column, as_index=False).agg(
        往来总数=('函证编号', 'count'),
        标准往来=('函证类型', lambda x: x.str.contains('往来', na=False).sum()),
        其他函证=('函证类型', lambda x: x.str.contains('其他', na=False).sum()),
        未发出=('函证状态', lambda x: (x == '未发出').sum()),
        已发出=('函证状态', lambda x: x.isin(['已发出', '已收回']).sum()),
        已签收_纸质=('发函签收时间', lambda x:
        x.apply(lambda y: 1 if validate_date_time(y) else 0).sum() if '发函签收时间' in df.columns else 0),
        已回函=('函证状态', lambda x: (x == '已收回').sum()),
        回函相符=('询证项回函结果', lambda x: (x == '相符').sum() if '询证项回函结果' in df.columns else 0),
        回函不符=('询证项回函结果',
                  lambda x: x.str.contains('不符', na=False).sum() if '询证项回函结果' in df.columns else 0)
    )

    return result


def process_bank_confirmation(input_path):
    """处理银行函证统计 - 确保项目名称保持原始格式"""
    try:
        print("=" * 70)
        print("开始处理银行函证统计...")

        # 读取数据
        df = pd.read_excel(input_path, dtype=str)
        print(f"数据读取完成，共{len(df)}行")

        # 筛选银行函证
        bank_df = df[df['函证类型'].isin(['银行', '银行-电子函证'])].copy()
        print(f"筛选出银行函证{len(bank_df)}行")

        # 处理空数据场景
        if len(bank_df) == 0:
            print("警告: 未找到银行函证数据，将跳过报表生成")
            return None, None

        # 处理基准日
        bank_df['函证基准日'] = pd.to_datetime(bank_df['函证基准日'], errors='coerce').dt.strftime(DATE_FORMAT)
        base_dates = [bd for bd in bank_df['函证基准日'].unique() if pd.notna(bd)]
        print(f"发现{len(base_dates)}个基准日: {', '.join(base_dates)}")

        # 按发函单位名称汇总统计（所有基准日）
        unit_summary = aggregate_bank_data(bank_df, '发函单位名称')
        unit_summary = unit_summary.rename(columns={'已签收_纸质': '已签收（纸质）'})

        # 计算百分比
        unit_summary['发函率'] = unit_summary.apply(
            lambda row: round(row['已发出'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
        unit_summary['回函率'] = unit_summary.apply(
            lambda row: round(row['已回函'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
        unit_summary['相符率'] = unit_summary.apply(
            lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

        # 调整列顺序
        column_order = [
            '发函单位名称', '函证总数', '银行函证', '纸质', '电子',
            '格式一', '格式二', '其他模版', '未发出', '已发出',
            '已签收（纸质）', '已回函', '回函相符', '回函不符',
            '发函率', '回函率', '相符率'
        ]
        unit_summary = unit_summary[[col for col in column_order if col in unit_summary.columns]]
        unit_summary_with_total = add_total_row(unit_summary.copy(), is_bank=True, group_by="发函单位名称")

        # 按项目名称汇总统计（所有基准日）
        project_summary = None
        project_summary_with_total = None
        if '项目名称' in bank_df.columns:
            # 关键修复：使用与发函单位名称相同的方法处理项目名称
            project_summary = aggregate_bank_data(bank_df, '项目名称')
            project_summary = project_summary.rename(columns={
                '已签收_纸质': '已签收（纸质）',
                '发函单位名称': '项目名称'  # 重命名分组列
            })

            # 计算百分比
            project_summary['发函率'] = project_summary.apply(
                lambda row: round(row['已发出'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
            project_summary['回函率'] = project_summary.apply(
                lambda row: round(row['已回函'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
            project_summary['相符率'] = project_summary.apply(
                lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

            # 关键修复：确保列顺序正确，项目名称在第一列
            # 创建新的列顺序，确保项目名称在第一列
            new_column_order = ['项目名称']  # 项目名称始终在第一列
            for col in column_order:
                if col != '发函单位名称' and col in project_summary.columns:
                    new_column_order.append(col)

            project_summary = project_summary[new_column_order]
            project_summary_with_total = add_total_row(project_summary.copy(), is_bank=True, group_by="项目名称")
            print("按项目名称统计完成")
        else:
            print("警告：数据中未找到'项目名称'列，跳过按项目名称统计")

        # 生成报表
        output_path = auto_generate_output_path(input_path, "银行函证")
        if output_path is None:
            return None, None

        wb = Workbook()
        # 删除默认创建的工作表
        wb.remove(wb.active)

        # 1. 按项目名称汇总统计（所有基准日）- 作为第一个sheet
        if project_summary_with_total is not None:
            print("生成按项目名称汇总统计表（作为第一个sheet）...")
            ws_project = wb.create_sheet("按项目名称汇总", 0)  # 插入到第一个位置
            title_project = f'银行函证进度统计报告 - 按项目名称汇总 - {datetime.now().strftime("%Y年%m月%d日")}'
            create_beautified_worksheet(ws_project, project_summary_with_total, title_project, is_bank=True,
                                        group_by="项目名称")

        # 2. 按发函单位汇总统计（所有基准日）
        print("生成按发函单位汇总统计表...")
        ws_unit = wb.create_sheet("按发函单位汇总")
        title_unit = f'银行函证进度统计报告 - 按发函单位汇总 - {datetime.now().strftime("%Y年%m月%d日")}'
        create_beautified_worksheet(ws_unit, unit_summary_with_total, title_unit, is_bank=True, group_by="发函单位名称")

        # 3. 分基准日统计（按发函单位和按项目名称）
        print("生成分基准日统计表...")
        for base_date in base_dates:
            base_date_df = bank_df[bank_df['函证基准日'] == base_date]
            print(f"正在处理基准日: {base_date}，共{len(base_date_df)}行数据")

            if len(base_date_df) == 0:
                continue

            # 按发函单位汇总
            base_date_summary_unit = aggregate_bank_data(base_date_df, '发函单位名称')
            base_date_summary_unit = base_date_summary_unit.rename(columns={'已签收_纸质': '已签收（纸质）'})

            # 计算百分比
            base_date_summary_unit['发函率'] = base_date_summary_unit.apply(
                lambda row: round(row['已发出'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
            base_date_summary_unit['回函率'] = base_date_summary_unit.apply(
                lambda row: round(row['已回函'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
            base_date_summary_unit['相符率'] = base_date_summary_unit.apply(
                lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

            base_date_summary_unit = base_date_summary_unit[
                [col for col in column_order if col in base_date_summary_unit.columns]]
            base_date_summary_unit_with_total = add_total_row(base_date_summary_unit.copy(), is_bank=True,
                                                              group_by="发函单位名称")

            # 创建工作表名称（按发函单位）
            sheet_name_unit = f"基准日_{base_date}_按发函单位"
            if len(sheet_name_unit) > 31:
                sheet_name_unit = sheet_name_unit[:31]

            ws_unit_detail = wb.create_sheet(sheet_name_unit)
            title_unit_detail = f'银行函证进度统计 - 基准日{base_date}（按发函单位） - {datetime.now().strftime("%Y年%m月%d日")}'
            create_beautified_worksheet(ws_unit_detail, base_date_summary_unit_with_total, title_unit_detail,
                                        is_bank=True, group_by="发函单位名称")

            # 按项目名称汇总（如果存在项目名称列）
            if '项目名称' in base_date_df.columns:
                base_date_summary_project = aggregate_bank_data(base_date_df, '项目名称')
                base_date_summary_project = base_date_summary_project.rename(columns={
                    '已签收_纸质': '已签收（纸质）',
                    '发函单位名称': '项目名称'  # 重命名分组列
                })

                # 计算百分比
                base_date_summary_project['发函率'] = base_date_summary_project.apply(
                    lambda row: round(row['已发出'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
                base_date_summary_project['回函率'] = base_date_summary_project.apply(
                    lambda row: round(row['已回函'] / row['银行函证'] * 100, 2) if row['银行函证'] > 0 else 0.0, axis=1)
                base_date_summary_project['相符率'] = base_date_summary_project.apply(
                    lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

                # 关键修复：确保列顺序正确，项目名称在第一列
                new_column_order = ['项目名称']  # 项目名称始终在第一列
                for col in column_order:
                    if col != '发函单位名称' and col in base_date_summary_project.columns:
                        new_column_order.append(col)

                base_date_summary_project = base_date_summary_project[new_column_order]
                base_date_summary_project_with_total = add_total_row(base_date_summary_project.copy(), is_bank=True,
                                                                     group_by="项目名称")

                # 创建工作表名称（按项目名称）
                sheet_name_project = f"基准日_{base_date}_按项目名称"
                if len(sheet_name_project) > 31:
                    sheet_name_project = sheet_name_project[:31]

                ws_project_detail = wb.create_sheet(sheet_name_project)
                title_project_detail = f'银行函证进度统计 - 基准日{base_date}（按项目名称） - {datetime.now().strftime("%Y年%m月%d日")}'
                create_beautified_worksheet(ws_project_detail, base_date_summary_project_with_total,
                                            title_project_detail, is_bank=True, group_by="项目名称")
            else:
                print(f"基准日{base_date}的数据中未找到'项目名称'列，跳过按项目名称汇总")

        # 保存工作簿
        wb.save(output_path)
        print(f"银行函证报表生成完成: {output_path}")

        return unit_summary_with_total, output_path

    except Exception as e:
        print(f"银行函证统计错误: {e}")
        traceback.print_exc()
        return None, None


def process_trade_confirmation(input_path):
    """处理往来函证统计 - 确保项目名称保持原始格式"""
    try:
        print("=" * 70)
        print("开始处理往来函证统计...")

        # 读取数据
        df = pd.read_excel(input_path, dtype=str)
        print(f"数据读取完成，共{len(df)}行")

        # 筛选往来函证
        trade_df = df[~df['函证类型'].isin(['银行', '银行-电子函证'])].copy()
        print(f"筛选出往来函证{len(trade_df)}行")

        # 处理空数据场景
        if len(trade_df) == 0:
            print("警告: 未找到往来函证数据，将跳过报表生成")
            return None, None

        # 按发函单位名称汇总统计
        unit_summary = aggregate_trade_data(trade_df, '发函单位名称')
        unit_summary = unit_summary.rename(columns={'已签收_纸质': '已签收（纸质）'})

        # 处理可能缺失的列
        for col in ['回函相符', '回函不符']:
            if col not in unit_summary.columns:
                unit_summary[col] = 0

        # 计算百分比
        unit_summary['发函率'] = unit_summary.apply(
            lambda row: round(row['已发出'] / row['往来总数'] * 100, 2) if row['往来总数'] > 0 else 0.0, axis=1)
        unit_summary['回函率'] = unit_summary.apply(
            lambda row: round(row['已回函'] / row['往来总数'] * 100, 2) if row['往来总数'] > 0 else 0.0, axis=1)
        unit_summary['相符率'] = unit_summary.apply(
            lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

        # 调整列顺序
        column_order = [
            '发函单位名称', '往来总数', '标准往来', '其他函证',
            '未发出', '已发出', '已签收（纸质）', '已回函',
            '回函相符', '回函不符', '发函率', '回函率', '相符率'
        ]
        unit_summary = unit_summary[[col for col in column_order if col in unit_summary.columns]]
        unit_summary_with_total = add_total_row(unit_summary.copy(), is_bank=False, group_by="发函单位名称")

        # 按项目名称汇总统计
        project_summary = None
        project_summary_with_total = None
        if '项目名称' in trade_df.columns:
            # 关键修复：使用与发函单位名称相同的方法处理项目名称
            project_summary = aggregate_trade_data(trade_df, '项目名称')
            project_summary = project_summary.rename(columns={
                '已签收_纸质': '已签收（纸质）',
                '发函单位名称': '项目名称'  # 重命名分组列
            })

            # 处理可能缺失的列
            for col in ['回函相符', '回函不符']:
                if col not in project_summary.columns:
                    project_summary[col] = 0

            # 计算百分比
            project_summary['发函率'] = project_summary.apply(
                lambda row: round(row['已发出'] / row['往来总数'] * 100, 2) if row['往来总数'] > 0 else 0.0, axis=1)
            project_summary['回函率'] = project_summary.apply(
                lambda row: round(row['已回函'] / row['往来总数'] * 100, 2) if row['往来总数'] > 0 else 0.0, axis=1)
            project_summary['相符率'] = project_summary.apply(
                lambda row: round(row['回函相符'] / row['已回函'] * 100, 2) if row['已回函'] > 0 else 0.0, axis=1)

            # 关键修复：确保列顺序正确，项目名称在第一列
            new_column_order = ['项目名称']  # 项目名称始终在第一列
            for col in column_order:
                if col != '发函单位名称' and col in project_summary.columns:
                    new_column_order.append(col)

            project_summary = project_summary[new_column_order]
            project_summary_with_total = add_total_row(project_summary.copy(), is_bank=False, group_by="项目名称")
            print("按项目名称统计完成")
        else:
            print("警告：数据中未找到'项目名称'列，跳过按项目名称统计")

        # 生成报表
        output_path = auto_generate_output_path(input_path, "往来函证")
        if output_path is None:
            return None, None

        wb = Workbook()
        # 删除默认创建的工作表
        wb.remove(wb.active)

        # 1. 按项目名称汇总统计 - 作为第一个sheet
        if project_summary_with_total is not None:
            print("生成按项目名称汇总统计表（作为第一个sheet）...")
            ws_project = wb.create_sheet("按项目名称汇总", 0)  # 插入到第一个位置
            title_project = f'往来函证进度统计报告 - 按项目名称汇总 - {datetime.now().strftime("%Y年%m月%d日")}'
            create_beautified_worksheet(ws_project, project_summary_with_total, title_project, is_bank=False,
                                        group_by="项目名称")

        # 2. 按发函单位汇总统计
        print("生成按发函单位汇总统计表...")
        ws_unit = wb.create_sheet("按发函单位汇总")
        title_unit = f'往来函证进度统计报告 - 按发函单位汇总 - {datetime.now().strftime("%Y年%m月%d日")}'
        create_beautified_worksheet(ws_unit, unit_summary_with_total, title_unit, is_bank=False,
                                    group_by="发函单位名称")

        # 保存工作簿
        wb.save(output_path)
        print(f"往来函证报表生成完成: {output_path}")

        return unit_summary_with_total, output_path

    except Exception as e:
        print(f"往来函证统计错误: {e}")
        traceback.print_exc()
        return None, None


def main(root=None):
    """主函数 - 增强容错性。root 为 runner 传入时，复用其窗口。"""
    need_mainloop = root is None
    if need_mainloop:
        root = tk.Tk()
        root.withdraw()
    else:
        root.withdraw()
    print("=" * 70)
    print("函证进度统计工具 - 修复项目名称显示问题")
    print("=" * 70)

    try:
        print("=== 调试：程序开始执行 ===")

        # 选择输入文件
        input_path = select_input_file(root)
        if not input_path:
            print("用户取消了文件选择")
            return

        if not Path(input_path).exists():
            raise FileNotFoundError(f"错误: 找不到输入文件: {input_path}")

        print(f"输入文件: {input_path}")

        # 处理银行函证
        bank_result, bank_output = process_bank_confirmation(input_path)
        bank_success = bank_result is not None and bank_output is not None

        # 处理往来函证
        trade_result, trade_output = process_trade_confirmation(input_path)
        trade_success = trade_result is not None and trade_output is not None

        # 准备结果消息
        messages = []
        if bank_success:
            messages.append(f"银行函证报告已生成: {bank_output}")
            messages.append("包含统计维度：")
            if '项目名称' in pd.read_excel(input_path, nrows=1).columns:
                messages.append("- 按项目名称汇总（所有基准日，第一个sheet）")
            else:
                messages.append("- 按项目名称汇总（数据中未找到'项目名称'列）")
            messages.append("- 按发函单位汇总（所有基准日）")
            messages.append("- 分基准日统计（按发函单位和按项目名称）")
        else:
            messages.append("未生成银行函证报告（可能没有银行函证数据）")

        if trade_success:
            messages.append(f"往来函证报告已生成: {trade_output}")
            messages.append("包含统计维度：")
            if '项目名称' in pd.read_excel(input_path, nrows=1).columns:
                messages.append("- 按项目名称汇总（第一个sheet）")
            else:
                messages.append("- 按项目名称汇总（数据中未找到'项目名称'列）")
            messages.append("- 按发函单位汇总")
        else:
            messages.append("未生成往来函证报告（可能没有往来函证数据）")

        # 显示结果
        if bank_success or trade_success:
            messagebox.showinfo("统计完成", "\n".join(messages))
            print("\n✓ 程序执行成功完成!")
        else:
            messagebox.showwarning("统计完成", "未找到任何函证数据！")
            print("\n⚠ 程序执行完成，但未找到任何函证数据。")

    except Exception as e:
        print(f"程序执行过程中发生错误: {e}")
        traceback.print_exc()
        messagebox.showerror("错误", f"程序执行失败: {str(e)}")
    finally:
        if need_mainloop:
            root.destroy()


if __name__ == "__main__":
    main()
    input("\n程序执行完毕，按回车键退出...")