from __future__ import annotations

import argparse
import re
from copy import copy
from datetime import date
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17324D"
TEAL = "0F7C80"
GOLD = "D6A83D"
LIGHT_GOLD = "FFF4D6"
PALE_TEAL = "E8F4F3"
PALE_BLUE = "EAF1F6"
LIGHT = "F4F7F9"
MID = "D9E2E8"
TEXT = "243746"
MUTED = "64717D"
WHITE = "FFFFFF"
GREEN = "2F7D5A"
PALE_GREEN = "E6F3EC"

FONT_NAME = "Microsoft YaHei"
THIN_GRAY = Side(style="thin", color=MID)
MEDIUM_NAVY = Side(style="medium", color=NAVY)
GRID_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def solid(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def set_cell_style(
    cell,
    *,
    fill: str | None = None,
    font_color: str = TEXT,
    bold: bool = False,
    size: int = 10,
    horizontal: str = "left",
    vertical: str = "center",
    wrap: bool = True,
    border: Border = GRID_BORDER,
):
    if fill:
        cell.fill = solid(fill)
    cell.font = Font(
        name=FONT_NAME,
        size=size,
        bold=bold,
        color=font_color,
    )
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
    )
    cell.border = border


def style_header_row(ws, row: int, start_col: int, end_col: int, fill: str = NAVY):
    for col in range(start_col, end_col + 1):
        set_cell_style(
            ws.cell(row, col),
            fill=fill,
            font_color=WHITE,
            bold=True,
            size=10,
            horizontal="center",
        )


def style_section_title(ws, row: int, start_col: int, end_col: int, fill: str = GOLD):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        set_cell_style(
            cell,
            fill=fill,
            font_color=NAVY,
            bold=True,
            size=11,
            border=Border(bottom=MEDIUM_NAVY),
        )


def apply_page_setup(ws, print_area: str, repeat_rows: str | None = None):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = print_area
    if repeat_rows:
        ws.print_title_rows = repeat_rows
    ws.oddFooter.left.text = "FY27 WP COE 服务方案"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Page &P / &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED


def style_source_sheet(ws):
    max_col = ws.max_column
    max_row = ws.max_row
    columns = source_column_map(ws)
    field_widths = {
        "client_name": 34,
        "engagement_code": 14,
        "engagement_name": 38,
        "outlook_hours": 14,
        "schedule_status": 13,
        "service_number": 27,
        "task_count": 14,
        "project_status": 15,
        "wp_eic": 19,
        "wp_fic": 18,
        "pre_start": 17,
        "pre_end": 17,
        "final_start": 17,
        "final_end": 17,
        "service_type": 34,
        "audit_eic": 25,
        "report_date": 16,
        "related_order": 28,
        "total_booking_hours": 18,
        "team": 23,
        "client_code": 16,
        "engagement_id": 23,
    }
    for column in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18
    for field, width in field_widths.items():
        column = columns.get(field)
        if column:
            ws.column_dimensions[get_column_letter(column)].width = width

    ws.row_dimensions[1].height = 34
    style_header_row(ws, 1, 1, max_col, NAVY)

    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 29
        row_fill = WHITE if row % 2 == 0 else LIGHT
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            set_cell_style(cell, fill=row_fill, size=9)
        for field in (
            "engagement_code", "outlook_hours", "task_count",
            "total_booking_hours", "client_code",
        ):
            column = columns.get(field)
            if column:
                ws.cell(row, column).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        for field in ("outlook_hours", "task_count", "total_booking_hours"):
            column = columns.get(field)
            if column:
                ws.cell(row, column).number_format = "#,##0.00"
        for field in ("service_number", "related_order"):
            column = columns.get(field)
            if column and ws.cell(row, column).value:
                ws.cell(row, column).font = Font(
                    name=FONT_NAME,
                    size=9,
                    bold=True,
                    color=TEAL,
                    underline="single",
                )
                ws.cell(row, column).fill = solid(PALE_TEAL)

    if max_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = GOLD if ws.title in {"AUD2026", "FY26"} else TEAL if ws.title == "IPO" else MUTED

    schedule_column = columns.get("schedule_status")
    project_column = columns.get("project_status")
    if max_row >= 2 and schedule_column:
        letter = get_column_letter(schedule_column)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{max_row}",
            FormulaRule(
                formula=[f'{letter}2="已完成"'],
                fill=solid(PALE_GREEN),
                font=Font(color=GREEN, bold=True),
            ),
        )
    if max_row >= 2 and project_column:
        letter = get_column_letter(project_column)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{max_row}",
            FormulaRule(
                formula=[f'{letter}2="项目承接"'],
                fill=solid(LIGHT_GOLD),
                font=Font(color=NAVY, bold=True),
            ),
        )

    apply_page_setup(ws, f"A1:{get_column_letter(max_col)}{max_row}", "1:1")


def style_service_sheet(ws):
    widths = {
        "A": 13,
        "B": 36,
        "C": 15,
        "D": 15,
        "E": 17,
        "F": 15,
        "G": 16,
        "H": 18,
        "I": 14,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = "A5"

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 42
    style_header_row(ws, 1, 1, 8, NAVY)
    for col in range(1, 9):
        set_cell_style(ws.cell(2, col), fill=WHITE, size=10, bold=col in (2, 3, 4))
    for row in (1, 2):
        for col in range(5, 9):
            set_cell_style(
                ws.cell(row, col),
                fill=WHITE,
                size=10,
                border=Border(),
            )
    ws["C2"].number_format = "#,##0.00"
    ws["D2"].number_format = "#,##0.00"
    if ws["I1"].value:
        set_cell_style(ws["I1"], fill=TEAL, font_color=WHITE, bold=True, horizontal="center")
        ws["I1"].font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE, underline="single")

    ws.row_dimensions[4].height = 32
    style_header_row(ws, 4, 1, 8, TEAL)
    set_cell_style(ws["I4"], fill=WHITE, border=Border())

    for row in range(5, 37):
        ws.row_dimensions[row].height = 27
        row_fill = WHITE if row % 2 else LIGHT
        for col in range(1, 9):
            set_cell_style(ws.cell(row, col), fill=row_fill, size=9)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        for col in (3, 4, 5, 6, 7, 8):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in (3, 4, 6):
            ws.cell(row, col).number_format = "General"
        for col in (5, 7, 8):
            ws.cell(row, col).number_format = "0.00"
        for col in (3, 4, 6):
            ws.cell(row, col).fill = solid(LIGHT_GOLD)
        ws.cell(row, 5).fill = solid(PALE_BLUE)
        ws.cell(row, 8).fill = solid(LIGHT)

    for col in range(1, 9):
        set_cell_style(ws.cell(37, col), fill=PALE_TEAL, bold=col in (4, 6, 7), size=9)
    ws["G37"].number_format = "#,##0.00"

    style_section_title(ws, 39, 1, 8, GOLD)
    ws.row_dimensions[39].height = 28
    style_header_row(ws, 40, 1, 8, NAVY)
    ws.row_dimensions[40].height = 30
    for row in range(41, 55):
        ws.row_dimensions[row].height = 26
        row_fill = WHITE if row % 2 else LIGHT
        for col in range(1, 9):
            set_cell_style(ws.cell(row, col), fill=row_fill, size=9)
        ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    style_section_title(ws, 55, 1, 8, TEAL)
    for col in range(1, 9):
        ws.cell(55, col).font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    ws.row_dimensions[55].height = 28
    for col in range(1, 9):
        set_cell_style(ws.cell(56, col), fill=PALE_TEAL, bold=col in (1, 2), size=9)
    ws["B56"].number_format = "#,##0.00"
    style_header_row(ws, 57, 1, 6, NAVY)
    for col in range(7, 9):
        set_cell_style(ws.cell(57, col), fill=LIGHT, size=9)

    for row in range(58, 62):
        ws.row_dimensions[row].height = 27
        row_fill = WHITE if row % 2 == 0 else LIGHT
        for col in range(1, 9):
            set_cell_style(ws.cell(row, col), fill=row_fill, size=9)
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).number_format = "#,##0.00"
        ws.cell(row, 4).number_format = "#,##0.00"
        ws.cell(row, 5).number_format = "#,##0.00"
        ws.cell(row, 6).number_format = "#,##0.00"

    for col in range(1, 9):
        set_cell_style(ws.cell(62, col), fill=LIGHT_GOLD, bold=col in (1, 2, 3, 6), size=9)
    ws["B62"].number_format = "0%"
    ws["C62"].number_format = "#,##0.00"
    ws["F62"].number_format = "#,##0.00"
    ws.row_dimensions[62].height = 29

    apply_page_setup(ws, "A1:H62", "1:4")


def parse_source_name(return_formula: object) -> str:
    text = str(return_formula or "")
    match = re.search(r"#'?(AUD2026|FY26|IPO)'?!", text, re.IGNORECASE)
    if not match:
        return "AUD2026"
    name = match.group(1).upper()
    return "AUD2026" if name == "FY26" else name


def quote_sheet_name(name: str) -> str:
    return name.replace("'", "''")


def hyperlink_display(value: object):
    if not isinstance(value, str) or not value.startswith("=HYPERLINK"):
        return value
    quoted = re.findall(r'"((?:[^"]|"")*)"', value)
    return quoted[-1].replace('""', '"') if quoted else value


SOURCE_FIELD_ALIASES = {
    "client_name": ("Client Name",),
    "engagement_code": ("Engagement Code",),
    "engagement_name": ("Engagement Name",),
    "outlook_hours": ("Outlook Hours",),
    "service_number": ("WP服务单编号",),
    "task_count": ("底稿任务数量",),
    "schedule_status": ("排班状态",),
    "project_status": ("项目状态",),
    "wp_eic": ("WP EIC",),
    "wp_fic": ("WP FIC", "WP FIC*"),
    "service_type": ("Service Type",),
    "audit_eic": ("Audit EIC",),
    "report_date": ("Audit Report Date",),
    "related_order": ("相关订单",),
    "pre_start": ("Booking Period Start-预审",),
    "pre_end": ("Booking Period End-预审",),
    "final_start": ("Booking Period Start-年审",),
    "final_end": ("Booking Period End-年审",),
    "total_booking_hours": ("Total Booking Hours",),
    "team": ("团队",),
    "client_code": ("Client Code",),
    "engagement_id": ("审计项目",),
}


def normalize_header(value: object) -> str:
    return (
        re.sub(r"\s+", "", str(value or ""))
        .replace("–", "-")
        .replace("—", "-")
        .casefold()
    )


def source_column_map(ws, required=()):
    """Return source field columns by header name, independent of column order."""
    headers = {
        normalize_header(ws.cell(1, column).value): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(1, column).value not in (None, "")
    }
    columns = {}
    for field, aliases in SOURCE_FIELD_ALIASES.items():
        for alias in aliases:
            column = headers.get(normalize_header(alias))
            if column:
                columns[field] = column
                break

    missing = [field for field in required if field not in columns]
    if missing:
        labels = [SOURCE_FIELD_ALIASES[field][0] for field in missing]
        raise ValueError(f"{ws.title}工作表缺少字段：{'、'.join(labels)}")
    return columns


def source_value(ws, row: int, columns: dict[str, int], field: str, default=None):
    column = columns.get(field)
    return ws.cell(row, column).value if column else default


def create_index_sheet(wb, service_sheets):
    if "服务方案索引" in wb.sheetnames:
        del wb["服务方案索引"]
    ws = wb.create_sheet("服务方案索引", 0)
    ws.sheet_properties.tabColor = GOLD
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    ws.merge_cells("A1:K1")
    ws["A1"] = "FY27 WP 服务方案清单"
    ws["A1"].fill = solid(NAVY)
    ws["A1"].font = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 46

    ws.merge_cells("A2:K2")
    ws["A2"] = "项目组展示版 · 服务单、相关订单、Section 与 SER 测算集中查看"
    ws["A2"].fill = solid(NAVY)
    ws["A2"].font = Font(name=FONT_NAME, size=10, color="DDE8F0")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 27

    aud2026_count = sum(1 for sheet in service_sheets if parse_source_name(sheet["I1"].value) == "AUD2026")
    ipo_count = len(service_sheets) - aud2026_count
    summary = [
        ("A4", "服务方案", "B4", len(service_sheets)),
        ("C4", "AUD2026项目", "D4", aud2026_count),
        ("E4", "IPO项目", "F4", ipo_count),
        ("G4", "生成日期", "H4", date.today().isoformat()),
    ]
    for label_ref, label, value_ref, value in summary:
        set_cell_style(ws[label_ref], fill=PALE_TEAL, bold=True, font_color=MUTED, size=9)
        set_cell_style(ws[value_ref], fill=WHITE, bold=True, font_color=NAVY, size=12, horizontal="center")
        ws[label_ref] = label
        ws[value_ref] = value
    source_info_by_service = {}
    for source_name in ("AUD2026", "FY26", "IPO"):
        if source_name not in wb.sheetnames:
            continue
        source_ws = wb[source_name]
        columns = source_column_map(
            source_ws, required=("service_number", "outlook_hours")
        )
        for row in range(2, source_ws.max_row + 1):
            service_number = hyperlink_display(
                source_value(source_ws, row, columns, "service_number")
            )
            if service_number:
                key = str(service_number).strip()
                if key not in source_info_by_service:
                    source_info_by_service[key] = {
                        "wp_fic": source_value(
                            source_ws, row, columns, "wp_fic", ""
                        ),
                        "outlook_hours": source_value(
                            source_ws, row, columns, "outlook_hours", ""
                        ),
                    }

    headers = [
        "序号", "来源", "项目名称", "WP服务单编号", "相关订单", "WP FIC",
        "预算Outlook Hours", "源表Outlook Hours", "差异", "核对结果", "查看服务方案",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(7, col).value = header
    style_header_row(ws, 7, 1, 11, TEAL)
    ws.row_dimensions[7].height = 32

    for index, service_ws in enumerate(service_sheets, 1):
        row = 7 + index
        source_name = parse_source_name(service_ws["I1"].value)
        safe_sheet = quote_sheet_name(service_ws.title)
        service_number = service_ws["B2"].value
        source_info = source_info_by_service.get(
            str(service_number).strip(), {}
        )
        has_section_data = any(
            service_ws.cell(section_row, col).value not in (None, "")
            for section_row in range(5, 37)
            for col in (3, 6)
        )
        values = [
            index,
            source_name,
            service_ws.title,
            service_number,
            service_ws["A2"].value,
            source_info.get("wp_fic", ""),
            f"='{safe_sheet}'!C2",
            source_info.get("outlook_hours", ""),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col).value = value
        if has_section_data:
            ws.cell(row, 9).value = f'=IF(OR(G{row}="",H{row}=""),"",G{row}-H{row})'
            ws.cell(row, 10).value = f'=IF(I{row}="","",IF(ABS(I{row})<=0.01,"一致","不一致"))'
        else:
            ws.cell(row, 9).value = None
            ws.cell(row, 10).value = "待补充Section"
        ws.cell(row, 11).value = f'=HYPERLINK("#\'{safe_sheet}\'!A1","打开")'

        row_fill = WHITE if row % 2 == 0 else LIGHT
        for col in range(1, 12):
            set_cell_style(ws.cell(row, col), fill=row_fill, size=9)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        for col in (7, 8, 9):
            ws.cell(row, col).number_format = "#,##0.00"
        ws.cell(row, 10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 11).fill = solid(PALE_TEAL)
        ws.cell(row, 11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 11).font = Font(name=FONT_NAME, size=9, bold=True, color=TEAL, underline="single")
        ws.row_dimensions[row].height = 28

    widths = {
        "A": 8, "B": 10, "C": 38, "D": 27, "E": 28, "F": 24,
        "G": 19, "H": 19, "I": 14, "J": 16, "K": 16,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:K{7 + len(service_sheets)}"
    apply_page_setup(ws, f"A1:K{7 + len(service_sheets)}", "7:7")


def format_workbook(input_path: Path, output_path: Path):
    wb = load_workbook(input_path, data_only=False)

    base_sheets = {
        "AUD2026", "AUD2025", "FY26", "FY25", "IPO", "IPO archive",
        "服务方案索引",
    }
    service_sheets = [ws for ws in wb.worksheets if ws.title not in base_sheets]

    for name in ("AUD2026", "AUD2025", "FY26", "FY25", "IPO", "IPO archive"):
        if name in wb.sheetnames:
            style_source_sheet(wb[name])

    for ws in service_sheets:
        style_service_sheet(ws)

    create_index_sheet(wb, service_sheets)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def validate_workbook(output_path: Path) -> dict[str, int]:
    wb = load_workbook(output_path, data_only=False, read_only=False)
    base_sheets = {
        "服务方案索引", "AUD2026", "AUD2025", "FY26", "FY25", "IPO",
        "IPO archive",
    }
    service_sheets = [ws for ws in wb.worksheets if ws.title not in base_sheets]
    service_errors = []

    for ws in service_sheets:
        sections = [ws.cell(row, 2).value for row in range(5, 37)]
        formulas = [ws.cell(row, 5).value for row in range(5, 37)]
        all_values = [cell.value for row in ws.iter_rows() for cell in row]
        ser_count = sum(
            1 for value in all_values if isinstance(value, str) and value.startswith("SER测算")
        )
        scope_count = sum(
            1 for value in all_values if isinstance(value, str) and "底稿服务范围" in value
        )
        if (
            len([value for value in sections if value]) != 32
            or ser_count != 1
            or scope_count != 0
            or not all(isinstance(value, str) and value.startswith("=") for value in formulas)
        ):
            service_errors.append(ws.title)

    index_ws = wb["服务方案索引"]
    index_rows = sum(
        1 for row in range(8, index_ws.max_row + 1) if index_ws.cell(row, 3).value
    )

    xml_errors = []
    with ZipFile(output_path) as archive:
        for name in archive.namelist():
            if name.endswith(".xml"):
                try:
                    ElementTree.fromstring(archive.read(name))
                except Exception:
                    xml_errors.append(name)

    if service_errors or xml_errors or index_rows != len(service_sheets):
        raise RuntimeError(
            f"生成后检查失败：服务方案异常 {len(service_errors)}，"
            f"XML异常 {len(xml_errors)}，索引 {index_rows}/{len(service_sheets)}。"
        )

    audit_sheet_name = "AUD2026" if "AUD2026" in wb.sheetnames else "FY26"
    return {
        "sheets": len(wb.sheetnames),
        "services": len(service_sheets),
        "index_rows": index_rows,
        "aud2026_rows": wb[audit_sheet_name].max_row - 1,
        "ipo_rows": wb["IPO"].max_row - 1,
        "ipo_archive_rows": (
            wb["IPO archive"].max_row - 1 if "IPO archive" in wb.sheetnames else 0
        ),
    }


def main():
    parser = argparse.ArgumentParser(description="将 FY27 WP 服务单格式化为项目组展示版。")
    parser.add_argument("--input", required=True, type=Path, help="输入服务单 Excel 文件")
    parser.add_argument("--output", required=True, type=Path, help="输出展示版 Excel 文件")
    args = parser.parse_args()
    format_workbook(args.input.resolve(), args.output.resolve())
    result = validate_workbook(args.output.resolve())
    print(args.output.resolve())
    print(
        "检查通过："
        f"{result['services']} 张服务方案，{result['index_rows']} 条索引，"
        f"AUD2026 {result['aud2026_rows']} 个项目，IPO {result['ipo_rows']} 个项目，"
        f"IPO archive {result['ipo_archive_rows']} 个项目。"
    )


if __name__ == "__main__":
    main()
