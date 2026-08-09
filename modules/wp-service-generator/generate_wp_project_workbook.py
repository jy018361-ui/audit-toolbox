from __future__ import annotations

import argparse
import importlib
import re
import warnings
from pathlib import Path

from openpyxl import load_workbook

import format_wp_workbook as workbook_formatter

workbook_formatter = importlib.reload(workbook_formatter)
create_index_sheet = workbook_formatter.create_index_sheet
style_service_sheet = workbook_formatter.style_service_sheet
style_source_sheet = workbook_formatter.style_source_sheet
validate_workbook = workbook_formatter.validate_workbook
source_column_map = workbook_formatter.source_column_map
source_value = workbook_formatter.source_value


SER_CONFIG_FILENAME = "SER配置.xlsx"
DEFAULT_SER_RULES = (
    {"role": "Manager", "hours_mix": 0.08, "ser_rate": 2733.0},
    {"role": "Senior", "hours_mix": 0.25, "ser_rate": 1199.0},
    {"role": "Staff", "hours_mix": 0.58, "ser_rate": 683.0},
    {"role": "Intern", "hours_mix": 0.09, "ser_rate": 173.0},
)
DEFAULT_SER_CONFIG = tuple(
    (rule["hours_mix"], rule["ser_rate"]) for rule in DEFAULT_SER_RULES
)


def load_ser_config(folder: Path):
    config_path = folder / SER_CONFIG_FILENAME
    if not config_path.exists():
        return DEFAULT_SER_CONFIG

    wb = load_workbook(config_path, data_only=True, read_only=False)
    ws = wb.active
    rows = []
    try:
        for row in range(2, ws.max_row + 1):
            mix_value = ws.cell(row, 2).value
            rate_value = ws.cell(row, 3).value
            if mix_value is None and rate_value is None:
                continue
            try:
                mix = float(mix_value)
                rate = float(rate_value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{SER_CONFIG_FILENAME}第{row}行比例或费率不是数字。") from exc
            if mix > 1:
                mix /= 100
            if mix <= 0 or rate <= 0:
                raise ValueError(f"{SER_CONFIG_FILENAME}第{row}行比例和费率必须大于0。")
            rows.append((mix, rate))
    finally:
        wb.close()

    if len(rows) != 4:
        raise ValueError(
            f"{SER_CONFIG_FILENAME}必须有4行配置，顺序为Manager、Senior、Staff、Intern。"
        )
    if abs(sum(mix for mix, _ in rows) - 1) > 0.0001:
        raise ValueError(f"{SER_CONFIG_FILENAME}的Hours占比合计必须为100%。")
    return tuple(rows)


def display_value(value):
    if not isinstance(value, str) or not value.startswith("=HYPERLINK"):
        return value
    quoted = re.findall(r'"((?:[^"]|"")*)"', value)
    return quoted[-1].replace('""', '"') if quoted else value


def number_value(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def normalize_order_number(value: object) -> str:
    return (
        re.sub(r"\s+", "", str(value or ""))
        .upper()
        .replace("–", "-")
        .replace("—", "-")
    )


def normalize_section_name(value: object) -> str:
    normalized = (
        re.sub(r"\s+", "", str(value or ""))
        .replace("（", "(")
        .replace("）", ")")
        .lower()
    )
    return re.sub(r"^u_exp(?:-?other)?(?=\()", "u_exp", normalized)


def safe_sheet_name(preferred: object, used_names: set[str]) -> str:
    base = re.sub(r"[\\/?*\[\]:]", " ", str(preferred or "服务方案")).strip()
    base = base[:31].strip() or "服务方案"
    candidate = base
    suffix_number = 2
    lowered = {name.lower() for name in used_names}
    while candidate.lower() in lowered:
        suffix = f"_{suffix_number}"
        candidate = f"{base[: 31 - len(suffix)].rstrip()}{suffix}"
        suffix_number += 1
    used_names.add(candidate)
    return candidate


def quote_sheet(name: str) -> str:
    return name.replace("'", "''")


def hyperlink_formula(sheet_name: str, target_cell: str, display: object) -> str:
    safe_sheet = quote_sheet(sheet_name)
    safe_display = str(display or "").replace('"', '""')
    return f'=HYPERLINK("#\'{safe_sheet}\'!{target_cell}","{safe_display}")'


def normalize_base_sheet_names(wb):
    if "AUD2026" not in wb.sheetnames and "FY26" in wb.sheetnames:
        wb["FY26"].title = "AUD2026"
    if "AUD2025" not in wb.sheetnames and "FY25" in wb.sheetnames:
        wb["FY25"].title = "AUD2025"
    if "AUD2026" not in wb.sheetnames:
        raise ValueError("找不到 AUD2026（或旧名称 FY26）工作表。")
    if "IPO" not in wb.sheetnames:
        raise ValueError("找不到 IPO 工作表。")


def find_template_sheet(wb):
    base_names = {"AUD2026", "AUD2025", "IPO", "服务方案索引", "服务方案生成"}
    candidates = []
    for ws in wb.worksheets:
        if ws.title in base_names:
            continue
        if ws["A1"].value == "相关订单" and ws["B4"].value == "Section":
            sections = [ws.cell(row, 2).value for row in range(5, 37)]
            if len([value for value in sections if value]) == 32:
                candidates.append(ws)
    if not candidates:
        raise ValueError("找不到包含 32 个 Section 的服务方案模板。")
    return candidates[0]


def booking_year(value: object) -> int | None:
    if hasattr(value, "year"):
        return int(value.year)
    match = re.match(r"\s*(\d{4})", str(value or ""))
    return int(match.group(1)) if match else None


def booking_year_month(value: object) -> tuple[int, int] | None:
    if hasattr(value, "year") and hasattr(value, "month"):
        return int(value.year), int(value.month)
    match = re.match(r"\s*(\d{4})[-/](\d{1,2})", str(value or ""))
    return (int(match.group(1)), int(match.group(2))) if match else None


def split_raw_service_orders(
    raw_path: Path,
    split_path: Path,
    template_path: Path | None = None,
    ipo_years: tuple[int, ...] = (2026, 2027),
):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Workbook contains no default style"
        )
        raw_book = load_workbook(raw_path, data_only=True, read_only=False)
    if "业务" not in raw_book.sheetnames:
        raise ValueError("新导出的 FY27 WP服务单中找不到‘业务’工作表。")
    raw_sheet = raw_book["业务"]
    headers = [
        raw_sheet.cell(1, column).value
        for column in range(1, raw_sheet.max_column + 1)
    ]
    header_map = {
        re.sub(r"\s+", "", str(value or "")): index
        for index, value in enumerate(headers)
    }
    required_headers = (
        "EngagementName",
        "OutlookHours",
        "BookingPeriodStart-预审",
        "BookingPeriodEnd-预审",
        "BookingPeriodStart-年审",
        "BookingPeriodEnd-年审",
        "WP服务单编号",
    )
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        raise ValueError("FY27 WP服务单缺少字段：" + "、".join(missing))

    groups = {"AUD2026": [], "IPO": [], "IPO archive": [], "AUD2025": []}
    excluded_ipo = []
    excluded_other = []
    allowed_ipo_years = set(ipo_years)
    for row in raw_sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        engagement_name = str(row[header_map["EngagementName"]] or "").strip()
        normalized_name = engagement_name.upper()
        if normalized_name.startswith("IPO"):
            start_values = (
                row[header_map["BookingPeriodStart-预审"]],
                row[header_map["BookingPeriodStart-年审"]],
            )
            end_values = (
                row[header_map["BookingPeriodEnd-预审"]],
                row[header_map["BookingPeriodEnd-年审"]],
            )
            start_years = {booking_year(value) for value in start_values}
            start_years.discard(None)
            if start_years & allowed_ipo_years:
                start_periods = [booking_year_month(value) for value in start_values]
                end_periods = [booking_year_month(value) for value in end_values]
                is_archive = any(
                    period and period[0] == 2026 and 1 <= period[1] <= 3
                    for period in start_periods
                ) or any(
                    period and period[0] == 2026 and period[1] <= 4
                    for period in end_periods
                )
                groups["IPO archive" if is_archive else "IPO"].append(row)
            else:
                excluded_ipo.append(
                    {
                        "engagement_name": engagement_name,
                        "service_number": row[header_map["WP服务单编号"]],
                        "start_years": sorted(start_years),
                    }
                )
            continue

        if re.search(r"(?:AUD|FY)\s*2025", normalized_name):
            groups["AUD2025"].append(row)
        elif re.search(r"(?:AUD|INT)\s*20(?:26|27)", normalized_name):
            groups["AUD2026"].append(row)
        else:
            years = [int(value) for value in re.findall(r"20\d{2}", normalized_name)]
            if years and min(years) <= 2025:
                groups["AUD2025"].append(row)
            elif years:
                groups["AUD2026"].append(row)
            else:
                excluded_other.append(
                    {
                        "engagement_name": engagement_name,
                        "service_number": row[header_map["WP服务单编号"]],
                    }
                )
    raw_book.close()

    if template_path is None:
        candidates = (
            raw_path.parent / "FY27+WP服务单.xlsx",
            raw_path.parent / "服务方案.xlsx",
        )
        template_path = next((path for path in candidates if path.exists()), None)
    if template_path is None or not template_path.exists():
        raise FileNotFoundError(
            "找不到服务方案模板。请将 FY27+WP服务单.xlsx 放在同一文件夹。"
        )

    source_book = load_workbook(template_path, data_only=False, read_only=False)
    normalize_base_sheet_names(source_book)
    template_sheet = find_template_sheet(source_book)
    for sheet in list(source_book.worksheets):
        if sheet is not template_sheet:
            source_book.remove(sheet)
    template_sheet.title = "_WP_TEMPLATE"
    template_sheet.sheet_state = "hidden"

    for index, sheet_name in enumerate(
        ("AUD2026", "IPO", "IPO archive", "AUD2025")
    ):
        sheet = source_book.create_sheet(sheet_name, index)
        sheet.append(headers)
        for row in groups[sheet_name]:
            values = list(row)
            outlook_index = header_map["OutlookHours"]
            outlook_value = values[outlook_index]
            if outlook_value not in (None, ""):
                try:
                    values[outlook_index] = float(
                        str(outlook_value).replace(",", "").strip()
                    )
                except (TypeError, ValueError):
                    pass
            sheet.append(values)

    source_book.active = 0
    split_path.parent.mkdir(parents=True, exist_ok=True)
    source_book.save(split_path)
    source_book.close()
    return {
        "split_file": str(split_path),
        "split_aud2026_rows": len(groups["AUD2026"]),
        "split_ipo_rows": len(groups["IPO"]),
        "split_ipo_archive_rows": len(groups["IPO archive"]),
        "split_aud2025_rows": len(groups["AUD2025"]),
        "excluded_ipo": excluded_ipo,
        "excluded_other": excluded_other,
        "ipo_years": tuple(sorted(allowed_ipo_years)),
    }


def collect_service_orders(wb):
    records = []
    seen = set()
    for source_name in ("AUD2026", "IPO"):
        ws = wb[source_name]
        columns = source_column_map(
            ws,
            required=("engagement_name", "outlook_hours", "service_number"),
        )
        for row in range(2, ws.max_row + 1):
            service_number = display_value(
                source_value(ws, row, columns, "service_number")
            )
            if not service_number or service_number in seen:
                continue
            seen.add(service_number)
            records.append(
                {
                    "source_sheet": source_name,
                    "source_row": row,
                    "engagement_name": source_value(
                        ws, row, columns, "engagement_name"
                    ),
                    "outlook_hours": number_value(
                        source_value(ws, row, columns, "outlook_hours")
                    ),
                    "service_number": service_number,
                    "task_count": number_value(
                        source_value(ws, row, columns, "task_count")
                    ),
                    "service_type": source_value(
                        ws, row, columns, "service_type"
                    ),
                    "audit_eic": source_value(ws, row, columns, "audit_eic"),
                    "report_date": source_value(
                        ws, row, columns, "report_date"
                    ),
                    "related_order": display_value(
                        source_value(ws, row, columns, "related_order")
                    ),
                    "pre_start": source_value(ws, row, columns, "pre_start"),
                    "pre_end": source_value(ws, row, columns, "pre_end"),
                    "final_start": source_value(ws, row, columns, "final_start"),
                    "final_end": source_value(ws, row, columns, "final_end"),
                    "sheet_name": "",
                }
            )
    return records


def load_section_details(section_list_path: Path, records):
    empty_result = {
        "details": {},
        "matched_orders": 0,
        "matched_rows": 0,
        "populated_rows": 0,
        "section_list_found": False,
    }
    if not section_list_path.exists():
        return empty_result

    target_orders = {
        normalize_order_number(record["service_number"]) for record in records
    }
    details = {}
    matched_rows = 0
    populated_rows = 0
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Workbook contains no default style"
        )
        section_book = load_workbook(
            section_list_path, read_only=True, data_only=True
        )
    section_sheet = section_book[section_book.sheetnames[0]]
    if section_sheet.max_row <= 1:
        section_book.close()
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore", message="Workbook contains no default style"
            )
            section_book = load_workbook(
                section_list_path, read_only=False, data_only=True
            )
        section_sheet = section_book[section_book.sheetnames[0]]

    headers = {
        re.sub(r"\s+", "", str(section_sheet.cell(1, col).value or "")): col
        for col in range(1, section_sheet.max_column + 1)
    }

    def find_column(exact=None, prefix=None):
        if exact and exact in headers:
            return headers[exact]
        if prefix:
            for header, column in headers.items():
                if header.startswith(prefix):
                    return column
        return None

    columns = {
        "section": find_column(exact="Section"),
        "entity": find_column(prefix="Entity数量"),
        "drafts": find_column(exact="底稿数量"),
        "budget": find_column(exact="预算调整"),
        "order": find_column(exact="所属WP服务单"),
    }
    missing = [name for name, column in columns.items() if column is None]
    if missing:
        raise ValueError(
            "FY27 Section List 缺少字段：" + "、".join(missing)
        )

    first_column = min(columns.values())
    last_column = max(columns.values())

    # 只读取所需字段所在的列区间，并且只保留服务单号命中的行。
    for values in section_sheet.iter_rows(
        min_row=2,
        min_col=first_column,
        max_col=last_column,
        values_only=True,
    ):
        def field(name):
            return values[columns[name] - first_column]

        order_key = normalize_order_number(field("order"))
        if order_key not in target_orders:
            continue
        section_key = normalize_section_name(field("section"))
        if not section_key:
            continue
        matched_rows += 1
        raw_values = {
            "entity": field("entity"),
            "drafts": field("drafts"),
            "budget": field("budget"),
        }
        if any(value not in (None, "") for value in raw_values.values()):
            populated_rows += 1
        order_details = details.setdefault(order_key, {})
        item = order_details.setdefault(
            section_key,
            {"entity": None, "drafts": None, "budget": None},
        )
        for name, value in raw_values.items():
            if value in (None, ""):
                continue
            item[name] = (item[name] or 0.0) + number_value(value)

    section_book.close()
    return {
        "details": details,
        "matched_orders": len(details),
        "matched_rows": matched_rows,
        "populated_rows": populated_rows,
        "section_list_found": True,
    }


def prepare_template(ws):
    if ws["H4"].value == "Section系统编号":
        ws.delete_cols(8, 1)
    for row in range(5, 37):
        value = ws.cell(row, 8).value
        if value in (None, ""):
            ws.cell(row, 8).value = None
            continue
        try:
            ws.cell(row, 8).value = float(str(value).replace(",", ""))
        except (TypeError, ValueError):
            ws.cell(row, 8).value = None
    if ws.max_row >= 55:
        ws.delete_rows(55, ws.max_row - 54)


def calculate_section_outlook_checks(records, section_details, template_ws):
    reference_by_section = {
        normalize_section_name(template_ws.cell(row, 2).value): number_value(
            template_ws.cell(row, 8).value
        )
        for row in range(5, 37)
        if template_ws.cell(row, 2).value
    }
    template_section_rows = 0
    populated_template_rows = 0
    compared = 0
    equal = 0
    differences = []

    for record in records:
        project_details = section_details.get(
            normalize_order_number(record["service_number"]), {}
        )
        total = 0.0
        has_outlook_data = False
        for section_key, item in project_details.items():
            if section_key not in reference_by_section:
                continue
            template_section_rows += 1
            if any(item[name] is not None for name in ("entity", "drafts", "budget")):
                populated_template_rows += 1
            if item["entity"] is None and item["budget"] is None:
                continue
            has_outlook_data = True
            section_hours = (
                (item["entity"] or 0.0) * reference_by_section[section_key]
                + (item["budget"] or 0.0)
            )
            total += round(section_hours, 2)

        if not has_outlook_data:
            continue
        calculated = round(total * 1.1, 2)
        source_value = round(record["outlook_hours"], 2)
        difference = round(calculated - source_value, 2)
        compared += 1
        if abs(difference) <= 0.01:
            equal += 1
        else:
            differences.append(
                {
                    "service_number": record["service_number"],
                    "engagement_name": record["engagement_name"],
                    "calculated": calculated,
                    "source": source_value,
                    "difference": difference,
                }
            )

    return {
        "template_section_rows": template_section_rows,
        "populated_template_rows": populated_template_rows,
        "outlook_compared": compared,
        "outlook_equal": equal,
        "outlook_differences": differences,
    }


def fill_service_sheet(ws, record, section_details, ser_config):
    ws["A2"] = record["related_order"]
    ws["B2"] = record["service_number"]
    for ref in ("E1", "F1", "G1", "H1", "E2", "F2", "G2", "H2"):
        ws[ref] = None
    ws["C1"] = "Outlook Hours"
    ws["C2"] = "=G37"
    ws["D1"] = "SER"
    ws["D2"] = "=F62"
    ws["I1"] = hyperlink_formula(
        record["source_sheet"], f"A{record['source_row']}", "返回源表"
    )
    ws["H4"] = "参考时间/Entity"

    project_sections = section_details.get(
        normalize_order_number(record["service_number"]), {}
    )
    for row in range(5, 37):
        imported = project_sections.get(
            normalize_section_name(ws.cell(row, 2).value)
        )
        if imported:
            ws.cell(row, 3).value = imported["entity"]
            ws.cell(row, 4).value = imported["drafts"]
            ws.cell(row, 6).value = imported["budget"]
        else:
            ws.cell(row, 3).value = None
            ws.cell(row, 4).value = None
            ws.cell(row, 6).value = None
        ws.cell(row, 5).value = (
            f'=IF(OR(C{row}="",H{row}=""),"",'
            f'ROUND(C{row}*IFERROR(VALUE(H{row}),0),2))'
        )
        ws.cell(row, 7).value = (
            f'=IF(AND(F{row}="",OR(C{row}="",H{row}="")),"",'
            f'ROUND(IF(OR(C{row}="",H{row}=""),0,C{row}*IFERROR(VALUE(H{row}),0))'
            f'+IFERROR(VALUE(F{row}),0),2))'
        )
    ws["G37"] = "=SUM(G5:G36)*1.1"
    ws["C41"] = record["pre_start"]
    ws["C42"] = record["pre_end"]
    ws["C47"] = record["final_start"]
    ws["C48"] = record["final_end"]

    ws["A55"] = "SER测算（计算上浮5%）"
    ws["A56"] = "Total Outlook Hours"
    ws["B56"] = "=G37"
    headers = (None, "Hours占比", "分配Hours", None, None, "SER金额")
    for col, header in enumerate(headers, 1):
        ws.cell(57, col).value = header

    for offset, (mix, rate) in enumerate(ser_config):
        row = 58 + offset
        ws.cell(row, 1).value = None
        ws.cell(row, 2).value = mix
        ws.cell(row, 3).value = f"=B{row}*$G$37"
        ws.cell(row, 4).value = rate
        ws.cell(row, 5).value = f"=D{row}*1.05"
        ws.cell(row, 6).value = f"=C{row}*E{row}"
    ws["A62"] = "合计"
    ws["B62"] = "=SUM(B58:B61)"
    ws["C62"] = "=SUM(C58:C61)"
    ws["F62"] = "=SUM(F58:F61)"


def generate(input_path: Path, output_path: Path):
    ser_config = load_ser_config(input_path.parent)
    split_result = {}
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Workbook contains no default style"
        )
        probe_book = load_workbook(input_path, data_only=True, read_only=False)
    is_raw_export = "业务" in probe_book.sheetnames and "AUD2026" not in probe_book.sheetnames
    probe_book.close()
    if is_raw_export:
        split_path = input_path.parent / "FY27+WP服务单_自动拆分.xlsx"
        split_result = split_raw_service_orders(input_path, split_path)
        input_path = split_path

    wb = load_workbook(input_path, data_only=False)
    normalize_base_sheet_names(wb)
    template_ws = find_template_sheet(wb)
    records = collect_service_orders(wb)
    if not records:
        raise ValueError("AUD2026 和 IPO 中没有找到 WP服务单编号。")
    section_result = load_section_details(
        input_path.parent / "FY27 section list.xlsx", records
    )
    section_details = section_result["details"]
    unmatched_section_orders = [
        record["service_number"]
        for record in records
        if normalize_order_number(record["service_number"])
        not in section_details
    ]

    keep_names = {
        "AUD2026", "AUD2025", "IPO", "IPO archive", template_ws.title
    }
    for ws in list(wb.worksheets):
        if ws.title not in keep_names:
            wb.remove(ws)

    prepare_template(template_ws)
    outlook_result = calculate_section_outlook_checks(
        records, section_details, template_ws
    )
    template_ws.title = "_WP_TEMPLATE"
    used_names = {name for name in wb.sheetnames if name != template_ws.title}
    for record in records:
        record["sheet_name"] = safe_sheet_name(record["engagement_name"], used_names)

    service_sheets = []
    for record in records:
        ws = wb.copy_worksheet(template_ws)
        ws.title = record["sheet_name"]
        fill_service_sheet(ws, record, section_details, ser_config)
        service_sheets.append(ws)
    wb.remove(template_ws)

    record_by_number = {record["service_number"]: record for record in records}
    for source_name in ("AUD2026", "IPO"):
        ws = wb[source_name]
        columns = source_column_map(ws, required=("service_number",))
        for row in range(2, ws.max_row + 1):
            service_number = display_value(
                source_value(ws, row, columns, "service_number")
            )
            record = record_by_number.get(service_number)
            if not record:
                continue
            ws.cell(row, columns["service_number"]).value = hyperlink_formula(
                record["sheet_name"], "A1", service_number
            )
            related_order = display_value(
                source_value(ws, row, columns, "related_order")
            )
            if related_order:
                ws.cell(row, columns["related_order"]).value = hyperlink_formula(
                    record["sheet_name"], "A1", related_order
                )

    for name in ("AUD2026", "AUD2025", "IPO", "IPO archive"):
        if name in wb.sheetnames:
            style_source_sheet(wb[name])
    for ws in service_sheets:
        style_service_sheet(ws)
    create_index_sheet(wb, service_sheets)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    result = validate_workbook(output_path)
    validate_formula_logic(output_path, section_details, ser_config)
    result.update(
        {
            "section_list_found": section_result["section_list_found"],
            "matched_section_orders": section_result["matched_orders"],
            "matched_section_rows": section_result["matched_rows"],
            "populated_section_rows": section_result["populated_rows"],
            "template_section_rows": outlook_result["template_section_rows"],
            "populated_template_rows": outlook_result["populated_template_rows"],
            "outlook_compared": outlook_result["outlook_compared"],
            "outlook_equal": outlook_result["outlook_equal"],
            "outlook_differences": outlook_result["outlook_differences"],
            "unmatched_section_orders": unmatched_section_orders,
            **split_result,
        }
    )
    return result


def validate_formula_logic(output_path: Path, section_details=None, ser_config=None):
    if ser_config is None:
        ser_config = load_ser_config(output_path.parent)
    wb = load_workbook(output_path, data_only=False, read_only=False)
    base_names = {
        "服务方案索引", "AUD2026", "AUD2025", "IPO", "IPO archive"
    }
    errors = []
    for ws in wb.worksheets:
        if ws.title in base_names:
            continue
        hidden_fields = [ws[ref].value for ref in ("E1", "F1", "G1", "H1", "E2", "F2", "G2", "H2")]
        expected = (
            all(value is None for value in hidden_fields)
            and ws["C1"].value == "Outlook Hours"
            and ws["C2"].value == "=G37"
            and ws["D1"].value == "SER"
            and ws["D2"].value == "=F62"
            and ws["G37"].value == "=SUM(G5:G36)*1.1"
            and ws["B56"].value == "=G37"
            and ws["H4"].value == "参考时间/Entity"
            and all(
                ws.cell(row, 7).value
                == f'=IF(AND(F{row}="",OR(C{row}="",H{row}="")),"",ROUND(IF(OR(C{row}="",H{row}=""),0,C{row}*IFERROR(VALUE(H{row}),0))+IFERROR(VALUE(F{row}),0),2))'
                for row in range(5, 37)
            )
            and all(
                ws.cell(row, 5).value
                == f'=IF(OR(C{row}="",H{row}=""),"",ROUND(C{row}*IFERROR(VALUE(H{row}),0),2))'
                for row in range(5, 37)
            )
            and all(
                ws.cell(row, 3).value == f"=B{row}*$G$37"
                for row in range(58, 62)
            )
            and all(ws.cell(row, 5).value == f"=D{row}*1.05" for row in range(58, 62))
            and all(ws.cell(row, 1).value is None for row in range(58, 62))
            and ws["D57"].value is None
            and ws["E57"].value is None
            and [ws.cell(row, 2).value for row in range(58, 62)]
            == [mix for mix, _ in ser_config]
            and [ws.cell(row, 4).value for row in range(58, 62)]
            == [rate for _, rate in ser_config]
        )
        if not expected:
            errors.append(ws.title)
        if section_details:
            project_details = section_details.get(
                normalize_order_number(ws["B2"].value), {}
            )
            for row in range(5, 37):
                imported = project_details.get(
                    normalize_section_name(ws.cell(row, 2).value)
                )
                if not imported:
                    continue
                actual = (
                    ws.cell(row, 3).value,
                    ws.cell(row, 4).value,
                    ws.cell(row, 6).value,
                )
                expected_values = (
                    imported["entity"],
                    imported["drafts"],
                    imported["budget"],
                )
                if actual != expected_values:
                    errors.append(f"{ws.title}!{row}")

    index_ws = wb["服务方案索引"]
    expected_headers = [
        "序号", "来源", "项目名称", "WP服务单编号",
        "相关订单", "WP FIC", "预算Outlook Hours", "源表Outlook Hours",
        "差异", "核对结果", "查看服务方案",
    ]
    actual_headers = [index_ws.cell(7, col).value for col in range(1, 12)]
    source_info_by_service = {}
    for source_name in ("AUD2026", "IPO"):
        source_ws = wb[source_name]
        columns = source_column_map(
            source_ws, required=("service_number", "outlook_hours")
        )
        for row in range(2, source_ws.max_row + 1):
            service_number = display_value(
                source_value(source_ws, row, columns, "service_number")
            )
            if service_number:
                key = str(service_number).strip()
                if key not in source_info_by_service:
                    source_info_by_service[key] = {
                        "wp_fic": source_value(
                            source_ws, row, columns, "wp_fic", ""
                        ),
                        "outlook_hours": source_value(
                            source_ws, row, columns, "outlook_hours", ""
                        ),
                    }
    index_mismatches = []
    for row in range(8, index_ws.max_row + 1):
        service_number = index_ws.cell(row, 4).value
        if not service_number:
            continue
        actual_wp_fic = str(index_ws.cell(row, 6).value or "").strip()
        source_info = source_info_by_service.get(
            str(service_number).strip(), {}
        )
        expected_wp_fic = str(
            source_info.get("wp_fic") or ""
        ).strip()
        actual_outlook = number_value(index_ws.cell(row, 8).value)
        expected_outlook = number_value(source_info.get("outlook_hours"))
        if (
            actual_wp_fic != expected_wp_fic
            or abs(actual_outlook - expected_outlook) > 0.001
        ):
            index_mismatches.append(row)
    if actual_headers != expected_headers or index_mismatches:
        raise RuntimeError(
            f"索引检查失败：表头一致={actual_headers == expected_headers}，"
            f"WP FIC不一致行={index_mismatches[:5]}"
        )
    if errors:
        raise RuntimeError(f"公式逻辑检查失败，共 {len(errors)} 张：{errors[:3]}")


def main():
    parser = argparse.ArgumentParser(
        description="读取FY27 WP服务单导出或AUD2026/IPO服务单并生成项目组展示版。"
    )
    parser.add_argument("--input", required=True, type=Path, help="原始服务单 Excel")
    parser.add_argument("--output", required=True, type=Path, help="项目组展示版 Excel")
    args = parser.parse_args()
    result = generate(args.input.resolve(), args.output.resolve())
    print(args.output.resolve())
    print(
        "检查通过："
        f"{result['services']} 张服务方案，{result['index_rows']} 条索引，"
        f"AUD2026 {result['aud2026_rows']} 个项目，IPO {result['ipo_rows']} 个项目，"
        f"IPO archive {result.get('ipo_archive_rows', 0)} 个项目。"
    )
    if result["section_list_found"]:
        print(
            "Section回填："
            f"匹配 {result['matched_section_orders']} 个服务单，"
            f"{result['matched_section_rows']} 条Section，"
            f"其中 {result['populated_section_rows']} 条有数量或预算数据。"
        )
        print(
            "Outlook核对："
            f"可核对 {result['outlook_compared']} 个项目，"
            f"一致 {result['outlook_equal']} 个，"
            f"不一致 {result['outlook_compared'] - result['outlook_equal']} 个。"
        )
        if result["unmatched_section_orders"]:
            print(
                "Section List未匹配服务单："
                + "；".join(result["unmatched_section_orders"])
            )
    else:
        print("未找到 FY27 section list.xlsx，Section数量字段保持空白。")


if __name__ == "__main__":
    main()
