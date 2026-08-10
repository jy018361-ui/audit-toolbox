import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import format_wp_workbook as formatter
from generate_wp_project_workbook import (
    DEFAULT_SER_CONFIG,
    collect_service_orders,
    fill_service_sheet,
    find_section_list_file,
    find_service_order_file,
)


SHUFFLED_HEADERS = [
    "WP FIC*",
    "相关订单",
    "Booking Period End-年审",
    "Service Type",
    "WP服务单编号",
    "Outlook Hours",
    "Audit EIC",
    "Engagement Name",
    "Booking Period Start-预审",
    "底稿任务数量",
    "Audit Report Date",
    "Booking Period End-预审",
    "Booking Period Start-年审",
]


SHUFFLED_VALUES = [
    "fic.user",
    "TEST-ORDER-001",
    "2027-04-30",
    "Audit/Working paper/WP审计底稿COE服务",
    "TEST-WP-001",
    123.45,
    "audit.eic",
    "AUD2026-12 Header Order Test",
    "2026-10-01",
    2,
    "2027-03-31",
    "2026-10-31",
    "2027-01-01",
]


def add_source_sheet(workbook, title):
    sheet = workbook.create_sheet(title)
    sheet.append(SHUFFLED_HEADERS)
    if title == "AUD2026":
        sheet.append(SHUFFLED_VALUES)
    return sheet


class HeaderBasedSourceReadingTests(unittest.TestCase):
    def test_input_files_are_found_by_keywords(self):
        folder = Path("test-inputs")
        service_order = folder / "8月导出的 WP 服务单 v2.xlsx"
        section_list = folder / "Client Section LIST final.xlsx"
        files = [
            service_order,
            section_list,
            folder / "FY27+WP服务单汇总.xlsx",
            folder / "~$临时 WP服务单.xlsx",
        ]
        with patch.object(Path, "iterdir", return_value=iter(files)), patch.object(
            Path, "is_file", return_value=True
        ):
            self.assertEqual(find_service_order_file(folder), service_order)
        with patch.object(Path, "iterdir", return_value=iter(files)), patch.object(
            Path, "is_file", return_value=True
        ):
            self.assertEqual(find_section_list_file(folder), section_list)

    def test_multiple_keyword_matches_are_rejected(self):
        folder = Path("test-inputs")
        files = [folder / "WP服务单 A.xlsx", folder / "WP服务单 B.xlsx"]
        with patch.object(Path, "iterdir", return_value=iter(files)), patch.object(
            Path, "is_file", return_value=True
        ):
            with self.assertRaisesRegex(ValueError, "多个可能的WP服务单"):
                find_service_order_file(folder)

    def test_collect_service_orders_uses_headers_not_positions(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        add_source_sheet(workbook, "AUD2026")
        add_source_sheet(workbook, "IPO")

        records = collect_service_orders(workbook)

        self.assertEqual(len(records), 1)
        record = records[0]
        self.assertEqual(record["engagement_name"], "AUD2026-12 Header Order Test")
        self.assertEqual(record["service_number"], "TEST-WP-001")
        self.assertEqual(record["outlook_hours"], 123.45)
        self.assertEqual(record["related_order"], "TEST-ORDER-001")
        self.assertEqual(record["service_type"], "Audit/Working paper/WP审计底稿COE服务")

    def test_index_uses_headers_not_positions(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        add_source_sheet(workbook, "AUD2026")
        add_source_sheet(workbook, "IPO")
        service = workbook.create_sheet("AUD2026 Test")
        service["A2"] = "TEST-ORDER-001"
        service["B2"] = "TEST-WP-001"
        service["C2"] = 123.45
        service["I1"] = '=HYPERLINK("#\'AUD2026\'!A2","返回原表")'
        service["C5"] = 1

        formatter.create_index_sheet(workbook, [service])

        index = workbook["服务方案索引"]
        self.assertEqual(index["F8"].value, "fic.user")
        self.assertEqual(index["H8"].value, 123.45)
        self.assertEqual(index["D8"].value, "TEST-WP-001")

    def test_service_sheet_shows_ser_roles_and_rate_headers(self):
        workbook = Workbook()
        service = workbook.active
        record = {
            "related_order": "TEST-ORDER-001",
            "service_number": "TEST-WP-001",
            "source_sheet": "AUD2026",
            "source_row": 2,
            "service_type": "Audit/Working paper",
            "task_count": 1,
            "audit_eic": "audit.eic",
            "report_date": "2027-03-31",
            "pre_start": "2026-10-01",
            "pre_end": "2026-10-31",
            "final_start": "2027-01-01",
            "final_end": "2027-04-30",
        }

        fill_service_sheet(service, record, {}, DEFAULT_SER_CONFIG)

        self.assertEqual(service["D57"].value, "bill rate")
        self.assertEqual(service["E57"].value, "上浮5%")
        self.assertEqual(
            [service.cell(row, 1).value for row in range(58, 62)],
            ["Manager", "Senior", "Staff", "Intern"],
        )
        self.assertEqual(
            [service.cell(row, 2).value for row in range(58, 62)],
            [0.08, 0.25, 0.58, 0.09],
        )


if __name__ == "__main__":
    unittest.main()
