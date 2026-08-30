from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "artifacts" / "TBJE完整性核对_三页导出样稿.xlsx"

EY_YELLOW = "FFE600"
INK = "1E2A32"
TEAL = "126E72"
LIGHT_TEAL = "E7F2F1"
LIGHT_YELLOW = "FFF9D6"
LIGHT_RED = "FDECEA"
GRID = "D9E2E1"
INPUT_BLUE = "0070C0"
WHITE = "FFFFFF"


def style_sheet(ws, widths):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:{get_column_letter(len(widths))}{ws.max_row}"
    ws.row_dimensions[1].height = 28
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(vertical="center")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=EY_YELLOW)
        cell.font = Font(name="Arial", size=14, bold=True, color=INK)
    for cell in ws[6]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows(min_row=6):
        for cell in row:
            cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=7):
        for cell in row:
            if cell.data_type != "f":
                cell.font = Font(name="Arial", size=10, color=INPUT_BLUE)


def add_intro(ws, title, note):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column or 1)
    ws["A1"] = title
    ws["A2"] = "样稿说明"
    ws["B2"] = note
    ws["A3"] = "容差"
    ws["B3"] = 0.01
    ws["B3"].number_format = "0.00"
    ws["B3"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    ws["B3"].comment = Comment("正式导出时记录核对所用容差。", "Codex")
    ws["A4"] = "数据口径"
    ws["B4"] = "示例数据；正式导出将保留源表行号、取数值与可复核公式。"
    ws["A2"].font = ws["A3"].font = ws["A4"].font = Font(
        name="Arial", size=10, bold=True, color=INK
    )
    ws["B2"].alignment = ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")


def make_rollforward(wb):
    ws = wb.active
    ws.title = "TB发生额与余额勾稽"
    headers = [
        "组", "源表行号", "科目编码", "科目名称", "期初余额", "TB借方发生额",
        "TB贷方发生额", "公式期末", "TB期末余额", "差异", "结论",
    ]
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.append(headers)
    rows = [
        ["04", 18, "1001", "库存现金", 100000, 420000, 370000, None, 150000],
        ["04", 27, "2202", "应付账款", -80000, 180000, 250000, None, -149999.5],
        ["08", 36, "6001", "主营业务收入", 0, 0, 500000, None, -500000],
    ]
    for index, row in enumerate(rows, start=7):
        ws.append(row + [None, None])
        ws[f"H{index}"] = f"=E{index}+F{index}-G{index}"
        ws[f"J{index}"] = f"=H{index}-I{index}"
        ws[f"K{index}"] = f'=IF(ABS(J{index})<=$B$3,"通过","差异")'
    add_intro(ws, "TB 发生额与余额勾稽", "逐科目验证：期初余额＋借方发生额－贷方发生额＝期末余额。")
    style_sheet(ws, [8, 11, 14, 24, 15, 16, 16, 15, 15, 14, 10])
    for row in range(7, ws.max_row + 1):
        for column in range(5, 11):
            ws.cell(row, column).number_format = "#,##0.00;[Red]-#,##0.00"
    ws.conditional_formatting.add(
        f"J7:J{ws.max_row}",
        CellIsRule(operator="greaterThan", formula=["$B$3"], fill=PatternFill("solid", fgColor=LIGHT_RED)),
    )


def make_tbje(wb):
    ws = wb.create_sheet("TB与JE发生额勾稽")
    headers = [
        "组", "主体", "科目编码", "科目名称", "出现在", "TB借方", "JE借方",
        "借方差异", "TB贷方", "JE贷方（已统一方向）", "贷方差异", "结论",
    ]
    for _ in range(6):
        ws.append([None] * len(headers))
    for column, value in enumerate(headers, start=1):
        ws.cell(6, column, value)
    rows = [
        ["04", "示例主体", "1001", "库存现金", "两边都有", 420000, 420000, None, 370000, 370000],
        ["04", "示例主体", "2202", "应付账款", "两边都有", 180000, 180000, None, 250000, 250000],
        ["08", "示例主体", "6602", "管理费用", "两边都有", 120000, 119500, None, 20000, 20000],
        ["08", "示例主体", "6001", "主营业务收入", "仅余额表有", 0, 0, None, 500000, 0],
    ]
    for index, row in enumerate(rows, start=7):
        ws.append(row + [None, None])
        ws[f"H{index}"] = f"=F{index}-G{index}"
        ws[f"K{index}"] = f"=I{index}-J{index}"
        ws[f"L{index}"] = f'=IF(MAX(ABS(H{index}),ABS(K{index}))<=$B$3,"通过","差异")'
    add_intro(
        ws,
        "TB 与 JE 发生额勾稽",
        "借、贷两侧分别对比。JE 贷方在引擎内统一为“正常贷方为正、红字冲销为负”，避免正负方向叠加。",
    )
    style_sheet(ws, [8, 16, 14, 24, 13, 15, 15, 14, 15, 21, 14, 10])
    for row in range(7, ws.max_row + 1):
        for column in range(6, 12):
            ws.cell(row, column).number_format = "#,##0.00;[Red]-#,##0.00"
    ws["J6"].comment = Comment(
        "如果源 JE 贷方列采用借正贷负口径，引擎会先翻回贷方正数；红字仍留在贷方侧冲减。",
        "Codex",
    )


def make_equation(wb):
    ws = wb.create_sheet("BS与PL勾稽")
    for _ in range(6):
        ws.append([None] * 7)
    for column, value in enumerate(
        ["时点", "会计要素", "归类金额", "平衡差异", "金额结论", "分类结论", "说明"],
        start=1,
    ):
        ws.cell(6, column, value)
    summary = [
        ("年初", "资产"), ("年初", "负债"), ("年初", "所有者权益"), ("年初", "成本"),
        ("年初", "损益"), ("年初", "合计（应为 0）"),
        ("年末", "资产"), ("年末", "负债"), ("年末", "所有者权益"), ("年末", "成本"),
        ("年末", "损益"), ("年末", "合计（应为 0）"),
    ]
    detail_start = 23
    detail_end = 30
    for index, (period, category) in enumerate(summary, start=7):
        ws.cell(index, 1, period)
        ws.cell(index, 2, category)
        if category.startswith("合计"):
            block_start = index - 5
            ws.cell(index, 3, f"=SUM(C{block_start}:C{index-1})")
            ws.cell(index, 4, f"=C{index}")
            ws.cell(index, 5, f'=IF(ABS(D{index})<=$B$3,"通过","差异")')
            ws.cell(index, 6, f'=IF(COUNTIF($F${detail_start}:$F${detail_end},"否")=0,"完整","待确认")')
            ws.cell(index, 7, "金额平衡与分类完整性分开判断")
        else:
            ws.cell(
                index,
                3,
                f'=SUMIFS($E${detail_start}:$E${detail_end},$A${detail_start}:$A${detail_end},A{index},$D${detail_start}:$D${detail_end},B{index},$F${detail_start}:$F${detail_end},"是")',
            )
    for column, value in enumerate(
        ["时点", "科目编码", "科目名称", "会计要素", "带符号余额", "是否纳入勾稽", "分类说明"],
        start=1,
    ):
        ws.cell(22, column, value)
    detail = [
        ["年初", "1001", "库存现金", "资产", 100000, "是", "按科目编码首位识别"],
        ["年初", "2202", "应付账款", "负债", -80000, "是", "按科目编码首位识别"],
        ["年初", "4001", "实收资本", "所有者权益", -20000, "是", "按科目编码首位识别"],
        ["年末", "1001", "库存现金", "资产", 150000, "是", "按科目编码首位识别"],
        ["年末", "2202", "应付账款", "负债", -150000, "是", "按科目编码首位识别"],
        ["年末", "6001", "主营业务收入", "损益", -500000, "是", "按科目编码首位识别"],
        ["年末", "6602", "管理费用", "损益", 500000, "是", "按科目编码首位识别"],
        ["年末", "X001", "自定义过渡科目", "未分类", 25000, "否", "编码无法自动归入会计要素"],
    ]
    for row in detail:
        ws.append(row)
    add_intro(
        ws,
        "BS 与 PL 勾稽",
        "按会计要素汇总带符号余额；金额是否为 0 与科目是否全部完成分类分别给结论。待分类科目保留在本页下方。",
    )
    style_sheet(ws, [10, 18, 20, 16, 16, 15, 32])
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A22:G{ws.max_row}"
    for cell in ws[22]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(7, ws.max_row + 1):
        ws.cell(row, 3).number_format = "#,##0.00;[Red]-#,##0.00"
        ws.cell(row, 4).number_format = "#,##0.00;[Red]-#,##0.00"
        ws.cell(row, 5).number_format = "#,##0.00;[Red]-#,##0.00"
    for row in (12, 18):
        for cell in ws[row]:
            cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL)
            cell.font = Font(name="Arial", size=10, bold=True, color=INK)
    for cell in ws[30]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    make_rollforward(wb)
    make_tbje(wb)
    make_equation(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    assert check.sheetnames == ["TB发生额与余额勾稽", "TB与JE发生额勾稽", "BS与PL勾稽"]
    assert sum(
        1
        for sheet in check.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ) >= 30
    check.close()
    print(OUTPUT)


if __name__ == "__main__":
    main()
