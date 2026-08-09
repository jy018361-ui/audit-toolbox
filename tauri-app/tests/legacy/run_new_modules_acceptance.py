from __future__ import annotations

import importlib
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas


ROOT = Path(__file__).resolve().parent.parents[2]
ARTIFACTS = ROOT / "tauri-app" / "tests" / "archive" / "new_modules_20260728"
AUDIPICK_DIR = ROOT / "modules" / "AudiPick"
ROLL_DIR = ROOT / "modules" / "audit-roll-forward"
WP_DIR = ROOT / "modules" / "wp-service-generator"


def make_contract_pdf() -> Path:
    target = ARTIFACTS / "AudiPick" / "样例收入合同.pdf"
    target.parent.mkdir(parents=True, exist_ok=True)
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    doc = canvas.Canvas(str(target), pagesize=A4)
    doc.setFont("STSong-Light", 14)
    lines = [
        "样例软件服务合同（仅用于功能验收）",
        "甲方：示例客户有限公司",
        "乙方：示例科技有限公司",
        "合同金额：人民币 110,000 元（含税）。",
        "履约期限：2026年1月1日至2026年12月31日。",
        "付款安排：签约后支付40%，验收后支付60%。",
        "履约义务：乙方提供软件实施、培训及一年期维护服务。",
        "验收条款：系统上线并完成用户验收测试后视为验收完成。",
        "违约责任：逾期交付每日按合同金额的0.05%支付违约金。",
    ]
    y = 800
    for line in lines:
        doc.drawString(60, y, line)
        y -= 32
    doc.save()
    return target


def import_from(module_dir: Path, module_name: str):
    sys.path.insert(0, str(module_dir))
    try:
        return importlib.import_module(module_name)
    finally:
        sys.path.pop(0)


def make_roll_workbook(path: Path, *, prior: bool) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "汇总"
    summary.append(["客户名称", "旧客户"])
    summary.append(["期末资产负债表日", datetime(2025, 12, 31)])
    summary.append(["记账本位币", "人民币"])
    summary.append(["适用会计准则", "企业会计准则"])
    summary.append(["PM", 0])
    summary.append(["TE", 0])
    summary.append(["SAD", 0])

    lead = wb.create_sheet("C.00 Lead")
    lead["B2"] = "客户名称"
    lead["C2"] = "旧客户"
    lead["B3"] = "资产负债表日"
    lead["C3"] = datetime(2025, 12, 31)
    lead["B4"] = "分析日期"
    lead["C4"] = datetime(2025, 1, 15)
    lead["B5"] = "TE"
    lead["C5"] = 0
    lead["B6"] = "SAD"
    lead["C6"] = 0
    lead["B9"] = "科目编码"
    lead["C9"] = "科目名称"
    lead["J9"] = "期末审定数"
    lead["K9"] = "上期末审定数"
    lead["B10"] = "1001"
    lead["C10"] = "银行存款"
    lead["J10"] = 123456.78 if prior else 0
    lead["K10"] = 0
    lead["B11"] = "合计"
    lead["J11"] = "=SUM(J10:J10)"
    lead["K11"] = "=SUM(K10:K10)"

    bkd = wb.create_sheet("C.00 BKD")
    bkd["A1"] = "开户银行"
    bkd["B1"] = "示例银行"
    wb.save(path)


def run_roll_forward() -> dict:
    target = ARTIFACTS / "Audit Roll Forward"
    template_dir = target / "templates"
    prior_dir = target / "prior"
    output_dir = target / "output"
    for folder in (template_dir, prior_dir, output_dir):
        folder.mkdir(parents=True, exist_ok=True)

    source_cfg = json.loads((ROLL_DIR / "subjects_config.json").read_text(encoding="utf-8"))
    config = {
        "version": "acceptance",
        "subjects": {"C": source_cfg["subjects"]["C"]},
    }
    config_path = target / "subjects_config.acceptance.json"
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    template_name = config["subjects"]["C"]["template_file"]
    template_path = template_dir / template_name
    prior_path = prior_dir / "C 货币资金 2025 样例公司.xlsx"
    make_roll_workbook(template_path, prior=False)
    make_roll_workbook(prior_path, prior=True)

    pmte = target / "PMTE_CRA_样例.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PMTE"
    ws.append(["公司", "Level", "RP", "PM", "TE", "SAD"])
    ws.append(["验收样例公司", "Low", 0.75, 1_000_000, 750_000, 50_000])
    cra = wb.create_sheet("CRA")
    cra.append(["", "C. 货币资金-存在性", "", "", "", "低", "", "", 0.75])
    wb.save(pmte)

    core = import_from(ROLL_DIR, "roll_forward_core")
    cra_support = import_from(ROLL_DIR, "cra_support")
    cra_text = (
        "科目名称\t认定\tCRA\t比例\t是否适用\n"
        "货币资金\t存在性\t低\t75%\tY\n"
        "货币资金\t完整性\t中等\t50%\tY\n"
        "固定资产\t存在性\t高\t25%\tY\n"
    )
    cra_records = cra_support.parse_cra_paste_text(cra_text, ["C"])
    results = core.process_multiple_subjects(
        ["C"],
        str(template_dir),
        str(prior_dir),
        str(pmte),
        "验收样例公司",
        "2026-12-31",
        str(output_dir),
        config_path=str(config_path),
        functional_currency="人民币",
        accounting_standard="企业会计准则",
        pm_value=1_000_000,
        te_value=750_000,
        sad_value=50_000,
        cra_records=cra_records,
        generate_summary=True,
    )
    code, success, message, output_path, warnings = results[0]
    if not success or not output_path:
        raise AssertionError(message)

    out = load_workbook(output_path, data_only=False)
    checks = {
        "subject": code,
        "success": success,
        "message": message,
        "output": output_path,
        "opening_balance": out["C.00 Lead"]["K10"].value,
        "company_name": out["汇总"]["B1"].value,
        "report_date": str(out["汇总"]["B2"].value),
        "has_summary": "Roll Forward Summary" in out.sheetnames,
        "cra_records": len(cra_records),
        "cra_write_records": sum(1 for item in cra_records if item.get("apply")),
        "warnings": list(warnings),
    }
    out.close()
    assert checks["opening_balance"] == 123456.78
    assert checks["company_name"] == "验收样例公司"
    assert checks["has_summary"]
    assert checks["cra_records"] == 3
    assert checks["cra_write_records"] == 2
    return checks


def run_wp_generator() -> dict:
    target = ARTIFACTS / "WP服务单"
    target.mkdir(parents=True, exist_ok=True)
    adapter = import_from(WP_DIR, "FY27_WP服务单生成工具")
    generator = import_from(WP_DIR, "generate_wp_project_workbook")
    template_path = adapter.ensure_template(target)
    template_book = load_workbook(template_path, data_only=False)
    source_ws = template_book["AUD2026"]
    headers = [source_ws.cell(1, col).value for col in range(1, source_ws.max_column + 1)]
    template_ws = generator.find_template_sheet(template_book)
    sections = [template_ws.cell(row, 2).value for row in range(5, 37)]
    reference_hours = [template_ws.cell(row, 8).value for row in range(5, 37)]
    template_book.close()

    raw = Workbook()
    ws = raw.active
    ws.title = "业务"
    ws.append(headers)
    header_map = {
        "".join(str(value or "").split()): index
        for index, value in enumerate(headers)
    }

    def add_row(name, order, hours, pre_start, pre_end, final_start, final_end):
        row = [None] * len(headers)
        values = {
            "EngagementName": name,
            "OutlookHours": hours,
            "BookingPeriodStart-预审": pre_start,
            "BookingPeriodEnd-预审": pre_end,
            "BookingPeriodStart-年审": final_start,
            "BookingPeriodEnd-年审": final_end,
            "WP服务单编号": order,
        }
        for key, value in values.items():
            row[header_map[key]] = value
        if len(row) >= 7:
            row[6] = row[6] or 3
        if len(row) >= 15:
            row[14] = "审计"
        if len(row) >= 16:
            row[15] = "验收经理"
        if len(row) >= 17:
            row[16] = datetime(2027, 3, 31)
        if len(row) >= 18:
            row[17] = f"REL-{order}"
        ws.append(row)

    add_row(
        "AUD 2026 验收项目",
        "WP-AUD-001",
        38.06,
        datetime(2026, 5, 1),
        datetime(2026, 6, 30),
        datetime(2026, 11, 1),
        datetime(2027, 3, 31),
    )
    add_row(
        "IPO 验收项目",
        "WP-IPO-001",
        12.32,
        datetime(2026, 6, 1),
        datetime(2026, 7, 31),
        datetime(2027, 1, 1),
        datetime(2027, 5, 31),
    )
    add_row(
        "IPO archive 验收项目",
        "WP-ARC-001",
        8,
        datetime(2026, 2, 1),
        datetime(2026, 3, 31),
        datetime(2026, 3, 1),
        datetime(2026, 4, 30),
    )
    add_row(
        "AUD 2025 历史项目",
        "WP-OLD-001",
        5,
        datetime(2025, 1, 1),
        datetime(2025, 2, 28),
        datetime(2025, 3, 1),
        datetime(2025, 4, 30),
    )
    raw_path = target / "FY27 WP服务单.xlsx"
    raw.save(raw_path)

    section_book = Workbook()
    section_ws = section_book.active
    section_ws.title = "Section List"
    section_ws.append(["Section", "Entity数量", "底稿数量", "预算调整", "所属WP服务单"])
    valid_sections = [
        (section, reference)
        for section, reference in zip(sections, reference_hours)
        if section
    ]
    first_section, _ = valid_sections[0]
    second_section, _ = valid_sections[1]
    section_ws.append([first_section, 1, 2, 0, "WP-AUD-001"])
    section_ws.append([second_section, 2, 1, 5, "WP-AUD-001"])
    section_ws.append([first_section, 2, 3, 2, "WP-IPO-001"])
    section_path = target / "FY27 section list.xlsx"
    section_book.save(section_path)

    result = adapter.run_generation(target)
    output_path = target / "FY27+WP服务单汇总.xlsx"
    generator.validate_formula_logic(output_path)
    out = load_workbook(output_path, data_only=False)
    formula_errors = []
    for sheet in out.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                    marker in cell.value
                    for marker in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}")
    service_sheets = [
        name for name in out.sheetnames
        if name not in {"服务方案索引", "AUD2026", "AUD2025", "IPO", "IPO archive"}
    ]
    out.close()
    assert result["services"] == 2
    assert result["split_aud2026_rows"] == 1
    assert result["split_ipo_rows"] == 1
    assert result["split_ipo_archive_rows"] == 1
    assert result["split_aud2025_rows"] == 1
    assert result["matched_section_orders"] == 2
    assert result["outlook_equal"] == 2
    assert not result["outlook_differences"]
    assert not formula_errors
    return {
        **result,
        "output": str(output_path),
        "service_sheets": service_sheets,
        "formula_errors": formula_errors,
    }


def main() -> int:
    ARTIFACTS.mkdir(parents=True, exist_ok=True)
    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "audipick": {
            "sample_pdf": str(make_contract_pdf()),
            "automated_tests": "run separately via npm test",
        },
    }
    failures = {}
    for name, runner in (
        ("audit_roll_forward", run_roll_forward),
        ("wp_service_generator", run_wp_generator),
    ):
        try:
            report[name] = runner()
        except Exception as exc:
            failures[name] = f"{type(exc).__name__}: {exc}"
    report["failures"] = failures
    report_path = ARTIFACTS / "acceptance_report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(report_path)
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
