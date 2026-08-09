from __future__ import annotations

import threading
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from audit_engine.confirmation_adapter import inspect_confirmation, process_confirmation
from audit_engine.errors import EngineError


@pytest.fixture()
def confirmation_file(tmp_path: Path) -> Path:
    path = tmp_path / "函证列表.xlsx"
    rows = [
        {
            "项目名称": "项目甲",
            "发函单位名称": "甲银行",
            "函证编号": "B-001",
            "函证类型": "银行",
            "发函模版": "银行格式一",
            "函证状态": "已收回",
            "发函签收时间": "2025-01-02 10:20:30",
            "询证项回函结果": "相符",
            "函证基准日": "2024-12-31",
        },
        {
            "项目名称": "项目甲",
            "发函单位名称": "乙银行",
            "函证编号": "B-002",
            "函证类型": "银行-电子函证",
            "发函模版": "银行格式二",
            "函证状态": "未发出",
            "发函签收时间": "",
            "询证项回函结果": "",
            "函证基准日": "2024-12-31",
        },
        {
            "项目名称": "项目乙",
            "发函单位名称": "客户A",
            "函证编号": "T-001",
            "函证类型": "往来询证函",
            "发函模版": "",
            "函证状态": "已收回",
            "发函签收时间": "2025-01-03 01:02:03",
            "询证项回函结果": "不符-金额差异",
            "函证基准日": "2024-12-31",
        },
        {
            "项目名称": "项目乙",
            "发函单位名称": "客户B",
            "函证编号": "T-002",
            "函证类型": "其他函证",
            "发函模版": "",
            "函证状态": "已发出",
            "发函签收时间": "bad date",
            "询证项回函结果": "",
            "函证基准日": "2025-06-30",
        },
    ]
    pd.DataFrame(rows).to_excel(path, index=False)
    return path


def test_inspect_reports_exact_legacy_dimensions_and_statistics(confirmation_file: Path):
    result = inspect_confirmation(confirmation_file, "both")

    assert result["dimensions"] == {"rows": 4, "columns": 9}
    assert result["missingColumns"] == []
    assert result["statistics"] == {
        "total": 4,
        "bank": 2,
        "trade": 2,
        "projects": 2,
        "units": 4,
        "baseDates": ["2024-12-31", "2025-06-30"],
    }
    assert result["willGenerate"] == {"bank": True, "trade": True}
    assert result["outputDirectory"].endswith("函证统计结果")


def test_mode_specific_validation_does_not_require_bank_only_columns_for_trade(tmp_path: Path):
    path = tmp_path / "往来.xlsx"
    pd.DataFrame(
        [{
            "函证类型": "往来",
            "函证编号": "1",
            "发函单位名称": "客户",
            "函证状态": "未发出",
            "发函签收时间": "",
            "询证项回函结果": "",
        }]
    ).to_excel(path, index=False)

    assert inspect_confirmation(path, "trade")["missingColumns"] == []
    assert inspect_confirmation(path, "both")["missingColumns"] == []
    assert inspect_confirmation(path, "bank")["missingColumns"] == []

    bank_path = tmp_path / "银行缺列.xlsx"
    pd.DataFrame([{"函证类型": "银行"}]).to_excel(bank_path, index=False)
    assert "函证基准日" in inspect_confirmation(bank_path, "bank")["missingColumns"]
    assert "发函模版" in inspect_confirmation(bank_path, "bank")["missingColumns"]

    result = process_confirmation(path, "both", lambda *_args: None, threading.Event())
    assert [report["status"] for report in result["reports"]] == ["skipped", "completed"]
    assert len(result["outputPaths"]) == 1


def test_both_reports_keep_legacy_sheets_totals_and_styling(confirmation_file: Path):
    events: list[tuple] = []
    result = process_confirmation(
        confirmation_file,
        "both",
        lambda *args: events.append(args),
        threading.Event(),
    )

    assert len(result["outputPaths"]) == 2
    assert [report["status"] for report in result["reports"]] == ["completed", "completed"]
    assert any("银行函证报告已生成" in event[3] for event in events)
    assert any("往来函证报告已生成" in event[3] for event in events)

    bank_path = next(Path(path) for path in result["outputPaths"] if "银行函证" in path)
    trade_path = next(Path(path) for path in result["outputPaths"] if "往来函证" in path)
    assert bank_path.parent.name == "函证统计结果"
    assert trade_path.parent == bank_path.parent

    bank = openpyxl.load_workbook(bank_path, data_only=False)
    assert bank.sheetnames[0:2] == ["按项目名称汇总", "按发函单位汇总"]
    assert "基准日_2024-12-31_按发函单位" in bank.sheetnames
    assert "基准日_2024-12-31_按项目名称" in bank.sheetnames
    unit = bank["按发函单位汇总"]
    assert unit.merged_cells.ranges
    assert unit["A1"].fill.fgColor.rgb.endswith("2E5B8F")
    assert unit.cell(unit.max_row, 1).value == "合计"
    assert unit.cell(unit.max_row, 2).value == 2
    assert len(unit.conditional_formatting) == 3
    bank.close()

    trade = openpyxl.load_workbook(trade_path, data_only=False)
    assert trade.sheetnames == ["按项目名称汇总", "按发函单位汇总"]
    unit = trade["按发函单位汇总"]
    assert unit.cell(unit.max_row, 1).value == "合计"
    assert unit.cell(unit.max_row, 2).value == 2
    assert len(unit.conditional_formatting) == 3
    trade.close()


def test_cancel_before_processing_is_cooperative(confirmation_file: Path):
    cancel = threading.Event()
    cancel.set()
    with pytest.raises(EngineError) as caught:
        process_confirmation(confirmation_file, "both", lambda *_args: None, cancel)
    assert caught.value.code == "JOB_CANCELLED"


def test_invalid_mode_is_rejected(confirmation_file: Path):
    with pytest.raises(EngineError) as caught:
        inspect_confirmation(confirmation_file, "unknown")
    assert caught.value.code == "CONFIRMATION_MODE_INVALID"
