import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
FA_LIST_ROOT = ROOT / "tools" / "fa_list"
if str(FA_LIST_ROOT) not in sys.path:
    sys.path.insert(0, str(FA_LIST_ROOT))

from exporter import Exporter
from summary_generator import SummaryGenerator


class SummaryNoiseDetectorTests(unittest.TestCase):
    """Cover the broadened noise detector. Old code only matched the cell value
    exactly equal to '合计/小计/total/subtotal'; in real-world源表 we also see
    variants like '总计/汇总/共计' and trailing-suffix subtotal rows like
    '运输工具合计/本年合计'."""

    def test_exact_keywords_still_match(self):
        for value in ("合计", "小计", "Total", "subtotal", " 合计 ", "合计:", "总计", "汇总", "共计"):
            with self.subTest(value=value):
                self.assertTrue(Exporter._is_summary_noise_value(value))

    def test_trailing_subtotal_suffix_is_noise(self):
        for value in (
            "运输工具合计",
            "电子设备 小计",
            "本年合计",
            "Office Subtotal",
        ):
            with self.subTest(value=value):
                self.assertTrue(Exporter._is_summary_noise_value(value))

    def test_real_category_names_are_not_flagged(self):
        for value in (
            "其他设备",
            "其他设备-工具/夹",   # 真实类别名，N 列就是这个，不能误伤
            "运输工具",
            "机器设备",
            "办公家具",
            "未分类",
            "智能回转库",
            None,
            "",
        ):
            with self.subTest(value=value):
                self.assertFalse(Exporter._is_summary_noise_value(value))

    def test_summary_generator_filter_uses_same_rules(self):
        # 同一份判定要在 summary_generator 里也生效，否则汇总表仍会捡起这类伪类别
        for value in ("合计", "运输工具合计", "总计"):
            with self.subTest(value=value):
                self.assertTrue(SummaryGenerator._is_summary_noise_value(value))
        for value in ("其他设备", "其他设备-工具/夹", "运输工具", "智能回转库"):
            with self.subTest(value=value):
                self.assertFalse(SummaryGenerator._is_summary_noise_value(value))

    def test_postprocess_summary_repositions_stray_total_column(self):
        """Reproduce the 'O 列出现合计' leak: summary_generator may emit a stray
        '合计' category column at the end if a 合计 row slipped past the filter.
        _postprocess_summary_sheet_only must move 合计 to column C and drop the
        stray duplicate."""
        wb = Workbook()
        ws = wb.active
        ws.title = "固定资产变动汇总表"
        # Mimic xlsxwriter output: row 1=header (合计 at C from totals + stray 合计 at end),
        # row 2=source labels (xlsxwriter writes 根据期初/期末卡片聚合 by default for categories),
        # rows 3+ = data
        ws.cell(1, 1).value = ""
        ws.cell(1, 2).value = ""
        ws.cell(1, 3).value = "合计"
        ws.cell(1, 4).value = "运输工具"
        ws.cell(1, 5).value = "其他设备-工具/夹"
        ws.cell(1, 6).value = "合计"   # stray duplicate at end — this is the bug
        ws.cell(2, 1).value = "变动项目"
        ws.cell(2, 2).value = "变动项目"
        ws.cell(2, 3).value = "计算"
        ws.cell(2, 4).value = "根据期初/期末卡片聚合"
        ws.cell(2, 5).value = "根据期初/期末卡片聚合"
        ws.cell(2, 6).value = "根据期初/期末卡片聚合"
        ws.cell(3, 1).value = "原值"
        ws.cell(3, 2).value = "年初余额"
        ws.cell(3, 3).value = 0
        ws.cell(3, 4).value = 100
        ws.cell(3, 5).value = 200
        ws.cell(3, 6).value = 300

        with tempfile.TemporaryDirectory() as tmp:
            file_path = str(Path(tmp) / "summary_repro.xlsx")
            wb.save(file_path)
            exporter = Exporter()
            exporter._postprocess_summary_sheet_only(file_path, summary_config=None)

            cleaned = load_workbook(file_path)
            cleaned_ws = cleaned["固定资产变动汇总表"]
            headers = [cleaned_ws.cell(1, c).value for c in range(1, cleaned_ws.max_column + 1)]
            self.assertEqual(headers[2], "合计")           # 合计 still at column C
            self.assertNotIn(
                "合计",
                headers[3:],
                msg="stray 合计 category column should be deleted from the right side",
            )
            # 合计列的第二行应为 "计算"，不应留着原来的 根据期初/期末卡片聚合
            self.assertEqual(cleaned_ws.cell(2, 3).value, "计算")

    def test_remove_summary_noise_rows_drops_total_label_row(self):
        exporter = Exporter()
        df = pd.DataFrame(
            {
                "资产类别": ["其他设备", "合计", "运输工具"],
                "资产编码": ["A1", "", "A2"],
                "原值": [100, 999, 200],
            }
        )
        cleaned = exporter._remove_summary_noise_rows(df, "合并数据", backup_sheet_name="固定资产变动汇总表")
        self.assertEqual(list(cleaned["资产类别"]), ["其他设备", "运输工具"])
        backup = getattr(exporter, "_summary_noise_backup", [])
        self.assertEqual(len(backup), 1)
        self.assertEqual(backup[0]["识别标识"], "合计")


if __name__ == "__main__":
    unittest.main()
