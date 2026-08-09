from __future__ import annotations

import tempfile
import threading
import unittest
import subprocess
import sys
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from launcher.llm_client import (
    LLMCombinedFAListResult,
    LLMMatchKeyReview,
    LLMSuggestion,
)
from audit_engine.handlers import (
    fa_export,
    fa_inspect,
    fa_match,
    fa_review,
    fa_supplement_inspect,
    fa_supplement_review,
)


class FaTauriExportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.begin = self.root / "begin.csv"
        self.end = self.root / "end.csv"
        self.output = self.root / "FA_List.xlsx"
        self.begin.write_text(
            "卡片编号,资产类别,资产名称,原值,累计折旧,开始使用日期,使用寿命,残值率\n"
            "A001,机器设备,设备甲,1000,200,2024-01-01,60,5%\n"
            "A002,电子设备,电脑乙,500,300,2023-01-01,36,5%\n",
            encoding="utf-8-sig",
        )
        self.end.write_text(
            "卡片编号,资产类别,资产名称,原值,累计折旧,开始使用日期,使用寿命,残值率\n"
            "A001,机器设备,设备甲,1000,400,2024-01-01,60,5%\n"
            "A003,运输设备,车辆丙,800,80,2026-01-01,48,5%\n",
            encoding="utf-8-sig",
        )
        self.params = {
            "beginPath": str(self.begin),
            "endPath": str(self.end),
            "beginKeys": ["卡片编号"],
            "endKeys": ["卡片编号"],
            "beginOriginalValue": "原值",
            "endOriginalValue": "原值",
            "beginDepreciation": "累计折旧",
            "endDepreciation": "累计折旧",
            "beginMapping": {
                "category": "资产类别",
                "name": "资产名称",
                "originalValue": "原值",
                "depreciation": "累计折旧",
                "startDate": "开始使用日期",
                "life": "使用寿命",
                "residualRate": "残值率",
            },
            "endMapping": {
                "category": "资产类别",
                "name": "资产名称",
                "originalValue": "原值",
                "depreciation": "累计折旧",
                "startDate": "开始使用日期",
                "life": "使用寿命",
                "residualRate": "残值率",
            },
            "outputPath": str(self.output),
            "balanceSheetDate": "2026-12-31",
        }

    def tearDown(self):
        self.temp.cleanup()

    @staticmethod
    def progress(*_args):
        return None

    def test_inspect_suggests_core_fields(self):
        result = fa_inspect(self.params)
        self.assertEqual(result["suggestedMapping"]["begin"]["matchKey"], "卡片编号")
        self.assertEqual(result["suggestedMapping"]["end"]["originalValue"], "原值")
        self.assertEqual(result["begin"]["displayName"], "begin.csv")

    def test_inspect_maps_residual_value_as_residual_rate_source(self):
        self.end.write_text(
            "卡片编号,资产类别,资产名称,原值,累计折旧,开始使用日期,使用寿命,残值\n"
            "A001,机器设备,设备甲,1000,200,2024-01-01,60,50\n",
            encoding="utf-8-sig",
        )
        result = fa_inspect(self.params)
        self.assertEqual(result["suggestedMapping"]["end"]["life"], "使用寿命")
        self.assertEqual(result["suggestedMapping"]["end"]["residualRate"], "残值")

    def test_llm_settings_loader_does_not_require_tkinter(self):
        script = (
            "import sys; sys.modules['tkinter'] = None; "
            "from launcher.llm_settings import load_llm_settings; "
            "assert isinstance(load_llm_settings(), dict)"
        )
        completed = subprocess.run(
            [sys.executable, "-c", script],
            cwd=Path(__file__).resolve().parent.parent,
            capture_output=True,
            text=True,
            timeout=20,
        )
        self.assertEqual(0, completed.returncode, completed.stderr)

    def test_review_is_optional_when_hub_llm_is_disabled(self):
        with patch("launcher.llm_settings.is_llm_enabled", return_value=False):
            result = fa_review(self.params)
        self.assertFalse(result["enabled"])
        self.assertTrue(result["passed"])

    def test_review_returns_original_llm_auto_fill_and_match_review(self):
        combined = LLMCombinedFAListResult(
            suggestions=[
                LLMSuggestion(
                    role="current_year_dep",
                    file_side="file2",
                    suggested_column="累计折旧",
                    confidence=0.9,
                    action="fill",
                    reason="样例字段",
                    review_warning="",
                )
            ],
            fa_review=[],
            match_review=LLMMatchKeyReview(
                status="ok",
                confidence=0.9,
                action="keep",
                reasons=[],
                suggested_file1_columns=["卡片编号"],
                suggested_file2_columns=["卡片编号"],
                suggestion_reason="",
            ),
        )
        with (
            patch("launcher.llm_settings.is_llm_enabled", return_value=True),
            patch("launcher.llm_settings.load_llm_settings", return_value={"api_key": "test"}),
            patch(
                "launcher.llm_client.generate_combined_fa_list_assistance",
                return_value=combined,
            ),
        ):
            result = fa_review(self.params)
        self.assertTrue(result["enabled"])
        self.assertEqual(result["autoApplied"][0]["role"], "current_year_dep")
        self.assertEqual(result["matchReview"]["action"], "keep")

    def test_review_explains_when_existing_mapping_needs_no_llm_fill(self):
        combined = LLMCombinedFAListResult(
            suggestions=[],
            fa_review=[],
            match_review=LLMMatchKeyReview(
                status="ok",
                confidence=0.95,
                action="keep",
                reasons=[],
                suggested_file1_columns=["卡片编号"],
                suggested_file2_columns=["卡片编号"],
                suggestion_reason="",
            ),
        )
        with (
            patch("launcher.llm_settings.is_llm_enabled", return_value=True),
            patch("launcher.llm_settings.load_llm_settings", return_value={"api_key": "test"}),
            patch(
                "launcher.llm_client.generate_combined_fa_list_assistance",
                return_value=combined,
            ),
        ):
            result = fa_review(self.params)
        self.assertEqual(
            "LLM 复核完成：现有脚本映射无需补充，匹配键已复核。",
            result["message"],
        )

    def test_supplement_review_is_optional_when_hub_llm_is_disabled(self):
        with patch("launcher.llm_settings.is_llm_enabled", return_value=False):
            result = fa_supplement_review({
                "addition": {"path": str(self.begin), "keys": ["卡片编号"]},
                "beginKeys": ["卡片编号"],
                "endKeys": ["卡片编号"],
            })
        self.assertFalse(result["enabled"])
        self.assertTrue(result["passed"])

    def test_inspect_detects_title_row_and_prefers_detail_sheet(self):
        begin = self.root / "begin_with_title.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["固定资产明细清单"])
        ws.append(["固定资产类别", "coding", "固定资产名称", "原值", "累计折旧"])
        ws.append(["机器设备", "1100000", "设备甲", 1000, 100])
        wb.save(begin)

        end = self.root / "end_with_summary.xlsx"
        wb = Workbook()
        summary = wb.active
        summary.title = "2512合计"
        summary.append(["固定资产汇总表"])
        summary.append(["公司", "公司代码", "资产类型描述", "原值合计"])
        summary.append(["公司甲", "A", "机器设备", 1000])
        detail = wb.create_sheet("2512")
        detail.append(["资产分类", "资产编码", "资产编码", "资产描述", "原值(期末)", "累计折旧"])
        detail.append(["机器设备", 0, "1100000", "设备甲", 1000, 200])
        detail.append(["机器设备", 0, "1100001", "设备乙", 2000, 300])
        detail.append(["机器设备", 0, "1100002", "设备丙", 3000, 400])
        wb.save(end)

        result = fa_inspect({"beginPath": str(begin), "endPath": str(end)})
        self.assertEqual(result["begin"]["detectedHeaderRow"], 2)
        self.assertEqual(result["end"]["selectedSheet"], "2512")
        self.assertEqual(result["suggestedMapping"]["begin"]["matchKey"], "coding")
        self.assertEqual(result["suggestedMapping"]["end"]["matchKey"], "资产编码.1")
        self.assertEqual(
            result["suggestedMapping"]["begin"]["matchKeys"],
            ["coding", "固定资产名称"],
        )
        self.assertEqual(
            result["suggestedMapping"]["end"]["matchKeys"],
            ["资产编码.1", "资产描述"],
        )
        self.assertEqual(result["suggestedMapping"]["end"]["category"], "资产分类")
        self.assertEqual(result["suggestedMapping"]["end"]["name"], "资产描述")

    def test_match_and_export_produce_real_workbook(self):
        preview = fa_match(self.params, self.progress, threading.Event())
        self.assertEqual(preview["stats"]["both"], 1)
        self.assertEqual(preview["stats"]["beginOnly"], 1)
        self.assertEqual(preview["stats"]["endOnly"], 1)
        result = fa_export(self.params, self.progress, threading.Event())
        self.assertEqual(result["outputPaths"], [str(self.output)])
        workbook = load_workbook(self.output, read_only=True, data_only=False)
        try:
            self.assertIn("FA List", workbook.sheetnames)
            self.assertIn("折旧期间", workbook.sheetnames)
            self.assertTrue(any("汇总" in name for name in workbook.sheetnames))
            headers = [cell.value for cell in next(workbook["合并数据"].iter_rows(max_row=1))]
            self.assertIn("卡片编号_期初", headers)
            self.assertIn("卡片编号_期末", headers)
            self.assertNotIn("卡片编号_文件1", headers)
        finally:
            workbook.close()

    def test_supplement_lists_are_normalized_aggregated_and_reported(self):
        addition = self.root / "addition.csv"
        disposal = self.root / "disposal.csv"
        addition.write_text(
            "卡片编号,新增方式,新增日期\n a003 ,购置,2026-01-01\nA003,转入,2026-02-01\nX999,购置,2026-03-01\n",
            encoding="utf-8-sig",
        )
        disposal.write_text(
            "卡片编号,处置方式,处置日期,处置原值,处置折旧\nA002,报废,2026-05-01,-200,-50\nA002,出售,2026-06-01,300,70\n",
            encoding="utf-8-sig",
        )
        inspect = fa_supplement_inspect({
            "path": str(addition),
            "referenceKeys": ["资产编码", "资产名称"],
        })
        self.assertEqual(inspect["suggestedMapping"]["additionMethod"], "新增方式")
        self.assertEqual(inspect["suggestedMapping"]["matchKeys"], ["卡片编号"])
        disposal_inspect = fa_supplement_inspect({
            "path": str(disposal),
            "referenceKeys": ["卡片编号"],
        })
        self.assertEqual(disposal_inspect["suggestedMapping"]["disposalOriginal"], "处置原值")
        self.assertEqual(
            disposal_inspect["suggestedMapping"]["disposalDepreciation"],
            "处置折旧",
        )
        params = {
            **self.params,
            "additionSupplement": {
                "path": str(addition), "keys": ["卡片编号"], "method": "新增方式", "date": "新增日期",
            },
            "disposalSupplement": {
                "path": str(disposal), "keys": ["卡片编号"], "method": "处置方式", "date": "处置日期",
                "originalValue": "处置原值", "depreciation": "处置折旧",
            },
        }
        preview = fa_match(params, self.progress, threading.Event())
        self.assertEqual(preview["stats"]["unmatchedAddition"], 1)
        self.assertEqual(preview["stats"]["unmatchedDisposal"], 0)
        row_a003 = next(row for row in preview["preview"] if row.get("卡片编号_文件2") == "A003")
        self.assertEqual(row_a003["新增方式_辅助_文件2"], "购置；转入")
        row_a002 = next(row for row in preview["preview"] if row.get("卡片编号_文件1") == "A002")
        self.assertEqual(row_a002["处置原值_辅助_文件1"], 500.0)
        exported = fa_export(params, self.progress, threading.Event())
        unmatched = self.output.parent / "[未匹配资产变动清单].xlsx"
        self.assertTrue(unmatched.is_file())
        self.assertIn("未匹配资产变动清单", exported["exportMessage"])


if __name__ == "__main__":
    unittest.main()
