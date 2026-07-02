import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
FA_LIST_ROOT = ROOT / "tools" / "fa_list"
if str(FA_LIST_ROOT) not in sys.path:
    sys.path.insert(0, str(FA_LIST_ROOT))

from file_handler import FileHandler


class FileHandlerSheetListTests(unittest.TestCase):
    def test_get_excel_sheets_excludes_hidden_xlsx_sheets(self):
        wb = Workbook()
        visible = wb.active
        visible.title = "可见主表"
        hidden = wb.create_sheet("隐藏底稿")
        hidden.sheet_state = "hidden"
        very_hidden = wb.create_sheet("深度隐藏")
        very_hidden.sheet_state = "veryHidden"
        wb.create_sheet("可见明细")

        with tempfile.TemporaryDirectory() as tmp:
            file_path = Path(tmp) / "sheets.xlsx"
            wb.save(file_path)

            success, error, sheets = FileHandler().get_excel_sheets(str(file_path))

        self.assertTrue(success, error)
        self.assertEqual(sheets, ["可见主表", "可见明细"])


if __name__ == "__main__":
    unittest.main()
