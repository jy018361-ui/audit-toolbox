from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from audit_engine.handlers import dispatch


def make_workbook(path: Path, *, prior: bool) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "汇总"
    summary.append(["客户名称", "旧客户"])
    summary.append(["期末资产负债表日", datetime(2025, 12, 31)])
    summary.append(["记账本位币", "人民币"])
    summary.append(["适用会计准则", "企业会计准则"])
    summary.append(["PM", 0])
    summary.append(["TE", 0])
    summary.append(["SAD", 0])
    lead = wb.create_sheet("C.00 Lead")
    lead["B2"], lead["C2"] = "客户名称", "旧客户"
    lead["B3"], lead["C3"] = "资产负债表日", datetime(2025, 12, 31)
    lead["B4"], lead["C4"] = "分析日期", datetime(2025, 1, 15)
    lead["B5"], lead["C5"] = "TE", 0
    lead["B6"], lead["C6"] = "SAD", 0
    lead["B9"], lead["C9"] = "科目编码", "科目名称"
    lead["J9"], lead["K9"] = "期末审定数", "上期末审定数"
    lead["B10"], lead["C10"] = "1001", "银行存款"
    lead["J10"], lead["K10"] = (123456.78 if prior else 0), 0
    lead["B11"], lead["J11"], lead["K11"] = "合计", "=SUM(J10:J10)", "=SUM(K10:K10)"
    bkd = wb.create_sheet("C.00 BKD")
    bkd["A1"], bkd["B1"] = "开户银行", "示例银行"
    wb.save(path)


class RollForwardTauriTest(unittest.TestCase):
    def test_detect_subjects_matches_legacy_filename_rules(self) -> None:
        with tempfile.TemporaryDirectory() as root_text:
            root = Path(root_text)
            for name in (
                "C 货币资金 2025.xlsx",
                "U_EXP 期间费用 2025.xlsx",
                "U_EXP VC&VD 费用底稿 2025.xlsx",
                "~$J1 临时文件.xlsx",
            ):
                (root / name).touch()
            result = dispatch("roll_forward.detect_subjects", {"priorPath": str(root)})
            self.assertIn("C", result["subjects"])
            self.assertIn("Uexp", result["subjects"])
            self.assertIn("UexpVCVD", result["subjects"])
            self.assertNotIn("J1", result["subjects"])
            self.assertEqual(result["scannedWorkbookCount"], 3)

    def test_detect_subjects_rejects_non_xlsx_single_file(self) -> None:
        with tempfile.TemporaryDirectory() as root_text:
            path = Path(root_text) / "prior.csv"
            path.touch()
            with self.assertRaises(Exception):
                dispatch("roll_forward.detect_subjects", {"priorPath": str(path)})

    def test_project_export_is_atomic_and_keeps_unicode(self) -> None:
        with tempfile.TemporaryDirectory() as root_text:
            output = Path(root_text) / "测试项目"
            result = dispatch(
                "roll_forward.project_export",
                {
                    "outputPath": str(output),
                    "project": {"project_name": "中文项目", "companies": [{"name": "甲公司"}]},
                },
            )
            exported = Path(result["outputPaths"][0])
            self.assertEqual(exported.suffix, ".auditproj")
            self.assertIn("中文项目", exported.read_text(encoding="utf-8"))
            self.assertFalse(list(exported.parent.glob(".*.tmp")))

    def test_catalog_cra_validation_and_process(self) -> None:
        with tempfile.TemporaryDirectory() as root_text:
            root = Path(root_text)
            template_dir, prior_dir, output_dir = root / "templates", root / "prior", root / "output"
            for folder in (template_dir, prior_dir, output_dir):
                folder.mkdir()
            catalog = dispatch("roll_forward.catalog", {})
            subject = next(row for row in catalog["subjects"] if row["code"] == "C")
            make_workbook(template_dir / subject["templateFile"], prior=False)
            make_workbook(prior_dir / "C 货币资金 2025 样例公司.xlsx", prior=True)
            cra = dispatch(
                "roll_forward.cra.parse",
                {
                    "text": "科目名称\t认定\tCRA\t比例\n货币资金\t存在性\t低\t75%",
                    "subjectCodes": ["C"],
                },
            )
            params = {
                "templateDir": str(template_dir),
                "priorDir": str(prior_dir),
                "outputDir": str(output_dir),
                "subjectCodes": ["C"],
                "companyName": "验收样例公司",
                "bsDate": "2026-12-31",
                "functionalCurrency": "人民币",
                "accountingStandard": "企业会计准则",
                "pmValue": 1_000_000,
                "teValue": 750_000,
                "sadValue": 50_000,
                "craRecords": cra["records"],
                "generateSummary": True,
            }
            check = dispatch("roll_forward.validate", params)
            self.assertTrue(check["valid"], check)
            self.assertFalse(check["llmRequested"])
            llm_missing = dispatch(
                "roll_forward.validate",
                {**params, "llmEnhanced": True, "__llmOptions": {}},
            )
            self.assertFalse(llm_missing["valid"])
            self.assertFalse(llm_missing["llmReady"])
            llm_ready = dispatch(
                "roll_forward.validate",
                {
                    **params,
                    "llmEnhanced": True,
                    "__llmOptions": {
                        "enabled": True,
                        "api_type": "openai",
                        "api_key": "test-only",
                        "model": "test-model",
                        "base_url": "https://example.invalid/v1",
                    },
                },
            )
            self.assertTrue(llm_ready["llmReady"])
            self.assertTrue(llm_ready["valid"])
            result = dispatch("roll_forward.process", params)
            self.assertTrue(result["results"][0]["success"], result)
            self.assertFalse(list(output_dir.glob("*.partial.xlsx")))
            output = Path(result["outputPaths"][0])
            wb = load_workbook(output, data_only=False)
            self.assertEqual(wb["C.00 Lead"]["K10"].value, 123456.78)
            self.assertIn("Roll Forward Summary", wb.sheetnames)
            wb.close()

    def test_invalid_date_is_reported_without_starting(self) -> None:
        with tempfile.TemporaryDirectory() as root_text:
            root = Path(root_text)
            for name in ("templates", "prior", "output"):
                (root / name).mkdir()
            result = dispatch(
                "roll_forward.validate",
                {
                    "templateDir": str(root / "templates"),
                    "priorDir": str(root / "prior"),
                    "outputDir": str(root / "output"),
                    "subjectCodes": ["C"],
                    "companyName": "公司",
                    "bsDate": "错误日期",
                },
            )
            self.assertFalse(result["valid"])
            self.assertFalse(result["dateValid"])
            compact = dispatch(
                "roll_forward.validate",
                {
                    "templateDir": str(root / "templates"),
                    "priorDir": str(root / "prior"),
                    "outputDir": str(root / "output"),
                    "subjectCodes": ["C"],
                    "companyName": "公司",
                    "bsDate": "20261231",
                },
            )
            self.assertTrue(compact["dateValid"])


if __name__ == "__main__":
    unittest.main()
