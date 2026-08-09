from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from audit_engine.errors import EngineError
from audit_engine.handlers import HANDLERS, dispatch
from audit_engine.jobs import JobManager


class TauriEngineContractTests(unittest.TestCase):
    @staticmethod
    def _strict_json_loads(payload: str):
        def reject_non_finite(value: str):
            raise ValueError(f"non-standard JSON number: {value}")

        return json.loads(payload, parse_constant=reject_non_finite)

    def test_jsonl_health_handshake(self):
        process = subprocess.Popen(
            [sys.executable, "-m", "audit_engine"],
            cwd=Path(__file__).resolve().parent.parent,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            text=True,
            encoding="utf-8",
        )
        try:
            self.assertEqual("ready", json.loads(process.stdout.readline())["type"])
            request = {"protocol": 1, "id": "health", "method": "system.health", "params": {}}
            process.stdin.write(json.dumps(request) + "\n"); process.stdin.flush()
            response = json.loads(process.stdout.readline())
            self.assertTrue(response["ok"])
            self.assertEqual("ok", response["result"]["status"])
        finally:
            process.stdin.close()
            process.wait(timeout=5)
            process.stdout.close()

    def test_jsonl_fa_inspect_with_blank_numeric_cells_is_strict_json(self):
        """Blank Excel cells must never become NaN on the Rust JSONL boundary."""
        with tempfile.TemporaryDirectory(prefix="tauri-fa-json-") as temp:
            root = Path(temp)
            paths = []
            for period in ("期初", "期末"):
                path = root / f"{period}.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "固定资产"
                sheet.append(["卡片编号", "资产名称", "原值"])
                sheet.append(["A001", "电脑", None])
                sheet.append(["A002", "打印机", 1200.0])
                workbook.save(path)
                paths.append(path)

            process = subprocess.Popen(
                [sys.executable, "-m", "audit_engine"],
                cwd=Path(__file__).resolve().parent.parent,
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                text=True,
                encoding="utf-8",
            )
            try:
                self.assertEqual("ready", self._strict_json_loads(process.stdout.readline())["type"])
                request = {
                    "protocol": 1,
                    "id": "fa-inspect",
                    "method": "fa.inspect",
                    "params": {"beginPath": str(paths[0]), "endPath": str(paths[1])},
                }
                process.stdin.write(json.dumps(request, ensure_ascii=False) + "\n")
                process.stdin.flush()
                response = self._strict_json_loads(process.stdout.readline())
                self.assertTrue(response["ok"])
                self.assertIsNone(response["result"]["begin"]["preview"][0][2])
                self.assertIsNone(response["result"]["end"]["preview"][0][2])
            finally:
                process.stdin.close()
                process.wait(timeout=10)
                process.stdout.close()

    def test_only_registered_methods_are_dispatched(self):
        expected = {
            "file_list", "wp", "confirmation", "fa",
            "roll_forward", "audipick",
        }
        self.assertEqual(expected, {name.split(".", 1)[0] for name in HANDLERS})
        with self.assertRaises(EngineError) as context:
            dispatch("shell.execute", {"command": "whoami"})
        self.assertEqual("METHOD_NOT_ALLOWED", context.exception.code)

    def test_file_list_scan_and_export(self):
        with tempfile.TemporaryDirectory(prefix="tauri-file-list-") as temp:
            root = Path(temp) / "中文目录"
            (root / "一级" / "二级").mkdir(parents=True)
            (root / "根目录.txt").write_text("root", encoding="utf-8")
            (root / "一级" / "二级" / "样例.txt").write_text("sample", encoding="utf-8")
            output = Path(temp) / "清单.xlsx"
            scan = dispatch("file_list.scan", {"sourceDir": str(root)})
            self.assertEqual(2, scan["fileCount"])
            self.assertEqual(2, scan["maxDepth"])
            result = dispatch(
                "file_list.export",
                {"sourceDir": str(root), "outputPath": str(output)},
                cancel=threading.Event(),
            )
            self.assertEqual([str(output)], result["outputPaths"])
            workbook = load_workbook(output, data_only=False)
            self.assertEqual("文件清单", workbook.active.title)
            self.assertEqual(3, workbook.active.max_row)
            self.assertTrue(any(cell.hyperlink is not None for row in workbook.active.iter_rows(min_row=2) for cell in row))
            workbook.close()

    def test_wp_validation_reports_missing_inputs(self):
        with tempfile.TemporaryDirectory(prefix="tauri-wp-") as temp:
            result = dispatch("wp.validate", {"folder": temp})
            self.assertFalse(result["valid"])
            self.assertEqual(["FY27 WP服务单.xlsx", "FY27 section list.xlsx"], result["missing"])

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_combines_multiple_workbooks(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-") as temp:
            root = Path(temp)
            sources = []
            for index in (1, 2):
                path = root / f"输入{index}.xlsx"
                workbook = Workbook(); sheet = workbook.active; sheet.title = "数据"
                sheet.append(["编号", "金额"]); sheet.append([index, index * 100]); workbook.save(path)
                sources.append(str(path))
            output = root / "合并.xlsx"
            result = dispatch("excel_merger.merge", {
                "inputPaths": sources, "outputPath": str(output), "outputMode": "one_sheet",
                "direction": "vertical", "sheetAction": "default", "addHyperlinks": False,
            }, cancel=threading.Event())
            self.assertEqual(2, result["inputFiles"])
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(["Merged"], workbook.sheetnames)
                self.assertEqual(4, workbook["Merged"].max_row)
                self.assertEqual("输入1.xlsx", workbook["Merged"]["A1"].value)
                self.assertEqual("编号", workbook["Merged"]["B1"].value)
            finally:
                workbook.close()

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_inspects_sheets_and_scans_folder(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-inspect-") as temp:
            root = Path(temp); nested = root / "子目录"; nested.mkdir()
            workbook = Workbook(); workbook.active.title = "明细"; workbook.create_sheet("汇总")
            path = nested / "双Sheet.xlsx"; workbook.save(path)
            (nested / "忽略.pdf").write_bytes(b"pdf")
            scan = dispatch("excel_merger.scan_folder", {"folder": str(root)})
            self.assertEqual([str(path)], scan["inputPaths"])
            expanded = dispatch("excel_merger.expand_paths", {"paths": [str(root), str(path)]})
            self.assertEqual([str(path)], expanded["inputPaths"])
            inspected = dispatch("excel_merger.inspect", {"inputPaths": scan["inputPaths"]})
            self.assertEqual(["明细", "汇总"], inspected["availableSheets"])
            self.assertIsNone(inspected["files"][0]["error"])

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_defaults_output_to_input_directory_and_avoids_overwrite(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-output-dir-") as temp:
            root = Path(temp); source = root / "输入.xlsx"
            workbook = Workbook(); workbook.active.append(["编号", "金额"]); workbook.active.append([1, 100]); workbook.save(source)
            params = {
                "inputPaths": [str(source)], "outputDirectory": str(root), "outputFormat": "xlsx",
                "outputMode": "one_sheet", "direction": "vertical", "sheetAction": "default",
                "addHyperlinks": False,
            }
            first = dispatch("excel_merger.merge", params, cancel=threading.Event())
            second = dispatch("excel_merger.merge", params, cancel=threading.Event())
            first_path = Path(first["outputPaths"][0]); second_path = Path(second["outputPaths"][0])
            self.assertEqual(root, first_path.parent); self.assertEqual(".xlsx", first_path.suffix)
            self.assertNotEqual(first_path, second_path); self.assertTrue(second_path.exists())

            csv_result = dispatch("excel_merger.merge", {**params, "outputFormat": "csv"}, cancel=threading.Event())
            csv_path = Path(csv_result["outputPaths"][0])
            self.assertEqual(root, csv_path.parent); self.assertEqual(".csv", csv_path.suffix)

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_horizontal_mode_and_sheet_selection(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-horizontal-") as temp:
            root = Path(temp); sources = []
            for index in (1, 2):
                path = root / f"横向{index}.xlsx"; workbook = Workbook(); detail = workbook.active; detail.title = "明细"
                detail.append(["编号", "金额"]); detail.append([index, index * 10])
                summary = workbook.create_sheet("汇总"); summary.append(["项目", "值"]); summary.append([f"P{index}", index])
                workbook.save(path); sources.append(str(path))
            output = root / "横向合并.xlsx"
            dispatch("excel_merger.merge", {
                "inputPaths": sources, "outputPath": str(output), "outputMode": "one_sheet",
                "direction": "horizontal", "sheetAction": "match_selected", "targetSheets": ["汇总"],
                "addHyperlinks": False,
            }, cancel=threading.Event())
            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook["合并结果"]
                self.assertEqual(4, sheet.max_column)
                self.assertEqual("项目", sheet.cell(2, 1).value)
                self.assertEqual("P1", sheet.cell(3, 1).value)
                self.assertEqual("P2", sheet.cell(3, 3).value)
            finally:
                workbook.close()

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_honors_preflight_cancellation(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-cancel-") as temp:
            path = Path(temp) / "输入.xlsx"; workbook = Workbook(); workbook.save(path)
            cancel = threading.Event(); cancel.set()
            with self.assertRaises(EngineError) as context:
                dispatch("excel_merger.merge", {
                    "inputPaths": [str(path)], "outputPath": str(Path(temp) / "输出.xlsx"),
                    "outputMode": "one_sheet", "direction": "vertical", "sheetAction": "default",
                }, cancel=cancel)
            self.assertEqual("JOB_CANCELLED", context.exception.code)

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_workbook_mode_preserves_excel_semantics(self):
        with tempfile.TemporaryDirectory(prefix="tauri-merger-workbook-") as temp:
            root = Path(temp); sources = []
            for index in (1, 2):
                path = root / f"底稿{index}.xlsx"; workbook = Workbook(); sheet = workbook.active; sheet.title = "Data"
                sheet["A1"] = "标题"; sheet.merge_cells("A1:B1")
                sheet["A1"].fill = PatternFill("solid", fgColor="FFFF00"); sheet["A1"].font = Font(bold=True)
                sheet["A2"] = index; sheet["B2"] = "=A2*10"
                workbook.create_sheet("不选择")["A1"] = "skip"; workbook.save(path); sources.append(str(path))
            output = root / "多Sheet合并.xlsx"
            result = dispatch("excel_merger.merge", {
                "inputPaths": sources, "outputPath": str(output), "outputMode": "one_workbook",
                "sheetAction": "default", "addHyperlinks": True,
            }, cancel=threading.Event())
            self.assertIsNone(result["fallbackWarning"])
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(["Reference", "底稿1", "底稿2"], workbook.sheetnames)
                copied = workbook["底稿1"]
                self.assertEqual("=A2*10", copied["B2"].value)
                self.assertEqual("FFFFFF00", copied["A1"].fill.fgColor.rgb)
                self.assertTrue(copied["A1"].font.bold)
                self.assertIn("A1:B1", {str(value) for value in copied.merged_cells.ranges})
                self.assertIsNotNone(workbook["Reference"]["B2"].hyperlink)
            finally:
                workbook.close()

    @unittest.skip("Excel 合并已迁移至 Rust worker")
    def test_excel_merger_runs_in_isolated_job_worker(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp); sources = []
            for index in (1, 2):
                path = root / f"隔离任务{index}.xlsx"
                workbook = Workbook(); sheet = workbook.active
                sheet.append(["编号", "金额"]); sheet.append([index, index * 100])
                workbook.save(path); sources.append(str(path))
            output = root / "隔离任务结果.xlsx"
            events = []; finished = threading.Event()

            def emit(event):
                events.append(event)
                if event.get("payload", {}).get("phase") in {"completed", "failed", "cancelled"}:
                    finished.set()

            manager = JobManager(emit)
            job_id = manager.start("excel_merger.merge", {
                "inputPaths": sources, "outputPath": str(output),
                "outputMode": "one_sheet", "direction": "vertical",
                "sheetAction": "default", "addHyperlinks": False,
            })
            self.assertTrue(finished.wait(30), "isolated worker did not finish")
            phases = [event["payload"]["phase"] for event in events]
            self.assertIn("queued", phases); self.assertIn("running", phases)
            self.assertEqual("completed", phases[-1]); self.assertTrue(output.exists())
            deadline = time.monotonic() + 2
            while job_id in manager.jobs and time.monotonic() < deadline:
                time.sleep(0.01)
            self.assertNotIn(job_id, manager.jobs)

    def test_errors_are_safe_for_frontend(self):
        error = EngineError("SAMPLE", "用户可读信息", detail="technical")
        payload = json.loads(json.dumps(error.as_dict(), ensure_ascii=False))
        self.assertEqual("SAMPLE", payload["code"])
        self.assertEqual("用户可读信息", payload["userMessage"])
        self.assertEqual(12, len(payload["diagnosticId"]))


if __name__ == "__main__":
    unittest.main()
