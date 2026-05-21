# -*- coding: utf-8 -*-
"""
Timesheet Pivot GUI

目标：
1. 扫描目标文件夹并让用户选择文件。
2. 支持多 Sheet 选择，默认第 1 行为标题行。
3. 参考 JE Net 交互，支持 by经理 / by项目 两套默认透视配置。
4. 支持条件筛选（字段 + 多值）。
5. 一键导出透视结果。
"""

from __future__ import annotations

import csv
import importlib.util
import os
import json
import shutil
import sys
import tempfile
import threading
import time
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence, Set, Tuple

import tkinter as tk
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
from openpyxl import load_workbook as ox_load_workbook
try:
    from python_calamine import load_workbook as cal_load_workbook
except Exception:
    cal_load_workbook = None
from tkinter import filedialog, messagebox, ttk
try:
    import polars as pl
except Exception:
    pl = None


DEFAULT_FOLDER = (
    r"\\Cnshausrfl025\025sha00001\G\GTH Assurance\!!! GDS Assurance\3. RM\Reporting相关\Report Data\Timesheet summary\FY26"
)
SUPPORTED_EXTS = {".xlsx", ".xlsm", ".xls", ".csv", ".txt"}
DEFAULT_FILTER_FIELD = "Department Name"
DEFAULT_FILTER_VALUE = "ASU Delivery Center ZZ-WP"
NO_FILTER_OPTION = "（无筛选）"
DEFAULT_EXPORT_FORMAT = ".xlsx"
EXPORT_TIMEOUT_SEC = 0
MAX_PIVOT_COLUMN_VALUES = 180
MANUAL_FILTER_CACHE = {
    "Department Name": [
        "FAAS-Financial&AccountingAdv",
        "ASU Delivery Center DL - WP",
        "ASU Delivery Center DL - AFS",
        "ASU Delivery Center JN - WP",
        "ASU Assurance support China",
        "ASU Delivery Center ZZ-WP",
        "ASU Delivery Center CD - FAAS",
        "ASU Delivery Center XA – WP",
        "ASU Delivery Center KM - WP",
        "ASU Delivery Center CS-WP",
        "ASU Delivery Center DL – Confirmation",
        "Assurance Development",
        "ASU Delivery Center DL - DDP",
        "ASU Delivery Center XA – DDP",
        "ASU Supp Resource&Produc Mgmt",
        "ASU Delivery Center DL",
        "ASU Delivery Center DL - Digital - Contractor",
        "ASU Delivery Center DL - CES",
        "ASU Delivery Center HZ - WP",
        "ASU Delivery Center DL - Lease",
        "ASU Delivery Center KM - ECL",
        "ASU Delivery Center KM - Lease",
        "ASU Delivery Center HZ - Digital",
        "ASU Delivery Center DL - Digital - Core",
        "Audit Assurance Digital",
        "Auto Digital",
        "FAAS Digital",
    ],
}


class UserCancelledError(Exception):
    pass


@dataclass
class DataSource:
    original_path: str
    data_path: str
    file_type: str
    sheet_name: str
    sep: str
    encoding: str
    header_row: int
    headers: List[str]


@dataclass
class TaskContext:
    start_ts: float
    timeout_sec: int
    cancel_event: Optional[threading.Event] = None
    report_stage: Optional[Callable[[str], None]] = None
    report_progress: Optional[Callable[[int, Optional[str]], None]] = None
    _last_progress: int = 0
    timings: Dict[str, float] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)

    def stage(self, text: str) -> None:
        if self.report_stage is not None:
            self.report_stage(text)

    def progress(self, pct: int, text: Optional[str] = None) -> None:
        v = max(0, min(100, int(pct)))
        if v < self._last_progress:
            v = self._last_progress
        self._last_progress = v
        if self.report_progress is not None:
            self.report_progress(v, text)

    def check_abort(self) -> None:
        if self.cancel_event is not None and self.cancel_event.is_set():
            raise UserCancelledError("用户取消操作。")
        if self.timeout_sec > 0 and (time.monotonic() - self.start_ts) > self.timeout_sec:
            raise TimeoutError(f"操作超时（>{self.timeout_sec}秒），请缩小筛选范围后重试。")

    def add_timing(self, key: str, sec: float) -> None:
        try:
            self.timings[key] = self.timings.get(key, 0.0) + max(0.0, float(sec))
        except Exception:
            pass

    def note(self, msg: str) -> None:
        try:
            text = str(msg).strip()
            if text and text not in self.notes:
                self.notes.append(text)
        except Exception:
            pass


@dataclass
class PostProcessState:
    columns: List[str]
    row_fields: List[str]
    idx_map: Dict[str, int] = field(default_factory=dict)
    repeat_indices: List[int] = field(default_factory=list)
    last_vals: Dict[int, str] = field(default_factory=dict)


def init_postprocess_state(columns: List[str], row_fields: Sequence[str]) -> PostProcessState:
    state = PostProcessState(
        columns=columns[:],
        row_fields=list(row_fields),
    )
    state.idx_map = {c: i for i, c in enumerate(columns)}
    state.repeat_indices = [state.idx_map[c] for c in row_fields if c in state.idx_map]
    state.last_vals = {i: "" for i in state.repeat_indices}
    return state


def consume_postprocess_row(state: PostProcessState, raw_row: Sequence[object]) -> List[List[object]]:
    row = list(raw_row)
    for idx in state.repeat_indices:
        v = "" if row[idx] is None else str(row[idx])
        if v == "":
            row[idx] = state.last_vals[idx]
        else:
            state.last_vals[idx] = v
    return [row]


def finalize_postprocess_state(state: PostProcessState) -> List[List[object]]:
    return []


class ProgressWindow(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        title: str = "正在处理",
        on_cancel: Optional[Callable[[], None]] = None,
    ):
        super().__init__(parent)
        self.title(title)
        self.geometry("520x180")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - 260
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 90
        self.geometry(f"+{x}+{y}")
        self._on_cancel = on_cancel

        self.lbl = tk.Label(self, text="正在处理数据，请稍候...", pady=10)
        self.lbl.pack()
        self.stage_var = tk.StringVar(value="准备中...")
        self.stage_lbl = tk.Label(self, textvariable=self.stage_var, fg="#1f4e79")
        self.stage_lbl.pack()
        self.pb = ttk.Progressbar(self, orient="horizontal", length=450, mode="determinate", maximum=100, value=0)
        self.pb.pack(pady=10)
        self.percent_var = tk.StringVar(value="0%")
        tk.Label(self, textvariable=self.percent_var, fg="#1f4e79").pack(pady=(0, 4))
        if self._on_cancel is not None:
            ttk.Button(self, text="取消", command=self._cancel).pack(pady=(0, 8))

    def set_stage(self, text: str) -> None:
        self.stage_var.set(text)
        self.update_idletasks()

    def set_progress(self, pct: int, text: Optional[str] = None) -> None:
        v = max(0, min(100, int(pct)))
        self.pb["value"] = v
        self.percent_var.set(f"{v}%")
        if text:
            self.stage_var.set(text)
        self.update_idletasks()

    def _cancel(self) -> None:
        if self._on_cancel is not None:
            try:
                self._on_cancel()
            except Exception:
                pass
        self.stage_var.set("正在取消，请稍候...")
        self.update_idletasks()

    def close(self) -> None:
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


class TimesheetPivotApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Timesheet Pivot Tool ( !!!导出需要5分钟!!!)")
        self._init_window_geometry()

        self.folder_var = tk.StringVar(value=DEFAULT_FOLDER)
        self.file_path_var = tk.StringVar(value="")
        self.header_row_var = tk.IntVar(value=1)
        self.source_info_var = tk.StringVar(value="未加载文件")
        self.agg_var = tk.StringVar(value="sum")
        self.fast_load_var = tk.BooleanVar(value=True)
        self.pyarrow_available = bool(importlib.util.find_spec("pyarrow"))
        self.polars_available = pl is not None and bool(importlib.util.find_spec("fastexcel"))
        self.source_df = None
        self.source_table_name = "source_data"
        self.cache_dir = Path(tempfile.gettempdir()) / "timesheet_polars_cache"
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.cache_manifest_path = self.cache_dir / "cache_manifest.json"
        self.cache_manifest = self._load_cache_manifest()
        self.source_loaded = False
        self.loaded_source_key = ""
        self.loaded_columns: Set[str] = set()
        self.loaded_full = False
        self.load_lock = threading.Lock()
        self.prewarm_thread: Optional[threading.Thread] = None
        self.prewarm_running = False
        self.prewarm_done = False
        self.preload_thread: Optional[threading.Thread] = None
        self.preload_running = False
        self.preload_done = False
        self.pivot_result_cache: Dict[str, Tuple[List[str], List[List[object]]]] = {}
        self.distinct_cache_tables: Dict[str, str] = {}

        self.data_source: Optional[DataSource] = None
        self.available_fields: List[str] = []
        self.row_fields: List[str] = []
        self.filters: Dict[str, Set[str]] = {}
        self.filter_values_cache: Dict[str, List[str]] = {}
        self.shadow_cache: Dict[str, str] = {}
        self.local_excel_copy_cache: Dict[str, str] = {}

        self.thread_event = threading.Event()
        self.user_sheet_choice: Optional[str] = None
        self.progress_win: Optional[ProgressWindow] = None
        self.busy = False
        self.current_cancel_event: Optional[threading.Event] = None

        self._build_scrollable_container()
        self._build_ui()
        self._bind_root_shortcuts()
        self.root.after(120, self._prompt_select_file_on_start)

    def _init_window_geometry(self) -> None:
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w = max(980, min(1360, int(sw * 0.78)))
        h = max(360, min(560, int(sh * 0.50)))
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")
        self.root.minsize(920, 320)

    def _build_scrollable_container(self) -> None:
        self.canvas = tk.Canvas(self.root, highlightthickness=0)
        self.v_scroll = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set)
        self.v_scroll.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.main = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.main, anchor="nw")
        self.main.bind("<Configure>", self._on_main_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_main_configure(self, _evt=None) -> None:
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, evt) -> None:
        self.canvas.itemconfigure(self.canvas_window, width=evt.width)

    def _on_mousewheel(self, evt) -> None:
        if self.canvas.winfo_exists():
            self.canvas.yview_scroll(int(-evt.delta / 120), "units")

    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 6}

        # 保留属性占位，兼容历史方法调用
        self.lb_filters = None
        self.filter_rows_ui: List[Dict[str, object]] = []
        self.txt_scan = None

        top = ttk.LabelFrame(self.main, text="1) 目标文件选择")
        top.pack(fill="x", padx=10, pady=(10, 6))
        ttk.Label(top, text="目标文件:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(top, textvariable=self.file_path_var, width=100).grid(row=0, column=1, columnspan=4, sticky="ew", **pad)
        ttk.Button(top, text="选择文件...", command=self._browse_file).grid(row=0, column=5, **pad)
        ttk.Button(top, text="加载文件", command=self._confirm_selected_file).grid(row=0, column=6, **pad)
        ttk.Checkbutton(top, text="快速加载(跳过全量扫描)", variable=self.fast_load_var).grid(
            row=1, column=1, columnspan=2, sticky="w", padx=8, pady=(0, 6)
        )
        ttk.Label(top, text="标题行(默认1):").grid(row=1, column=5, sticky="e", padx=8, pady=(0, 6))
        ttk.Spinbox(top, from_=1, to=50, textvariable=self.header_row_var, width=6).grid(
            row=1, column=6, sticky="w", padx=8, pady=(0, 6)
        )
        top.grid_columnconfigure(1, weight=1)

        info = ttk.Frame(self.main)
        info.pack(fill="x", padx=10, pady=(2, 6))
        ttk.Label(info, textvariable=self.source_info_var).pack(side="left")

        filt = ttk.LabelFrame(self.main, text="2) 条件筛选（字段 + 筛选信息）")
        filt.pack(fill="x", padx=10, pady=(6, 6))
        head = ttk.Frame(filt)
        head.pack(fill="x", padx=8, pady=(8, 4))
        head.grid_columnconfigure(0, weight=1)
        head.grid_columnconfigure(1, weight=1)
        ttk.Label(head, text="筛选字段").grid(row=0, column=0, sticky="w")
        ttk.Label(head, text="筛选信息").grid(row=0, column=1, sticky="w")
        self.filter_rows_host = ttk.Frame(filt)
        self.filter_rows_host.pack(fill="x", padx=8, pady=(0, 8))
        self.filter_rows_host.grid_columnconfigure(0, weight=1)
        self.filter_rows_ui = []
        self._add_filter_row_ui()

        bottom = ttk.LabelFrame(self.main, text="3) 一键导出（默认 by经理 + by项目）")
        bottom.pack(fill="x", padx=10, pady=(6, 10))
        export_btn = tk.Button(
            bottom,
            text="导出默认双sheet",
            command=self._export_default_dual,
            bg="#0A66C2",
            fg="white",
            activebackground="#084E97",
            activeforeground="white",
            font=("Microsoft YaHei UI", 12, "bold"),
            padx=28,
            pady=8,
            relief="raised",
            bd=2,
        )
        export_btn.pack(side="left", padx=8, pady=8)
        ttk.Label(bottom, text="将按默认字段配置导出到同一个Excel：by经理 + by项目").pack(side="left", padx=12, pady=8)

    def _open_pivot_window(self, mode: str) -> None:
        messagebox.showinfo("提示", "数据透视配置窗口已停用，请在首页直接导出默认双sheet。")

    def _bind_root_shortcuts(self) -> None:
        self.root.bind("<Control-a>", self._select_all_for_focus)
        self.root.bind("<Control-A>", self._select_all_for_focus)

    def _select_all_for_focus(self, _evt=None):
        w = self.root.focus_get()
        if isinstance(w, tk.Listbox):
            w.selection_set(0, tk.END)
            return "break"
        if isinstance(w, tk.Entry):
            w.selection_range(0, tk.END)
            w.icursor(tk.END)
            return "break"
        if isinstance(w, tk.Text):
            w.tag_add(tk.SEL, "1.0", tk.END)
            w.mark_set(tk.INSERT, tk.END)
            return "break"
        if isinstance(w, ttk.Treeview):
            for item in w.get_children():
                w.selection_add(item)
            return "break"
        return None

    def _browse_file(self) -> None:
        init_dir = self.folder_var.get().strip() or os.getcwd()
        path = filedialog.askopenfilename(
            title="选择目标文件",
            initialdir=init_dir,
            filetypes=[("支持文件", "*.xlsx *.xlsm *.xls *.csv *.txt"), ("All files", "*.*")],
        )
        if path:
            self.file_path_var.set(path)
            self.folder_var.set(str(Path(path).parent))
            self._confirm_selected_file()

    def _prompt_select_file_on_start(self) -> None:
        if self.file_path_var.get().strip():
            return
        self._browse_file()

    def _confirm_selected_file(self) -> None:
        if self.busy:
            return
        path = self.file_path_var.get().strip()
        if not path:
            messagebox.showwarning("提示", "请先选择目标文件。")
            return
        if not os.path.isfile(path):
            messagebox.showerror("错误", f"文件不存在：\n{path}")
            return

        header_row = self.header_row_var.get()
        if header_row < 1:
            messagebox.showwarning("提示", "标题行必须大于等于 1。")
            return

        self._run_background(
            title="正在读取文件并建立缓存...",
            worker=lambda: self._prepare_source(path, header_row, bool(self.fast_load_var.get())),
            on_success=self._on_source_loaded,
        )

    def _run_background(self, title: str, worker: Callable[[], object], on_success: Callable[[object], None]) -> None:
        if self.busy:
            return
        self.busy = True
        self._show_progress(title)

        def run():
            err = None
            result = None
            try:
                result = worker()
            except Exception as ex:
                err = ex
            finally:
                self.root.after(0, lambda: self._finish_background(err, result, on_success))

        threading.Thread(target=run, daemon=True).start()

    def _finish_background(self, err: Optional[Exception], result: object, on_success: Callable[[object], None]) -> None:
        self.busy = False
        self._close_progress()
        if err is not None:
            if isinstance(err, UserCancelledError):
                messagebox.showinfo("已取消", "操作已取消。")
                return
            traceback.print_exc()
            messagebox.showerror("错误", str(err))
            return
        on_success(result)

    def _show_progress(self, title: str, on_cancel: Optional[Callable[[], None]] = None) -> None:
        self.root.update_idletasks()
        self.progress_win = ProgressWindow(self.root, title=title, on_cancel=on_cancel)
        self.progress_win.lift()

    def _update_progress_stage(self, text: str) -> None:
        if self.progress_win is not None:
            try:
                self.progress_win.set_stage(text)
            except Exception:
                pass

    def _update_progress_value(self, pct: int, text: Optional[str] = None) -> None:
        if self.progress_win is not None:
            try:
                self.progress_win.set_progress(pct, text=text)
            except Exception:
                pass

    def _close_progress(self) -> None:
        if self.progress_win is not None:
            try:
                self.progress_win.close()
            except Exception:
                pass
            self.progress_win = None

    def _prepare_source(self, path: str, header_row: int, fast_load: bool) -> Tuple[DataSource, str]:
        p = Path(path)
        ext = p.suffix.lower()
        file_type = "csv"
        data_path = str(p)
        sheet_name = ""
        sep = ","
        encoding = "utf-8-sig"

        if ext in {".xlsx", ".xlsm", ".xls"}:
            file_type = "excel"
            sheet_names = self._list_sheet_names(str(p))
            if not sheet_names:
                raise RuntimeError("未找到任何 sheet。")
            if len(sheet_names) > 1:
                selected_sheet = self._ask_sheet_from_worker(sheet_names)
                if not selected_sheet:
                    raise UserCancelledError("未选择 sheet。")
                sheet_name = selected_sheet
            else:
                sheet_name = sheet_names[0]
            # Excel 直接入 DuckDB，不再中转 CSV
            data_path = str(p)
            sep = ","
            encoding = "utf-8-sig"
        else:
            file_type = "csv"
            sep, encoding = self._detect_sep_encoding(str(p))

        source = DataSource(
            original_path=str(p),
            data_path=data_path,
            file_type=file_type,
            sheet_name=sheet_name,
            sep=sep,
            encoding=encoding,
            header_row=header_row,
            headers=[],
        )
        headers = self._read_headers_from_source(source)
        if not headers:
            raise RuntimeError("未读取到表头，请检查标题行设置。")
        source.headers = headers

        # 切换来源后重置装载状态，后续按需懒加载
        self.source_loaded = False
        self.loaded_source_key = ""
        self.loaded_columns = set()
        self.loaded_full = False

        # 第一阶段固定仅做表头读取与预览，不做全量扫描/缓存构建
        summary = self._quick_profile_from_source(source, headers)
        return source, summary

    def _ensure_focus_cache(self, source: DataSource) -> None:
        focus_cols = self._normalize_required_cols(source, self._default_focus_fields())
        if not focus_cols:
            return
        focus_cache = self._source_cache_parquet(source, focus_cols)
        if focus_cache.exists():
            return
        try:
            df = self._read_source_polars(source, focus_cols)
            focus_cache.parent.mkdir(parents=True, exist_ok=True)
            df.write_parquet(str(focus_cache), compression="zstd")
            self._record_cache_manifest(source, focus_cache, "polars_focus_precut")
            return
            # csv/txt：直接按列投影生成精简缓存
            expr = self._duck_source_expr(source)
            select_cols = ", ".join(self._duck_quote_ident(c) for c in focus_cols)
            tmp_tbl = f"focus_{abs(hash(str(focus_cache)))}"
            with self.load_lock:
                self.db.execute(f"DROP TABLE IF EXISTS {tmp_tbl}")
                self.db.execute(f"CREATE TABLE {tmp_tbl} AS SELECT {select_cols} FROM {expr}")
                p = self._duck_quote_literal(str(focus_cache))
                self.db.execute(f"COPY {tmp_tbl} TO {p} (FORMAT PARQUET, COMPRESSION ZSTD)")
                self.db.execute(f"DROP TABLE IF EXISTS {tmp_tbl}")
            self._record_cache_manifest(source, focus_cache, "duckdb_focus_precut")
        except Exception:
            # 预裁剪失败不阻断主流程，后续仍按现有链路加载
            pass

    def _list_sheet_names(self, excel_path: str) -> List[str]:
        if cal_load_workbook is not None:
            try:
                wb = cal_load_workbook(excel_path)
                try:
                    return list(wb.sheet_names)
                finally:
                    wb.close()
            except Exception:
                pass
        wb = ox_load_workbook(excel_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def _ask_sheet_from_worker(self, sheet_names: Sequence[str]) -> Optional[str]:
        self.thread_event.clear()
        self.user_sheet_choice = None
        self.root.after(0, lambda: self._show_sheet_dialog(sheet_names))
        self.thread_event.wait()
        return self.user_sheet_choice

    def _show_sheet_dialog(self, sheet_names: Sequence[str]) -> None:
        top = tk.Toplevel(self.root)
        top.title("请选择 Sheet")
        top.geometry("440x160")
        top.resizable(False, False)
        top.transient(self.root)
        top.grab_set()
        top.lift()
        top.after(20, top.focus_force)

        ttk.Label(top, text="该文件包含多个 Sheet，请选择用于透视的数据表：").pack(anchor="w", padx=12, pady=(12, 8))
        cb = ttk.Combobox(top, state="readonly", values=list(sheet_names), width=52)
        cb.pack(anchor="w", padx=12)
        cb.set(sheet_names[0])

        btns = ttk.Frame(top)
        btns.pack(fill="x", pady=12, padx=12)

        def on_ok():
            self.user_sheet_choice = cb.get()
            top.destroy()
            self.thread_event.set()

        def on_cancel():
            self.user_sheet_choice = None
            top.destroy()
            self.thread_event.set()

        ttk.Button(btns, text="确认", command=on_ok).pack(side="right")
        ttk.Button(btns, text="取消", command=on_cancel).pack(side="right", padx=(0, 8))
        top.protocol("WM_DELETE_WINDOW", on_cancel)
        top.bind("<Escape>", lambda _e: on_cancel())

    def _ensure_shadow_csv(self, excel_path: str, sheet_name: str) -> str:
        st = os.stat(excel_path)
        key = f"{excel_path}|{st.st_mtime_ns}|{sheet_name}"
        cached = self.shadow_cache.get(key)
        if cached and os.path.exists(cached):
            return cached

        safe_sheet = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in sheet_name)[:40]
        base = Path(excel_path).stem
        ts = int(time.time())
        temp_name = f"shadow_{base}_{safe_sheet}_{ts}.csv"
        csv_path = str(Path(tempfile.gettempdir()) / temp_name)

        if cal_load_workbook is not None:
            try:
                wb = cal_load_workbook(excel_path)
                try:
                    ws = wb.get_sheet_by_name(sheet_name)
                    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f)
                        for row in ws.iter_rows():
                            writer.writerow([self._cell_to_text(v) for v in row])
                finally:
                    wb.close()
                self.shadow_cache[key] = csv_path
                return csv_path
            except Exception:
                pass

        wb = ox_load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([self._cell_to_text(v) for v in row])
        finally:
            wb.close()

        self.shadow_cache[key] = csv_path
        return csv_path

    @staticmethod
    def _cell_to_text(v: object) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(v, date):
            return v.strftime("%Y-%m-%d")
        return str(v)

    def _duck_quote_ident(self, name: str) -> str:
        return '"' + str(name).replace('"', '""') + '"'

    def _duck_quote_literal(self, v: str) -> str:
        return "'" + str(v).replace("'", "''") + "'"

    def _duck_normalize_encoding(self, enc: str) -> str:
        e = (enc or "").lower().replace("_", "-")
        if e in {"utf-8", "utf-8-sig", "utf8"}:
            return "utf-8"
        if e in {"utf-16", "utf-16-le", "utf16le"}:
            return "utf-16"
        if e in {"utf-16-be", "utf16be"}:
            return "utf-16"
        if e in {"gbk", "cp936", "gb18030"}:
            return "gbk"
        return "utf-8"

    def _duck_source_expr(self, source: DataSource) -> str:
        path = self._duck_quote_literal(source.data_path)
        delim = self._duck_quote_literal(source.sep)
        enc = self._duck_quote_literal(self._duck_normalize_encoding(source.encoding))
        skip = max(0, source.header_row - 1)
        return (
            f"read_csv({path}, delim={delim}, header=true, skip={skip}, all_varchar=true, "
            f"encoding={enc}, sample_size=-1)"
        )

    def _read_source_polars(
        self,
        source: DataSource,
        required_cols: Optional[Sequence[str]] = None,
    ):
        if pl is None:
            raise RuntimeError("polars 不可用。")
        cols = self._normalize_required_cols(source, required_cols)
        if source.file_type == "excel":
            cache_path = self._source_cache_parquet(source, None if not cols else cols)
            if not cache_path.exists():
                self._build_parquet_from_excel_polars(source, cache_path, None if not cols else cols)
                self._record_cache_manifest(source, cache_path, "polars_calamine")
            return pl.read_parquet(str(cache_path))

        read_kwargs = {
            "separator": source.sep,
            "has_header": True,
            "skip_rows": max(0, source.header_row - 1),
            "infer_schema_length": 0,
            "ignore_errors": True,
        }
        if source.encoding.lower().replace("_", "-") in {"utf-8", "utf-8-sig", "utf8"}:
            df = pl.read_csv(source.data_path, encoding="utf8-lossy", **read_kwargs)
        else:
            import pandas as pd
            df_pd = pd.read_csv(
                source.data_path,
                sep=source.sep,
                encoding=source.encoding,
                skiprows=max(0, source.header_row - 1),
                dtype=str,
                keep_default_na=False,
            )
            df = pl.from_pandas(df_pd)
        df = df.select([pl.col(c).cast(pl.Utf8, strict=False).alias(c) for c in df.columns])
        if cols:
            df = df.select([c for c in cols if c in df.columns])
        return df

    def _active_df(self):
        if self.source_df is None:
            raise RuntimeError("数据源尚未载入。")
        return self.source_df

    def _apply_filters_polars(self, df, filters: Dict[str, Set[str]]):
        for field, vals in filters.items():
            if field in df.columns and vals:
                df = df.filter(pl.col(field).cast(pl.Utf8, strict=False).fill_null("").is_in([str(v) for v in vals]))
        return df

    def _compute_pivot_polars(
        self,
        row_fields: Sequence[str],
        col_field: str,
        value_field: str,
        filters: Dict[str, Set[str]],
        agg: str,
    ) -> Tuple[List[str], List[List[object]]]:
        df = self._apply_filters_polars(self._active_df(), filters)
        if value_field not in df.columns:
            raise RuntimeError(f"值字段不存在：{value_field}")
        metric = (
            pl.col(value_field).cast(pl.Float64, strict=False).fill_null(0.0).alias("__metric")
            if agg == "sum"
            else (pl.col(value_field).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars() != "").cast(pl.Int64).alias("__metric")
        )
        df = df.with_columns(metric)
        if col_field:
            if col_field not in df.columns:
                col_field = ""
            else:
                col_values = sorted(
                    ["" if v is None else str(v) for v in df.select(pl.col(col_field).cast(pl.Utf8, strict=False).fill_null("").unique()).to_series().to_list()]
                )
                if len(col_values) > MAX_PIVOT_COLUMN_VALUES:
                    raise RuntimeError(
                        f"列字段 {col_field} 去重值过多（{len(col_values)}），请先筛选后重试，或清空列字段。当前上限为 {MAX_PIVOT_COLUMN_VALUES}。"
                    )
                grouped = (
                    df.with_columns(pl.col(col_field).cast(pl.Utf8, strict=False).fill_null("").alias("__col"))
                    .group_by(list(row_fields) + ["__col"])
                    .agg(pl.sum("__metric").alias("__metric"))
                )
                try:
                    wide = grouped.pivot(index=list(row_fields), on="__col", values="__metric", aggregate_function="sum")
                except TypeError:
                    wide = grouped.pivot(index=list(row_fields), columns="__col", values="__metric", aggregate_function="sum")
                if "" in wide.columns:
                    wide = wide.rename({"": "<空白>"})
                metric_cols = [c for c in wide.columns if c not in row_fields]
                wide = wide.select(list(row_fields) + sorted(metric_cols))
                wide = wide.fill_null(0)
                columns = list(wide.columns)
                rows = [list(r) for r in wide.iter_rows()]
                return columns, rows

        grouped = df.group_by(list(row_fields)).agg(pl.sum("__metric").alias(value_field))
        grouped = grouped.sort(list(row_fields))
        columns = list(grouped.columns)
        rows = [list(r) for r in grouped.iter_rows()]
        return columns, rows

    def _source_cache_key(self, source: DataSource) -> str:
        st = os.stat(source.original_path)
        return f"{source.original_path}|{st.st_mtime_ns}|{st.st_size}|{source.sheet_name}|{source.header_row}|{source.file_type}"

    def _normalize_required_cols(self, source: DataSource, required_cols: Optional[Sequence[str]]) -> List[str]:
        if not required_cols:
            return []
        req = {str(x) for x in required_cols if str(x).strip()}
        if not req:
            return []
        return [c for c in source.headers if c in req]

    def _source_cache_parquet(self, source: DataSource, required_cols: Optional[Sequence[str]] = None) -> Path:
        key = self._source_cache_key(source)
        norm = self._normalize_required_cols(source, required_cols)
        col_sig = "ALL" if not norm else "COLS:" + "|".join(norm)
        safe = str(abs(hash(f"{key}|{col_sig}")))
        base = Path(source.original_path).stem
        return self.cache_dir / f"{base}_{safe}.parquet"

    def _manifest_entry_key(self, source: DataSource) -> str:
        return self._source_cache_key(source)

    def _load_cache_manifest(self) -> Dict[str, Dict[str, object]]:
        if not self.cache_manifest_path.exists():
            return {}
        try:
            with open(self.cache_manifest_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
        return {}

    def _save_cache_manifest(self) -> None:
        try:
            with open(self.cache_manifest_path, "w", encoding="utf-8") as f:
                json.dump(self.cache_manifest, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _record_cache_manifest(self, source: DataSource, cache_path: Path, engine: str) -> None:
        try:
            st = os.stat(source.original_path)
            self.cache_manifest[self._manifest_entry_key(source)] = {
                "cache_path": str(cache_path),
                "engine": engine,
                "source_path": source.original_path,
                "sheet_name": source.sheet_name,
                "header_row": source.header_row,
                "source_mtime_ns": st.st_mtime_ns,
                "source_size": st.st_size,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            self._save_cache_manifest()
        except Exception:
            pass

    def _cache_engine_hint(self, source: DataSource) -> str:
        entry = self.cache_manifest.get(self._manifest_entry_key(source), {})
        if isinstance(entry, dict):
            return str(entry.get("engine", "unknown"))
        return "unknown"

    def _distinct_cache_path(self, source: DataSource, col: str) -> Path:
        key = self._source_cache_key(source)
        safe = str(abs(hash(f"{key}|distinct|{col}")))
        return self.cache_dir / f"distinct_{safe}.parquet"

    def _read_headers_from_source(self, source: DataSource) -> List[str]:
        if source.file_type == "excel":
            # 首行表头读取优先 openpyxl：对超大 xlsx 该路径明显更快、更稳定
            wb = ox_load_workbook(source.original_path, read_only=True, data_only=True)
            try:
                ws = wb[source.sheet_name] if source.sheet_name else wb.worksheets[0]
                target = max(1, int(source.header_row))
                row_iter = ws.iter_rows(min_row=target, max_row=target, values_only=True)
                header_row = next(row_iter, None)
                if header_row is None:
                    # 少数异常场景回退为常规遍历
                    row_iter = ws.iter_rows(values_only=True)
                    for _ in range(max(0, target - 1)):
                        next(row_iter, None)
                    header_row = next(row_iter, None)
                if header_row is None:
                    return []
                headers = self._normalize_headers(list(header_row))
                return headers
            finally:
                wb.close()

        with open(source.data_path, "r", encoding=source.encoding, newline="") as f:
            reader = csv.reader(f, delimiter=source.sep)
            for _ in range(max(0, source.header_row - 1)):
                next(reader, None)
            row = next(reader, None)
            if row is None:
                return []
            return self._normalize_headers(row)

    def _load_source_from_cache_or_origin(
        self,
        source: DataSource,
        required_cols: Optional[Sequence[str]] = None,
        ctx: Optional[TaskContext] = None,
    ) -> None:
        norm_cols = self._normalize_required_cols(source, required_cols)
        need_full = len(norm_cols) == 0
        t_all = time.perf_counter()
        with self.load_lock:
            if ctx is not None:
                ctx.progress(10, "阶段 1/5：检查数据缓存...")
            cache_path = self._source_cache_parquet(source, None if need_full else norm_cols)
            full_cache_path = self._source_cache_parquet(source, None)
            if cache_path.exists():
                self.source_df = pl.read_parquet(str(cache_path))
                self._record_cache_manifest(source, cache_path, self._cache_engine_hint(source) or "polars_cache")
            elif (not need_full) and full_cache_path.exists():
                df = pl.read_parquet(str(full_cache_path))
                keep = [c for c in norm_cols if c in df.columns]
                self.source_df = df.select(keep)
                cache_path.parent.mkdir(parents=True, exist_ok=True)
                self.source_df.write_parquet(str(cache_path), compression="zstd")
                self._record_cache_manifest(source, cache_path, "project_from_full_parquet")
            elif need_full and full_cache_path.exists():
                self.source_df = pl.read_parquet(str(full_cache_path))
                self._record_cache_manifest(source, full_cache_path, self._cache_engine_hint(source) or "polars_cache_full")
            else:
                self.source_df = self._read_source_polars(source, None if need_full else norm_cols)
                cache_path.parent.mkdir(parents=True, exist_ok=True)
                self.source_df.write_parquet(str(cache_path), compression="zstd")
                self._record_cache_manifest(source, cache_path, "polars_load")
            self.source_loaded = True
            self.loaded_source_key = self._source_cache_key(source)
            tbl_cols = list(self.source_df.columns) if self.source_df is not None else []
            self.loaded_columns = set(tbl_cols)
            self.loaded_full = set(source.headers).issubset(self.loaded_columns)
            if ctx is not None:
                ctx.progress(30, "阶段 1/5：数据源装载完成")
                ctx.add_timing("stage1_total_prepare_source", time.perf_counter() - t_all)
            return
        with self.load_lock:
            t_lock = time.perf_counter()
            if ctx is not None:
                ctx.progress(10, "阶段 1/5：检查数据缓存...")
            self.db.execute(f"DROP TABLE IF EXISTS {self.source_table_name}")
            cache_path = self._source_cache_parquet(source, None if need_full else norm_cols)
            full_cache_path = self._source_cache_parquet(source, None)
            if ctx is not None:
                ctx.add_timing("stage1_lock_wait_and_cache_check", time.perf_counter() - t_lock)
            # Excel 先统一落标准全量 parquet，再做子集投影
            if source.file_type == "excel" and self.polars_available and not full_cache_path.exists():
                try:
                    t = time.perf_counter()
                    if ctx is not None:
                        ctx.progress(14, "阶段 1/5：首次将 Excel 转为标准 Parquet...")
                    self._build_parquet_from_excel_polars(source, full_cache_path, required_cols=None)
                    self._record_cache_manifest(source, full_cache_path, "polars_calamine_full")
                    if ctx is not None:
                        ctx.add_timing("stage1_excel_to_full_parquet", time.perf_counter() - t)
                except Exception as ex:
                    if ctx is not None:
                        ctx.note(f"快路径失败(全量Parquet): {type(ex).__name__}: {ex}")
            if cache_path.exists():
                t = time.perf_counter()
                if ctx is not None:
                    ctx.progress(18, "阶段 1/5：命中列级缓存，正在装载...")
                p = self._duck_quote_literal(str(cache_path))
                self.db.execute(f"CREATE TABLE {self.source_table_name} AS SELECT * FROM read_parquet({p})")
                self._record_cache_manifest(source, cache_path, self._cache_engine_hint(source) or "parquet_cache")
                if ctx is not None:
                    ctx.add_timing("stage1_load_subset_cache", time.perf_counter() - t)
            elif (not need_full) and full_cache_path.exists():
                # 子集缓存不存在时，优先从全量parquet投影，避免再次回源读取Excel
                t = time.perf_counter()
                if ctx is not None:
                    ctx.progress(18, "阶段 1/5：命中全量缓存，正在按需裁剪列...")
                p_full = self._duck_quote_literal(str(full_cache_path))
                select_cols = ", ".join(self._duck_quote_ident(c) for c in norm_cols)
                self.db.execute(
                    f"CREATE TABLE {self.source_table_name} AS SELECT {select_cols} FROM read_parquet({p_full})"
                )
                try:
                    p_sub = self._duck_quote_literal(str(cache_path))
                    self.db.execute(f"COPY {self.source_table_name} TO {p_sub} (FORMAT PARQUET, COMPRESSION ZSTD)")
                    self._record_cache_manifest(source, cache_path, "project_from_full_parquet")
                except Exception:
                    pass
                if ctx is not None:
                    ctx.add_timing("stage1_project_from_full_parquet", time.perf_counter() - t)
            elif need_full and full_cache_path.exists():
                t = time.perf_counter()
                if ctx is not None:
                    ctx.progress(18, "阶段 1/5：命中全量缓存，正在装载...")
                p_full = self._duck_quote_literal(str(full_cache_path))
                self.db.execute(f"CREATE TABLE {self.source_table_name} AS SELECT * FROM read_parquet({p_full})")
                self._record_cache_manifest(source, full_cache_path, self._cache_engine_hint(source) or "parquet_cache_full")
                if ctx is not None:
                    ctx.add_timing("stage1_load_full_cache", time.perf_counter() - t)
            else:
                built_with_polars = False
                if source.file_type == "excel":
                    if not self.polars_available:
                        raise RuntimeError(
                            "Excel 快速引擎不可用：请安装 polars + fastexcel。\n"
                            f"当前解释器: {sys.executable}\n"
                            "建议执行:\n"
                            f"\"{sys.executable}\" -m pip install polars fastexcel"
                        )
                    try:
                        t = time.perf_counter()
                        if ctx is not None:
                            ctx.progress(16, "阶段 1/5：正在读取 Excel 并构建列缓存...")
                        self._build_parquet_from_excel_polars(source, cache_path, None if need_full else norm_cols)
                        built_with_polars = cache_path.exists()
                        if ctx is not None:
                            ctx.add_timing("stage1_excel_to_subset_parquet", time.perf_counter() - t)
                    except Exception as ex:
                        if ctx is not None:
                            ctx.note(f"快路径失败(列级Parquet): {type(ex).__name__}: {ex}")
                        built_with_polars = False

                if built_with_polars:
                    t = time.perf_counter()
                    if ctx is not None:
                        ctx.progress(22, "阶段 1/5：列缓存构建完成，正在装载...")
                    p = self._duck_quote_literal(str(cache_path))
                    self.db.execute(f"CREATE TABLE {self.source_table_name} AS SELECT * FROM read_parquet({p})")
                    self._record_cache_manifest(source, cache_path, "polars_calamine")
                    if ctx is not None:
                        ctx.add_timing("stage1_load_built_subset_cache", time.perf_counter() - t)
                else:
                    if source.file_type == "excel":
                        raise RuntimeError("Excel 快路径失败，已禁用回退路径。请检查快路径诊断信息。")
                    t = time.perf_counter()
                    self._init_duckdb_source(source)
                    try:
                        if need_full:
                            p = self._duck_quote_literal(str(cache_path))
                            self.db.execute(f"COPY {self.source_table_name} TO {p} (FORMAT PARQUET, COMPRESSION ZSTD)")
                            self._record_cache_manifest(source, cache_path, "duckdb_copy")
                    except Exception:
                        pass
                    if ctx is not None:
                        ctx.add_timing("stage1_fallback_load_source", time.perf_counter() - t)
            self.source_loaded = True
            self.loaded_source_key = self._source_cache_key(source)
            try:
                tbl_cols = [str(r[1]) for r in self.db.execute(f"PRAGMA table_info('{self.source_table_name}')").fetchall()]
            except Exception:
                tbl_cols = source.headers[:] if need_full else norm_cols[:]
            self.loaded_columns = set(tbl_cols)
            self.loaded_full = set(source.headers).issubset(self.loaded_columns)
            if ctx is not None:
                ctx.progress(30, "阶段 1/5：数据源装载完成")
                ctx.add_timing("stage1_total_prepare_source", time.perf_counter() - t_all)

    def _build_parquet_from_excel_polars(
        self,
        source: DataSource,
        cache_path: Path,
        required_cols: Optional[Sequence[str]] = None,
    ) -> None:
        if pl is None:
            raise RuntimeError("polars 不可用。")
        if not source.sheet_name:
            raise RuntimeError("未指定 sheet。")
        if not source.headers:
            raise RuntimeError("未识别到表头。")
        cols = self._normalize_required_cols(source, required_cols)
        if not cols:
            cols = source.headers[:]
        idx_map = {c: i for i, c in enumerate(source.headers)}
        sel_idx = [idx_map[c] for c in cols if c in idx_map]

        read_path = self._maybe_local_excel_copy(source.original_path)
        # 优先使用 fastexcel；打包环境缺少该扩展时回退到 python-calamine/openpyxl。
        try:
            df = pl.read_excel(
                read_path,
                sheet_name=source.sheet_name,
                engine="calamine",
                has_header=False,
                columns=sel_idx if sel_idx else None,
                infer_schema_length=0,
                drop_empty_rows=False,
                drop_empty_cols=False,
                raise_if_empty=True,
            )
            data_df = df.slice(source.header_row)
        except Exception:
            data_df = self._read_excel_rows_as_polars(source, cols)
        n = len(cols)
        if data_df.width < n:
            for i in range(data_df.width, n):
                data_df = data_df.with_columns(pl.lit(None).alias(f"__pad_{i}"))
        if data_df.width > n:
            data_df = data_df.select(data_df.columns[:n])

        rename_map = {old: cols[i] for i, old in enumerate(data_df.columns[:n])}
        data_df = data_df.rename(rename_map)
        data_df = data_df.select([pl.col(c).cast(pl.Utf8, strict=False).alias(c) for c in cols])

        cache_path.parent.mkdir(parents=True, exist_ok=True)
        data_df.write_parquet(str(cache_path), compression="zstd")

    def _read_excel_rows_as_polars(self, source: DataSource, cols: Sequence[str]):
        if pl is None:
            raise RuntimeError("polars 不可用。")
        idx_map = {c: i for i, c in enumerate(source.headers)}
        sel_idx = [idx_map[c] for c in cols if c in idx_map]
        rows: List[List[str]] = []

        def keep_values(values: Sequence[object]) -> List[str]:
            out = []
            for i in sel_idx:
                v = values[i] if i < len(values) else None
                out.append(self._cell_to_text(v))
            return out

        if cal_load_workbook is not None:
            try:
                wb = cal_load_workbook(source.original_path)
                try:
                    ws = wb.get_sheet_by_name(source.sheet_name) if source.sheet_name else wb.get_sheet_by_index(0)
                    for ridx, row in enumerate(ws.iter_rows()):
                        if ridx <= source.header_row - 1:
                            continue
                        rows.append(keep_values(list(row)))
                finally:
                    wb.close()
                return pl.DataFrame(rows, schema=list(cols), orient="row")
            except Exception:
                rows.clear()

        wb = ox_load_workbook(source.original_path, read_only=True, data_only=True)
        try:
            ws = wb[source.sheet_name] if source.sheet_name else wb.worksheets[0]
            for ridx, row in enumerate(ws.iter_rows(values_only=True)):
                if ridx <= source.header_row - 1:
                    continue
                rows.append(keep_values(list(row or [])))
        finally:
            wb.close()
        return pl.DataFrame(rows, schema=list(cols), orient="row")

    def _maybe_local_excel_copy(self, path: str) -> str:
        p = Path(path)
        if not str(p).startswith("\\\\"):
            return path
        st = os.stat(path)
        key = f"{path}|{st.st_mtime_ns}|{st.st_size}"
        hit = self.local_excel_copy_cache.get(key)
        if hit and os.path.exists(hit):
            return hit
        local_name = f"ts_local_{abs(hash(key))}{p.suffix}"
        local_path = str(self.cache_dir / local_name)
        shutil.copy2(path, local_path)
        self.local_excel_copy_cache[key] = local_path
        return local_path

    def _ensure_duckdb_source_loaded(
        self,
        required_cols: Optional[Sequence[str]] = None,
        ctx: Optional[TaskContext] = None,
    ) -> None:
        if not self.data_source:
            raise RuntimeError("请先加载数据源。")
        key = self._source_cache_key(self.data_source)
        norm_cols = self._normalize_required_cols(self.data_source, required_cols)
        need_full = len(norm_cols) == 0

        if self.source_loaded and self.loaded_source_key == key:
            if self.loaded_full:
                if ctx is not None:
                    ctx.progress(30, "阶段 1/5：已命中内存数据源")
                return
            if not need_full and set(norm_cols).issubset(self.loaded_columns):
                if ctx is not None:
                    ctx.progress(30, "阶段 1/5：已命中列级内存数据源")
                return
            if not need_full:
                merged = list(dict.fromkeys(list(self.loaded_columns) + norm_cols))
                self._load_source_from_cache_or_origin(self.data_source, required_cols=merged, ctx=ctx)
                return
            # 需要全量时，升级为全量载入
            self._load_source_from_cache_or_origin(self.data_source, required_cols=None, ctx=ctx)
            return
        self._load_source_from_cache_or_origin(
            self.data_source,
            required_cols=None if need_full else norm_cols,
            ctx=ctx,
        )

    def _start_prewarm(self) -> None:
        if not self.data_source:
            return
        if self.prewarm_running:
            return
        # 仅在已有缓存时才做后台预热，避免首次加载时后台任务长时间占用
        cache_path = self._source_cache_parquet(self.data_source, self._default_focus_fields())
        if not cache_path.exists():
            self.prewarm_done = False
            self.prewarm_running = False
            return
        source = self.data_source
        self.prewarm_running = True
        self.prewarm_done = False
        self.distinct_cache_tables = {}

        def run():
            try:
                self._load_source_from_cache_or_origin(source, required_cols=self._default_focus_fields())
                self._build_distinct_cache_tables(source)
                self.prewarm_done = True
            except Exception:
                pass
            finally:
                self.prewarm_running = False

        self.prewarm_thread = threading.Thread(target=run, daemon=True)
        self.prewarm_thread.start()

    def _build_distinct_cache_tables(self, source: DataSource) -> None:
        if not self.source_loaded:
            return
        if self.source_df is None:
            return
        watched = self._default_focus_fields()
        for col in watched:
            if col not in source.headers:
                continue
            if col not in self.source_df.columns:
                continue
            cache_path = self._distinct_cache_path(source, col)
            try:
                vals_df = (
                    self.source_df.select(pl.col(col).cast(pl.Utf8, strict=False).fill_null("").alias("v"))
                    .unique()
                    .sort("v")
                )
                vals_df.write_parquet(str(cache_path), compression="zstd")
                self.distinct_cache_tables[col] = str(cache_path)
            except Exception:
                pass
        return

    def _init_duckdb_source(self, source: DataSource) -> None:
        self.db.execute(f"DROP TABLE IF EXISTS {self.source_table_name}")
        if source.file_type == "excel":
            self._load_excel_sheet_to_duckdb(source)
        else:
            expr = self._duck_source_expr(source)
            self.db.execute(f"CREATE TABLE {self.source_table_name} AS SELECT * FROM {expr}")

    @staticmethod
    def _normalize_headers(raw_headers: Sequence[object]) -> List[str]:
        out: List[str] = []
        used: Dict[str, int] = {}
        for i, h in enumerate(raw_headers, start=1):
            name = str(h).strip() if h is not None else ""
            if not name:
                name = f"Column_{i}"
            base = name
            if base in used:
                used[base] += 1
                name = f"{base}_{used[base]}"
            else:
                used[base] = 1
            out.append(name)
        return out

    def _load_excel_sheet_to_duckdb(self, source: DataSource) -> None:
        if cal_load_workbook is not None:
            try:
                wb = cal_load_workbook(source.original_path)
                try:
                    ws = wb.get_sheet_by_name(source.sheet_name) if source.sheet_name else wb.get_sheet_by_index(0)
                    row_iter = ws.iter_rows()
                    skip_n = max(0, source.header_row - 1)
                    for _ in range(skip_n):
                        next(row_iter, None)
                    header_row = next(row_iter, None)
                    if header_row is None:
                        raise RuntimeError("未读取到表头行。")

                    headers = self._normalize_headers(list(header_row))
                    source.headers = headers
                    cols_ddl = ", ".join(f'{self._duck_quote_ident(c)} VARCHAR' for c in headers)
                    self.db.execute(f"CREATE TABLE {self.source_table_name} ({cols_ddl})")

                    n = len(headers)
                    placeholders = ",".join(["?"] * n)
                    ins_sql = f"INSERT INTO {self.source_table_name} VALUES ({placeholders})"
                    batch: List[Tuple[str, ...]] = []
                    batch_size = 5000
                    for row in row_iter:
                        vals = list(row) if row is not None else []
                        if len(vals) < n:
                            vals.extend([None] * (n - len(vals)))
                        elif len(vals) > n:
                            vals = vals[:n]
                        tup = tuple(self._cell_to_text(v) for v in vals)
                        batch.append(tup)
                        if len(batch) >= batch_size:
                            self.db.executemany(ins_sql, batch)
                            batch.clear()
                    if batch:
                        self.db.executemany(ins_sql, batch)
                    return
                finally:
                    wb.close()
            except Exception:
                pass

        wb = ox_load_workbook(source.original_path, read_only=True, data_only=True)
        try:
            ws = wb[source.sheet_name] if source.sheet_name else wb.worksheets[0]
            row_iter = ws.iter_rows(values_only=True)
            skip_n = max(0, source.header_row - 1)
            for _ in range(skip_n):
                next(row_iter, None)
            header_row = next(row_iter, None)
            if header_row is None:
                raise RuntimeError("未读取到表头行。")

            headers = self._normalize_headers(list(header_row))
            source.headers = headers
            cols_ddl = ", ".join(f'{self._duck_quote_ident(c)} VARCHAR' for c in headers)
            self.db.execute(f"CREATE TABLE {self.source_table_name} ({cols_ddl})")

            n = len(headers)
            placeholders = ",".join(["?"] * n)
            ins_sql = f"INSERT INTO {self.source_table_name} VALUES ({placeholders})"
            batch: List[Tuple[str, ...]] = []
            batch_size = 5000
            for row in row_iter:
                vals = list(row) if row is not None else []
                if len(vals) < n:
                    vals.extend([None] * (n - len(vals)))
                elif len(vals) > n:
                    vals = vals[:n]
                tup = tuple(self._cell_to_text(v) for v in vals)
                batch.append(tup)
                if len(batch) >= batch_size:
                    self.db.executemany(ins_sql, batch)
                    batch.clear()
            if batch:
                self.db.executemany(ins_sql, batch)
        finally:
            wb.close()

    def _read_headers(self, _data_path: str, _sep: str, _encoding: str, _header_row: int) -> List[str]:
        sql = f"SELECT * FROM {self.source_table_name} LIMIT 0"
        cur = self.db.execute(sql)
        return [d[0] for d in cur.description]

    def _run_deep_scan(self) -> None:
        if not self.data_source:
            messagebox.showwarning("提示", "请先加载文件。")
            return
        headers = self.data_source.headers[:]
        self._run_background(
            title="正在执行全量深度扫描...",
            worker=lambda: self._scan_profile("", "", "", 1, headers),
            on_success=lambda summary: self._set_scan_text(summary),  # type: ignore[arg-type]
        )

    def _scan_profile(self, _data_path: str, _sep: str, _encoding: str, _header_row: int, headers: Sequence[str]) -> str:
        self._ensure_duckdb_source_loaded(required_cols=self._default_focus_fields())
        df = self._active_df()
        lines: List[str] = []
        lines.append("扫描完成")
        lines.append(f"- 总行数(不含表头): {df.height:,}")
        lines.append(f"- 总列数: {len(headers)}")
        lines.append(f"- 表头: {', '.join(headers[:12])}{' ...' if len(headers) > 12 else ''}")
        for c in ["Employee GPN", "Employee Name", "Engagement Code"]:
            if c in df.columns:
                vals = df.select(pl.col(c).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars().alias(c))
                vals = vals.filter(pl.col(c) != "")
                lines.append(f"- {c} 去重: {vals.select(pl.col(c).n_unique()).item():,}")
        if "Hours" in df.columns:
            h_txt = pl.col("Hours").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars()
            h_num = pl.col("Hours").cast(pl.Float64, strict=False)
            h_non_null, h_invalid, h_sum = df.select(
                (h_txt != "").sum(),
                ((h_txt != "") & h_num.is_null()).sum(),
                h_num.fill_null(0).sum(),
            ).row(0)
            lines.append(f"- Hours 非空: {int(h_non_null or 0):,}")
            lines.append(f"- Hours 无法解析: {int(h_invalid or 0):,}")
            lines.append(f"- Hours 合计: {float(h_sum or 0):,.2f}")
        for c in self._default_focus_fields():
            if c in df.columns:
                lines.append(self._duck_top_line(c, 8))
        return "\n".join(lines)
        t = self.source_table_name
        lines: List[str] = []
        total_rows = int(self.db.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0])
        lines.append("扫描完成")
        lines.append(f"- 总行数(不含表头): {total_rows:,}")
        lines.append(f"- 总列数: {len(headers)}")
        lines.append(f"- 表头: {', '.join(headers[:12])}{' ...' if len(headers) > 12 else ''}")

        def distinct_non_empty(col: str) -> int:
            q = self._duck_quote_ident(col)
            return int(
                self.db.execute(
                    f"SELECT COUNT(DISTINCT NULLIF(TRIM(COALESCE({q},'')), '')) FROM {t}"
                ).fetchone()[0]
            )

        for c in ["Employee GPN", "Employee Name", "Engagement Code"]:
            if c in headers:
                lines.append(f"- {c} 去重: {distinct_non_empty(c):,}")

        if "Hours" in headers:
            qh = self._duck_quote_ident("Hours")
            h_non_null, h_invalid, h_sum = self.db.execute(
                f"""
                SELECT
                  SUM(CASE WHEN TRIM(COALESCE({qh},'')) <> '' THEN 1 ELSE 0 END),
                  SUM(CASE WHEN TRIM(COALESCE({qh},'')) <> '' AND TRY_CAST({qh} AS DOUBLE) IS NULL THEN 1 ELSE 0 END),
                  SUM(COALESCE(TRY_CAST({qh} AS DOUBLE), 0))
                FROM {t}
                """
            ).fetchone()
            lines.append(f"- Hours 非空: {int(h_non_null or 0):,}")
            lines.append(f"- Hours 无法解析: {int(h_invalid or 0):,}")
            lines.append(f"- Hours 合计: {float(h_sum or 0):,.2f}")

        date_exprs = [
            "TRY_CAST({q} AS TIMESTAMP)",
            "TRY_STRPTIME({q}, '%Y-%m-%d %H:%M:%S')",
            "TRY_STRPTIME({q}, '%Y-%m-%d')",
            "TRY_STRPTIME({q}, '%Y/%m/%d')",
            "TRY_STRPTIME({q}, '%Y%m%d')",
        ]
        for c in ["Transaction Cycle Date", "Proceed Time"]:
            if c in headers:
                q = self._duck_quote_ident(c)
                expr = "COALESCE(" + ", ".join(x.format(q=q) for x in date_exprs) + ")"
                mn, mx = self.db.execute(f"SELECT MIN({expr}), MAX({expr}) FROM {t}").fetchone()
                if mn and mx:
                    lines.append(f"- {c} 范围: {str(mn)[:10]} ~ {str(mx)[:10]}")

        for c in self._default_focus_fields():
            if c in headers:
                lines.append(self._duck_top_line(c, 8))
        return "\n".join(lines)

    def _default_focus_fields(self) -> List[str]:
        # 扫描/预热仅聚焦透视默认字段，避免遍历无关列
        base = [
            "COE Manager",
            "Employee Name",
            "Employee Rank Name",
            "Engagement Name",
            "Engagement Code",
            "Engagement Type",
            "Time Type Desc",
            "Employee GPN",
            "COE Senior",
            "Transaction Cycle Date",
            "Hours",
            DEFAULT_FILTER_FIELD,
        ]
        out: List[str] = []
        seen: Set[str] = set()
        for c in base:
            if c and c not in seen:
                seen.add(c)
                out.append(c)
        return out

    def _quick_profile_from_source(self, source: DataSource, headers: Sequence[str]) -> str:
        cache_path = self._source_cache_parquet(source)
        focus_cache = self._source_cache_parquet(source, self._default_focus_fields())
        lines: List[str] = []
        lines.append("快速加载完成")
        lines.append("- 已启用懒加载：当前仅解析表头，尚未全量读取数据。")
        lines.append("- 当你执行深度扫描、筛选值加载、或导出时，才会自动入库。")
        if source.file_type == "excel":
            lines.append(
                f"- Excel 加载引擎: {'polars+fastexcel(calamine)' if self.polars_available else 'python-calamine(主路径)/openpyxl(回退)'}"
            )
            lines.append("- 数据链路: Excel -> 标准Parquet -> DuckDB 计算")
        if cache_path.exists():
            lines.append(f"- 缓存引擎: {self._cache_engine_hint(source)}")
        lines.append(f"- 精简缓存状态: {'已命中' if focus_cache.exists() else '未命中'}")
        lines.append(f"- 总列数: {len(headers)}")
        lines.append(f"- 表头: {', '.join(headers[:12])}{' ...' if len(headers) > 12 else ''}")
        lines.append(f"- 本地缓存状态: {'已命中' if cache_path.exists() else '未命中（首次会慢）'}")
        lines.append(f"- 缓存路径: {cache_path}")
        return "\n".join(lines)

    def _duck_top_line(self, col: str, n: int = 8) -> str:
        self._ensure_duckdb_source_loaded(required_cols=[col])
        df = self._active_df()
        if col not in df.columns:
            return f"- {col} Top: "
        rows = (
            df.select(pl.col(col).cast(pl.Utf8, strict=False).fill_null("").alias("v"))
            .with_columns(pl.when(pl.col("v") == "").then(pl.lit("NULL")).otherwise(pl.col("v")).alias("v"))
            .group_by("v")
            .len(name="c")
            .sort("c", descending=True)
            .head(n)
            .iter_rows()
        )
        parts = [f"{('<空白>' if str(v)=='' else str(v))}: {int(c):,}" for v, c in rows]
        return f"- {col} Top: " + " | ".join(parts)
        t = self.source_table_name
        q = self._duck_quote_ident(col)
        rows = self.db.execute(
            f"""
            SELECT COALESCE(NULLIF({q}, ''), 'NULL') AS v, COUNT(*) AS c
            FROM {t}
            GROUP BY 1
            ORDER BY c DESC
            LIMIT {n}
            """
        ).fetchall()
        parts = [f"{('<空白>' if str(v)=='' else str(v))}: {int(c):,}" for v, c in rows]
        return f"- {col} Top: " + " | ".join(parts)

    def _on_source_loaded(self, payload: object) -> None:
        source, summary = payload  # type: ignore[misc]
        self.data_source = source
        self.available_fields = source.headers[:]
        self.filter_values_cache.clear()
        self.pivot_result_cache.clear()
        self.filters.clear()
        self._refresh_filter_listbox()

        info = Path(source.original_path).name
        if source.sheet_name:
            info += f" | Sheet: {source.sheet_name}"
        info += f" | 标题行: {source.header_row}"
        self.source_info_var.set(info)
        self._set_scan_text(summary)
        self._start_prewarm()
        self._start_background_preload()
        # 首页筛选控件刷新（重置为默认一行）
        host = getattr(self, "filter_rows_host", None)
        if host is not None and hasattr(host, "winfo_exists") and host.winfo_exists():
            for child in host.winfo_children():
                child.destroy()
            self.filter_rows_ui = []
            self._add_filter_row_ui()

    def _start_background_preload(self) -> None:
        if not self.data_source:
            return
        if self.preload_running:
            return
        self.preload_running = True
        self.preload_done = False
        src = self.data_source
        self.source_info_var.set(f"{self.source_info_var.get()} | 后台预读中...")

        def run():
            try:
                required = []
                for c in self._build_default_row_fields("by_manager"):
                    if c and c not in required:
                        required.append(c)
                for c in self._build_default_row_fields("by_project"):
                    if c and c not in required:
                        required.append(c)
                col = self._default_column_field()
                if col and col not in required:
                    required.append(col)
                if "Hours" in src.headers and "Hours" not in required:
                    required.append("Hours")
                if DEFAULT_FILTER_FIELD in src.headers and DEFAULT_FILTER_FIELD not in required:
                    required.append(DEFAULT_FILTER_FIELD)
                self._ensure_duckdb_source_loaded(required_cols=required)
                # 预热默认筛选字段 distinct（若用户常用筛选）
                if DEFAULT_FILTER_FIELD in src.headers:
                    try:
                        self._get_unique_values(DEFAULT_FILTER_FIELD)
                    except Exception:
                        pass
                self.preload_done = True
                self.root.after(0, lambda: self.source_info_var.set(f"{self.source_info_var.get().replace(' | 后台预读中...', '')} | 后台预读完成"))
            except Exception:
                self.root.after(0, lambda: self.source_info_var.set(self.source_info_var.get().replace(" | 后台预读中...", "")))
            finally:
                self.preload_running = False

        self.preload_thread = threading.Thread(target=run, daemon=True)
        self.preload_thread.start()

    def _set_scan_text(self, content: str) -> None:
        if self.txt_scan is None:
            return
        self.txt_scan.configure(state="normal")
        self.txt_scan.delete("1.0", tk.END)
        self.txt_scan.insert("1.0", content)
        self.txt_scan.configure(state="disabled")

    def _build_default_row_fields(self, mode: str) -> List[str]:
        manager_base = [
            "COE Manager",
            "Employee Name",
            "Employee Rank Name",
            "Engagement Name",
            "Engagement Code",
            "Engagement Type",
            "Time Type Desc",
            "Employee GPN",
            "COE Senior",
        ]
        rows = [c for c in manager_base if c in self.available_fields]
        if not rows:
            rows = self.available_fields[: min(4, len(self.available_fields))]

        if mode == "by_manager":
            # 固定第2顺位为 Employee Name（存在时）
            if "Employee Name" in rows:
                rows.remove("Employee Name")
                insert_at = 1 if len(rows) >= 1 else 0
                rows.insert(insert_at, "Employee Name")

        if mode == "by_project":
            # 保持 by项目第一位 Engagement Name、第二位 Employee Name
            if "Engagement Name" in rows:
                rows.remove("Engagement Name")
                rows.insert(0, "Engagement Name")
            # 固定第2顺位为 Employee Name（存在时）
            if "Employee Name" in rows:
                rows.remove("Employee Name")
                insert_at = 1 if len(rows) >= 1 else 0
                rows.insert(insert_at, "Employee Name")
        return rows

    def _add_filter_row_ui(self) -> None:
        host = getattr(self, "filter_rows_host", None)
        if host is None or not host.winfo_exists():
            return
        row_idx = len(self.filter_rows_ui)
        row = ttk.Frame(host)
        row.grid(row=row_idx, column=0, sticky="ew", pady=2)
        row.grid_columnconfigure(0, weight=3)
        row.grid_columnconfigure(1, weight=3)

        field_var = tk.StringVar(value="")
        value_var = tk.StringVar(value="")
        field_values = [NO_FILTER_OPTION] + self.available_fields[:]
        cb_field = ttk.Combobox(row, state="readonly", textvariable=field_var, values=field_values)
        cb_field.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        cb_value = ttk.Combobox(row, state="readonly", textvariable=value_var, values=[NO_FILTER_OPTION])
        cb_value.grid(row=0, column=1, sticky="ew", padx=(0, 0))

        row_data: Dict[str, object] = {
            "frame": row,
            "field_var": field_var,
            "value_var": value_var,
            "cb_field": cb_field,
            "cb_value": cb_value,
            "load_token": 0,
        }
        self.filter_rows_ui.append(row_data)

        def on_field_selected(_evt=None):
            field = field_var.get().strip()
            if not field or field == NO_FILTER_OPTION:
                cb_value.configure(values=[NO_FILTER_OPTION])
                value_var.set(NO_FILTER_OPTION)
                return
            self._load_filter_values_async(row_data, field)

        cb_field.bind("<<ComboboxSelected>>", on_field_selected)

        # 首行默认预填（仍允许用户改为“无筛选”）
        if row_idx == 0:
            if DEFAULT_FILTER_FIELD in self.available_fields:
                field_var.set(DEFAULT_FILTER_FIELD)
                if DEFAULT_FILTER_FIELD in MANUAL_FILTER_CACHE:
                    vals = [NO_FILTER_OPTION] + [self._display_value(v) for v in MANUAL_FILTER_CACHE[DEFAULT_FILTER_FIELD]]
                    cb_value.configure(values=vals)
                    if DEFAULT_FILTER_VALUE in vals:
                        value_var.set(DEFAULT_FILTER_VALUE)
                    else:
                        value_var.set(vals[0] if vals else NO_FILTER_OPTION)
                else:
                    cb_value.configure(values=[NO_FILTER_OPTION, DEFAULT_FILTER_VALUE])
                    value_var.set(DEFAULT_FILTER_VALUE)
            else:
                field_var.set(NO_FILTER_OPTION)
                cb_value.configure(values=[NO_FILTER_OPTION])
                value_var.set(NO_FILTER_OPTION)

    def _remove_filter_row_ui(self, row_data: Dict[str, object]) -> None:
        if row_data in self.filter_rows_ui:
            self.filter_rows_ui.remove(row_data)
        frame = row_data.get("frame")
        if isinstance(frame, ttk.Frame) and frame.winfo_exists():
            frame.destroy()
        self._repack_filter_rows_ui()
        if not self.filter_rows_ui:
            self._add_filter_row_ui()

    def _repack_filter_rows_ui(self) -> None:
        for idx, item in enumerate(self.filter_rows_ui):
            frame = item.get("frame")
            if isinstance(frame, ttk.Frame) and frame.winfo_exists():
                frame.grid_configure(row=idx)

    def _collect_filters_from_rows(self) -> Dict[str, Set[str]]:
        filters: Dict[str, Set[str]] = {}
        for item in self.filter_rows_ui:
            field_var = item.get("field_var")
            value_var = item.get("value_var")
            if not isinstance(field_var, tk.StringVar) or not isinstance(value_var, tk.StringVar):
                continue
            field = field_var.get().strip()
            value_disp = value_var.get().strip()
            if not field or not value_disp:
                continue
            if field == NO_FILTER_OPTION or value_disp == NO_FILTER_OPTION:
                continue
            if value_disp in {"正在加载...", "（无可用筛选值）"}:
                continue
            value = "" if value_disp == "<空白>" else value_disp
            filters.setdefault(field, set()).add(value)
        return filters

    def _load_filter_values_async(
        self,
        row_data: Dict[str, object],
        field: str,
        preferred_value: Optional[str] = None,
    ) -> None:
        cb_value = row_data.get("cb_value")
        value_var = row_data.get("value_var")
        field_var = row_data.get("field_var")
        frame = row_data.get("frame")
        if not isinstance(cb_value, ttk.Combobox) or not isinstance(value_var, tk.StringVar):
            return
        if not isinstance(field_var, tk.StringVar):
            return
        if not isinstance(frame, ttk.Frame):
            return

        token = time.time_ns()
        row_data["load_token"] = token
        cb_value.configure(values=["正在加载..."])
        value_var.set("正在加载...")

        def worker():
            try:
                # 1) 先读持久化 distinct 缓存（最快）
                vals = self._read_distinct_cache_values(field)
                status = "ok"
                if not vals:
                    # 2) 若正在后台预热，避免在筛选下拉无限等待
                    if self.prewarm_running and not self.source_loaded:
                        status = "warming"
                    else:
                        # 3) 没有预热阻塞时再走常规查询
                        vals = self._get_unique_values(field)
                err = None
            except Exception as ex:
                vals = []
                status = "err"
                err = ex

            def apply_result():
                if not frame.winfo_exists():
                    return
                if row_data.get("load_token") != token:
                    return
                if field_var.get().strip() != field:
                    return
                if err is not None:
                    cb_value.configure(values=[NO_FILTER_OPTION])
                    value_var.set(NO_FILTER_OPTION)
                    return
                if status == "warming":
                    cb_value.configure(values=[NO_FILTER_OPTION, "后台预热中，请稍后重试"])
                    value_var.set("后台预热中，请稍后重试")
                    return
                disp = [NO_FILTER_OPTION] + [self._display_value(v) for v in vals]
                if len(disp) == 1:
                    disp.append("（无可用筛选值）")
                cb_value.configure(values=disp)
                if preferred_value and preferred_value in disp:
                    value_var.set(preferred_value)
                else:
                    value_var.set(NO_FILTER_OPTION)

            self.root.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    def _add_filter(self) -> None:
        if not self.data_source:
            messagebox.showwarning("提示", "请先加载文件。")
            return
        field = self._choose_field_dialog("新增筛选字段")
        if not field:
            return
        self._load_and_edit_filter(field)

    def _edit_filter(self) -> None:
        sel = self.lb_filters.curselection()
        if not sel:
            messagebox.showwarning("提示", "请先选中一条筛选条件。")
            return
        field = list(self.filters.keys())[sel[0]]
        self._load_and_edit_filter(field)

    def _remove_filter(self) -> None:
        sel = self.lb_filters.curselection()
        if not sel:
            return
        field = list(self.filters.keys())[sel[0]]
        self.filters.pop(field, None)
        self._refresh_filter_listbox()

    def _clear_filters(self) -> None:
        self.filters.clear()
        self._refresh_filter_listbox()

    def _choose_field_dialog(self, title: str) -> Optional[str]:
        if not self.available_fields:
            return None

        result: Dict[str, Optional[str]] = {"field": None}
        top = tk.Toplevel(self.root)
        top.title(title)
        top.geometry("420x150")
        top.resizable(False, False)
        top.transient(self.root)
        top.grab_set()
        top.lift()
        top.after(20, top.focus_force)

        ttk.Label(top, text="请选择筛选字段：").pack(anchor="w", padx=12, pady=(12, 8))
        cb = ttk.Combobox(top, values=self.available_fields, state="readonly", width=44)
        cb.pack(anchor="w", padx=12)
        cb.set(self.available_fields[0])

        def on_ok():
            result["field"] = cb.get()
            top.destroy()

        def on_cancel():
            top.destroy()

        f = ttk.Frame(top)
        f.pack(fill="x", padx=12, pady=12)
        ttk.Button(f, text="确认", command=on_ok).pack(side="right")
        ttk.Button(f, text="取消", command=on_cancel).pack(side="right", padx=(0, 8))
        top.bind("<Escape>", lambda _e: on_cancel())
        top.protocol("WM_DELETE_WINDOW", on_cancel)

        self.root.wait_window(top)
        return result["field"]

    def _load_and_edit_filter(self, field: str) -> None:
        self._run_background(
            title=f"正在读取筛选值：{field}",
            worker=lambda: self._get_unique_values(field),
            on_success=lambda vals: self._show_filter_values_dialog(field, vals),  # type: ignore[arg-type]
        )

    def _get_unique_values(self, field: str) -> List[str]:
        if field in MANUAL_FILTER_CACHE:
            vals = MANUAL_FILTER_CACHE[field]
            self.filter_values_cache[field] = vals[:]
            return vals[:]
        if field in self.filter_values_cache:
            return self.filter_values_cache[field]
        if not self.data_source:
            raise RuntimeError("数据源未加载。")
        dpath = self._distinct_cache_path(self.data_source, field)
        if dpath.exists():
            try:
                values = pl.read_parquet(str(dpath)).select(pl.col("v").cast(pl.Utf8, strict=False).fill_null("")).head(20000).to_series().to_list()
                values = sorted([str(v) for v in values], key=lambda x: (x == "", x.lower()))
                self.filter_values_cache[field] = values
                return values
            except Exception:
                pass
        self._ensure_duckdb_source_loaded(required_cols=[field])
        df = self._active_df()
        if field not in df.columns:
            return []
        values_df = (
            df.select(pl.col(field).cast(pl.Utf8, strict=False).fill_null("").alias("v"))
            .unique()
            .sort("v")
        )
        try:
            values_df.write_parquet(str(dpath), compression="zstd")
        except Exception:
            pass
        values = sorted([str(v) for v in values_df.head(20000).to_series().to_list()], key=lambda x: (x == "", x.lower()))
        self.filter_values_cache[field] = values
        return values
        if not self.data_source:
            raise RuntimeError("数据源未加载。")

        # 优先命中持久化 distinct 缓存（无需全表加载）
        dpath = self._distinct_cache_path(self.data_source, field)
        if dpath.exists():
            con = duckdb.connect(database=":memory:")
            try:
                p = self._duck_quote_literal(str(dpath))
                rows = con.execute(f"SELECT v FROM read_parquet({p}) LIMIT 20000").fetchall()
                values = sorted([str(r[0]) for r in rows], key=lambda x: (x == "", x.lower()))
                self.filter_values_cache[field] = values
                return values
            finally:
                con.close()

        self._ensure_duckdb_source_loaded(required_cols=[field])
        with self.load_lock:
            if field in self.distinct_cache_tables:
                tbl = self.distinct_cache_tables[field]
                rows = self.db.execute(f"SELECT v FROM {tbl} LIMIT 20000").fetchall()
            else:
                q = self._duck_quote_ident(field)
                rows = self.db.execute(
                    f"""
                    SELECT DISTINCT COALESCE({q}, '') AS v
                    FROM {self.source_table_name}
                    LIMIT 20000
                    """
                ).fetchall()
        values = sorted([str(r[0]) for r in rows], key=lambda x: (x == "", x.lower()))
        self.filter_values_cache[field] = values
        return values

    def _read_distinct_cache_values(self, field: str) -> List[str]:
        if not self.data_source:
            return []
        dpath = self._distinct_cache_path(self.data_source, field)
        if not dpath.exists():
            return []
        try:
            rows = pl.read_parquet(str(dpath)).select(pl.col("v").cast(pl.Utf8, strict=False).fill_null("")).head(20000).to_series().to_list()
            return sorted([str(r) for r in rows], key=lambda x: (x == "", x.lower()))
        except Exception:
            return []
        con = duckdb.connect(database=":memory:")
        try:
            p = self._duck_quote_literal(str(dpath))
            rows = con.execute(f"SELECT v FROM read_parquet({p}) LIMIT 20000").fetchall()
            return sorted([str(r[0]) for r in rows], key=lambda x: (x == "", x.lower()))
        finally:
            con.close()

    def _show_filter_values_dialog(self, field: str, values: List[str]) -> None:
        selected = set(self.filters.get(field, set()))
        top = tk.Toplevel(self.root)
        top.title(f"筛选值选择 - {field}")
        top.geometry("520x560")
        top.transient(self.root)
        top.grab_set()
        top.lift()
        top.after(20, top.focus_force)

        search_var = tk.StringVar(value="")
        ttk.Label(top, text=f"字段：{field}（可多选）").pack(anchor="w", padx=10, pady=(10, 6))
        ttk.Entry(top, textvariable=search_var).pack(fill="x", padx=10)

        box_frame = ttk.Frame(top)
        box_frame.pack(fill="both", expand=True, padx=10, pady=8)
        lb = tk.Listbox(box_frame, selectmode=tk.EXTENDED, exportselection=False)
        lb.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(box_frame, orient="vertical", command=lb.yview)
        sb.pack(side="left", fill="y")
        lb.configure(yscrollcommand=sb.set)

        current_actuals: List[str] = []
        selected_live = set(selected)

        def sync_selected_from_ui():
            for i in lb.curselection():
                if 0 <= i < len(current_actuals):
                    selected_live.add(current_actuals[i])

        def render():
            sync_selected_from_ui()
            kw = search_var.get().strip().lower()
            lb.delete(0, tk.END)
            current_actuals.clear()
            for v in values:
                dv = self._display_value(v)
                if kw and kw not in dv.lower():
                    continue
                current_actuals.append(v)
                lb.insert(tk.END, dv)
                if v in selected_live:
                    lb.selection_set(tk.END)

        def on_search(*_):
            render()

        search_var.trace_add("write", on_search)
        render()

        btns = ttk.Frame(top)
        btns.pack(fill="x", padx=10, pady=(0, 10))

        def pick_all():
            lb.selection_set(0, tk.END)

        def clear_all():
            lb.selection_clear(0, tk.END)

        def on_ok():
            sync_selected_from_ui()
            new_sel = set(selected_live)
            if new_sel:
                self.filters[field] = new_sel
            elif field in self.filters:
                self.filters.pop(field, None)
            self._refresh_filter_listbox()
            top.destroy()

        def on_cancel():
            top.destroy()

        ttk.Button(btns, text="全选", command=pick_all).pack(side="left")
        ttk.Button(btns, text="清空选择", command=clear_all).pack(side="left", padx=(6, 0))
        ttk.Button(btns, text="确认", command=on_ok).pack(side="right")
        ttk.Button(btns, text="取消", command=on_cancel).pack(side="right", padx=(0, 8))
        top.bind("<Escape>", lambda _e: on_cancel())
        top.bind("<Control-a>", lambda _e: (pick_all(), "break"))
        top.bind("<Control-A>", lambda _e: (pick_all(), "break"))
        top.protocol("WM_DELETE_WINDOW", on_cancel)

    @staticmethod
    def _display_value(v: str) -> str:
        return "<空白>" if v == "" else v

    def _refresh_filter_listbox(self) -> None:
        if self.lb_filters is None:
            return
        self.lb_filters.delete(0, tk.END)
        for f, vals in self.filters.items():
            arr = list(vals)
            head = ", ".join(self._display_value(x) for x in arr[:3])
            more = f" (+{len(arr) - 3})" if len(arr) > 3 else ""
            self.lb_filters.insert(tk.END, f"{f}: {head}{more}")

    def _export_pivot(self) -> None:
        self._export_default_dual()

    def _export_default_dual(self) -> None:
        if not self.data_source:
            messagebox.showwarning("提示", "请先加载文件。")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = filedialog.asksaveasfilename(
            title="导出默认双sheet数据透视",
            defaultextension=".xlsx",
            initialfile=f"Timesheet_Default_Dual_{ts}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not save_path:
            return
        filters = self._collect_filters_from_rows()
        self._run_dual_export_background(save_path, filters)

    def _default_column_field(self) -> str:
        if "Transaction Cycle Date" in self.available_fields:
            return "Transaction Cycle Date"
        if "Month" in self.available_fields:
            return "Month"
        return ""

    def _run_dual_export_background(self, save_path: str, filters: Dict[str, Set[str]]) -> None:
        if self.busy:
            return
        self.busy = True
        cancel_event = threading.Event()
        self.current_cancel_event = cancel_event
        self._show_progress("正在生成并导出双sheet数据透视...", on_cancel=cancel_event.set)
        self._update_progress_value(2, "阶段 1/4：准备数据源...")

        def report_stage(msg: str) -> None:
            self.root.after(0, lambda: self._update_progress_stage(msg))

        def report_progress(pct: int, text: Optional[str]) -> None:
            self.root.after(0, lambda: self._update_progress_value(pct, text))

        ctx = TaskContext(
            start_ts=time.monotonic(),
            timeout_sec=EXPORT_TIMEOUT_SEC,
            cancel_event=cancel_event,
            report_stage=report_stage,
            report_progress=report_progress,
        )

        def run() -> None:
            err = None
            result = None
            try:
                result = self._build_and_export_dual_default(save_path, filters, ctx)
            except Exception as ex:
                err = ex
            finally:
                self.root.after(0, lambda: self._finish_background(err, result, self._on_export_done))

        threading.Thread(target=run, daemon=True).start()

    def _run_export_background(self, cfg: Dict) -> None:
        if self.busy:
            return
        self.busy = True
        cancel_event = threading.Event()
        self.current_cancel_event = cancel_event
        self._show_progress(
            "正在生成并导出数据透视...",
            on_cancel=cancel_event.set,
        )
        self._update_progress_value(2, "阶段 1/5：准备数据源...")

        def report_stage(msg: str) -> None:
            self.root.after(0, lambda: self._update_progress_stage(msg))
        def report_progress(pct: int, text: Optional[str]) -> None:
            self.root.after(0, lambda: self._update_progress_value(pct, text))
        ctx = TaskContext(
            start_ts=time.monotonic(),
            timeout_sec=EXPORT_TIMEOUT_SEC,
            cancel_event=cancel_event,
            report_stage=report_stage,
            report_progress=report_progress,
        )

        def run() -> None:
            err = None
            result = None
            try:
                result = self._build_and_export_pivot(
                    cfg,
                    ctx=ctx,
                )
            except Exception as ex:
                err = ex
            finally:
                self.root.after(0, lambda: self._finish_background(err, result, self._on_export_done))

        threading.Thread(target=run, daemon=True).start()

    def _build_and_export_pivot(
        self,
        cfg: Dict,
        ctx: Optional[TaskContext] = None,
    ) -> Dict[str, object]:
        if ctx is None:
            ctx = TaskContext(start_ts=time.monotonic(), timeout_sec=EXPORT_TIMEOUT_SEC)

        ctx.stage("阶段 1/5：准备数据源...")
        ctx.progress(5, "阶段 1/5：准备数据源...")
        ctx.check_abort()
        if not self.data_source:
            raise RuntimeError("数据源不存在。")
        src = self.data_source
        row_fields = [f for f in cfg["row_fields"] if f in src.headers]
        if not row_fields:
            raise RuntimeError("行字段为空或无效。")

        col_field = cfg["column_field"] if cfg["column_field"] in src.headers else ""
        value_field = cfg["value_field"]
        if value_field not in src.headers:
            raise RuntimeError(f"值字段不存在：{value_field}")

        filters = {k: v for k, v in cfg["filters"].items() if k in src.headers and len(v) > 0}
        agg = cfg["agg"]
        if agg not in {"sum", "count"}:
            agg = "sum"

        required_cols: List[str] = []
        for c in row_fields:
            if c and c not in required_cols:
                required_cols.append(c)
        if col_field and col_field not in required_cols:
            required_cols.append(col_field)
        if value_field and value_field not in required_cols:
            required_cols.append(value_field)
        for f in filters.keys():
            if f and f not in required_cols:
                required_cols.append(f)
        ctx.progress(18, "阶段 1/5：加载所需列...")
        self._ensure_duckdb_source_loaded(required_cols=required_cols, ctx=ctx)
        ctx.progress(30, "阶段 1/5：数据源就绪")
        ctx.check_abort()

        save_ext = Path(str(cfg["save_path"])).suffix.lower()
        csv_stream_mode = save_ext == ".csv"

        cache_key = self._pivot_cache_key(src, cfg, filters, row_fields, col_field, value_field, agg)
        cached = None if csv_stream_mode else self.pivot_result_cache.get(cache_key)
        if cached is None:
            ctx.stage("阶段 2/5：使用 Polars 计算透视结果...")
            ctx.progress(45, "阶段 2/5：使用 Polars 计算透视结果...")
            columns, rows = self._compute_pivot_result(src, cfg)
            ctx.check_abort()
            ctx.stage("阶段 5/5：写出结果文件...")
            ctx.progress(90, "阶段 5/5：写出结果文件...")
            self._write_output(columns, rows, cfg, src)
            ctx.progress(100, "100% 完成")
            elapsed = time.monotonic() - ctx.start_ts
            return {
                "save_path": cfg["save_path"],
                "rows": len(rows),
                "cols": len(columns),
                "elapsed_sec": elapsed,
                "timings": dict(ctx.timings),
                "notes": list(ctx.notes),
            }
        if cached is not None:
            columns, rows = cached
            ctx.check_abort()
            ctx.stage("阶段 5/5：写出结果文件（缓存命中）...")
            ctx.progress(85, "阶段 5/5：写出结果文件（缓存命中）...")
            self._write_output(columns, [r[:] for r in rows], cfg, src)
            ctx.progress(100, "100% 完成")
            elapsed = time.monotonic() - ctx.start_ts
            return {
                "save_path": cfg["save_path"],
                "rows": len(rows),
                "cols": len(columns),
                "elapsed_sec": elapsed,
                "timings": dict(ctx.timings),
                "notes": list(ctx.notes),
            }
        else:
            ctx.stage("阶段 2/5：应用筛选并准备聚合...")
            ctx.progress(40, "阶段 2/5：应用筛选并准备聚合...")
            where_clause = self._build_where_clause(filters)
            metric_expr = (
                f"COALESCE(TRY_CAST({self._duck_quote_ident(value_field)} AS DOUBLE),0)"
                if agg == "sum"
                else f"CASE WHEN TRIM(COALESCE({self._duck_quote_ident(value_field)},''))<>'' THEN 1 ELSE 0 END"
            )
            t = self.source_table_name
            rq = [self._duck_quote_ident(c) for c in row_fields]
            ctx.check_abort()

            if col_field:
                cq = self._duck_quote_ident(col_field)
                col_values = [
                    str(r[0] if r[0] is not None else "")
                    for r in self.db.execute(
                        f"SELECT DISTINCT COALESCE({cq}, '') FROM {t} {where_clause} ORDER BY 1"
                    ).fetchall()
                ]
                if len(col_values) > MAX_PIVOT_COLUMN_VALUES:
                    raise RuntimeError(
                        f"列字段 {col_field} 去重值过多（{len(col_values)}），"
                        f"请先筛选后重试，或清空列字段。当前上限为 {MAX_PIVOT_COLUMN_VALUES}。"
                    )
                select_cols = ", ".join(rq)
                pivot_parts = []
                for v in col_values:
                    lit = self._duck_quote_literal(v)
                    alias = self._duck_quote_ident(v if v else "<空白>")
                    pivot_parts.append(f"SUM(CASE WHEN __col={lit} THEN metric ELSE 0 END) AS {alias}")
                pivot_expr = ",\n  ".join(pivot_parts) if pivot_parts else "SUM(metric) AS metric"
                sql = f"""
                    WITH base AS (
                      SELECT {select_cols}, COALESCE({cq}, '') AS __col, {metric_expr} AS metric
                      FROM {t}
                      {where_clause}
                    )
                    SELECT {select_cols},
                      {pivot_expr}
                    FROM base
                    GROUP BY {select_cols}
                    ORDER BY {select_cols}
                """
            else:
                select_cols = ", ".join(rq)
                sql = f"""
                    SELECT {select_cols},
                           SUM({metric_expr}) AS {self._duck_quote_ident(value_field)}
                    FROM {t}
                    {where_clause}
                    GROUP BY {select_cols}
                    ORDER BY {select_cols}
                """

            ctx.stage("阶段 3/5：执行透视聚合计算...")
            ctx.progress(55, "阶段 3/5：执行透视聚合计算...")
            ctx.check_abort()
            cur = self.db.execute(sql)
            columns = [d[0] for d in cur.description]
            if csv_stream_mode:
                ctx.stage("阶段 4/5：边算边写 CSV（仅重复填充）...")
                ctx.progress(68, "阶段 4/5：边算边写 CSV（仅重复填充）...")
                rows_written = self._stream_query_to_csv_with_postprocess(
                    cur=cur,
                    columns=columns,
                    row_fields=row_fields,
                    cfg=cfg,
                    ctx=ctx,
                )
                ctx.progress(100, "100% 完成")
                elapsed = time.monotonic() - ctx.start_ts
                return {
                    "save_path": cfg["save_path"],
                    "rows": rows_written,
                    "cols": len(columns),
                    "elapsed_sec": elapsed,
                    "timings": dict(ctx.timings),
                    "notes": list(ctx.notes),
                }
            else:
                rows = [list(r) for r in cur.fetchall()]
                ctx.stage(f"阶段 4/5：聚合完成，共 {len(rows):,} 行，处理重复填充...")
                ctx.progress(78, f"阶段 4/5：处理 {len(rows):,} 行...")
                ctx.check_abort()
                rows = self._apply_postprocess_rows(rows, columns, row_fields)

                self.pivot_result_cache[cache_key] = (columns[:], [r[:] for r in rows])
                ctx.stage("阶段 5/5：写出结果文件...")
                ctx.progress(90, "阶段 5/5：写出结果文件...")
                ctx.check_abort()
                self._write_output(columns, rows, cfg, src)
                ctx.progress(100, "100% 完成")
                elapsed = time.monotonic() - ctx.start_ts
                return {
                    "save_path": cfg["save_path"],
                    "rows": len(rows),
                    "cols": len(columns),
                    "elapsed_sec": elapsed,
                    "timings": dict(ctx.timings),
                    "notes": list(ctx.notes),
                }

    def _compute_pivot_result(self, src: DataSource, cfg: Dict) -> Tuple[List[str], List[List[object]]]:
        row_fields = [f for f in cfg["row_fields"] if f in src.headers]
        if not row_fields:
            raise RuntimeError("行字段为空或无效。")
        col_field = cfg["column_field"] if cfg["column_field"] in src.headers else ""
        value_field = cfg["value_field"]
        if value_field not in src.headers:
            raise RuntimeError(f"值字段不存在：{value_field}")
        filters = {k: v for k, v in cfg["filters"].items() if k in src.headers and len(v) > 0}
        agg = cfg["agg"] if cfg["agg"] in {"sum", "count"} else "sum"

        cache_key = self._pivot_cache_key(src, cfg, filters, row_fields, col_field, value_field, agg)
        cached = self.pivot_result_cache.get(cache_key)
        if cached is not None:
            columns, rows = cached
            return columns[:], [r[:] for r in rows]
        columns, rows = self._compute_pivot_polars(row_fields, col_field, value_field, filters, agg)
        rows = self._apply_postprocess_rows(rows, columns, row_fields)
        self.pivot_result_cache[cache_key] = (columns[:], [r[:] for r in rows])
        return columns, rows

        where_clause = self._build_where_clause(filters)
        metric_expr = (
            f"COALESCE(TRY_CAST({self._duck_quote_ident(value_field)} AS DOUBLE),0)"
            if agg == "sum"
            else f"CASE WHEN TRIM(COALESCE({self._duck_quote_ident(value_field)},''))<>'' THEN 1 ELSE 0 END"
        )
        t = self.source_table_name
        rq = [self._duck_quote_ident(c) for c in row_fields]
        if col_field:
            cq = self._duck_quote_ident(col_field)
            col_values = [
                str(r[0] if r[0] is not None else "")
                for r in self.db.execute(
                    f"SELECT DISTINCT COALESCE({cq}, '') FROM {t} {where_clause} ORDER BY 1"
                ).fetchall()
            ]
            if len(col_values) > MAX_PIVOT_COLUMN_VALUES:
                raise RuntimeError(
                    f"列字段 {col_field} 去重值过多（{len(col_values)}），"
                    f"请先筛选后重试，或清空列字段。当前上限为 {MAX_PIVOT_COLUMN_VALUES}。"
                )
            select_cols = ", ".join(rq)
            pivot_parts = []
            for v in col_values:
                lit = self._duck_quote_literal(v)
                alias = self._duck_quote_ident(v if v else "<空白>")
                pivot_parts.append(f"SUM(CASE WHEN __col={lit} THEN metric ELSE 0 END) AS {alias}")
            pivot_expr = ",\n  ".join(pivot_parts) if pivot_parts else "SUM(metric) AS metric"
            sql = f"""
                WITH base AS (
                  SELECT {select_cols}, COALESCE({cq}, '') AS __col, {metric_expr} AS metric
                  FROM {t}
                  {where_clause}
                )
                SELECT {select_cols},
                  {pivot_expr}
                FROM base
                GROUP BY {select_cols}
                ORDER BY {select_cols}
            """
        else:
            select_cols = ", ".join(rq)
            sql = f"""
                SELECT {select_cols},
                       SUM({metric_expr}) AS {self._duck_quote_ident(value_field)}
                FROM {t}
                {where_clause}
                GROUP BY {select_cols}
                ORDER BY {select_cols}
            """
        cur = self.db.execute(sql)
        columns = [d[0] for d in cur.description]
        rows = [list(r) for r in cur.fetchall()]
        rows = self._apply_postprocess_rows(rows, columns, row_fields)
        self.pivot_result_cache[cache_key] = (columns[:], [r[:] for r in rows])
        return columns, rows

    def _build_and_export_dual_default(self, save_path: str, filters: Dict[str, Set[str]], ctx: TaskContext) -> Dict[str, object]:
        if not self.data_source:
            raise RuntimeError("数据源不存在。")
        src = self.data_source
        value_field = "Hours" if "Hours" in src.headers else (src.headers[0] if src.headers else "")
        if not value_field:
            raise RuntimeError("未找到可用值字段。")
        col_field = self._default_column_field()

        cfg_mgr = {
            "mode": "by_manager",
            "row_fields": self._build_default_row_fields("by_manager"),
            "column_field": col_field,
            "value_field": value_field,
            "agg": "sum",
            "filters": filters,
        }
        cfg_prj = {
            "mode": "by_project",
            "row_fields": self._build_default_row_fields("by_project"),
            "column_field": col_field,
            "value_field": value_field,
            "agg": "sum",
            "filters": filters,
        }

        required_cols: List[str] = []
        for cfg in (cfg_mgr, cfg_prj):
            for c in cfg["row_fields"]:
                if c and c not in required_cols:
                    required_cols.append(c)
            if cfg["column_field"] and cfg["column_field"] not in required_cols:
                required_cols.append(cfg["column_field"])
            if cfg["value_field"] and cfg["value_field"] not in required_cols:
                required_cols.append(cfg["value_field"])
        for f in filters.keys():
            if f and f not in required_cols:
                required_cols.append(f)

        ctx.progress(18, "阶段 1/4：加载所需列...")
        self._ensure_duckdb_source_loaded(required_cols=required_cols, ctx=ctx)
        ctx.progress(35, "阶段 2/4：计算 by经理...")
        cols_mgr, rows_mgr = self._compute_pivot_result(src, cfg_mgr)
        ctx.progress(62, "阶段 3/4：计算 by项目...")
        cols_prj, rows_prj = self._compute_pivot_result(src, cfg_prj)

        ctx.progress(85, "阶段 4/4：写出双sheet Excel...")
        wb = xlsxwriter.Workbook(save_path)
        ws_mgr = wb.add_worksheet("by经理")
        ws_prj = wb.add_worksheet("by项目")

        def on_mgr(done: int, total: int) -> None:
            pct = 85 + int((done / max(total, 1)) * 6)
            ctx.progress(pct, f"阶段 4/4：写出 by经理... {done:,}/{total:,}")

        def on_prj(done: int, total: int) -> None:
            pct = 91 + int((done / max(total, 1)) * 8)
            ctx.progress(pct, f"阶段 4/4：写出 by项目... {done:,}/{total:,}")

        self._write_excel_sheet(ws_mgr, cols_mgr, rows_mgr, cfg_mgr, wb, progress_cb=on_mgr)
        self._write_excel_sheet(ws_prj, cols_prj, rows_prj, cfg_prj, wb, progress_cb=on_prj)
        wb.close()
        ctx.progress(100, "100% 完成")

        elapsed = time.monotonic() - ctx.start_ts
        return {
            "save_path": save_path,
            "rows": len(rows_mgr) + len(rows_prj),
            "cols": max(len(cols_mgr), len(cols_prj)),
            "elapsed_sec": elapsed,
            "timings": dict(ctx.timings),
            "notes": list(ctx.notes),
        }

    def _pivot_cache_key(
        self,
        src: DataSource,
        cfg: Dict,
        filters: Dict[str, Set[str]],
        row_fields: Sequence[str],
        col_field: str,
        value_field: str,
        agg: str,
    ) -> str:
        src_key = self._source_cache_key(src)
        payload = {
            "src_key": src_key,
            "mode": cfg.get("mode", ""),
            "row_fields": list(row_fields),
            "col_field": col_field,
            "value_field": value_field,
            "agg": agg,
            "filters": {k: sorted(list(v)) for k, v in sorted(filters.items(), key=lambda x: x[0])},
        }
        return json.dumps(payload, ensure_ascii=False, sort_keys=True)

    def _build_where_clause(self, filters: Dict[str, Set[str]]) -> str:
        if not filters:
            return ""
        clauses = []
        for field, vals in filters.items():
            if not vals:
                continue
            q = self._duck_quote_ident(field)
            lits = ", ".join(self._duck_quote_literal(v) for v in sorted(vals))
            clauses.append(f"COALESCE({q}, '') IN ({lits})")
        if not clauses:
            return ""
        return "WHERE " + " AND ".join(clauses)

    def _repeat_fill_row_labels_rows(
        self, rows: List[List[object]], columns: List[str], row_fields: Sequence[str]
    ) -> List[List[object]]:
        state = init_postprocess_state(columns, row_fields)
        out: List[List[object]] = []
        for row in rows:
            out.extend(consume_postprocess_row(state, row))
        return out

    def _apply_postprocess_rows(
        self,
        rows: List[List[object]],
        columns: List[str],
        row_fields: Sequence[str],
    ) -> List[List[object]]:
        state = init_postprocess_state(columns=columns, row_fields=row_fields)
        out: List[List[object]] = []
        for row in rows:
            out.extend(consume_postprocess_row(state, row))
        out.extend(finalize_postprocess_state(state))
        return out

    def _stream_query_to_csv_with_postprocess(
        self,
        cur,
        columns: List[str],
        row_fields: Sequence[str],
        cfg: Dict,
        ctx: TaskContext,
    ) -> int:
        save_path = str(cfg["save_path"])
        state = init_postprocess_state(columns=columns, row_fields=row_fields)
        rows_written = 0
        batch_size = 5000
        next_report_at = 50000

        def write_row(writer_obj, row_obj: List[object]) -> None:
            writer_obj.writerow(["" if v is None else v for v in row_obj])

        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            while True:
                ctx.check_abort()
                batch = cur.fetchmany(batch_size)
                if not batch:
                    break
                for raw in batch:
                    emitted_rows = consume_postprocess_row(state, raw)
                    for row in emitted_rows:
                        write_row(writer, row)
                        rows_written += 1
                    if rows_written >= next_report_at:
                        ctx.stage(f"阶段 4/5：已写出 {rows_written:,} 行...")
                        pct = min(95, 68 + (rows_written // 50000) * 3)
                        ctx.progress(int(pct), f"阶段 4/5：已写出 {rows_written:,} 行...")
                        next_report_at += 50000

                ctx.check_abort()

            for tail_row in finalize_postprocess_state(state):
                write_row(writer, tail_row)
                rows_written += 1

        return rows_written

    def _write_output(self, columns: List[str], rows: List[List[object]], cfg: Dict, src: DataSource) -> None:
        save_path = str(cfg["save_path"])
        ext = Path(save_path).suffix.lower()
        if ext == ".csv":
            self._write_csv(columns, rows, save_path)
            return
        self._write_excel(columns, rows, cfg, src)

    def _write_csv(self, columns: List[str], rows: List[List[object]], save_path: str) -> None:
        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(columns)
            for r in rows:
                w.writerow(["" if v is None else v for v in r])

    def _write_excel_sheet(
        self,
        ws,
        columns: List[str],
        rows: List[List[object]],
        cfg: Dict,
        wb,
        progress_cb: Optional[Callable[[int, int], None]] = None,
    ) -> None:
        num_fmt = wb.add_format({"num_format": "#,##0.00"})
        int_fmt = wb.add_format({"num_format": "#,##0"})
        hdr_fmt = wb.add_format({"bold": True})
        row_field_set = set(cfg.get("row_fields", []))

        for cidx, c in enumerate(columns):
            ws.write(0, cidx, c, hdr_fmt)
        for ridx, row in enumerate(rows, start=1):
            for cidx, val in enumerate(row):
                is_row_field = columns[cidx] in row_field_set
                if is_row_field:
                    ws.write(ridx, cidx, "" if val is None else val)
                else:
                    try:
                        fv = float(val) if val not in (None, "") else 0.0
                        ws.write_number(ridx, cidx, fv, int_fmt if cfg.get("agg") == "count" else num_fmt)
                    except Exception:
                        ws.write(ridx, cidx, "" if val is None else val)
            if progress_cb is not None and (ridx % 400 == 0 or ridx == len(rows)):
                try:
                    progress_cb(ridx, len(rows))
                except Exception:
                    pass

        # J2: 冻结首行 + A~I列
        ws.freeze_panes(1, 9)
        ws.autofilter(0, 0, max(len(rows), 1), max(len(columns) - 1, 0))
        for cidx, c in enumerate(columns):
            base_width = max(12, min(38, len(str(c)) + 4))
            width = base_width if cidx == 0 else max(6, base_width / 2)
            if c not in row_field_set:
                ws.set_column(cidx, cidx, width, int_fmt if cfg.get("agg") == "count" else num_fmt)
            else:
                ws.set_column(cidx, cidx, width)
        # 动态分组：从J列开始，直到“最后4列之前”的最后一列
        start_idx = 9  # J
        end_idx = len(columns) - 5  # 预留最后4列不分组
        if end_idx >= start_idx:
            start_col = xl_col_to_name(start_idx)
            end_col = xl_col_to_name(end_idx)
            ws.set_column(f"{start_col}:{end_col}", None, None, {"level": 1, "hidden": True})
            collapse_idx = end_idx + 1
            if collapse_idx < len(columns):
                collapse_col = xl_col_to_name(collapse_idx)
                ws.set_column(f"{collapse_col}:{collapse_col}", None, None, {"collapsed": True})

    def _write_excel(self, columns: List[str], rows: List[List[object]], cfg: Dict, src: DataSource) -> None:
        save_path = cfg["save_path"]
        wb = xlsxwriter.Workbook(save_path)
        ws = wb.add_worksheet("Pivot")
        ws_cfg = wb.add_worksheet("Config")
        self._write_excel_sheet(ws, columns, rows, cfg, wb)

        cfg_rows = self._build_config_sheet_rows(cfg, src)
        hdr_fmt = wb.add_format({"bold": True})
        ws_cfg.write(0, 0, "参数", hdr_fmt)
        ws_cfg.write(0, 1, "值", hdr_fmt)
        for r, (k, v) in enumerate(cfg_rows, start=1):
            ws_cfg.write(r, 0, k)
            ws_cfg.write(r, 1, v)
        ws_cfg.freeze_panes(1, 0)
        ws_cfg.set_column(0, 0, 28)
        ws_cfg.set_column(1, 1, 120)

        wb.close()

    @staticmethod
    def _build_config_sheet_rows(cfg: Dict, src: DataSource) -> List[Tuple[str, str]]:
        filters_str = []
        for k, vals in cfg["filters"].items():
            preview = ", ".join(list(vals)[:20])
            if len(vals) > 20:
                preview += f" ...(+{len(vals)-20})"
            filters_str.append(f"{k}: {preview}")
        return [
            ("来源文件", src.original_path),
            ("来源Sheet", src.sheet_name or "-"),
            ("导出模式", cfg["mode"]),
            ("行字段", " | ".join(cfg["row_fields"])),
            ("列字段", cfg["column_field"] or "-"),
            ("值字段", cfg["value_field"]),
            ("汇总方式", cfg["agg"]),
            ("筛选条件", " || ".join(filters_str) if filters_str else "(无)"),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

    def _on_export_done(self, payload: object) -> None:
        d = payload  # type: ignore[assignment]
        timings = d.get("timings", {}) if isinstance(d, dict) else {}
        notes = d.get("notes", []) if isinstance(d, dict) else []
        stage1_lines: List[str] = []
        if isinstance(timings, dict):
            items = [(k, float(v)) for k, v in timings.items() if str(k).startswith("stage1_")]
            items.sort(key=lambda x: x[1], reverse=True)
            name_map = {
                "stage1_lock_wait_and_cache_check": "缓存检查",
                "stage1_excel_to_full_parquet": "Excel转全量Parquet",
                "stage1_load_subset_cache": "装载列级缓存",
                "stage1_project_from_full_parquet": "全量缓存裁剪列",
                "stage1_load_full_cache": "装载全量缓存",
                "stage1_excel_to_subset_parquet": "Excel转列级Parquet",
                "stage1_load_built_subset_cache": "装载新建列缓存",
                "stage1_fallback_load_source": "回退路径加载",
                "stage1_total_prepare_source": "阶段一总计",
            }
            for k, sec in items[:5]:
                stage1_lines.append(f"- {name_map.get(k, k)}: {sec:.2f}s")
        extra = ""
        if stage1_lines:
            extra = "\n\n阶段一细分耗时（Top5）：\n" + "\n".join(stage1_lines)
        warn = ""
        if isinstance(notes, list) and notes:
            warn_lines = [f"- {str(x)}" for x in notes[:3]]
            warn = "\n\n快路径诊断：\n" + "\n".join(warn_lines)
        messagebox.showinfo(
            "导出完成",
            f"已导出：\n{d['save_path']}\n\n结果规模：{d['rows']:,} 行 × {d['cols']:,} 列\n耗时：{float(d.get('elapsed_sec', 0.0)):.1f} 秒{extra}{warn}",
        )

    @staticmethod
    def _detect_sep_encoding(path: str) -> Tuple[str, str]:
        with open(path, "rb") as f:
            head = f.read(4)
        if head.startswith(b"\xff\xfe"):
            enc = "utf-16-le"
        elif head.startswith(b"\xfe\xff"):
            enc = "utf-16-be"
        elif head.startswith(b"\xef\xbb\xbf"):
            enc = "utf-8-sig"
        else:
            enc = "utf-8-sig"

        candidates = [enc, "utf-8-sig", "gb18030", "cp936", "utf-16-le", "utf-16-be"]
        seps = ["\t", ",", ";", "|"]

        best = (",", enc, -1)
        for e in candidates:
            try:
                with open(path, "r", encoding=e, errors="strict") as f:
                    lines = []
                    for _ in range(50):
                        line = f.readline()
                        if not line:
                            break
                        if line.strip():
                            lines.append(line.rstrip("\n\r"))
                    if not lines:
                        continue
                for sep in seps:
                    counts = [len(x.split(sep)) for x in lines]
                    score = max(counts)
                    if score > best[2]:
                        best = (sep, e, score)
            except Exception:
                continue
        return best[0], best[1]


def main() -> None:
    root = tk.Tk()
    app = TimesheetPivotApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
