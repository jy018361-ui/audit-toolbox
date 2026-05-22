"""
瀵煎嚭妯″潡
鏀寔瀵煎嚭涓篍xcel鍜孋SV鏍煎紡
"""
import pandas as pd
import os
import re
from datetime import date
from typing import List, Optional, Dict
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from config import DEFAULT_EXPORT_FORMAT, DEPRECIATION_FORMULA_ROW_LIMIT, DEPRECIATION_FORMULA_SAMPLE_ROWS
from summary_generator import SummaryGenerator
from sheet_generator import SheetGenerator
from pivot_engine import PivotEngine


class Exporter:
    """Exporter"""
    
    def __init__(self):
        self.export_progress_callback = None
        self.summary_generator = SummaryGenerator()
        self.sheet_generator = SheetGenerator()
        self.pivot_engine = PivotEngine()
        self._export_notes = []
        self._template_map_cache = {}

    @staticmethod
    def _make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Ensure duplicate headers do not turn df[col] into a DataFrame."""
        if not isinstance(df, pd.DataFrame) or df.empty or df.columns.is_unique:
            return df
        out = df.copy()
        seen = {}
        new_cols = []
        for col in out.columns:
            text = str(col)
            count = seen.get(text, 0)
            if count == 0:
                new_cols.append(col)
            else:
                new_cols.append(f"{text}_{count + 1}")
            seen[text] = count + 1
        out.columns = new_cols
        return out
    
    def export_dataframe(
        self,
        df: pd.DataFrame,
        file_path: str,
        selected_columns: Optional[List[str]] = None,
        format: str = DEFAULT_EXPORT_FORMAT,
        pivot_df: Optional[pd.DataFrame] = None,
        full_df: Optional[pd.DataFrame] = None,
        summary_config: Optional[Dict] = None  # 姹囨€昏〃閰嶇疆
    ) -> tuple:
        """
        瀵煎嚭DataFrame鍒版枃浠?        
        Args:
            df: 瑕佸鍑虹殑DataFrame
            file_path: 瀵煎嚭鏂囦欢璺緞
            selected_columns: 瑕佸鍑虹殑鍒楀垪琛紙None琛ㄧず瀵煎嚭鎵€鏈夊垪锛?            format: 瀵煎嚭鏍煎紡 ('xlsx' 鎴?'csv')
            
        Returns:
            tuple: (鎴愬姛鏍囧織, 閿欒娑堟伅)
        """
        try:
            if df is None or df.empty:
                return False, "data is empty"
            
            if selected_columns:
                missing_cols = [col for col in selected_columns if col not in df.columns]
                if missing_cols:
                    return False, f"浠ヤ笅鍒椾笉瀛樺湪: {', '.join(missing_cols)}"
                
                df_export = df[selected_columns].copy()
            else:
                df_export = df.copy()
            df_export = self._make_unique_columns(df_export)
            
            # 纭繚鐩綍瀛樺湪
            os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)
            
            # 鏍规嵁鏍煎紡瀵煎嚭
            full_df_for_lists = self._make_unique_columns(full_df if full_df is not None else df)
            
            if format.lower() == 'xlsx':
                return self._export_excel(df_export, file_path, pivot_df, full_df_for_lists, summary_config)
            elif format.lower() == 'csv':
                return self._export_csv(df_export, file_path, pivot_df, full_df_for_lists, summary_config)
            else:
                return False, f"涓嶆敮鎸佺殑瀵煎嚭鏍煎紡: {format}"
                
        except Exception as e:
            return False, f"瀵煎嚭鏃跺嚭閿? {str(e)}"
    
    def _export_excel(self, df: pd.DataFrame, file_path: str, pivot_df: Optional[pd.DataFrame] = None,
                       full_df: Optional[pd.DataFrame] = None, summary_config: Optional[Dict] = None) -> tuple:
        """导出 Excel（统一使用 xlsxwriter）。"""
        try:
            self._export_notes = []
            self._template_map_cache = {}
            if not file_path.endswith('.xlsx'):
                file_path = os.path.splitext(file_path)[0] + '.xlsx'

            data_for_lists = full_df if full_df is not None else df
            correction_warnings = []
            sheets = []
            used_sheet_names = set()

            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                wb = writer.book
                fmts = self._build_xlsx_formats(wb)

                # 合并数据
                sheet_name = self._reserve_sheet_name('合并数据', used_sheet_names)
                merged_df, _, _, _ = self._enhance_duplicate_display(df, None, None, None, summary_config=summary_config)
                merged_df_for_pivot = merged_df.copy()
                summary_input_df = merged_df_for_pivot
                merged_df = self._coerce_sheet_df_for_write(merged_df, sheet_name, summary_config)
                # 先写，后续若生成了FA/新增清单则再按重复ID场景回写增强展示
                merged_df.to_excel(writer, index=False, sheet_name=sheet_name)
                self._format_sheet_xlsxwriter(writer, sheet_name, merged_df, fmts, summary_config=summary_config)
                sheets.append(sheet_name)

                # 数据透视表
                pivot_for_export = self._rebuild_pivot_from_export_config(merged_df_for_pivot, summary_config)
                if pivot_for_export is None:
                    pivot_for_export = pivot_df
                if pivot_for_export is not None and not pivot_for_export.empty:
                    pivot_export_df = self._prepare_pivot_export_df(pivot_for_export)
                    pivot_sheet = self._reserve_sheet_name('数据透视表', used_sheet_names)
                    pivot_export_df = self._coerce_sheet_df_for_write(pivot_export_df, pivot_sheet, summary_config)
                    pivot_export_df.to_excel(writer, index=False, sheet_name=pivot_sheet, merge_cells=False)
                    self._format_sheet_xlsxwriter(
                        writer,
                        pivot_sheet,
                        pivot_export_df,
                        fmts,
                        summary_config=summary_config,
                        merge_total_ab=True,
                    )
                    sheets.append(pivot_sheet)

                if summary_config and summary_input_df is not None:
                    summary_field_mapping = summary_config.get('field_mapping', {}) or {}
                    field_mapping = summary_config.get('field_mapping', {})
                    self.sheet_generator.set_config(
                        field_mapping=field_mapping,
                        match_col=summary_config.get('match_col'),
                        match_col2=summary_config.get('match_col2'),
                        match_cols=summary_config.get('match_cols'),
                        match_cols2=summary_config.get('match_cols2'),
                        category_col=summary_config.get('category_col'),
                        original_value_col1=summary_config.get('original_value_col1'),
                        original_value_col2=summary_config.get('original_value_col2'),
                        depreciation_col1=summary_config.get('depreciation_col1'),
                        depreciation_col2=summary_config.get('depreciation_col2'),
                        use_supplement_lists=bool(summary_config.get('use_supplement_lists')),
                    )
                    disp_success, disp_msg, disp_df = self.sheet_generator.generate_disposal_list(data_for_lists)
                    try:
                        success, msg, _summary_data = self.summary_generator.generate_summary(
                            summary_input_df,
                            category_col=summary_config.get('category_col'),
                            category_col1=summary_config.get('category_col1'),
                            category_col2=summary_config.get('category_col2'),
                            original_value_col1=summary_config.get('original_value_col1'),
                            original_value_col2=summary_config.get('original_value_col2'),
                            depreciation_col1=summary_config.get('depreciation_col1'),
                            depreciation_col2=summary_config.get('depreciation_col2'),
                            file1_display_name=summary_config.get('file1_display_name', '期初'),
                            file2_display_name=summary_config.get('file2_display_name', '期末'),
                            pivot_row_field_for_log=summary_config.get('pivot_row_field'),
                            addition_method_col1=summary_field_mapping.get('addition_method_col1'),
                            addition_method_col2=summary_field_mapping.get('addition_method_col2'),
                            disposal_method_col1=summary_field_mapping.get('disposal_method_col1'),
                            disposal_method_col2=summary_field_mapping.get('disposal_method_col2'),
                            disposal_orig_col1=summary_field_mapping.get('disposal_orig_col1'),
                            disposal_orig_col2=summary_field_mapping.get('disposal_orig_col2'),
                            disposal_dep_col1=summary_field_mapping.get('disposal_dep_col1'),
                            disposal_dep_col2=summary_field_mapping.get('disposal_dep_col2'),
                            match_col=summary_config.get('match_col'),
                            match_cols=summary_config.get('match_cols'),
                            disposal_bkd_df=disp_df,
                            extended_mode=bool(summary_config.get('extended_summary_mode')),
                            use_supplement_lists=bool(summary_config.get('use_supplement_lists', True)),
                        )
                        if success:
                            summary_sheet = self._reserve_sheet_name('固定资产变动汇总表', used_sheet_names)
                            if self._write_summary_sheet_xlsxwriter(writer, summary_sheet, fmts):
                                sheets.append(summary_sheet)
                    except Exception as e:
                        print(f"生成固定资产变动汇总表出错: {str(e)}")

                    try:
                        self.sheet_generator.set_config(
                            field_mapping=field_mapping,
                            match_col=summary_config.get('match_col'),
                            match_col2=summary_config.get('match_col2'),
                            match_cols=summary_config.get('match_cols'),
                            match_cols2=summary_config.get('match_cols2'),
                            category_col=summary_config.get('category_col'),
                            original_value_col1=summary_config.get('original_value_col1'),
                            original_value_col2=summary_config.get('original_value_col2'),
                            depreciation_col1=summary_config.get('depreciation_col1'),
                            depreciation_col2=summary_config.get('depreciation_col2'),
                            use_supplement_lists=bool(summary_config.get('use_supplement_lists')),
                        )

                        fa_success, fa_msg, fa_df = self._build_fa_list_from_merged_file2(merged_df, summary_config)
                        if fa_success and fa_df is not None and not fa_df.empty:
                            # 仅增强FA展示；合并数据sheet已在前面写入，避免重复写盘
                            _, fa_df, _, _ = self._enhance_duplicate_display(pd.DataFrame(), fa_df, None, None, summary_config=summary_config)
                            fa_sheet = self._reserve_sheet_name('FA List', used_sheet_names)
                            fa_df = self._coerce_sheet_df_for_write(fa_df, fa_sheet, summary_config)
                            fa_df.to_excel(writer, index=False, sheet_name=fa_sheet)
                            self._format_sheet_xlsxwriter(writer, fa_sheet, fa_df, fmts, summary_config=summary_config)
                            sheets.append(fa_sheet)
                        if fa_msg:
                            for line in str(fa_msg).split('\n'):
                                if "【" in line and line not in correction_warnings:
                                    correction_warnings.append(line)

                        add_success, add_msg, add_df = self.sheet_generator.generate_addition_list(data_for_lists)
                        if add_df is None or add_df.empty:
                            add_df = pd.DataFrame(columns=["资产类别", "固定资产编号", "固定资产名称", "入账开始日期", "使用寿命(月)", "残值率", "新增方式", "新增时间", "增加类型", "原值增加"])
                            if not add_success:
                                correction_warnings.append(f"生成新增清单_BKD失败: {add_msg}")
                        if not self._has_user_addition_mapping(summary_config):
                            add_df = self._apply_addition_placeholders(add_df)
                        add_df = self._reorder_addition_tail_columns(add_df)
                        # 仅增强新增清单展示；合并数据sheet保持单次写入
                        _, _, add_df, _ = self._enhance_duplicate_display(pd.DataFrame(), None, add_df, None, summary_config=summary_config)
                        add_sheet = self._reserve_sheet_name('新增清单_BKD', used_sheet_names)
                        add_df = self._coerce_sheet_df_for_write(add_df, add_sheet, summary_config)
                        add_df.to_excel(writer, index=False, sheet_name=add_sheet)
                        self._format_sheet_xlsxwriter(writer, add_sheet, add_df, fmts, summary_config=summary_config)
                        sheets.append(add_sheet)
                        if add_msg:
                            for line in str(add_msg).split('\n'):
                                if "【" in line and line not in correction_warnings:
                                    correction_warnings.append(line)

                        if disp_success and disp_df is not None and not disp_df.empty:
                            if not self._has_user_disposal_mapping(summary_config):
                                disp_df = self._apply_placeholder_tail_columns(
                                    disp_df,
                                    placeholder_headers=["[处置方式?]", "[处置时间?]", "[处置原值?]", "[处置折旧?]"]
                                )
                            _, _, _, disp_df = self._enhance_duplicate_display(pd.DataFrame(), None, None, disp_df, summary_config=summary_config)
                            disp_sheet = self._reserve_sheet_name('处置清单_BKD', used_sheet_names)
                            disp_df = self._coerce_sheet_df_for_write(disp_df, disp_sheet, summary_config)
                            disp_df.to_excel(writer, index=False, sheet_name=disp_sheet)
                            self._format_sheet_xlsxwriter(writer, disp_sheet, disp_df, fmts, summary_config=summary_config)
                            sheets.append(disp_sheet)
                        if disp_msg:
                            for line in str(disp_msg).split('\n'):
                                if "【" in line and line not in correction_warnings:
                                    correction_warnings.append(line)

                        dep_df = self._build_depreciation_period_df(data_for_lists, summary_config, fa_df=fa_df)
                        if dep_df is not None and not dep_df.empty:
                            dep_sheet = self._reserve_sheet_name('折旧期间', used_sheet_names)
                            dep_df = self._coerce_sheet_df_for_write(dep_df, dep_sheet, summary_config)
                            dep_df.to_excel(writer, index=False, sheet_name=dep_sheet)
                            self._format_sheet_xlsxwriter(writer, dep_sheet, dep_df, fmts, summary_config=summary_config)
                            sheets.append(dep_sheet)
                    except Exception as e:
                        import traceback
                        error_msg = f"生成FA List/BKD清单时出错: {str(e)}\n{traceback.format_exc()}"
                        print(error_msg)
                        correction_warnings.append(f"错误: {str(e)}")

                msg = f"成功导出到: {file_path}"
            if len(sheets) > 1:
                msg += f"\n(包含{len(sheets)}个sheet: {', '.join(sheets)})"
            for note in getattr(self, "_export_notes", []):
                if note not in correction_warnings:
                    correction_warnings.append(note)
            if correction_warnings:
                msg += "\n\n===CORRECTION_WARNINGS===\n" + "\n".join(correction_warnings)
            if summary_config and summary_config.get("has_unmatched_supplement"):
                add_unmatched = summary_config.get("unmatched_add_df")
                disp_unmatched = summary_config.get("unmatched_disp_df")
                has_unmatched_rows = (
                    isinstance(add_unmatched, pd.DataFrame) and not add_unmatched.empty
                ) or (
                    isinstance(disp_unmatched, pd.DataFrame) and not disp_unmatched.empty
                )
                if has_unmatched_rows:
                    ok_unmatched, info_unmatched = self._export_unmatched_change_workbook(
                        file_path, add_unmatched, disp_unmatched
                    )
                    if ok_unmatched:
                        msg += f"\n未匹配资产变动清单导出到: {info_unmatched}"
                        if "===CORRECTION_WARNINGS===" not in msg:
                            msg += "\n\n===CORRECTION_WARNINGS==="
                        msg += f"\n【未匹配资产变动清单】检测到补充清单存在未匹配记录，已导出：{info_unmatched}"
                    else:
                        if "===CORRECTION_WARNINGS===" not in msg:
                            msg += "\n\n===CORRECTION_WARNINGS==="
                        msg += f"\n【未匹配资产变动清单】检测到补充清单存在未匹配记录，但导出失败：{info_unmatched}"
            return True, msg
        except Exception as e:
            return False, f"导出Excel时出错: {str(e)}"
    def _export_csv(self, df: pd.DataFrame, file_path: str, pivot_df: Optional[pd.DataFrame] = None, full_df: Optional[pd.DataFrame] = None, summary_config: Optional[Dict] = None, encoding: str = 'utf-8-sig') -> tuple:
        """???CSV???"""
        try:
            if not file_path.endswith('.csv'):
                file_path = os.path.splitext(file_path)[0] + '.csv'
            
            base_path = os.path.splitext(file_path)[0]
            files_created = []
            
            # 浼樺寲CSV瀵煎嚭鍙傛暟浠ユ彁楂橀€熷害
            # 浣跨敤index=False閬垮厤鍐欏叆绱㈠紩
            # 浣跨敤encoding='utf-8-sig'纭繚Excel鍏煎鎬?            # 瀵逛簬澶ф枃浠讹紝鍙互鑰冭檻浣跨敤chunksize锛屼絾杩欓噷涓轰簡绠€鍖栵紝鍏堜笉浣跨敤
            csv_kwargs = {
                'index': False,
                'encoding': encoding,
                'mode': 'w',  # 鏄庣‘鎸囧畾鍐欏叆妯″紡
                'lineterminator': '\n'  # 浣跨敤Unix椋庢牸鐨勬崲琛岀锛堟洿蹇級
            }
            
            # 重复ID场景下先做主表展示增强，再导出CSV
            df, _, _, _ = self._enhance_duplicate_display(df, None, None, None, summary_config=summary_config)
            # 瀵煎嚭涓绘暟鎹埌CSV
            df.to_csv(file_path, **csv_kwargs)
            files_created.append(file_path)
            msg = f"鎴愬姛瀵煎嚭鍒? {file_path}"
            
            # 濡傛灉鏈夋暟鎹€忚琛紝瀵煎嚭鍒板崟鐙殑鏂囦欢
            pivot_for_export = self._rebuild_pivot_from_export_config(df, summary_config)
            if pivot_for_export is None:
                pivot_for_export = pivot_df
            if pivot_for_export is not None and not pivot_for_export.empty:
                pivot_file_path = f"{base_path}_鏁版嵁閫忚琛?csv"
                pivot_kwargs = csv_kwargs.copy()
                pivot_kwargs['index'] = True
                pivot_for_export.to_csv(pivot_file_path, **pivot_kwargs)
                files_created.append(pivot_file_path)
                msg += f"\n鏁版嵁閫忚琛ㄥ鍑哄埌: {pivot_file_path}"
            
            data_for_lists = full_df if full_df is not None else df
            summary_input_df = df
            if summary_config and summary_input_df is not None:
                field_mapping = summary_config.get('field_mapping', {})
                self.sheet_generator.set_config(
                    field_mapping=field_mapping,
                    match_col=summary_config.get('match_col'),  # 鍚戝悗鍏煎
                    match_col2=summary_config.get('match_col2'),  # 鍚戝悗鍏煎
                    match_cols=summary_config.get('match_cols'),  # 澶氬垪鏍煎紡
                    match_cols2=summary_config.get('match_cols2'),  # 澶氬垪鏍煎紡
                    category_col=summary_config.get('category_col'),
                    original_value_col1=summary_config.get('original_value_col1'),
                    original_value_col2=summary_config.get('original_value_col2'),
                    depreciation_col1=summary_config.get('depreciation_col1'),
                    depreciation_col2=summary_config.get('depreciation_col2'),
                    use_supplement_lists=bool(summary_config.get('use_supplement_lists')),
                )
                disp_success, disp_msg, disp_df = self.sheet_generator.generate_disposal_list(data_for_lists)
                try:
                    summary_field_mapping = summary_config.get('field_mapping', {}) or {}
                    # 鐢熸垚姹囨€昏〃
                    success, summary_msg, summary_data = self.summary_generator.generate_summary(
                        summary_input_df,
                        category_col=summary_config.get('category_col'),
                        category_col1=summary_config.get('category_col1'),
                        category_col2=summary_config.get('category_col2'),
                        original_value_col1=summary_config.get('original_value_col1'),
                        original_value_col2=summary_config.get('original_value_col2'),
                        depreciation_col1=summary_config.get('depreciation_col1'),
                        depreciation_col2=summary_config.get('depreciation_col2'),
                        file1_display_name=summary_config.get('file1_display_name', '鏈熷垵'),
                        file2_display_name=summary_config.get('file2_display_name', '鏈熸湯'),
                        pivot_row_field_for_log=summary_config.get('pivot_row_field'),
                        addition_method_col1=summary_field_mapping.get('addition_method_col1'),
                        addition_method_col2=summary_field_mapping.get('addition_method_col2'),
                        disposal_method_col1=summary_field_mapping.get('disposal_method_col1'),
                        disposal_method_col2=summary_field_mapping.get('disposal_method_col2'),
                        disposal_orig_col1=summary_field_mapping.get('disposal_orig_col1'),
                        disposal_orig_col2=summary_field_mapping.get('disposal_orig_col2'),
                        disposal_dep_col1=summary_field_mapping.get('disposal_dep_col1'),
                        disposal_dep_col2=summary_field_mapping.get('disposal_dep_col2'),
                        match_col=summary_config.get('match_col'),
                        match_cols=summary_config.get('match_cols'),
                        disposal_bkd_df=disp_df,
                        extended_mode=bool(summary_config.get('extended_summary_mode')),
                        use_supplement_lists=bool(summary_config.get('use_supplement_lists', True)),
                    )
                    
                    if success and summary_data:
                        summary_df = self.summary_generator.get_summary_dataframe()
                        if summary_df is not None and not summary_df.empty:
                            summary_path = f"{base_path}_鍥哄畾璧勪骇鍙樺姩姹囨€昏〃.csv"
                            summary_df.to_csv(summary_path, **csv_kwargs)
                            files_created.append(summary_path)
                            msg += f"\n鍥哄畾璧勪骇鍙樺姩姹囨€昏〃瀵煎嚭鍒? {summary_path}"
                except Exception as e:
                    print(f"鐢熸垚姹囨€昏〃鏃跺嚭閿? {str(e)}")
                
                try:
                    self.sheet_generator.set_config(
                        field_mapping=field_mapping,
                        match_col=summary_config.get('match_col'),  # 鍚戝悗鍏煎
                        match_col2=summary_config.get('match_col2'),  # 鍚戝悗鍏煎
                        match_cols=summary_config.get('match_cols'),  # 澶氬垪鏍煎紡
                        match_cols2=summary_config.get('match_cols2'),  # 澶氬垪鏍煎紡
                        category_col=summary_config.get('category_col'),
                        original_value_col1=summary_config.get('original_value_col1'),
                        original_value_col2=summary_config.get('original_value_col2'),
                        depreciation_col1=summary_config.get('depreciation_col1'),
                        depreciation_col2=summary_config.get('depreciation_col2'),
                        use_supplement_lists=bool(summary_config.get('use_supplement_lists')),
                    )
                    
                    # 鏀堕泦绾犲亸璀﹀憡锛堢敤浜庡脊绐楋級
                    correction_warnings = []
                    
                    # 鐢熸垚FA List
                    fa_success, fa_msg, fa_df = self._build_fa_list_from_merged_file2(df, summary_config)
                    if fa_success and fa_df is not None and not fa_df.empty:
                        fa_path = f"{base_path}_FA List.csv"
                        fa_df.to_csv(fa_path, **csv_kwargs)
                        files_created.append(fa_path)
                        msg += f"\nFA List瀵煎嚭鍒? {fa_path}"
                    
                    if fa_msg:
                        for line in fa_msg.split('\n'):
                            if "??" in line:
                                if line not in correction_warnings:
                                    correction_warnings.append(line)
                    
                    # 鐢熸垚鏂板娓呭崟_BKD
                    add_success, add_msg, add_df = self.sheet_generator.generate_addition_list(data_for_lists)
                    # 纭繚鏂板娓呭崟_BKD涓€瀹氫細琚鍑猴紙鍗充娇澶辫触涔熷鍑虹┖琛級
                    if add_df is None:
                        # 濡傛灉鐢熸垚澶辫触锛屽垱寤哄甫琛ㄥご鐨勭┖DataFrame
                        add_df = pd.DataFrame(columns=["????", "??????", "??????", "??????", "????(?)", "???", "????"])
                        if not add_success:
                            msg += f"\n鐢熸垚鏂板娓呭崟_BKD澶辫触: {add_msg}"
                            print(f"鐢熸垚鏂板娓呭崟_BKD澶辫触: {add_msg}")
                    elif add_df.empty:
                        # 濡傛灉鏁版嵁涓虹┖锛岀‘淇濇湁琛ㄥご
                        add_df = pd.DataFrame(columns=["????", "??????", "??????", "??????", "????(?)", "???", "????"])

                    if not self._has_user_addition_mapping(summary_config):
                        add_df = self._apply_addition_placeholders(add_df)
                    # CSV导出同样做重复ID展示增强
                    _, _, add_df, _ = self._enhance_duplicate_display(pd.DataFrame(), None, add_df, None, summary_config=summary_config)
                    add_df = self._reorder_addition_tail_columns(add_df)
                    
                    add_path = f"{base_path}_鏂板娓呭崟_BKD.csv"
                    add_df.to_csv(add_path, **csv_kwargs)
                    files_created.append(add_path)
                    msg += f"\n鏂板娓呭崟_BKD瀵煎嚭鍒? {add_path}"
                    
                    if add_msg:
                        for line in add_msg.split('\n'):
                            if "??" in line:
                                if line not in correction_warnings:
                                    correction_warnings.append(line)
                    # 濡傛灉娌℃湁鏁版嵁锛屼篃璁板綍鎻愮ず
                    if '娌℃湁鍘熷€煎鍔犵殑璁板綍' in add_msg:
                        msg += "\n??: ????_BKD??????????????????"
                    
                    # 鐢熸垚澶勭疆娓呭崟_BKD
                    if disp_success and disp_df is not None and not disp_df.empty:
                        if not self._has_user_disposal_mapping(summary_config):
                            disp_df = self._apply_placeholder_tail_columns(
                                disp_df,
                                placeholder_headers=["[处置方式?]", "[处置时间?]", "[处置原值?]", "[处置折旧?]"]
                            )
                        _, _, _, disp_df = self._enhance_duplicate_display(pd.DataFrame(), None, None, disp_df, summary_config=summary_config)
                        disp_path = f"{base_path}_澶勭疆娓呭崟_BKD.csv"
                        disp_df.to_csv(disp_path, **csv_kwargs)
                        files_created.append(disp_path)
                        msg += f"\n澶勭疆娓呭崟_BKD瀵煎嚭鍒? {disp_path}"
                    
                    try:
                        dep_msg = self._create_depreciation_period_csv(base_path, data_for_lists, summary_config, csv_kwargs, files_created)
                        if dep_msg:
                            msg += dep_msg
                    except Exception as e:
                        print(f"鐢熸垚鎶樻棫鏈熼棿琛ㄦ牸鏃跺嚭閿? {str(e)}")
                        
                    if correction_warnings:
                        msg += "\n\n===CORRECTION_WARNINGS===\n" + "\n".join(correction_warnings)
                        
                except Exception as e:
                    print(f"鐢熸垚FA List/BKD娓呭崟鏃跺嚭閿? {str(e)}")
            
            return True, msg
            
        except Exception as e:
            return False, f"瀵煎嚭CSV鏃跺嚭閿? {str(e)}"
    
    def _create_depreciation_period_sheet(self, writer, df: pd.DataFrame, summary_config: Dict):
        """
        ????????
        """
        from openpyxl.utils.dataframe import dataframe_to_rows
        import json
        
        # 鑾峰彇鐢ㄦ埛閰嶇疆鐨勫瓧娈碉紙浣跨敤宸叉牸寮忓寲鐨勫垪鍚嶏級
        original_value_col1 = summary_config.get('original_value_col1')  # 鏂囦欢1鍘熷€硷紙鍊煎瓧娈碉級
        original_value_col2 = summary_config.get('original_value_col2')  # 鏂囦欢2鍘熷€硷紙鍊煎瓧娈碉級
        field_mapping = summary_config.get('field_mapping', {})
        # 琛屽瓧娈碉細璧勪骇绫诲埆 + 浣跨敤瀵垮懡锛堝畬鍏ㄤ娇鐢ㄧ敤鎴锋槧灏勭殑瀛楁锛屼笉鍋氫换浣曡嚜鍔ㄥ尮閰嶏級
        category_col1 = field_mapping.get('category_col1')  # 鏂囦欢1璧勪骇绫诲埆
        category_col2 = field_mapping.get('category_col2')  # 鏂囦欢2璧勪骇绫诲埆
        life_col1 = field_mapping.get('life_col1')          # 鏂囦欢1浣跨敤瀵垮懡
        life_col2 = field_mapping.get('life_col2')          # 鏂囦欢2浣跨敤瀵垮懡
        
        # #region agent log
        try:
            with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": "B",
                    "location": "exporter.py:_create_depreciation_period_sheet:entry",
                    "message": "??????????",
                    "data": {
                        "original_value_col1": original_value_col1,
                        "original_value_col2": original_value_col2,
                        "category_col1": category_col1,
                        "category_col2": category_col2,
                        "life_col1": life_col1,
                        "life_col2": life_col2,
                        "df_columns_count": len(df.columns) if df is not None else 0,
                        "df_rows": len(df) if df is not None else 0,
                        "sample_columns": list(df.columns)[:10] if df is not None else []
                    },
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + '\n')
        except Exception:
            pass
        # #endregion
        
        # 妫€鏌ュ繀闇€鐨勫瓧娈垫槸鍚﹀瓨鍦?        # 琛屽瓧娈碉細category_col1/2 + life_col1/2锛涘€煎瓧娈碉細original_value_col1/2
        if not (original_value_col1 and original_value_col2 and
                category_col1 and category_col2 and
                life_col1 and life_col2):
            # #region agent log
            try:
                with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "run1",
                        "hypothesisId": "B",
                        "location": "exporter.py:_create_depreciation_period_sheet:field_check",
                        "message": "???????",
                        "data": {
                            "has_orig_col1": bool(original_value_col1),
                            "has_cat_col1": bool(category_col1),
                            "has_life_col1": bool(life_col1),
                            "has_orig_col2": bool(original_value_col2),
                            "has_cat_col2": bool(category_col2),
                            "has_life_col2": bool(life_col2)
                        },
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + '\n')
            except Exception:
                pass
            # #endregion
            return
        
        if (original_value_col1 not in df.columns or
                category_col1 not in df.columns or
                life_col1 not in df.columns):
            # #region agent log
            try:
                with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "run1",
                        "hypothesisId": "B",
                        "location": "exporter.py:_create_depreciation_period_sheet:column_check1",
                        "message": "??1????DataFrame????",
                        "data": {
                            "original_value_col1_in_df": original_value_col1 in df.columns,
                            "category_col1_in_df": category_col1 in df.columns,
                            "life_col1_in_df": life_col1 in df.columns,
                            "all_cols_with_orig": [col for col in df.columns if '??' in str(col)][:5],
                            "all_cols_with_cat": [col for col in df.columns if '璧勪骇' in str(col) or '绫诲埆' in str(col) or '澶х被' in str(col)][:5],
                            "all_cols_with_life": [col for col in df.columns if '瀵垮懡' in str(col) or '骞撮檺' in str(col)][:5]
                        },
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + '\n')
            except Exception:
                pass
            # #endregion
            return
        if (original_value_col2 not in df.columns or
                category_col2 not in df.columns or
                life_col2 not in df.columns):
            # #region agent log
            try:
                with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({
                        "sessionId": "debug-session",
                        "runId": "run1",
                        "hypothesisId": "B",
                        "location": "exporter.py:_create_depreciation_period_sheet:column_check2",
                        "message": "??2????DataFrame????",
                        "data": {
                            "original_value_col2_in_df": original_value_col2 in df.columns,
                            "category_col2_in_df": category_col2 in df.columns,
                            "life_col2_in_df": life_col2 in df.columns
                        },
                        "timestamp": int(__import__('time').time() * 1000)
                    }, ensure_ascii=False) + '\n')
            except Exception:
                pass
            # #endregion
            return
        
        # 鍒涘缓sheet
        wb = writer.book
        ws = wb.create_sheet('折旧期间')
        
        df_work = df.copy()
        
        temp_value_col1 = '__temp_orig1__'
        temp_value_col2 = '__temp_orig2__'
        if original_value_col1 in df_work.columns:
            df_work[temp_value_col1] = df_work[original_value_col1]
        if original_value_col2 in df_work.columns:
            df_work[temp_value_col2] = df_work[original_value_col2]
        
        # ===================== 鍚堝苟琛紙缃《锛?=====================
        merged_pivot_success, merged_pivot_msg, merged_pivot_df = self.pivot_engine.create_pivot_table(
            df_work,
            index=[category_col1, category_col2, life_col1, life_col2],
            values=[temp_value_col1, temp_value_col2],
            aggfunc='sum'
        )
        
        current_row = 1
        if merged_pivot_success and merged_pivot_df is not None and not merged_pivot_df.empty:
            merged_pivot_df = merged_pivot_df.rename(
                columns={
                    temp_value_col1: original_value_col1,
                    temp_value_col2: original_value_col2
                },
                errors='ignore'
            )
            ws.cell(row=current_row, column=1, value="???????1 + ??2 ??????")
            current_row += 1
            
            for r_idx, row in enumerate(dataframe_to_rows(merged_pivot_df, index=True, header=True), current_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            merged_end_row = current_row + len(merged_pivot_df) + 1  # +1 for header
            current_row = merged_end_row + 4  # 闂撮殧3琛?        
        # ===================== 鏂囦欢1閫忚琛?=====================
        # 鍒涘缓绗竴涓暟鎹€忚琛紙鏂囦欢1锛?        # 娉ㄦ剰锛氫笉鑳藉皢鍚屼竴鍒楁棦浣滀负index鍙堜綔涓簐alues锛屽€煎瓧娈典娇鐢ㄤ复鏃跺垪锛岃瀛楁浣跨敤"璧勪骇绫诲埆 + 浣跨敤瀵垮懡"
        
        # #region agent log
        try:
            with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": "B",
                    "location": "exporter.py:_create_depreciation_period_sheet:pivot1",
                    "message": "鍒涘缓绗竴涓€忚琛ㄥ墠",
                        "data": {
                            "index_cols": [category_col1, life_col1],
                        "values_col": temp_value_col1,
                        "df_rows": len(df),
                            "cat_col1_non_null": df[category_col1].notna().sum() if category_col1 in df.columns else 0,
                        "orig_col1_non_null": df[original_value_col1].notna().sum() if original_value_col1 in df.columns else 0,
                        "life_col1_non_null": df[life_col1].notna().sum() if life_col1 in df.columns else 0
                    },
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + '\n')
        except Exception:
            pass
        # #endregion
        
        pivot1_success, pivot1_msg, pivot1_df = self.pivot_engine.create_pivot_table(
            df_work,
            index=[category_col1, life_col1],
            values=[temp_value_col1],
            aggfunc='sum'
        )
        
        # #region agent log
        try:
            with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": "B",
                    "location": "exporter.py:_create_depreciation_period_sheet:pivot1_result",
                    "message": "??????????",
                    "data": {
                        "pivot1_success": pivot1_success,
                        "pivot1_msg": pivot1_msg[:200] if pivot1_msg else None,
                        "pivot1_df_is_none": pivot1_df is None,
                        "pivot1_df_empty": pivot1_df.empty if pivot1_df is not None else None,
                        "pivot1_df_rows": len(pivot1_df) if pivot1_df is not None else 0
                    },
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + '\n')
        except Exception:
            pass
        # #endregion
        
        if pivot1_success and pivot1_df is not None:
            pivot1_df = pivot1_df.rename(columns={temp_value_col1: original_value_col1}, errors='ignore')
            ws.cell(row=current_row, column=1, value="???1?????")
            current_row += 1
            
            for r_idx, row in enumerate(dataframe_to_rows(pivot1_df, index=True, header=True), current_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 璁板綍绗竴涓〃鐨勭粨鏉熻
            pivot1_end_row = current_row + len(pivot1_df) + 1  # +1 for header
            current_row = pivot1_end_row + 4  # 闂撮殧3琛?        
        if original_value_col2 not in df_work.columns and original_value_col2:
            df_work[temp_value_col2] = df_work[original_value_col2]
        
        # #region agent log
        try:
            with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": "B",
                    "location": "exporter.py:_create_depreciation_period_sheet:pivot2",
                    "message": "鍒涘缓绗簩涓€忚琛ㄥ墠",
                        "data": {
                            "index_cols": [category_col2, life_col2],
                        "values_col": temp_value_col2,
                            "cat_col2_non_null": df[category_col2].notna().sum() if category_col2 in df.columns else 0,
                        "orig_col2_non_null": df[original_value_col2].notna().sum() if original_value_col2 in df.columns else 0,
                        "life_col2_non_null": df[life_col2].notna().sum() if life_col2 in df.columns else 0
                    },
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + '\n')
        except Exception:
            pass
        # #endregion
        
        pivot2_success, pivot2_msg, pivot2_df = self.pivot_engine.create_pivot_table(
            df_work,
            index=[category_col2, life_col2],
            values=[temp_value_col2],
            aggfunc='sum'
        )
        
        # #region agent log
        try:
            with open(r'c:\Users\Administrator\Downloads\FA\.cursor\debug.log', 'a', encoding='utf-8') as f:
                f.write(json.dumps({
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": "B",
                    "location": "exporter.py:_create_depreciation_period_sheet:pivot2_result",
                    "message": "??????????",
                    "data": {
                        "pivot2_success": pivot2_success,
                        "pivot2_msg": pivot2_msg[:200] if pivot2_msg else None,
                        "pivot2_df_is_none": pivot2_df is None,
                        "pivot2_df_empty": pivot2_df.empty if pivot2_df is not None else None,
                        "pivot2_df_rows": len(pivot2_df) if pivot2_df is not None else 0
                    },
                    "timestamp": int(__import__('time').time() * 1000)
                }, ensure_ascii=False) + '\n')
        except Exception:
            pass
        # #endregion
        
        if pivot2_success and pivot2_df is not None:
            pivot2_df = pivot2_df.rename(columns={temp_value_col2: original_value_col2}, errors='ignore')
            ws.cell(row=current_row, column=1, value="???2?????")
            current_row += 1
            
            for r_idx, row in enumerate(dataframe_to_rows(pivot2_df, index=True, header=True), current_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    def _create_depreciation_period_csv(self, base_path: str, df: pd.DataFrame, summary_config: Dict, csv_kwargs: dict, files_created: list) -> str:
        """
        鍒涘缓鎶樻棫鏈熼棿琛ㄦ牸锛圕SV鏍煎紡锛夛紝鍒涘缓涓や釜鍗曠嫭鐨凜SV鏂囦欢
        
        Returns:
            str: 瀵煎嚭娑堟伅
        """
        msg_parts = []
        
        # 鑾峰彇鐢ㄦ埛閰嶇疆鐨勫瓧娈碉紙浣跨敤宸叉牸寮忓寲鐨勫垪鍚嶏級
        original_value_col1 = summary_config.get('original_value_col1')  # 鏂囦欢1鍘熷€硷紙鍊煎瓧娈碉級
        original_value_col2 = summary_config.get('original_value_col2')  # 鏂囦欢2鍘熷€硷紙鍊煎瓧娈碉級
        field_mapping = summary_config.get('field_mapping', {})
        # 琛屽瓧娈碉細璧勪骇绫诲埆 + 浣跨敤瀵垮懡锛堝畬鍏ㄤ娇鐢ㄧ敤鎴锋槧灏勭殑瀛楁锛屼笉鍋氫换浣曡嚜鍔ㄥ尮閰嶏級
        category_col1 = field_mapping.get('category_col1')  # 鏂囦欢1璧勪骇绫诲埆
        category_col2 = field_mapping.get('category_col2')  # 鏂囦欢2璧勪骇绫诲埆
        life_col1 = field_mapping.get('life_col1')          # 鏂囦欢1浣跨敤瀵垮懡
        life_col2 = field_mapping.get('life_col2')          # 鏂囦欢2浣跨敤瀵垮懡
        
        # 妫€鏌ュ繀闇€鐨勫瓧娈垫槸鍚﹀瓨鍦?        # 琛屽瓧娈碉細category_col1/2 + life_col1/2锛涘€煎瓧娈碉細original_value_col1/2
        if not (original_value_col1 and original_value_col2 and
                category_col1 and category_col2 and
                life_col1 and life_col2):
            return ""
        
        if (original_value_col1 not in df.columns or
                category_col1 not in df.columns or
                life_col1 not in df.columns):
            return ""
        if (original_value_col2 not in df.columns or
                category_col2 not in df.columns or
                life_col2 not in df.columns):
            return ""
        
        df_work = df.copy()
        
        temp_value_col1 = '__temp_orig1__'
        if original_value_col1 in df_work.columns:
            df_work[temp_value_col1] = df_work[original_value_col1]
        
        pivot1_success, pivot1_msg, pivot1_df = self.pivot_engine.create_pivot_table(
            df_work,
            index=[category_col1, life_col1],
            values=[temp_value_col1],
            aggfunc='sum'
        )
        
        if pivot1_success and pivot1_df is not None:
            pivot1_df = pivot1_df.rename(columns={temp_value_col1: original_value_col1}, errors='ignore')
            pivot1_path = f"{base_path}_鎶樻棫鏈熼棿_鏂囦欢1.csv"
            pivot1_df.to_csv(pivot1_path, **csv_kwargs)
            files_created.append(pivot1_path)
            msg_parts.append(f"\n鎶樻棫鏈熼棿锛堟枃浠?锛夊鍑哄埌: {pivot1_path}")
        
        temp_value_col2 = '__temp_orig2__'
        if original_value_col2 in df_work.columns:
            df_work[temp_value_col2] = df_work[original_value_col2]
        
        pivot2_success, pivot2_msg, pivot2_df = self.pivot_engine.create_pivot_table(
            df_work,
            index=[category_col2, life_col2],
            values=[temp_value_col2],
            aggfunc='sum'
        )
        
        if pivot2_success and pivot2_df is not None:
            pivot2_df = pivot2_df.rename(columns={temp_value_col2: original_value_col2}, errors='ignore')
            pivot2_path = f"{base_path}_鎶樻棫鏈熼棿_鏂囦欢2.csv"
            pivot2_df.to_csv(pivot2_path, **csv_kwargs)
            files_created.append(pivot2_path)
            msg_parts.append(f"\n鎶樻棫鏈熼棿锛堟枃浠?锛夊鍑哄埌: {pivot2_path}")
        
        return "".join(msg_parts)

    def _prepare_pivot_export_df(self, pivot_df: pd.DataFrame) -> pd.DataFrame:
        """Prepare pivot output per business rules before writing to Excel."""
        out = pivot_df.copy()
        out = self._make_unique_columns(out)
        if isinstance(out, pd.DataFrame) and not isinstance(out.index, pd.RangeIndex):
            out = out.reset_index()
        if out.shape[1] < 2:
            return out

        # Flatten any complex column labels for stable downstream handling.
        flat_cols = []
        for c in out.columns:
            if isinstance(c, tuple):
                flat_cols.append("_".join([str(x) for x in c if str(x) != ""]).strip("_"))
            else:
                flat_cols.append(str(c))
        out.columns = flat_cols

        # Drop synthetic index columns introduced by reset_index on a plain RangeIndex-like source.
        for idx_like in ("index", "level_0"):
            if idx_like in out.columns:
                idx_num = pd.to_numeric(out[idx_like], errors="coerce")
                if idx_num.notna().all():
                    out = out.drop(columns=[idx_like])
        if out.shape[1] < 2:
            return out

        # Business rule: swap first two visible columns, then process by final A/B semantics.
        swapped_cols = [out.columns[1], out.columns[0]] + list(out.columns[2:])
        out = out[swapped_cols]
        # Business rule: swap E/F columns (1-based), i.e. index 4 and 5 in 0-based.
        if out.shape[1] >= 6:
            ef_cols = list(out.columns)
            ef_cols[4], ef_cols[5] = ef_cols[5], ef_cols[4]
            out = out[ef_cols]
        col_a, col_b = out.columns[0], out.columns[1]
        out[col_a] = out[col_a].replace("", pd.NA).ffill()

        def _norm(v):
            return re.sub(r"\s+", "", str(v)).strip()

        def _is_unc(v):
            return _norm(v) == "未分类"

        def _is_total_or_subtotal(v):
            s = _norm(v)
            return ("合计" in s) or ("小计" in s)

        # Remove rows where both A and B are 未分类.
        both_unc = out[col_a].apply(_is_unc) & out[col_b].apply(_is_unc)
        out = out.loc[~both_unc].copy()

        # First-pass remove rows where both A and B are 合计/小计.
        both_total_or_sub = out[col_a].apply(_is_total_or_subtotal) & out[col_b].apply(_is_total_or_subtotal)
        out = out.loc[~both_total_or_sub].copy()

        # If B is 未分类 and A is not, aggregate under A.
        b_unc = out[col_b].apply(_is_unc)
        if bool(b_unc.any()):
            out[col_b] = out[col_b].astype(object)
            out.loc[b_unc, col_b] = out.loc[b_unc, col_a]

        # If A is 未分类 and B is not, aggregate under B.
        a_unc_b_not_unc = out[col_a].apply(_is_unc) & (~out[col_b].apply(_is_unc))
        if bool(a_unc_b_not_unc.any()):
            out[col_a] = out[col_a].astype(object)
            out.loc[a_unc_b_not_unc, col_a] = out.loc[a_unc_b_not_unc, col_b]

        # Second-pass remove after B was replaced by A (covers "A=合计, B=未分类" -> "A/B=合计").
        both_total_or_sub = out[col_a].apply(_is_total_or_subtotal) & out[col_b].apply(_is_total_or_subtotal)
        out = out.loc[~both_total_or_sub].copy()

        # Normalize numeric columns and aggregate duplicate (A,B) rows.
        numeric_cols = []
        for c in out.columns:
            if c in (col_a, col_b):
                continue
            converted = pd.to_numeric(out[c], errors="coerce")
            if converted.notna().sum() > 0:
                out[c] = converted.fillna(0)
                numeric_cols.append(c)

        agg_map = {c: "sum" for c in numeric_cols}
        for c in out.columns:
            if c in (col_a, col_b):
                continue
            if c not in agg_map:
                agg_map[c] = "first"
        out = out.groupby([col_a, col_b], as_index=False, sort=False, dropna=False).agg(agg_map)

        # Sort by B column descending (except total row which is appended later).
        out = out.sort_values(
            by=col_b,
            ascending=False,
            na_position="last",
            kind="mergesort",
            key=lambda s: s.astype(str),
        ).reset_index(drop=True)

        # Append total row.
        total_row = {col_a: "total", col_b: ""}
        for c in out.columns:
            if c in (col_a, col_b):
                continue
            total_row[c] = ""
        out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
        return out

    def _rebuild_pivot_from_export_config(
        self,
        df: Optional[pd.DataFrame],
        summary_config: Optional[Dict],
    ) -> Optional[pd.DataFrame]:
        """按导出时透视配置重算，确保与当前导出口径一致。"""
        if df is None or df.empty:
            return None
        pc = (summary_config or {}).get("pivot_export_config") or {}
        if not isinstance(pc, dict):
            return None

        index = [c for c in (pc.get("index") or []) if c in df.columns]
        if not index:
            return None
        columns = [c for c in (pc.get("columns") or []) if c in df.columns]
        values = [c for c in (pc.get("values") or []) if c in df.columns]
        aggfunc = str(pc.get("aggfunc") or "sum")

        ok, _msg, pv = self.pivot_engine.create_pivot_table(
            df,
            index=index,
            columns=(columns if columns else None),
            values=(values if values else None),
            aggfunc=aggfunc,
        )
        if ok and pv is not None and not pv.empty:
            return pv
        return None

    def _reserve_sheet_name(self, preferred: str, used: set) -> str:
        safe = re.sub(r'[\\/*?:\\[\\]]', '_', str(preferred or 'Sheet')).strip()
        if not safe:
            safe = "Sheet"
        safe = safe[:31]
        candidate = safe
        i = 1
        while candidate in used:
            suffix = f"_{i}"
            candidate = f"{safe[:31 - len(suffix)]}{suffix}"
            i += 1
        used.add(candidate)
        return candidate

    def _build_xlsx_formats(self, workbook):
        return {
            "header_left": workbook.add_format({"bold": True, "align": "left", "valign": "vcenter"}),
            "header_left_file2": workbook.add_format({"bold": True, "align": "left", "valign": "vcenter", "bg_color": "#E6E6E6"}),
            "text_left": workbook.add_format({"align": "left", "valign": "vcenter"}),
            "text_left_file2": workbook.add_format({"align": "left", "valign": "vcenter", "bg_color": "#E6E6E6"}),
            "num_right": workbook.add_format({"align": "right", "valign": "vcenter", "num_format": "#,##0"}),
            "num_right_file2": workbook.add_format({"align": "right", "valign": "vcenter", "num_format": "#,##0", "bg_color": "#E6E6E6"}),
            "pct_right": workbook.add_format({"align": "right", "valign": "vcenter", "num_format": "0.00%"}),
            "pct_right_file2": workbook.add_format({"align": "right", "valign": "vcenter", "num_format": "0.00%", "bg_color": "#E6E6E6"}),
            "border_only": workbook.add_format({"border": 1, "border_color": "#000000"}),
            "group_gray": workbook.add_format({"bg_color": "#E6E6E6"}),
            "summary_light_blue": workbook.add_format({"bold": True, "align": "left", "valign": "vcenter", "bg_color": "#DDEBF7"}),
        }

    def _get_force_numeric_headers(self, summary_config: Optional[Dict] = None) -> set:
        headers = {"原值", "累计折旧", "净值", "原值增加", "处置原值", "处置折旧", "本年折旧", "原值减少", "年初累计折旧"}
        if summary_config:
            for k in ("original_value_col1", "original_value_col2", "depreciation_col1", "depreciation_col2"):
                v = summary_config.get(k)
                if v:
                    headers.add(v)
            fm = summary_config.get("field_mapping", {}) or {}
            for k in ("disposal_orig_col1", "disposal_orig_col2", "disposal_dep_col1", "disposal_dep_col2"):
                v = fm.get(k)
                if v:
                    headers.add(v)
        return headers

    def _sheet_numeric_header_hints(self, sheet_name: str) -> set:
        hints = {
            "FA List": {"原值", "累计折旧", "净值", "本年折旧"},
            "新增清单_BKD": {"原值增加"},
            "处置清单_BKD": {"原值减少", "年初累计折旧", "本年折旧", "净值", "处置原值", "处置折旧"},
        }
        return hints.get(sheet_name, set())

    def _get_percent_headers(self, sheet_name: str) -> set:
        if sheet_name in ("FA List", "新增清单_BKD", "处置清单_BKD"):
            return {"残值率"}
        return set()

    def _get_width_limits(self, sheet_name: str):
        # 明细与透视页保持更紧凑，避免“看起来过宽”
        if sheet_name in ("合并数据", "数据透视表", "折旧期间"):
            return 8, 26
        return 8, 45

    def _detect_residual_header_modes(self, df: pd.DataFrame, sheet_name: str, summary_config: Optional[Dict] = None) -> Dict[str, str]:
        """
        主数据“残值率”列展示规则：
        - 若检测到该列本质为残值（存在数值 > 100），按千分位整数显示（number）
        - 否则按百分比显示（percent）
        """
        modes = {}
        if sheet_name != "合并数据" or df is None or df.empty:
            return modes

        candidates = []
        fm = (summary_config or {}).get("field_mapping", {}) or {}
        for k in ("residual_col1", "residual_col2"):
            c = fm.get(k)
            if c and c in df.columns:
                candidates.append(c)
        for c in df.columns:
            if "残值率" in str(c) and c not in candidates:
                candidates.append(c)

        for c in candidates:
            s = pd.to_numeric(df[c].map(self._to_number), errors="coerce").dropna()
            if s.empty:
                continue
            modes[c] = "number" if float(s.max()) > 100 else "percent"
        return modes

    def _coerce_sheet_df_for_write(self, df: pd.DataFrame, sheet_name: str, summary_config: Optional[Dict] = None) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        out = df.copy()
        out = self._make_unique_columns(out)
        if out.empty:
            return out

        if sheet_name == "处置清单_BKD":
            out = out.rename(columns={"原值": "原值减少", "累计折旧": "年初累计折旧"})
            if "入账开始日期" in out.columns:
                out["入账开始日期"] = out["入账开始日期"].map(self.sheet_generator._format_date_only)

        force_numeric_headers = self._get_force_numeric_headers(summary_config) | self._sheet_numeric_header_hints(sheet_name)
        percent_headers = self._get_percent_headers(sheet_name)
        residual_modes = self._detect_residual_header_modes(out, sheet_name, summary_config)
        is_dep_sheet_like = (sheet_name == "折旧期间") or ({"判断结果", "影响当年金额"}.issubset(set(out.columns)))

        for idx, col in enumerate(out.columns):
            header = str(col)
            if header in residual_modes:
                if residual_modes[header] == "percent":
                    series = out[col].map(self._to_number)
                    series = series.apply(lambda x: x / 100.0 if x is not None and pd.notna(x) and abs(x) > 1 else x)
                    out[col] = pd.to_numeric(series, errors="coerce")
                else:
                    out[col] = pd.to_numeric(out[col].map(self._to_number), errors="coerce")
            elif header in percent_headers:
                series = out[col].map(self._to_number)
                series = series.apply(lambda x: x / 100.0 if x is not None and pd.notna(x) and abs(x) > 1 else x)
                out[col] = pd.to_numeric(series, errors="coerce")
            elif is_dep_sheet_like and ("残值率" in header or idx in (4, 5)):
                series = out[col].map(self._to_number)
                series = series.apply(lambda x: x / 100.0 if x is not None and pd.notna(x) and abs(x) > 1 else x)
                out[col] = pd.to_numeric(series, errors="coerce")
            elif header in force_numeric_headers:
                out[col] = pd.to_numeric(out[col].map(self._to_number), errors="coerce")

        if sheet_name == "数据透视表" and out.shape[1] > 2:
            for col in out.columns[2:]:
                out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
        return out

    def _estimate_column_width(
        self,
        series: pd.Series,
        header: str,
        is_numeric: bool = False,
        is_percent: bool = False,
        min_width: int = 8,
        max_width: int = 45,
    ) -> int:
        max_len = 0
        sample = series.head(10)
        for v in sample:
            if pd.isna(v):
                continue
            if is_percent and isinstance(v, (int, float)):
                text = f"{v:.2%}"
            elif is_numeric and isinstance(v, (int, float)):
                text = f"{v:,.0f}"
            else:
                text = str(v)
            if len(text) > max_len:
                max_len = len(text)
        return min(max(max_len + 2, min_width), max_width)

    def _xlsx_col_letter(self, col_idx: int) -> str:
        letters = ""
        n = int(col_idx)
        while n >= 0:
            n, r = divmod(n, 26)
            letters = chr(65 + r) + letters
            n -= 1
        return letters

    def _depreciation_bs_date_literal(self, summary_config: Optional[Dict] = None) -> str:
        """Return the balance sheet date used by embedded depreciation formulas."""
        for key in ("balance_sheet_date", "bs_date", "depreciation_bs_date"):
            value = (summary_config or {}).get(key)
            if value:
                text = str(value).strip()
                if text:
                    return text.replace("/", "-").replace(".", "-")
        return f"{date.today().year}-12-31"

    def _depreciation_formula_rows_to_write(self, sheet_name: str, row_count: int) -> int:
        if row_count <= int(DEPRECIATION_FORMULA_ROW_LIMIT):
            return row_count
        formula_rows = min(row_count, int(DEPRECIATION_FORMULA_SAMPLE_ROWS))
        note = (
            f"【导出提速】{sheet_name} 共{row_count}行，超过{DEPRECIATION_FORMULA_ROW_LIMIT}行，"
            f"折旧测算公式仅写入前{formula_rows}行；如需全量计算，请在Excel中向下填充公式。"
        )
        if note not in self._export_notes:
            self._export_notes.append(note)
        return formula_rows

    def _append_depreciation_formula_block(
        self,
        ws,
        sheet_name: str,
        df: pd.DataFrame,
        fmts: Dict,
        summary_config: Optional[Dict] = None,
    ) -> int:
        """Append depreciation testing formulas to FA List / disposal sheets."""
        if df is None or df.empty or sheet_name not in ("FA List", "处置清单_BKD"):
            return 0

        headers = [str(c) for c in df.columns]
        if sheet_name == "FA List":
            source = {
                "start_date": "入账开始日期",
                "life": "使用寿命(月)",
                "residual": "残值率",
                "original": "原值",
                "book_accumulated": "累计折旧",
                "book_current": "本年折旧",
                "cutoff": None,
            }
            separator_idx = 12  # M列空白分隔
        else:
            source = {
                "start_date": "入账开始日期",
                "life": "使用寿命(月)",
                "residual": "残值率",
                "original": "原值减少",
                "book_accumulated": "处置折旧",
                "book_current": "本年折旧",
                "cutoff": "处置时间",
            }
            separator_idx = 15  # P列空白分隔

        required = [v for k, v in source.items() if k != "cutoff" and v]
        missing = [header for header in required if header not in headers]
        if missing:
            return 0

        separator_idx = max(len(headers), separator_idx)
        start_idx = separator_idx + 1
        result_headers = [
            "月折旧额",
            "本年应计提折旧月份",
            "累计折旧月份",
            "测算的当年折旧",
            "测算的累计折旧",
            "账面本年折旧",
            "差异_本年折旧",
            "差异_累计折旧",
        ]

        ws.set_column(separator_idx, separator_idx, 3, fmts["text_left"])
        for offset, header in enumerate(result_headers):
            ws.write(0, start_idx + offset, header, fmts["header_left"])

        start_date_col = self._xlsx_col_letter(headers.index(source["start_date"]))
        life_col = self._xlsx_col_letter(headers.index(source["life"]))
        residual_col = self._xlsx_col_letter(headers.index(source["residual"]))
        original_col = self._xlsx_col_letter(headers.index(source["original"]))
        accumulated_col = self._xlsx_col_letter(headers.index(source["book_accumulated"]))
        current_source_col = self._xlsx_col_letter(headers.index(source["book_current"]))
        cutoff_name = source.get("cutoff")
        cutoff_col = self._xlsx_col_letter(headers.index(cutoff_name)) if cutoff_name in headers else None

        monthly_col = self._xlsx_col_letter(start_idx)
        current_months_col = self._xlsx_col_letter(start_idx + 1)
        accumulated_months_col = self._xlsx_col_letter(start_idx + 2)
        estimated_current_col = self._xlsx_col_letter(start_idx + 3)
        estimated_accumulated_col = self._xlsx_col_letter(start_idx + 4)
        book_current_col = self._xlsx_col_letter(start_idx + 5)
        diff_current_col = self._xlsx_col_letter(start_idx + 6)
        diff_accumulated_col = self._xlsx_col_letter(start_idx + 7)
        bs_date_literal = self._depreciation_bs_date_literal(summary_config)

        def _row_formula(row_idx: int, kind: str) -> str:
            row_ref = str(row_idx)
            d = f"{start_date_col}{row_ref}"
            life = f"{life_col}{row_ref}"
            residual = f"{residual_col}{row_ref}"
            original = f"{original_col}{row_ref}"
            accumulated = f"{accumulated_col}{row_ref}"
            rate_expr = f'IF({residual}="",0,IF({residual}>1,{residual}/100,{residual}))'
            dep_start_expr = f"EDATE(DATE(YEAR({d}),MONTH({d}),1),1)"
            dep_end_expr = f"EDATE({dep_start_expr},{life}-1)"
            bs_expr = f'DATEVALUE("{bs_date_literal}")'
            bs_month_expr = f"DATE(YEAR({bs_expr}),MONTH({bs_expr}),1)"
            if cutoff_col:
                cutoff = f"{cutoff_col}{row_ref}"
                effective_month_expr = (
                    f'IF(OR({cutoff}="",ISBLANK({cutoff})),{bs_month_expr},'
                    f'MIN({bs_month_expr},DATE(YEAR({cutoff}),MONTH({cutoff}),1)))'
                )
            else:
                effective_month_expr = bs_month_expr
            year_start_expr = f"DATE(YEAR({bs_expr}),1,1)"
            month_diff_current = (
                f"(YEAR(MIN({dep_end_expr},{effective_month_expr}))-YEAR(MAX({dep_start_expr},{year_start_expr})))*12+"
                f"MONTH(MIN({dep_end_expr},{effective_month_expr}))-MONTH(MAX({dep_start_expr},{year_start_expr}))+1"
            )
            month_diff_acc = (
                f"(YEAR(MIN({dep_end_expr},{effective_month_expr}))-YEAR({dep_start_expr}))*12+"
                f"MONTH(MIN({dep_end_expr},{effective_month_expr}))-MONTH({dep_start_expr})+1"
            )
            if kind == "monthly":
                return f'=IFERROR(ROUND({original}*(1-{rate_expr})/{life},2),"")'
            if kind == "current_months":
                return f'=IF(OR({d}="",{life}<=0),0,MAX(0,IF(MIN({dep_end_expr},{effective_month_expr})<MAX({dep_start_expr},{year_start_expr}),0,{month_diff_current})))'
            if kind == "accumulated_months":
                return f'=IF(OR({d}="",{life}<=0),0,MAX(0,IF(MIN({dep_end_expr},{effective_month_expr})<{dep_start_expr},0,{month_diff_acc})))'
            if kind == "estimated_current":
                return f'=IF(OR(LEN({monthly_col}{row_ref})=0,LEN({current_months_col}{row_ref})=0),"",ROUND({monthly_col}{row_ref}*{current_months_col}{row_ref},2))'
            if kind == "estimated_accumulated":
                return f'=IF(OR(LEN({monthly_col}{row_ref})=0,LEN({accumulated_months_col}{row_ref})=0),"",ROUND({monthly_col}{row_ref}*{accumulated_months_col}{row_ref},2))'
            if kind == "book_current":
                return f"={current_source_col}{row_ref}"
            if kind == "diff_current":
                return f"={book_current_col}{row_ref}-{estimated_current_col}{row_ref}"
            if kind == "diff_accumulated":
                return f"={accumulated}-{estimated_accumulated_col}{row_ref}"
            return '=""'

        formula_kinds = [
            "monthly",
            "current_months",
            "accumulated_months",
            "estimated_current",
            "estimated_accumulated",
            "book_current",
            "diff_current",
            "diff_accumulated",
        ]
        row_count = int(len(df))
        formula_rows = self._depreciation_formula_rows_to_write(sheet_name, row_count)
        for row_idx in range(1, formula_rows + 1):
            excel_row = row_idx + 1
            for offset, kind in enumerate(formula_kinds):
                fmt = fmts["num_right"] if offset not in (1, 2) else fmts["text_left"]
                ws.write_formula(row_idx, start_idx + offset, _row_formula(excel_row, kind), fmt)

        for offset in (0, 3, 4, 5, 6, 7):
            ws.set_column(start_idx + offset, start_idx + offset, 14, fmts["num_right"])
        for offset in (1, 2):
            ws.set_column(start_idx + offset, start_idx + offset, 16, fmts["text_left"])
        return start_idx + len(result_headers)

    def _is_file2_column_header(self, header: str, summary_config: Optional[Dict]) -> bool:
        h = str(header or "")
        if "_文件2" in h:
            return True
        file2_display_name = (summary_config or {}).get("file2_display_name")
        if file2_display_name and f"_{file2_display_name}" in h:
            return True
        return False

    def _format_sheet_xlsxwriter(
        self,
        writer,
        sheet_name: str,
        df: pd.DataFrame,
        fmts: Dict,
        summary_config: Optional[Dict] = None,
        merge_total_ab: bool = False,
    ):
        ws = writer.sheets[sheet_name]
        row_count = int(len(df))
        col_count = int(len(df.columns))
        if col_count <= 0:
            return

        ws.set_row(0, None, fmts["header_left"])
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, max(row_count, 0), col_count - 1)

        force_numeric_headers = self._get_force_numeric_headers(summary_config) | self._sheet_numeric_header_hints(sheet_name)
        percent_headers = self._get_percent_headers(sheet_name)
        residual_modes = self._detect_residual_header_modes(df, sheet_name, summary_config)
        is_dep_sheet_like = (sheet_name == "折旧期间") or ({"判断结果", "影响当年金额"}.issubset(set(df.columns)))
        min_w, max_w = self._get_width_limits(sheet_name)
        numeric_cols = set()
        percent_cols = set()

        for col_idx, col in enumerate(df.columns):
            header = str(col)
            series = df.iloc[:, col_idx]
            is_file2_col = sheet_name == "合并数据" and self._is_file2_column_header(header, summary_config)
            if header in residual_modes:
                if residual_modes[header] == "percent":
                    percent_cols.add(col_idx)
                else:
                    numeric_cols.add(col_idx)
            elif header in percent_headers:
                percent_cols.add(col_idx)
            elif is_dep_sheet_like and ("残值率" in header or col_idx in (4, 5)):
                percent_cols.add(col_idx)
            elif header in force_numeric_headers:
                numeric_cols.add(col_idx)
            elif sheet_name == "数据透视表" and col_idx >= 2:
                numeric_cols.add(col_idx)
            elif pd.api.types.is_numeric_dtype(series) and not pd.api.types.is_bool_dtype(series):
                numeric_cols.add(col_idx)

            width = self._estimate_column_width(
                series,
                header,
                is_numeric=(col_idx in numeric_cols),
                is_percent=(col_idx in percent_cols),
                min_width=min_w,
                max_width=max_w,
            )
            if col_idx in percent_cols:
                ws.set_column(col_idx, col_idx, width, fmts["pct_right_file2"] if is_file2_col else fmts["pct_right"])
            elif col_idx in numeric_cols:
                ws.set_column(col_idx, col_idx, width, fmts["num_right_file2"] if is_file2_col else fmts["num_right"])
            else:
                ws.set_column(col_idx, col_idx, width, fmts["text_left_file2"] if is_file2_col else fmts["text_left"])

        # 强制标题行左对齐（覆盖 pandas 默认 header 样式）
        for col_idx, col in enumerate(df.columns):
            header = str(col)
            if sheet_name == "折旧期间" and header.strip() == "计算过程":
                header = "计算过程=年末原值*(1-年末残值率)/年末寿命-年末原值*(1-年初残值率)/年初寿命"
            is_file2_col = sheet_name == "合并数据" and self._is_file2_column_header(header, summary_config)
            ws.write(0, col_idx, header, fmts["header_left_file2"] if is_file2_col else fmts["header_left"])

        if sheet_name == "处置清单_BKD" and row_count > 0:
            headers = [str(c) for c in df.columns]
            required = ("年初累计折旧", "本年折旧", "处置折旧")
            if all(h in headers for h in required):
                opening_dep_col = headers.index("年初累计折旧")
                current_dep_col = headers.index("本年折旧")
                disposal_dep_col = headers.index("处置折旧")
                opening_dep_letter = self._xlsx_col_letter(opening_dep_col)
                disposal_dep_letter = self._xlsx_col_letter(disposal_dep_col)
                formula_rows = self._depreciation_formula_rows_to_write(sheet_name, row_count)
                for row_idx in range(1, formula_rows + 1):
                    excel_row = row_idx + 1
                    formula = f'=IFERROR({disposal_dep_letter}{excel_row}-{opening_dep_letter}{excel_row},"")'
                    cached_value = df.iloc[row_idx - 1, current_dep_col]
                    if pd.isna(cached_value):
                        cached_value = ""
                    ws.write_formula(row_idx, current_dep_col, formula, fmts["num_right"], cached_value)

        formula_col_count = self._append_depreciation_formula_block(
            ws, sheet_name, df, fmts, summary_config=summary_config
        )
        if formula_col_count:
            col_count = max(col_count, formula_col_count)
            ws.autofilter(0, 0, max(row_count, 0), col_count - 1)

        if merge_total_ab and row_count > 0 and col_count >= 2:
            if str(df.iloc[-1, 0]).strip().lower() == "total":
                row_idx = row_count
                # total 行用公式求和，避免写死数值
                for col_idx in range(2, col_count):
                    col_letter = self._xlsx_col_letter(col_idx)
                    formula = f"=SUM({col_letter}2:{col_letter}{row_count})"
                    ws.write_formula(row_idx, col_idx, formula, fmts["num_right"])
                ws.merge_range(row_idx, 0, row_idx, 1, "total", fmts["text_left"])

        ws.conditional_format(
            0, 0, max(row_count, 0), col_count - 1,
            {"type": "no_errors", "format": fmts["border_only"]}
        )

        if sheet_name == "数据透视表" and row_count > 0:
            # 按B列分组交替灰白底色；若B=未分类则回退按A分组
            ws.conditional_format(
                1, 0, row_count, col_count - 1,
                {
                    "type": "formula",
                    "criteria": '=MOD(SUMPRODUCT(--(IF($B$2:$B2="未分类",$A$2:$A2,$B$2:$B2)<>IF($B$1:$B1="未分类",$A$1:$A1,$B$1:$B1)))-1,2)=1',
                    "format": fmts["group_gray"],
                },
            )
        elif sheet_name == "折旧期间" and row_count > 0:
            ws.conditional_format(
                1, 0, row_count, col_count - 1,
                {
                    "type": "formula",
                    "criteria": "=MOD(SUMPRODUCT(--($A$2:$A2<>$A$1:$A1))-1,2)=1",
                    "format": fmts["group_gray"],
                },
            )

    def _write_summary_sheet_xlsxwriter(self, writer, sheet_name: str, fmts: Dict) -> bool:
        summary_data = getattr(self.summary_generator, "summary_data", None)
        if not summary_data:
            return False

        def _safe_num(v) -> float:
            """将汇总值安全转换为可写入 xlsxwriter 的数字。"""
            try:
                if v is None or pd.isna(v):
                    return 0.0
            except Exception:
                pass
            try:
                return float(v)
            except Exception:
                return 0.0

        ws = writer.book.add_worksheet(sheet_name)
        writer.sheets[sheet_name] = ws

        categories = summary_data.get("categories", [])
        file2_only_categories = set(summary_data.get("file2_only_categories", []))
        data = summary_data.get("data", {})
        row_defs = summary_data.get("row_defs", [])
        ws.write(0, 0, "", fmts["header_left"])
        ws.write(0, 1, "", fmts["header_left"])
        for i, cat in enumerate(categories):
            cat_fmt = fmts["summary_light_blue"] if cat in file2_only_categories else fmts["header_left"]
            ws.write(0, i + 2, cat, cat_fmt)
            ws.write(1, i + 2, "账面数", fmts["header_left"])

        start_row = 2
        for idx, row_def in enumerate(row_defs):
            row = start_row + idx
            section = row_def.get("section", "")
            item = row_def.get("item", "")
            ws.write(row, 0, section, fmts["text_left"])
            ws.write(row, 1, item, fmts["text_left"])
            for i, cat in enumerate(categories):
                col = i + 2
                cat_data = data.get(cat, {})
                if row_def.get("kind") == "formula":
                    value = self.summary_generator._calc_formula_value(cat_data, row_def.get("key"))
                else:
                    value = cat_data.get(row_def.get("key"), 0.0)
                # 导出稳健性：统一转为数字，避免 pd.NA/NaN/空值导致 xlsxwriter 报错
                ws.write_number(row, col, _safe_num(value), fmts["num_right"])

        section_start = None
        section_text = None
        for idx, row_def in enumerate(row_defs):
            row = start_row + idx
            section = row_def.get("section", "")
            if section:
                if section_start is not None and section_text is not None:
                    prev = row - 1
                    if prev > section_start:
                        ws.merge_range(section_start, 0, prev, 0, section_text, fmts["text_left"])
                section_start = row
                section_text = section
        if section_start is not None and section_text is not None:
            end_row = start_row + len(row_defs) - 1
            if end_row > section_start:
                ws.merge_range(section_start, 0, end_row, 0, section_text, fmts["text_left"])

        ws.set_column(0, 0, 14, fmts["text_left"])
        ws.set_column(1, 1, 30, fmts["text_left"])
        for i in range(len(categories)):
            ws.set_column(i + 2, i + 2, 16, fmts["num_right"])

        ws.freeze_panes(2, 2)
        max_row = max(start_row + len(row_defs) - 1, 1)
        max_col = max(len(categories) + 1, 1)
        ws.conditional_format(
            0, 0, max_row, max_col,
            {"type": "no_errors", "format": fmts["border_only"]}
        )
        return True

    def _build_depreciation_period_df(
        self,
        df: pd.DataFrame,
        summary_config: Optional[Dict],
        fa_df: Optional[pd.DataFrame] = None,
    ) -> Optional[pd.DataFrame]:
        if df is None or df.empty or not summary_config:
            return None

        original_value_col1 = summary_config.get('original_value_col1')
        original_value_col2 = summary_config.get('original_value_col2')
        field_mapping = summary_config.get('field_mapping', {}) or {}
        category_col1 = field_mapping.get('category_col1')
        category_col2 = field_mapping.get('category_col2')
        life_col1 = field_mapping.get('life_col1')
        life_col2 = field_mapping.get('life_col2')
        residual_col1 = field_mapping.get('residual_col1')
        residual_col2 = field_mapping.get('residual_col2')

        required = [original_value_col1, original_value_col2, category_col1, category_col2, life_col1, life_col2]
        if not all(required):
            return None
        if any(col not in df.columns for col in required):
            return None

        def _correct_life_series(src: pd.Series) -> pd.Series:
            num = pd.to_numeric(src, errors='coerce')
            valid = num.dropna()
            if not valid.empty and float(valid.max()) < 30:
                return num * 12
            return num

        df_work = df.copy()
        temp_cat1_col = "__temp_cat1__"
        temp_cat2_col = "__temp_cat2__"
        temp_value_col1 = '__temp_orig1__'
        temp_value_col2 = '__temp_orig2__'
        temp_life_col1 = '__temp_life1__'
        temp_life_col2 = '__temp_life2__'
        temp_res1_col = '__temp_res1__'
        temp_res2_col = '__temp_res2__'
        temp_res_amt1_col = '__temp_res_amt1__'
        temp_res_amt2_col = '__temp_res_amt2__'

        def _norm_text(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return str(v).strip()

        def _is_unc(v) -> bool:
            sv = _norm_text(v).replace(" ", "").upper()
            return sv in ("", "未分类", "N/A", "#N/A", "#N/AN/A", "<NA>", "NA", "NAN", "NONE", "NULL", "-", "--")

        df_work[temp_cat1_col] = df_work[category_col1].map(_norm_text)
        df_work[temp_cat2_col] = df_work[category_col2].map(_norm_text)
        df_work[temp_value_col1] = pd.to_numeric(df_work[original_value_col1].map(self._to_number), errors='coerce').fillna(0)
        df_work[temp_value_col2] = pd.to_numeric(df_work[original_value_col2].map(self._to_number), errors='coerce').fillna(0)
        df_work[temp_life_col1] = _correct_life_series(df_work[life_col1])
        df_work[temp_life_col2] = _correct_life_series(df_work[life_col2])

        # B未分类/A非未分类 -> 按 A+C 查找已有聚合对象并归并（无对象则保留，后续可见N/A）
        # A未分类/B非未分类 -> 仅当“B分类在年初已存在”时，才将A回填为B并做同寿命聚合；
        # 否则A保持未分类。
        mask_b_unc = df_work[temp_cat2_col].map(_is_unc) & (~df_work[temp_cat1_col].map(_is_unc))
        mask_a_unc = df_work[temp_cat1_col].map(_is_unc) & (~df_work[temp_cat2_col].map(_is_unc))
        start_cat_set = set(
            df_work.loc[
                (~df_work[temp_cat1_col].map(_is_unc)) & (df_work[temp_value_col1] != 0),
                temp_cat1_col,
            ].astype(str).tolist()
        )
        mask_a_unc_fill = mask_a_unc & df_work[temp_cat2_col].astype(str).isin(start_cat_set)
        if bool(mask_a_unc_fill.any()):
            df_work.loc[mask_a_unc_fill, temp_cat1_col] = df_work.loc[mask_a_unc_fill, temp_cat2_col]

        # 先用“B非未分类”的行构建(A, C)->(B, D)映射；B缺失行仅在有映射时才并入
        b_map = {}
        src_mask = (~df_work[temp_cat2_col].map(_is_unc)) & (~df_work[temp_cat1_col].map(_is_unc))
        for _, r in df_work.loc[src_mask, [temp_cat1_col, temp_life_col1, temp_cat2_col, temp_life_col2]].iterrows():
            k = (r[temp_cat1_col], r[temp_life_col1])
            if k not in b_map and (not _is_unc(r[temp_cat2_col])):
                b_map[k] = (r[temp_cat2_col], r[temp_life_col2])

        if bool(mask_b_unc.any()):
            idxs = df_work.index[mask_b_unc].tolist()
            for idx in idxs:
                k = (df_work.at[idx, temp_cat1_col], df_work.at[idx, temp_life_col1])
                tgt = b_map.get(k)
                if tgt is not None:
                    df_work.at[idx, temp_cat2_col] = tgt[0]
                    df_work.at[idx, temp_life_col2] = tgt[1]

        # 折旧期间聚合口径：仅对“允许回填A”的行，同步寿命到D，便于同类型同寿命聚合
        if bool(mask_a_unc_fill.any()):
            df_work.loc[mask_a_unc_fill, temp_life_col1] = df_work.loc[mask_a_unc_fill, temp_life_col2]

        # C/D 互相回填
        l1_nan_l2_ok = df_work[temp_life_col1].isna() & df_work[temp_life_col2].notna()
        l2_nan_l1_ok = df_work[temp_life_col2].isna() & df_work[temp_life_col1].notna()
        if bool(l1_nan_l2_ok.any()):
            df_work.loc[l1_nan_l2_ok, temp_life_col1] = df_work.loc[l1_nan_l2_ok, temp_life_col2]
        if bool(l2_nan_l1_ok.any()):
            df_work.loc[l2_nan_l1_ok, temp_life_col2] = df_work.loc[l2_nan_l1_ok, temp_life_col1]

        # 残值率口径调整为：净残值合计 / 原值合计
        # 其中净残值可来自：
        # - 每单项资产 原值 * 净残值率
        # - 或每单项资产“净残值”直接相加（触发纠偏时按净残值字段处理）
        def _to_rate_amount_series(raw_series: pd.Series, base_orig_series: pd.Series) -> tuple[pd.Series, pd.Series]:
            raw_num = pd.to_numeric(raw_series.map(self._to_number), errors='coerce')
            base_num = pd.to_numeric(base_orig_series.map(self._to_number), errors='coerce').fillna(0)
            correction_triggered = bool((raw_num > 100).fillna(False).any())
            rates = []
            amounts = []
            for rv, ov in zip(raw_num.tolist(), base_num.tolist()):
                if pd.isna(rv):
                    rates.append(float("nan"))
                    amounts.append(0.0)
                    continue
                rv = float(rv)
                ov = float(ov) if pd.notna(ov) else 0.0
                if correction_triggered:
                    # 触发纠偏后整列按“净残值金额”口径，避免同列金额/百分比混算
                    amt = max(0.0, rv)
                    rr = (amt / ov) if ov > 0 else float("nan")
                elif rv <= 1:
                    rr = rv
                    amt = max(0.0, ov * rr)
                else:
                    rr = rv / 100.0
                    amt = max(0.0, ov * rr)
                rr = max(0.0, min(float(rr), 1.0)) if pd.notna(rr) else float("nan")
                rates.append(rr)
                amounts.append(float(amt))
            return (
                pd.Series(rates, index=raw_series.index, dtype='float64'),
                pd.Series(amounts, index=raw_series.index, dtype='float64'),
            )

        if residual_col1 and residual_col1 in df_work.columns:
            r1, a1 = _to_rate_amount_series(df_work[residual_col1], df_work[original_value_col1])
            df_work[temp_res1_col] = r1
            df_work[temp_res_amt1_col] = a1
        else:
            df_work[temp_res1_col] = pd.Series([float("nan")] * len(df_work), index=df_work.index, dtype='float64')
            df_work[temp_res_amt1_col] = pd.Series([0.0] * len(df_work), index=df_work.index, dtype='float64')

        if residual_col2 and residual_col2 in df_work.columns:
            r2, a2 = _to_rate_amount_series(df_work[residual_col2], df_work[original_value_col2])
            df_work[temp_res2_col] = r2
            df_work[temp_res_amt2_col] = a2
        else:
            df_work[temp_res2_col] = pd.Series([float("nan")] * len(df_work), index=df_work.index, dtype='float64')
            df_work[temp_res_amt2_col] = pd.Series([0.0] * len(df_work), index=df_work.index, dtype='float64')

        key_cols = [temp_cat1_col, temp_cat2_col, temp_life_col1, temp_life_col2]
        out = (
            df_work.groupby(key_cols, as_index=False, dropna=False)
            .agg({
                temp_value_col1: "sum",
                temp_value_col2: "sum",
                temp_res_amt1_col: "sum",
                temp_res_amt2_col: "sum",
            })
        )
        # 组内残值率 = 净残值合计 / 原值合计
        out[temp_res1_col] = out.apply(
            lambda r: (float(r[temp_res_amt1_col]) / float(r[temp_value_col1])) if float(r[temp_value_col1]) != 0 else 0.0,
            axis=1
        )
        out[temp_res2_col] = out.apply(
            lambda r: (float(r[temp_res_amt2_col]) / float(r[temp_value_col2])) if float(r[temp_value_col2]) != 0 else 0.0,
            axis=1
        )
        out = out.rename(
            columns={
                temp_cat1_col: category_col1,
                temp_cat2_col: category_col2,
                temp_life_col1: life_col1,
                temp_life_col2: life_col2,
                temp_value_col1: original_value_col1,
                temp_value_col2: original_value_col2,
                temp_res1_col: (residual_col1 or "文件1残值率"),
                temp_res2_col: (residual_col2 or "文件2残值率"),
            },
            errors='ignore'
        )
        out = out.drop(columns=[temp_res_amt1_col, temp_res_amt2_col], errors='ignore')
        out = self._make_unique_columns(out)
        if out is None or out.empty:
            return None

        # 分组后再做一轮A/B兜底，避免输出N/A
        if out.shape[1] >= 4:
            col_a = out.columns[0]
            col_b = out.columns[1]
            col_c = out.columns[2]
            col_d = out.columns[3]
            # 显示口径：未分类/空值统一显示为“未分类”
            a_unc_all = out[col_a].map(_is_unc)
            b_unc_all = out[col_b].map(_is_unc)
            if bool(a_unc_all.any()):
                out.loc[a_unc_all, col_a] = "未分类"
            if bool(b_unc_all.any()):
                out.loc[b_unc_all, col_b] = "未分类"
            b_unc = out[col_b].map(_is_unc) & (~out[col_a].map(_is_unc))
            a_unc = out[col_a].map(_is_unc) & (~out[col_b].map(_is_unc))
            if bool(b_unc.any()):
                out.loc[b_unc, col_b] = out.loc[b_unc, col_a]
                out.loc[b_unc, col_d] = out.loc[b_unc, col_c]
            # 业务规则：当年末有分类而年初无分类时，年初A列保持“未分类”，不再回填为年末分类
            both_unc = out[col_a].map(_is_unc) & out[col_b].map(_is_unc)
            if bool(both_unc.any()):
                out.loc[both_unc, col_a] = "未分类"
                out.loc[both_unc, col_b] = "未分类"

        # 列顺序：A/B/C/D + 文件1残值率 + 文件2残值率 + 原值(文件1/文件2)
        if out.shape[1] >= 8:
            ordered = [
                out.columns[0], out.columns[1], out.columns[2], out.columns[3],
                out.columns[6], out.columns[7], out.columns[4], out.columns[5]
            ]
            out = out[ordered]

        # 数值列标准化
        for c in out.columns[2:]:
            out[c] = pd.to_numeric(out[c], errors='coerce')

        # 删除 C-E 同时为0的行（按当前列位置：C=2, D=3, E=4）
        if out.shape[1] >= 5:
            cde = out.iloc[:, [2, 3, 4]].apply(lambda s: pd.to_numeric(s, errors='coerce').fillna(0))
            keep = ~(cde.eq(0).all(axis=1))
            out = out.loc[keep].reset_index(drop=True)

        # G/H 辅助列
        if out.shape[1] >= 6:
            life1_col = out.columns[2]
            life2_col = out.columns[3]
            res1_col = out.columns[4]
            res2_col = out.columns[5]
            orig1_col = out.columns[6] if out.shape[1] >= 8 else None
            orig2_col = out.columns[7] if out.shape[1] >= 8 else out.columns[5]
            orig1_num = pd.to_numeric(out[orig1_col], errors='coerce').fillna(0) if orig1_col else pd.Series([0.0] * len(out), index=out.index)
            orig2_num = pd.to_numeric(out[orig2_col], errors='coerce').fillna(0)
            # 展示口径：若对应原值为0，则对应使用寿命显示为0
            out.loc[orig1_num == 0, life1_col] = 0
            out.loc[orig2_num == 0, life2_col] = 0
            life1_num = pd.to_numeric(out[life1_col], errors='coerce')
            life2_num = pd.to_numeric(out[life2_col], errors='coerce')
            res1_num = pd.to_numeric(out[res1_col], errors='coerce').fillna(0)
            res2_num = pd.to_numeric(out[res2_col], errors='coerce').fillna(0)

            # 规则：
            # 1) 年初原值=0且年末原值!=0：检查“本资产类型+年末寿命”是否在年初已存在；
            #    - 已存在 -> 一致
            #    - 不存在 -> 不一致（并计算影响当年金额）
            # 2) 其他“年初/年末原值任一为0” -> 一致
            # 3) 两侧残值率均非0，且绝对差>0.5% -> 不一致
            # 4) 其他按寿命是否一致判断
            zero_orig = (orig1_num == 0) | (orig2_num == 0)
            new_end_mask = (orig1_num == 0) & (orig2_num != 0)
            res_diff_conflict = (res1_num != 0) & (res2_num != 0) & ((res1_num - res2_num).abs() > 0.005)
            same_num = life1_num.notna() & life2_num.notna() & (life1_num == life2_num)
            same_txt = out[life1_col].astype(str).str.strip().eq(out[life2_col].astype(str).str.strip())
            life_consistent = same_num | same_txt
            cat1_col = out.columns[0]
            cat2_col = out.columns[1]

            def _asset_type_keys(row) -> set:
                keys = set()
                c1 = str(row.get(cat1_col, "")).strip()
                c2 = str(row.get(cat2_col, "")).strip()
                if c1 and (not _is_unc(c1)):
                    keys.add(c1)
                if c2 and (not _is_unc(c2)):
                    keys.add(c2)
                return keys

            def _life_token(v) -> str:
                n = pd.to_numeric(pd.Series([v]), errors='coerce').iloc[0]
                if pd.notna(n):
                    return f"N:{float(n):.10g}"
                s = str(v).strip()
                return f"S:{s}" if s else ""

            life1_tokens = out[life1_col].map(_life_token)
            life2_tokens = out[life2_col].map(_life_token)
            start_life_by_type: Dict[str, set] = {}
            for idx, row in out.iterrows():
                if float(orig1_num.loc[idx]) == 0:
                    continue
                t = life1_tokens.loc[idx]
                keys = _asset_type_keys(row)
                if (not keys) or (not t):
                    continue
                for k in keys:
                    if k not in start_life_by_type:
                        start_life_by_type[k] = set()
                    start_life_by_type[k].add(t)

            life_exists_in_start = pd.Series(False, index=out.index)
            for idx, row in out.loc[new_end_mask].iterrows():
                t = life2_tokens.loc[idx]
                if not t:
                    continue
                keys = _asset_type_keys(row)
                for k in keys:
                    if t in start_life_by_type.get(k, set()):
                        life_exists_in_start.loc[idx] = True
                        break

            new_life_in_end_mask = new_end_mask & (~life_exists_in_start)
            other_zero_mask = zero_orig & (~new_end_mask)
            end_only_category_mask = out[cat1_col].map(_is_unc) & (~out[cat2_col].map(_is_unc))
            out["判断结果"] = "不一致"
            out.loc[life_consistent, "判断结果"] = "一致"
            out.loc[res_diff_conflict, "判断结果"] = "不一致"
            out.loc[other_zero_mask, "判断结果"] = "一致"
            out.loc[new_end_mask & life_exists_in_start, "判断结果"] = "一致"
            out.loc[new_life_in_end_mask, "判断结果"] = "不一致"
            out.loc[end_only_category_mask, "判断结果"] = "待确认"

            def _calc_impact(row):
                if str(row.get("判断结果", "")).strip() != "不一致":
                    return 0.0
                y_end_orig = pd.to_numeric(pd.Series([row.get(orig2_col)]).map(self._to_number), errors='coerce').iloc[0]
                y_end_life = pd.to_numeric(pd.Series([row.get(life2_col)]), errors='coerce').iloc[0]
                y_start_life = pd.to_numeric(pd.Series([row.get(life1_col)]), errors='coerce').iloc[0]
                rr_start = pd.to_numeric(pd.Series([row.get(res1_col)]), errors='coerce').iloc[0]
                rr_end = pd.to_numeric(pd.Series([row.get(res2_col)]), errors='coerce').iloc[0]
                if row.name in new_life_in_end_mask.index and bool(new_life_in_end_mask.loc[row.name]):
                    if pd.isna(y_start_life) or y_start_life == 0:
                        y_start_life = y_end_life
                if pd.isna(y_end_orig) or pd.isna(y_end_life) or pd.isna(y_start_life) or y_end_life == 0 or y_start_life == 0:
                    return 0.0
                rr_start = 0.0 if pd.isna(rr_start) else max(0.0, min(float(rr_start), 1.0))
                rr_end = 0.0 if pd.isna(rr_end) else max(0.0, min(float(rr_end), 1.0))
                # 新口径：年末原值*(1-年末残值率)/年末寿命 - 年末原值*(1-年初残值率)/年初寿命
                return (
                    float(y_end_orig) * (1.0 - rr_end) / float(y_end_life)
                    - float(y_end_orig) * (1.0 - rr_start) / float(y_start_life)
                )

            out["影响当年金额"] = out.apply(_calc_impact, axis=1)
            def _impact_expr(row):
                if str(row.get("判断结果", "")).strip() != "不一致":
                    return ""
                y_end_orig = pd.to_numeric(pd.Series([row.get(orig2_col)]).map(self._to_number), errors='coerce').iloc[0]
                y_end_life = pd.to_numeric(pd.Series([row.get(life2_col)]), errors='coerce').iloc[0]
                y_start_life = pd.to_numeric(pd.Series([row.get(life1_col)]), errors='coerce').iloc[0]
                rr_start = pd.to_numeric(pd.Series([row.get(res1_col)]), errors='coerce').iloc[0]
                rr_end = pd.to_numeric(pd.Series([row.get(res2_col)]), errors='coerce').iloc[0]
                impact = row.get("影响当年金额", 0.0)
                if row.name in new_life_in_end_mask.index and bool(new_life_in_end_mask.loc[row.name]):
                    if pd.isna(y_start_life) or y_start_life == 0:
                        y_start_life = y_end_life
                if pd.isna(y_end_orig) or pd.isna(y_end_life) or pd.isna(y_start_life) or y_end_life == 0 or y_start_life == 0:
                    return ""
                rr_start = 0.0 if pd.isna(rr_start) else max(0.0, min(float(rr_start), 1.0))
                rr_end = 0.0 if pd.isna(rr_end) else max(0.0, min(float(rr_end), 1.0))
                return f"{float(y_end_orig):,.2f}*(1-{float(rr_end):.4f})/{float(y_end_life):.4f}-{float(y_end_orig):,.2f}*(1-{float(rr_start):.4f})/{float(y_start_life):.4f}={float(impact):,.2f}"
            out["计算过程"] = out.apply(_impact_expr, axis=1)
        return out

    def set_progress_callback(self, callback):
        """璁剧疆杩涘害鍥炶皟鍑芥暟"""
        self.export_progress_callback = callback
    
    def _update_progress(self, progress: int, message: str = ""):
        """鏇存柊杩涘害"""
        if self.export_progress_callback:
            self.export_progress_callback(progress, message)

    def _beautify_workbook(self, wb, summary_sheet_name: str = "固定资产变动汇总表", summary_config: Optional[Dict] = None):
        """??????????"""
        original_col1 = summary_config.get('original_value_col1') if summary_config else None
        original_col2 = summary_config.get('original_value_col2') if summary_config else None
        for ws in wb.worksheets:
            if ws.title == summary_sheet_name:
                ws.freeze_panes = "C3"
                self._left_align_sheet(ws)
                self._apply_black_border_sheet(ws)
                continue
            if ws.title == "折旧期间" and ws.max_row >= 2:
                ws.delete_rows(1, 2)
                self._fill_down_columns(ws, start_col=1, end_col=4)
                if ws.max_row >= 1:
                    if ws.cell(row=1, column=5).value in (None, ""):
                        ws.cell(row=1, column=5, value=original_col1 or "文件1原值")
                    if ws.cell(row=1, column=6).value in (None, ""):
                        ws.cell(row=1, column=6, value=original_col2 or "文件2原值")
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=1, column=c).value
                        if str(v).strip() == "计算过程":
                            ws.cell(
                                row=1,
                                column=c,
                                value="计算过程=年末原值*(1-年末残值率)/年末寿命-年末原值*(1-年初残值率)/年初寿命",
                            )
                            break
            self._beautify_sheet(ws, freeze_panes="A2", summary_config=summary_config)
            self._apply_black_border_sheet(ws)

    def _beautify_sheet(self, ws, freeze_panes: str = "A2", summary_config: Optional[Dict] = None):
        """???????????"""
        if ws.max_row < 1 or ws.max_column < 1:
            return

        # Use column-level styles to keep export responsive on large sheets.
        sample_end = min(ws.max_row, 10)
        for col in range(1, ws.max_column + 1):
            head_cell = ws.cell(row=1, column=col)
            if not isinstance(head_cell, MergedCell):
                head_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            col_dim = ws.column_dimensions[col_letter]
            col_dim.alignment = Alignment(horizontal="left", vertical="center")

            max_len = 0
            non_empty = 0
            numeric_count = 0
            for row in range(1, sample_end + 1):
                value = ws.cell(row=row, column=col).value
                if value is None or value == "":
                    continue
                non_empty += 1
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    numeric_count += 1
                max_len = max(max_len, len(str(value)))

            col_dim.width = min(max(max_len + 2, 10), 42)
            if non_empty > 0 and numeric_count / non_empty >= 0.9:
                col_dim.number_format = "#,##0"

        # Force numeric format for mapped original/depreciation fields (no text output).
        self._apply_forced_numeric_formats(ws, summary_config)

        # Sheet-specific business formats.
        if ws.title in ("FA List", "新增清单_BKD"):
            self._apply_percent_format_by_header(ws, {"残值率"})
        if ws.title == "处置清单_BKD":
            self._apply_numeric_format_by_header(ws, {"净值"}, "#,##0")

        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.freeze_panes = freeze_panes

    def _left_align_sheet(self, ws):
        """将sheet内容统一为左对齐，保留其他对齐属性。"""
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].alignment = Alignment(horizontal="left", vertical="center")
            head_cell = ws.cell(row=1, column=col)
            if not isinstance(head_cell, MergedCell):
                head_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _to_number(self, value):
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s or s.lower() in ("nan", "none"):
            return None
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1]
        s = s.replace(",", "").replace("%", "").replace(" ", "")
        try:
            n = float(s)
            return -n if neg else n
        except Exception:
            return None

    def _apply_numeric_format_by_header(self, ws, headers: set, fmt: str):
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header not in headers:
                continue
            ws.column_dimensions[get_column_letter(col)].number_format = fmt
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                num = self._to_number(cell.value)
                if num is not None:
                    cell.value = num
                    cell.number_format = fmt
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _apply_percent_format_by_header(self, ws, headers: set):
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header not in headers:
                continue
            ws.column_dimensions[get_column_letter(col)].number_format = "0.00%"
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                num = self._to_number(cell.value)
                if num is not None:
                    if abs(num) > 1:
                        num = num / 100.0
                    cell.value = num
                    cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _apply_forced_numeric_formats(self, ws, summary_config: Optional[Dict] = None):
        force_headers = set()
        if summary_config:
            for k in ("original_value_col1", "original_value_col2", "depreciation_col1", "depreciation_col2"):
                v = summary_config.get(k)
                if v:
                    force_headers.add(v)
            fm = summary_config.get("field_mapping") or {}
            for k in ("disposal_orig_col1", "disposal_orig_col2", "disposal_dep_col1", "disposal_dep_col2"):
                v = fm.get(k)
                if v:
                    force_headers.add(v)
        force_headers.update({"原值", "累计折旧", "原值增加", "净值", "处置原值", "处置折旧", "原值减少", "年初累计折旧"})
        self._apply_numeric_format_by_header(ws, force_headers, "#,##0")

    def _has_user_addition_mapping(self, summary_config: Optional[Dict]) -> bool:
        fm = (summary_config or {}).get('field_mapping', {}) or {}
        return any(
            fm.get(k) for k in (
                'addition_method_col1', 'addition_method_col2',
                'addition_date_col1', 'addition_date_col2'
            )
        )

    def _has_user_disposal_mapping(self, summary_config: Optional[Dict]) -> bool:
        fm = (summary_config or {}).get('field_mapping', {}) or {}
        return any(
            fm.get(k) for k in (
                'disposal_method_col1', 'disposal_method_col2',
                'disposal_date_col1', 'disposal_date_col2',
                'disposal_orig_col1', 'disposal_orig_col2',
                'disposal_dep_col1', 'disposal_dep_col2'
            )
        )

    def _apply_placeholder_tail_columns(self, df: pd.DataFrame, placeholder_headers: List[str]) -> pd.DataFrame:
        if df is None:
            return df
        out = df.copy()
        n = len(placeholder_headers)
        if out.shape[1] < n:
            for h in placeholder_headers:
                out[h] = h
            return out
        start = out.shape[1] - n
        cols = list(out.columns)
        for i, h in enumerate(placeholder_headers):
            old_col = cols[start + i]
            out = out.rename(columns={old_col: h})
            out[h] = h
        return out

    def _apply_addition_placeholders(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        无用户新增清单映射时：
        - 新增方式/新增时间使用问号占位
        - 原值增加保留真实数据，并放到占位列前面
        """
        if df is None:
            return df
        out = df.copy()

        if "[原值增加?]" in out.columns and "原值增加" not in out.columns:
            out = out.rename(columns={"[原值增加?]": "原值增加"})

        rename_map = {}
        if "新增方式" in out.columns:
            rename_map["新增方式"] = "[新增方式?]"
        if "新增时间" in out.columns:
            rename_map["新增时间"] = "[新增时间?]"
        if rename_map:
            out = out.rename(columns=rename_map)

        for h in ("[新增方式?]", "[新增时间?]"):
            if h not in out.columns:
                out[h] = h
            else:
                out[h] = h

        if "增加类型" not in out.columns:
            out["增加类型"] = ""
        return self._reorder_addition_tail_columns(out)

    def _reorder_addition_tail_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """新增清单中新增方式/新增时间（含占位列）固定放在最后两列。"""
        if df is None:
            return df
        out = df.copy()
        cols = list(out.columns)
        tail_candidates = ["新增方式", "新增时间", "[新增方式?]", "[新增时间?]"]
        tail_cols = [c for c in tail_candidates if c in cols]
        if not tail_cols:
            return out
        head_cols = [c for c in cols if c not in tail_cols]
        return out[head_cols + tail_cols]

    def _build_fa_list_from_merged_file2(self, merged_df: pd.DataFrame, summary_config: Optional[Dict]) -> tuple:
        """FA List 直接取自主数据中的文件2列（展示口径一致）。"""
        try:
            if merged_df is None or merged_df.empty:
                return False, "主数据为空", None
            sc = summary_config or {}
            fm = sc.get("field_mapping") or {}
            use_supplement_lists = bool(sc.get("use_supplement_lists", True))

            category_col = fm.get("category_col2")
            match_col = sc.get("match_col2")
            name_col = fm.get("name_col2")
            date_col = fm.get("date_col2")
            life_col = fm.get("life_col2")
            residual_col = fm.get("residual_col2")
            current_year_dep_col = fm.get("current_year_dep_col2")
            orig_col = sc.get("original_value_col2")
            dep_col = sc.get("depreciation_col2")

            # 过滤出包含文件2数据的行
            work = merged_df.copy()
            if "数据来源" in work.columns:
                src = work["数据来源"].astype(str)
                file1_name = sc.get("file1_display_name") or "文件1"
                keep = (~src.str.contains(f"仅{file1_name}", regex=False)) | (src == "两文件都有")
                work = work[keep].copy()
            if work.empty:
                return True, "FA List生成成功，共0条记录", pd.DataFrame(
                    columns=["资产类别", "固定资产编号", "固定资产名称", "入账开始日期", "使用寿命(月)", "残值率", "原值", "累计折旧", "净值", "已提足折旧", "提足折旧时间", "本年折旧"]
                )

            # 缺失时使用空列，保证结构稳定
            def _col_or_blank(col_name: Optional[str]):
                if col_name and col_name in work.columns:
                    return work[col_name]
                return pd.Series([""] * len(work), index=work.index)

            # 本年折旧口径：
            # - FA List 优先取年末文件映射的“本年折旧”列；
            # - 年末本年折旧为空时，退回 -(累计折旧变动)；
            # - 原值减少卡片若有处置清单，则按处置清单_BKD 公式口径覆盖。
            if current_year_dep_col and current_year_dep_col in work.columns:
                mapped_current_year_dep_series = pd.to_numeric(
                    work[current_year_dep_col].map(self._to_number), errors="coerce"
                )
            else:
                mapped_current_year_dep_series = pd.Series([float("nan")] * len(work), index=work.index)
            if "累计折旧变动" in work.columns:
                dep_change_series = pd.to_numeric(
                    work["累计折旧变动"].map(self._to_number), errors="coerce"
                )
            else:
                dep_change_series = pd.Series([float("nan")] * len(work), index=work.index)
            base_current_year_dep = -dep_change_series

            disposal_dep_lookup: Dict = {}
            if use_supplement_lists and "原值变动类型" in merged_df.columns:
                try:
                    disposal_idx_in_merged = merged_df.index[
                        merged_df["原值变动类型"].astype(str) == "原值减少"
                    ].tolist()
                    disp_ok, _disp_msg, disp_df_temp = self.sheet_generator.generate_disposal_list(merged_df)
                    if (
                        disp_ok
                        and disp_df_temp is not None
                        and not disp_df_temp.empty
                        and "本年折旧" in disp_df_temp.columns
                        and len(disp_df_temp) == len(disposal_idx_in_merged)
                    ):
                        for i, orig_idx in enumerate(disposal_idx_in_merged):
                            disposal_dep_lookup[orig_idx] = disp_df_temp.iloc[i].get("本年折旧", "")
                except Exception:
                    disposal_dep_lookup = {}

            def _is_blank_dep_value(v) -> bool:
                if v is None:
                    return True
                try:
                    if pd.isna(v):
                        return True
                except Exception:
                    pass
                if isinstance(v, str):
                    s = v.strip()
                    return s == "" or s == "[需客户提供]"
                return False

            def _calc_current_year_dep(idx):
                try:
                    is_disposal = ("原值变动类型" in work.columns) and (
                        str(work.at[idx, "原值变动类型"]) == "原值减少"
                    )
                except Exception:
                    is_disposal = False
                mapped_val = (
                    mapped_current_year_dep_series.loc[idx]
                    if idx in mapped_current_year_dep_series.index
                    else float("nan")
                )
                base_val = (
                    base_current_year_dep.loc[idx]
                    if idx in base_current_year_dep.index
                    else float("nan")
                )
                if not _is_blank_dep_value(mapped_val):
                    base_val = mapped_val
                if not use_supplement_lists:
                    return base_val
                if not is_disposal:
                    return base_val
                v = disposal_dep_lookup.get(idx, "")
                if _is_blank_dep_value(v):
                    return base_val
                try:
                    return float(v)
                except Exception:
                    return base_val

            current_year_dep_series = pd.Series(
                [_calc_current_year_dep(idx) for idx in work.index],
                index=work.index,
            )

            out = pd.DataFrame({
                "资产类别": _col_or_blank(category_col),
                "固定资产编号": _col_or_blank(match_col if (match_col and match_col in work.columns) else None),
                "固定资产名称": _col_or_blank(name_col),
                "入账开始日期": _col_or_blank(date_col),
                "使用寿命(月)": _col_or_blank(life_col),
                "残值率": _col_or_blank(residual_col),
                "原值": _col_or_blank(orig_col),
                "累计折旧": _col_or_blank(dep_col),
                "本年折旧": current_year_dep_series,
            })
            # 若单列匹配不存在，退回“匹配列”展示
            if ("固定资产编号" not in out.columns) or out["固定资产编号"].isna().all() or (out["固定资产编号"].astype(str).str.strip() == "").all():
                if "匹配列" in work.columns:
                    out["固定资产编号"] = work["匹配列"]

            def _to_num(v):
                if v is None:
                    return 0.0
                try:
                    if pd.isna(v):
                        return 0.0
                except Exception:
                    pass
                s = str(v).replace(",", "").strip()
                if s == "":
                    return 0.0
                try:
                    return float(s)
                except Exception:
                    return 0.0

            # 保持原FA List口径：残值率纠偏（若列中存在>100，按“残值/原值”计算）
            if "残值率" in out.columns and "原值" in out.columns:
                try:
                    res_num = pd.to_numeric(out["残值率"], errors="coerce")
                    if bool(res_num.notna().any()) and float(res_num.max()) > 100:
                        orig_num = pd.to_numeric(out["原值"], errors="coerce")
                        corrected = res_num / orig_num.replace(0, pd.NA)
                        out["残值率"] = corrected.fillna(0.0)
                except Exception:
                    pass

            orig_num = out["原值"].map(_to_num)
            dep_num = out["累计折旧"].map(_to_num)
            out["净值"] = orig_num - dep_num.abs()
            out["已提足折旧"] = [
                self.sheet_generator._is_fully_depreciated(net, orig, residual)
                for net, orig, residual in zip(out["净值"], out["原值"], out["残值率"])
            ]
            out["提足折旧时间"] = [
                self.sheet_generator._calculate_depreciation_end_date(start_date, life)
                for start_date, life in zip(out["入账开始日期"], out["使用寿命(月)"])
            ]

            # 去除完全空白ID行（若有）
            if "固定资产编号" in out.columns:
                id_series = out["固定资产编号"].astype(str).str.strip()
                out = out[id_series != ""].copy()

            return True, f"FA List生成成功，共{len(out)}条记录", out
        except Exception as e:
            return False, f"FA List生成失败: {str(e)}", None

    @staticmethod
    def _first_non_empty(series: pd.Series):
        for v in series:
            if v is None:
                continue
            if pd.isna(v):
                continue
            s = str(v).strip()
            if s != "":
                return v
        return None

    def _fill_display_fields_by_duplicate_id(
        self,
        df: pd.DataFrame,
        id_col: str,
        fill_cols: List[str],
        keep_first_only_cols: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """
        在重复ID组内，仅补齐展示字段（不改匹配/计算口径）。
        keep_first_only_cols: 这些列仅首行保留值，其余行清空（用于新增清单原值增加）。
        """
        if df is None or df.empty or id_col not in df.columns:
            return df
        out = df.copy()
        grp = out[id_col].astype(str).fillna("")
        dup_mask = grp.duplicated(keep=False) & (grp != "")
        if not bool(dup_mask.any()):
            return out

        valid_fill_cols = [c for c in fill_cols if c in out.columns and c != id_col]
        for col in valid_fill_cols:
            fill_map = out.loc[dup_mask].groupby(grp[dup_mask])[col].apply(self._first_non_empty)
            miss_mask = dup_mask & (
                out[col].isna() |
                (out[col].astype(str).str.strip() == "")
            )
            mapped = grp[miss_mask].map(fill_map)
            mapped = mapped.apply(lambda v: self._coerce_for_series_dtype(out[col], v))
            out.loc[miss_mask, col] = mapped

        if keep_first_only_cols:
            for col in keep_first_only_cols:
                if col not in out.columns:
                    continue
                order_in_grp = out.loc[dup_mask].groupby(grp[dup_mask]).cumcount()
                clear_mask = pd.Series(False, index=out.index)
                clear_mask.loc[dup_mask] = order_in_grp > 0
                out.loc[clear_mask, col] = self._coerce_for_series_dtype(out[col], "")

        return out

    def _sync_group_first_values(
        self,
        df: pd.DataFrame,
        id_col: str,
        sync_cols: List[str],
        only_when_blank: bool = False,
    ) -> pd.DataFrame:
        """重复ID组内按首行值同步指定列。"""
        if df is None or df.empty or id_col not in df.columns:
            return df
        out = df.copy()
        grp = out[id_col].astype(str).fillna("")
        dup_mask = grp.duplicated(keep=False) & (grp != "")
        if not bool(dup_mask.any()):
            return out

        valid_cols = [c for c in sync_cols if c in out.columns]
        for col in valid_cols:
            first_map = out.loc[dup_mask].groupby(grp[dup_mask])[col].apply(
                lambda s: s.iloc[0] if len(s) > 0 else ""
            )
            if only_when_blank:
                tgt_mask = dup_mask & (out[col].isna() | (out[col].astype(str).str.strip() == ""))
            else:
                tgt_mask = dup_mask
            mapped = grp[tgt_mask].map(first_map)
            mapped = mapped.apply(lambda v: self._coerce_for_series_dtype(out[col], v))
            out.loc[tgt_mask, col] = mapped
        return out

    @staticmethod
    def _is_blank_value(v) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        s = str(v).strip()
        return s == "" or s.lower() == "nan"

    @staticmethod
    def _coerce_for_series_dtype(series: pd.Series, value):
        """按目标列 dtype 做赋值兼容，避免 StringDtype 列写入 int 报错。"""
        try:
            if pd.api.types.is_string_dtype(series.dtype):
                if value is None:
                    return pd.NA
                try:
                    if pd.isna(value):
                        return pd.NA
                except Exception:
                    pass
                return str(value)
        except Exception:
            pass
        return value

    @staticmethod
    def _format_match_component(v) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        s = str(v).strip()
        if s.lower() == "nan":
            return ""
        return s

    def _build_match_key_from_row(self, row, match_cols: List[str]) -> str:
        parts = [self._format_match_component(row.get(c, "")) for c in match_cols]
        return " | ".join(parts)

    def _build_match_key_series_from_cols(self, df: pd.DataFrame, cols: List[str]) -> pd.Series:
        if df is None or df.empty or not cols:
            return pd.Series([""] * (0 if df is None else len(df)), index=(None if df is None else df.index))
        valid = [c for c in cols if c in df.columns]
        if not valid:
            return pd.Series([""] * len(df), index=df.index)
        formatted = pd.DataFrame(
            {col: df[col].map(self._format_match_component) for col in valid},
            index=df.index,
        )
        return formatted.agg(" | ".join, axis=1)

    def _source_life_year_mode(self, source_df: Optional[pd.DataFrame], source_life_col: Optional[str]) -> bool:
        """依据源数据判断是否应按“年->月”纠偏。"""
        if not isinstance(source_df, pd.DataFrame) or source_df.empty or not source_life_col or source_life_col not in source_df.columns:
            return False
        vals = pd.to_numeric(source_df[source_life_col], errors="coerce").dropna()
        return bool(len(vals) > 0 and (vals < 30).all())

    def _convert_life_lt30_to_month(self, df: pd.DataFrame, col_name: str) -> pd.DataFrame:
        """仅把<30的寿命值按年转月，避免对已是月的值重复转换。"""
        if df is None or df.empty or col_name not in df.columns:
            return df
        out = df.copy()
        nums = pd.to_numeric(out[col_name], errors="coerce")
        mask = nums.notna() & (nums < 30)
        if not bool(mask.any()):
            return out
        converted = (nums[mask] * 12.0)
        converted = converted.apply(lambda v: int(v) if v == int(v) else v)
        if pd.api.types.is_string_dtype(out[col_name].dtype):
            converted = converted.apply(lambda v: str(v) if pd.notna(v) else pd.NA)
        out.loc[mask, col_name] = converted
        return out

    @staticmethod
    def _strip_display_suffix(col_name: Optional[str], display_name: Optional[str]) -> Optional[str]:
        if not col_name:
            return None
        s = str(col_name)
        if display_name:
            suf = f"_{display_name}"
            if s.endswith(suf):
                return s[: -len(suf)]
        if s.endswith("_文件2"):
            return s[:-4]
        if s.endswith("_文件1"):
            return s[:-4]
        return s

    def _build_template_map_from_source(
        self,
        source_df: Optional[pd.DataFrame],
        match_cols: List[str],
        output_to_source_cols: Dict[str, str],
        include_blank_key: bool = False,
    ) -> Dict[str, List[Dict[str, object]]]:
        templates: Dict[str, List[Dict[str, object]]] = {}
        if not isinstance(source_df, pd.DataFrame) or source_df.empty:
            return templates
        valid_match_cols = [c for c in (match_cols or []) if c in source_df.columns]
        if not valid_match_cols:
            return templates
        if not output_to_source_cols:
            return templates

        cache_key = (
            id(source_df),
            tuple(valid_match_cols),
            tuple(sorted((str(k), str(v)) for k, v in output_to_source_cols.items())),
            bool(include_blank_key),
        )
        cached = self._template_map_cache.get(cache_key)
        if cached is not None:
            return cached

        source_cols = set(source_df.columns)
        for row in source_df.to_dict("records"):
            key = self._build_match_key_from_row(row, valid_match_cols)
            if (not include_blank_key) and self._is_blank_value(key):
                continue
            item = {}
            for out_col, src_col in output_to_source_cols.items():
                if src_col and src_col in source_cols:
                    item[out_col] = row.get(src_col, "")
            templates.setdefault(key, []).append(item)
        self._template_map_cache[cache_key] = templates
        return templates

    def _expand_rows_for_template_cardinality(
        self,
        df: pd.DataFrame,
        id_col: str,
        template_map: Dict[str, List[Dict[str, object]]],
        include_blank_key: bool = False,
    ) -> pd.DataFrame:
        """
        仅用于展示：当某匹配键下模板行数 > 当前行数时，补足空白展示行。
        不改原有匹配/计算结果，仅保证文件2卡片可完整还原展示。
        """
        if df is None or df.empty or id_col not in df.columns or not template_map:
            return df
        out = df.copy()
        grp = out[id_col].apply(self._format_match_component)
        counts = grp.value_counts(dropna=False)

        add_rows = []
        for key, tpl_list in template_map.items():
            if (not include_blank_key) and self._is_blank_value(key):
                continue
            need = max(0, len(tpl_list) - int(counts.get(key, 0)))
            if need <= 0:
                continue
            for _ in range(need):
                row = {c: "" for c in out.columns}
                row[id_col] = key
                add_rows.append(row)

        if not add_rows:
            return out
        add_df = pd.DataFrame(add_rows, columns=out.columns)
        out = pd.concat([out, add_df], ignore_index=True)
        return out

    def _fill_duplicate_rows_from_template_order(
        self,
        df: pd.DataFrame,
        id_col: str,
        fill_cols: List[str],
        template_map: Dict[str, List[Dict[str, object]]],
        overwrite_first_row: bool = False,
        overwrite_all_rows: bool = False,
        repeat_last_when_short: bool = True,
        include_blank_key: bool = False,
    ) -> pd.DataFrame:
        """按同ID组顺序从模板回填，仅填空值，不改非空值。"""
        if df is None or df.empty or id_col not in df.columns or not template_map:
            return df
        out = df.copy()
        grp = out[id_col].apply(self._format_match_component)
        dup_mask = grp.duplicated(keep=False) if include_blank_key else (grp.duplicated(keep=False) & (grp != ""))
        if not bool(dup_mask.any()):
            return out

        valid_fill_cols = [c for c in fill_cols if c in out.columns and c != id_col]
        if not valid_fill_cols:
            return out

        group_indices = out.loc[dup_mask].groupby(grp[dup_mask], sort=False).groups
        for key, idx_group in group_indices.items():
            if (not include_blank_key) and self._is_blank_value(key):
                continue
            tpl_list = template_map.get(key, [])
            if not tpl_list:
                continue
            idx_list = list(idx_group)
            for pos, idx in enumerate(idx_list):
                if pos < len(tpl_list):
                    tpl = tpl_list[pos]
                else:
                    if not repeat_last_when_short:
                        continue
                    tpl = tpl_list[-1]
                for col in valid_fill_cols:
                    cur = out.at[idx, col]
                    v = tpl.get(col, "")
                    if self._is_blank_value(v):
                        continue
                    v = self._coerce_for_series_dtype(out[col], v)
                    if overwrite_all_rows:
                        out.at[idx, col] = v
                    elif pos == 0 and overwrite_first_row:
                        out.at[idx, col] = v
                    elif self._is_blank_value(cur):
                        out.at[idx, col] = v
        return out

    def _enhance_duplicate_display(
        self,
        merged_df: pd.DataFrame,
        fa_df: Optional[pd.DataFrame],
        add_df: Optional[pd.DataFrame],
        disp_df: Optional[pd.DataFrame] = None,
        summary_config: Optional[Dict] = None,
    ):
        """统一增强重复ID场景展示：补齐描述字段，不改变计算口径。"""
        merged_out = merged_df
        fa_out = fa_df
        add_out = add_df
        disp_out = disp_df
        sc = summary_config or {}
        fm_fmt = sc.get("field_mapping") or {}
        fm_raw = sc.get("source_field_mapping_raw") or {}
        source_file1_df = sc.get("source_file1_df")
        source_file2_df = sc.get("source_file2_df")
        source_match_cols1_raw = sc.get("source_match_cols1_raw") or []
        source_match_cols2_raw = sc.get("source_match_cols2_raw") or []
        source_original_value_col2_raw = sc.get("source_original_value_col2_raw")
        source_depreciation_col2_raw = sc.get("source_depreciation_col2_raw")
        source_life_col2_raw = fm_raw.get("life_col2")

        if isinstance(merged_out, pd.DataFrame) and not merged_out.empty and "匹配列" in merged_out.columns:
            file2_display_name = sc.get("file2_display_name")
            # 优先按“原始文件2同ID组顺序”回填资产卡片字段（含原值/累计折旧），避免简单复制首行
            original_value_col2_fmt = sc.get("original_value_col2")
            depreciation_col2_fmt = sc.get("depreciation_col2")
            merged_out_to_source_cols = {}
            # 1) 先覆盖用户映射的核心字段
            for k in ("category_col2", "name_col2", "date_col2", "life_col2", "residual_col2"):
                out_col = fm_fmt.get(k)
                src_col = fm_raw.get(k)
                if out_col and src_col and out_col in merged_out.columns:
                    merged_out_to_source_cols[out_col] = src_col
            if original_value_col2_fmt and original_value_col2_fmt in merged_out.columns:
                src_orig2 = source_original_value_col2_raw or self._strip_display_suffix(original_value_col2_fmt, file2_display_name)
                if src_orig2:
                    merged_out_to_source_cols[original_value_col2_fmt] = src_orig2
            if depreciation_col2_fmt and depreciation_col2_fmt in merged_out.columns:
                src_dep2 = source_depreciation_col2_raw or self._strip_display_suffix(depreciation_col2_fmt, file2_display_name)
                if src_dep2:
                    merged_out_to_source_cols[depreciation_col2_fmt] = src_dep2

            # 2) 扩展到“主数据中全部文件2来源列”：自动按列名反推原始文件2列
            if isinstance(source_file2_df, pd.DataFrame) and not source_file2_df.empty:
                source_cols = set(source_file2_df.columns)
                for out_col in merged_out.columns:
                    if out_col == "匹配列":
                        continue
                    if not self._is_file2_column_header(str(out_col), sc):
                        continue
                    if out_col in merged_out_to_source_cols:
                        continue
                    src_guess = self._strip_display_suffix(str(out_col), file2_display_name)
                    if src_guess in source_cols:
                        merged_out_to_source_cols[out_col] = src_guess
                        continue
                    # 处理类似 “列名_文件2_1 / 列名_<显示名>_1” 场景
                    if file2_display_name:
                        m = re.match(rf"^(.*)_{re.escape(file2_display_name)}(?:_\d+)?$", str(out_col))
                        if m and m.group(1) in source_cols:
                            merged_out_to_source_cols[out_col] = m.group(1)
                            continue
                    m2 = re.match(r"^(.*)_文件2(?:_\d+)?$", str(out_col))
                    if m2 and m2.group(1) in source_cols:
                        merged_out_to_source_cols[out_col] = m2.group(1)

            merged_tpl_map = self._build_template_map_from_source(
                source_file2_df,
                source_match_cols2_raw,
                merged_out_to_source_cols,
            )
            # 展示层补齐：确保“文件2有多少卡片，主数据文件2展示就有多少行”
            merged_out = self._expand_rows_for_template_cardinality(
                merged_out,
                "匹配列",
                merged_tpl_map,
            )
            merged_out = self._fill_duplicate_rows_from_template_order(
                merged_out,
                "匹配列",
                list(merged_out_to_source_cols.keys()),
                merged_tpl_map,
                overwrite_first_row=True,
                overwrite_all_rows=True,
                repeat_last_when_short=False,
            )

            # 其余（非文件2还原列）再做常规补空
            restore_cols = set(merged_out_to_source_cols.keys())
            merged_fill_cols = [
                c for c in merged_out.columns
                if c != "匹配列"
                and c not in restore_cols
                and any(k in str(c) for k in ("类别", "名称", "日期", "时间", "方式", "寿命", "来源"))
            ]
            merged_out = self._fill_display_fields_by_duplicate_id(
                merged_out, "匹配列", merged_fill_cols
            )
            merged_out = self._sync_group_first_values(
                merged_out,
                "匹配列",
                ["数据来源", "匹配列", "原值变动类型", "累计折旧变动类型"],
                only_when_blank=False,
            )

        if isinstance(fa_out, pd.DataFrame) and not fa_out.empty and "固定资产编号" in fa_out.columns:
            fa_out_to_source_cols = {
                "资产类别": fm_raw.get("category_col2"),
                "固定资产名称": fm_raw.get("name_col2"),
                "入账开始日期": fm_raw.get("date_col2"),
                "使用寿命(月)": fm_raw.get("life_col2"),
                "原值": source_original_value_col2_raw,
                "累计折旧": source_depreciation_col2_raw,
            }
            # 优先用“与主数据一致的复合键”做分配，保证FA与主数据文件2明细一致
            fa_work = fa_out.copy()
            fa_id_col = "固定资产编号"
            fa_tpl_map = {}
            used_composite_key = False

            fa_col_by_raw = {
                (source_match_cols2_raw[0] if source_match_cols2_raw else None): "固定资产编号",
                fm_raw.get("name_col2"): "固定资产名称",
                fm_raw.get("date_col2"): "入账开始日期",
                fm_raw.get("life_col2"): "使用寿命(月)",
                fm_raw.get("residual_col2"): "残值率",
            }
            composite_fa_cols = []
            can_build_composite = bool(source_match_cols2_raw)
            for raw_col in (source_match_cols2_raw or []):
                mapped_fa_col = fa_col_by_raw.get(raw_col)
                if not mapped_fa_col or mapped_fa_col not in fa_work.columns:
                    can_build_composite = False
                    break
                composite_fa_cols.append(mapped_fa_col)

            if can_build_composite and composite_fa_cols:
                fa_work["__fa_match_key__"] = self._build_match_key_series_from_cols(fa_work, composite_fa_cols)
                fa_id_col = "__fa_match_key__"
                fa_tpl_map = self._build_template_map_from_source(
                    source_file2_df,
                    source_match_cols2_raw,
                    fa_out_to_source_cols,
                )
                used_composite_key = True
            else:
                # 回退：单键（固定资产编号）口径
                fa_ids = set(
                    fa_work["固定资产编号"]
                    .apply(self._format_match_component)
                    .replace("", pd.NA)
                    .dropna()
                    .tolist()
                )
                key_candidates = []
                if source_match_cols2_raw:
                    key_candidates.append([source_match_cols2_raw[0]])
                    key_candidates.append(list(source_match_cols2_raw))
                best_map = {}
                best_score = -1
                for key_cols in key_candidates:
                    cur_map = self._build_template_map_from_source(
                        source_file2_df,
                        key_cols,
                        fa_out_to_source_cols,
                    )
                    score = sum(1 for k in fa_ids if k in cur_map)
                    if score > best_score:
                        best_score = score
                        best_map = cur_map
                fa_tpl_map = best_map

            fa_work = self._expand_rows_for_template_cardinality(
                fa_work,
                fa_id_col,
                fa_tpl_map,
            )
            if used_composite_key and "__fa_match_key__" in fa_work.columns:
                # 补行后重算复合键（新增空行需要带键值）
                fa_work["__fa_match_key__"] = self._build_match_key_series_from_cols(fa_work, composite_fa_cols)
            fa_work = self._fill_duplicate_rows_from_template_order(
                fa_work,
                fa_id_col,
                list(fa_out_to_source_cols.keys()),
                fa_tpl_map,
                overwrite_first_row=True,
                overwrite_all_rows=True,
                repeat_last_when_short=False,
            )
            if "__fa_match_key__" in fa_work.columns:
                fa_work = fa_work.drop(columns=["__fa_match_key__"], errors="ignore")
            fa_out = fa_work
            # 若源文件2寿命列判定为“年”，则对FA表中<30的寿命值转月（避免重复转换）
            if "使用寿命(月)" in fa_out.columns and self._source_life_year_mode(source_file2_df, source_life_col2_raw):
                fa_out = self._convert_life_lt30_to_month(fa_out, "使用寿命(月)")
            # 回填后同步净值，保证与原值/累计折旧一致
            if {"原值", "累计折旧", "净值"}.issubset(set(fa_out.columns)):
                def _to_num(v):
                    if v is None:
                        return 0.0
                    try:
                        if pd.isna(v):
                            return 0.0
                    except Exception:
                        pass
                    s = str(v).replace(",", "").strip()
                    if s == "":
                        return 0.0
                    try:
                        return float(s)
                    except Exception:
                        return 0.0
                orig_num = fa_out["原值"].map(_to_num)
                dep_num = fa_out["累计折旧"].map(_to_num)
                fa_out["净值"] = orig_num - dep_num.abs()
            # 回填后同步重算“已提足折旧/提足折旧时间”，避免沿用旧值
            if {"已提足折旧", "净值", "原值", "残值率"}.issubset(set(fa_out.columns)):
                try:
                    fa_out["已提足折旧"] = [
                        self.sheet_generator._is_fully_depreciated(net, orig, residual)
                        for net, orig, residual in zip(fa_out["净值"], fa_out["原值"], fa_out["残值率"])
                    ]
                except Exception:
                    pass
            if {"提足折旧时间", "入账开始日期", "使用寿命(月)"}.issubset(set(fa_out.columns)):
                try:
                    fa_out["提足折旧时间"] = [
                        self.sheet_generator._calculate_depreciation_end_date(start_date, life)
                        for start_date, life in zip(fa_out["入账开始日期"], fa_out["使用寿命(月)"])
                    ]
                except Exception:
                    pass

        if isinstance(add_out, pd.DataFrame) and not add_out.empty and "固定资产编号" in add_out.columns:
            add_out_to_source_cols = {
                "资产类别": fm_raw.get("category_col2"),
                "固定资产名称": fm_raw.get("name_col2"),
                "入账开始日期": fm_raw.get("date_col2"),
                "使用寿命(月)": fm_raw.get("life_col2"),
                "残值率": fm_raw.get("residual_col2"),
                "新增方式": fm_raw.get("addition_method_col2"),
                "新增时间": fm_raw.get("addition_date_col2") or fm_raw.get("date_col2"),
            }
            add_tpl_map = self._build_template_map_from_source(
                source_file2_df,
                source_match_cols2_raw,
                add_out_to_source_cols,
            )
            add_work = add_out.copy()
            add_id_col = "固定资产编号"
            # 多匹配列时使用复合键分组，避免“仅第一列ID”导致原值增加被误清空
            if source_match_cols2_raw and len(source_match_cols2_raw) > 1:
                raw_to_add_col = {
                    source_match_cols2_raw[0]: "固定资产编号",
                    fm_raw.get("category_col2"): "资产类别",
                    fm_raw.get("name_col2"): "固定资产名称",
                    fm_raw.get("date_col2"): "入账开始日期",
                    fm_raw.get("life_col2"): "使用寿命(月)",
                    fm_raw.get("residual_col2"): "残值率",
                }
                composite_cols = []
                can_build = True
                for rc in source_match_cols2_raw:
                    c = raw_to_add_col.get(rc)
                    if not c or c not in add_work.columns:
                        can_build = False
                        break
                    composite_cols.append(c)
                if can_build and composite_cols:
                    add_work["__add_match_key__"] = self._build_match_key_series_from_cols(add_work, composite_cols)
                    add_id_col = "__add_match_key__"

            add_work = self._fill_duplicate_rows_from_template_order(
                add_work,
                add_id_col,
                list(add_out_to_source_cols.keys()),
                add_tpl_map,
                overwrite_all_rows=True,
            )
            # 若源文件2寿命列判定为“年”，则对新增清单中<30的寿命值转月（避免重复转换）
            if "使用寿命(月)" in add_work.columns and self._source_life_year_mode(source_file2_df, source_life_col2_raw):
                add_work = self._convert_life_lt30_to_month(add_work, "使用寿命(月)")
            add_work = self._fill_display_fields_by_duplicate_id(
                add_work, add_id_col, [], keep_first_only_cols=["原值增加"]
            )
            if "__add_match_key__" in add_work.columns:
                add_work = add_work.drop(columns=["__add_match_key__"], errors="ignore")
            add_out = add_work

        if isinstance(disp_out, pd.DataFrame) and not disp_out.empty and "固定资产编号" in disp_out.columns:
            disp_out_to_source_cols = {
                "资产类别": fm_raw.get("category_col1"),
                "固定资产名称": fm_raw.get("name_col1"),
                "入账开始日期": fm_raw.get("date_col1"),
                "使用寿命(月)": fm_raw.get("life_col1"),
                "残值率": fm_raw.get("residual_col1"),
                "本年折旧": fm_raw.get("current_year_dep_col1"),
                "处置方式": fm_raw.get("disposal_method_col1"),
                "处置时间": fm_raw.get("disposal_date_col1"),
                "处置原值": fm_raw.get("disposal_orig_col1"),
                "处置折旧": fm_raw.get("disposal_dep_col1"),
            }
            disp_tpl_map = self._build_template_map_from_source(
                source_file1_df,
                source_match_cols1_raw,
                disp_out_to_source_cols,
            )
            disp_out = self._fill_duplicate_rows_from_template_order(
                disp_out,
                "固定资产编号",
                list(disp_out_to_source_cols.keys()),
                disp_tpl_map,
            )

        return merged_out, fa_out, add_out, disp_out

    def _export_unmatched_change_workbook(self, main_file_path: str, add_df: Optional[pd.DataFrame], disp_df: Optional[pd.DataFrame]):
        """导出未匹配资产变动清单（两个sheet放在一个工作簿）。"""
        try:
            out_dir = os.path.dirname(main_file_path) if os.path.dirname(main_file_path) else "."
            out_path = os.path.join(out_dir, "[未匹配资产变动清单].xlsx")
            add_out = add_df.copy() if isinstance(add_df, pd.DataFrame) else pd.DataFrame()
            disp_out = disp_df.copy() if isinstance(disp_df, pd.DataFrame) else pd.DataFrame()
            with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
                add_out.to_excel(writer, sheet_name="未匹配新增清单", index=False)
                disp_out.to_excel(writer, sheet_name="未匹配处置清单", index=False)
            return True, out_path
        except Exception as e:
            return False, str(e)

    def _fill_down_columns(self, ws, start_col: int = 1, end_col: int = 4):
        last_vals = {c: None for c in range(start_col, end_col + 1)}
        for row in range(1, ws.max_row + 1):
            marker = ws.cell(row=row, column=start_col).value
            if isinstance(marker, str) and marker.startswith("【"):
                last_vals = {c: None for c in range(start_col, end_col + 1)}
                continue
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if v is None or v == "":
                    if last_vals[col] is not None:
                        cell.value = last_vals[col]
                else:
                    last_vals[col] = v

    def _apply_black_border_sheet(self, ws):
        """Apply black thin border only to the outer boundary of used range."""
        thin_black = Side(style="thin", color="000000")
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 1 or max_col < 1:
            return

        def set_sides(cell, left=None, right=None, top=None, bottom=None):
            if isinstance(cell, MergedCell):
                return
            b = cell.border
            cell.border = Border(
                left=left if left is not None else b.left,
                right=right if right is not None else b.right,
                top=top if top is not None else b.top,
                bottom=bottom if bottom is not None else b.bottom,
                diagonal=b.diagonal,
                diagonal_direction=b.diagonal_direction,
                outline=b.outline,
                vertical=b.vertical,
                horizontal=b.horizontal,
            )

        # Top and bottom edges
        for col in range(1, max_col + 1):
            set_sides(ws.cell(row=1, column=col), top=thin_black)
            set_sides(ws.cell(row=max_row, column=col), bottom=thin_black)

        # Left and right edges
        for row in range(1, max_row + 1):
            set_sides(ws.cell(row=row, column=1), left=thin_black)
            set_sides(ws.cell(row=row, column=max_col), right=thin_black)
