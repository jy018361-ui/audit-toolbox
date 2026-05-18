"""
汇总表生成器模块：生成固定资产变动汇总表（支持按新增/处置方式分拆与重分类）。
"""
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class SummaryGenerator:
    """汇总表生成器 - 生成固定资产变动汇总表"""

    def __init__(self):
        self.summary_data = None
        self.categories: List[str] = []

    @staticmethod
    def _safe_numeric_series(series: pd.Series) -> pd.Series:
        def convert(val):
            if pd.isna(val):
                return 0.0
            try:
                if isinstance(val, str):
                    val = val.replace(",", "").replace(" ", "").strip()
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        return series.apply(convert)

    @staticmethod
    def _normalize_text(val, default=""):
        if val is None or pd.isna(val):
            return default
        s = str(val).strip()
        return s if s else default

    @staticmethod
    def _pick_series(
        df: pd.DataFrame,
        primary_col: Optional[str],
        secondary_col: Optional[str],
        default: str = "",
    ) -> pd.Series:
        def nonempty(v):
            if v is None or pd.isna(v):
                return None
            s = str(v).strip()
            return s if s else None

        if primary_col in df.columns and secondary_col in df.columns:
            return df.apply(
                lambda row: nonempty(row.get(primary_col))
                or nonempty(row.get(secondary_col))
                or default,
                axis=1,
            )
        if primary_col in df.columns:
            return df[primary_col].apply(lambda v: nonempty(v) or default)
        if secondary_col in df.columns:
            return df[secondary_col].apply(lambda v: nonempty(v) or default)
        return pd.Series([default] * len(df), index=df.index)

    @staticmethod
    def _ordered_unique(values: List[str]) -> List[str]:
        seen = set()
        ordered = []
        for v in values:
            if v not in seen:
                seen.add(v)
                ordered.append(v)
        return ordered

    def _build_row_defs(
        self,
        file1_name: str,
        file2_name: str,
        addition_methods: List[str],
        disposal_methods: List[str],
    ) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = [
            {"section": "原值", "item": f"{file1_name}原值", "kind": "metric", "key": "file1_original_value"},
            {"section": "", "item": "原值增加", "kind": "metric", "key": "original_increase_total"},
        ]
        for method in addition_methods:
            rows.append(
                {
                    "section": "",
                    "item": f"  其中-新增方式:{method}",
                    "kind": "metric",
                    "key": f"original_increase_by_method::{method}",
                }
            )

        rows.extend(
            [
                {"section": "", "item": "原值减少", "kind": "metric", "key": "original_decrease_total"},
            ]
        )
        for method in disposal_methods:
            disposal_method_label = (
                "非处置变动（含计提折旧）"
                if method == "其他(非处置变动)"
                else method
            )
            rows.append(
                {
                    "section": "",
                    "item": f"  其中-{disposal_method_label}",
                    "kind": "metric",
                    "key": f"original_decrease_by_method::{method}",
                }
            )

        rows.extend(
            [
                {"section": "", "item": "原值重分类", "kind": "metric", "key": "original_reclass"},
                {"section": "", "item": f"{file2_name}原值", "kind": "formula", "key": "formula_file2_orig"},
                {"section": "累计折旧", "item": f"{file1_name}累计折旧", "kind": "metric", "key": "file1_depreciation"},
                {"section": "", "item": "累计折旧变动净额", "kind": "metric", "key": "depreciation_adjustment_total"},
            ]
        )
        for method in disposal_methods:
            depreciation_method_label = (
                "非处置变动（含计提折旧）"
                if method == "其他(非处置变动)"
                else method
            )
            rows.append(
                {
                    "section": "",
                    "item": f"  其中-{depreciation_method_label}",
                    "kind": "metric",
                    "key": f"depreciation_adjustment_by_method::{method}",
                }
            )

        rows.extend(
            [
                {"section": "", "item": "累计折旧重分类", "kind": "metric", "key": "depreciation_reclass"},
                {"section": "", "item": f"{file2_name}累计折旧", "kind": "formula", "key": "formula_file2_dep"},
                {"section": "净值(NBV)", "item": "年初余额", "kind": "formula", "key": "formula_nbv_start"},
                {"section": "", "item": "年末余额", "kind": "formula", "key": "formula_nbv_end"},
            ]
        )
        return rows

    @staticmethod
    def _build_row_defs_legacy(file1_name: str, file2_name: str) -> List[Dict[str, str]]:
        return [
            {"section": "原值", "item": f"{file1_name}原值", "kind": "metric", "key": "file1_original_value"},
            {"section": "", "item": "原值增加", "kind": "metric", "key": "original_increase_total"},
            {"section": "", "item": "原值减少", "kind": "metric", "key": "original_decrease_total"},
            {"section": "", "item": "原值重分类", "kind": "metric", "key": "original_reclass"},
            {"section": "", "item": f"{file2_name}原值", "kind": "formula", "key": "formula_file2_orig"},
            {"section": "累计折旧", "item": f"{file1_name}累计折旧", "kind": "metric", "key": "file1_depreciation"},
            {"section": "", "item": "累计折旧变动净额", "kind": "metric", "key": "depreciation_adjustment_total"},
            {"section": "", "item": "累计折旧重分类", "kind": "metric", "key": "depreciation_reclass"},
            {"section": "", "item": f"{file2_name}累计折旧", "kind": "formula", "key": "formula_file2_dep"},
            {"section": "净值(NBV)", "item": "年初余额", "kind": "formula", "key": "formula_nbv_start"},
            {"section": "", "item": "年末余额", "kind": "formula", "key": "formula_nbv_end"},
        ]

    @staticmethod
    def _calc_formula_value(cat_data: Dict[str, float], key: str) -> float:
        file2_orig = (
            cat_data.get("file1_original_value", 0.0)
            + cat_data.get("original_increase_total", 0.0)
            - cat_data.get("original_decrease_total", 0.0)
            + cat_data.get("original_reclass", 0.0)
        )
        file2_dep = (
            cat_data.get("file1_depreciation", 0.0)
            + cat_data.get("depreciation_adjustment_total", 0.0)
            + cat_data.get("depreciation_reclass", 0.0)
        )
        if key == "formula_file2_orig":
            return file2_orig
        if key == "formula_file2_dep":
            return file2_dep
        if key == "formula_nbv_start":
            return cat_data.get("file1_original_value", 0.0) - abs(cat_data.get("file1_depreciation", 0.0))
        if key == "formula_nbv_end":
            return file2_orig - abs(file2_dep)
        return 0.0

    def _filter_zero_rows(
        self,
        row_defs: List[Dict[str, str]],
        data: Dict[str, Dict[str, float]],
        categories: List[str],
    ) -> List[Dict[str, str]]:
        """移除所有类别账面数均为0的行。"""
        kept: List[Dict[str, str]] = []
        for row_def in row_defs:
            all_zero = True
            for cat in categories:
                cat_data = data.get(cat, {})
                if row_def["kind"] == "formula":
                    v = self._calc_formula_value(cat_data, row_def["key"])
                else:
                    v = cat_data.get(row_def["key"], 0.0)
                if abs(float(v)) > 1e-6:
                    all_zero = False
                    break
            if not all_zero:
                kept.append(row_def)
        return kept

    def _is_category_all_zero(
        self,
        category: str,
        row_defs: List[Dict[str, str]],
        data: Dict[str, Dict[str, float]],
    ) -> bool:
        """判断指定分类在当前展示行上是否全为0。"""
        cat_data = data.get(category, {})
        for row_def in row_defs:
            if row_def["kind"] == "formula":
                v = self._calc_formula_value(cat_data, row_def["key"])
            else:
                v = cat_data.get(row_def["key"], 0.0)
            if abs(float(v)) > 1e-6:
                return False
        return True

    def _reconcile_reclass_to_closing(
        self,
        category_data: Dict[str, defaultdict],
        categories: List[str],
        target_file2_orig: pd.Series,
        target_file2_dep: pd.Series,
    ) -> None:
        """
        将“原值重分类/累计折旧重分类”并入现有口径，使各分类期末值与主数据口径对齐。
        目标：
        1) 各分类期末原值、期末累计折旧与主数据一致；
        2) 重分类净额保持0（至少消除浮点残差）。
        """
        target_orig_map = {self._normalize_text(k, "未分类"): float(v) for k, v in target_file2_orig.items()}
        target_dep_map = {self._normalize_text(k, "未分类"): float(v) for k, v in target_file2_dep.items()}

        for cat in categories:
            cat_data = category_data[cat]
            base_orig = (
                float(cat_data.get("file1_original_value", 0.0))
                + float(cat_data.get("original_increase_total", 0.0))
                - float(cat_data.get("original_decrease_total", 0.0))
            )
            base_dep = (
                float(cat_data.get("file1_depreciation", 0.0))
                + float(cat_data.get("depreciation_adjustment_total", 0.0))
            )
            cat_data["original_reclass"] = float(target_orig_map.get(cat, 0.0)) - base_orig
            cat_data["depreciation_reclass"] = float(target_dep_map.get(cat, 0.0)) - base_dep

        # 消除浮点残差，保持重分类净额=0（不改业务口径，仅修正计算精度）
        anchor = "未分类" if "未分类" in categories else (categories[0] if categories else None)
        if anchor:
            total_orig_reclass = sum(float(category_data[c].get("original_reclass", 0.0)) for c in categories)
            total_dep_reclass = sum(float(category_data[c].get("depreciation_reclass", 0.0)) for c in categories)
            if abs(total_orig_reclass) <= 1e-6:
                category_data[anchor]["original_reclass"] -= total_orig_reclass
            if abs(total_dep_reclass) <= 1e-6:
                category_data[anchor]["depreciation_reclass"] -= total_dep_reclass

    def generate_summary(
        self,
        df: pd.DataFrame,
        category_col: str,
        original_value_col1: str,
        original_value_col2: str,
        depreciation_col1: str,
        depreciation_col2: str,
        file1_display_name: str = "期初",
        file2_display_name: str = "期末",
        category_col1: Optional[str] = None,
        category_col2: Optional[str] = None,
        pivot_row_field_for_log: Optional[str] = None,
        addition_method_col1: Optional[str] = None,
        addition_method_col2: Optional[str] = None,
        disposal_method_col1: Optional[str] = None,
        disposal_method_col2: Optional[str] = None,
        disposal_orig_col1: Optional[str] = None,
        disposal_orig_col2: Optional[str] = None,
        disposal_dep_col1: Optional[str] = None,
        disposal_dep_col2: Optional[str] = None,
        extended_mode: bool = False,
        use_supplement_lists: bool = True,
    ) -> Tuple[bool, str, Optional[Dict]]:
        del pivot_row_field_for_log  # 仅为兼容旧调用参数

        try:
            if df is None or df.empty:
                return False, "数据为空，无法生成汇总表", None

            work_df = df.copy()

            if category_col not in work_df.columns:
                if category_col1 and category_col1 in work_df.columns:
                    category_col = category_col1
                elif category_col2 and category_col2 in work_df.columns:
                    category_col = category_col2
                else:
                    return False, f"未找到有效的资产类别列: {category_col}", None

            if not extended_mode:
                final_cat_series = self._pick_series(work_df, category_col1 or category_col, category_col2, "未分类")
                if category_col1 and category_col1 in work_df.columns:
                    cat1_raw_series = work_df[category_col1].apply(lambda v: self._normalize_text(v, "未分类"))
                else:
                    cat1_raw_series = pd.Series(["未分类"] * len(work_df), index=work_df.index)
                if category_col2 and category_col2 in work_df.columns:
                    cat2_raw_series = work_df[category_col2].apply(lambda v: self._normalize_text(v, "未分类"))
                else:
                    cat2_raw_series = pd.Series(["未分类"] * len(work_df), index=work_df.index)
                orig1 = self._safe_numeric_series(work_df[original_value_col1]) if original_value_col1 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
                orig2 = self._safe_numeric_series(work_df[original_value_col2]) if original_value_col2 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
                dep1 = self._safe_numeric_series(work_df[depreciation_col1]) if depreciation_col1 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
                dep2 = self._safe_numeric_series(work_df[depreciation_col2]) if depreciation_col2 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)

                if "原值变动" in work_df.columns:
                    orig_change = self._safe_numeric_series(work_df["原值变动"])
                else:
                    orig_change = orig1 - orig2

                if "累计折旧变动" in work_df.columns:
                    dep_change = self._safe_numeric_series(work_df["累计折旧变动"])
                else:
                    dep_change = dep1 - dep2

                if "原值变动类型" in work_df.columns:
                    change_type = work_df["原值变动类型"].astype(str)
                else:
                    change_type = orig_change.apply(lambda x: "原值减少" if x > 0 else ("原值增加" if x < 0 else "原值不变"))

                category_candidates: List[str] = []
                category_candidates.extend(final_cat_series.tolist())
                category_candidates.extend(cat1_raw_series.tolist())
                category_candidates.extend(cat2_raw_series.tolist())
                categories = [self._normalize_text(c, "未分类") for c in category_candidates]
                self.categories = self._ordered_unique([c for c in categories if c])
                if not self.categories:
                    return False, "未找到有效的资产类别", None

                category_data: Dict[str, defaultdict] = defaultdict(lambda: defaultdict(float))
                for cat in self.categories:
                    mask = final_cat_series == cat
                    cat_orig1 = orig1[mask]
                    cat_dep1 = dep1[mask]
                    cat_orig_change = orig_change[mask]
                    cat_dep_change = dep_change[mask]
                    cat_change_type = change_type[mask]

                    category_data[cat]["file1_original_value"] = float(cat_orig1.sum())
                    category_data[cat]["original_increase_total"] = float(cat_orig_change[cat_change_type == "原值增加"].abs().sum())
                    category_data[cat]["original_decrease_total"] = float(cat_orig_change[cat_change_type == "原值减少"].abs().sum())
                    category_data[cat]["file1_depreciation"] = float(cat_dep1.sum())
                    category_data[cat]["depreciation_adjustment_total"] = float((-cat_dep_change).sum())
                    category_data[cat]["original_reclass"] = 0.0
                    category_data[cat]["depreciation_reclass"] = 0.0

                # 重分类：年初分类 != 年末分类，将期初原值/累计折旧按年初转出、年末转入
                changed_mask = (cat1_raw_series != "未分类") & (cat2_raw_series != "未分类") & (cat1_raw_series != cat2_raw_series)
                changed_df = pd.DataFrame(
                    {
                        "cat1": cat1_raw_series[changed_mask],
                        "cat2": cat2_raw_series[changed_mask],
                        "orig1": orig1[changed_mask],
                        "dep1": dep1[changed_mask],
                    }
                )
                if not changed_df.empty:
                    out_orig = changed_df.groupby("cat1")["orig1"].sum()
                    in_orig = changed_df.groupby("cat2")["orig1"].sum()
                    out_dep = changed_df.groupby("cat1")["dep1"].sum()
                    in_dep = changed_df.groupby("cat2")["dep1"].sum()
                    for cat, val in out_orig.items():
                        category_data[cat]["original_reclass"] -= float(val)
                    for cat, val in in_orig.items():
                        category_data[cat]["original_reclass"] += float(val)
                    for cat, val in out_dep.items():
                        category_data[cat]["depreciation_reclass"] -= float(val)
                    for cat, val in in_dep.items():
                        category_data[cat]["depreciation_reclass"] += float(val)

                # 将重分类并入口径：确保各分类期末值与主数据（按年末分类）一致
                target_file2_orig = orig2.groupby(cat2_raw_series).sum()
                target_file2_dep = dep2.groupby(cat2_raw_series).sum()
                self._reconcile_reclass_to_closing(
                    category_data,
                    self.categories,
                    target_file2_orig,
                    target_file2_dep,
                )

                row_defs = self._build_row_defs_legacy(file1_display_name, file2_display_name)
                data_dict = {cat: dict(vals) for cat, vals in category_data.items()}
                row_defs = self._filter_zero_rows(row_defs, data_dict, self.categories)
                # 去掉“未分类”但全0的无效展示列
                self.categories = [
                    cat for cat in self.categories
                    if not (cat == "未分类" and self._is_category_all_zero(cat, row_defs, data_dict))
                ]
                file1_categories = set()
                file2_categories = set()
                if category_col1 and category_col1 in work_df.columns:
                    file1_categories = {
                        self._normalize_text(v, "未分类")
                        for v in work_df[category_col1].tolist()
                        if self._normalize_text(v, "未分类")
                    }
                if category_col2 and category_col2 in work_df.columns:
                    file2_categories = {
                        self._normalize_text(v, "未分类")
                        for v in work_df[category_col2].tolist()
                        if self._normalize_text(v, "未分类")
                    }
                file2_only_categories = [
                    cat for cat in self.categories
                    if cat in file2_categories and cat not in file1_categories
                ]
                summary = {
                    "categories": self.categories,
                    "file1_display_name": file1_display_name,
                    "file2_display_name": file2_display_name,
                    "data": data_dict,
                    "row_defs": row_defs,
                    "addition_methods": [],
                    "disposal_methods": [],
                    "extended_mode": False,
                    "file2_only_categories": file2_only_categories,
                }
                self.summary_data = summary
                return True, f"汇总表生成成功，共 {len(self.categories)} 个资产类别", summary

            cat1_series = self._pick_series(work_df, category_col1, category_col2, "未分类")
            cat2_series = self._pick_series(work_df, category_col2, category_col1, "未分类")
            final_cat_series = self._pick_series(work_df, category_col, category_col2 or category_col1, "未分类")
            # 重分类判定使用“原始年初/年末分类”而非互相回填，避免把真实分类差异抹平。
            if category_col1 in work_df.columns:
                cat1_raw_series = work_df[category_col1].apply(lambda v: self._normalize_text(v, "未分类"))
            else:
                cat1_raw_series = pd.Series(["未分类"] * len(work_df), index=work_df.index)
            if category_col2 in work_df.columns:
                cat2_raw_series = work_df[category_col2].apply(lambda v: self._normalize_text(v, "未分类"))
            else:
                cat2_raw_series = pd.Series(["未分类"] * len(work_df), index=work_df.index)

            # 规范化关键金额列
            orig1 = self._safe_numeric_series(work_df[original_value_col1]) if original_value_col1 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
            orig2 = self._safe_numeric_series(work_df[original_value_col2]) if original_value_col2 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
            dep1 = self._safe_numeric_series(work_df[depreciation_col1]) if depreciation_col1 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)
            dep2 = self._safe_numeric_series(work_df[depreciation_col2]) if depreciation_col2 in work_df.columns else pd.Series([0.0] * len(work_df), index=work_df.index)

            if "原值变动" in work_df.columns:
                orig_change = self._safe_numeric_series(work_df["原值变动"])
            else:
                orig_change = orig1 - orig2

            if "累计折旧变动" in work_df.columns:
                dep_change = self._safe_numeric_series(work_df["累计折旧变动"])
            else:
                dep_change = dep1 - dep2

            if "原值变动类型" in work_df.columns:
                change_type = work_df["原值变动类型"].astype(str)
            else:
                change_type = orig_change.apply(lambda x: "原值减少" if x > 0 else ("原值增加" if x < 0 else "原值不变"))

            has_add_method_mapping = bool(
                (addition_method_col1 and addition_method_col1 in work_df.columns)
                or (addition_method_col2 and addition_method_col2 in work_df.columns)
            )
            use_supplement_lists = bool(use_supplement_lists)
            has_disp_method_mapping = use_supplement_lists and bool(
                (disposal_method_col1 and disposal_method_col1 in work_df.columns)
                or (disposal_method_col2 and disposal_method_col2 in work_df.columns)
            )
            add_method_series = (
                self._pick_series(work_df, addition_method_col2, addition_method_col1, "未标注新增方式")
                if has_add_method_mapping
                else None
            )
            disp_method_series = (
                self._pick_series(work_df, disposal_method_col1, disposal_method_col2, "未标注处置方式")
                if has_disp_method_mapping
                else None
            )
            has_disposal_amount_mapping = use_supplement_lists and bool(
                (disposal_orig_col1 and disposal_orig_col1 in work_df.columns)
                or (disposal_orig_col2 and disposal_orig_col2 in work_df.columns)
                or (disposal_dep_col1 and disposal_dep_col1 in work_df.columns)
                or (disposal_dep_col2 and disposal_dep_col2 in work_df.columns)
            )

            def _to_numeric_or_nan(v):
                if v is None or pd.isna(v):
                    return pd.NA
                try:
                    if isinstance(v, str):
                        v = v.replace(",", "").replace(" ", "").strip()
                        if v == "":
                            return pd.NA
                    return float(v)
                except Exception:
                    return pd.NA

            def _align_by_ref_sign(value_series: pd.Series, ref_series: pd.Series, default_negative: bool = False) -> pd.Series:
                """金额取绝对值后按参考序列方向定符号，避免用户上传正负方向不一致。"""
                out = value_series.abs()
                sign = ref_series.apply(lambda x: -1.0 if x < 0 else (1.0 if x > 0 else (-1.0 if default_negative else 1.0)))
                return out * sign

            method_mapped_mask = pd.Series([False] * len(work_df), index=work_df.index)
            method_unmarked_mask = pd.Series([True] * len(work_df), index=work_df.index)
            if has_disp_method_mapping and disp_method_series is not None:
                def _is_unmarked_method(v) -> bool:
                    s = self._normalize_text(v, "")
                    if not s:
                        return True
                    s_lower = s.lower()
                    return s in {"0", "0.0", "未标注处置方式"} or s_lower in {"nan", "none", "null"}

                method_unmarked_mask = disp_method_series.apply(_is_unmarked_method)
                method_mapped_mask = ~method_unmarked_mask

            disp_orig_amount = None
            disp_dep_amount = None
            if has_disposal_amount_mapping:
                disp_orig_raw = self._pick_series(work_df, disposal_orig_col1, disposal_orig_col2, "")
                disp_dep_raw = self._pick_series(work_df, disposal_dep_col1, disposal_dep_col2, "")
                disp_orig_amount = disp_orig_raw.apply(_to_numeric_or_nan)
                disp_dep_amount = disp_dep_raw.apply(_to_numeric_or_nan)

            category_data: Dict[str, defaultdict] = defaultdict(lambda: defaultdict(float))

            # 重分类：年初分类 != 年末分类，将期初原值/累计折旧做净额重分类
            changed_mask = (cat1_raw_series != "未分类") & (cat2_raw_series != "未分类") & (cat1_raw_series != cat2_raw_series)
            changed_df = pd.DataFrame(
                {
                    "cat1": cat1_raw_series[changed_mask],
                    "cat2": cat2_raw_series[changed_mask],
                    "orig1": orig1[changed_mask],
                    "dep1": dep1[changed_mask],
                }
            )
            if not changed_df.empty:
                out_orig = changed_df.groupby("cat1")["orig1"].sum()
                in_orig = changed_df.groupby("cat2")["orig1"].sum()
                out_dep = changed_df.groupby("cat1")["dep1"].sum()
                in_dep = changed_df.groupby("cat2")["dep1"].sum()
                for cat, val in out_orig.items():
                    category_data[cat]["original_reclass"] -= float(val)
                for cat, val in in_orig.items():
                    category_data[cat]["original_reclass"] += float(val)
                for cat, val in out_dep.items():
                    category_data[cat]["depreciation_reclass"] -= float(val)
                for cat, val in in_dep.items():
                    category_data[cat]["depreciation_reclass"] += float(val)

            # 期初余额应包含全部年初卡片；重分类作为单独变动列展示
            # 若剔除重分类卡片，会导致“年初原值”在扩展模式下偏小。
            opening_mask = pd.Series([True] * len(work_df), index=work_df.index)
            opening_df = pd.DataFrame(
                {
                    "cat": cat1_series[opening_mask].where(cat1_series[opening_mask] != "未分类", final_cat_series[opening_mask]),
                    "orig1": orig1[opening_mask],
                    "dep1": dep1[opening_mask],
                }
            )
            if not opening_df.empty:
                open_orig = opening_df.groupby("cat")["orig1"].sum()
                open_dep = opening_df.groupby("cat")["dep1"].sum()
                for cat, val in open_orig.items():
                    category_data[cat]["file1_original_value"] += float(val)
                for cat, val in open_dep.items():
                    category_data[cat]["file1_depreciation"] += float(val)

            # 原值增加按新增方式分拆（优先按年末分类）
            inc_mask = change_type == "原值增加"
            inc_data = {
                "cat": cat2_series.where(cat2_series != "未分类", final_cat_series),
                "amount": orig_change.abs(),
            }
            if has_add_method_mapping and add_method_series is not None:
                inc_data["method"] = add_method_series
            inc_df = pd.DataFrame(inc_data)[inc_mask]
            addition_methods: List[str] = []
            if not inc_df.empty:
                inc_total = inc_df.groupby("cat")["amount"].sum()
                for cat, val in inc_total.items():
                    category_data[cat]["original_increase_total"] += float(val)
                if has_add_method_mapping and "method" in inc_df.columns:
                    addition_methods = self._ordered_unique([self._normalize_text(v, "未标注新增方式") for v in inc_df["method"].tolist()])
                    inc_method = inc_df.groupby(["cat", "method"])["amount"].sum()
                    for (cat, method), val in inc_method.items():
                        m = self._normalize_text(method, "未标注新增方式")
                        category_data[cat][f"original_increase_by_method::{m}"] += float(val)

            # 原值减少按处置方式分拆（优先按年初分类）
            dec_mask = change_type == "原值减少"
            # 总额口径固定对齐处置清单_BKD.原值：abs(原值变动)
            dec_total_base = pd.DataFrame(
                {
                    "cat": cat1_series.where(cat1_series != "未分类", final_cat_series),
                    "amount": orig_change.abs(),
                }
            )[dec_mask]
            dec_amount = orig_change.abs()
            dec_data = {
                "cat": cat1_series.where(cat1_series != "未分类", final_cat_series),
                "amount": dec_amount,
            }
            if has_disp_method_mapping and disp_method_series is not None:
                dec_data["method"] = disp_method_series
            dec_df = pd.DataFrame(dec_data)[dec_mask]
            disposal_methods: List[str] = []
            if not dec_df.empty:
                dec_total = dec_total_base.groupby("cat")["amount"].sum()
                for cat, val in dec_total.items():
                    category_data[cat]["original_decrease_total"] += float(val)
                # 方式分拆口径（统一）：
                # - 映射处置方式：按映射方式分拆，空/0归入“未标注处置方式”
                # - 未映射处置方式（含未上传处置清单）：全部归入“未标注处置方式”
                # - 金额优先用映射金额；未映射/无效时回退主数据金额
                dec_method_df = dec_df.copy()
                if has_disp_method_mapping and "method" in dec_method_df.columns:
                    invalid_method_mask = method_unmarked_mask.loc[dec_method_df.index]
                    dec_method_df.loc[invalid_method_mask, "method"] = "未标注处置方式"
                else:
                    dec_method_df["method"] = "未标注处置方式"

                dec_method_df["amount"] = dec_method_df["amount"].abs()
                if has_disposal_amount_mapping and disp_orig_amount is not None:
                    mapped_amt = disp_orig_amount.loc[dec_method_df.index]
                    mapped_abs = mapped_amt.abs()
                    valid_amt_mask = mapped_amt.notna() & (mapped_abs > 1e-12)
                    dec_method_df.loc[valid_amt_mask, "amount"] = mapped_abs.loc[valid_amt_mask]

                disposal_methods = self._ordered_unique([self._normalize_text(v, "未标注处置方式") for v in dec_method_df["method"].tolist()])
                dec_method = dec_method_df.groupby(["cat", "method"])["amount"].sum()
                for (cat, method), val in dec_method.items():
                    m = self._normalize_text(method, "未标注处置方式")
                    category_data[cat][f"original_decrease_by_method::{m}"] += float(val)
                # 方式拆分与总额不一致时，差异挤入“当期先新增后减少”
                residual_method_key = "当期先新增后减少"
                if residual_method_key not in disposal_methods:
                    disposal_methods.append(residual_method_key)
                for cat, val in dec_total.items():
                    split_sum = 0.0
                    for method in disposal_methods:
                        split_sum += category_data[cat].get(f"original_decrease_by_method::{method}", 0.0)
                    residual = float(val) - float(split_sum)
                    if abs(residual) > 1e-6:
                        category_data[cat][f"original_decrease_by_method::{residual_method_key}"] += residual
            # 未上传处置清单时，确保“未标注处置方式”在汇总表中展示
            if dec_mask.any() and not has_disp_method_mapping:
                if "未标注处置方式" not in disposal_methods:
                    disposal_methods.insert(0, "未标注处置方式")

            # 累计折旧变动净额：保留原口径（按分类对 -累计折旧变动 求和）
            dep_adj_df = pd.DataFrame(
                {
                    "cat": final_cat_series,
                    "amount": -dep_change,
                }
            )
            dep_adj_total = dep_adj_df.groupby("cat")["amount"].sum()
            for cat, val in dep_adj_total.items():
                category_data[cat]["depreciation_adjustment_total"] += float(val)

            # 累计折旧变动净额分拆（优先按处置方式；若存在非处置差额，归入其他方式）
            dep_adj_method_df = pd.DataFrame(
                {
                    "cat": cat1_series.where(cat1_series != "未分类", final_cat_series),
                }
            )[dec_mask]
            if not dep_adj_method_df.empty:
                if has_disp_method_mapping and disp_method_series is not None:
                    dep_adj_method_df["method"] = disp_method_series.loc[dep_adj_method_df.index]
                    invalid_method_mask = method_unmarked_mask.loc[dep_adj_method_df.index]
                    dep_adj_method_df.loc[invalid_method_mask, "method"] = "未标注处置方式"
                else:
                    dep_adj_method_df["method"] = "未标注处置方式"

                # 基准金额（按展示口径使用正数）：
                # 与"处置清单_BKD"口径一致：按"减少原值占年初原值的比例 × 期初累计折旧"分摊
                # - 非原值修改（整张处置）：比例=1，等价于 abs(期初累计折旧)
                # - 原值修改（部分处置）：比例 = abs(原值变动) / abs(年初原值)
                # - 取不到年初原值或为 0 时退回全额（比例=1），避免误算
                idx_for_dep = dep_adj_method_df.index
                full_dep_abs = dep1.loc[idx_for_dep].abs().astype(float)
                orig1_abs = orig1.loc[idx_for_dep].abs().astype(float)
                orig_change_abs = orig_change.loc[idx_for_dep].abs().astype(float)
                ratio = pd.Series(1.0, index=idx_for_dep, dtype="float64")
                positive_orig1 = orig1_abs > 0
                ratio.loc[positive_orig1] = (
                    orig_change_abs.loc[positive_orig1] / orig1_abs.loc[positive_orig1]
                ).clip(upper=1.0)
                base_dep_abs = (full_dep_abs * ratio).astype(float)
                dep_adj_method_df["amount"] = base_dep_abs
                if has_disposal_amount_mapping and disp_dep_amount is not None:
                    dep_amt = disp_dep_amount.loc[dep_adj_method_df.index]
                    dep_abs = dep_amt.abs()
                    valid_dep_mask = dep_amt.notna() & (dep_abs > 1e-12)
                    dep_adj_method_df.loc[valid_dep_mask, "amount"] = dep_abs.loc[valid_dep_mask]

                if not disposal_methods:
                    disposal_methods = self._ordered_unique([self._normalize_text(v, "未标注处置方式") for v in dep_adj_method_df["method"].tolist()])
                dep_adj_method = dep_adj_method_df.groupby(["cat", "method"])["amount"].sum()
                for (cat, method), val in dep_adj_method.items():
                    m = self._normalize_text(method, "未标注处置方式")
                    category_data[cat][f"depreciation_adjustment_by_method::{m}"] += float(val)

            # “其他(非处置变动)”口径：
            # 按每个类别计算：-累计折旧协调金额 - 各处置方式分拆合计
            # 其中累计折旧协调金额对应 depreciation_adjustment_total
            if disposal_methods:
                residual_key = "其他(非处置变动)"
                need_residual_method = False
                for cat, cat_data in category_data.items():
                    split_sum = 0.0
                    for method in disposal_methods:
                        split_sum += cat_data.get(f"depreciation_adjustment_by_method::{method}", 0.0)
                    total = cat_data.get("depreciation_adjustment_total", 0.0)
                    residual = -float(total) - float(split_sum)
                    if abs(residual) > 1e-6:
                        cat_data[f"depreciation_adjustment_by_method::{residual_key}"] += residual
                        need_residual_method = True
                if need_residual_method and residual_key not in disposal_methods:
                    disposal_methods.append(residual_key)

            # 将重分类并入口径：确保各分类期末值与主数据（按年末分类）一致
            target_file2_orig = orig2.groupby(cat2_raw_series).sum()
            target_file2_dep = dep2.groupby(cat2_raw_series).sum()
            # 先形成完整分类集合，保证目标分类都能进入重分类平衡
            category_candidates: List[str] = []
            category_candidates.extend(cat1_series.tolist())
            category_candidates.extend(cat2_series.tolist())
            category_candidates.extend(final_cat_series.tolist())
            category_candidates.extend(list(category_data.keys()))
            categories = [self._normalize_text(c, "未分类") for c in category_candidates if self._normalize_text(c, "未分类")]
            self.categories = self._ordered_unique(categories) or ["未分类"]
            self._reconcile_reclass_to_closing(
                category_data,
                self.categories,
                target_file2_orig,
                target_file2_dep,
            )

            # 汇总分类列表（上面已初始化 self.categories，这里仅兜底）
            if not self.categories:
                self.categories = ["未分类"]

            # 补齐默认键，避免取值缺失
            default_keys = [
                "file1_original_value",
                "original_increase_total",
                "original_decrease_total",
                "original_reclass",
                "file1_depreciation",
                "depreciation_adjustment_total",
                "depreciation_reclass",
            ]
            for cat in self.categories:
                for key in default_keys:
                    _ = category_data[cat][key]
                for method in addition_methods:
                    _ = category_data[cat][f"original_increase_by_method::{method}"]
                for method in disposal_methods:
                    _ = category_data[cat][f"original_decrease_by_method::{method}"]
                    _ = category_data[cat][f"depreciation_adjustment_by_method::{method}"]

            row_defs = self._build_row_defs(
                file1_display_name,
                file2_display_name,
                addition_methods,
                disposal_methods,
            )
            data_dict = {cat: dict(vals) for cat, vals in category_data.items()}
            row_defs = self._filter_zero_rows(row_defs, data_dict, self.categories)
            # 去掉“未分类”但全0的无效展示列
            self.categories = [
                cat for cat in self.categories
                if not (cat == "未分类" and self._is_category_all_zero(cat, row_defs, data_dict))
            ]
            # 蓝色标记应基于“原始年初/年末分类”，不能使用互相回填后的 cat1/cat2。
            # 否则会把仅年末存在的类别误判为两边都存在，导致漏标。
            file1_categories = {
                self._normalize_text(v, "未分类")
                for v in cat1_raw_series.tolist()
                if self._normalize_text(v, "未分类")
            }
            file2_categories = {
                self._normalize_text(v, "未分类")
                for v in cat2_raw_series.tolist()
                if self._normalize_text(v, "未分类")
            }
            file2_only_categories = [
                cat for cat in self.categories
                if cat in file2_categories and cat not in file1_categories
            ]

            summary = {
                "categories": self.categories,
                "file1_display_name": file1_display_name,
                "file2_display_name": file2_display_name,
                "data": data_dict,
                "row_defs": row_defs,
                "addition_methods": addition_methods,
                "disposal_methods": disposal_methods,
                "extended_mode": True,
                "file2_only_categories": file2_only_categories,
            }

            self.summary_data = summary
            return True, f"汇总表生成成功，共 {len(self.categories)} 个资产类别", summary
        except Exception as e:
            return False, f"生成汇总表失败: {str(e)}", None

    def create_summary_worksheet(self, wb: Workbook, sheet_name: str = "固定资产变动汇总表") -> bool:
        """
        在Excel工作簿中创建汇总表worksheet。
        """
        try:
            if not self.summary_data:
                return False

            ws = wb.create_sheet(title=sheet_name)
            categories = self.summary_data["categories"]
            data = self.summary_data["data"]
            row_defs = self.summary_data.get("row_defs", [])

            header_font = Font(bold=True)
            normal_font = Font(bold=False)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            section_fill = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")

            ws.cell(row=1, column=1, value="")
            ws.cell(row=1, column=2, value="")
            for i, cat in enumerate(categories):
                col = i + 3
                cell = ws.cell(row=1, column=col, value=cat)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            ws.cell(row=2, column=1, value="")
            ws.cell(row=2, column=2, value="")
            for i in range(len(categories)):
                col = i + 3
                cell = ws.cell(row=2, column=col, value="账面数")
                cell.alignment = center_align
                cell.border = thin_border

            start_row = 3
            for idx, row_def in enumerate(row_defs):
                row_num = start_row + idx
                ws.cell(row=row_num, column=1, value=row_def["section"])
                ws.cell(row=row_num, column=1).font = header_font
                ws.cell(row=row_num, column=1).fill = section_fill
                ws.cell(row=row_num, column=1).alignment = center_align
                ws.cell(row=row_num, column=1).border = thin_border

                ws.cell(row=row_num, column=2, value=row_def["item"])
                ws.cell(row=row_num, column=2).font = normal_font
                ws.cell(row=row_num, column=2).alignment = left_align
                ws.cell(row=row_num, column=2).border = thin_border

                for i, cat in enumerate(categories):
                    col = i + 3
                    cat_data = data.get(cat, {})
                    if row_def["kind"] == "formula":
                        value = self._calc_formula_value(cat_data, row_def["key"])
                    else:
                        value = cat_data.get(row_def["key"], 0.0)
                    # 展示口径：累计折旧变动净额下“其中-处置方式”按绝对值显示
                    ws.cell(row=row_num, column=col, value=value)
                    ws.cell(row=row_num, column=col).border = thin_border
                    ws.cell(row=row_num, column=col).number_format = "#,##0.00"
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right")

            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 30
            for i in range(len(categories)):
                ws.column_dimensions[get_column_letter(i + 3)].width = 16
            ws.freeze_panes = "C3"

            # 按 section 合并首列单元格
            section_start = None
            last_section = None
            for idx, row_def in enumerate(row_defs):
                row_num = start_row + idx
                section = row_def["section"]
                if section:
                    if last_section is not None and section_start is not None:
                        prev_row = row_num - 1
                        if prev_row > section_start:
                            ws.merge_cells(f"A{section_start}:A{prev_row}")
                    last_section = section
                    section_start = row_num
            if last_section is not None and section_start is not None:
                end_row = start_row + len(row_defs) - 1
                if end_row > section_start:
                    ws.merge_cells(f"A{section_start}:A{end_row}")

            return True
        except Exception as e:
            print(f"创建汇总表worksheet失败: {str(e)}")
            return False

    def get_summary_dataframe(self) -> Optional[pd.DataFrame]:
        """获取汇总表DataFrame（用于预览/CSV导出）。"""
        if not self.summary_data:
            return None

        try:
            categories = self.summary_data["categories"]
            data = self.summary_data["data"]
            row_defs = self.summary_data.get("row_defs", [])
            rows = []

            for row_def in row_defs:
                row = {"大类": row_def["section"], "项目": row_def["item"]}
                for cat in categories:
                    cat_data = data.get(cat, {})
                    if row_def["kind"] == "formula":
                        value = self._calc_formula_value(cat_data, row_def["key"])
                    else:
                        value = cat_data.get(row_def["key"], 0.0)
                    # 与Excel输出保持一致：累计折旧变动净额下“其中-处置方式”按绝对值显示
                    row[cat] = value
                rows.append(row)

            return pd.DataFrame(rows)
        except Exception as e:
            print(f"生成汇总DataFrame失败: {str(e)}")
            return None
