import math
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FA_SHEET_NAME = "FA List"
DISPOSAL_SHEET_NAME = "处置清单_BKD"
DEFAULT_OUTPUT_SUFFIX = "_折旧测算"
RESULT_HEADERS = [
    "月折旧额",
    "本年应计提折旧月份",
    "累计折旧月份",
    "测算的当年折旧",
    "测算的累计折旧",
    "本年折旧",
    "差异_本年折旧",
    "差异_累计折旧",
]

LOG_FILE = Path(__file__).with_name("fa_depreciation_audit.log")


@dataclass
class WorkbookChoice:
    workbook_path: Path
    fa_sheet: str
    balance_sheet_date: date
    fa_field_mapping: dict[str, str]
    disposal_field_mapping: dict[str, str] | None = None


FA_ROLE_LABELS = {
    "asset_category": "资产类别",
    "asset_id": "固定资产编号",
    "asset_name": "固定资产名称",
    "start_date": "入账开始日期",
    "useful_life": "使用寿命(月)",
    "residual_rate": "残值率",
    "original_value": "原值",
    "accumulated_dep": "累计折旧",
    "current_year_dep": "本年折旧",
}

FA_FIELD_KEYWORDS = {
    "asset_category": {
        "exact": ["资产类别", "固定资产类别", "资产大类", "类别", "大类"],
        "contain": ["类别", "分类", "资产类"],
    },
    "asset_id": {
        "exact": ["固定资产编号", "资产编号", "资产编码", "卡片编码", "卡片编号"],
        "contain": ["编号", "编码"],
    },
    "asset_name": {
        "exact": ["固定资产名称", "资产名称", "名称", "资产描述"],
        "contain": ["名称", "描述", "资产名"],
    },
    "start_date": {
        "exact": ["入账开始日期", "入账日期", "开始日期", "购置日期", "取得日期", "启用日期", "资本化日期"],
        "contain": ["日期", "时间"],
    },
    "useful_life": {
        "exact": ["使用寿命(月)", "使用寿命", "预计寿命", "使用年限"],
        "contain": ["寿命", "年限", "计划使用"],
    },
    "residual_rate": {
        "exact": ["残值率", "预计残值率", "净残值率"],
        "contain": ["残值"],
    },
    "original_value": {
        "exact": ["原值", "原值减少", "资产原值", "固定资产原值"],
        "contain": ["成本", "入账价值", "原值减少"],
    },
    "accumulated_dep": {
        "exact": ["累计折旧", "年初累计折旧", "年末累计折旧", "期末累计折旧"],
        "contain": ["累计折旧", "年初累计折旧"],
    },
    "current_year_dep": {
        "exact": ["本年折旧", "年折旧额", "本期折旧"],
        "contain": ["本年折旧"],
    },
}


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def write_log(message: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with LOG_FILE.open("a", encoding="utf-8-sig") as fh:
        fh.write(f"[{timestamp}] {message}\n")


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return 0.0
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return float(text)
    except Exception:
        return 0.0


def normalize_rate(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str) and value.strip().endswith("%"):
        return to_number(value) / 100.0
    rate = to_number(value)
    if rate > 1:
        rate = rate / 100.0
    if rate < 0:
        return 0.0
    return rate


def month_floor(dt: date) -> date:
    return date(dt.year, dt.month, 1)


def add_months(month_start: date, months: int) -> date:
    year = month_start.year + (month_start.month - 1 + months) // 12
    month = (month_start.month - 1 + months) % 12 + 1
    return date(year, month, 1)


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def depreciation_start_month(start_date: date):
    if start_date is None:
        return None
    return add_months(month_floor(start_date), 1)


def count_overlap_months(start_month: date, end_month: date, range_start: date, range_end: date) -> int:
    if not start_month or not end_month or range_end < range_start:
        return 0
    actual_start = max(start_month, range_start)
    actual_end = min(end_month, range_end)
    if actual_end < actual_start:
        return 0
    return (actual_end.year - actual_start.year) * 12 + (actual_end.month - actual_start.month) + 1


def format_number(value: float):
    if value is None:
        return None
    return round(float(value), 2)


def get_sheet_headers(ws):
    return [normalize_header(cell.value) for cell in ws[1]]


def build_header_index(headers):
    result = {}
    for idx, header in enumerate(headers, start=1):
        if header and header not in result:
            result[header] = idx
    return result


def auto_map_column(columns, exact_keywords, contain_keywords=None, excluded_keywords=None):
    contain_keywords = contain_keywords or []
    excluded_keywords = excluded_keywords or []

    def _allowed(col_name: str) -> bool:
        return not any(keyword in col_name for keyword in excluded_keywords)

    for column in columns:
        column_text = str(column)
        if _allowed(column_text) and column_text in exact_keywords:
            return column

    for column in columns:
        column_text = str(column)
        if not _allowed(column_text):
            continue
        if any(keyword in column_text for keyword in exact_keywords):
            return column

    for column in columns:
        column_text = str(column)
        if not _allowed(column_text):
            continue
        if any(keyword in column_text for keyword in contain_keywords):
            return column

    return ""


def suggest_fa_field_mapping(headers):
    mapping = {}
    for role_id, label in FA_ROLE_LABELS.items():
        keywords = FA_FIELD_KEYWORDS[role_id]
        excluded = ["剩余"] if role_id == "useful_life" else []
        mapping[label] = auto_map_column(headers, keywords["exact"], keywords["contain"], excluded)
    return mapping


def normalize_key_value(value) -> str:
    text = normalize_header(value)
    if not text:
        return ""
    return text.lower().replace(" ", "")


def build_composite_key_from_values(values):
    parts = [normalize_key_value(value) for value in values]
    if not any(parts):
        return None
    return tuple(parts)


class BalanceDateDialog:
    def __init__(self, workbook_path: Path, fa_sheet: str, fa_field_mapping: dict[str, str], disposal_field_mapping: dict[str, str] | None = None):
        self.workbook_path = workbook_path
        self.fa_sheet = fa_sheet
        self.fa_field_mapping = fa_field_mapping
        self.disposal_field_mapping = disposal_field_mapping
        self.result = None

        self.window = tk.Tk()
        self.window.title("折旧测算参数确认")
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.window.attributes("-topmost", True)

        self.file_var = tk.StringVar(value=self.workbook_path.name)
        self.fa_sheet_var = tk.StringVar(value=self.fa_sheet)
        self.bs_date_var = tk.StringVar(value=f"{date.today().year}-12-31")

        frame = ttk.Frame(self.window, padding=12)
        frame.grid(row=0, column=0, sticky="nsew")
        ttk.Label(frame, text="目标文件").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(frame, textvariable=self.file_var, state="readonly", width=44).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(frame, text="FA List Sheet").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(frame, textvariable=self.fa_sheet_var, state="readonly", width=44).grid(row=1, column=1, sticky="ew", pady=4)
        ttk.Label(frame, text="资产负债表日").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(frame, textvariable=self.bs_date_var, width=24).grid(row=2, column=1, sticky="w", pady=4)
        ttk.Label(frame, text="本工具将直接使用 FA List 中已有的“本年折旧”列。", foreground="#666666").grid(row=3, column=0, columnspan=2, sticky="w", pady=(4, 8))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, sticky="e")
        ttk.Button(btn_frame, text="确定", command=self.on_ok).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btn_frame, text="取消", command=self.on_cancel).grid(row=0, column=1)

        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x_pos = (self.window.winfo_screenwidth() - width) // 2
        y_pos = (self.window.winfo_screenheight() - height) // 2
        self.window.geometry(f"+{x_pos}+{y_pos}")
        self.window.lift()
        self.window.focus_force()
        self.window.after(300, lambda: self.window.attributes("-topmost", False))

    def show(self):
        self.window.mainloop()
        return self.result

    def on_ok(self):
        try:
            bs_date = datetime.strptime(self.bs_date_var.get().strip(), "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("日期错误", "资产负债表日请按 YYYY-MM-DD 填写。", parent=self.window)
            return
        self.result = WorkbookChoice(
            workbook_path=self.workbook_path,
            fa_sheet=self.fa_sheet,
            balance_sheet_date=bs_date,
            fa_field_mapping=self.fa_field_mapping,
            disposal_field_mapping=self.disposal_field_mapping,
        )
        self.close()

    def on_cancel(self):
        self.result = None
        self.close()

    def close(self):
        self.window.destroy()


class FieldMappingDialog:
    def __init__(self, workbook_path: Path, sheet_name: str, dialog_title: str):
        self.workbook_path = workbook_path
        self.default_sheet_name = sheet_name
        self.result = None
        self.window = tk.Tk()
        self.window.title(dialog_title)
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.window.attributes("-topmost", True)

        self.sheet_var = tk.StringVar(value=sheet_name)
        self.mapping_vars = {label: tk.StringVar() for label in FA_ROLE_LABELS.values()}
        self.mapping_combos = {}
        self.workbook = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)

        self._build()
        self._init_defaults()
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x_pos = (self.window.winfo_screenwidth() - width) // 2
        y_pos = (self.window.winfo_screenheight() - height) // 2
        self.window.geometry(f"+{x_pos}+{y_pos}")
        self.window.lift()
        self.window.focus_force()
        self.window.after(300, lambda: self.window.attributes("-topmost", False))

    def _build(self):
        frame = ttk.Frame(self.window, padding=12)
        frame.grid(row=0, column=0, sticky="nsew")

        ttk.Label(frame, text="目标 Sheet").grid(row=0, column=0, sticky="w", pady=4)
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.sheet_var, state="readonly", width=42)
        self.sheet_combo.grid(row=0, column=1, sticky="ew", pady=4)
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda _e: self._refresh_mapping())

        for row_idx, label in enumerate(FA_ROLE_LABELS.values(), start=1):
            ttk.Label(frame, text=label).grid(row=row_idx, column=0, sticky="w", pady=4)
            combo = ttk.Combobox(frame, textvariable=self.mapping_vars[label], state="readonly", width=42)
            combo.grid(row=row_idx, column=1, sticky="ew", pady=4)
            self.mapping_combos[label] = combo

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=len(FA_ROLE_LABELS) + 1, column=0, columnspan=2, sticky="e", pady=(8, 0))
        ttk.Button(btn_frame, text="确定", command=self.on_ok).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btn_frame, text="取消", command=self.on_cancel).grid(row=0, column=1)

    def _init_defaults(self):
        self.sheet_combo["values"] = self.workbook.sheetnames
        if self.default_sheet_name in self.workbook.sheetnames:
            self.sheet_var.set(self.default_sheet_name)
        elif self.workbook.sheetnames:
            self.sheet_var.set(self.workbook.sheetnames[0])
        self._refresh_mapping()

    def _refresh_mapping(self):
        sheet_name = self.sheet_var.get()
        headers = [header for header in get_sheet_headers(self.workbook[sheet_name]) if header]
        suggestions = suggest_fa_field_mapping(headers)
        for label, combo in self.mapping_combos.items():
            combo["values"] = headers
            current_value = self.mapping_vars[label].get()
            suggested = suggestions.get(label, "")
            if current_value in headers:
                continue
            self.mapping_vars[label].set(suggested or (headers[0] if headers else ""))

    def show(self):
        self.window.mainloop()
        return self.result

    def on_ok(self):
        mapping = {label: var.get() for label, var in self.mapping_vars.items()}
        required = ["固定资产编号", "入账开始日期", "使用寿命(月)", "残值率", "原值", "累计折旧", "本年折旧"]
        missing = [label for label in required if not mapping.get(label)]
        if missing:
            messagebox.showerror("映射不完整", f"以下字段需要确认：{', '.join(missing)}", parent=self.window)
            return
        self.result = (self.sheet_var.get(), mapping)
        self.close()

    def on_cancel(self):
        self.result = None
        self.close()

    def close(self):
        try:
            self.workbook.close()
        except Exception:
            pass
        self.window.destroy()


def choose_workbook(root, initial_dir: Path):
    selected = filedialog.askopenfilename(
        parent=root,
        title="请选择目标文件",
        initialdir=str(initial_dir),
        filetypes=[("Excel 文件", "*.xlsx;*.xlsm"), ("所有文件", "*.*")],
    )
    if not selected:
        return None
    return Path(selected)


def show_error(title: str, text: str):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(title, text, parent=root)
    root.destroy()


def show_info(title: str, text: str):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, text, parent=root)
    root.destroy()


def load_sheet_as_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalize_header(value) for value in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def build_lookup(data_ws, id_columns: list[str], value_column: str):
    headers, rows = load_sheet_as_rows(data_ws)
    header_index = build_header_index(headers)
    id_indexes = [header_index[column] - 1 for column in id_columns]
    value_idx = header_index[value_column] - 1
    lookup = {}
    for row in rows:
        key_values = [row[idx] if idx < len(row) else "" for idx in id_indexes]
        asset_key = build_composite_key_from_values(key_values)
        if not asset_key or asset_key in lookup:
            continue
        lookup[asset_key] = row[value_idx] if value_idx < len(row) else None
    return lookup


def compute_row_values(start_date_value, life_value, residual_value, original_value, accumulated_value, q_value, balance_sheet_date: date):
    start_date = parse_date(start_date_value)
    life_months = int(to_number(life_value)) if to_number(life_value) > 0 else 0
    residual_rate = normalize_rate(residual_value)
    original_cost = to_number(original_value)
    accumulated_dep = to_number(accumulated_value)
    has_q_value = normalize_header(q_value) != ""
    current_year_dep = to_number(q_value) if has_q_value else None

    if not start_date or life_months <= 0:
        delta_current = (current_year_dep - 0.0) if current_year_dep is not None else None
        return [None, 0, 0, 0.0, 0.0, format_number(current_year_dep) if current_year_dep is not None else None, format_number(delta_current) if delta_current is not None else None, format_number(accumulated_dep)]

    dep_start = depreciation_start_month(start_date)
    dep_end = add_months(dep_start, life_months - 1)
    bs_month = month_floor(balance_sheet_date)
    year_start = date(balance_sheet_date.year, 1, 1)

    monthly_dep = 0.0
    depreciable_base = max(original_cost * (1 - residual_rate), 0.0)
    if life_months > 0:
        monthly_dep = depreciable_base / life_months

    current_year_months = count_overlap_months(dep_start, dep_end, year_start, bs_month)
    accumulated_months = count_overlap_months(dep_start, dep_end, dep_start, bs_month)
    estimated_current = min(monthly_dep * current_year_months, depreciable_base)
    estimated_accumulated = min(monthly_dep * accumulated_months, depreciable_base)
    delta_current = (current_year_dep - estimated_current) if current_year_dep is not None else None
    delta_accumulated = accumulated_dep - estimated_accumulated

    return [
        format_number(monthly_dep),
        current_year_months,
        accumulated_months,
        format_number(estimated_current),
        format_number(estimated_accumulated),
        format_number(current_year_dep) if current_year_dep is not None else None,
        format_number(delta_current) if delta_current is not None else None,
        format_number(delta_accumulated),
    ]


def process_workbook(choice: WorkbookChoice):
    write_log(f"process_workbook start: {choice.workbook_path.name}")
    wb = load_workbook(choice.workbook_path, keep_links=False)
    try:
        def append_measurement_block(ws, source_mapping: dict[str, str], sheet_label: str, cutoff_date_col: str | None = None):
            headers = get_sheet_headers(ws)
            header_index = build_header_index(headers)
            required_columns = {
                "入账开始日期": source_mapping["入账开始日期"],
                "使用寿命(月)": source_mapping["使用寿命(月)"],
                "残值率": source_mapping["残值率"],
                "原值": source_mapping["原值"],
                "累计折旧": source_mapping["累计折旧"],
                "本年折旧": source_mapping["本年折旧"],
            }
            missing = [label for label, column_name in required_columns.items() if column_name not in header_index]
            if missing:
                write_log(f"{sheet_label} skipped, missing columns: {missing}")
                return False

            start_column = ws.max_column + 1
            for offset, header in enumerate(RESULT_HEADERS):
                ws.cell(row=1, column=start_column + offset, value=header)

            bs_date_literal = choice.balance_sheet_date.strftime("%Y-%m-%d")
            start_date_col_letter = get_column_letter(header_index[source_mapping["入账开始日期"]])
            life_col_letter = get_column_letter(header_index[source_mapping["使用寿命(月)"]])
            residual_col_letter = get_column_letter(header_index[source_mapping["残值率"]])
            original_col_letter = get_column_letter(header_index[source_mapping["原值"]])
            accumulated_col_letter = get_column_letter(header_index[source_mapping["累计折旧"]])
            current_year_dep_source_letter = get_column_letter(header_index[source_mapping["本年折旧"]])
            cutoff_col_letter = get_column_letter(header_index[cutoff_date_col]) if cutoff_date_col and cutoff_date_col in header_index else None

            monthly_col = get_column_letter(start_column)
            current_months_col = get_column_letter(start_column + 1)
            accumulated_months_col = get_column_letter(start_column + 2)
            estimated_current_col = get_column_letter(start_column + 3)
            estimated_accumulated_col = get_column_letter(start_column + 4)
            current_year_dep_col = get_column_letter(start_column + 5)
            diff_current_col = get_column_letter(start_column + 6)
            diff_accumulated_col = get_column_letter(start_column + 7)

            def row_formula(row_idx: int, kind: str) -> str:
                row_ref = str(row_idx)
                d = f"{start_date_col_letter}{row_ref}"
                e = f"{life_col_letter}{row_ref}"
                f = f"{residual_col_letter}{row_ref}"
                h = f"{accumulated_col_letter}{row_ref}"
                rate_expr = f'IF({f}="",0,IF({f}>1,{f}/100,{f}))'
                dep_start_expr = f"EDATE(DATE(YEAR({d}),MONTH({d}),1),1)"
                dep_end_expr = f"EDATE({dep_start_expr},{e}-1)"
                bs_expr = f'DATEVALUE("{bs_date_literal}")'
                bs_month_expr = f"DATE(YEAR({bs_expr}),MONTH({bs_expr}),1)"
                if cutoff_col_letter:
                    cutoff_expr = f"{cutoff_col_letter}{row_ref}"
                    effective_month_expr = (
                        f'IF(OR({cutoff_expr}="",ISBLANK({cutoff_expr})),{bs_month_expr},'
                        f'MIN({bs_month_expr},DATE(YEAR({cutoff_expr}),MONTH({cutoff_expr}),1)))'
                    )
                else:
                    effective_month_expr = bs_month_expr
                year_start_expr = f"DATE(YEAR({bs_expr}),1,1)"
                month_diff_current = (
                    f"(YEAR(MIN({dep_end_expr},{effective_month_expr}))-YEAR(MAX({dep_start_expr},{year_start_expr})))*12+"
                    f"MONTH(MIN({dep_end_expr},{effective_month_expr}))-MONTH(MAX({dep_start_expr},{year_start_expr}))+1"
                )
                month_diff_acc = (
                    f"(YEAR(MIN({dep_end_expr},{effective_month_expr}))-YEAR({dep_start_expr}))*12+"
                    f"MONTH(MIN({dep_end_expr},{effective_month_expr}))-MONTH({dep_start_expr})+1"
                )
                monthly_formula = f'IFERROR(ROUND({original_col_letter}{row_ref}*(1-{rate_expr})/{e},2),"")'
                current_months_formula = (
                    f'IF(OR({d}="",{e}<=0),0,MAX(0,IF(MIN({dep_end_expr},{effective_month_expr})<MAX({dep_start_expr},{year_start_expr}),0,{month_diff_current})))'
                )
                accumulated_months_formula = (
                    f'IF(OR({d}="",{e}<=0),0,MAX(0,IF(MIN({dep_end_expr},{effective_month_expr})<{dep_start_expr},0,{month_diff_acc})))'
                )
                if kind == "monthly":
                    return f"={monthly_formula}"
                if kind == "current_months":
                    return f"={current_months_formula}"
                if kind == "accumulated_months":
                    return f"={accumulated_months_formula}"
                if kind == "estimated_current":
                    return f'=IF(OR(LEN({monthly_col}{row_ref})=0,LEN({current_months_col}{row_ref})=0),"",ROUND({monthly_col}{row_ref}*{current_months_col}{row_ref},2))'
                if kind == "estimated_accumulated":
                    return f'=IF(OR(LEN({monthly_col}{row_ref})=0,LEN({accumulated_months_col}{row_ref})=0),"",ROUND({monthly_col}{row_ref}*{accumulated_months_col}{row_ref},2))'
                if kind == "current_year_dep":
                    return f"={current_year_dep_source_letter}{row_ref}"
                if kind == "diff_current":
                    return f"={current_year_dep_col}{row_ref}-{estimated_current_col}{row_ref}"
                if kind == "diff_accumulated":
                    return f"={h}-{estimated_accumulated_col}{row_ref}"
                return ""

            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=start_column + 0, value=row_formula(row_idx, "monthly"))
                ws.cell(row=row_idx, column=start_column + 1, value=row_formula(row_idx, "current_months"))
                ws.cell(row=row_idx, column=start_column + 2, value=row_formula(row_idx, "accumulated_months"))
                ws.cell(row=row_idx, column=start_column + 3, value=row_formula(row_idx, "estimated_current"))
                ws.cell(row=row_idx, column=start_column + 4, value=row_formula(row_idx, "estimated_accumulated"))
                ws.cell(row=row_idx, column=start_column + 5, value=row_formula(row_idx, "current_year_dep"))
                ws.cell(row=row_idx, column=start_column + 6, value=row_formula(row_idx, "diff_current"))
                ws.cell(row=row_idx, column=start_column + 7, value=row_formula(row_idx, "diff_accumulated"))

            for col_idx in (start_column, start_column + 3, start_column + 4, start_column + 5, start_column + 6, start_column + 7):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'
            write_log(f"{sheet_label} measurement appended from col {get_column_letter(start_column)}")
            return True

        fa_mapping = {
            "入账开始日期": choice.fa_field_mapping["入账开始日期"],
            "使用寿命(月)": choice.fa_field_mapping["使用寿命(月)"],
            "残值率": choice.fa_field_mapping["残值率"],
            "原值": choice.fa_field_mapping["原值"],
            "累计折旧": choice.fa_field_mapping["累计折旧"],
            "本年折旧": choice.fa_field_mapping["本年折旧"],
        }
        fa_ws = wb[choice.fa_sheet]
        append_measurement_block(fa_ws, fa_mapping, choice.fa_sheet)

        if DISPOSAL_SHEET_NAME in wb.sheetnames:
            disposal_ws = wb[DISPOSAL_SHEET_NAME]
            disposal_mapping = choice.disposal_field_mapping or {
                "入账开始日期": "入账开始日期",
                "使用寿命(月)": "使用寿命(月)",
                "残值率": "残值率",
                "原值": "原值减少",
                "累计折旧": "年初累计折旧",
                "本年折旧": "本年折旧",
            }
            append_measurement_block(disposal_ws, disposal_mapping, DISPOSAL_SHEET_NAME, cutoff_date_col="处置时间")

        output_path = choice.workbook_path.with_name(f"{choice.workbook_path.stem}{DEFAULT_OUTPUT_SUFFIX}{choice.workbook_path.suffix}")
        wb.save(output_path)
        write_log(f"process_workbook saved: {output_path.name}")
        return output_path
    finally:
        wb.close()


def main():
    write_log("main start")
    root = tk.Tk()
    root.withdraw()

    base_dir = Path(__file__).resolve().parent.parent
    workbook_path = choose_workbook(root, base_dir)
    if not workbook_path:
        write_log("main cancelled at file choose")
        root.destroy()
        return
    write_log(f"main selected workbook: {workbook_path}")
    if workbook_path.name.startswith("~$"):
        root.destroy()
        show_error("文件错误", "不能选择 Excel 临时锁定文件，请重新选择正式文件。")
        return
    try:
        root.config(cursor="watch")
        root.update_idletasks()
        wb = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)
        has_fa_sheet = FA_SHEET_NAME in wb.sheetnames
        wb.close()
        write_log(f"main workbook opened, has_fa_sheet={has_fa_sheet}")
    except Exception as exc:
        root.config(cursor="")
        root.destroy()
        write_log(f"main workbook open error: {exc}")
        show_error("打开失败", f"目标文件无法读取：{exc}")
        return
    finally:
        root.config(cursor="")
        root.destroy()
    if not has_fa_sheet:
        show_error("文件错误", "所选文件中未找到 FA List sheet，请重新选择。")
        return

    try:
        write_log("main creating FA FieldMappingDialog")
        mapping_dialog = FieldMappingDialog(workbook_path, FA_SHEET_NAME, "FA List 字段映射确认")
    except Exception as exc:
        write_log(f"main FA FieldMappingDialog error: {exc}")
        show_error("加载失败", f"初始化 FA List 映射窗口失败：{exc}")
        return
    field_mapping_result = mapping_dialog.show()
    if field_mapping_result is None:
        write_log("main cancelled at FA field mapping dialog")
        return

    fa_sheet, fa_field_mapping = field_mapping_result
    disposal_field_mapping = None

    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)
        has_disposal_sheet = DISPOSAL_SHEET_NAME in wb.sheetnames
        wb.close()
    except Exception:
        has_disposal_sheet = False

    if has_disposal_sheet:
        try:
            write_log("main creating Disposal FieldMappingDialog")
            disposal_dialog = FieldMappingDialog(workbook_path, DISPOSAL_SHEET_NAME, "处置清单_BKD 字段映射确认")
        except Exception as exc:
            write_log(f"main Disposal FieldMappingDialog error: {exc}")
            show_error("加载失败", f"初始化处置清单映射窗口失败：{exc}")
            return
        disposal_mapping_result = disposal_dialog.show()
        if disposal_mapping_result is None:
            write_log("main cancelled at disposal field mapping dialog")
            return
        _disposal_sheet, disposal_field_mapping = disposal_mapping_result

    try:
        write_log("main creating BalanceDateDialog")
        dialog = BalanceDateDialog(workbook_path, fa_sheet, fa_field_mapping, disposal_field_mapping)
    except Exception as exc:
        write_log(f"main BalanceDateDialog error: {exc}")
        show_error("加载失败", f"初始化参数窗口失败：{exc}")
        return
    choice = dialog.show()
    if choice is None:
        write_log("main cancelled at dialog")
        return

    try:
        output_path = process_workbook(choice)
    except Exception as exc:
        write_log(f"main process error: {exc}")
        show_error("处理失败", str(exc))
        return

    write_log(f"main completed: {output_path}")
    show_info(
        "处理完成",
        f"已生成测算结果文件：\n{output_path}",
    )


if __name__ == "__main__":
    main()
