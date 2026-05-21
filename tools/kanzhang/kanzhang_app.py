import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import os
import threading
import gc
import tempfile
import time
import math
import numpy as np 
import webbrowser
import re
import csv
from datetime import datetime
import json
from contextlib import contextmanager

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ==========================================
# 0. 库检测
# ==========================================
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

try:
    import python_calamine
    EXCEL_ENGINE = 'calamine'
except ImportError:
    EXCEL_ENGINE = None

try:
    import polars as pl
    HAS_POLARS = True
except ImportError:
    pl = None
    HAS_POLARS = False


class ExportCancelled(Exception):
    pass


def _fit_toplevel_to_screen(win, width, height, min_width=None, min_height=None, margin_x=80, margin_y=120):
    try:
        screen_w = win.winfo_screenwidth()
        screen_h = win.winfo_screenheight()
        actual_w = min(width, max(320, screen_w - margin_x))
        actual_h = min(height, max(240, screen_h - margin_y))
        pos_x = max(20, (screen_w - actual_w) // 2)
        pos_y = max(20, (screen_h - actual_h) // 2)
        win.geometry(f"{actual_w}x{actual_h}+{pos_x}+{pos_y}")
        if min_width and min_height:
            win.minsize(min(min_width, actual_w), min(min_height, actual_h))
    except Exception:
        win.geometry(f"{width}x{height}")


class ExportPerfTracer:
    def __init__(self, log_path, context=None):
        self.log_path = log_path
        self.context = context or {}
        self.events = []
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        self._write("SESSION_START", {"session_id": self.session_id, **self.context})

    def _write(self, typ, payload):
        rec = {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": typ,
            "session_id": self.session_id,
            **payload,
        }
        line = json.dumps(rec, ensure_ascii=False)
        try:
            with open(self.log_path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def event(self, name, **fields):
        rec = {"name": name, **fields}
        self.events.append(rec)
        self._write("EVENT", rec)

    @contextmanager
    def step(self, name, **fields):
        t0 = time.perf_counter()
        self.event(f"{name}::start", **fields)
        try:
            yield
        finally:
            self.event(f"{name}::end", elapsed_s=round(time.perf_counter() - t0, 6), **fields)

    def close(self):
        slow = [e for e in self.events if "elapsed_s" in e]
        slow.sort(key=lambda x: x.get("elapsed_s", 0), reverse=True)
        self._write("SESSION_TOP", {"top_slowest": slow[:20]})
        self._write("SESSION_END", {"event_count": len(self.events)})

# ==========================================
# 1. 进度弹窗
# ==========================================
class ProgressWindow(tk.Toplevel):
    """导出/处理进度提示，固定足够高度避免按钮与进度条重叠。"""

    _BG = "#f4f6f8"
    _TITLE_FG = "#1b2a33"
    _MSG_FG = "#4a5a63"
    _FONT = ("Microsoft YaHei UI", 10)
    _TITLE_FONT = ("Microsoft YaHei UI", 11, "bold")

    def __init__(self, parent, title="正在处理", message=None, on_cancel=None, cancellable=False):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.on_cancel = on_cancel
        self.cancellable = cancellable
        self.configure(bg=self._BG)

        display_msg = (message or title or "正在处理数据，请稍候...").strip()
        win_h = 188 if cancellable else 148
        _fit_toplevel_to_screen(self, 440, win_h, min_width=400, min_height=win_h)

        outer = tk.Frame(self, bg=self._BG, padx=22, pady=18)
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(0, weight=1)

        tk.Label(
            outer,
            text=title,
            bg=self._BG,
            fg=self._TITLE_FG,
            font=self._TITLE_FONT,
            anchor="w",
            justify="left",
            wraplength=380,
        ).grid(row=0, column=0, sticky="ew")

        self.lbl = tk.Label(
            outer,
            text=display_msg,
            bg=self._BG,
            fg=self._MSG_FG,
            font=self._FONT,
            anchor="w",
            justify="left",
            wraplength=380,
        )
        self.lbl.grid(row=1, column=0, sticky="ew", pady=(8, 12))

        self.pb = ttk.Progressbar(outer, orient="horizontal", length=360, mode="indeterminate")
        self.pb.grid(row=2, column=0, sticky="ew", pady=(0, 14))
        self.pb.start(12)

        if self.cancellable:
            btn_row = tk.Frame(outer, bg=self._BG)
            btn_row.grid(row=3, column=0, sticky="e")
            ttk.Button(btn_row, text="终止导出", width=14, command=self._on_cancel).pack(side=tk.RIGHT)
            self.protocol("WM_DELETE_WINDOW", self._on_cancel)
            self.bind("<Escape>", lambda e: self._on_cancel())
        else:
            self.protocol("WM_DELETE_WINDOW", lambda: None)

        self.update_idletasks()
        try:
            px = parent.winfo_rootx() + max(0, (parent.winfo_width() - self.winfo_width()) // 2)
            py = parent.winfo_rooty() + max(0, (parent.winfo_height() - self.winfo_height()) // 2)
            self.geometry(f"+{px}+{py}")
        except Exception:
            pass

    def set_message(self, text: str) -> None:
        try:
            if self.lbl.winfo_exists():
                self.lbl.config(text=str(text))
        except Exception:
            pass

    def _on_cancel(self):
        if callable(self.on_cancel):
            self.on_cancel()
        else:
            self.close()

    def close(self):
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()

# ==========================================
# 2. 支持拖拽的列表框
# ==========================================
class DraggableListbox(tk.Listbox):
    def __init__(self, master, app_ref=None, dialog_ref=None, box_type=None, **kw):
        super().__init__(master, **kw)
        self.app = app_ref        
        self.dialog = dialog_ref  
        self.box_type = box_type  
        self._selection_anchor = None

        self.bind('<Button-1>', self.on_click)
        self.bind('<Shift-Button-1>', self.on_click)
        self.bind('<Control-Button-1>', self.on_click)
        self.bind('<B1-Motion>', self.on_drag)
        self.bind('<Shift-B1-Motion>', self.on_drag)
        self.bind('<Control-B1-Motion>', self.on_drag)
        self.bind('<ButtonRelease-1>', self.on_drop)
        self.bind('<Control-a>', self._select_all)
        self.bind('<Control-A>', self._select_all)
        self.cur_selection_values = []

    def on_click(self, event):
        idx = self.nearest(event.y)
        if idx is None:
            return "break"

        # 自定义选择逻辑，避免按住拖动时触发列表的批量复选
        is_shift = (event.state & 0x0001) != 0
        is_ctrl = (event.state & 0x0004) != 0

        current_sel = self.curselection()
        if is_shift:
            if self._selection_anchor is None:
                if current_sel:
                    self._selection_anchor = current_sel[0]
                else:
                    self._selection_anchor = idx
            start = min(self._selection_anchor, idx)
            end = max(self._selection_anchor, idx)
            self.selection_clear(0, tk.END)
            self.selection_set(start, end)
        elif is_ctrl:
            if idx in self.curselection():
                self.selection_clear(idx)
            else:
                self.selection_set(idx)
            self._selection_anchor = idx
        else:
            # 若已有多选并点击到已选项，则保留多选以便拖拽
            if idx not in current_sel or len(current_sel) <= 1:
                self.selection_clear(0, tk.END)
                self.selection_set(idx)
            self._selection_anchor = idx

        self.activate(idx)
        self.cur_selection_indices = self.curselection()
        self.cur_selection_values = [self.get(i) for i in self.cur_selection_indices]
        return "break"

    def on_drag(self, event):
        self.config(cursor="hand2")
        # 阻止 Listbox 默认的拖动多选行为
        if self.dialog or self.app:
            return "break"

    def on_drop(self, event):
        self.config(cursor="")
        x, y = self.winfo_pointerxy()
        
        # 场景1: 透视表弹窗拖拽
        if self.dialog:
            target_lb = None
            for lb in [self.dialog.lb_src, self.dialog.lb_rows, self.dialog.lb_cols, self.dialog.lb_vals]:
                wx, wy = lb.winfo_rootx(), lb.winfo_rooty()
                ww, wh = lb.winfo_width(), lb.winfo_height()
                if wx <= x <= wx+ww and wy <= y <= wy+wh:
                    target_lb = lb
                    break
            
            if target_lb and target_lb != self:
                for val in self.cur_selection_values:
                    # 简单逻辑：如果是从源拖出则添加，否则移动
                    if self == self.dialog.lb_src:
                        if val not in target_lb.get(0, tk.END): target_lb.insert(tk.END, val)
                    elif target_lb == self.dialog.lb_src:
                        idx_opts = [i for i, v in enumerate(self.get(0, tk.END)) if v == val]
                        for i in reversed(idx_opts): self.delete(i)
                    else:
                        if val not in target_lb.get(0, tk.END): target_lb.insert(tk.END, val)
                        idx_opts = [i for i, v in enumerate(self.get(0, tk.END)) if v == val]
                        for i in reversed(idx_opts): self.delete(i)
            return

        # 场景2: 主界面筛选框拖拽
        if self.app:
            target_type = None
            widgets = {
                'source': self.app.shuttle_list_left,
                'target': self.app.shuttle_list_right,
                'exclude': self.app.shuttle_list_exclude
            }
            for w_type, widget in widgets.items():
                if not widget: continue
                wx, wy = widget.winfo_rootx(), widget.winfo_rooty()
                ww, wh = widget.winfo_width(), widget.winfo_height()
                if wx <= x <= wx+ww and wy <= y <= wy+wh:
                    target_type = w_type
                    break
            
            if target_type and target_type != self.box_type and self.cur_selection_values:
                self.app.handle_drag_drop(self.cur_selection_values, target_type)

    def _select_all(self, event=None):
        try:
            self.selection_set(0, tk.END)
        except Exception:
            pass
        return "break"

# ==========================================
# 3. 透视表设计器
# ==========================================
class PivotDesignerDialog:
    def __init__(self, parent, all_columns, predefined_calc_cols, defaults):
        self.top = tk.Toplevel(parent)
        self.top.title("📊 透视表配置 (支持拖拽)")
        _fit_toplevel_to_screen(self.top, 850, 600, min_width=720, min_height=500)
        self.top.transient(parent)
        self.top.grab_set()
        self.result = None
        self.action = "cancel" 
        self.all_cols = all_columns + predefined_calc_cols
        self.top.bind("<Escape>", lambda e: self.on_cancel())
        self.top.lift()
        self.top.after(10, lambda: self.top.focus_force())
        self._bind_ctrl_a()

        btn_frame = tk.Frame(self.top, pady=12, padx=16, bg="#eef2f5")
        btn_frame.pack(side="bottom", fill="x")
        ttk.Button(btn_frame, text="正常导出（含套表）", width=18, command=self.on_confirm).pack(side="right", padx=(10, 0))
        ttk.Button(btn_frame, text="快速导出（仅明细）", width=18, command=self.on_skip).pack(side="right")
        ttk.Button(btn_frame, text="取消", width=10, command=self.on_cancel).pack(side="left")

        tip_bar = tk.Frame(self.top, bg="#e8f4f8", padx=12, pady=8)
        tip_bar.pack(fill="x")
        tk.Label(
            tip_bar,
            text="提示：可将左侧字段拖拽到行、列、值区域；行与列不能同时为空，至少需要一个数值字段。",
            bg="#e8f4f8",
            fg="#205860",
            font=("Microsoft YaHei UI", 9),
            anchor="w",
            justify="left",
            wraplength=780,
        ).pack(fill="x")
        main_frame = tk.Frame(self.top, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)

        frame_src = tk.LabelFrame(main_frame, text="可用字段 (源)", padx=5, pady=5)
        frame_src.pack(side="left", fill="both", expand=True, padx=5)
        self.lb_src = DraggableListbox(frame_src, dialog_ref=self, selectmode="single", bg="#f9f9f9")
        self.lb_src.pack(side="left", fill="both", expand=True)
        sb_src = ttk.Scrollbar(frame_src, orient="vertical", command=self.lb_src.yview)
        sb_src.pack(side="right", fill="y")
        self.lb_src.config(yscrollcommand=sb_src.set)
        for c in self.all_cols: self.lb_src.insert(tk.END, c)

        frame_mid = tk.Frame(main_frame, padx=5)
        frame_mid.pack(side="left", fill="y")
        ttk.Button(frame_mid, text="行 (Rows) >", command=lambda: self.add_item(self.lb_rows)).pack(fill="x", pady=4)
        ttk.Button(frame_mid, text="列 (Cols) >", command=lambda: self.add_item(self.lb_cols)).pack(fill="x", pady=4)
        ttk.Button(frame_mid, text="值 (Values) >", command=lambda: self.add_item(self.lb_vals)).pack(fill="x", pady=4)
        ttk.Separator(frame_mid, orient="horizontal").pack(fill="x", pady=30)
        ttk.Button(frame_mid, text="❌ 移除", command=self.remove_item).pack(fill="x", pady=4)

        frame_dest = tk.Frame(main_frame)
        frame_dest.pack(side="right", fill="both", expand=True, padx=5)
        
        tk.Label(frame_dest, text="1. 行字段 (Index)", fg="#1565C0").pack(anchor="w")
        self.lb_rows = DraggableListbox(frame_dest, dialog_ref=self, height=6, bg="white"); self.lb_rows.pack(fill="x", expand=True)
        
        tk.Label(frame_dest, text="2. 列字段 (Cols)", fg="#2E7D32").pack(anchor="w")
        self.lb_cols = DraggableListbox(frame_dest, dialog_ref=self, height=5, bg="white"); self.lb_cols.pack(fill="x", expand=True)
        
        tk.Label(frame_dest, text="3. 数值字段 (Values)", fg="#C62828").pack(anchor="w")
        self.lb_vals = DraggableListbox(frame_dest, dialog_ref=self, height=5, bg="white"); self.lb_vals.pack(fill="x", expand=True)

        if defaults:
            for item in defaults.get('rows', []): self.lb_rows.insert(tk.END, item)
            for item in defaults.get('cols', []): self.lb_cols.insert(tk.END, item)
            for item in defaults.get('vals', []): self.lb_vals.insert(tk.END, item)

    def _bind_ctrl_a(self):
        def _select_all(event=None):
            w = self.top.focus_get()
            if isinstance(w, tk.Listbox):
                w.selection_set(0, tk.END)
                return "break"
            if isinstance(w, ttk.Treeview):
                for item in w.get_children():
                    w.selection_add(item)
                return "break"
            if isinstance(w, (tk.Entry, ttk.Entry)):
                w.select_range(0, tk.END)
                w.icursor(tk.END)
                return "break"
            if isinstance(w, tk.Text):
                w.tag_add(tk.SEL, "1.0", tk.END)
                w.mark_set(tk.INSERT, tk.END)
                return "break"
            return "break"
        self.top.bind("<Control-a>", _select_all)
        self.top.bind("<Control-A>", _select_all)

    def add_item(self, target_lb):
        sels = self.lb_src.curselection()
        for i in sels: 
            val = self.lb_src.get(i)
            if val not in target_lb.get(0, tk.END): target_lb.insert(tk.END, val)
    def remove_item(self):
        for t in [self.lb_rows, self.lb_cols, self.lb_vals]:
            for i in reversed(t.curselection()): t.delete(i)
    def on_confirm(self):
        rows = list(self.lb_rows.get(0, tk.END))
        cols = list(self.lb_cols.get(0, tk.END))
        vals = list(self.lb_vals.get(0, tk.END))
        if not rows and not cols: return messagebox.showwarning("提示", "行和列不能同时为空")
        if not vals: return messagebox.showwarning("提示", "至少需要一个数值字段")
        self.result = {"index": rows, "columns": cols, "values": vals}
        self.action = "pivot"
        self.top.destroy()
    def on_skip(self):
        self.result = None
        self.action = "skip"
        self.top.destroy()
    def on_cancel(self):
        self.result = None
        self.action = "cancel"
        self.top.destroy()

# ==========================================
# 4. 主程序类 (V2 智能修正完整版)
# ==========================================
class AuditApp_V70_2:
    def __init__(self, root):
        self.root = root
        self.root.title("看账小工具 by CSDC")
        self.root.geometry("1280x850")
        style = ttk.Style(); style.theme_use('clam')
        style.configure("Slim.Horizontal.TScrollbar", arrowsize=10, width=10)
        
        self.file_path = None
        self.real_xlsx_path = None
        self.df_preview = None
        self.full_columns = []
        
        # 筛选容器
        self.target_accounts = set()
        self.exclude_accounts = set()
        self.target_batches = []
        self.active_batch_idx = 0
        
        self.column_mapping = {} 
        self.cached_accounts = None 
        self.current_sheet_name = 0
        self.header_row_idx = 0
        self.thread_event = threading.Event()
        self.user_save_path = None
        self.user_selected_cols = None 
        self.user_sheet_choice = None
        self.user_split_count = None 
        self.user_continue_detail = None
        self.pivot_config = None 
        self.progress_win = None 
        self.export_cancel_event = threading.Event()
        self.export_in_progress = False

        self.shadow_running = False
        self.shadow_ready = False
        self.shadow_error = None
        self.shadow_csv_path = None
        self.shadow_parquet_path = None
        self.shadow_source_path = None
        self.shadow_task_id = 0
        self.shadow_event = threading.Event()

        self.full_cache_df = None
        self.full_cache_running = False
        self.full_cache_ready = False
        self.full_cache_error = None
        self.full_cache_task_id = 0
        self.full_cache_event = threading.Event()
        self.full_cache_header_idx = None
        self.full_cache_source_tag = None
        
        self.shuttle_top = None
        self.shuttle_list_left = None
        self.shuttle_list_right = None
        self.shuttle_list_exclude = None
        self.shuttle_search_var = None
        self.shuttle_batch_list = None

        self.final_csv_sep = None
        self.final_csv_enc = None
        self.final_csv_sig = None
        self.preview_max_cols = 100
        self.perf_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "导入耗时日志.log")
        self.export_perf_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "导出耗时日志.jsonl")
        
        # 单位名称映射字典，用于存储单位名称的映射关系
        self.entity_map = {}
        
        self.ROLES = {
            "唯一识别码 (ID)": "role_id",
            "科目名称": "role_acc",
            "🏢 公司/主体 (可选)": "role_entity",
            "日期 (可选)": "role_date",
            "凭证摘要": "role_summary",
            "方案B-借方金额": "role_dr",
            "方案B-贷方金额": "role_cr",
            "方案A-金额列": "role_amt",
            "方案A-方向列": "role_dir"
        }
        self.KEYWORDS = {
            "role_id": ["Je number", "jenumber", "凭证编号", "凭证号", "凭证编码", "reference", "单据号"],
            "role_acc": ["科目名称", "科目描述", "account", "gl account", "总账科目", "GL Account Name"],
            "role_entity": ["公司名称", "公司", "单位名称", "单位", "主体", "entity", "company", "bukrs", "co code", "Business unit", "businessunit"],
            "role_date": ["日期", "date", "过账日期", "posting", "docdate", "postdate", "凭证日期", "effective date", "effectivedate"], 
            "role_summary": ["摘要", "描述", "行项目文本", "Description", "JE Line Description", "凭证摘要", "description"],
            
            # --- 方案B：借贷分列 (添加无空格小写版本以匹配清洗后的表头) ---
            "role_dr": [
                "functionaldebitamount", "debitamount", # <--- 对应 Functional Debit Amount
                "Debit Amount", "借方金额", "借项金额", "本币借项金额", "借方", "Functional Debit Amount"
            ],
            "role_cr": [
                "functionalcreditamount", "creditamount", # <--- 对应 Functional Credit Amount
                "Credit Amount", "贷方金额", "贷项金额", "本币贷项金额", "贷方", "Functional Credit Amount"
            ],
            
            # --- 方向列 (添加 debitcredit 以匹配 Debit_Credit) ---
            "role_dir": [
                "debitcredit", "Debit_Credit", # <--- 核心修复点：添加 debitcredit
                "方向", "借贷", "d/c", "dc", "记账码", "bschl", "p/k", "s/h", "s/h"
            ],
            
            # --- 方案A：单列金额 ---
            "role_amt": [
                "functionalamount", # <--- 对应 Functional Amount
                "金额", "Functional Amount", "amount", "本币金额"
            ]
        }
        if not HAS_XLSXWRITER: self.root.after(1000, lambda: messagebox.showwarning("提示", "未检测到 xlsxwriter库"))
        if not HAS_POLARS:
            raise RuntimeError("缺少依赖 polars。按当前性能方案，系统禁止回退到 pandas 路径。请先安装：pip install polars")
        self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)
        self.setup_ui()

    def _cleanup_shadow_files_on_exit(self):
        temp_dir = tempfile.gettempdir()
        try:
            for fn in os.listdir(temp_dir):
                low = fn.lower()
                if not low.startswith("shadow_"):
                    continue
                if not (low.endswith(".parquet") or low.endswith(".csv")):
                    continue
                p = os.path.join(temp_dir, fn)
                try:
                    os.remove(p)
                except Exception:
                    pass
        except Exception:
            pass

    def _on_app_close(self):
        try:
            self._cleanup_shadow_files_on_exit()
        finally:
            self.root.destroy()

    def _bind_esc_close(self, top, on_close=None):
        # 强制弹窗获取焦点，确保 Esc 可用
        try:
            top.transient(self.root)
            top.grab_set()
        except Exception:
            pass
        top.lift()
        top.after(10, lambda: top.focus_force())
        if on_close:
            top.bind("<Escape>", lambda e: on_close())
        else:
            top.bind("<Escape>", lambda e: top.destroy())
        self._bind_ctrl_a(top)

    def _bind_ctrl_a(self, top, callback=None):
        def _select_all(event=None):
            if callback:
                callback()
                return "break"
            w = top.focus_get()
            if isinstance(w, tk.Listbox):
                w.selection_set(0, tk.END)
                return "break"
            if isinstance(w, ttk.Treeview):
                for item in w.get_children():
                    w.selection_add(item)
                return "break"
            if isinstance(w, (tk.Entry, ttk.Entry)):
                w.select_range(0, tk.END)
                w.icursor(tk.END)
                return "break"
            if isinstance(w, tk.Text):
                w.tag_add(tk.SEL, "1.0", tk.END)
                w.mark_set(tk.INSERT, tk.END)
                return "break"
            return "break"
        top.bind("<Control-a>", _select_all)
        top.bind("<Control-A>", _select_all)

    def setup_ui(self):
        frame_ctrl = tk.LabelFrame(self.root, text=" 核心操作流程 ", padx=10, pady=8, font=("微软雅黑", 10, "bold"), fg="#333")
        frame_ctrl.pack(fill="x", padx=10, pady=5)
        
        f_1 = tk.Frame(frame_ctrl); f_1.pack(side="left", padx=5)
        ttk.Button(f_1, text="📂 1. 加载数据", command=self.load_file).pack(side="left")
        ttk.Separator(frame_ctrl, orient="vertical").pack(side="left", fill="y", padx=15)
        
        f_2 = tk.Frame(frame_ctrl); f_2.pack(side="left", padx=5)
        self.btn_filter = ttk.Button(f_2, text="🔍 2. 科目筛选", command=self.prepare_filter_data, state="disabled")
        self.btn_filter.pack(side="left")
        ttk.Separator(frame_ctrl, orient="vertical").pack(side="left", fill="y", padx=15)

        f_3 = tk.Frame(frame_ctrl); f_3.pack(side="left", padx=5)
        self.btn_run = ttk.Button(f_3, text="🚀 3. 导出 & 透视", command=self.start_process_flow, state="disabled")
        self.btn_run.pack(side="left")

        self.lbl_scheme_status = tk.Label(frame_ctrl, text="等待加载...", fg="gray", font=("Arial", 9, "bold"), width=52, anchor="w")
        self.lbl_scheme_status.pack(side="left", padx=20, fill="x", expand=True)

        f_opts = tk.Frame(frame_ctrl); f_opts.pack(side="right", padx=10)
        self.var_mark_je = tk.BooleanVar(value=True)
        ttk.Checkbutton(f_opts, text="启用正负数智能标记", variable=self.var_mark_je).pack(anchor="e")
        self.var_mark_loss = tk.BooleanVar(value=True)
        ttk.Checkbutton(f_opts, text="标记损益结转凭证", variable=self.var_mark_loss).pack(anchor="e")
        # --- 已删除: "按借正贷负处理(方案A)" 勾选框 ---
        self.var_partial_col = tk.BooleanVar(value=False)
        ttk.Checkbutton(f_opts, text="仅导出部分列", variable=self.var_partial_col).pack(anchor="e")

        frame_desc = tk.Frame(self.root, bg="#FFF8E1", padx=10, pady=8, relief="groove", borderwidth=1)
        frame_desc.pack(fill="x", padx=10, pady=(0, 5))

        # 左侧文本区域
        left_frame = tk.Frame(frame_desc, bg="#FFF8E1")
        left_frame.pack(side="left", fill="both", expand=True)

        info_text = (
            "ℹ️ 自动配置方案说明：\n"
            "   【方案 A】：适用于JE仅有'金额列'，或'金额列'加'方向列'。\n"
            "   【方案 B】：适用于JE有'借方'和'贷方'两列。\n"
            "   💡 快捷操作：【右键设置标题行，左键标题映射】/【复选多列作为凭证唯一识别码】/【标题自动映射】/【勾选开启“损益结转”识别】\n" 
            "   💡 支持导出套表：【月度透视表】/【对方科目表】/【凭证类型表】/【剔除科目[薪酬.折旧...]明细表】\n"
            "   💡 核心功能：【对方科目分析】/【凭证类型分析】/【正负数标记】\n"
            "   💡 支持选择csv格式导出，可大幅提升写入速度" 
        )

# === 修改开始：将长文本拆分为多行，以便调整间距 ===
        for line in info_text.strip().split('\n'):
            tk.Label(
                left_frame, 
                text=line, 
                bg="#FFF8E1", 
                justify="left", 
                fg="#5D4037", 
                font=("宋体", 9, "normal")
            ).pack(anchor="w", pady=2)  # <--- 修改这里的 pady 值（例如改为 2, 4, 5）来调整行间距
        # === 修改结束 ===

        # 右侧反馈功能区域 - 与左边字体和样式保持一致
        right_frame = tk.Frame(frame_desc, bg="#FFF8E1", padx=10, pady=0)  # 调整内边距，移除底部间距
        right_frame.pack(side="right", anchor="ne")  # 移除fill="y"，减少垂直占用

        # 点赞按钮 - 标准化尺寸和样式
        ttk.Button(right_frame, text="👍 认可", command=self.send_like_email, width=8).pack(fill="x", pady=2, padx=0)  # 统一按钮宽度和间距

        # 建议按钮 - 标准化尺寸和样式
        ttk.Button(right_frame, text="💡 建议", command=self.send_suggestion_email, width=8).pack(fill="x", pady=2, padx=0)  # 统一按钮宽度和间距

        # 提示文字 - 与左侧注释字体保持一致
        tk.Label(right_frame, text="", bg="#FFF8E1", fg="#0D47A1", font=("宋体", 9, "normal")).pack(anchor="center", pady=2)

        frame_info = tk.Frame(self.root, padx=10); frame_info.pack(fill="x")
        self.lbl_file = tk.Label(frame_info, text="当前文件: (未加载)", fg="gray"); self.lbl_file.pack(side="left")
        self.lbl_status = tk.Label(frame_info, text="", fg="blue"); self.lbl_status.pack(side="right")

        frame_table = tk.Frame(self.root); frame_table.pack(fill="both", expand=True, padx=7, pady=3)
        self.tree = ttk.Treeview(frame_table, show='headings', height=10)
        vsb = ttk.Scrollbar(frame_table, orient="vertical", command=self.tree.yview)
        hsb = tk.Scrollbar(frame_table, orient="horizontal", command=self.tree.xview, width=10, highlightthickness=0)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns'); hsb.grid(row=1, column=0, sticky='ew')
        frame_table.grid_rowconfigure(0, weight=1); frame_table.grid_columnconfigure(0, weight=1)
        
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="⬇️ 设为标题行 (重新加载)", command=self.set_row_as_header)
        self.tree.bind("<Button-3>", lambda e: (self.tree.identify_row(e.y) and self.tree.selection_set(self.tree.identify_row(e.y)), self.context_menu.post(e.x_root, e.y_root)))
        self._bind_global_select_all()

        def _tree_click_clear(event):
            # 常规点击时清除全选，仅保留当前点击项
            if (event.state & 0x0001) or (event.state & 0x0004):
                return
            item = self.tree.identify_row(event.y)
            if item:
                self.tree.selection_set(item)
        self.tree.bind("<Button-1>", _tree_click_clear, add="+")
    
    # === 反馈功能 ===
    def send_like_email(self):
        """发送点赞/认可邮件"""
        mailto_url = (
            "mailto:Melody.BT.Liu@cn.ey.com;April.YL.Wang@cn.ey.com"
            "?cc=John.SX.Yan@cn.ey.com"
            "&subject=内容认可反馈"
            "&body=工具实用，继续加油！"
        )
        webbrowser.open(mailto_url)
    
    def send_suggestion_email(self):
        """发送建议/反馈邮件"""
        mailto_url = (
            "mailto:Melody.BT.Liu@cn.ey.com;April.YL.Wang@cn.ey.com"
            "?cc=John.SX.Yan@cn.ey.com"
            "&subject=用户建议与反馈"
            "&body=[bug反馈或你的建议或想法]"
        )
        webbrowser.open(mailto_url)

    def _bind_global_select_all(self):
        def _select_all(event=None):
            w = self.root.focus_get()
            if isinstance(w, tk.Listbox):
                w.selection_set(0, tk.END)
                return "break"
            if isinstance(w, ttk.Treeview):
                for item in w.get_children():
                    w.selection_add(item)
                return "break"
            if isinstance(w, (tk.Entry, ttk.Entry)):
                w.select_range(0, tk.END)
                w.icursor(tk.END)
                return "break"
            if isinstance(w, tk.Text):
                w.tag_add(tk.SEL, "1.0", tk.END)
                w.mark_set(tk.INSERT, tk.END)
                return "break"
            return "break"
        self.root.bind_all("<Control-a>", _select_all)
        self.root.bind_all("<Control-A>", _select_all)

    def _apply_output_formatting(
        self,
        writer,
        df_target,
        t_cols,
        num_cols,
        map_inv,
        voucher_type_df,
        vt_type_col,
        vt_id_col,
        v_acc_col,
        pivot_res,
        voucher_type_strict_df=None,
        vt_type_col_strict=None,
        vt_id_col_strict=None,
        target_accounts=None,
    ):
        workbook = writer.book
        fmt_header = workbook.add_format({'bold': True})
        fmt_header_blue = workbook.add_format({'bold': True, 'bg_color': '#D9EAF7'})
        fmt_gray = workbook.add_format({'bg_color': '#E0E0E0'})
        fmt_white = workbook.add_format({'bg_color': '#FFFFFF'})
        fmt_num = workbook.add_format({'num_format': '#,##0'})
        fmt_bold = workbook.add_format({'bold': True})
        fmt_border = workbook.add_format({'border': 1})

        # 目标科目列表（用于条件加粗）
        target_sheet = None
        target_range = None
        target_range_norm = None
        effective_targets = set(target_accounts if target_accounts is not None else self.target_accounts)
        if effective_targets:
            target_sheet = "_targets"
            ws_targets = workbook.add_worksheet(target_sheet)
            targets_sorted = sorted(effective_targets)
            for i, val in enumerate(targets_sorted):
                ws_targets.write(i, 0, val)
                norm_val = re.sub(r"\s*-\s*", "-", str(val)).strip()
                ws_targets.write(i, 1, norm_val)
            ws_targets.hide()
            target_range = f"'{target_sheet}'!$A$1:$A${len(targets_sorted)}"
            target_range_norm = f"'{target_sheet}'!$B$1:$B${len(targets_sorted)}"

        def _col_letter(idx: int) -> str:
            # 0-based index to Excel column letter
            letters = ""
            n = idx
            while n >= 0:
                n, r = divmod(n, 26)
                letters = chr(65 + r) + letters
                n = n - 1
            return letters

        def _norm_expr(expr: str) -> str:
            return f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({expr}," - ","-")," -","-"),"- ","-")'

        # 凭证明细：筛选、冻结首行、加粗、辅助列灰底、金额格式
        if "凭证明细" in writer.sheets:
            ws = writer.sheets["凭证明细"]
            rows = len(df_target)
            cols = len(t_cols)
            ws.freeze_panes(1, 0)
            ws.set_row(0, None, fmt_header)
            if rows > 0 and cols > 0:
                ws.autofilter(0, 0, rows, cols - 1)
                ws.conditional_format(0, 0, rows, cols - 1, {'type': 'no_errors', 'format': fmt_border})

            # 金额列格式
            for c in num_cols:
                if c in t_cols:
                    ci = t_cols.index(c)
                    ws.set_column(ci, ci, None, fmt_num)

            # 辅助列灰底
            aux_cols = []
            if self.var_mark_loss.get():
                aux_cols.append('【损益结转】')
            if self.var_mark_je.get():
                aux_cols = ['【辅助_绝对值】', '【辅助_符号】', '【智能匹配状态】'] + (
                    ['【损益结转】'] if self.var_mark_loss.get() else []
                )
            if rows > 0:
                for c in aux_cols:
                    if c in t_cols:
                        ci = t_cols.index(c)
                        ws.conditional_format(1, ci, rows, ci, {'type': 'no_errors', 'format': fmt_gray})

            # 目标科目加粗：仅科目列 + 金额列(方案优先级)
            if target_range_norm and rows > 0:
                acc_cols = [c for c in map_inv.get('role_acc', []) if c in t_cols]
                if acc_cols:
                    # 多列科目时用 " - " 合并，并进行规范化后匹配
                    acc_letters = [_col_letter(t_cols.index(c)) for c in acc_cols]
                    acc_expr = f"{acc_letters[0]}2"
                    for lt in acc_letters[1:]:
                        acc_expr = f'{acc_expr}&" - "&{lt}2'
                    acc_expr_norm = _norm_expr(acc_expr)
                    formula = f"=COUNTIF({target_range_norm},{acc_expr_norm})>0"

                    bold_cols = []
                    bold_cols.extend(acc_cols)
                    if map_inv.get('role_dr') and map_inv.get('role_cr'):
                        bold_cols.extend([c for c in map_inv.get('role_dr', []) if c in t_cols])
                        bold_cols.extend([c for c in map_inv.get('role_cr', []) if c in t_cols])
                    elif map_inv.get('role_amt'):
                        bold_cols.extend([c for c in map_inv.get('role_amt', []) if c in t_cols])

                    for col_name in dict.fromkeys(bold_cols):
                        ci = t_cols.index(col_name)
                        ws.conditional_format(1, ci, rows, ci, {'type': 'formula', 'criteria': formula, 'format': fmt_bold})

        # 凭证类型：筛选、冻结首行、加粗、交替底色、YYYY-MM灰底、金额格式
        def _format_voucher_type_sheet(sheet_name, vt_df, type_col_hint, id_col_hint):
            if sheet_name not in writer.sheets or vt_df is None:
                return
            ws = writer.sheets[sheet_name]
            vt_out = vt_df.copy()
            vt_out.columns = self._flatten_columns(vt_out.columns)
            drop_cols = [c for c in vt_out.columns if str(c).startswith("科目名称-类型_")]
            if drop_cols:
                vt_out = vt_out.drop(columns=drop_cols)
            vt_out = self._reorder_voucher_type_columns(vt_out, type_col_hint, id_col_hint, v_acc_col)
            vt_cols = list(vt_out.columns)
            rows = len(vt_out)
            cols = len(vt_cols)
            ws.freeze_panes(1, 0)
            ws.set_row(0, None, fmt_header)
            if rows > 0 and cols > 0:
                ws.autofilter(0, 0, rows, cols - 1)
                ws.conditional_format(0, 0, rows, cols - 1, {'type': 'no_errors', 'format': fmt_border})

            type_col = type_col_hint if type_col_hint in vt_cols else "科目名称-类型"
            id_col = id_col_hint if id_col_hint in vt_cols else "唯一识别码"
            acc_col = v_acc_col if (v_acc_col in vt_cols) else (map_inv.get('role_acc', ["科目名称"])[0])

            def _is_month_col(x):
                s = str(x).strip()
                return len(s) == 7 and s[4] == '-' and s[:4].isdigit() and s[5:].isdigit()

            month_cols = [c for c in vt_cols if _is_month_col(c)]

            # 交替底色（按类型分组，且不覆盖YYYY-MM列）
            if type_col in vt_cols and rows > 0:
                col_data = vt_out[type_col]
                if isinstance(col_data, pd.DataFrame):
                    col_data = col_data.iloc[:, 0]
                type_vals = col_data.astype(str).tolist()
                first_month_idx = vt_cols.index(month_cols[0]) if month_cols else cols
                start = 0
                group_idx = 0
                for i in range(1, len(type_vals) + 1):
                    if i == len(type_vals) or type_vals[i] != type_vals[start]:
                        fmt = fmt_gray if (group_idx % 2 == 1) else fmt_white
                        ws.conditional_format(
                            start + 1, 0, i, max(0, first_month_idx - 1),
                            {'type': 'no_errors', 'format': fmt}
                        )
                        group_idx += 1
                        start = i

            # YYYY-MM列数值格式 + 表头淡蓝
            for c in month_cols:
                ci = vt_cols.index(c)
                ws.set_column(ci, ci, None, fmt_num)
                ws.write(0, ci, c, fmt_header_blue)

            # 目标科目加粗：仅科目列 + 净额列(所有 #_净额(Net))
            if target_range_norm and rows > 0 and acc_col in vt_cols:
                ci = vt_cols.index(acc_col)
                col_letter = _col_letter(ci)
                acc_expr_norm = _norm_expr(f"{col_letter}2")
                formula = f"=COUNTIF({target_range_norm},{acc_expr_norm})>0"
                net_cols = [c for c in vt_cols if "#_净额(Net)" in str(c)]
                for col_name in dict.fromkeys([acc_col] + net_cols):
                    ci = vt_cols.index(col_name)
                    ws.conditional_format(1, ci, rows, ci, {'type': 'formula', 'criteria': formula, 'format': fmt_bold})

            # 数值列格式（除文本列外）
            text_cols = {type_col, id_col, "摘要", acc_col}
            for ci, c in enumerate(vt_cols):
                if c in month_cols:
                    continue
                if c not in text_cols:
                    ws.set_column(ci, ci, None, fmt_num)

            # 合并唯一识别码
            if id_col in vt_cols and rows > 0:
                id_idx = vt_cols.index(id_col)
                col_data = vt_out[id_col]
                if isinstance(col_data, pd.DataFrame):
                    col_data = col_data.iloc[:, 0]
                id_vals = col_data.astype(str).tolist()
                start = 0
                for i in range(1, len(id_vals) + 1):
                    if i == len(id_vals) or id_vals[i] != id_vals[start]:
                        if i - start > 1:
                            ws.merge_range(start + 1, id_idx, i, id_idx, id_vals[start])
                        start = i

        _format_voucher_type_sheet("凭证类型-宽松", voucher_type_df, vt_type_col, vt_id_col)
        _format_voucher_type_sheet("凭证类型-严格", voucher_type_strict_df, vt_type_col_strict, vt_id_col_strict)

        # 透视分析：金额格式
        if "透视分析" in writer.sheets and pivot_res is not None:
            ws = writer.sheets["透视分析"]
            pivot_out = pivot_res.reset_index()
            pivot_out.columns = self._flatten_pivot_columns(pivot_out.columns)
            p_cols = list(pivot_out.columns)
            rows = len(pivot_out)
            cols = len(p_cols)
            ws.set_row(0, None, fmt_header)
            if rows > 0 and cols > 0:
                ws.freeze_panes(1, 0)
                ws.autofilter(0, 0, rows, cols - 1)
                ws.conditional_format(0, 0, rows, cols - 1, {'type': 'no_errors', 'format': fmt_border})
                for ci, c in enumerate(p_cols):
                    if ci == 0:
                        continue
                    ws.set_column(ci, ci, None, fmt_num)
                # 目标科目整行加粗
                if target_range_norm:
                    acc_cols = map_inv.get('role_acc', [])
                    acc_cols = [c for c in acc_cols if c in p_cols]
                    if acc_cols:
                        acc_letters = [_col_letter(p_cols.index(c)) for c in acc_cols]
                        acc_expr = f"{acc_letters[0]}2"
                        for lt in acc_letters[1:]:
                            acc_expr = f'{acc_expr}&" - "&{lt}2'
                        acc_expr_norm = _norm_expr(acc_expr)
                        formula = f"=COUNTIF({target_range_norm},{acc_expr_norm})>0"
                        ws.conditional_format(1, 0, rows, cols - 1, {'type': 'formula', 'criteria': formula, 'format': fmt_bold})

        # 全局边框：对所有有内容的sheet范围加黑色边框
        for _name, _ws in writer.sheets.items():
            try:
                rmax = _ws.dim_rowmax
                cmax = _ws.dim_colmax
                if rmax is not None and cmax is not None and rmax >= 0 and cmax >= 0:
                    _ws.conditional_format(0, 0, rmax, cmax, {'type': 'no_errors', 'format': fmt_border})
            except Exception:
                pass

        # 隐藏“凭证”sheet
        if "凭证" in writer.sheets:
            writer.sheets["凭证"].hide()

    def _get_file_signature(self, path):
        try:
            st = os.stat(path)
            return (os.path.abspath(path), st.st_size, st.st_mtime_ns)
        except Exception:
            return (os.path.abspath(path), None, None)

    def _read_text_file_with_fallback(self, h_idx, final_sep, final_enc, has_bom=False, nrows=None, usecols=None, chunksize=None):
        quoting = 3 if final_sep == '\t' else 0
        effective_h_idx = h_idx

        if effective_h_idx == 0:
            try:
                with open(self.file_path, 'r', encoding=final_enc, errors='ignore') as f:
                    lines = []
                    for _ in range(10):
                        line = f.readline()
                        if not line:
                            break
                        lines.append(line.strip())
                if lines:
                    max_cols_line = max(lines, key=lambda line: len(line.split(final_sep)))
                    effective_h_idx = lines.index(max_cols_line)
                    print(f"DEBUG: 自动检测到标题行位置: {effective_h_idx + 1}")
            except Exception:
                pass

        read_kwargs = dict(
            header=effective_h_idx,
            sep=final_sep,
            encoding=final_enc,
            nrows=nrows,
            usecols=usecols,
            chunksize=chunksize,
            dtype=str,
            on_bad_lines='skip',
            quoting=quoting,
            skip_blank_lines=True,
        )

        try:
            # 默认使用 C 引擎，失败后再回退 Python 引擎
            return pd.read_csv(self.file_path, **read_kwargs)
        except Exception as e:
            print(f"DEBUG: C引擎读取失败，回退python: {e}")
            try:
                return pd.read_csv(self.file_path, engine='python', **read_kwargs)
            except Exception as fallback_e:
                print(f"DEBUG: Python引擎读取失败，尝试保守参数: {fallback_e}")
                try:
                    return pd.read_csv(
                        self.file_path, header=effective_h_idx, sep=None, encoding=final_enc, engine='python',
                        nrows=nrows, usecols=usecols, chunksize=chunksize, dtype=str,
                        on_bad_lines='skip', quoting=3, skip_blank_lines=True
                    )
                except Exception as last_e:
                    print(f"DEBUG: 保守参数失败，进入最终fallback: {last_e}")
                    if self.file_path.lower().endswith('.txt'):
                        fallback_enc = 'utf-16-le' if has_bom else 'utf-8-sig'
                        return pd.read_csv(
                            self.file_path, header=3, sep='\t', encoding=fallback_enc, engine='python',
                            nrows=nrows, usecols=usecols, chunksize=chunksize, dtype=str,
                            on_bad_lines='skip', quoting=3, skip_blank_lines=True, skiprows=5
                        )
                    return pd.read_csv(
                        self.file_path, header=effective_h_idx, sep='\t', encoding=final_enc, engine='python',
                        nrows=nrows, usecols=usecols, chunksize=chunksize, dtype=str,
                        on_bad_lines='skip', quoting=3
                    )

    # === [核心修复] SAP兼容 & 暴力分隔符探测 ===
    def _universal_loader(self, h_idx, nrows=None, usecols=None, chunksize=None):
        # 检查文件是否为文本文件（包括CSV和TXT，支持大小写）
        file_ext = self.file_path.lower()
        is_text_file = file_ext.endswith('.csv') or file_ext.endswith('.txt')
        is_parquet_file = file_ext.endswith('.parquet')
        if is_parquet_file:
            return self._read_shadow_parquet(h_idx, nrows=nrows, usecols=usecols, chunksize=chunksize)

        if is_text_file:
            if self.real_xlsx_path is not None:
                return pd.read_csv(self.file_path, header=h_idx, sep=',', encoding='utf-8-sig',
                                   nrows=nrows, usecols=usecols, chunksize=chunksize, dtype=str, on_bad_lines='skip')

            file_sig = self._get_file_signature(self.file_path)
            if self.final_csv_sig == file_sig and self.final_csv_sep and self.final_csv_enc:
                return self._read_text_file_with_fallback(
                    h_idx, self.final_csv_sep, self.final_csv_enc,
                    has_bom=self.final_csv_enc.lower().startswith('utf-16'),
                    nrows=nrows, usecols=usecols, chunksize=chunksize
                )

            best_sep = ','
            best_enc = 'utf-8-sig'
            max_cols = 1
            has_bom = False
            try:
                with open(self.file_path, 'rb') as f:
                    raw_header = f.read(4)
                    if raw_header.startswith(b'\xff\xfe'):
                        best_enc = 'utf-16-le'  # 明确使用UTF-16 LE
                        has_bom = True
                    elif raw_header.startswith(b'\xfe\xff'):
                        best_enc = 'utf-16-be'  # 明确使用UTF-16 BE
                        has_bom = True
                    elif raw_header.startswith(b'\xef\xbb\xbf'):
                        best_enc = 'utf-8-sig'
                        has_bom = True
            except Exception:
                pass

            # 增加SAP常用编码，优先使用检测到的BOM编码
            if has_bom:
                encodings_to_try = [best_enc, 'utf-16', 'utf-8-sig', 'gb18030', 'cp936', 'iso-8859-1', 'latin1']
            else:
                encodings_to_try = ['utf-8-sig', 'utf-16-le', 'utf-16', 'gb18030', 'cp936', 'iso-8859-1', 'latin1']
            # 优先检测制表符，这是SAP导出CSV的常见分隔符
            separators = ['\t', ',', ';', '|']
            final_sep = '\t'  # 优先使用制表符作为默认分隔符
            final_enc = best_enc
            found_valid_structure = False

            # 记录所有可能的分隔符和对应的列数
            sep_candidates = []

            # 优先使用BOM检测结果
            if has_bom:
                # 如果检测到BOM，直接使用对应的编码，不尝试其他编码
                enc = best_enc
                try:
                    with open(self.file_path, 'r', encoding=enc) as f:  # 不使用errors='ignore'，确保编码正确
                        sample_lines = []
                        # 跳过开头的空行，最多跳过10行
                        skip_count = 0
                        while skip_count < 10:
                            line = f.readline()
                            if not line:
                                break
                            if line.strip():
                                sample_lines.append(line.strip())
                                break
                            skip_count += 1

                        # 继续读取更多样本行，总共最多500行
                        for _ in range(499):
                            line = f.readline()
                            if not line:
                                break
                            if line.strip():
                                sample_lines.append(line.strip())

                        if sample_lines:
                            for sep in separators:
                                col_counts = []
                                for line in sample_lines:
                                    if sep == '\t':
                                        cols = line.split(sep)
                                        col_counts.append(len(cols))
                                    elif sep == ' ':
                                        continue
                                    else:
                                        cols = line.split(sep)
                                        col_counts.append(len(cols))

                                current_max = max(col_counts) if col_counts else 0
                                current_avg = sum(col_counts) / len(col_counts) if col_counts else 0

                                sep_candidates.append({
                                    'sep': sep,
                                    'enc': enc,
                                    'max_cols': current_max,
                                    'avg_cols': current_avg,
                                    'valid': current_max > 1  # 至少要有2列才有效
                                })

                                if current_max > max_cols:
                                    max_cols = current_max
                                    final_sep = sep
                                    final_enc = enc
                                    found_valid_structure = True
                except Exception as e:
                    print(f"DEBUG: BOM编码读取失败: {e}")
                    found_valid_structure = False

            # 如果BOM编码读取失败，或者没有BOM，尝试其他编码
            if not found_valid_structure:
                for enc in encodings_to_try:
                    try:
                        with open(self.file_path, 'r', encoding=enc) as f:  # 不使用errors='ignore'
                            sample_lines = []
                            skip_count = 0
                            while skip_count < 10:
                                line = f.readline()
                                if not line:
                                    break
                                if line.strip():
                                    sample_lines.append(line.strip())
                                    break
                                skip_count += 1

                            for _ in range(499):
                                line = f.readline()
                                if not line:
                                    break
                                if line.strip():
                                    sample_lines.append(line.strip())

                            if not sample_lines:
                                continue

                            for sep in separators:
                                col_counts = []
                                for line in sample_lines:
                                    if sep == ' ':
                                        continue
                                    cols = line.split(sep)
                                    col_counts.append(len(cols))

                                current_max = max(col_counts) if col_counts else 0
                                current_avg = sum(col_counts) / len(col_counts) if col_counts else 0

                                sep_candidates.append({
                                    'sep': sep,
                                    'enc': enc,
                                    'max_cols': current_max,
                                    'avg_cols': current_avg,
                                    'valid': current_max > 1  # 至少要有2列才有效
                                })

                                if current_max > max_cols:
                                    max_cols = current_max
                                    final_sep = sep
                                    final_enc = enc
                                    found_valid_structure = True
                    except Exception:
                        continue

            # 如果没有找到有效分隔符，尝试从候选列表中选择最佳方案
            if not found_valid_structure and sep_candidates:
                sep_candidates.sort(key=lambda x: (x['valid'], x['avg_cols'], x['max_cols']), reverse=True)
                best_candidate = sep_candidates[0]
                final_sep = best_candidate['sep']
                final_enc = best_candidate['enc']

            # 强制使用BOM检测到的编码
            if has_bom:
                final_enc = best_enc

            # 智能选择分隔符
            if has_bom and (best_enc == 'utf-16-le' or best_enc == 'utf-16-be'):
                final_sep = '\t'  # SAP UTF-16文件通常使用制表符
            elif self.file_path.endswith('.csv') or self.file_path.endswith('.CSV'):
                if max_cols <= 1:  # 如果检测到的列数太少，使用逗号作为默认
                    final_sep = ','
            elif max_cols <= 1 and final_sep != '\t':
                final_sep = '\t'

            self.final_csv_enc = final_enc
            self.final_csv_sep = final_sep
            self.final_csv_sig = file_sig

            print(f"DEBUG: 文本文件处理 - 检测到的分隔符: '{final_sep}', 编码: '{final_enc}', 最大列数: {max_cols}")
            return self._read_text_file_with_fallback(
                h_idx, final_sep, final_enc, has_bom=has_bom,
                nrows=nrows, usecols=usecols, chunksize=chunksize
            )
        else:
            return pd.read_excel(self.file_path, sheet_name=self.current_sheet_name, header=h_idx, nrows=nrows, usecols=usecols, dtype=str, engine=EXCEL_ENGINE)

    def _read_shadow_parquet(self, h_idx, nrows=None, usecols=None, chunksize=None):
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"shadow Parquet不存在：{self.file_path}")

        header_idx = max(0, int(h_idx))
        header_pl = pl.read_parquet(self.file_path, n_rows=header_idx + 1)
        if header_pl.height == 0 or header_idx >= header_pl.height:
            return pd.DataFrame()

        raw_cols = list(header_pl.columns)
        header_vals = header_pl.row(header_idx)
        seen = {}
        out_cols = []
        for i, v in enumerate(header_vals):
            if v is None or str(v).strip() == "":
                base = f"Unnamed: {i}"
            else:
                base = str(v)
            if base not in seen:
                seen[base] = 0
                out_cols.append(base)
            else:
                seen[base] += 1
                out_cols.append(f"{base}.{seen[base]}")

        selected_pairs = list(zip(raw_cols, out_cols))
        if usecols is not None:
            use_set = set(usecols)
            selected_pairs = [(raw, out) for raw, out in selected_pairs if out in use_set]
        if not selected_pairs:
            return pd.DataFrame(columns=list(usecols or []))

        selected_raw_cols = [raw for raw, _ in selected_pairs]
        selected_out_cols = [out for _, out in selected_pairs]

        def _collect_slice(offset, length=None):
            scan = pl.scan_parquet(self.file_path).select(selected_raw_cols).slice(header_idx + 1 + offset, length)
            part = scan.collect()
            part.columns = selected_out_cols
            return self._to_pandas_df_safe(part.fill_null("")).astype(str)

        if chunksize:
            def _gen():
                start = 0
                step = int(chunksize)
                remaining = None if nrows is None else int(nrows)
                while remaining is None or remaining > 0:
                    take = step if remaining is None else min(step, remaining)
                    chunk_df = _collect_slice(start, take)
                    if chunk_df.empty:
                        break
                    yield chunk_df
                    got = len(chunk_df)
                    start += got
                    if remaining is not None:
                        remaining -= got
                    if got < take:
                        break
            return _gen()

        return _collect_slice(0, None if nrows is None else int(nrows))

    def _create_shadow_parquet(self, source_path=None, sheet_name=None):
        temp_dir = tempfile.gettempdir()
        src = source_path or self.file_path
        sheet = self.current_sheet_name if sheet_name is None else sheet_name
        base_name = os.path.basename(src).split('.')[0]
        parquet_path = os.path.join(temp_dir, f"shadow_{base_name}_{int(time.time())}.parquet")

        if EXCEL_ENGINE == 'calamine':
            wb = python_calamine.CalamineWorkbook.from_path(src)
            try:
                sh = wb.get_sheet_by_name(sheet) if isinstance(sheet, str) else wb.get_sheet_by_index(int(sheet))
                all_rows = []
                max_cols = 0
                for row in sh.iter_rows():
                    vals = ["" if v is None else str(v) for v in row]
                    max_cols = max(max_cols, len(vals))
                    all_rows.append(vals)
                if max_cols == 0:
                    pl.DataFrame().write_parquet(parquet_path)
                else:
                    norm_rows = [r + [""] * (max_cols - len(r)) for r in all_rows]
                    col_names = [f"column_{i}" for i in range(max_cols)]
                    pl.DataFrame(norm_rows, schema=col_names, orient="row").write_parquet(parquet_path)
            finally:
                wb.close()
        else:
            df = pd.read_excel(src, sheet_name=sheet, header=None, dtype=str, engine=EXCEL_ENGINE)
            self._to_polars_df_safe(df.fillna("")).write_parquet(parquet_path)
            del df
            gc.collect()
        return parquet_path

    def _create_shadow_csv(self, source_path=None, sheet_name=None):
        return self._create_shadow_parquet(source_path=source_path, sheet_name=sheet_name)

    def _load_xlsx_preview_openpyxl(self, h_idx, nrows=50):
        max_preview_cols = max(1, int(self.preview_max_cols))

        def _trim_and_cap(row):
            vals = list(row)
            last = -1
            for i, v in enumerate(vals):
                if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() != "":
                    last = i
            if last >= 0:
                vals = vals[:last + 1]
            else:
                vals = []
            return vals[:max_preview_cols]

        if load_workbook is None:
            raise RuntimeError("openpyxl 不可用")

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            ws = wb[self.current_sheet_name] if isinstance(self.current_sheet_name, str) else wb.worksheets[int(self.current_sheet_name)]
            min_row = max(1, int(h_idx) + 1)
            max_row = min_row + max(0, int(nrows))
            rows_iter = ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=1,
                max_col=max_preview_cols,
                values_only=True
            )

            header_row = next(rows_iter, None)
            if header_row is None:
                return pd.DataFrame()

            header_vals_raw = _trim_and_cap(header_row)
            max_cols = len(header_vals_raw)
            data_rows = []
            for _ in range(max(0, nrows)):
                row = next(rows_iter, None)
                if row is None:
                    break
                row_vals = _trim_and_cap(row)
                max_cols = max(max_cols, len(row_vals))
                data_rows.append(row_vals)

            header_vals = header_vals_raw + [None] * (max_cols - len(header_vals_raw))
            col_names = []
            seen = {}
            for i, v in enumerate(header_vals):
                if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "":
                    base = f"Unnamed: {i}"
                else:
                    base = str(v)
                if base not in seen:
                    seen[base] = 0
                    col_names.append(base)
                else:
                    seen[base] += 1
                    col_names.append(f"{base}.{seen[base]}")

            normalized_rows = []
            for r in data_rows:
                normalized_rows.append(r + [None] * (max_cols - len(r)))

            return pd.DataFrame(normalized_rows, columns=col_names)
        finally:
            wb.close()

    def _reset_shadow_state(self):
        self.shadow_task_id += 1
        self.shadow_running = False
        self.shadow_ready = False
        self.shadow_error = None
        self.shadow_csv_path = None
        self.shadow_parquet_path = None
        self.shadow_source_path = None
        self.shadow_event = threading.Event()

    def _reset_full_cache_state(self):
        self.full_cache_task_id += 1
        self.full_cache_df = None
        self.full_cache_running = False
        self.full_cache_ready = False
        self.full_cache_error = None
        self.full_cache_event = threading.Event()
        self.full_cache_header_idx = None
        self.full_cache_source_tag = None

    def _current_source_tag(self):
        src = self.real_xlsx_path if self.real_xlsx_path else self.file_path
        return (src, int(self.header_row_idx), str(self.current_sheet_name))

    def _start_full_cache_background(self):
        if self.full_cache_running:
            return
        if self.file_path is None:
            return

        task_id = self.full_cache_task_id
        header_idx = int(self.header_row_idx)
        source_tag = self._current_source_tag()
        self.full_cache_running = True
        self.full_cache_ready = False
        self.full_cache_error = None
        self.full_cache_event.clear()

        def _worker(tid, h_idx, src_tag):
            try:
                # Excel场景下先等待shadow转换，确保后续全量读取走shadow文件
                if self.real_xlsx_path is None and self.file_path and self.file_path.lower().endswith(('.xlsx', '.xls')):
                    if not self.shadow_running and not self.shadow_ready:
                        self._start_shadow_parquet_background()
                    if self.shadow_running:
                        self.shadow_event.wait()
                    if self.shadow_error:
                        raise Exception(f"后台转换shadow失败：{self.shadow_error}")

                t0 = time.perf_counter()
                df_full = self._universal_loader(h_idx)
                if tid != self.full_cache_task_id:
                    return
                self.full_cache_df = df_full
                self.full_cache_ready = True
                self.full_cache_running = False
                self.full_cache_error = None
                self.full_cache_header_idx = h_idx
                self.full_cache_source_tag = src_tag
                self.full_cache_event.set()
                self._perf_log(f"后台全量缓存完成: {time.perf_counter() - t0:.3f}s, rows={len(df_full)}")
            except Exception as e:
                if tid != self.full_cache_task_id:
                    return
                self.full_cache_df = None
                self.full_cache_ready = False
                self.full_cache_running = False
                self.full_cache_error = str(e)
                self.full_cache_event.set()
                self._perf_log(f"后台全量缓存失败: {e}")

        threading.Thread(target=_worker, args=(task_id, header_idx, source_tag), daemon=True).start()

    def _start_shadow_parquet_background(self):
        if self.real_xlsx_path is not None:
            return
        if not self.file_path.lower().endswith(('.xlsx', '.xls')):
            return
        if self.shadow_running or self.shadow_ready:
            return

        task_id = self.shadow_task_id
        source_path = self.file_path
        self.shadow_running = True
        self.shadow_ready = False
        self.shadow_error = None
        self.shadow_source_path = source_path
        self.shadow_event.clear()

        sheet_name = self.current_sheet_name

        def _worker(tid, src, sht):
            out_file = None
            try:
                t_shadow = time.perf_counter()
                out_file = self._create_shadow_parquet(source_path=src, sheet_name=sht)
                elapsed = time.perf_counter() - t_shadow
                if tid != self.shadow_task_id:
                    if out_file and os.path.exists(out_file):
                        try:
                            os.remove(out_file)
                        except Exception:
                            pass
                    return
                self.real_xlsx_path = src
                self.file_path = out_file
                self.shadow_csv_path = out_file if out_file.lower().endswith(".csv") else None
                self.shadow_parquet_path = out_file if out_file.lower().endswith(".parquet") else None
                self.shadow_ready = True
                self.shadow_running = False
                self.shadow_error = None
                self.shadow_event.set()
                self._perf_log(f"后台转shadow完成: {elapsed:.3f}s, 输出={out_file}")
            except Exception as e:
                if tid != self.shadow_task_id:
                    return
                self.shadow_ready = False
                self.shadow_running = False
                self.shadow_error = str(e)
                self.shadow_event.set()
                self._perf_log(f"后台转shadow失败: {e}")

        threading.Thread(target=_worker, args=(task_id, source_path, sheet_name), daemon=True).start()

    def _start_shadow_csv_background(self):
        self._start_shadow_parquet_background()

    def request_export_cancel(self):
        if self.export_in_progress:
            self.export_cancel_event.set()
            if self.progress_win:
                try:
                    self.progress_win.set_message("正在终止导出，请稍候...")
                except Exception:
                    pass

    def _raise_if_export_cancelled(self):
        if self.export_cancel_event.is_set():
            raise ExportCancelled("用户已终止导出")

    def show_progress(self, msg="处理中...", allow_cancel=False):
        if self.progress_win:
            self.progress_win.destroy()
        self.progress_win = ProgressWindow(
            self.root,
            title=msg,
            message=msg,
            on_cancel=self.request_export_cancel if allow_cancel else None,
            cancellable=allow_cancel,
        )
        self.root.update()
    def hide_progress(self):
        if self.progress_win: self.progress_win.close(); self.progress_win = None

    def _reset_filter_state(self):
        self.target_accounts.clear()
        self.exclude_accounts.clear()
        self.target_batches = []
        self.active_batch_idx = 0

    def _ensure_default_batch(self):
        if not self.target_batches:
            self.target_batches = [{"name": "批次1", "accounts": set(self.target_accounts)}]
            self.active_batch_idx = 0
        if self.active_batch_idx >= len(self.target_batches):
            self.active_batch_idx = max(0, len(self.target_batches) - 1)

    def _sync_active_batch_from_target_accounts(self):
        self._ensure_default_batch()
        self.target_batches[self.active_batch_idx]["accounts"] = set(self.target_accounts)

    def _load_target_accounts_from_active_batch(self):
        self._ensure_default_batch()
        self.target_accounts = set(self.target_batches[self.active_batch_idx]["accounts"])

    def _get_effective_batches(self):
        batches = []
        if self.target_batches:
            for b in self.target_batches:
                accs = set(b.get("accounts", set()))
                if accs:
                    batches.append((str(b.get("name") or "未命名批次"), accs))
        if batches:
            return batches
        if self.target_accounts:
            return [("批次1", set(self.target_accounts))]
        return [("全部", set())]

    def _combine_account_key(self, frame, acc_cols):
        if len(acc_cols) == 1:
            return frame[acc_cols[0]].fillna("")
        out = frame[acc_cols[0]].fillna("").astype(str)
        for c in acc_cols[1:]:
            out = out.str.cat(frame[c].fillna("").astype(str), sep=" - ")
        return out

    def _perf_log(self, msg):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {msg}"
        try:
            print(line, flush=True)
        except Exception:
            pass
        try:
            with open(self.perf_log_path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass

    def set_row_as_header(self):
        sel = self.tree.selection()
        if not sel: return
        idx = self.tree.index(sel[0])
        # Treeview展示的是“当前标题行之后”的数据行，所以要 +1 才是实际标题行索引
        self.header_row_idx += idx + 1
        self.lbl_status.config(text=f"⏳ 重载中... 标题行: {self.header_row_idx+1}")
        self.cached_accounts = None; self._reset_filter_state(); self._reset_full_cache_state()
        self.show_progress("重新读取..."); threading.Thread(target=self.process_load, args=(False,)).start()

    def load_file(self):
        path = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx *.xls *.csv *.txt *.TXT")])
        if not path: return
        self.file_path = path; self.header_row_idx = 0; self.current_sheet_name = 0
        self.real_xlsx_path = None; self.final_csv_sep = None; self.final_csv_enc = None; self.final_csv_sig = None
        self._reset_shadow_state()
        self._reset_full_cache_state()
        self._perf_log("=" * 50)
        self._perf_log(f"开始加载文件: {path}")
        self.lbl_file.config(text=f"文件: {os.path.basename(path)}", fg="black")
        self.cached_accounts = None; self._reset_filter_state()
        self.show_progress("读取结构..."); threading.Thread(target=self.process_load, args=(True,)).start()

    def process_load(self, check_sheet=True):
        try:
            t0 = time.perf_counter()
            need_shadow = False
            if check_sheet and (self.file_path.endswith('.xlsx') or self.file_path.endswith('.xls')):
                try:
                    t_sheet = time.perf_counter()
                    xls = pd.ExcelFile(self.file_path, engine=EXCEL_ENGINE)
                    self._perf_log(f"ExcelFile读取sheet列表耗时: {time.perf_counter() - t_sheet:.3f}s, sheet数={len(xls.sheet_names)}")
                    if len(xls.sheet_names) > 1:
                        self.thread_event.clear(); self.root.after(0, lambda: self._ask_sheet_on_main(xls.sheet_names)); self.thread_event.wait()
                        if not self.user_sheet_choice: return self.root.after(0, self.hide_progress)
                        self.current_sheet_name = self.user_sheet_choice
                    else: self.current_sheet_name = xls.sheet_names[0]
                    need_shadow = True
                except: pass
            if self.file_path.lower().endswith('.xlsx') and self.real_xlsx_path is None:
                t_preview = time.perf_counter()
                df = self._load_xlsx_preview_openpyxl(self.header_row_idx, nrows=50)
                self._perf_log(f"XLSX预览50行耗时: {time.perf_counter() - t_preview:.3f}s")
            else:
                t_preview = time.perf_counter()
                df = self._universal_loader(self.header_row_idx, nrows=50)
                self._perf_log(f"通用读取预览50行耗时: {time.perf_counter() - t_preview:.3f}s")
            if df.empty: raise Exception("文件读取为空")
            df = df.dropna(how='all', axis=1)
            self.df_preview = df.fillna("")
            self.full_columns = list(df.columns)
            self.auto_map_columns()
            if need_shadow:
                self._start_shadow_parquet_background()
            # 所有格式均在导入阶段后台做一次全量缓存
            self._start_full_cache_background()
            self._perf_log(f"加载流程总耗时: {time.perf_counter() - t0:.3f}s")
            self.root.after(0, self.build_table)
        except Exception as e: msg = str(e); self.root.after(0, lambda: messagebox.showerror("错误", msg))
        finally: self.root.after(0, self.hide_progress)

    def _ask_sheet_on_main(self, names):
        self.hide_progress(); top = tk.Toplevel(self.root); _fit_toplevel_to_screen(top, 300, 150, min_width=280, min_height=140)
        cb = ttk.Combobox(top, values=names, state="readonly"); cb.pack(pady=20); cb.current(0)
        def ok(): self.user_sheet_choice = cb.get(); top.destroy(); self.show_progress(); self.thread_event.set()
        top.protocol("WM_DELETE_WINDOW", lambda: (setattr(self, 'user_sheet_choice', None), top.destroy(), self.thread_event.set()))
        self._bind_esc_close(top, lambda: (setattr(self, 'user_sheet_choice', None), top.destroy(), self.thread_event.set()))
        ttk.Button(top, text="确定", command=ok).pack()

    def auto_map_columns(self):
        self.column_mapping = {} 
        cols_lower = {str(c): str(c).lower().replace("_", "").replace(" ", "") for c in self.full_columns}
        def add_role(role, col):
            if col not in self.column_mapping: self.column_mapping[col] = set()
            self.column_mapping[col].add(role)
        # 优先映射摘要，避免 Description 被科目名称抢占
        for role in ["role_id", "role_summary", "role_entity", "role_date", "role_acc"]: 
            c = self._find_col_strict(role, cols_lower); 
            if c:
                # 若列已映射为摘要，则不再映射为科目名称
                if role == "role_acc" and c in self.column_mapping and "role_summary" in self.column_mapping.get(c, set()):
                    pass
                else:
                    add_role(role, c)
        c_dr, c_cr = self._find_col_strict("role_dr", cols_lower), self._find_col_strict("role_cr", cols_lower)
        c_dir = self._find_col_strict("role_dir", cols_lower)
        if c_dr and c_cr:
            # 同一列表头同时含有借贷语义时，按方案A方向列处理，不再误判为方案B借贷分列
            if c_dr == c_cr and self._is_combined_dr_cr_header(c_dr, cols_lower):
                add_role("role_dir", c_dr)
                c_amt = self._find_col_strict("role_amt", cols_lower)
                if c_amt:
                    add_role("role_amt", c_amt)
            else:
                add_role("role_dr", c_dr); add_role("role_cr", c_cr)
                # 方案B存在时，不再自动映射方案A金额列
        else:
            c_amt = self._find_col_strict("role_amt", cols_lower)
            if c_amt: add_role("role_amt", c_amt)
        if c_dir: add_role("role_dir", c_dir)
        self.root.after(0, self.update_scheme_status)

    def _find_col_strict(self, role, cols_lower):
        for kw in self.KEYWORDS[role]:
            for col in self.full_columns:
                if kw == cols_lower[str(col)]: return col
        for kw in self.KEYWORDS[role]:
            for col in self.full_columns:
                if kw in cols_lower[str(col)]: return col
        return None

    def _is_combined_dr_cr_header(self, col, cols_lower):
        normalized = cols_lower.get(str(col), str(col).lower().replace("_", "").replace(" ", ""))
        has_dr = ("借" in normalized) or ("debit" in normalized)
        has_cr = ("贷" in normalized) or ("credit" in normalized)
        return has_dr and has_cr

    def build_table(self):
        self.tree.delete(*self.tree.get_children()); self.tree["columns"] = self.full_columns
        prefix_map = {"role_id": "【🔑ID】", "role_acc": "【📘科目】", "role_entity": "【🏢公司】", 
                      "role_date": "【📅日期】", "role_summary": "【📝摘要】", "role_dr": "【➕借方】", "role_cr": "【➖贷方】", 
                      "role_amt": "【💰金额】", "role_dir": "【🧭方向】"}
        for col in self.full_columns:
            txt = str(col)
            if col in self.column_mapping:
                tags = "".join([prefix_map.get(r, "") for r in self.column_mapping[col]])
                txt = f"{tags}{col}"
            self.tree.heading(col, text=txt, command=lambda c=col: self.on_header_click(c))
            self.tree.column(col, width=120, stretch=False)
        for _, row in self.df_preview.iterrows(): self.tree.insert("", "end", values=list(row))
        self.btn_filter.config(state="normal"); self.btn_run.config(state="normal")
        display_name = os.path.basename(self.real_xlsx_path) if self.real_xlsx_path else os.path.basename(self.file_path)
        if self.header_row_idx == 0: self.lbl_status.config(text="⚠️ 请右键设置标题行", fg="red")
        else: self.lbl_status.config(text=f"✅ 加载完成 [{display_name}]", fg="green")
        self.update_scheme_status()

    def update_scheme_status(self):
        map_inv = {v: [] for v in self.ROLES.values()}
        for col, roles in self.column_mapping.items():
            for r in roles: map_inv[r].append(col) if r in map_inv else None

        role_name = {
            "role_id": "ID",
            "role_acc": "科目",
            "role_amt": "金额",
            "role_dr": "借方",
            "role_cr": "贷方",
            "role_summary": "凭证摘要",
            "role_date": "日期",
            "role_dir": "方向",
        }

        required_a = ["role_id", "role_acc", "role_amt"]
        required_b = ["role_id", "role_acc", "role_dr", "role_cr"]

        missing_a = [role_name[r] for r in required_a if not map_inv.get(r)]
        missing_b = [role_name[r] for r in required_b if not map_inv.get(r)]

        ready_a = (len(missing_a) == 0)
        ready_b = (len(missing_b) == 0)

        if ready_b:
            status, color = "🟢 就绪 (方案B: 借贷分列)", "#2E7D32"
            optional_notice = ["role_summary", "role_date"]
        elif ready_a:
            status, color = "🟢 就绪 (方案A: 单列金额)", "#2E7D32"
            optional_notice = ["role_summary", "role_date", "role_dir"]
        else:
            status, color = "🔴 未就绪", "red"
            status += f" ｜ 方案A缺少: {'、'.join(missing_a)}"
            status += f"；方案B缺少: {'、'.join(missing_b)}"
            optional_notice = ["role_summary", "role_date"]

        optional_missing = [role_name[r] for r in optional_notice if not map_inv.get(r)]

        if optional_missing:
            status += f" ｜ 可补充: {'、'.join(optional_missing)}"

        # 防止状态文本过长挤压右侧勾选项显示
        max_len = 72
        status_show = status if len(status) <= max_len else (status[:max_len - 1] + "…")
        self.lbl_scheme_status.config(text=status_show, fg=color)

    def on_header_click(self, col):
        top = tk.Toplevel(self.root)
        top.title("勾选角色")
        _fit_toplevel_to_screen(top, 340, 460, min_width=320, min_height=360)
        try:
            top.transient(self.root)
        except Exception:
            pass

        button_bar = tk.Frame(top, padx=10, pady=10)
        button_bar.pack(side=tk.BOTTOM, fill=tk.X)

        body = tk.Frame(top)
        body.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        tk.Label(body, text="勾选角色:", pady=10).pack()

        canvas = tk.Canvas(body, highlightthickness=0)
        scrollbar = ttk.Scrollbar(body, orient=tk.VERTICAL, command=canvas.yview)
        role_frame = tk.Frame(canvas)
        role_window = canvas.create_window((0, 0), window=role_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _sync_scrollregion(event=None):
            canvas.itemconfig(role_window, width=canvas.winfo_width())
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        role_frame.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        role_frame.bind("<MouseWheel>", _on_mousewheel)
        current_roles = self.column_mapping.get(col, set())
        vars_map = {}
        for label, role_code in self.ROLES.items():
            var = tk.BooleanVar(value=(role_code in current_roles))
            cb = tk.Checkbutton(role_frame, text=label, variable=var)
            cb.pack(anchor="w", padx=20, pady=2)
            cb.bind("<MouseWheel>", _on_mousewheel)
            vars_map[role_code] = var
        def confirm():
            new = {r for r, v in vars_map.items() if v.get()}
            if new: self.column_mapping[col] = new
            elif col in self.column_mapping: del self.column_mapping[col]
            self.cached_accounts = None; self._reset_filter_state()
            self.build_table(); self.update_scheme_status(); top.destroy()
        self._bind_esc_close(top)
        tk.Button(button_bar, text="取消", command=top.destroy, width=10).pack(side=tk.RIGHT, padx=5)
        tk.Button(button_bar, text="确定", command=confirm, width=10).pack(side=tk.RIGHT, padx=5)

    # ===============================================
    # 4. 筛选逻辑
    # ===============================================
    def prepare_filter_data(self):
        if self.cached_accounts is not None: return self.open_shuttle_dialog()
        
        acc_cols = [c for c, roles in self.column_mapping.items() if "role_acc" in roles]
        if not acc_cols: return messagebox.showerror("错误", "请先标记【科目名称】")
        
        self.btn_filter.config(text="⏳...", state="disabled"); self.show_progress("扫描科目...")
        threading.Thread(target=self.scan_core, args=(acc_cols,)).start()

    def scan_core(self, acc_cols):
        try:
            accs = set()
            chunk = 50000
            if self.real_xlsx_path is None and self.file_path.lower().endswith(('.xlsx', '.xls')):
                if not self.shadow_running and not self.shadow_ready:
                    self._start_shadow_parquet_background()
                if self.shadow_running:
                    t_wait = time.perf_counter()
                    self.root.after(0, lambda: self.progress_win and self.progress_win.set_message("后台正在转换 shadow 文件，请稍候..."))
                    self.shadow_event.wait()
                    self._perf_log(f"筛选前等待后台转shadow耗时: {time.perf_counter() - t_wait:.3f}s")
                if self.shadow_error:
                    raise Exception(f"后台转换shadow失败：{self.shadow_error}")
                if self.real_xlsx_path is None and not self.shadow_ready:
                    raise Exception("后台转换shadow尚未完成，请稍后重试")

            is_text_file = self.file_path.lower().endswith(('.csv', '.txt', '.parquet'))
            t_scan = time.perf_counter()

            if is_text_file:
                reader = self._universal_loader(self.header_row_idx, usecols=acc_cols, chunksize=chunk)
                for c in reader: 
                    if len(acc_cols) > 1:
                        combined = self._combine_account_key(c, acc_cols)
                        accs.update(combined.unique())
                    else:
                        accs.update(c[acc_cols[0]].dropna().unique())
            else:
                df = self._universal_loader(self.header_row_idx, usecols=acc_cols)
                if len(acc_cols) > 1:
                    combined = self._combine_account_key(df, acc_cols)
                    accs = set(combined.unique())
                else:
                    accs = set(df[acc_cols[0]].dropna().unique())
                del df; gc.collect()
            
            self.cached_accounts = sorted(list(accs))
            self._perf_log(f"科目扫描总耗时: {time.perf_counter() - t_scan:.3f}s, 科目数={len(self.cached_accounts)}")
            self.root.after(0, lambda: [self.btn_filter.config(text="🔍 科目筛选", state="normal"), self.hide_progress(), self.open_shuttle_dialog()])
        except Exception as e:
            msg = str(e); self.root.after(0, lambda: [self.hide_progress(), messagebox.showerror("失败", msg)])

    def refresh_filter_data(self):
        acc_cols = [c for c, roles in self.column_mapping.items() if "role_acc" in roles]
        if not acc_cols: return messagebox.showerror("错误", "找不到【科目名称】列")
        self.shuttle_btn_refresh.config(state="disabled", text="⏳..."); self.show_progress("重新扫描...")
        threading.Thread(target=self.scan_core, args=(acc_cols,)).start()

    def handle_drag_drop(self, items, target_type):
        self.target_accounts.difference_update(items)
        self.exclude_accounts.difference_update(items)
        if target_type == 'target': self.target_accounts.update(items)
        elif target_type == 'exclude': self.exclude_accounts.update(items)
        self._sync_active_batch_from_target_accounts()
        self.update_shuttle_ui()

    def _refresh_batch_list_ui(self):
        if not self.shuttle_batch_list:
            return
        self.shuttle_batch_list.delete(0, tk.END)
        self._ensure_default_batch()
        for i, b in enumerate(self.target_batches):
            cnt = len(b.get("accounts", set()))
            self.shuttle_batch_list.insert(tk.END, f"{i+1}. {b.get('name', '未命名批次')} ({cnt})")
        if self.target_batches:
            self.shuttle_batch_list.selection_clear(0, tk.END)
            self.shuttle_batch_list.selection_set(self.active_batch_idx)

    def _switch_active_batch(self, idx):
        if idx < 0:
            return
        self._sync_active_batch_from_target_accounts()
        self._ensure_default_batch()
        if idx >= len(self.target_batches):
            return
        self.active_batch_idx = idx
        self._load_target_accounts_from_active_batch()
        self._refresh_batch_list_ui()
        self.update_shuttle_ui()

    def open_shuttle_dialog(self):
        if self.shuttle_top and self.shuttle_top.winfo_exists(): self.shuttle_top.lift(); return
        self._ensure_default_batch()
        self._load_target_accounts_from_active_batch()
        top = tk.Toplevel(self.root)
        _fit_toplevel_to_screen(top, 980, 700, min_width=760, min_height=520)
        self.shuttle_top = top
        self._bind_esc_close(top)
        f_bottom = tk.Frame(top, padx=10, pady=10)
        f_bottom.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Button(f_bottom, text="取消", width=12, command=top.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(
            f_bottom,
            text="确定并导出",
            width=20,
            command=lambda: [self._sync_active_batch_from_target_accounts(), top.destroy(), self.start_process_flow()],
        ).pack(side=tk.RIGHT)

        f_top = tk.Frame(top, bg="#f0f0f0", pady=5); f_top.pack(side=tk.TOP, fill=tk.X)
        self.shuttle_btn_refresh = ttk.Button(f_top, text="🔄 刷新", command=self.refresh_filter_data)
        self.shuttle_btn_refresh.pack(side="left", padx=10)
        tk.Label(f_top, text="🔍 搜索:", bg="#f0f0f0").pack(side="left")
        self.shuttle_search_var = tk.StringVar(); ttk.Entry(f_top, textvariable=self.shuttle_search_var).pack(side="left", fill="x", expand=True)
        
        f_main = tk.Frame(top); f_main.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=5)
        f_main.grid_columnconfigure(0, weight=1)
        f_main.grid_columnconfigure(1, weight=0)
        f_main.grid_columnconfigure(2, weight=1)
        f_main.grid_rowconfigure(0, weight=0)
        f_main.grid_rowconfigure(1, weight=1)
        f_main.grid_rowconfigure(2, weight=1)

        f_left = tk.LabelFrame(f_main, text="待选科目 (可拖拽)")
        f_left.grid(row=0, column=0, rowspan=3, sticky="nsew")
        self.shuttle_list_left = DraggableListbox(f_left, self, None, 'source', selectmode="extended")
        self.shuttle_list_left.pack(fill="both", expand=True)

        f_batch = tk.LabelFrame(f_main, text="📦 目标批次")
        f_batch.grid(row=0, column=2, sticky="nsew", pady=(0, 5), padx=(5, 0))
        self.shuttle_batch_list = tk.Listbox(f_batch, height=4, exportselection=False)
        self.shuttle_batch_list.pack(fill="x", padx=5, pady=5)
        f_batch_btn = tk.Frame(f_batch)
        f_batch_btn.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Button(
            f_batch_btn,
            text="新增批次",
            command=lambda: [
                self._sync_active_batch_from_target_accounts(),
                self.target_batches.append({"name": f"批次{len(self.target_batches)+1}", "accounts": set()}),
                setattr(self, "active_batch_idx", len(self.target_batches) - 1),
                self._load_target_accounts_from_active_batch(),
                self._refresh_batch_list_ui(),
                self.update_shuttle_ui(),
            ],
        ).pack(side="left")
        ttk.Button(
            f_batch_btn,
            text="重命名",
            command=lambda: (
                self._ensure_default_batch(),
                self.target_batches.__setitem__(
                    self.active_batch_idx,
                    {
                        "name": (
                            simpledialog.askstring(
                                "批次名称",
                                "请输入批次名称：",
                                initialvalue=self.target_batches[self.active_batch_idx].get("name", f"批次{self.active_batch_idx+1}"),
                                parent=top,
                            )
                            or self.target_batches[self.active_batch_idx].get("name", f"批次{self.active_batch_idx+1}")
                        ).strip() or self.target_batches[self.active_batch_idx].get("name", f"批次{self.active_batch_idx+1}"),
                        "accounts": set(self.target_batches[self.active_batch_idx].get("accounts", set())),
                    },
                ),
                self._refresh_batch_list_ui(),
            ),
        ).pack(side="left", padx=5)
        ttk.Button(
            f_batch_btn,
            text="删除批次",
            command=lambda: (
                self._sync_active_batch_from_target_accounts(),
                len(self.target_batches) > 1 and self.target_batches.pop(self.active_batch_idx),
                setattr(self, "active_batch_idx", min(self.active_batch_idx, max(0, len(self.target_batches) - 1))),
                self._ensure_default_batch(),
                self._load_target_accounts_from_active_batch(),
                self._refresh_batch_list_ui(),
                self.update_shuttle_ui(),
            ),
        ).pack(side="left")
        
        f_r_top = tk.LabelFrame(f_main, text="🟢 目标匹配 (智能对冲)", fg="green")
        f_r_top.grid(row=1, column=2, sticky="nsew", padx=(5, 0))
        self.shuttle_list_right = DraggableListbox(f_r_top, self, None, 'target', selectmode="extended")
        self.shuttle_list_right.pack(fill="both", expand=True)
        
        f_r_bot = tk.LabelFrame(f_main, text="🔴 剔除/例外 (独立导出)", fg="red")
        f_r_bot.grid(row=2, column=2, sticky="nsew", padx=(5, 0), pady=(5, 0))
        self.shuttle_list_exclude = DraggableListbox(f_r_bot, self, None, 'exclude', selectmode="extended")
        self.shuttle_list_exclude.pack(fill="both", expand=True)

        def move(target_name):
            items = [self.shuttle_list_left.get(i) for i in self.shuttle_list_left.curselection()]
            self.handle_drag_drop(items, target_name)
        
        def remove(from_where):
            if from_where == 'target':
                items = [self.shuttle_list_right.get(i) for i in self.shuttle_list_right.curselection()]
            else:
                items = [self.shuttle_list_exclude.get(i) for i in self.shuttle_list_exclude.curselection()]
            self.handle_drag_drop(items, 'source')

        f_btn_spacer = tk.Frame(f_main, width=160)
        f_btn_spacer.grid(row=0, column=1, sticky="nsew", padx=8)
        f_btn_target = tk.Frame(f_main, width=160)
        f_btn_target.grid(row=1, column=1, sticky="n", padx=8, pady=(10, 0))
        f_btn_exclude = tk.Frame(f_main, width=160)
        f_btn_exclude.grid(row=2, column=1, sticky="n", padx=8, pady=(10, 0))

        ttk.Button(f_btn_target, text="移入目标 >", width=18, command=lambda: move('target')).pack(pady=8, fill="x")
        ttk.Button(f_btn_target, text="移除 (从目标)", width=18, command=lambda: remove('target')).pack(pady=8, fill="x")
        ttk.Button(f_btn_exclude, text="移入剔除 >", width=18, command=lambda: move('exclude')).pack(pady=8, fill="x")
        ttk.Button(f_btn_exclude, text="移除 (从剔除)", width=18, command=lambda: remove('exclude')).pack(pady=8, fill="x")

        self.shuttle_search_var.trace("w", lambda *a: self.update_shuttle_ui())
        self.update_shuttle_ui()
        self._refresh_batch_list_ui()
        self.shuttle_batch_list.bind(
            "<<ListboxSelect>>",
            lambda e: (
                (lambda sel: self._switch_active_batch(sel[0]) if sel else None)(self.shuttle_batch_list.curselection())
            ),
        )

        def _shuttle_select_all(event=None):
            w = top.focus_get()
            if isinstance(w, tk.Listbox):
                w.selection_set(0, tk.END)
                return "break"
            if self.shuttle_list_left:
                self.shuttle_list_left.selection_set(0, tk.END)
            return "break"
        top.bind("<Control-a>", _shuttle_select_all)
        top.bind("<Control-A>", _shuttle_select_all)

    def update_shuttle_ui(self):
        if not self.shuttle_list_left: return
        self._sync_active_batch_from_target_accounts()
        self.shuttle_list_left.delete(0, tk.END); self.shuttle_list_right.delete(0, tk.END); self.shuttle_list_exclude.delete(0, tk.END)
        for x in sorted(list(self.target_accounts)): self.shuttle_list_right.insert(tk.END, x)
        for x in sorted(list(self.exclude_accounts)): self.shuttle_list_exclude.insert(tk.END, x)
        if not self.cached_accounts: return
        kw = self.shuttle_search_var.get().lower()
        c=0
        for x in self.cached_accounts:
            if x not in self.target_accounts and x not in self.exclude_accounts and (not kw or kw in str(x).lower()):
                self.shuttle_list_left.insert(tk.END, x); c+=1; 
                if c>1000: break

    # ==========================================
    # 5. 流程核心
    # ==========================================
    def _ask_split_count_on_main(self, suggested_count, total_rows):
        self.hide_progress() 
        msg = f"检测到导出数据量巨大 ({total_rows} 行)。\n\n为保证写入速度，透视表将单独存为Excel。\n所有明细数据将拆分为 {suggested_count} 个部分。"
        val = simpledialog.askinteger("极速写入模式", msg, initialvalue=suggested_count, minvalue=1, maxvalue=100, parent=self.root)
        self.user_split_count = val
        self.thread_event.set()
        if val: self.show_progress(f"正在全速写入 {val} 个分块...")

    def _ask_continue_detail_on_main(self, suite_path):
        self.hide_progress()
        name = os.path.basename(suite_path)
        msg = f"阶段1已完成，套表已导出：\n{name}\n\n是否继续导出凭证明细？"
        self.user_continue_detail = messagebox.askyesno("阶段1完成", msg, parent=self.root)
        self.thread_event.set()

    def start_process_flow(self):
        self._sync_active_batch_from_target_accounts()
        map_inv = {v: [] for v in self.ROLES.values()}
        for col, roles in self.column_mapping.items():
            for r in roles: map_inv[r].append(col) if r in map_inv else None
        
        if not map_inv['role_id'] or not map_inv['role_acc']: return messagebox.showerror("提示", "需标记ID和科目")
        
        # 逻辑判断
        logic = "B" if (map_inv['role_dr'] and map_inv['role_cr']) else ("A" if map_inv['role_amt'] else None)
        if not logic: return messagebox.showerror("提示", "缺少金额列")
        
        # === 修改开始 ===
        calc = ["#_净额(Net)"]
        
        # 逻辑：优先查找“金额”列，如果没找到，再查找“借贷”列，最后才用“净额”兜底
        default_vals = []
        if map_inv['role_amt']:
            default_vals.append(map_inv['role_amt'][0])
        elif map_inv['role_dr'] and map_inv['role_cr']:
            default_vals.append(map_inv['role_dr'][0])
            default_vals.append(map_inv['role_cr'][0])
        else:
            default_vals.append("#_净额(Net)")

        defaults = {
            'rows': map_inv['role_entity'] + map_inv['role_acc'], 
            'cols': [], 
            'vals': default_vals
        }
        # === 修改结束 ===
        
        # 默认不再预填“日期”为列字段
        if logic == "A" and map_inv['role_dir']: defaults['cols'].append(map_inv['role_dir'][0])

        # 方案B也改为弹窗确认（默认值预填）
        if logic == "B":
            acc_col = map_inv['role_acc'][0]
            # 方案B默认行字段仅保留科目列（不再合并ID）
            defaults['rows'] = [acc_col]
            defaults['cols'] = []
            if map_inv.get('role_dir'):
                defaults['cols'].append(map_inv['role_dir'][0])
            # 值字段：优先借/贷，没有则用金额
            if map_inv['role_dr'] and map_inv['role_cr']:
                defaults['vals'] = [map_inv['role_dr'][0], map_inv['role_cr'][0]]
            elif map_inv['role_amt']:
                defaults['vals'] = [map_inv['role_amt'][0]]
            else:
                defaults['vals'] = ["#_净额(Net)"]

        dlg = PivotDesignerDialog(self.root, self.full_columns, calc, defaults)
        self.root.wait_window(dlg.top)
        if dlg.action == "cancel": return
        self.pivot_config = dlg.result if dlg.action == "pivot" else None
        batches = self._get_effective_batches()
        threading.Thread(target=self.run_export, args=(map_inv, logic, batches)).start()

    # ==============================
    # 透视/凭证输出辅助函数
    # ==============================
    def _make_join_col(self, df, cols):
        if len(cols) == 1:
            return cols[0]
        join_col = "-".join(cols)
        if join_col not in df.columns:
            df[join_col] = df[cols].fillna("").astype(str).agg("-".join, axis=1)
        return join_col

    def _get_voucher_id_cols(self, map_inv, df=None):
        cols = []
        for c in map_inv.get('role_entity', []):
            if df is None or c in df.columns:
                cols.append(c)
        for c in map_inv.get('role_id', []):
            if df is None or c in df.columns:
                cols.append(c)
        return list(dict.fromkeys(cols)) or list(map_inv.get('role_id', []))

    def _safe_convert_date(self, series):
        s_clean = series.fillna("").astype(str).str.replace(r'\.0$', '', regex=True)
        s_num = pd.to_numeric(s_clean, errors='coerce')
        mask_8digit = s_num.notna() & (s_num > 19000000) & (s_num < 21000000)
        dt_8digit = pd.to_datetime(s_num[mask_8digit].astype(int).astype(str), format='%Y%m%d', errors='coerce')
        mask_serial = s_num.notna() & (s_num > 1000) & (s_num < 100000)
        dt_serial = pd.to_datetime(s_num[mask_serial], unit='D', origin='1899-12-30', errors='coerce')
        dt_std = pd.to_datetime(s_clean[~mask_serial & ~mask_8digit], errors='coerce')
        final_dt = dt_serial.reindex(series.index).combine_first(dt_8digit).combine_first(dt_std)
        return final_dt.dt.strftime('%Y-%m').fillna("Unknown")

    def _to_polars_df_safe(self, frame):
        if frame is None:
            return pl.DataFrame()
        if len(frame.columns) == 0:
            return pl.DataFrame()
        data = {}
        for c in frame.columns:
            vals = []
            for v in frame[c].tolist():
                try:
                    vals.append(None if pd.isna(v) else v)
                except Exception:
                    vals.append(v)
            data[str(c)] = vals
        return pl.DataFrame(data)

    def _to_pandas_df_safe(self, pl_df):
        try:
            return pl_df.to_pandas(use_pyarrow_extension_array=False)
        except Exception:
            return pd.DataFrame(pl_df.to_dict(as_series=False), columns=pl_df.columns)

    def _pl_clean_num_series(self, series):
        vals = ["" if v is None else str(v) for v in series.tolist()]
        ps = pl.Series("v", vals)
        ps = ps.str.replace_all(r'[,"]', "").cast(pl.Float64, strict=False).fill_null(0.0)
        return pd.Series(ps.to_list(), index=series.index)

    def _ensure_net_column_polars(self, frame, map_inv, logic):
        if frame is None or frame.empty:
            if frame is None:
                return frame
            out = frame.copy()
            if "__net__" not in out.columns:
                out["__net__"] = []
            return out
        if "__net__" in frame.columns:
            return frame

        out = frame.copy()
        voucher_id_cols = self._get_voucher_id_cols(map_inv, out)
        id_col = self._make_join_col(out, voucher_id_cols) if voucher_id_cols else None

        if logic == "B":
            dr_col = map_inv['role_dr'][0]
            cr_col = map_inv['role_cr'][0]
            val_dr = self._pl_clean_num_series(out[dr_col])
            val_cr = self._pl_clean_num_series(out[cr_col])
            both_mask = (val_dr != 0) & (val_cr != 0)
            raw_amt = np.where(both_mask, val_dr - val_cr, np.where(val_dr != 0, val_dr, val_cr))
            raw_amt = pd.Series(raw_amt, index=out.index)

            sample_id = None
            if id_col and id_col in out.columns:
                for vid in pd.unique(out[id_col]):
                    m = (out[id_col] == vid)
                    if (val_dr[m] != 0).any() and (val_cr[m] != 0).any():
                        sample_id = vid
                        break
            is_already_balanced = False
            if sample_id is not None:
                sample_mask = (out[id_col] == sample_id)
                sample_sum = raw_amt[sample_mask].sum()
                is_already_balanced = (abs(sample_sum) < 0.01) and (sample_mask.sum() > 1)
            out["__net__"] = raw_amt if is_already_balanced else (val_dr - val_cr)
            return out

        amt_col = map_inv['role_amt'][0]
        raw_amt = self._pl_clean_num_series(out[amt_col])
        is_already_balanced = False
        sample_id = None
        if id_col and id_col in out.columns:
            if map_inv.get('role_dir'):
                dir_col = map_inv['role_dir'][0]
                dir_vals = ["" if v is None else str(v) for v in out[dir_col].tolist()]
                is_credit_s = pl.Series("d", dir_vals).str.contains(r'(?:贷|贷方|Credit|Cr\b|^C$|^H$|[-−])', literal=False)
                is_credit = pd.Series(is_credit_s.fill_null(False).to_list(), index=out.index)
                is_debit = (~is_credit) & out[dir_col].astype(str).str.strip().ne('')
                for vid in pd.unique(out[id_col]):
                    m = (out[id_col] == vid)
                    if is_credit[m].any() and is_debit[m].any():
                        sample_id = vid
                        break
            if sample_id is None:
                sample_id = out[id_col].iloc[0] if len(out) > 0 else None
            if sample_id is not None:
                sample_mask = (out[id_col] == sample_id)
                sample_sum = raw_amt[sample_mask].sum()
                is_already_balanced = (abs(sample_sum) < 0.01) and (sample_mask.sum() > 1)

        if is_already_balanced:
            out["__net__"] = raw_amt
        else:
            if map_inv.get('role_dir'):
                dir_col = map_inv['role_dir'][0]
                dir_vals = ["" if v is None else str(v) for v in out[dir_col].tolist()]
                is_credit_s = pl.Series("d", dir_vals).str.contains(r'(?:贷|贷方|Credit|Cr\b|^C$|^H$|[-−])', literal=False)
                is_credit = pd.Series(is_credit_s.fill_null(False).to_list(), index=out.index)
                out["__net__"] = np.where(is_credit, raw_amt * -1, raw_amt)
            else:
                out["__net__"] = raw_amt
        return out

    def _flatten_columns(self, cols):
        out = []
        for c in cols:
            if isinstance(c, tuple):
                parts = [str(x) for x in c if x not in (None, "", "nan")]
                out.append("-".join(parts) if parts else "")
            else:
                out.append(str(c))
        return out

    def _flatten_pivot_columns(self, cols):
        out = []
        for c in cols:
            if isinstance(c, tuple):
                parts = [str(x) for x in c if x not in (None, "", "nan")]
                if len(parts) >= 2:
                    # 透视表默认层级是 值字段->列字段，这里调整为 列字段->值字段
                    out.append("-".join(parts[1:] + [parts[0]]))
                else:
                    out.append("-".join(parts) if parts else "")
            else:
                out.append(str(c))
        return out

    def _reorder_voucher_type_columns(self, df, type_col=None, id_col=None, acc_col=None):
        cols = list(df.columns)

        def _pick(preferred, fallback):
            if preferred and preferred in cols:
                return preferred
            if fallback in cols:
                return fallback
            return None

        type_c = _pick(type_col, "科目名称-类型")
        id_c = _pick(id_col, "唯一识别码")
        acc_c = _pick(acc_col, "科目名称")

        ordered = []
        for c in [type_c, id_c, "摘要", acc_c]:
            if c and c in cols and c not in ordered:
                ordered.append(c)

        def _is_month_col(x):
            s = str(x).strip()
            return len(s) == 7 and s[4] == '-' and s[:4].isdigit() and s[5:].isdigit()

        net_cols = [c for c in cols if "#_净额(Net)" in str(c)]
        month_cols = [c for c in cols if _is_month_col(c)]
        rest = [c for c in cols if c not in ordered and c not in net_cols and c not in month_cols]
        new_cols = ordered + net_cols + month_cols + rest
        return df[new_cols]

    def _collect_summary_map(self, df, id_col_name, summary_col):
        if not summary_col or summary_col not in df.columns:
            return {}
        id_vals = df[id_col_name].astype(str)
        summaries = df[summary_col].fillna("").astype(str)
        summary_map = {}
        for vid, txt in zip(id_vals, summaries):
            if not txt:
                continue
            bucket = summary_map.setdefault(vid, [])
            if txt not in bucket:
                bucket.append(txt)
        return summary_map

    def build_voucher_pivot(self, df_target, map_inv, logic):
        if df_target is None or df_target.empty:
            return None, None, None
        id_cols = self._get_voucher_id_cols(map_inv, df_target)
        acc_cols = map_inv['role_acc']
        id_join_col = self._make_join_col(df_target, id_cols)
        acc_join_col = self._make_join_col(df_target, acc_cols)

        # 凭证类型统一使用净额Net
        val_fields = ["#_净额(Net)"]

        col_fields = []
        if map_inv.get('role_dir'):
            col_fields.append(map_inv['role_dir'][0])

        p_df = self._ensure_net_column_polars(df_target.copy(), map_inv, logic)
        p_df['#_净额(Net)'] = p_df['__net__']
        need_cols = [id_join_col, acc_join_col] + (col_fields if col_fields else []) + ['#_净额(Net)']
        need_cols = [c for c in need_cols if c in p_df.columns]
        pl_df = self._to_polars_df_safe(p_df[need_cols])
        agg_col = '#_净额(Net)'
        if col_fields:
            gb_cols = [id_join_col, acc_join_col] + col_fields
            grp_pd = self._to_pandas_df_safe(
                pl_df.group_by(gb_cols)
                .agg(pl.col(agg_col).cast(pl.Float64, strict=False).fill_null(0.0).sum().alias(agg_col))
            )
            v_pivot = pd.pivot_table(
                grp_pd,
                index=[id_join_col, acc_join_col],
                columns=col_fields if col_fields else None,
                values=agg_col,
                aggfunc='sum'
            ).fillna(0)
        else:
            grp_pd = self._to_pandas_df_safe(
                pl_df.group_by([id_join_col, acc_join_col])
                .agg(pl.col(agg_col).cast(pl.Float64, strict=False).fill_null(0.0).sum().alias(agg_col))
            )
            v_pivot = grp_pd.set_index([id_join_col, acc_join_col])
        return v_pivot, id_join_col, acc_join_col

    def build_voucher_type_pivot(self, v_pivot, df_target, map_inv, loss_ids=None, mode="normal", target_accounts=None):
        if v_pivot is None or not isinstance(v_pivot.index, pd.MultiIndex):
            return None, None, None, []
        if loss_ids is None:
            loss_ids = set()

        id_col_name = v_pivot.index.names[0]
        acc_col_name = v_pivot.index.names[1]
        cols = list(v_pivot.columns)

        # 未筛选目标科目时，不生成“凭证类型”
        effective_targets = set(target_accounts if target_accounts is not None else self.target_accounts)
        if not effective_targets:
            return None, None, None, []

        # 损益结转：先按ID过滤，再做汇总
        df_target_f = df_target
        if loss_ids:
            loss_ids_str = {str(x) for x in loss_ids}
            v_pivot = v_pivot[~v_pivot.index.get_level_values(id_col_name).astype(str).isin(loss_ids_str)]
            if v_pivot.empty:
                return None, None, None, []
            id_join_col = self._make_join_col(df_target_f, self._get_voucher_id_cols(map_inv, df_target_f))
            df_target_f = df_target_f[~df_target_f[id_join_col].astype(str).isin(loss_ids_str)].copy()

        norm_cache = {}
        def _norm_acc(val):
            s = str(val)
            if s in norm_cache:
                return norm_cache[s]
            v = re.sub(r"\s*-\s*", "-", s).strip()
            norm_cache[s] = v
            return v

        target_acc_norm = {_norm_acc(v) for v in effective_targets}
        _v_reset = v_pivot.reset_index(drop=False)
        _idx_names = list(v_pivot.index.names)
        _idx_names = [n if n is not None else f"__idx_{i}__" for i, n in enumerate(_idx_names)]
        _v_reset.columns = _idx_names + list(v_pivot.columns)
        pl_tmp = self._to_polars_df_safe(_v_reset)
        val_cols = [c for c in _v_reset.columns if c not in _idx_names]
        if val_cols:
            pl_tmp = pl_tmp.with_columns([
                pl.col(c).cast(pl.Float64, strict=False).fill_null(0.0).alias(c) for c in val_cols
            ])
        _tmp_pd = self._to_pandas_df_safe(pl_tmp)
        v_pivot_num = _tmp_pd.set_index(_idx_names)
        v_pivot_num.index.names = v_pivot.index.names


        # 1) 生成同类凭证分组：
        #    - “基准集合”= 全部科目集合的最小集合（基于现有凭证）
        #    - 仅当存在对应基准集合时才归类（避免无基准时的误合并）
        #    - 方向一致性：仅比较目标科目符号（以输出口径为准，净额为0的不计）
        voucher_info = []
        for vid, sub in v_pivot_num.groupby(level=id_col_name, sort=False):
            if isinstance(sub.index, pd.MultiIndex):
                sub = sub.droplevel(id_col_name)
            # 以输出口径为准：净额四舍五入到2位，小于等于0视为0，不计入科目集合
            net_vals = sub.sum(axis=1).round(2)
            nz_mask = net_vals.abs() > 0
            sub_nz = sub.loc[nz_mask]
            acc_list = [str(x) for x in sub_nz.index.tolist()]
            acc_set = set(acc_list)
            full_keys_set = frozenset(acc_set)
            signs = np.sign(net_vals.loc[sub_nz.index].to_numpy()).astype(int)
            target_keys = set()
            target_signs = {}
            for acc, s in zip(acc_list, signs):
                if _norm_acc(acc) in target_acc_norm:
                    target_keys.add(acc)
                    target_signs[acc] = int(s)
            target_keys_set = frozenset(target_keys)
            voucher_info.append({
                "vid": vid,
                "vid_str": str(vid),
                "acc_set": acc_set,
                "acc_size": len(acc_set),
                "full_keys_set": full_keys_set,
                "target_signs": target_signs,
                "target_keys": target_keys,
                "target_keys_set": target_keys_set,
            })

        # 凭证类型只分析命中目标科目的凭证。没有目标科目的凭证即使被前序数据带入，
        # 也不能参与类型归并，否则会出现“类型无目标科目”或金额只像代表凭证的错觉。
        voucher_info = [info for info in voucher_info if info["target_keys_set"]]
        if not voucher_info:
            return None, None, None, []

        n = len(voucher_info)
        parent = list(range(n))

        def _find(x):
            while parent[x] != x:
                parent[x] = parent[parent[x]]
                x = parent[x]
            return x

        def _union(a, b):
            ra, rb = _find(a), _find(b)
            if ra != rb:
                parent[rb] = ra

        def _set_order_key(s):
            # 优先元素更多，其次按字典序（稳定）
            return (-len(s), tuple(sorted([str(x) for x in s])))

        def _compatible_signs(group_map, acc_map):
            # 仅要求共同出现的目标科目符号一致
            for acc, s in acc_map.items():
                if acc in group_map and group_map[acc] != s:
                    return False
            return True

        def _minimal_sets(set_list):
            uniq = []
            seen = set()
            for s in set_list:
                if not s or s in seen:
                    continue
                seen.add(s)
                uniq.append(s)
            uniq.sort(key=lambda s: (len(s), tuple(sorted([str(x) for x in s]))))
            mins = []
            for s in uniq:
                if any(m.issubset(s) for m in mins):
                    continue
                mins.append(s)
            return mins

        if mode == "strict":
            # 严格：按“同一ID下目标科目数量”分流
            # - 目标科目数>1：用目标集合做最小集
            # - 目标科目数=1：用全科目集合做最小集

            target_set_keys = [
                info["target_keys_set"]
                for info in voucher_info
                if info["target_keys_set"] and len(info["target_keys_set"]) > 1
            ]
            base_target_sets = _minimal_sets(target_set_keys)
            base_target_sets_desc = sorted(base_target_sets, key=_set_order_key)
            base_target_set_lookup = set(base_target_sets)

            full_set_keys = [
                info["full_keys_set"]
                for info in voucher_info
                if info["full_keys_set"] and len(info["target_keys_set"]) == 1
            ]
            base_full_sets = _minimal_sets(full_set_keys)
            base_full_sets_desc = sorted(base_full_sets, key=_set_order_key)
            base_full_set_lookup = set(base_full_sets)

            def _pick_base_target_set(tset):
                for b in base_target_sets_desc:
                    if b.issubset(tset):
                        return b
                return None

            def _pick_base_full_set(fset):
                for b in base_full_sets_desc:
                    if b.issubset(fset):
                        return b
                return None

            def _has_base_set(info):
                tset = info["target_keys_set"]
                if not tset:
                    return False
                if len(tset) > 1:
                    return _pick_base_target_set(tset) is not None
                fset = info["full_keys_set"]
                if not fset:
                    return False
                return _pick_base_full_set(fset) is not None

            # 仅作用于“存在 base_set 的凭证”
            primary_idxs = {i for i, info in enumerate(voucher_info) if _has_base_set(info)}

            # 基准集合分组：仅“集合恰好等于最小集合”的凭证可建基准组
            base_groups_target = {}
            if base_target_sets:
                for idx, info in enumerate(voucher_info):
                    tset = info["target_keys_set"]
                    if not tset or len(tset) <= 1 or tset not in base_target_set_lookup:
                        continue
                    groups = base_groups_target.setdefault(tset, [])
                    compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                    if len(compatible) == 0:
                        groups.append({
                            "members": [idx],
                            "sign_map": dict(info["target_signs"]),
                        })

            base_groups_full = {}
            if base_full_sets:
                for idx, info in enumerate(voucher_info):
                    tset = info["target_keys_set"]
                    if not tset or len(tset) != 1:
                        continue
                    fset = info["full_keys_set"]
                    if not fset or fset not in base_full_set_lookup:
                        continue
                    groups = base_groups_full.setdefault(fset, [])
                    compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                    if len(compatible) == 0:
                        groups.append({
                            "members": [idx],
                            "sign_map": dict(info["target_signs"]),
                        })

            # 仅当存在“基准集合”时归类
            for idx in primary_idxs:
                info = voucher_info[idx]
                tset = info["target_keys_set"]
                if not tset:
                    continue
                if len(tset) > 1:
                    base_tset = _pick_base_target_set(tset)
                    if base_tset is None:
                        continue
                    groups = base_groups_target.get(base_tset, [])
                else:
                    fset = info["full_keys_set"]
                    base_fset = _pick_base_full_set(fset) if fset else None
                    if base_fset is None:
                        continue
                    groups = base_groups_full.get(base_fset, [])

                compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                if len(compatible) != 1:
                    continue
                g = compatible[0]
                _union(idx, g["members"][0])
                if idx not in g["members"]:
                    g["members"].append(idx)
                for acc, s in info["target_signs"].items():
                    if acc not in g["sign_map"]:
                        g["sign_map"][acc] = s

            # 兜底：同一全科目集合且目标科目数=1时，若符号兼容则直接合并
            one_target_idxs = [i for i, info in enumerate(voucher_info) if len(info["target_keys_set"]) == 1]
            if one_target_idxs:
                by_full = {}
                for i in one_target_idxs:
                    fset = voucher_info[i]["full_keys_set"]
                    by_full.setdefault(fset, []).append(i)
                for fset, idxs in by_full.items():
                    if len(idxs) <= 1:
                        continue
                    base = idxs[0]
                    for j in idxs[1:]:
                        if _compatible_signs(voucher_info[base]["target_signs"], voucher_info[j]["target_signs"]) and \
                           _compatible_signs(voucher_info[j]["target_signs"], voucher_info[base]["target_signs"]):
                            _union(base, j)

        else:
            # 识别最小“基准集合”（目标科目集合优先）
            target_set_keys = [info["target_keys_set"] for info in voucher_info if info["target_keys_set"]]
            base_target_sets = _minimal_sets(target_set_keys)
            base_target_sets_desc = sorted(base_target_sets, key=_set_order_key)

            def _pick_base_target_set(tset):
                for b in base_target_sets_desc:
                    if b.issubset(tset):
                        return b
                return None

            def _has_base_target_set(info):
                tset = info["target_keys_set"]
                if not tset:
                    return False
                return _pick_base_target_set(tset) is not None

            # 第二阶段仅作用于“存在 base_target_set 的凭证”（不含 fallback）
            primary_idxs = {i for i, info in enumerate(voucher_info) if _has_base_target_set(info)}

            # 识别最小“基准集合”（全部科目集合）- 仅基于 primary_idxs
            full_set_keys = [voucher_info[i]["full_keys_set"] for i in primary_idxs if voucher_info[i]["full_keys_set"]]
            base_full_sets = _minimal_sets(full_set_keys)
            base_full_sets_desc = sorted(base_full_sets, key=_set_order_key)

            def _pick_base_full_set(fset):
                for b in base_full_sets_desc:
                    if b.issubset(fset):
                        return b
                return None

            # 基准集合分组：先按目标科目集合建组（组内仅比较目标科目符号）
            base_groups_target = {}
            if base_target_sets:
                for idx, info in enumerate(voucher_info):
                    tset = info["target_keys_set"]
                    base_tset = _pick_base_target_set(tset) if tset else None
                    if not tset or base_tset is None:
                        continue
                    groups = base_groups_target.setdefault(base_tset, [])
                    compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                    if len(compatible) == 0:
                        groups.append({
                            "members": [idx],
                            "sign_map": dict(info["target_signs"]),
                        })

                # 仅当存在“基准集合”时归类
                for idx, info in enumerate(voucher_info):
                    tset = info["target_keys_set"]
                    base_tset = _pick_base_target_set(tset) if tset else None
                    if not tset or base_tset is None:
                        continue
                    groups = base_groups_target.get(base_tset, [])
                    compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                    if len(compatible) != 1:
                        continue
                    g = compatible[0]
                    _union(idx, g["members"][0])
                    if idx not in g["members"]:
                        g["members"].append(idx)
                    for acc, s in info["target_signs"].items():
                        if acc not in g["sign_map"]:
                            g["sign_map"][acc] = s

            # 基准集合分组：再按全部科目集合建组（第二阶段：仅作用于存在 base_target_set 的凭证）
            base_groups_full = {}
            # 先用“全科目包含基准集合”的凭证建组
            for idx in primary_idxs:
                info = voucher_info[idx]
                fset = info["full_keys_set"]
                base_fset = _pick_base_full_set(fset) if fset else None
                if not fset or base_fset is None:
                    continue
                groups = base_groups_full.setdefault(base_fset, [])
                compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                if len(compatible) == 0:
                    groups.append({
                        "members": [idx],
                        "sign_map": dict(info["target_signs"]),
                    })

            # 仅当存在“基准集合”时归类（第二阶段）
            for idx in primary_idxs:
                info = voucher_info[idx]
                fset = info["full_keys_set"]
                base_fset = _pick_base_full_set(fset) if fset else None
                if not fset or base_fset is None:
                    continue
                groups = base_groups_full.get(base_fset, [])
                compatible = [g for g in groups if _compatible_signs(g["sign_map"], info["target_signs"])]
                if len(compatible) != 1:
                    continue
                g = compatible[0]
                _union(idx, g["members"][0])
                if idx not in g["members"]:
                    g["members"].append(idx)
                for acc, s in info["target_signs"].items():
                    if acc not in g["sign_map"]:
                        g["sign_map"][acc] = s

        groups = {}
        for idx, info in enumerate(voucher_info):
            root = _find(idx)
            groups.setdefault(root, []).append(info["vid_str"])
        type_groups = list(groups.values())
        # 使用字符串键，保证后续map一致
        rep_map = {vid: group[0] for group in type_groups for vid in group}

        # 2) 汇总同类凭证
        reset = v_pivot_num.reset_index()
        reset_id = reset[id_col_name].astype(str)
        reset["__rep_id__"] = reset_id.map(rep_map)
        reset = reset[reset["__rep_id__"].notna()].copy()
        if reset.empty:
            return None, None, None, []
        grouped = reset.groupby(["__rep_id__", acc_col_name], dropna=False).sum(numeric_only=True)
        grouped = grouped.rename_axis(index={"__rep_id__": id_col_name})

        # 3) 为目标科目生成类型（以整凭证签名区分）
        orig = v_pivot_num
        orig_ids = orig.index.get_level_values(id_col_name).astype(str)
        orig_acc = orig.index.get_level_values(acc_col_name).astype(str)
        orig_net = orig.sum(axis=1).round(2).to_numpy()

        accs_per_vid = {}
        for oid, acc, net in zip(orig_ids, orig_acc, orig_net):
            if abs(net) <= 0:
                continue
            accs_per_vid.setdefault(oid, set()).add(acc)

        # 类型编号基于“目标科目”，但标签应用到同类型的全部科目行
        acc_sig_set = {}
        for oid, accs in accs_per_vid.items():
            rep_id = rep_map.get(str(oid))
            if rep_id is None:
                continue
            for acc in accs:
                if _norm_acc(acc) not in target_acc_norm:
                    continue
                acc_sig_set.setdefault(acc, set()).add(rep_id)

        type_rank_map = {}
        for acc, rep_ids in acc_sig_set.items():
            for idx, rep_id in enumerate(sorted(rep_ids, key=lambda x: str(x)), start=1):
                type_rank_map[(acc, rep_id)] = idx

        # 以类型(rep_id)为单位生成“科目名称-类型”，应用到该类型下所有科目
        label_map = {}
        for group in type_groups:
            rep_id = group[0]
            accs_in_group = set()
            for oid in group:
                accs_in_group.update(accs_per_vid.get(oid, set()))
            accs_target = [a for a in accs_in_group if _norm_acc(a) in target_acc_norm]
            labels = []
            for acc in sorted(accs_target, key=lambda x: str(x)):
                t_idx = type_rank_map.get((acc, rep_id), 1)
                labels.append(f"{acc}-类型{t_idx}")
            if labels:
                label_map[str(rep_id)] = " | ".join(labels)

        # 4) 汇总摘要（去重后取前三条）
        summary_col = map_inv.get('role_summary', [None])[0] if map_inv.get('role_summary') else None
        summary_map = {}
        if summary_col and summary_col in df_target_f.columns:
            id_join_col = self._make_join_col(df_target_f, self._get_voucher_id_cols(map_inv, df_target_f))
            summary_map = self._collect_summary_map(df_target_f, id_join_col, summary_col)

        rep_summaries = {}
        for group in type_groups:
            rep_id = str(group[0])
            buf = []
            for oid in group:
                for txt in summary_map.get(str(oid), []):
                    if txt not in buf:
                        buf.append(txt)
                    if len(buf) >= 3:
                        break
                if len(buf) >= 3:
                    break
            rep_summaries[rep_id] = " | ".join(buf)

        # 如果存在方向列，则将方向展开为行，插入“方向”列
        dir_col_name = map_inv.get('role_dir', [None])[0] if map_inv.get('role_dir') else None
        has_dir = False
        if dir_col_name and isinstance(grouped.columns, pd.MultiIndex):
            has_dir = True
            names = list(grouped.columns.names)
            if dir_col_name in names:
                dir_level = names.index(dir_col_name)
            else:
                dir_level = len(names) - 1
            stacked = grouped.stack(level=dir_level)
            stacked.index.names = [id_col_name, acc_col_name, dir_col_name]
            flat = stacked.reset_index()
        else:
            flat = grouped.reset_index()
        # flatten columns to avoid merge level mismatch
        flat.columns = self._flatten_columns(flat.columns)
        # 确保列名唯一
        seen = {}
        new_cols = []
        for c in flat.columns:
            if c not in seen:
                seen[c] = 0
                new_cols.append(c)
            else:
                seen[c] += 1
                new_cols.append(f"{c}_{seen[c]}")
        flat.columns = new_cols
        # 避免已有同名列导致不唯一
        if "科目名称-类型" in flat.columns:
            flat = flat.drop(columns=["科目名称-类型"])
        # 先准备类型列名（稍后写入）
        type_col_name = "科目名称-类型"
        if "摘要" in flat.columns:
            flat = flat.drop(columns=["摘要"])
        flat["摘要"] = flat[id_col_name].astype(str).map(rep_summaries).fillna("")

        # 5) 合并YYYY-MM分布列
        month_cols = []
        date_cols = map_inv.get('role_date', [])
        if date_cols and date_cols[0] in df_target_f.columns and '__net__' in df_target_f.columns:
            date_col = date_cols[0]
            df_m = df_target_f.copy()
            id_join_col = self._make_join_col(df_m, self._get_voucher_id_cols(map_inv, df_m))
            df_m['__month__'] = self._safe_convert_date(df_m[date_col])
            df_m['__rep_id__'] = df_m[id_join_col].astype(str).map(rep_map)
            df_m = df_m[df_m['__rep_id__'].notna()]
            df_m['__net__'] = pd.to_numeric(df_m['__net__'], errors='coerce').fillna(0)
            # 按（rep_id, 科目, 方向）汇总月份
            acc_key = acc_col_name
            if acc_key not in df_m.columns:
                acc_key = self._make_join_col(df_m, map_inv['role_acc'])
                if acc_key != acc_col_name:
                    df_m[acc_col_name] = df_m[acc_key]
                    acc_key = acc_col_name
            index_keys = ['__rep_id__', acc_key]
            if has_dir and dir_col_name in df_m.columns:
                index_keys.append(dir_col_name)
            month_pivot = pd.pivot_table(
                df_m,
                index=index_keys,
                columns='__month__',
                values='__net__',
                aggfunc='sum'
            ).fillna(0)
            # 仅保留YYYY-MM格式
            month_cols = [c for c in month_pivot.columns if isinstance(c, str) and len(c) == 7 and c[4] == '-']
            if month_cols:
                month_pivot = month_pivot[month_cols].reset_index().rename(columns={'__rep_id__': id_col_name})
                merge_keys = [id_col_name, acc_col_name]
                if has_dir and dir_col_name in flat.columns and dir_col_name in month_pivot.columns:
                    merge_keys.append(dir_col_name)
                flat = flat.merge(month_pivot, on=merge_keys, how='left')

        # 6) 去除损益结转后重新编号：保证同科目从类型1开始
        rep_id_series = flat[id_col_name]
        if isinstance(rep_id_series, pd.DataFrame):
            rep_id_series = rep_id_series.iloc[:, 0]
        remaining_rep_ids = set(rep_id_series.astype(str).tolist())
        acc_sig_set_rem = {}
        for group in type_groups:
            rep_id = str(group[0])
            if rep_id not in remaining_rep_ids:
                continue
            accs_in_group = set()
            for oid in group:
                accs_in_group.update(accs_per_vid.get(oid, set()))
            for acc in accs_in_group:
                if _norm_acc(acc) not in target_acc_norm:
                    continue
                acc_sig_set_rem.setdefault(acc, set()).add(rep_id)
        type_rank_map_rem = {}
        for acc, rep_ids in acc_sig_set_rem.items():
            for idx, rep_id in enumerate(sorted(rep_ids, key=lambda x: str(x)), start=1):
                type_rank_map_rem[(acc, rep_id)] = idx
        label_map_final = {}
        for group in type_groups:
            rep_id = str(group[0])
            if rep_id not in remaining_rep_ids:
                continue
            accs_in_group = set()
            for oid in group:
                accs_in_group.update(accs_per_vid.get(oid, set()))
            accs_target = [a for a in accs_in_group if _norm_acc(a) in target_acc_norm]
            labels = []
            for acc in sorted(accs_target, key=lambda x: str(x)):
                t_idx = type_rank_map_rem.get((acc, rep_id), 1)
                labels.append(f"{acc}-类型{t_idx}")
            if labels:
                label_map_final[rep_id] = " | ".join(labels)

        # 写入“科目名称-类型”
        if type_col_name in flat.columns:
            flat = flat.drop(columns=[type_col_name])
        flat[type_col_name] = rep_id_series.astype(str).map(label_map_final).fillna("")
        # 去掉可能残留的“科目名称-类型_*”
        dup_type_cols = [c for c in flat.columns if str(c).startswith("科目名称-类型") and c != type_col_name]
        if dup_type_cols:
            flat = flat.drop(columns=dup_type_cols)

        # 8) 调整列顺序：科目名称-类型置前，方向列紧随科目名称
        base_cols = [type_col_name, id_col_name, "摘要", acc_col_name]
        if has_dir and dir_col_name in flat.columns:
            base_cols.append(dir_col_name)
        value_cols = [c for c in flat.columns if c not in base_cols]
        flat = flat[base_cols + value_cols]

        # 仅保留净额Net列 + YYYY-MM列（移除原借贷/金额列）
        keep_cols = [c for c in flat.columns if "#_净额(Net)" in str(c)]
        def _is_month_col_keep(x):
            s = str(x).strip()
            return len(s) == 7 and s[4] == '-' and s[:4].isdigit() and s[5:].isdigit()
        month_keep = [c for c in flat.columns if _is_month_col_keep(c)]
        if keep_cols or month_keep:
            non_value_cols = [c for c in flat.columns if c not in value_cols]
            flat = flat[non_value_cols + keep_cols + month_keep]

        # 8.1) 同一科目+同一类型合并列示（净额为0不展示）
        num_cols = [c for c in flat.columns if c in keep_cols or c in month_keep]
        def _first_non_empty(s):
            for v in s:
                if str(v).strip():
                    return v
            return ""
        group_keys = [type_col_name, acc_col_name]
        if has_dir and dir_col_name in flat.columns:
            group_keys.append(dir_col_name)
        agg = {}
        for c in num_cols:
            agg[c] = "sum"
        for c in [id_col_name, "摘要"]:
            if c in flat.columns:
                agg[c] = _first_non_empty
        # 保留分组列本身
        for c in group_keys:
            if c not in agg:
                agg[c] = "first"
        flat = flat.groupby(group_keys, dropna=False, as_index=False).agg(agg)
        if num_cols:
            flat = flat[(flat[num_cols].abs().sum(axis=1) != 0)]

        # 9) 排序：按科目名称 + 类型数字（降序）
        tmp_acc = "__tmp_type_acc__"
        tmp_num = "__tmp_type_num__"
        col_data = flat[type_col_name]
        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]
        flat[tmp_acc] = col_data.astype(str).apply(
            lambda x: x.split("-类型")[0] if "-类型" in x else x
        )
        def _type_num(x):
            if "-类型" in x:
                try:
                    return int(x.split("-类型")[-1])
                except Exception:
                    return 0
            return 0
        flat[tmp_num] = col_data.astype(str).apply(_type_num)
        flat = flat.sort_values(by=[tmp_acc, tmp_num], ascending=[False, False], kind='stable')
        flat = flat.drop(columns=[tmp_acc, tmp_num])
        # DEBUG: 输出列名用于排查重复
        try:
            cols_list = list(flat.columns)
            dups = [c for c in cols_list if cols_list.count(c) > 1]
            print(f"DEBUG[凭证类型-构建]: 列名={cols_list}", flush=True)
            if dups:
                print(f"DEBUG[凭证类型-构建]: 重复列名={sorted(set(dups))}", flush=True)
        except Exception:
            pass
        return flat, id_col_name, type_col_name, month_cols

    def _build_batch_save_path(self, base_path, batch_name, idx):
        base_name, ext = os.path.splitext(base_path)
        safe_batch = re.sub(r'[\\/:*?"<>|]+', "_", str(batch_name)).strip()
        safe_batch = re.sub(r"\s+", "_", safe_batch) or f"批次{idx}"
        return f"{base_name}_{idx:02d}_{safe_batch}{ext}"

    def _build_default_save_name(self, ext=".csv"):
        src = self.real_xlsx_path or self.file_path or ""
        stem = os.path.splitext(os.path.basename(src))[0] if src else "未命名"
        parts = ["看账导出", stem]
        if self.current_sheet_name not in (None, 0, "0"):
            parts.append(f"工作表{self.current_sheet_name}")
        parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
        raw_name = "_".join(str(p).strip() for p in parts if str(p).strip())
        safe_name = re.sub(r'[\\/:*?"<>|]+', "_", raw_name)
        safe_name = re.sub(r"\s+", "_", safe_name).strip("._ ") or "看账导出"
        return f"{safe_name}{ext}"

    def _default_save_initialdir(self):
        src = self.real_xlsx_path or self.file_path
        if src:
            folder = os.path.dirname(os.path.abspath(src))
            if os.path.isdir(folder):
                return folder
        return None

    def _preprocess_and_filter_with_polars(self, df, map_inv):
        if not HAS_POLARS:
            raise RuntimeError("polars不可用")

        id_cols = self._get_voucher_id_cols(map_inv, df)
        acc_cols = [c for c in map_inv.get('role_acc', []) if c in df.columns]
        if not id_cols or any(c not in df.columns for c in id_cols):
            raise RuntimeError("ID列缺失，无法使用polars快速路径")
        if not acc_cols:
            raise RuntimeError("科目列缺失，无法使用polars快速路径")

        mapped_cols = []
        for role_key, r_cols in map_inv.items():
            if role_key in ("role_amt", "role_dr", "role_cr"):
                continue
            mapped_cols.extend([c for c in r_cols if c in df.columns])
        mapped_cols = list(dict.fromkeys(mapped_cols))

        amt_cols = []
        amt_cols += [c for c in map_inv.get('role_amt', []) if c in df.columns]
        amt_cols += [c for c in map_inv.get('role_dr', []) if c in df.columns]
        amt_cols += [c for c in map_inv.get('role_cr', []) if c in df.columns]
        amt_cols = list(dict.fromkeys(amt_cols))

        pl_df = self._to_polars_df_safe(df)
        all_cols = list(pl_df.columns)

        def _blank_to_null_expr(col_name):
            return pl.when(
                pl.col(col_name).is_null() | (pl.col(col_name).cast(pl.Utf8).str.strip_chars() == "")
            ).then(None).otherwise(pl.col(col_name).cast(pl.Utf8))

        if mapped_cols:
            repl_exprs = [_blank_to_null_expr(c).alias(c) for c in mapped_cols if c in all_cols]
            if repl_exprs:
                pl_df = pl_df.with_columns(repl_exprs)
                pl_df = pl_df.with_columns([pl.col(c).forward_fill().alias(c) for c in mapped_cols if c in all_cols])

            non_empty_exprs = [pl.col(c).is_not_null() for c in mapped_cols if c in all_cols]
            if non_empty_exprs:
                has_any = non_empty_exprs[0]
                for e in non_empty_exprs[1:]:
                    has_any = has_any | e
                pl_df = pl_df.filter(has_any)

        for c in id_cols:
            pl_df = pl_df.with_columns(_blank_to_null_expr(c).forward_fill().alias(c))

        id_present = pl.col(id_cols[0]).is_not_null()
        for c in id_cols[1:]:
            id_present = id_present & pl.col(c).is_not_null()

        if amt_cols:
            amt_non_empty = _blank_to_null_expr(amt_cols[0]).is_not_null()
            for c in amt_cols[1:]:
                amt_non_empty = amt_non_empty | _blank_to_null_expr(c).is_not_null()
            pl_df = pl_df.filter(id_present | amt_non_empty)
        else:
            pl_df = pl_df.filter(id_present)

        temp_filter_col = "__temp_filter_key__"
        if len(acc_cols) > 1:
            pl_df = pl_df.with_columns(
                pl.concat_str([pl.col(c).fill_null("").cast(pl.Utf8) for c in acc_cols], separator=" - ").alias(temp_filter_col)
            )
        else:
            pl_df = pl_df.with_columns(pl.col(acc_cols[0]).fill_null("").cast(pl.Utf8).alias(temp_filter_col))

        if self.target_accounts:
            target_rows = pl_df.filter(pl.col(temp_filter_col).is_in(list(self.target_accounts)))
            if target_rows.height > 0:
                target_ids = target_rows.select(id_cols).unique()
                df_target_pl = pl_df.join(target_ids, on=id_cols, how="inner")
            else:
                df_target_pl = pl_df.head(0)
        else:
            df_target_pl = pl_df.clone()

        if self.exclude_accounts:
            df_exclude_pl = pl_df.filter(pl.col(temp_filter_col).is_in(list(self.exclude_accounts)))
        else:
            df_exclude_pl = pl_df.head(0)

        if temp_filter_col in df_target_pl.columns:
            df_target_pl = df_target_pl.drop(temp_filter_col)
        if temp_filter_col in df_exclude_pl.columns:
            df_exclude_pl = df_exclude_pl.drop(temp_filter_col)
        if temp_filter_col in pl_df.columns:
            pl_df = pl_df.drop(temp_filter_col)

        # 后续只需要全量数据的列顺序，避免把清洗后的大表再复制回 pandas。
        df_out = pd.DataFrame(columns=pl_df.columns)
        df_target = self._to_pandas_df_safe(df_target_pl)
        df_exclude = self._to_pandas_df_safe(df_exclude_pl)

        for out_df in (df_out, df_target, df_exclude):
            if not out_df.empty:
                out_df.fillna("", inplace=True)

        df_target.reset_index(drop=True, inplace=True)
        if not df_exclude.empty:
            df_exclude.reset_index(drop=True, inplace=True)

        return df_out, df_target, df_exclude, acc_cols

    def run_export(self, map_inv, logic, batches=None):
        batches = batches or self._get_effective_batches()
        if not batches:
            batches = [("全部", set())]
        self.export_cancel_event.clear()
        self.export_in_progress = True

        try:
            if len(batches) == 1:
                name, accs = batches[0]
                self.target_accounts = set(accs)
                return self._run_export_single(map_inv, logic, batch_name=name, ask_save=True, save_path_override=None, export_cols_override=None, show_done_msg=True)

            self.btn_run.config(state="disabled")
            first_name, first_accs = batches[0]
            self.target_accounts = set(first_accs)
            ok, first_save_path, first_export_cols, first_msg = self._run_export_single(
                map_inv, logic, batch_name=first_name, ask_save=True, save_path_override=None, export_cols_override=None, show_done_msg=False
            )
            if not ok:
                return
            msgs = [f"{first_name}: {first_msg}"]
            for i, (batch_name, batch_accs) in enumerate(batches[1:], start=2):
                self._raise_if_export_cancelled()
                self.target_accounts = set(batch_accs)
                cur_path = self._build_batch_save_path(first_save_path, batch_name, i)
                ok, _, _, cur_msg = self._run_export_single(
                    map_inv, logic, batch_name=batch_name, ask_save=False, save_path_override=cur_path, export_cols_override=first_export_cols, show_done_msg=False
                )
                if not ok:
                    return
                msgs.append(f"{batch_name}: {cur_msg}")
            self.root.after(0, lambda: messagebox.showinfo("完成", "批量导出完成：\n" + "\n".join(msgs)))
        except ExportCancelled:
            self.root.after(0, lambda: messagebox.showinfo("已终止", "导出已按你的操作终止。"))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
        finally:
            self.export_in_progress = False
            self.root.after(0, self.hide_progress)
            self.root.after(0, lambda: self.btn_run.config(state="normal"))

    def _run_export_single(self, map_inv, logic, batch_name=None, ask_save=True, save_path_override=None, export_cols_override=None, show_done_msg=True):
        tracer = ExportPerfTracer(
            self.export_perf_log_path,
            context={"batch_name": batch_name or "default", "logic": logic, "ask_save": ask_save},
        )
        try:
            self.btn_run.config(state="disabled")
            self._raise_if_export_cancelled()
            tracer.event("run_export_single_begin")
            
            # ==========================================
            # [调整点 1] 将选列逻辑移到最前面
            # ==========================================
            export_cols = export_cols_override
            # 只有当勾选了"仅导出部分列"时才触发
            t_choose_cols = time.perf_counter()
            if export_cols is None and self.var_partial_col.get():
                self.thread_event.clear()
                # 在主线程弹出选列窗口
                self.root.after(0, self._ask_columns_ui)
                # 等待用户操作完成
                self.thread_event.wait()
                
                # 如果用户没选或者直接关闭了窗口，则终止流程
                if not self.user_selected_cols: 
                    self.root.after(0, lambda: self.btn_run.config(state="normal"))
                    return False, None, None, "已取消"
                export_cols = self.user_selected_cols
            tracer.event("choose_export_cols", elapsed_s=round(time.perf_counter() - t_choose_cols, 6), selected_count=(len(export_cols) if export_cols else 0))

            t_pick_save = time.perf_counter()
            if ask_save:
                self.thread_event.clear()
                self.root.after(0, self._ask_save_ui)
                self.thread_event.wait()
                if not self.user_save_path:
                    return False, None, export_cols, "已取消"
            else:
                self.user_save_path = save_path_override
                if not self.user_save_path:
                    return False, None, export_cols, "缺少输出路径"
            self._raise_if_export_cancelled()
            tracer.event("resolve_save_path", elapsed_s=round(time.perf_counter() - t_pick_save, 6), save_path=self.user_save_path)

            base_name, ext = os.path.splitext(self.user_save_path)
            is_csv = (ext.lower() == '.csv')

            # ==========================================
            # [调整点 2] 选完列后，才开始显示进度条和读数据
            # ==========================================
            self.show_progress("正在读取全量数据...", allow_cancel=True)

            t_read_all = time.perf_counter()
            use_cached = False
            cache_tag = self._current_source_tag()
            if self.full_cache_running:
                self._perf_log("导出前等待后台全量缓存完成...")
                self.full_cache_event.wait()

            if self.full_cache_ready and self.full_cache_df is not None:
                if self.full_cache_header_idx == int(self.header_row_idx) and self.full_cache_source_tag == cache_tag:
                    df = self.full_cache_df.copy(deep=False)
                    use_cached = True
                else:
                    df = self._universal_loader(self.header_row_idx)
            else:
                if self.full_cache_error:
                    self._perf_log(f"后台全量缓存不可用，改为导出前读取: {self.full_cache_error}")
                df = self._universal_loader(self.header_row_idx)

            self._raise_if_export_cancelled()
            tracer.event(
                "read_full_data",
                elapsed_s=round(time.perf_counter() - t_read_all, 6),
                rows=int(len(df)),
                cols=int(len(df.columns)),
                from_cache=bool(use_cached),
            )
            id_cols = self._get_voucher_id_cols(map_inv, df)

            # === 新增：对所有已映射字段做向上填充（ffill），全空行剔除 ===
            t_preprocess = time.perf_counter()
            df, df_target, df_exclude, acc_cols = self._preprocess_and_filter_with_polars(df, map_inv)
            tracer.event(
                "preprocess_filter_polars",
                rows=int(len(df)),
                target_rows=int(len(df_target)),
                exclude_rows=int(len(df_exclude)),
            )

            # 索引对齐
            df_target.reset_index(drop=True, inplace=True)
            if not df_exclude.empty: df_exclude.reset_index(drop=True, inplace=True)
            tracer.event("preprocess_and_filter", elapsed_s=round(time.perf_counter() - t_preprocess, 6), target_rows=int(len(df_target)), exclude_rows=int(len(df_exclude)))

            t_logic_calc = time.perf_counter()
            df_target = self._ensure_net_column_polars(df_target, map_inv, logic)
            tracer.event("logic_calc_net", elapsed_s=round(time.perf_counter() - t_logic_calc, 6), logic=logic)

            # 日期列按ID向上填充（处理合并单元格/空白）
            if map_inv.get('role_date'):
                date_col = map_inv['role_date'][0]
                if date_col in df_target.columns:
                    id_cols = self._get_voucher_id_cols(map_inv, df_target)
                    df_target[date_col] = (
                        df_target.groupby(id_cols)[date_col]
                        .ffill()
                        .bfill()
                    )

            # === 损益结转标记（凭证明细辅助列） ===
            loss_ids = set()
            loss_col = "【损益结转】"
            if self.var_mark_loss.get():
                if map_inv.get('role_acc'):
                    acc_cols = map_inv['role_acc']
                    loss_mask = False
                    for c in acc_cols:
                        if c in df_target.columns:
                            loss_mask = loss_mask | df_target[c].astype(str).str.contains(r'本年利润|未分配利润', na=False)
                    if loss_mask is not False and loss_mask.any():
                        id_join_col_for_loss = self._make_join_col(df_target, self._get_voucher_id_cols(map_inv, df_target))
                        loss_ids = set(df_target.loc[loss_mask, id_join_col_for_loss].astype(str).unique())
                        df_target[loss_col] = df_target[id_join_col_for_loss].astype(str).isin(loss_ids).map(lambda x: "损益结转" if x else "")
                    else:
                        df_target[loss_col] = ""
                else:
                    df_target[loss_col] = ""

            voucher_pivot = None
            v_id_col = None
            v_acc_col = None
            voucher_type_df = None
            vt_id_col = None
            vt_type_col = None
            vt_month_cols = None
            voucher_type_strict_df = None
            vt_id_col_s = None
            vt_type_col_s = None
            vt_month_cols_s = None

            need_voucher_suite = True
            t_voucher_suite = time.perf_counter()
            if need_voucher_suite:
                # === 生成“凭证”与“凭证类型”透视 ===
                t_vs1 = time.perf_counter()
                voucher_pivot, v_id_col, v_acc_col = self.build_voucher_pivot(df_target, map_inv, logic)
                elapsed_vs1 = time.perf_counter() - t_vs1
                tracer.event("build_voucher_pivot", elapsed_s=round(elapsed_vs1, 6), ok=bool(voucher_pivot is not None))
                if voucher_pivot is not None:
                    t_vs2 = time.perf_counter()
                    voucher_type_df, vt_id_col, vt_type_col, vt_month_cols = self.build_voucher_type_pivot(
                        voucher_pivot, df_target, map_inv, loss_ids, target_accounts=self.target_accounts
                    )
                    elapsed_vs2 = time.perf_counter() - t_vs2
                    tracer.event("build_voucher_type_loose", elapsed_s=round(elapsed_vs2, 6), ok=bool(voucher_type_df is not None))
                    # 始终生成严格表，避免套表中缺少“凭证类型-严格”
                    t_vs3 = time.perf_counter()
                    voucher_type_strict_df, vt_id_col_s, vt_type_col_s, vt_month_cols_s = self.build_voucher_type_pivot(
                        voucher_pivot, df_target, map_inv, loss_ids, mode="strict", target_accounts=self.target_accounts
                    )
                    tracer.event("build_voucher_type_strict", elapsed_s=round(time.perf_counter() - t_vs3, 6), ok=bool(voucher_type_strict_df is not None))
            tracer.event("build_voucher_suite", elapsed_s=round(time.perf_counter() - t_voucher_suite, 6), enabled=bool(need_voucher_suite))

            pivot_res = None
            pivot_err = None
            pivot_path = None
            t_pivot = time.perf_counter()
            if self.pivot_config:
                self.show_progress("生成透视表...", allow_cancel=True)
                p_df = df_target.copy()

                # 方案B：多ID合并列，仅用于透视表
                if logic == "B":
                    id_cols = self._get_voucher_id_cols(map_inv, p_df)
                    if len(id_cols) > 1:
                        id_join_col = "-".join(id_cols)
                        if id_join_col not in p_df.columns:
                            p_df[id_join_col] = p_df[id_cols].fillna("").astype(str).agg("-".join, axis=1)

                # 透视分析剔除损益结转凭证
                if self.var_mark_loss.get() and loss_ids:
                    id_join_col_for_pivot = self._make_join_col(p_df, self._get_voucher_id_cols(map_inv, p_df))
                    p_df = p_df[~p_df[id_join_col_for_pivot].astype(str).isin(loss_ids)].copy()

                # 确保透视表行字段存在（缺失则补空值）
                idx_fields = self.pivot_config.get('index') or []
                for f in idx_fields:
                    if f not in p_df.columns:
                        p_df[f] = ""
                
                # 确保透视前已有净额列
                p_df = self._ensure_net_column_polars(p_df, map_inv, logic)
                
                # 创建净额列用于透视表
                p_df['#_净额(Net)'] = p_df['__net__']

                # 过滤不存在的value字段，兜底为净额
                values_cfg = [v for v in (self.pivot_config.get('values') or []) if v in p_df.columns]
                if not values_cfg:
                    values_cfg = ['#_净额(Net)']
                self.pivot_config['values'] = values_cfg
                
                # 转换用户选择的所有value字段为数值类型
                for val_col in values_cfg:
                    if val_col != '#_净额(Net)':  # 净额已经是数值类型
                        p_df[val_col] = self._pl_clean_num_series(p_df[val_col])

                date_cols_marked = map_inv.get('role_date', [])
                def safe_convert_date(series):
                    s_clean = series.fillna("").astype(str).str.replace(r'\.0$', '', regex=True)
                    s_num = pd.to_numeric(s_clean, errors='coerce')
                    mask_8digit = s_num.notna() & (s_num > 19000000) & (s_num < 21000000)
                    dt_8digit = pd.to_datetime(s_num[mask_8digit].astype(int).astype(str), format='%Y%m%d', errors='coerce')
                    mask_serial = s_num.notna() & (s_num > 1000) & (s_num < 100000)
                    dt_serial = pd.to_datetime(s_num[mask_serial], unit='D', origin='1899-12-30', errors='coerce')
                    dt_std = pd.to_datetime(s_clean[~mask_serial & ~mask_8digit], errors='coerce')
                    final_dt = dt_serial.reindex(series.index).combine_first(dt_8digit).combine_first(dt_std)
                    return final_dt.dt.strftime('%Y-%m').fillna("Unknown")

                columns_cfg = [c for c in (self.pivot_config.get('columns') or []) if c in p_df.columns]
                final_cols = []
                for col_name in columns_cfg:
                    if col_name in date_cols_marked:
                        new_col = col_name + "(月)"
                        p_df[new_col] = safe_convert_date(p_df[col_name])
                        final_cols.append(new_col)
                    else:
                        final_cols.append(col_name)
                
                try:
                    columns_for_pivot = final_cols if final_cols else None
                    idx_cols = [c for c in (self.pivot_config.get('index') or []) if c in p_df.columns]
                    val_cols = [c for c in (self.pivot_config.get('values') or []) if c in p_df.columns]
                    col_cols = [c for c in (columns_for_pivot or []) if c in p_df.columns] if columns_for_pivot else []
                    if not idx_cols or not val_cols:
                        raise ValueError("透视配置缺少有效的行字段或值字段")
                    selected_cols = list(dict.fromkeys(idx_cols + col_cols + val_cols))
                    pl_tmp = self._to_polars_df_safe(p_df[selected_cols])
                    pl_tmp = pl_tmp.with_columns([
                        pl.col(c).cast(pl.Float64, strict=False).fill_null(0.0).alias(c) for c in val_cols
                    ])
                    gb_cols = idx_cols + col_cols
                    agg_exprs = [pl.col(c).sum().alias(c) for c in val_cols]
                    gb_pd = self._to_pandas_df_safe(pl_tmp.group_by(gb_cols).agg(agg_exprs))
                    if col_cols:
                        pivot_res = pd.pivot_table(
                            gb_pd,
                            index=idx_cols,
                            columns=col_cols if len(col_cols) > 1 else col_cols[0],
                            values=val_cols,
                            aggfunc='sum'
                        ).fillna(0)
                    else:
                        pivot_res = gb_pd.groupby(idx_cols, as_index=True)[val_cols].sum()
                    pivot_err = None
                except Exception as e:
                    pivot_err = str(e)
                    print(f"透视表生成错误: {e}")
                    pivot_res = None
                    # 兜底：仅按行字段生成净额透视，避免完全无表
                    try:
                        pivot_res = pd.pivot_table(
                            p_df, index=self.pivot_config['index'], values=['#_净额(Net)'], aggfunc='sum'
                        ).fillna(0)
                        pivot_err = None
                    except Exception as e2:
                        pivot_err = str(e2)
                        print(f"透视表兜底失败: {e2}")
                        pivot_res = None
            tracer.event("build_pivot", elapsed_s=round(time.perf_counter() - t_pivot, 6), has_pivot=bool(self.pivot_config), pivot_ok=bool(pivot_res is not None))
            suite_written = False
            suite_path = f"{base_name}_套表.xlsx"
            suite_enabled = any([
                voucher_pivot is not None,
                voucher_type_df is not None,
                voucher_type_strict_df is not None,
                pivot_res is not None,
            ])
            if suite_enabled:
                self._raise_if_export_cancelled()
                self.show_progress("阶段1/2：导出套表...", allow_cancel=True)
                t_write_suite = time.perf_counter()
                tracer.event(
                    "suite_stage_begin",
                    target_rows=int(len(df_target)),
                    has_voucher=bool(voucher_pivot is not None),
                    has_type_loose=bool(voucher_type_df is not None),
                    has_type_strict=bool(voucher_type_strict_df is not None),
                    has_pivot=bool(pivot_res is not None),
                )
                num_cols_for_suite = []
                num_cols_for_suite += map_inv.get('role_amt', []) if map_inv.get('role_amt') else []
                num_cols_for_suite += map_inv.get('role_dr', []) if map_inv.get('role_dr') else []
                num_cols_for_suite += map_inv.get('role_cr', []) if map_inv.get('role_cr') else []
                num_cols_for_suite = list(dict.fromkeys(num_cols_for_suite))
                final_cols_for_suite = [c for c in df.columns if (not export_cols or c in export_cols)]
                with pd.ExcelWriter(suite_path, engine='xlsxwriter') as writer_suite:
                    if voucher_pivot is not None:
                        t_sheet = time.perf_counter()
                        v_out = voucher_pivot.reset_index()
                        v_out.columns = self._flatten_columns(v_out.columns)
                        v_out.to_excel(writer_suite, sheet_name="凭证", index=False)
                        tracer.event("suite_sheet_voucher", elapsed_s=round(time.perf_counter() - t_sheet, 6), rows=int(len(v_out)))
                    if voucher_type_df is not None:
                        t_sheet = time.perf_counter()
                        vt_out = voucher_type_df.copy()
                        vt_out.columns = self._flatten_columns(vt_out.columns)
                        drop_cols = [c for c in vt_out.columns if str(c).startswith("科目名称-类型_")]
                        if drop_cols:
                            vt_out = vt_out.drop(columns=drop_cols)
                        vt_out = self._reorder_voucher_type_columns(vt_out, vt_type_col, vt_id_col, v_acc_col)
                        vt_out.to_excel(writer_suite, sheet_name="凭证类型-宽松", index=False)
                        tracer.event("suite_sheet_type_loose", elapsed_s=round(time.perf_counter() - t_sheet, 6), rows=int(len(vt_out)))
                    if voucher_type_strict_df is not None:
                        t_sheet = time.perf_counter()
                        vt_out_s = voucher_type_strict_df.copy()
                        vt_out_s.columns = self._flatten_columns(vt_out_s.columns)
                        drop_cols_s = [c for c in vt_out_s.columns if str(c).startswith("科目名称-类型_")]
                        if drop_cols_s:
                            vt_out_s = vt_out_s.drop(columns=drop_cols_s)
                        vt_out_s = self._reorder_voucher_type_columns(vt_out_s, vt_type_col_s, vt_id_col_s, v_acc_col)
                        vt_out_s.to_excel(writer_suite, sheet_name="凭证类型-严格", index=False)
                        tracer.event("suite_sheet_type_strict", elapsed_s=round(time.perf_counter() - t_sheet, 6), rows=int(len(vt_out_s)))
                    if pivot_res is not None:
                        t_sheet = time.perf_counter()
                        pivot_out = pivot_res.reset_index()
                        pivot_out.columns = self._flatten_pivot_columns(pivot_out.columns)
                        pivot_out.to_excel(writer_suite, sheet_name="透视分析", index=False)
                        tracer.event("suite_sheet_pivot", elapsed_s=round(time.perf_counter() - t_sheet, 6), rows=int(len(pivot_out)))

                    t_fmt = time.perf_counter()
                    self._apply_output_formatting(
                        writer_suite,
                        df_target,
                        final_cols_for_suite,
                        num_cols_for_suite,
                        map_inv,
                        voucher_type_df,
                        vt_type_col,
                        vt_id_col,
                        v_acc_col,
                        pivot_res,
                        voucher_type_strict_df,
                        vt_type_col_s,
                        vt_id_col_s,
                        target_accounts=self.target_accounts,
                    )
                    tracer.event("suite_formatting", elapsed_s=round(time.perf_counter() - t_fmt, 6))
                    # 双保险：确保“凭证”sheet在套表中隐藏
                    if "凭证" in writer_suite.sheets:
                        try:
                            # 先激活可见sheet，避免Excel对活动sheet隐藏限制
                            for _sn in ["凭证类型-宽松", "凭证类型-严格", "透视分析"]:
                                if _sn in writer_suite.sheets:
                                    writer_suite.sheets[_sn].activate()
                                    break
                            writer_suite.sheets["凭证"].hide()
                        except Exception:
                            pass
                suite_elapsed = round(time.perf_counter() - t_write_suite, 6)
                tracer.event("write_suite_workbook", elapsed_s=suite_elapsed, path=suite_path)
                if is_csv:
                    tracer.event("write_csv_pivot_suite", elapsed_s=suite_elapsed, path=suite_path)
                suite_written = True
                self._raise_if_export_cancelled()

            if self.var_mark_je.get():
                t_je_mark = time.perf_counter()
                if not df_target.empty:
                    self.show_progress("执行 JE 智能标记...", allow_cancel=True)
                    acc_col_prime = acc_cols[0]
                    tracer.event("je_mark_begin", target_rows=int(len(df_target)), target_accounts=int(len(self.target_accounts)))
                    
                    # 获取用户映射的单位名称列（role_entity）
                    entity_col = map_inv.get('role_entity', [None])[0] if map_inv.get('role_entity') else None
                    
                    # 1. 创建目标科目掩码
                    # 重新计算temp_filter_col，用于识别目标科目行
                    temp_filter_col = "__temp_filter_key__"
                    if len(acc_cols) > 1:
                        df_target[temp_filter_col] = self._combine_account_key(df_target, acc_cols)
                    else:
                        df_target[temp_filter_col] = df_target[acc_cols[0]]
                    tracer.event("je_mark_prepare_filter_col", multi_acc=bool(len(acc_cols) > 1))
                    
                    # 创建目标科目掩码：仅对目标科目行执行匹配（损益结转不参与）
                    is_target_account = df_target[temp_filter_col].isin(self.target_accounts)
                    if self.var_mark_loss.get() and loss_ids:
                        loss_id_col = self._make_join_col(df_target, self._get_voucher_id_cols(map_inv, df_target))
                        is_target_account = is_target_account & ~df_target[loss_id_col].astype(str).isin(loss_ids)
                    tracer.event("je_mark_target_mask", target_hit_rows=int(is_target_account.sum()))
                    
                    # 初始化辅助列，避免string dtype导致数值赋值报错
                    cols_to_sync = ['【辅助_绝对值】', '【辅助_符号】', '【智能匹配状态】']
                    df_target['【辅助_绝对值】'] = np.nan
                    df_target['【辅助_符号】'] = ""
                    df_target['【智能匹配状态】'] = ""
                    # 损益结转行保持空白不参与匹配
                    if self.var_mark_loss.get() and loss_ids:
                        loss_id_col = self._make_join_col(df_target, self._get_voucher_id_cols(map_inv, df_target))
                        loss_mask = df_target[loss_id_col].astype(str).isin(loss_ids)
                        # 避免向 float 列写入空字符串导致 dtype 报错
                        df_target.loc[loss_mask, '【辅助_绝对值】'] = np.nan
                        df_target.loc[loss_mask, ['【辅助_符号】', '【智能匹配状态】']] = ""
                    
                    # 2. 仅对目标科目子集执行匹配
                    if is_target_account.any():
                        # 提取目标科目行
                        target_rows = df_target[is_target_account].copy()
                        tracer.event("je_mark_target_extract", rows=int(len(target_rows)))
                        
                        # 修复：方案B下使用原始借贷列进行匹配
                        # 匹配规则：同一科目，绝对值相等，且一个是正数另外一个是负数
                        # 核心逻辑：借方保持原符号，贷方保持原符号（不取反）
                        if logic == "B":
                            dr_col = map_inv['role_dr'][0]
                            cr_col = map_inv['role_cr'][0]
                            def clean_num(s): return pd.to_numeric(s.astype(str).str.replace(r'[,"]','', regex=True), errors='coerce').fillna(0.0)
                            val_dr = clean_num(target_rows[dr_col])
                            val_cr = clean_num(target_rows[cr_col])
                            
                            # 方案B的特殊处理：创建用于匹配的金额列
                            # 核心逻辑：借方保持原符号，贷方保持原符号（不取反）
                            # 这样，借方100和贷方-90会变成+100和-90，可以正确匹配
                            # 借方-90和贷方90会变成-90和+90，可以正确匹配
                            
                            # 创建用于匹配的金额列
                            # 如果有借方，使用借方（保持原符号）
                            # 如果没有借方但有贷方，使用贷方（保持原符号，不取反）
                            target_rows['__match_amt__'] = np.where(
                                val_dr != 0, 
                                val_dr,  # 有借方时，使用借方（保持原符号，可能是正数或负数）
                                val_cr   # 没有借方但有贷方时，使用贷方（保持原符号，可能是正数或负数）
                            )
                            
                            # 对于同时有借方和贷方的记录，使用净额
                            both_mask = (val_dr != 0) & (val_cr != 0)
                            if both_mask.any():
                                net_amt = val_dr[both_mask] - val_cr[both_mask]
                                target_rows.loc[both_mask, '__match_amt__'] = net_amt
                            
                            # 执行智能匹配，使用匹配金额列
                            id_join_col_for_match = self._make_join_col(target_rows, self._get_voucher_id_cols(map_inv, target_rows))
                            df_calc = self.apply_je_2_0_matching(target_rows, acc_col_prime, '__match_amt__', entity_col, id_join_col_for_match, '__net__')
                            tracer.event("je_mark_match_done", logic=logic, mode="B", rows=int(len(target_rows)))
                            
                            # 删除临时列
                            if '__match_amt__' in target_rows.columns:
                                del target_rows['__match_amt__']
                        else:
                            # 方案A：有方向列时，匹配阶段按「贷方负数=正数」处理；否则直接用净额
                            if map_inv.get('role_dir'):
                                dir_col = map_inv['role_dir'][0]
                                amt_col = map_inv['role_amt'][0]
                                def clean_num(s): return pd.to_numeric(s.astype(str).str.replace(r'[,"]','', regex=True), errors='coerce').fillna(0.0)
                                raw_amt = clean_num(target_rows[amt_col])
                                # 扩展贷方识别：包含“贷/贷方/CR/C/H/负号”
                                cr_regex = r'(?:贷|贷方|Credit|Cr\b|^C$|^H$|[-−])'
                                is_credit = target_rows[dir_col].astype(str).str.contains(cr_regex, case=False, regex=True, na=False)
                                # 贷方且金额<0 → 匹配用金额取绝对值（当正数）；否则用净额
                                target_rows['__match_amt__'] = np.where(is_credit & (raw_amt < 0), raw_amt.abs(), target_rows['__net__'])
                                id_join_col_for_match = self._make_join_col(target_rows, self._get_voucher_id_cols(map_inv, target_rows))
                                df_calc = self.apply_je_2_0_matching(target_rows, acc_col_prime, '__match_amt__', entity_col, id_join_col_for_match, '__net__')
                                tracer.event("je_mark_match_done", logic=logic, mode="A_dir", rows=int(len(target_rows)))
                                if '__match_amt__' in target_rows.columns:
                                    del target_rows['__match_amt__']
                            else:
                                id_join_col_for_match = self._make_join_col(target_rows, self._get_voucher_id_cols(map_inv, target_rows))
                                df_calc = self.apply_je_2_0_matching(target_rows, acc_col_prime, '__net__', entity_col, id_join_col_for_match, '__net__')
                                tracer.event("je_mark_match_done", logic=logic, mode="A_no_dir", rows=int(len(target_rows)))
                        
                        if not df_calc.empty:
                            # 将匹配结果合并回原数据框
                            # 使用索引对齐，确保只有目标科目行被更新
                            for col in cols_to_sync:
                                df_target.loc[is_target_account, col] = df_calc[col]
                            # 方案A有方向列时，匹配阶段用了 __match_amt__（贷方负数取绝对值），
                            # 会导致【辅助_符号】/【智能匹配状态】偏离真实符号。这里统一按净额符号重算。
                            if logic == "A" and map_inv.get("role_dir") and is_target_account.any():
                                # 用净额作为最终符号来源，保证“贷方负数”保持负号，同时不丢失计提端
                                net_full = df_target["__net__"]
                                df_target.loc[is_target_account, "【辅助_绝对值】"] = net_full.loc[is_target_account].abs().round(2)
                                df_target.loc[is_target_account, "【辅助_符号】"] = np.where(
                                    net_full.loc[is_target_account] >= 0, "正数", "负数"
                                )
                                df_target.loc[is_target_account, "【智能匹配状态】"] = "未匹配"

                                # 重新按“科目+绝对值(+单位)”进行成对匹配，确保计提/冲销成对输出
                                group_keys = [acc_col_prime, "【辅助_绝对值】"]
                                tmp_entity_col = None
                                if entity_col and entity_col in df_target.columns:
                                    tmp_entity_col = "__tmp_entity__"
                                    df_target.loc[is_target_account, tmp_entity_col] = (
                                        df_target.loc[is_target_account, entity_col]
                                        .astype(str)
                                        .map(self.entity_map)
                                        .fillna(df_target.loc[is_target_account, entity_col].astype(str))
                                    )
                                    group_keys.append(tmp_entity_col)

                                grouped = df_target.loc[is_target_account].groupby(group_keys, sort=False, as_index=True)
                                rematch_groups = 0
                                for _, group in grouped:
                                    rematch_groups += 1
                                    pos = group[group["【辅助_符号】"] == "正数"].index.tolist()
                                    neg = group[group["【辅助_符号】"] == "负数"].index.tolist()
                                    pairs_to_match = min(len(pos), len(neg))
                                    if pairs_to_match > 0:
                                        df_target.loc[pos[:pairs_to_match], "【智能匹配状态】"] = "已匹配-计提"
                                        df_target.loc[neg[:pairs_to_match], "【智能匹配状态】"] = "已匹配-冲销"
                                tracer.event("je_mark_a_rematch", groups=int(rematch_groups))

                                if tmp_entity_col and tmp_entity_col in df_target.columns:
                                    df_target.drop(columns=[tmp_entity_col], inplace=True)
                    
                    # 3. 移除临时列
                    if temp_filter_col in df_target.columns:
                        del df_target[temp_filter_col]
                
                # 确保所有辅助列都存在并填充空值
                for c in ['【辅助_绝对值】', '【辅助_符号】', '【智能匹配状态】']:
                    if c not in df_target.columns:
                        df_target[c] = np.nan if c == '【辅助_绝对值】' else ""
                    else:
                        if c == '【辅助_绝对值】':
                            df_target[c] = pd.to_numeric(df_target[c], errors='coerce')
                        else:
                            df_target[c] = df_target[c].fillna('')
                tracer.event("je_mark", elapsed_s=round(time.perf_counter() - t_je_mark, 6), enabled=True, target_rows=int(len(df_target)))
            else:
                tracer.event("je_mark", elapsed_s=0.0, enabled=False, target_rows=int(len(df_target)))

            # 数值字段格式化（凭证明细/剔除明细）
            t_num_clean = time.perf_counter()
            num_cols = []
            num_cols += map_inv.get('role_amt', []) if map_inv.get('role_amt') else []
            num_cols += map_inv.get('role_dr', []) if map_inv.get('role_dr') else []
            num_cols += map_inv.get('role_cr', []) if map_inv.get('role_cr') else []
            num_cols = list(dict.fromkeys(num_cols))
            for c in num_cols:
                if c in df_target.columns:
                    df_target[c] = self._pl_clean_num_series(df_target[c])
                if not df_exclude.empty and c in df_exclude.columns:
                    df_exclude[c] = self._pl_clean_num_series(df_exclude[c])
            tracer.event("num_clean", elapsed_s=round(time.perf_counter() - t_num_clean, 6), num_col_count=len(num_cols))

            if '__net__' in df_target.columns: del df_target['__net__']

            self.hide_progress(); self.thread_event.clear()
            self._raise_if_export_cancelled()
            limit_threshold = 1000000
            
            t_split_decision = time.perf_counter()
            parts = 1
            if len(df_target) > limit_threshold:
                suggested = math.ceil(len(df_target) / 900000)
                self.root.after(0, lambda: self._ask_split_count_on_main(suggested, len(df_target)))
                self.thread_event.wait()
                if not self.user_split_count: 
                    self.root.after(0, lambda: self.btn_run.config(state="normal"))
                    return False, None, export_cols, "已取消"
                parts = self.user_split_count
            tracer.event("split_decision", elapsed_s=round(time.perf_counter() - t_split_decision, 6), parts=int(parts), target_rows=int(len(df_target)))

            self.show_progress("写入文件...", allow_cancel=True)
            
            # 使用最前面选好的 export_cols 变量
            final_cols = [c for c in df.columns if (not export_cols or c in export_cols)]
            aux_cols = []
            if self.var_mark_loss.get():
                aux_cols.append('【损益结转】')
            if self.var_mark_je.get():
                aux_cols = ['【辅助_绝对值】', '【辅助_符号】', '【智能匹配状态】'] + (
                    ['【损益结转】'] if self.var_mark_loss.get() else []
                )
            t_cols = (aux_cols + final_cols) if aux_cols else final_cols
            
            if is_csv:
                t_write_csv = time.perf_counter()
                target_path = f"{base_name}_凭证明细.csv"
                exclude_path = f"{base_name}_剔除明细.csv"
                tracer.event("write_csv_begin", target_rows=int(len(df_target)), cols=int(len(t_cols)), parts=int(parts))
                stream_chunk_size = 200000
                
                if parts > 1:
                    chunk_size = math.ceil(len(df_target) / parts)
                    for i in range(parts):
                        t_chunk = time.perf_counter()
                        self._raise_if_export_cancelled()
                        start, end = i*chunk_size, (i+1)*chunk_size
                        sub = df_target.iloc[start:end]
                        p = f"{base_name}_凭证明细_Part{i+1}.csv"
                        sub.to_csv(p, index=False, encoding='utf-8-sig', columns=t_cols)
                        tracer.event("write_csv_chunk", chunk=i + 1, rows=int(len(sub)), elapsed_s=round(time.perf_counter() - t_chunk, 6), path=p)
                    msg = f"已分切{parts}个CSV"
                elif len(df_target) > stream_chunk_size:
                    total_rows = len(df_target)
                    chunk_idx = 0
                    for start in range(0, total_rows, stream_chunk_size):
                        t_chunk = time.perf_counter()
                        self._raise_if_export_cancelled()
                        end = min(start + stream_chunk_size, total_rows)
                        sub = df_target.iloc[start:end]
                        write_mode = 'w' if chunk_idx == 0 else 'a'
                        write_header = (chunk_idx == 0)
                        sub.to_csv(
                            target_path,
                            index=False,
                            encoding='utf-8-sig',
                            columns=t_cols,
                            mode=write_mode,
                            header=write_header,
                        )
                        chunk_idx += 1
                        tracer.event(
                            "write_csv_stream_chunk",
                            chunk=int(chunk_idx),
                            rows=int(len(sub)),
                            elapsed_s=round(time.perf_counter() - t_chunk, 6),
                            path=target_path,
                        )
                    msg = "CSV导出成功"
                else:
                    t_chunk = time.perf_counter()
                    self._raise_if_export_cancelled()
                    df_target.to_csv(target_path, index=False, encoding='utf-8-sig', columns=t_cols)
                    tracer.event("write_csv_target", rows=int(len(df_target)), elapsed_s=round(time.perf_counter() - t_chunk, 6), path=target_path)
                    msg = "CSV导出成功"

                if not df_exclude.empty:
                    t_ex = time.perf_counter()
                    self._raise_if_export_cancelled()
                    df_exclude.to_csv(exclude_path, index=False, encoding='utf-8-sig', columns=final_cols)
                    tracer.event("write_csv_exclude", rows=int(len(df_exclude)), elapsed_s=round(time.perf_counter() - t_ex, 6), path=exclude_path)
                if pivot_res is not None and (not suite_written):
                    t_pivot_xlsx = time.perf_counter()
                    self._raise_if_export_cancelled()
                    pivot_path = f"{base_name}_透视分析.xlsx"
                    with pd.ExcelWriter(pivot_path, engine='xlsxwriter') as pw:
                        pivot_out = pivot_res.reset_index()
                        pivot_out.columns = self._flatten_pivot_columns(pivot_out.columns)
                        pivot_out.to_excel(pw, sheet_name="透视分析", index=False)
                        if voucher_type_df is not None:
                            vt_out = voucher_type_df.copy()
                            vt_out.columns = self._flatten_columns(vt_out.columns)
                            # 删除科目名称-类型_1 等冗余列
                            drop_cols = [c for c in vt_out.columns if str(c).startswith("科目名称-类型_")]
                            if drop_cols:
                                vt_out = vt_out.drop(columns=drop_cols)
                            vt_out = self._reorder_voucher_type_columns(vt_out, vt_type_col, vt_id_col, v_acc_col)
                            vt_out.to_excel(pw, sheet_name="凭证类型-宽松", index=False)
                        if voucher_type_strict_df is not None:
                            vt_out_s = voucher_type_strict_df.copy()
                            vt_out_s.columns = self._flatten_columns(vt_out_s.columns)
                            # 删除科目名称-类型_1 等冗余列
                            drop_cols_s = [c for c in vt_out_s.columns if str(c).startswith("科目名称-类型_")]
                            if drop_cols_s:
                                vt_out_s = vt_out_s.drop(columns=drop_cols_s)
                            vt_out_s = self._reorder_voucher_type_columns(vt_out_s, vt_type_col_s, vt_id_col_s, v_acc_col)
                            vt_out_s.to_excel(pw, sheet_name="凭证类型-严格", index=False)
                        # 应用与Excel导出一致的格式
                        self._apply_output_formatting(
                            pw,
                            df_target,
                            t_cols,
                            num_cols,
                            map_inv,
                            voucher_type_df,
                            vt_type_col,
                            vt_id_col,
                            v_acc_col,
                            pivot_res,
                            voucher_type_strict_df,
                            vt_type_col_s,
                            vt_id_col_s,
                            target_accounts=self.target_accounts,
                        )
                    tracer.event("write_csv_pivot_suite", elapsed_s=round(time.perf_counter() - t_pivot_xlsx, 6), path=pivot_path)
                    msg += " (+透视表)"
                elif suite_written:
                    msg += " (+套表)"
                tracer.event("write_csv_total", elapsed_s=round(time.perf_counter() - t_write_csv, 6))
            else:
                split_export = suite_enabled
                suite_path = f"{base_name}_套表.xlsx" if split_export else None

                if split_export:
                    self.show_progress("阶段2/2：导出凭证明细...", allow_cancel=True)

                t_write_detail = time.perf_counter()
                tracer.event("write_excel_begin", target_rows=int(len(df_target)), cols=int(len(t_cols)), parts=int(parts))
                with pd.ExcelWriter(self.user_save_path, engine='xlsxwriter') as writer:
                    if parts > 1:
                        chunk_size = math.ceil(len(df_target) / parts)
                        for i in range(parts):
                            t_chunk = time.perf_counter()
                            self._raise_if_export_cancelled()
                            start, end = i*chunk_size, (i+1)*chunk_size
                            sub = df_target.iloc[start:end]
                            sn = f"凭证明细_{i+1}" if i > 0 else "凭证明细"
                            sub.to_excel(writer, sheet_name=sn, index=False, columns=t_cols)
                            tracer.event("write_excel_detail_chunk", chunk=i + 1, rows=int(len(sub)), elapsed_s=round(time.perf_counter() - t_chunk, 6), sheet=sn)
                    else:
                        t_chunk = time.perf_counter()
                        self._raise_if_export_cancelled()
                        df_target.to_excel(writer, sheet_name="凭证明细", index=False, columns=t_cols)
                        tracer.event("write_excel_detail_main", rows=int(len(df_target)), elapsed_s=round(time.perf_counter() - t_chunk, 6), sheet="凭证明细")

                    if not df_exclude.empty:
                        if len(df_exclude) > 1000000:
                            ex_p = math.ceil(len(df_exclude)/900000)
                            ex_c = math.ceil(len(df_exclude)/ex_p)
                            for i in range(ex_p):
                                t_ex = time.perf_counter()
                                self._raise_if_export_cancelled()
                                s, e = i*ex_c, (i+1)*ex_c
                                sn = f"剔除明细_{i+1}" if i>0 else "剔除明细"
                                df_exclude.iloc[s:e].to_excel(writer, sheet_name=sn, index=False, columns=final_cols)
                                tracer.event("write_excel_exclude_chunk", chunk=i + 1, rows=int(e - s), elapsed_s=round(time.perf_counter() - t_ex, 6), sheet=sn)
                        else:
                            t_ex = time.perf_counter()
                            self._raise_if_export_cancelled()
                            df_exclude.to_excel(writer, sheet_name="剔除明细", index=False, columns=final_cols)
                            tracer.event("write_excel_exclude_main", rows=int(len(df_exclude)), elapsed_s=round(time.perf_counter() - t_ex, 6), sheet="剔除明细")

                    self._apply_output_formatting(
                        writer,
                        df_target,
                        t_cols,
                        num_cols,
                        map_inv,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        target_accounts=self.target_accounts,
                    )
                tracer.event("write_excel_detail_total", elapsed_s=round(time.perf_counter() - t_write_detail, 6), path=self.user_save_path)

                if split_export:
                    msg = f"已完成两阶段导出：套表[{os.path.basename(suite_path)}] + 明细[{os.path.basename(self.user_save_path)}]"
                else:
                    msg = "Excel导出成功"
            
            if pivot_err:
                msg += " (透视表失败: " + str(pivot_err) + ")"
            if batch_name:
                msg = f"[{batch_name}] {msg}"
            if show_done_msg:
                self.root.after(0, lambda: messagebox.showinfo("完成", msg))
            tracer.event("run_export_single_end", success=True, message=msg)
            return True, self.user_save_path, export_cols, msg
        except ExportCancelled:
            msg = "导出已终止"
            if show_done_msg:
                self.root.after(0, lambda: messagebox.showinfo("已终止", msg))
            tracer.event("run_export_single_end", success=False, cancelled=True, message=msg)
            return False, self.user_save_path, export_cols_override, msg
        except Exception as e:
            msg = str(e)
            # 强制显示错误信息到弹窗
            self.root.after(0, lambda: messagebox.showerror("错误", msg))
            tracer.event("run_export_single_end", success=False, error=msg)
            return False, self.user_save_path, export_cols_override, msg
        finally:
            tracer.close()
            self.root.after(0, self.hide_progress)
            self.root.after(0, lambda: self.btn_run.config(state="normal"))

    def apply_je_2_0_matching(self, df, subject_col, amt_col, entity_col=None, id_col=None, net_col=None):
        """
        执行JE 智能标记，参考JE正负数抓取的逻辑
        
        参数：
        - df: 待匹配的数据框
        - subject_col: 科目列名
        - amt_col: 金额列名
        - entity_col: 单位名称列名（可选）
        """
        # 修复浮点精度问题，添加四舍五入，保留两位小数
        df['【辅助_绝对值】'] = df[amt_col].abs().round(2)
        df['【辅助_符号】'] = np.where(df[amt_col] >= 0, '正数', '负数')
        df['【智能匹配状态】'] = '未匹配'
        
        # 确定分组键：包含单位名称（如果存在）、科目和绝对值
        group_keys = [subject_col, '【辅助_绝对值】']
        
        # 添加单位名称作为分组条件（如果存在）
        if entity_col and entity_col in df.columns:
            # 创建映射后的单位名称列
            # 使用entity_map字典进行映射，如果找不到映射关系则使用原始单位名称
            df['【辅助_映射单位】'] = df[entity_col].astype(str).map(self.entity_map).fillna(df[entity_col].astype(str))
            # 添加到分组键
            group_keys.append('【辅助_映射单位】')
        
        # 使用稳定的分组方式
        grouped = df.groupby(group_keys, sort=False, as_index=True)
        pos_matched_idxs = []
        neg_matched_idxs = []
        match_count = 0
        
        for _, group in grouped:
            # 获取正负数索引（避免重复子表过滤）
            sign_series = group['【辅助_符号】']
            pos = sign_series.index[sign_series.eq('正数')].tolist()
            neg = sign_series.index[sign_series.eq('负数')].tolist()
            
            # 匹配条件：同一组内存在正负数
            # 参考JE正负数抓取的逻辑，仅当组内存在正负两种符号时才匹配
            has_opposite_signs = len(group['【辅助_符号】'].unique()) > 1
            
            if has_opposite_signs:
                # 按数量匹配，确保公平匹配
                pairs_to_match = min(len(pos), len(neg))
                
                if pairs_to_match > 0:
                    # 传统正负匹配情况
                    pos_matched_idxs.extend(pos[:pairs_to_match])
                    neg_matched_idxs.extend(neg[:pairs_to_match])
                    match_count += pairs_to_match
        
        # 更新匹配状态
        if pos_matched_idxs: 
            df.loc[pos_matched_idxs, '【智能匹配状态】'] = '已匹配-计提'
        if neg_matched_idxs: 
            df.loc[neg_matched_idxs, '【智能匹配状态】'] = '已匹配-冲销'

        # 第二轮：跨行匹配（按ID汇总后匹配）
        cross_match_count = 0
        if id_col and id_col in df.columns:
            remaining = df[df['【智能匹配状态】'] == '未匹配']
            if not remaining.empty:
                net_col_use = net_col if (net_col and net_col in remaining.columns) else amt_col
                cross_group_keys = [subject_col]
                if entity_col and entity_col in remaining.columns:
                    cross_group_keys.append('【辅助_映射单位】')

                for _, g in remaining.groupby(cross_group_keys, sort=False):
                    # 按ID汇总净额
                    id_groups = g.groupby(id_col, sort=False).groups
                    id_sums = g.groupby(id_col, sort=False)[net_col_use].sum()
                    pos_ids = {}
                    neg_ids = {}
                    for vid, total in id_sums.items():
                        total_2 = round(float(total), 2)
                        if total_2 > 0:
                            pos_ids.setdefault(total_2, []).append(vid)
                        elif total_2 < 0:
                            neg_ids.setdefault(-total_2, []).append(vid)

                    # 同绝对值成对匹配
                    for abs_val, pos_list in pos_ids.items():
                        neg_list = neg_ids.get(abs_val, [])
                        pairs = min(len(pos_list), len(neg_list))
                        if pairs <= 0:
                            continue
                        for i in range(pairs):
                            pos_id = pos_list[i]
                            neg_id = neg_list[i]
                            # 仅标记当前科目(+单位)分组内的行
                            pos_idx = id_groups.get(pos_id, [])
                            neg_idx = id_groups.get(neg_id, [])
                            if len(pos_idx) > 0:
                                df.loc[list(pos_idx), '【智能匹配状态】'] = '跨行已匹配-计提'
                            if len(neg_idx) > 0:
                                df.loc[list(neg_idx), '【智能匹配状态】'] = '跨行已匹配-冲销'
                            cross_match_count += 1

        print(f"【JE 匹配】共找到 {match_count} 对匹配记录；跨行匹配 {cross_match_count} 对")
        return df

    def _ask_columns_ui(self):
        top = tk.Toplevel(self.root); _fit_toplevel_to_screen(top, 600, 600, min_width=480, min_height=420)
        f_btn = tk.Frame(top, padx=10, pady=10, bg="#f0f0f0"); f_btn.pack(side=tk.BOTTOM, fill=tk.X)
        cvs = tk.Canvas(top); sb = ttk.Scrollbar(top, orient="vertical", command=cvs.yview)
        frm = tk.Frame(cvs); cvs.create_window((0,0), window=frm, anchor="nw")
        cvs.configure(yscrollcommand=sb.set); cvs.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
        frm.bind("<Configure>", lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        
        vars = {c: tk.BooleanVar(value=True) for c in self.full_columns}
        for i, (col, var) in enumerate(vars.items()):
            r, c = divmod(i, 3); tk.Checkbutton(frm, text=col, variable=var).grid(row=r, column=c, sticky="w")
        
        def ok(): self.user_selected_cols = [c for c,v in vars.items() if v.get()]; top.destroy(); self.thread_event.set()
        def select_all_cols():
            for v in vars.values():
                v.set(True)
        self._bind_esc_close(top, lambda: (setattr(self, 'user_selected_cols', None), top.destroy(), self.thread_event.set()))
        self._bind_ctrl_a(top, callback=select_all_cols)
        ttk.Button(f_btn, text="取消", width=12, command=lambda: (setattr(self, 'user_selected_cols', None), top.destroy(), self.thread_event.set())).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(f_btn, text="确定", width=12, command=ok).pack(side=tk.RIGHT)
        top.protocol("WM_DELETE_WINDOW", lambda: (setattr(self, 'user_selected_cols', None), top.destroy(), self.thread_event.set()))

    def _ask_save_ui(self):
        # 默认导出格式改为CSV，提高导出速度
        initialdir = self._default_save_initialdir()
        kwargs = {
            "defaultextension": ".csv",
            "filetypes": [("CSV", "*.csv"), ("Excel", "*.xlsx")],
            "initialfile": self._build_default_save_name(".csv"),
        }
        if initialdir:
            kwargs["initialdir"] = initialdir
        self.user_save_path = filedialog.asksaveasfilename(**kwargs)
        self.thread_event.set()

def main(parent=None):
    if parent is not None:
        win = tk.Toplevel(parent)
        app = AuditApp_V70_2(win)
        win.wait_window()
    else:
        root = tk.Tk()
        app = AuditApp_V70_2(root)
        root.mainloop()


if __name__ == "__main__":
    main()
